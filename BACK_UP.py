from __future__ import print_function

from xml.sax import ErrorHandler
from xmlrpc.client import Error
import multiprocessing as mp
import tkinter.simpledialog as simpledialog
import tkinter.messagebox as messagebox
import socket
# from socket import socket, AF_INET, SOCK_STREAM
# socket
from typing import List, Optional
import serial.tools.list_ports
from serial import Serial
# serial
import tkinter as tk  # ウィンドウ作成用
# from tkinter import filedialog  # ファイルを開くダイアログ用
from tkinter import *
from tkinter import ttk
import tkinter.font as ft
from tkinter import scrolledtext
from openpyxl import load_workbook
from tkinter import scrolledtext
from PIL import Image, ImageTk, ImageOps
import numpy as np  # アフィン変換行列演算用

import time
import cv2
import math
import datetime
import os
import struct
from copy import deepcopy
from random import randint
import display_error
################ ROBOT SIMULATION ######################

import copy
from doctest import master
from OpenGL import GLU, GL
from OpenGL.GL import *
from OpenGL.arrays import vbo
import kinematics as kin
import dynamics as dyn
from graphics_objects import GroundGraphics
from graphics_objects import RobotGraphics
from graphics_objects import GraphicsOptions
import robot_defs
import transformations as tf
import inverse_kinematics as ik
from pyopengltk import OpenGLFrame
#################################### ROBOT ANNI ######################
from multiprocessing.resource_sharer import stop
from os import execv

from tkinter.ttk import *
from ttkthemes import ThemedStyle
from tkinter import messagebox
from matplotlib import pyplot as plt
from pygrabber.dshow_graph import FilterGraph
from tkinter import filedialog as fd
from functools import partial
import sys
import pickle
import threading
import tkinter.messagebox
import webbrowser

import pathlib

from numpy import mean
from os import path

DIR = pathlib.Path(__file__).parent.resolve()

#############################################

HOST = '192.168.0.10'
PORT = 8501
HOST1 = '192.168.0.11'
PORT1 = 8501

MAX_MESSAGE = 2048
NUM_THREAD = 1

CHR_CAN = '\18'
CHR_EOT = '\04'

# ボーレートのリスト
BAUDRATES = [2400, 4800, 9600, 19200, 38400, 56000, 57600, 115200]
# デリミタのリスト
DELIMITERS = {'[CR]': '\r', '[LF]': '\n', '[CRLF]': '\r\n', '[None]': ''}

# デリミタのリスト(履歴表示用)
REV_DELIMITER = {'\r': '[CR]', '\n': '[LF]', '\r\n': '[CRLF]', '': ''}
MODE_RUN = ["MODE_IO", "MODE_PULSE", "MODE_PC"]
NUM_MOTOR = ["ONE_MOTOR", "TWO_MOTOR"]
################################# MSP ###############################
MSP_HEADER = "$M>"
MSP_MOTOR_MEOBIET = 199
MSP_MOTOR_MODE = 100
MSP_MOTOR_POS = 101
MSP_MOTOR_VEL = 102
MSP_MOTOR_ACCEL_DECCEL = 103
MSP_MOTOR_VOL = 104
MSP_MOTOR_VOLLIMIT = 105
MSP_MOTOR_PID_P = 106
MSP_MOTOR_PID_VEL_P = 107
MSP_MOTOR_PID_VEL_I = 108
MSP_MOTOR_PID_VEL_D = 109

MSP_MOTOR_MOTOR_SET = 1
MSP_MOTOR_MODE_SET = 2
MSP_MOTOR_SAVE_SET = 3
MSP_MOTOR_RESET_SET = 4
MSP_MOTOR_RUN_SET = 5
MSP_MOTOR_ERROR_SET = 6  # new
MSP_MOTOR_NC_SET = 7  # new
MSP_MOTOR_DATAPOS1_SET = 8
MSP_MOTOR_DATAPOS2_SET = 9
MSP_MOTOR_PARASAVE1_SET = 34
MSP_MOTOR_PARASAVE2_SET = 35

MSP_MOTOR_POS11_SET = 10
MSP_MOTOR_POS12_SET = 11
MSP_MOTOR_POS21_SET = 12
MSP_MOTOR_POS22_SET = 13

MSP_MOTOR_VEL11_SET = 14
MSP_MOTOR_VEL12_SET = 15
MSP_MOTOR_VEL21_SET = 16
MSP_MOTOR_VEL22_SET = 17

MSP_MOTOR_ACCEL1_SET = 18
MSP_MOTOR_ACCEL2_SET = 19
MSP_MOTOR_DECCEL1_SET = 20
MSP_MOTOR_DECCEL2_SET = 21

MSP_MOTOR_VOL_M1_SET = 22
MSP_MOTOR_VOL_M2_SET = 23

MSP_MOTOR_VOLLIMIT_M1_SET = 24
MSP_MOTOR_VOLLIMIT_M2_SET = 25
MSP_MOTOR_PID_P_M1_SET = 26
MSP_MOTOR_PID_P_M2_SET = 27
MSP_MOTOR_PID_VEL_P_M1_SET = 28
MSP_MOTOR_PID_VEL_P_M2_SET = 29
MSP_MOTOR_PID_VEL_I_M1_SET = 30
MSP_MOTOR_PID_VEL_I_M2_SET = 31
MSP_MOTOR_PID_VEL_D_M1_SET = 32
MSP_MOTOR_PID_VEL_D_M2_SET = 33
MSP_MOTOR_FB_ICHI = 50
MSP_MOTOR_FB_VEL_ACC_DEC_M1 = 51
MSP_MOTOR_FB_VEL_ACC_DEC_M2 = 52

MSP_MOTOR_IDOU_1_SET = 60
MSP_MOTOR_IDOU_2_SET = 61
MSP_MOTOR_ACC_DEC_1_SET = 62
MSP_MOTOR_ACC_DEC_2_SET = 63
MSP_MOTOR_SPEED_1_SET = 64
MSP_MOTOR_SPEED_2_SET = 65

MSP_MOTOR_DISABLE_MOTOR_SET = 66
MSP_MOTOR_RESO1_SET = 67
MSP_MOTOR_RESO2_SET = 68
MSP_MOTOR_POLE_1_SET = 69
MSP_MOTOR_POLE_2_SET = 70
MSP_MOTOR_MODE_RUN = 71
MSP_MOTOR_MOTION_CONTROL1 = 72
MSP_MOTOR_MOTION_CONTROL2 = 73

IDLE = 0
HEADER_START = 1
HEADER_M = 2
HEADER_ARROW = 3
HEADER_SIZE = 4
HEADER_CMD = 5
HEADER_ERR = 6

c_state = IDLE
err_rcvd = False

checksum = 0
cmd = 0
offset = 0
dataSize = 0
inBuf = bytearray(256)

mode = 0
p = 0
payload = []
present = 0

time2 = 0
time3 = 0
time4 = 0
time5 = 0
time6 = 0

import decimal

df = decimal.Decimal('0.00')
global error, run_mode_val, cur_motor_number
global motor1_resolution, motor2_resolution, motor1_pole, motor2_pole
global data_motor1, data_motor2, control_motion_1, control_motion_2
global current_pos_11, current_pos_12, current_pos_21, current_pos_22
global current_vel11, current_vel12, current_vel21, current_vel22
global current_accel1, current_accel2, current_deccel1, current_deccel2
global current_vol_in1, current_vol_in2
global current_vol_limit1, current_vol_limit2
global pid_pos_p1, pid_pos_p2
global pid_vel_p1, pid_vel_p2
global pid_vel_i1, pid_vel_i2
global pid_vel_d1, pid_vel_d2
global position_now_motor1, position_now_motor2
global now_velocity_1, now_acceleration_1, now_decceleration_1
global now_velocity_2, now_acceleration_2, now_decceleration_2


class ROBOT_GRAPHIC(OpenGLFrame):

    def initgl(self):

        GL.glLoadIdentity()
        GLU.gluPerspective(45, (self.width / self.height), 0.1, 50.0)
        GL.glTranslatef(0.0, 0.0, -5)
        GL.glClearColor(0.6, 0.7, 1, 1.0)
        ######
        # Timing variables:
        self.sim_freq = 100.0
        self.render_freq = 30.0

        self.spin_time = time.time()
        self.spin_time_prev = self.spin_time

        self.render_time = time.time()
        self.render_time_prev = self.spin_time

        self.dt = 1.0 / self.sim_freq  # ideal, to be computed in loop

        # Kinematics and dynamics objects:
        self.robot = kin.RobotParams()
        self.robot_state = kin.RobotState(self.robot)

        self.sim_time = 0.0
        self.kin = kin.Kinematics(self.robot)  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.kin_des = kin.Kinematics(self.robot)  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.dyn = dyn.RecursiveNewtonEuler(self.robot, self.robot_state, self.kin)  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!

        self.integrator_type = robot_defs.INTEGRATOR_TYPE_NONE  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.controller_type = robot_defs.CONTROLLER_TYPE_PD  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.kin.initialize(tf.HomogeneousTransform(), np.zeros(6), np.zeros(6), self.robot_state.joint_state, self.dt)  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.ik = ik.InverseKinematics(self.robot, self.robot_state.joint_state)  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.initial_endeff_pos = self.kin.h_tf_links[-1].t()  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!
        self.initial_endeff_ori = self.kin.h_tf_links[-1].R()  # chu y !!!!!!!!!!!!!!!!!!!!!!!!!!

        # Camera spherical coordinates:
        self.eye_r = 3.0
        self.eye_th = 0.3 * np.pi
        self.eye_phi = 0.5

        self.eye_pos = np.array([self.eye_r * np.sin(self.eye_th) * np.cos(self.eye_phi),
                                 self.eye_r * np.sin(self.eye_th) * np.sin(self.eye_phi),
                                 self.eye_r * np.cos(self.eye_th)])
        self.center_pos = np.array([0.0, 0.0, 0.0])

        self.ground_graphics = GroundGraphics(length=20.0, width=20.0)

        self.robot_graphics = RobotGraphics()

        # General class for encapsulating graphics options, eg set draw
        # in wireframe, draw joint axes etc:
        self.graphics_options = GraphicsOptions(self.robot.n_dofs)  # self.robot.n_dofs
        self.initializeGL()
        self.resizeGL(500, 500)

    def update_view(self):
        self.eye_pos = np.array([self.eye_r * np.sin(self.eye_th) * np.cos(self.eye_phi),
                                 self.eye_r * np.sin(self.eye_th) * np.sin(self.eye_phi),
                                 self.eye_r * np.cos(self.eye_th)])
        look_vec = (self.center_pos - self.eye_pos) / np.linalg.norm(self.center_pos - self.eye_pos)
        up_vec = np.array([0.0, 0.0, 1.0])
        right_vec = np.cross(look_vec, up_vec)
        glLoadIdentity()
        GLU.gluLookAt(*np.concatenate((self.eye_pos, self.center_pos, up_vec)))

    def initializeGL(self):

        # Convenience function, calls glClearColor under the hood.
        # QColor is specified as RGB ints (0-255).  Specify this clear
        # color once and call glClear(GL_COLOR_BUFFER_BIT) before each
        # round of rendering (in paintGL):
        GL.glClearColor(100, 100, 100, 1)  # a grey background

        # Initialize the cube vertices:
        self.initGeometry()

        # Enable the depth buffer:
        glEnable(GL_DEPTH_TEST)

    def resizeGL(self, width, height):

        # Prevent the window height from being set to zero:
        if height == 0: height = 1

        # Set the affine transform converting 'display' to 'screen' coords:.  By using
        # the same width and height passed to the resizeGL function, we resize objects
        # to the new window size.
        glViewport(0, 0, width, height)

        # Set the target matrix stack to the projection matrix stack:
        glMatrixMode(GL_PROJECTION)

        # Replace the current matrix on the stack with the identity (homogeneous tf):
        glLoadIdentity()

        # Set up a perspective projection matrix:
        fov = 45.0  # field of view angle in y-direction (degrees)
        aspect = width / float(height)  # aspect ratio, determines field of view in x-direction
        zNear = 0.1  # distance from viewer to near clipping plane (+ve)
        zFar = 100.0  # distance from viewer to far clipping plane (+ve)
        GLU.gluPerspective(fov, aspect, zNear, zFar)

        # Set the target matrix stack to the modelview matrix stack:
        glMatrixMode(GL_MODELVIEW)

        # Create the initial view using the initial eye position:
        self.update_view()

    def paintGL(self):

        # Clear depth and color buffers in preparations for new rendering:
        glClear(GL_COLOR_BUFFER_BIT | GL_DEPTH_BUFFER_BIT)
        self.ground_graphics.render()
        self.robot_graphics.render(self.kin, self.graphics_options)

        # Draw the desired state using transparency:
        # if (app.bool_draw_wireframe.get()==True):
        #     self.graphics_options.set_draw_wireframe(True)
        # else:self.graphics_options.set_draw_wireframe(False)
        # if (app.bool_draw_motion_vectors.get()==True):
        #     self.graphics_options.set_draw_motion_vectors(True)
        # else:self.graphics_options.set_draw_motion_vectors(False)
        # if (app.bool_use_alpha.get()==True):
        #     self.graphics_options.set_use_alpha(True)
        # else:self.graphics_options.set_use_alpha(False)
        # if (app.bool_draw_joint_frame.get()==True):
        #     self.graphics_options.draw_joint_frame(True)
        # else:self.graphics_options.draw_joint_frame(False)
        # if (app.bool_draw_all_joint_frames.get()==True):
        #     self.graphics_options.set_draw_wireframe(True)
        # else:self.graphics_options.set_draw_wireframe(False)

        self.graphics_options.set_use_alpha(True)
        self.robot_graphics.render(self.kin_des, self.graphics_options)
        self.graphics_options.set_use_alpha(False)

        # glPushMatrix()
        # self.renderText(0.0, 0.0, 0.0, 'text!')
        # glPopMatrix()

    def initGeometry(self):
        pass

    def spin(self):
        self.spin_time_prev = self.spin_time
        self.spin_time = time.time()
        self.dt = self.spin_time - self.spin_time_prev
        self.sim_time += self.dt
        self.fps = 1.0 / self.dt

        # Compute joint trajectories:#hieu
        # amp = 0.5
        # freq = 0.1
        # for i in range(self.robot.n_dofs):
        #     self.robot_state.joint_state_des[i].th = amp*np.sin(2.0*np.pi*freq*self.sim_time)
        #     self.robot_state.joint_state_des[i].thd = 2.0*np.pi*freq*amp*np.cos(2.0*np.pi*freq*self.sim_time)
        #     self.robot_state.joint_state_des[i].thdd = -2.0*np.pi*freq*2.0*np.pi*freq*amp*np.sin(2.0*np.pi*freq*self.sim_time)
        #     print(self.robot_state.joint_state_des[i].th)
        # app.operation_main()
        # Update kinematics from the current joint state:
        self.kin.update(tf.HomogeneousTransform(), np.zeros(6), np.zeros(6), self.robot_state.joint_state, self.dt)
        # Update desired kinematics:
        self.kin_des.update(tf.HomogeneousTransform(), np.zeros(6), np.zeros(6), self.robot_state.joint_state_des, self.dt)

        # Compute joint state from differential IK:
        amp = 0.1
        freq = 0.2
        self.robot_state.joint_state_des = self.ik.update(self.kin_des,
                                                          self.initial_endeff_pos +
                                                          np.array([amp * np.cos(2.0 * np.pi * freq * self.sim_time), 0.0, amp * np.sin(2.0 * np.pi * freq * self.sim_time)]),
                                                          self.initial_endeff_ori,
                                                          np.array([-2.0 * np.pi * freq * amp * np.sin(2.0 * np.pi * freq * self.sim_time), 0.0, 2.0 * np.pi * freq * amp * np.cos(2.0 * np.pi * freq * self.sim_time)]),
                                                          np.zeros(3),
                                                          self.dt)

        # Update dynamics quantities before calling class methods:
        self.dyn.update(self.robot_state.joint_state)

        if self.controller_type == robot_defs.CONTROLLER_TYPE_PD:
            # PD controller:
            for i in range(self.robot.n_dofs):
                self.robot_state.joint_state_des[i].u = 0.0 * (
                        self.robot_state.joint_state_des[i].th - self.robot_state.joint_state[i].th) + \
                                                        0.0 * (self.robot_state.joint_state_des[i].thd - self.robot_state.joint_state[i].thd)

        elif self.controller_type == robot_defs.CONTROLLER_TYPE_GRAV_COMP_PD:
            # Gravity Compensation + PD Controller:
            tau_grav_comp = self.dyn.compute_grav_comp(self.robot_state.joint_state_des)
            for i in range(self.robot.n_dofs):
                self.robot_state.joint_state_des[i].u = tau_grav_comp[i] + 5000.0 * (
                        self.robot_state.joint_state_des[i].th - self.robot_state.joint_state[i].th) + \
                                                        200.0 * (self.robot_state.joint_state_des[i].thd - self.robot_state.joint_state[i].thd)


        # Inverse dynamics (using qddot) + PD Controller:
        elif self.controller_type == robot_defs.CONTROLLER_TYPE_INVDYN_PD:
            tau_invdyn = self.dyn.compute_joint_torques(np.zeros(6),
                                                        np.array([0.0, 0.0, 9.81, 0.0, 0.0, 0.0]),
                                                        self.robot_state.joint_state_des)
            for i in range(self.robot.n_dofs):
                self.robot_state.joint_state_des[i].u = tau_invdyn[i] + 5000.0 * (
                        self.robot_state.joint_state_des[i].th - self.robot_state.joint_state[i].th) + \
                                                        200.0 * (self.robot_state.joint_state_des[i].thd - self.robot_state.joint_state[i].thd)

        elif self.controller_type == robot_defs.CONTROLLER_TYPE_NONE:
            for i in range(self.robot.n_dofs):
                self.robot_state.joint_state_des[i].u = 0.0

        else:
            print('WARNING >> Invalid controller type.')
            for i in range(self.robot.n_dofs):
                self.robot_state.joint_state_des[i].u = 0.0
        # print(self.controller_type)#hieu
        # Check for contacts with the ground and apply reaction forces if necessary:
        for i, joint in enumerate(self.robot_state.joint_state):
            if self.kin.h_tf[i].t()[2] < 0.0:
                joint.fext = np.array([0.0,
                                       0.0,
                                       (robot_defs.FLOOR_STIFF * self.kin.h_tf[i].t()[2] +
                                        -robot_defs.FLOOR_DAMP * self.kin.link_vel[i][2])])
            else:
                joint.fext = np.zeros(3)

        self.robot_state.joint_state = self.integrate_dynamics(self.robot_state.joint_state,
                                                               self.robot_state.joint_state_des, self.dt)

        # self.parent.statusBar().showMessage('sim_freq: ' + str(self.fps))

        self.render_time = time.time()
        if ((self.render_time - self.render_time_prev) >= (1.0 / self.render_freq)):
            # self.update()
            self.render_time_prev = self.render_time

    def integrate_dynamics(self, joint_state, joint_state_des, dt):
        """ Integrates the dynamics forward in joint space by dt. """
        # print(robot_defs.INTEGRATOR_TYPE_EULER)
        # self.integrator_type=0
        # print(joint_state_des)
        # Create a deep copy of the joint state to modify for integration:
        jstate = copy.deepcopy(joint_state)

        # Update the joint state by computing the joint-space dynamics and integrating:
        if self.integrator_type == robot_defs.INTEGRATOR_TYPE_EULER:
            qddot = self.dyn.compute_qddot(joint_state_des)
            for i in range(self.robot.n_dofs):
                jstate[i].thdd = qddot[i]
                jstate[i].thd = joint_state[i].thd + self.dt * jstate[i].thdd
                jstate[i].th = jstate[i].th + self.dt * jstate[i].thd

        elif self.integrator_type == robot_defs.INTEGRATOR_TYPE_RK4:
            k1 = np.zeros(2 * self.robot.n_dofs)
            k2 = np.zeros(2 * self.robot.n_dofs)
            k3 = np.zeros(2 * self.robot.n_dofs)
            k4 = np.zeros(2 * self.robot.n_dofs)

            # k1 = h * f(tn, yn):
            k1[:self.robot.n_dofs] = dt * np.array([j.thd for j in jstate])
            k1[self.robot.n_dofs:] = dt * self.dyn.compute_qddot(joint_state_des)

            # k2 = h * f(tn + h/2, yn + k1/2):
            for i in range(self.robot.n_dofs):
                jstate[i].th = joint_state[i].th + 0.5 * k1[i]
                jstate[i].thd = joint_state[i].thd + 0.5 * k1[i + self.robot.n_dofs]
            k2[:self.robot.n_dofs] = dt * np.array([j.thd for j in jstate])
            k2[self.robot.n_dofs:] = dt * self.dyn.compute_qddot(joint_state_des)

            # k3 = h * f(tn + h/2, yn + k2/2):
            for i in range(self.robot.n_dofs):
                jstate[i].th = joint_state[i].th + 0.5 * k2[i]
                jstate[i].thd = joint_state[i].thd + 0.5 * k2[i + self.robot.n_dofs]
            k3[:self.robot.n_dofs] = dt * np.array([j.thd for j in jstate])
            k3[self.robot.n_dofs:] = dt * self.dyn.compute_qddot(joint_state_des)

            # k4 = h * f(tn + h/2, yn + k3):
            for i in range(self.robot.n_dofs):
                jstate[i].th = joint_state[i].th + k3[i]
                jstate[i].thd = joint_state[i].thd + k3[i + self.robot.n_dofs]
            k4[:self.robot.n_dofs] = dt * np.array([j.thd for j in jstate])
            k4[self.robot.n_dofs:] = dt * self.dyn.compute_qddot(joint_state_des)

            # Finally, compute the joint state from RK4 intermediate variables:
            # y_{n+1} = y_{n} + (1/6)*(k1 + 2*k2 + 2*k3 + k4)
            for i in range(self.robot.n_dofs):
                jstate[i].th = joint_state[i].th + \
                               (1.0 / 6.0) * (k1[i] + 2 * k2[i] + 2 * k3[i] + k4[i])
                jstate[i].thd = joint_state[i].thd + \
                                (1.0 / 6.0) * (k1[i + self.robot.n_dofs] + 2 * k2[i + self.robot.n_dofs] + \
                                               2 * k3[i + self.robot.n_dofs] + k4[i + self.robot.n_dofs])
                jstate[i].thdd = (1.0 / dt) * k1[i + self.robot.n_dofs]

        elif self.integrator_type == robot_defs.INTEGRATOR_TYPE_NONE:
            pass  # returns the joint_state copy unmodified

        return jstate

    def redraw(self):

        self.robot_state.joint_state[0].set_th((2 * np.pi / 200) * app.val1.get())
        self.robot_state.joint_state[1].set_th((2 * np.pi / 200) * app.val2.get())
        self.robot_state.joint_state[2].set_th((2 * np.pi / 200) * app.val3.get())
        self.robot_state.joint_state[3].set_th((2 * np.pi / 200) * app.val4.get())
        self.robot_state.joint_state[4].set_th((2 * np.pi / 200) * app.val5.get())
        self.robot_state.joint_state[5].set_th((2 * np.pi / 200) * app.val6.get())
        self.robot_state.joint_state[6].set_th((2 * np.pi / 200) * app.val7.get())
        self.spin()
        self.paintGL()
        #################


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)

        self.IP_PLC = '192.168.0.10'
        self.PORT_PLC = 8501
        self.pack()
        # self.member1 = Winzoom()
        # self.root.iconbitmap('c:/gui/codemy.ico')
        # font
        self.my_font = ft.Font(self.master, family="System", size=14, weight="bold")
        self.sys_font = ft.Font(self.master, family="System", size=30, weight="bold")
        self.sysbol_font = ft.Font(self.master, family="System", size=45)
        self.small_font = ft.Font(self.master, family="System", size=6)
        #
        self.my_title = "ROBOT_ARM"  # タイトル
        self.back_color = "#DBFFFF"  # 背景色
        # self.pil_img = Image.open("label4.PNG")
        self.status = str("normal")
        style = ttk.Style(root)
        # ウィンドウの設定
        self.master.title(self.my_title)  # タイトル
        self.master.geometry("1080x740")  # サイズ
        # Import the tcl file
        # root.tk.call("source", "forest-light.tcl")

        # Set the theme with the theme_use method
        # style.theme_use("forest-light")
        # self.master.geometry("%dx%d" % (root.winfo_screenwidth(), root.winfo_screenheight()))
        # self.slave.geometry("400x400")
        ################################################################## ROBOT SIMULator

        self.robot_ = kin.RobotParams()
        self.robot_state_ = kin.RobotState(self.robot_)

        self.val1 = tk.IntVar()
        self.val2 = tk.IntVar()
        self.val3 = tk.IntVar()
        self.val4 = tk.IntVar()
        self.val5 = tk.IntVar()
        self.val6 = tk.IntVar()
        self.val7 = tk.IntVar()
        root.bind("<Key>", self.key_handler)

        self.val1.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[0].th)
        self.val2.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[1].th)
        self.val3.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[2].th)
        self.val4.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[3].th)
        self.val5.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[4].th)
        self.val6.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[5].th)
        self.val7.set((200 / (2 * np.pi)) * self.robot_state_.joint_state[6].th)

        ############################################################################

        self.point = 0
        self.i_couter = 0
        self.i_couter_1 = 0
        self.i_couter_2 = 0
        # Create workbook EXCEL instance
        # エクセルファイルのロード
        excel_path = 'data.xlsx'
        self.workbook = load_workbook(filename=excel_path, read_only=False)
        self.ws = self.workbook.active
        # シートのロード
        self.sheet = self.workbook['Sheet1']
        self.cam_edit_hani = False
        self.capture = cv2.VideoCapture(0)
        self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        # Scale（オプションをいくつか設定）
        self.loop_display = 0
        self.pil_image = None  # 表示する画像データ
        self.filename = None  # 最後に開いた画像ファイル名
        self.DM1000_var = tk.DoubleVar()
        self.DM1002_var = tk.DoubleVar()
        self.DM1004_var = tk.DoubleVar()
        self.DM1006_var = tk.DoubleVar()
        self.DM1008_var = tk.DoubleVar()
        self.DM1010_var = tk.DoubleVar()
        self.DM1012_var = tk.DoubleVar()
        self.DM1014_var = tk.DoubleVar()
        self.DM1016_var = tk.DoubleVar()
        self.DM1018_var = tk.DoubleVar()
        self.DM1020_var = tk.DoubleVar()
        self.now_area_check = tk.StringVar(value="00")
        self.now_chuvi_check = tk.StringVar(value="00")
        self.prty_val = 'N'
        self.create_menu()  # メニューの作成
        # self.create_treeview()
        self.creat_textbox()
        self.create_widget()  # ウィジェットの作成
        self.robot_variable()
        self.create_robot_program()
        self.create_robot_widget()
        self.robot_tab_0_widget()
        self.robot_tab_1_widget()
        self.robot_tab_2_widget()
        self.robot_tab_3_widget()
        self.robot_tab_4_widget()
        self.robot_Cam_Search_widget()
        self.create_log_page()
        self.create_camera_processing()
        self.count = 0
        self.record = 0
        self.Camera_check = 0
        self.disp_id = None
        self.MSP_id = None
        self.MSP_id_1 = None
        self.MSP_id_2 = None
        self.comport = None
        self.communicate_on = False
        self.have_contour = 0
        self.ser = Serial()
        self.ser.close()
        self.ser_IO = Serial()
        self.ser_IO.close()
        self.ser_msp = Serial()
        self.ser_msp.close()
        self.ser_msp_1 = Serial()
        self.ser_msp_1.close()
        self.ser_msp_2 = Serial()
        self.ser_msp_2.close()
        self.refresh_serial_ports()
        self.refresh_serial_M1_2()
        self.refresh_serial_M3_4()
        self.refresh_serial_M5_6()
        self.recv_id = None
        self.draw_id = None
        self.status = 0
        self.error_status = 0
        self.toggle_blink = 0
        self.syudou_kirikaeis_on = True
        self.ryru_hansoubelt3is_on = True
        self.ryru_hansoubelt4is_on = True
        self.ryru_checkis_on = True
        self.ryru_check1is_on = True
        self.ryru_check2is_on = True
        self.ryru_check9is_on = True
        self.ryru_check10is_on = True
        self.ryru_hariis_on = True
        self.ryru_hari1is_on = True
        self.ryru_hari2is_on = True
        self.ryru_hari3is_on = True
        self.ryru_hari4is_on = True
        self.ryru_hari5is_on = True
        self.ryru_hari6is_on = True
        self.ryru_hari7is_on = True
        self.ryru_hari8is_on = True
        #
        self.ryru_check_bar1is_on = True
        self.ryru_check_bar2is_on = True
        self.ryru_check_bar3is_on = True
        self.ryru_check_bar4is_on = True
        self.ryru_check_bar5is_on = True
        self.ryru_check_bar6is_on = True
        self.ryru_check_bar7is_on = True
        self.ryru_check_bar9is_on = True
        self.hakokyokyu2is_on = True
        self.hakokyokyu3is_on = True
        self.hakokyokyu4is_on = True
        self.hakokyokyu5is_on = True
        self.hakokyokyu6is_on = True
        self.hakokyokyu7is_on = True
        self.hakokyokyu8is_on = True
        self.hakokyokyu9is_on = True
        self.hakokyokyu10is_on = True
        self.hakokyokyu11_1is_on = True
        self.hakokyokyu11_2is_on = True
        self.hakokyokyu11_3is_on = True
        self.hakokyokyu12is_on = True
        self.hakokyokyu13is_on = True
        self.hakokyokyu17is_on = True
        self.hakokyokyu18is_on = True
        self.hakokyokyu19is_on = True
        self.hakokyokyu20is_on = True
        self.hakokyokyu21is_on = True
        self.hakokyokyu22is_on = True
        self.hakokyokyu23is_on = True
        self.hakokyokyu24is_on = True
        self.hakokyokyu27is_on = True
        self.hakokyokyu29is_on = True
        self.hakokyokyu30is_on = True
        self.hakokyokyu31is_on = True
        self.hakokyokyu33is_on = True
        self.hakokyokyu34is_on = True
        self.hakokyokyu35is_on = True
        self.print_riruumuis_on = True
        self.print_hakoumuis_on = True
        self.capture_switchis_on = False
        #
        self.time_lock = 0
        self.update_main_id = None
        self.RIRU_NUMBER = tk.DoubleVar()  ##ERROR STATUS
        self.KIKAI_var = tk.DoubleVar()  ##KIKAI STATUS
        self.FLAG_Barcode = tk.DoubleVar()  ##FLAG STATUS
        self.FLAG_Camera = tk.DoubleVar()  ##SENSOR STATUS
        self.YOBI_var = tk.DoubleVar()  ##YOBI STATUS
        self.print_data1_var = tk.DoubleVar()  ##print1 STATUS
        self.print_data2_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data3_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data4_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data5_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data6_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data7_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data8_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data9_var = tk.DoubleVar()  ##print2 STATUS
        self.send_flag = 0
        self.data_old = ""
        self.data_old1 = ""
        self.data_old2 = ""
        self.data_old_a = ""
        self.data_old_b = ""
        self.data_old_c = ""
        self.scanned = False
        self.barcoder_data = ""
        self.ONE_scan = False
        self.cv_image = None
        self.read_defaut = False
        self.zoom_load_defaut = False
        # CAMERA
        self.Camera_kido = 0
        self.mat_affine = np.eye(3)
        self.mat_affine_main = np.eye(3)
        self.mat_affine_sita = np.eye(3)
        self.mat_affine_hoshu1 = np.eye(3)
        self.mat_affine_hoshu2 = np.eye(3)
        self.pt1 = (20, 20)
        self.pt2 = (600, 400)
        self.edit_pt1 = False
        self.edit_pt2 = False

        self.pt1_specal = (100, 200)
        self.pt2_specal = (100, 300)
        self.pt3_specal = (200, 250)
        self.edit_pt1_specal = False
        self.edit_pt2_specal = False
        self.edit_pt3_specal = False
        self.pt1_kaku = (200, 200)
        self.pt2_kaku = (400, 200)
        self.pt3_kaku = (400, 400)
        self.pt4_kaku = (200, 400)
        self.edit_pt1_kaku = False
        self.edit_pt2_kaku = False
        self.edit_pt3_kaku = False
        self.edit_pt4_kaku = False
        self.pt1_polygol5 = (200, 200)
        self.pt2_polygol5 = (300, 100)
        self.pt3_polygol5 = (400, 200)
        self.pt4_polygol5 = (300, 300)
        self.pt5_polygol5 = (200, 400)
        self.edit_pt1_polygol5 = False
        self.edit_pt2_polygol5 = False
        self.edit_pt3_polygol5 = False
        self.edit_pt4_polygol5 = False
        self.edit_pt5_polygol5 = False

        self.so_fiter = 1
        self.setting_lock = 1
        self.tab_select = 0
        self.PLC_CN_OK = 0
        self.stt = 0
        self.last_day_on = '2024_07_25'
        self.hako_tori_kakunin = 0
        self.now_area = 0
        self.now_chuvi = 0

    # -------------------------------------------------------------------------------
    def day_log(self):

        self.day_month_year = datetime.datetime.now().strftime('%Y_%m_%d')
        if self.day_month_year != self.last_day_on:
            self.stt = self.stt + 1
            self.last_day_on = self.day_month_year
            self.write_data(1, 13, self.stt)
            self.write_data(2, 13, self.last_day_on)

        self.name_log = self.day_month_year + '_' + str(self.stt) + '.txt'
        print(self.name_log)

    def uint_to_float(self, uint32_value):
        packed = struct.pack('I', uint32_value)  # 'I' is the format for unsigned 32-bit integer
        float_value = struct.unpack('f', packed)[0]  # 'f' is the format for a 32-bit float
        return float_value

    # ------------------------------------SCREEN HAICHI-----------------------------------
    def menu_open_clicked(self, event=None):
        print("「ファイルを開く」が選択された")
        self.filename = fd.askopenfilename(
            filetypes=[("file", ".xlsx .xlsm .xltx .xltm")],  # ファイルフィルタ
            title="ファイルを開く",
            initialdir="./"  # 自分自身のディレクトリ
        )
        self.excel_path = self.filename
        self.workbook = load_workbook(filename=self.excel_path, read_only=False)
        self.ws = self.workbook.active
        # シートのロード
        self.sheet = self.workbook['Sheet1']

        print(self.filename)
        # print(self.read_data(1,1))
        ###  LOAD DATA  TAB 1 ####################
        '''self.resolution_h_var.set(self.read_data(1, 1))
        self.resolution_w_var.set(self.read_data(2, 1))
        self.preview_check_var.set(self.read_data(3, 1))
        self.hdr_var.set(self.read_data(4, 1))
        self.saturation_var.set(self.read_data(5, 1))
        self.contrast_var.set(self.read_data(6, 1))
        self.sharpness_var.set(self.read_data(7, 1))
        self.brightness_var.set(self.read_data(8, 1))
        # hdr
        self.num_hdr_var.set(self.read_data(9, 1, ))
        self.stops_hdr_above_var.set(self.read_data(10, 1))
        self.stops_hdr_below_var.set(self.read_data(11, 1))
        self.hdr_gamma_var.set(self.read_data(12, 1))'''

        #

    def menu_reload_clicked(self, event=None):
        self.status = str("disabled")

        ''''''

    def menu_quit_clicked(self):
        # ウィンドウを閉じる
        self.master.destroy()

        # -------------------------------------------------------------------------------

    def creat_textbox(self):
        '''lbl_sock = tk.Label(self.master, text="PLC socket")
        lbl_sock.place(x=5, y=250)
        lbl_comp = tk.Label(self.master, text="Serial")
        lbl_comp.place(x=5, y=420)
        self.Text_sock_recieved = Text(self.master, height=11, width=28, background='#856ff8')
        self.Text_sock_recieved.place(x=0, y=270)
        self.received_text = tk.Text(self.master, height=11, width=28, background='#856ff8')
        self.received_text.place(x=0, y=440)
        clear_button = tk.Button(self.master, text="履歴クリア", command=self.clear_received_text)
        clear_button.place(x=137, y=560)'''

    def create_treeview(self):
        left_frame = tk.Frame(self.master, relief=tk.SUNKEN, bd=2, width=200, height=500)
        left_frame.pack(side=tk.LEFT, anchor=NW, fill=tk.X)
        # Creating treeview window
        '''self.treeview = ttk.Treeview(left_frame)

        # Calling pack method on the treeview
        self.treeview.pack(side=tk.RIGHT)
        self.treeview.insert('', '0', 'item1',
                             text='Motor parameter', values=1)

        # Inserting child
        self.treeview.insert('', '1', 'item2',
                             text='Motor_asix1', values=2)
        self.treeview.insert('', '2', 'item3',
                             text='Motor_asix2', values=3)
        self.treeview.insert('', 'end', 'item4',
                             text='Motor_asix3', values=4)

        # Inserting more than one attribute of an item
        self.treeview.insert('item2', 'end', 'Orignal',
                             text='Vol input asix1', values=5)
        self.treeview.insert('item2', 'end', 'fix_pixel',
                             text='基板_Area', values=6)
        self.treeview.insert('item2', 'end', 'fix_point',
                             text='記号_Area', values=7)
        self.treeview.insert('item2', 'end', 'fix_grey_pixel',
                             text='C_面_Area', values=8)
        self.treeview.insert('item3', 'end', 'point check1',
                             text='Vol input asix2', values=9)
        self.treeview.insert('item3', 'end', 'point check2',
                             text='point check　●', values=10)
        self.treeview.insert('item3', 'end', 'point check3',
                             text='point check　▲', values=11)
        self.treeview.insert('item3', 'end', 'point check4',
                             text='point check　C面', values=12)
        self.treeview.insert('item4', 'end', 'Check all Raw_Colum',
                             text='Check all Raw Colum', values=13)
        self.treeview.insert('item4', 'end', 'Check center',
                             text='Check all center', values=14)
        self.treeview.insert('item4', 'end', 'Check パタン',
                             text='Check パタン', values=15)

        # Placing each child items in parent widget
        self.treeview.move('item2', 'item1', 'end')
        self.treeview.move('item3', 'item1', 'end')
        self.treeview.move('item4', 'item1', 'end')
        self.treeview.bind('<<TreeviewSelect>>', self.item_selected)'''

    def item_selected(self, event):
        for selected_item in self.treeview.selection():
            item = self.treeview.item(selected_item)
            self.record = item['values']
            # show a message
            # print(self.record)

    def create_menu(self):
        self.menu_bar = tk.Menu(self)  # Menuクラスからmenu_barインスタンスを生成

        self.file_menu = tk.Menu(self.menu_bar, tearoff=tk.OFF)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        self.file_menu.add_command(label="Open", command=self.menu_open_clicked, accelerator="Ctrl+O")
        self.file_menu.add_command(label="ReLoad_data", command=self.auto_load)
        self.file_menu.add_command(label="Save_data", command=self.save_data)
        self.file_menu.add_separator()  # セパレーターを追加
        self.file_menu.add_command(label="Exit", command=self.menu_quit_clicked)

        self.menu_bar.bind_all("<Control-o>", self.menu_open_clicked)  # ファイルを開くのショートカット(Ctrol-Oボタン)

        self.master.config(menu=self.menu_bar)  # メニューバーの配置

    def create_widget(self):
        # Variable
        font = ('Yu Gothic UI', 11)
        style = ttk.Style()
        style.configure('TNotebook.Tab', font=font, padding=(8, 12))
        self.text_var = tk.StringVar(value="A")
        self.error_code = tk.IntVar()
        self.error_code.set(0)
        self.riru_ok_var = tk.StringVar(value="E")
        self.riru_ng_var = tk.StringVar(value="D")
        self.hako_ok_var = tk.StringVar(value="C")
        self.lot_number = tk.StringVar(value="B")
        self.BAR_check = tk.StringVar(value="F")
        self.IP_PLC = tk.StringVar()
        self.IP_PLC.set("192.168.0.10")
        self.PORT_PLC = tk.IntVar()
        self.PORT_PLC.set(8501)
        self.port_var = tk.StringVar(root)
        self.baudrate_var = tk.StringVar(root)
        self.baudrate_var.set(BAUDRATES[4])
        self.delimiter_var = tk.StringVar(root)
        self.delimiter_var.set(REV_DELIMITER['\r'])
        ##msp
        self.port_var_msp = tk.StringVar(root)
        self.baudrate_var_msp = tk.StringVar(root)
        self.baudrate_var_msp.set(BAUDRATES[4])
        self.delimiter_var_msp = tk.StringVar(root)
        self.delimiter_var_msp.set(REV_DELIMITER['\r'])
        ##msp1
        self.port_var_msp_1 = tk.StringVar(root)
        self.baudrate_var_msp_1 = tk.StringVar(root)
        self.baudrate_var_msp_1.set(BAUDRATES[4])
        self.delimiter_var_msp_1 = tk.StringVar(root)
        self.delimiter_var_msp_1.set(REV_DELIMITER['\r'])
        ##msp2
        self.port_var_msp_2 = tk.StringVar(root)
        self.baudrate_var_msp_2 = tk.StringVar(root)
        self.baudrate_var_msp_2.set(BAUDRATES[4])
        self.delimiter_var_msp_2 = tk.StringVar(root)
        self.delimiter_var_msp_2.set(REV_DELIMITER['\r'])
        #
        self.mode_run_var = tk.StringVar(root)
        self.mode_run_var.set(MODE_RUN[2])
        self.num_motor_var = tk.StringVar(root)
        self.num_motor_var.set(NUM_MOTOR[1])
        #
        self.mode_run_var_1 = tk.StringVar(root)
        self.mode_run_var_1.set(MODE_RUN[2])
        self.num_motor_var_1 = tk.StringVar(root)
        self.num_motor_var_1.set(NUM_MOTOR[1])
        #
        self.mode_run_var_2 = tk.StringVar(root)
        self.mode_run_var_2.set(MODE_RUN[2])
        self.num_motor_var_2 = tk.StringVar(root)
        self.num_motor_var_2.set(NUM_MOTOR[1])
        #
        self.text_var1 = tk.StringVar(value="000")
        self.text_var2 = tk.StringVar(value="000")
        self.DM1000_var.set(0)
        self.DM1002_var.set(0)
        self.riru_henko = tk.IntVar()
        self.FRAME_WIDTH_var = tk.DoubleVar()
        self.FRAME_HEIGHT_var = tk.DoubleVar()
        self.FPS_var = tk.DoubleVar()
        self.BRIGHTNESS_var = tk.DoubleVar()
        self.CONTRAST_var = tk.DoubleVar()
        self.SATURATION_var = tk.DoubleVar()
        self.HUE_var = tk.DoubleVar()
        self.GAIN_var = tk.DoubleVar()
        self.EXPOSE_var = tk.DoubleVar()
        self.auto_exposure_var = tk.IntVar()
        self.auto_exposure_var.set(0)
        self.auto_focus_var = tk.IntVar()
        self.auto_focus_var.set(0)
        self.canny1_var = tk.DoubleVar()
        self.canny1_var.set(100)
        self.canny2_var = tk.DoubleVar()
        self.canny2_var.set(500)
        self.gaussion_var = tk.DoubleVar()
        self.Blur_var = tk.DoubleVar()
        self.bold_var = tk.IntVar()
        self.debold_var = tk.IntVar()
        self.bold_var1 = tk.IntVar()
        self.debold_var1 = tk.IntVar()
        self.filtermap = tk.StringVar()
        self.filtermap1 = tk.StringVar()
        self.filtermap2 = tk.StringVar()
        self.filtermap3 = tk.StringVar()
        self.filtermap4 = tk.StringVar()
        self.check_hani_ok = tk.IntVar()

        self.check_patan1 = tk.IntVar()
        self.check_patan2 = tk.IntVar()
        self.check_patan3 = tk.IntVar()
        self.check_patan4 = tk.IntVar()
        self.check_patan5 = tk.IntVar()
        self.check_patan6 = tk.IntVar()
        self.check_hani_ok.set(1)
        self.colormap = tk.StringVar()
        self.boder_filter_map_sl = tk.StringVar()
        self.scale_var = tk.DoubleVar()
        self.scale_var.set(50)
        self.adaptiveThreshold_var = tk.DoubleVar()
        self.boder_filter_map_sl1 = tk.StringVar()
        #####################################################
        # ステータスバー相当(親に追加)
        self.statusbar = tk.Frame(self.master)
        self.mouse_position = tk.Label(self.statusbar, relief=tk.SUNKEN, text="mouse position")  # マウスの座標
        self.image_position = tk.Label(self.statusbar, relief=tk.SUNKEN, text="image position")  # 画像の座標
        self.label_space = tk.Label(self.statusbar, relief=tk.SUNKEN)  # 隙間を埋めるだけ
        self.image_info = tk.Label(self.statusbar, relief=tk.SUNKEN, text="info")  # 画像情報
        self.mouse_position.pack(side=tk.LEFT)
        self.image_position.pack(side=tk.LEFT)
        self.label_space.pack(side=tk.LEFT, expand=True, fill=tk.X)
        self.image_info.pack(side=tk.RIGHT)
        self.statusbar.pack(side=tk.BOTTOM, fill=tk.X)
        #####################WINDOW2##################################
        # ButtonWINDOW2

        #####################################################
        # 右側フレーム（画像処理用ボタン配置用）
        right_frame = tk.Frame(self.master, relief=tk.SUNKEN, bd=2, width=1080)
        right_frame.pack(side=tk.RIGHT, fill=tk.Y)

        self.tab_control = ttk.Notebook(right_frame, height=768, width=1080)
        self.tab0 = ttk.Frame(self.tab_control)
        self.tab1 = ttk.Frame(self.tab_control)
        self.tab10 = ttk.Frame(self.tab_control)
        self.tab2 = ttk.Frame(self.tab_control)
        self.tab3 = ttk.Frame(self.tab_control)
        self.tab4 = ttk.Frame(self.tab_control)
        self.tab5 = ttk.Frame(self.tab_control)
        self.tab6 = ttk.Frame(self.tab_control)
        self.tab7 = ttk.Frame(self.tab_control)
        self.tab9 = ttk.Frame(self.tab_control)
        self.tab8 = ttk.Frame(self.tab_control)

        self.tab_control.add(self.tab0, state=self.status, text='メイン')
        self.tab_control.add(self.tab1, state=self.status, text='ドライバー設定')
        self.tab_control.add(self.tab10, state=self.status, text='ロボット_PG')
        self.tab_control.add(self.tab2, state=self.status, text='ロボット')
        self.tab_control.add(self.tab3, state=self.status, text='保守２')
        self.tab_control.add(self.tab4, state=self.status, text='設定１')
        self.tab_control.add(self.tab5, state=self.status, text='設定２')
        self.tab_control.add(self.tab6, state=self.status, text='通信')
        self.tab_control.add(self.tab7, state=self.status, text='カメラ1')
        self.tab_control.add(self.tab9, state=self.status, text='カメラ2')
        self.tab_control.add(self.tab8, state=self.status, text='点検')
        self.tab_control.place(x=0, y=0)
        self.tab_control.bind("<<NotebookTabChanged>>", self.gettab_select)

        # TAB 0======================================================================

        lbl_error = tk.Label(self.tab0, text="EEROR")
        lbl_error.place(x=505, y=0)

        self.syudou_kirikae = tk.Button(right_frame, command=self.syudou_kirikae_switch, height=2, width=11, text="BT_3", bg="gray")
        self.karabako_okuri = tk.Button(right_frame, command=lambda: self.Socket_send("ST R3011"), height=2, width=11, text="BT_2", bg="#fdd2b9")
        self.riru_haraidasi = tk.Button(right_frame, command=lambda: self.Socket_send("ST R3012"), height=2, width=11, text="BT_1", bg="#faa755")
        self.syudou_kirikae.place(x=916, y=4)
        self.karabako_okuri.place(x=810, y=4)
        #self.riru_haraidasi.place(x=702, y=4)

        self.frm = ROBOT_GRAPHIC(self.tab0, height=500, width=500)
        self.frm.animate = 10
        self.frm.place(x=0, y=0)
        ####
        self.canvas = tk.Canvas(self.tab0, background="#C3C2C4", width=500, height=340)  #
        self.canvas_sita = tk.Canvas(self.tab0, background="#C3C2C4", width=500, height=330)  #
        self.canvas_hoshu1 = tk.Canvas(self.tab2, background="#C3C2C4", width=1020, height=630)  #
        self.canvas_hoshu2 = tk.Canvas(self.tab3, background="#C3C2C4", width=1020, height=630)  #
        # self.canvas.place(x=0, y=0)
        self.canvas_sita.place(x=0, y=340)
        # self.canvas_hoshu1.place(x=0, y=0)
        self.canvas_hoshu2.place(x=0, y=0)

        self.riru_number_couter_lb = tk.LabelFrame(self.tab0, text="GROUP0", labelanchor="nw", width=100, height=130, font=self.small_font)
        self.riru_number_couter_lb.propagate(False)
        self.PLC_error_text = scrolledtext.ScrolledText(self.tab0, wrap=tk.WORD, width=27, height=8, font=("Arial", 19))
        self.PLC_error_annai = scrolledtext.ScrolledText(self.tab0, wrap=tk.WORD, width=27, height=7, font=("Arial", 19))
        self.lbl_riru_number = tk.Label(self.riru_number_couter_lb, textvariable=self.text_var, font=("Arial", 60, "bold"), bg="yellow")
        self.clear_error = tk.Button(self.tab0, text="履歴クリア", command=self.clear_error_text, width=16, height=2, bg="gray")
        lbl_code = Label(self.tab0, text="エラーCODE")
        txt_address = tk.Entry(self.tab0, justify=tk.RIGHT, width=15, textvariable=self.error_code)
        self.start_button = tk.Button(self.tab0, text="START", height=3, width=6, bd=6, bg='gray', font=self.sys_font, fg="lightgreen", command=lambda: self.Socket_send("ST R3000"))
        self.stop_button = tk.Button(self.tab0, text="STOP", height=3, width=6, bd=6, bg='gray', font=self.sys_font, fg="red", command=lambda: self.Socket_send("ST R3001"))
        self.buzz_button = tk.Button(self.tab0, wraplength=90, text="LOT RESET", height=3, width=6, bd=6, bg='gray', font=self.sys_font, fg="blue", command=lambda: self.Socket_send("ST R3005"))
        self.reset_button = tk.Button(self.tab0, text="RESET", height=3, width=6, bd=6, bg='gray', font=self.sys_font, fg="yellow", command=self.reset_kikai)
        # position
        self.riru_number_couter_lb.place(x=910, y=320)
        self.PLC_error_text.place(x=505, y=20)
        self.PLC_error_annai.place(x=505, y=250)
        self.lbl_riru_number.place(x=16, y=3)
        self.clear_error.place(x=872, y=450)
        lbl_code.place(x=580, y=460)
        txt_address.place(x=650, y=460)
        self.start_button.place(x=510, y=490)
        self.stop_button.place(x=630, y=490)
        self.buzz_button.place(x=750, y=490)
        self.reset_button.place(x=870, y=490)

        self.labelhonsu1 = tk.LabelFrame(self.tab0, text="GROUP3", labelanchor="nw", width=100, height=80)
        self.labelhonsu2 = tk.LabelFrame(self.tab0, text="GROUP5", labelanchor="nw", width=100, height=80)
        self.labelhonsu3 = tk.LabelFrame(self.tab0, text="GROUP4", labelanchor="nw", width=100, height=80)
        self.labelhonsu4 = tk.LabelFrame(self.tab0, text="GROUP6", labelanchor="nw", width=100, height=80)
        self.ARM1_2 = tk.LabelFrame(self.tab1, text="SETTING_ARM_1_2", labelanchor="nw", width=1000, height=200, font=self.small_font)
        self.ARM3_4 = tk.LabelFrame(self.tab1, text="SETTING_ARM_3_4", labelanchor="nw", width=1000, height=200, font=self.small_font)
        self.ARM5_6 = tk.LabelFrame(self.tab1, text="SETTING_ARM_5_6", labelanchor="nw", width=1000, height=200, font=self.small_font)
        self.labelhonsu1.propagate(False)
        self.labelhonsu2.propagate(False)
        self.labelhonsu3.propagate(False)
        self.labelhonsu4.propagate(False)
        self.ARM1_2.propagate(False)
        self.ARM3_4.propagate(False)
        self.ARM5_6.propagate(False)
        #########################################
        self.para_edit_1_is_on = False
        self.para_edit_2_is_on = False
        self.para_edit_3_is_on = False
        self.pid_vp_1 = tk.DoubleVar()
        self.pid_vp_1.set(0.00)
        self.pid_pp_1 = tk.DoubleVar()
        self.pid_pp_1.set(0.00)
        self.pid_vi_1 = tk.DoubleVar()
        self.pid_vi_1.set(0.00)
        self.pid_vd_1 = tk.DoubleVar()
        self.pid_vd_1.set(0.00)
        self.volin_1 = tk.DoubleVar()
        self.volin_1.set(0.00)
        self.vollimit_1 = tk.DoubleVar()
        self.vollimit_1.set(0.00)
        self.pid_vp_2 = tk.DoubleVar()
        self.pid_vp_2.set(0.00)
        self.pid_pp_2 = tk.DoubleVar()
        self.pid_pp_2.set(0.00)
        self.pid_vi_2 = tk.DoubleVar()
        self.pid_vi_2.set(0.00)
        self.pid_vd_2 = tk.DoubleVar()
        self.pid_vd_2.set(0.00)
        self.volin_2 = tk.DoubleVar()
        self.volin_2.set(0.00)
        self.vollimit_2 = tk.DoubleVar()
        self.vollimit_2.set(0.00)
        self.pid_vp_3 = tk.DoubleVar()
        self.pid_vp_3.set(0.00)
        self.pid_pp_3 = tk.DoubleVar()
        self.pid_pp_3.set(0.00)
        self.pid_vi_3 = tk.DoubleVar()
        self.pid_vi_3.set(0.00)
        self.pid_vd_3 = tk.DoubleVar()
        self.pid_vd_3.set(0.00)
        self.volin_3 = tk.DoubleVar()
        self.volin_3.set(0.00)
        self.vollimit_3 = tk.DoubleVar()
        self.vollimit_3.set(0.00)
        self.pid_vp_4 = tk.DoubleVar()
        self.pid_vp_4.set(0.00)
        self.pid_pp_4 = tk.DoubleVar()
        self.pid_pp_4.set(0.00)
        self.pid_vi_4 = tk.DoubleVar()
        self.pid_vi_4.set(0.00)
        self.pid_vd_4 = tk.DoubleVar()
        self.pid_vd_4.set(0.00)
        self.volin_4 = tk.DoubleVar()
        self.volin_4.set(0.00)
        self.vollimit_4 = tk.DoubleVar()
        self.vollimit_4.set(0.00)
        self.pid_vp_5 = tk.DoubleVar()
        self.pid_vp_5.set(0.00)
        self.pid_pp_5 = tk.DoubleVar()
        self.pid_pp_5.set(0.00)
        self.pid_vi_5 = tk.DoubleVar()
        self.pid_vi_5.set(0.00)
        self.pid_vd_5 = tk.DoubleVar()
        self.pid_vd_5.set(0.00)
        self.volin_5 = tk.DoubleVar()
        self.volin_5.set(0.00)
        self.vollimit_5 = tk.DoubleVar()
        self.vollimit_5.set(0.00)
        self.pid_vp_6 = tk.DoubleVar()
        self.pid_vp_6.set(0.00)
        self.pid_pp_6 = tk.DoubleVar()
        self.pid_pp_6.set(0.00)
        self.pid_vi_6 = tk.DoubleVar()
        self.pid_vi_6.set(0.00)
        self.pid_vd_6 = tk.DoubleVar()
        self.pid_vd_6.set(0.00)
        self.volin_6 = tk.DoubleVar()
        self.volin_6.set(0.00)
        self.vollimit_6 = tk.DoubleVar()
        self.vollimit_6.set(0.00)
        ############################### SETTING_ARM_1_2 ###############################
        lbl_mode = Label(self.ARM1_2, text="MODE_選択")
        self.mode_menu = tk.OptionMenu(self.ARM1_2, self.mode_run_var, *MODE_RUN)  # self.mode_run_var/self.num_motor_var
        lbl_num_motor = Label(self.ARM1_2, text="NUM_MOTOR_選択")
        self.num_motor = tk.OptionMenu(self.ARM1_2, self.num_motor_var, *NUM_MOTOR)
        self.toggle_servo_1 = tk.Button(self.ARM1_2, height=1, width=10, text="SERVO_1(Off)", background="#FFC652")
        self.toggle_servo_2 = tk.Button(self.ARM1_2, height=1, width=10, text="SERVO_2(Off)", background="#FFC652")
        self.group_pid_1 = tk.LabelFrame(self.ARM1_2, text="arm_1_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_1.propagate(False)
        lbl_pid_p1 = Label(self.group_pid_1, text="PID_position:")
        self.spinbox_PID_PP_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_pp_1)
        lbl_pid_vp1 = Label(self.group_pid_1, text="PID_(P)velocity:")
        self.spinbox_PID_VP_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vp_1)
        lbl_vol_in1 = Label(self.group_pid_1, text="VOL_input:")
        self.spinbox_VOL_IN1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8, textvariable=self.volin_1)
        lbl_pid_vi1 = Label(self.group_pid_1, text="PID_(I)velocity:")
        self.spinbox_PID_VI_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vi_1)
        lbl_vol_limit1 = Label(self.group_pid_1, text="VOL_limit:")
        self.spinbox_VOL_LIMIT1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8, textvariable=self.vollimit_1)
        lbl_pid_vd1 = Label(self.group_pid_1, text="PID_(D)velocity:")
        self.spinbox_PID_VD_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.001, width=8, textvariable=self.pid_vd_1)
        #
        self.group_pid_2 = tk.LabelFrame(self.ARM1_2, text="arm_2_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_2.propagate(False)
        lbl_pid_p2 = Label(self.group_pid_2, text="PID_position:")
        self.spinbox_PID_PP_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_pp_2)
        lbl_pid_vp2 = Label(self.group_pid_2, text="PID_(P)velocity:")
        self.spinbox_PID_VP_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vp_2)
        lbl_vol_in2 = Label(self.group_pid_2, text="VOL_input:")
        self.spinbox_VOL_IN2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8, textvariable=self.volin_2)
        lbl_pid_vi2 = Label(self.group_pid_2, text="PID_(I)velocity:")
        self.spinbox_PID_VI_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vi_2)
        lbl_vol_limit2 = Label(self.group_pid_2, text="VOL_limit:")
        self.spinbox_VOL_LIMIT2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8, textvariable=self.vollimit_2)
        lbl_pid_vd2 = Label(self.group_pid_2, text="PID_(D)velocity:")
        self.spinbox_PID_VD_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.001, width=8, textvariable=self.pid_vd_2)

        self.save_para_m1 = tk.Button(self.ARM1_2, height=1, width=6, text="Save M1", background="#AD8EE3", command=self.save_para_1)
        self.save_para_m2 = tk.Button(self.ARM1_2, height=1, width=6, text="Save M2", background="#AD8EE3", command=self.save_para_2)
        self.edit_para_m1_2 = tk.Button(self.ARM1_2, height=1, width=4, text="Edit", background="#AD8EE3", command=self.edit_1_switch)
        # 1
        lbl_mode.place(x=560, y=-5)
        self.mode_menu.place(x=640, y=-10)
        lbl_num_motor.place(x=760, y=-5)
        self.num_motor.place(x=870, y=-10)
        self.toggle_servo_1.place(x=150, y=-5)
        self.toggle_servo_2.place(x=250, y=-5)
        self.group_pid_1.place(x=450, y=20)
        lbl_pid_p1.place(x=0, y=0)
        self.spinbox_PID_PP_1.place(x=85, y=0)
        lbl_pid_vp1.place(x=0, y=30)
        self.spinbox_PID_VP_1.place(x=85, y=30)
        lbl_vol_in1.place(x=150, y=0)
        self.spinbox_VOL_IN1.place(x=235, y=0)
        lbl_pid_vi1.place(x=150, y=30)
        self.spinbox_PID_VI_1.place(x=235, y=30)
        lbl_vol_limit1.place(x=300, y=0)
        self.spinbox_VOL_LIMIT1.place(x=385, y=0)
        lbl_pid_vd1.place(x=300, y=30)
        self.spinbox_PID_VD_1.place(x=385, y=30)
        # 2
        self.group_pid_2.place(x=450, y=100)
        lbl_pid_p2.place(x=0, y=0)
        self.spinbox_PID_PP_2.place(x=85, y=0)
        lbl_pid_vp2.place(x=0, y=30)
        self.spinbox_PID_VP_2.place(x=85, y=30)
        lbl_vol_in2.place(x=150, y=0)
        self.spinbox_VOL_IN2.place(x=235, y=0)
        lbl_pid_vi2.place(x=150, y=30)
        self.spinbox_PID_VI_2.place(x=235, y=30)
        lbl_vol_limit2.place(x=300, y=0)
        self.spinbox_VOL_LIMIT2.place(x=385, y=0)
        lbl_pid_vd2.place(x=300, y=30)
        self.spinbox_PID_VD_2.place(x=385, y=30)

        self.save_para_m1.place(x=941, y=120)
        self.save_para_m2.place(x=941, y=148)
        self.edit_para_m1_2.place(x=945, y=90)
        #################################################################################
        ############################### SETTING_ARM_3_4 ###############################
        lbl_mode_1 = Label(self.ARM3_4, text="MODE_選択")
        self.mode_menu_1 = tk.OptionMenu(self.ARM3_4, self.mode_run_var_1, *MODE_RUN)  # self.mode_run_var/self.num_motor_var
        lbl_num_motor_1 = Label(self.ARM3_4, text="NUM_MOTOR_選択")
        self.num_motor_1 = tk.OptionMenu(self.ARM3_4, self.num_motor_var_1, *NUM_MOTOR)
        self.toggle_servo_3 = tk.Button(self.ARM3_4, height=1, width=10, text="SERVO_3(Off)", background="#FFC652")
        self.toggle_servo_4 = tk.Button(self.ARM3_4, height=1, width=10, text="SERVO_4(Off)", background="#FFC652")
        self.group_pid_3 = tk.LabelFrame(self.ARM3_4, text="arm_3_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_3.propagate(False)
        lbl_pid_p3 = Label(self.group_pid_3, text="PID_position:")
        self.spinbox_PID_PP_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_pp_3)
        lbl_pid_vp3 = Label(self.group_pid_3, text="PID_(P)velocity:")
        self.spinbox_PID_VP_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vp_3)
        lbl_vol_in3 = Label(self.group_pid_3, text="VOL_input:")
        self.spinbox_VOL_IN3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8, textvariable=self.volin_3)
        lbl_pid_vi3 = Label(self.group_pid_3, text="PID_(I)velocity:")
        self.spinbox_PID_VI_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vi_3)
        lbl_vol_limit3 = Label(self.group_pid_3, text="VOL_limit:")
        self.spinbox_VOL_LIMIT3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8, textvariable=self.vollimit_3)
        lbl_pid_vd3 = Label(self.group_pid_3, text="PID_(D)velocity:")
        self.spinbox_PID_VD_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.001, width=8, textvariable=self.pid_vd_3)
        #
        self.group_pid_4 = tk.LabelFrame(self.ARM3_4, text="arm_4_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_4.propagate(False)
        lbl_pid_p4 = Label(self.group_pid_4, text="PID_position:")
        self.spinbox_PID_PP_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_pp_4)
        lbl_pid_vp4 = Label(self.group_pid_4, text="PID_(P)velocity:")
        self.spinbox_PID_VP_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vp_4)
        lbl_vol_in4 = Label(self.group_pid_4, text="VOL_input:")
        self.spinbox_VOL_IN4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8, textvariable=self.volin_4)
        lbl_pid_vi4 = Label(self.group_pid_4, text="PID_(I)velocity:")
        self.spinbox_PID_VI_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vi_4)
        lbl_vol_limit4 = Label(self.group_pid_4, text="VOL_limit:")
        self.spinbox_VOL_LIMIT4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8, textvariable=self.vollimit_4)
        lbl_pid_vd4 = Label(self.group_pid_4, text="PID_(D)velocity:")
        self.spinbox_PID_VD_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.001, width=8, textvariable=self.pid_vd_4)

        self.save_para_m3 = tk.Button(self.ARM3_4, height=1, width=6, text="Save M3", background="#AD8EE3", command=self.save_para_3)
        self.save_para_m4 = tk.Button(self.ARM3_4, height=1, width=6, text="Save M4", background="#AD8EE3", command=self.save_para_4)
        self.edit_para_m3_4 = tk.Button(self.ARM3_4, height=1, width=4, text="Edit", background="#AD8EE3", command=self.edit_2_switch)
        # 3
        lbl_mode_1.place(x=560, y=-5)
        self.mode_menu_1.place(x=640, y=-10)
        lbl_num_motor_1.place(x=760, y=-5)
        self.num_motor_1.place(x=870, y=-10)
        self.toggle_servo_3.place(x=150, y=-5)
        self.toggle_servo_4.place(x=250, y=-5)
        self.group_pid_3.place(x=450, y=20)
        lbl_pid_p3.place(x=0, y=0)
        self.spinbox_PID_PP_3.place(x=85, y=0)
        lbl_pid_vp3.place(x=0, y=30)
        self.spinbox_PID_VP_3.place(x=85, y=30)
        lbl_vol_in3.place(x=150, y=0)
        self.spinbox_VOL_IN3.place(x=235, y=0)
        lbl_pid_vi3.place(x=150, y=30)
        self.spinbox_PID_VI_3.place(x=235, y=30)
        lbl_vol_limit3.place(x=300, y=0)
        self.spinbox_VOL_LIMIT3.place(x=385, y=0)
        lbl_pid_vd3.place(x=300, y=30)
        self.spinbox_PID_VD_3.place(x=385, y=30)
        # 4
        self.group_pid_4.place(x=450, y=100)
        lbl_pid_p4.place(x=0, y=0)
        self.spinbox_PID_PP_4.place(x=85, y=0)
        lbl_pid_vp4.place(x=0, y=30)
        self.spinbox_PID_VP_4.place(x=85, y=30)
        lbl_vol_in4.place(x=150, y=0)
        self.spinbox_VOL_IN4.place(x=235, y=0)
        lbl_pid_vi4.place(x=150, y=30)
        self.spinbox_PID_VI_4.place(x=235, y=30)
        lbl_vol_limit4.place(x=300, y=0)
        self.spinbox_VOL_LIMIT4.place(x=385, y=0)
        lbl_pid_vd4.place(x=300, y=30)
        self.spinbox_PID_VD_4.place(x=385, y=30)

        self.save_para_m3.place(x=941, y=120)
        self.save_para_m4.place(x=941, y=148)
        self.edit_para_m3_4.place(x=945, y=90)
        #################################################################################
        ############################### SETTING_ARM_5_6 #################################
        lbl_mode_2 = Label(self.ARM5_6, text="MODE_選択")
        self.mode_menu_2 = tk.OptionMenu(self.ARM5_6, self.mode_run_var_2, *MODE_RUN)  # self.mode_run_var/self.num_motor_var
        lbl_num_motor_2 = Label(self.ARM5_6, text="NUM_MOTOR_選択")
        self.num_motor_2 = tk.OptionMenu(self.ARM5_6, self.num_motor_var_2, *NUM_MOTOR)
        self.toggle_servo_5 = tk.Button(self.ARM5_6, height=1, width=10, text="SERVO_5(Off)", background="#FFC652")
        self.toggle_servo_6 = tk.Button(self.ARM5_6, height=1, width=10, text="SERVO_6(Off)", background="#FFC652")
        self.group_pid_5 = tk.LabelFrame(self.ARM5_6, text="arm_5_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_5.propagate(False)
        lbl_pid_p5 = Label(self.group_pid_5, text="PID_position:")
        self.spinbox_PID_PP_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_pp_5)
        lbl_pid_vp5 = Label(self.group_pid_5, text="PID_(P)velocity:")
        self.spinbox_PID_VP_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vp_5)
        lbl_vol_in5 = Label(self.group_pid_5, text="VOL_input:")
        self.spinbox_VOL_IN5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8, textvariable=self.volin_5)
        lbl_pid_vi5 = Label(self.group_pid_5, text="PID_(I)velocity:")
        self.spinbox_PID_VI_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vi_5)
        lbl_vol_limit5 = Label(self.group_pid_5, text="VOL_limit:")
        self.spinbox_VOL_LIMIT5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8, textvariable=self.vollimit_5)
        lbl_pid_vd5 = Label(self.group_pid_5, text="PID_(D)velocity:")
        self.spinbox_PID_VD_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.001, width=8, textvariable=self.pid_vd_5)
        #
        self.group_pid_6 = tk.LabelFrame(self.ARM5_6, text="arm_6_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_6.propagate(False)
        lbl_pid_p6 = Label(self.group_pid_6, text="PID_position:")
        self.spinbox_PID_PP_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_pp_6)
        lbl_pid_vp6 = Label(self.group_pid_6, text="PID_(P)velocity:")
        self.spinbox_PID_VP_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vp_6)
        lbl_vol_in6 = Label(self.group_pid_6, text="VOL_input:")
        self.spinbox_VOL_IN6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8, textvariable=self.volin_6)
        lbl_pid_vi6 = Label(self.group_pid_6, text="PID_(I)velocity:")
        self.spinbox_PID_VI_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8, textvariable=self.pid_vi_6)
        lbl_vol_limit6 = Label(self.group_pid_6, text="VOL_limit:")
        self.spinbox_VOL_LIMIT6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8, textvariable=self.vollimit_6)
        lbl_pid_vd6 = Label(self.group_pid_6, text="PID_(D)velocity:")
        self.spinbox_PID_VD_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.001, width=8, textvariable=self.pid_vd_6)

        self.save_para_m5 = tk.Button(self.ARM5_6, height=1, width=6, text="Save M5", background="#AD8EE3", command=self.save_para_5)
        self.save_para_m6 = tk.Button(self.ARM5_6, height=1, width=6, text="Save M6", background="#AD8EE3", command=self.save_para_6)
        self.edit_para_m5_6 = tk.Button(self.ARM5_6, height=1, width=4, text="Edit", background="#AD8EE3", command=self.edit_3_switch)
        # 6
        lbl_mode_2.place(x=560, y=-5)
        self.mode_menu_2.place(x=640, y=-10)
        lbl_num_motor_2.place(x=760, y=-5)
        self.num_motor_2.place(x=870, y=-10)
        self.toggle_servo_5.place(x=150, y=-5)
        self.toggle_servo_6.place(x=250, y=-5)
        self.group_pid_6.place(x=450, y=100)
        lbl_pid_p6.place(x=0, y=0)
        self.spinbox_PID_PP_6.place(x=85, y=0)
        lbl_pid_vp6.place(x=0, y=30)
        self.spinbox_PID_VP_6.place(x=85, y=30)
        lbl_vol_in6.place(x=150, y=0)
        self.spinbox_VOL_IN6.place(x=235, y=0)
        lbl_pid_vi6.place(x=150, y=30)
        self.spinbox_PID_VI_6.place(x=235, y=30)
        lbl_vol_limit6.place(x=300, y=0)
        self.spinbox_VOL_LIMIT6.place(x=385, y=0)
        lbl_pid_vd6.place(x=300, y=30)
        self.spinbox_PID_VD_6.place(x=385, y=30)
        # 5
        self.group_pid_5.place(x=450, y=20)
        lbl_pid_p5.place(x=0, y=0)
        self.spinbox_PID_PP_5.place(x=85, y=0)
        lbl_pid_vp5.place(x=0, y=30)
        self.spinbox_PID_VP_5.place(x=85, y=30)
        lbl_vol_in5.place(x=150, y=0)
        self.spinbox_VOL_IN5.place(x=235, y=0)
        lbl_pid_vi5.place(x=150, y=30)
        self.spinbox_PID_VI_5.place(x=235, y=30)
        lbl_vol_limit5.place(x=300, y=0)
        self.spinbox_VOL_LIMIT5.place(x=385, y=0)
        lbl_pid_vd5.place(x=300, y=30)
        self.spinbox_PID_VD_5.place(x=385, y=30)

        self.save_para_m5.place(x=941, y=120)
        self.save_para_m6.place(x=941, y=148)
        self.edit_para_m5_6.place(x=945, y=90)
        #################################################################################
        self.labelhonsu1_counter = tk.Label(self.labelhonsu1, textvariable=self.riru_ok_var, font=("Arial", 30, "bold"), )
        self.labelhonsu2_counter = tk.Label(self.labelhonsu2, textvariable=self.hako_ok_var, font=("Arial", 30, "bold"), )
        self.labelhonsu3_counter = tk.Label(self.labelhonsu3, textvariable=self.riru_ng_var, font=("Arial", 30, "bold"), )
        self.labelhonsu4_counter = tk.Label(self.labelhonsu4, textvariable=self.lot_number, font=("Arial", 30, "bold"), )
        self.labelhonsu1.place(x=910, y=0)
        self.labelhonsu1_counter.place(x=0, y=3)
        self.labelhonsu2.place(x=910, y=160)
        self.labelhonsu2_counter.place(x=0, y=3)
        self.labelhonsu3.place(x=910, y=80)
        self.labelhonsu3_counter.place(x=0, y=3)
        self.labelhonsu4.place(x=910, y=240)
        self.labelhonsu4_counter.place(x=0, y=3)
        self.ARM1_2.place(x=0, y=0)
        self.ARM3_4.place(x=0, y=200)
        self.ARM5_6.place(x=0, y=400)

        self.reset_button.bind("<ButtonPress-1>", self.reset_button_on_click)
        self.reset_button.bind("<ButtonRelease-1>", self.reset_button_off_click)
        # TAB 6======================================================================
        self.labelconnect = tk.LabelFrame(self.tab6, text="PLC address", labelanchor="nw", width=300, height=170)
        self.labelconnect.propagate(False)
        lbl_box = Label(self.labelconnect, text="Local address IP")
        txt_address = tk.Entry(self.labelconnect, justify=tk.RIGHT, width=15, textvariable=self.IP_PLC)
        lbl_box1 = Label(self.labelconnect, text="Port")
        txt_port = tk.Entry(self.labelconnect, justify=tk.RIGHT, width=15, textvariable=self.PORT_PLC)
        Sock_connect_button = tk.Button(self.labelconnect, text="PLC_Connect", command=self.socket_connect)
        self.PLC_stt = tk.Text(self.labelconnect, height=2, width=6, font=("Arial", 15), background='white')
        #
        self.labelconnect.place(x=0, y=0)
        lbl_box.place(x=0, y=5)
        txt_address.place(x=100, y=5)
        lbl_box1.place(x=0, y=30)
        txt_port.place(x=100, y=30)
        Sock_connect_button.place(x=0, y=70)
        self.PLC_stt.place(x=200, y=90)
        # シリアルポート選択メニュー

        self.labelserial = tk.LabelFrame(self.tab6, text="Serial_setting", labelanchor="nw", width=250, height=170)
        self.plc_mode = tk.LabelFrame(self.tab6, text="GROUP7", labelanchor="nw", width=340, height=170)
        self.dengen_lb = tk.LabelFrame(self.tab6, text="GROUP8", labelanchor="nw", width=170, height=170)
        self.end_lb = tk.LabelFrame(self.tab6, text="プログラム終了", labelanchor="nw", width=170, height=170)
        self.labelserial.propagate(False)
        self.plc_mode.propagate(False)
        self.dengen_lb.propagate(False)
        self.end_lb.propagate(False)

        ##################### SERIAL ######################################
        lbl_box31 = Label(self.labelserial, text="COM 選択")
        self.port_menu = tk.OptionMenu(self.labelserial, self.port_var, 'None')
        self.port_menu.config(height=1, width=5, bg='gray')
        self.refresh_port_bt = tk.Button(self.labelserial, height=1, width=7, text="Refresh", command=self.refresh_serial_ports)
        lbl_box3 = Label(self.labelserial, text="ボーレート選択")
        self.baudrate_menu = tk.OptionMenu(self.labelserial, self.baudrate_var, *BAUDRATES)
        lbl_box4 = Label(self.labelserial, text="デリミタ選択")
        self.delimiter_menu = tk.OptionMenu(self.labelserial, self.delimiter_var, *DELIMITERS.keys())
        self.delimiter_menu.config(height=1, width=5)
        lbl_box5 = Label(self.labelserial, text="接続")
        self.toggle_button = tk.Button(self.labelserial, height=1, width=9, text="Connect", command=self.toggle_serial_port)
        self.entry_text = tk.Entry(self.labelserial, width=25)
        Serial_send_button = tk.Button(self.labelserial, height=1, width=5, text="送信", command=self.send_data)

        self.labelserial.place(x=0, y=170)
        self.plc_mode.place(x=650, y=0)
        self.dengen_lb.place(x=300, y=0)
        self.end_lb.place(x=470, y=0)
        lbl_box31.place(x=0, y=0)
        self.port_menu.place(x=100, y=0)
        self.refresh_port_bt.place(x=180, y=0)
        lbl_box3.place(x=0, y=30)
        self.baudrate_menu.place(x=100, y=30)
        lbl_box4.place(x=0, y=60)
        self.delimiter_menu.place(x=100, y=60)
        lbl_box5.place(x=0, y=90)
        self.toggle_button.place(x=100, y=90)
        self.entry_text.place(x=0, y=120)
        Serial_send_button.place(x=170, y=120)

        self.displayserial = tk.LabelFrame(self.tab6, text="IN<>OUT", labelanchor="nw", width=500, height=250)
        self.displayserial.propagate(False)
        lbl_sock = tk.Label(self.displayserial, text="PLC socket")
        lbl_comp = tk.Label(self.displayserial, text="Serial")
        self.Text_sock_recieved = Text(self.displayserial, height=13, width=28, background='#C2DEFA')
        self.received_text = tk.Text(self.displayserial, height=13, width=28, background='#C2DEFA')
        clear_button = tk.Button(self.displayserial, text="履歴クリア", command=self.clear_received_text)
        self.connect_mode = tk.Button(self.plc_mode, wraplength=90, text="PROGRAM RUN", height=5, width=12, bd=6, bg='#019a66', font=self.small_font, fg="black", command=self.Communication_serial_run)
        self.program_mode = tk.Button(self.plc_mode, wraplength=90, text="プログラム  停止", height=5, width=12, bd=6, bg='#ef4123', font=self.small_font, fg="black", command=self.Communication_serial_stop)
        self.dengen_bt = tk.Button(self.dengen_lb, wraplength=90, text="電源OFF", height=5, width=12, bd=6, bg='#ef4123', font=self.small_font, fg="black", command=self.ending_program)
        self.end_bt = tk.Button(self.end_lb, wraplength=90, text="プログラム   終了", height=5, width=12, bd=6, bg='pink', font=self.small_font, fg="black", command=self.end_thread)
        #
        self.displayserial.place(x=0, y=345)
        lbl_sock.place(x=0, y=0)
        lbl_comp.place(x=250, y=0)
        self.Text_sock_recieved.place(x=0, y=20)
        self.received_text.place(x=250, y=20)
        clear_button.place(x=250, y=200)
        self.connect_mode.place(x=20, y=20)
        self.program_mode.place(x=200, y=20)
        self.dengen_bt.place(x=30, y=20)
        self.end_bt.place(x=30, y=20)
        ################################################MSP_arm1_2 SERIAL Group
        self.labelserial_msp = tk.LabelFrame(self.tab6, text="Serial_ARM_1-2_setting", labelanchor="nw", width=250, height=170)
        self.labelserial_msp.propagate(False)
        lbl_box_msp = Label(self.labelserial_msp, text="COM 選択")
        self.port_menu_msp = tk.OptionMenu(self.labelserial_msp, self.port_var_msp, 'None')
        self.port_menu_msp.config(height=1, width=5, bg='gray')
        self.refresh_port_bt_msp = tk.Button(self.labelserial_msp, height=1, width=7, text="Refresh", command=self.refresh_serial_M1_2)
        lbl_box_msp3 = Label(self.labelserial_msp, text="ボーレート選択")
        self.baudrate_menu_msp = tk.OptionMenu(self.labelserial_msp, self.baudrate_var_msp, *BAUDRATES)
        lbl_box_msp4 = Label(self.labelserial_msp, text="デリミタ選択")
        self.delimiter_menu_msp = tk.OptionMenu(self.labelserial_msp, self.delimiter_var_msp, *DELIMITERS.keys())
        self.delimiter_menu_msp.config(height=1, width=5)
        lbl_box_msp5 = Label(self.labelserial_msp, text="接続")
        self.toggle_button_msp = tk.Button(self.labelserial_msp, height=1, width=9, text="Connect", command=self.toggle_serial_connect_M1_2)
        self.entry_text_msp = tk.Entry(self.labelserial_msp, width=25)
        Serial_send_button_msp = tk.Button(self.labelserial_msp, height=1, width=5, text="送信", command=self.send_data_M1_2)

        self.labelserial_msp.place(x=255, y=170)
        lbl_box_msp.place(x=0, y=0)
        self.port_menu_msp.place(x=100, y=0)
        self.refresh_port_bt_msp.place(x=180, y=0)
        lbl_box_msp3.place(x=0, y=30)
        self.baudrate_menu_msp.place(x=100, y=30)
        lbl_box_msp4.place(x=0, y=60)
        self.delimiter_menu_msp.place(x=100, y=60)
        lbl_box_msp5.place(x=0, y=90)
        self.toggle_button_msp.place(x=100, y=90)
        self.entry_text_msp.place(x=0, y=120)
        Serial_send_button_msp.place(x=170, y=120)
        ##
        ################################################MSP_arm3_4 SERIAL Group
        self.labelserial_msp_1 = tk.LabelFrame(self.tab6, text="Serial_ARM_3-4_setting", labelanchor="nw", width=250, height=170)
        self.labelserial_msp_1.propagate(False)
        lbl_box_msp_1 = Label(self.labelserial_msp_1, text="COM 選択")
        self.port_menu_msp_1 = tk.OptionMenu(self.labelserial_msp_1, self.port_var_msp_1, 'None')
        self.port_menu_msp_1.config(height=1, width=5, bg='gray')
        self.refresh_port_bt_msp_1 = tk.Button(self.labelserial_msp_1, height=1, width=7, text="Refresh", command=self.refresh_serial_M3_4)
        lbl_box_msp3_1 = Label(self.labelserial_msp_1, text="ボーレート選択")
        self.baudrate_menu_msp_1 = tk.OptionMenu(self.labelserial_msp_1, self.baudrate_var_msp_1, *BAUDRATES)
        lbl_box_msp4_1 = Label(self.labelserial_msp_1, text="デリミタ選択")
        self.delimiter_menu_msp_1 = tk.OptionMenu(self.labelserial_msp_1, self.delimiter_var_msp_1, *DELIMITERS.keys())
        self.delimiter_menu_msp_1.config(height=1, width=5)
        lbl_box_msp5_1 = Label(self.labelserial_msp_1, text="接続")
        self.toggle_button_msp_1 = tk.Button(self.labelserial_msp_1, height=1, width=9, text="Connect", command=self.toggle_serial_connect_M3_4)
        self.entry_text_msp_1 = tk.Entry(self.labelserial_msp_1, width=25)
        Serial_send_button_msp_1 = tk.Button(self.labelserial_msp_1, height=1, width=5, text="送信", command=self.send_data_M3_4)

        self.labelserial_msp_1.place(x=505, y=170)
        lbl_box_msp_1.place(x=0, y=0)
        self.port_menu_msp_1.place(x=100, y=0)
        self.refresh_port_bt_msp_1.place(x=180, y=0)
        lbl_box_msp3_1.place(x=0, y=30)
        self.baudrate_menu_msp_1.place(x=100, y=30)
        lbl_box_msp4_1.place(x=0, y=60)
        self.delimiter_menu_msp_1.place(x=100, y=60)
        lbl_box_msp5_1.place(x=0, y=90)
        self.toggle_button_msp_1.place(x=100, y=90)
        self.entry_text_msp_1.place(x=0, y=120)
        Serial_send_button_msp_1.place(x=170, y=120)
        ###
        ################################################MSP_arm5_6 SERIAL Group
        self.labelserial_msp_2 = tk.LabelFrame(self.tab6, text="Serial_ARM_5-6_setting", labelanchor="nw", width=250, height=170)
        self.labelserial_msp_2.propagate(False)
        lbl_box_msp_2 = Label(self.labelserial_msp_2, text="COM 選択")
        self.port_menu_msp_2 = tk.OptionMenu(self.labelserial_msp_2, self.port_var_msp_2, 'None')
        self.port_menu_msp_2.config(height=1, width=5, bg='gray')
        self.refresh_port_bt_msp_2 = tk.Button(self.labelserial_msp_2, height=1, width=7, text="Refresh", command=self.refresh_serial_M5_6)
        lbl_box_msp3_2 = Label(self.labelserial_msp_2, text="ボーレート選択")
        self.baudrate_menu_msp_2 = tk.OptionMenu(self.labelserial_msp_2, self.baudrate_var_msp_2, *BAUDRATES)
        lbl_box_msp4_2 = Label(self.labelserial_msp_2, text="デリミタ選択")
        self.delimiter_menu_msp_2 = tk.OptionMenu(self.labelserial_msp_2, self.delimiter_var_msp_2, *DELIMITERS.keys())
        self.delimiter_menu_msp_2.config(height=1, width=5)
        lbl_box_msp5_2 = Label(self.labelserial_msp_2, text="接続")
        self.toggle_button_msp_2 = tk.Button(self.labelserial_msp_2, height=1, width=9, text="Connect", command=self.toggle_serial_connect_M5_6)
        self.entry_text_msp_2 = tk.Entry(self.labelserial_msp_2, width=25)
        Serial_send_button_msp_2 = tk.Button(self.labelserial_msp_2, height=1, width=5, text="送信", command=self.send_data_M5_6)

        self.labelserial_msp_2.place(x=755, y=170)
        lbl_box_msp_2.place(x=0, y=0)
        self.port_menu_msp_2.place(x=100, y=0)
        self.refresh_port_bt_msp_2.place(x=180, y=0)
        lbl_box_msp3_2.place(x=0, y=30)
        self.baudrate_menu_msp_2.place(x=100, y=30)
        lbl_box_msp4_2.place(x=0, y=60)
        self.delimiter_menu_msp_2.place(x=100, y=60)
        lbl_box_msp5_2.place(x=0, y=90)
        self.toggle_button_msp_2.place(x=100, y=90)
        self.entry_text_msp_2.place(x=0, y=120)
        Serial_send_button_msp_2.place(x=170, y=120)
        ##
        self.msp_group_serial = tk.LabelFrame(self.tab6, text="MSP_protocol", labelanchor="nw", width=250, height=250)
        self.msp_group_serial.propagate(False)
        lbl_msp_bt1 = tk.Label(self.msp_group_serial, text="MSP send request")
        lbl_msp_recv = tk.Label(self.msp_group_serial, text="MSP recever")
        lbl_msp_command = tk.Label(self.msp_group_serial, text="MSP Command")
        self.send_msp_text = tk.Entry(self.msp_group_serial, width=20, background='#C2DEFA')
        send_msp_button = tk.Button(self.msp_group_serial, text="MSP Cmd")
        self.Text_msp_recieved = Text(self.msp_group_serial, height=8, width=20, background='#C2DEFA')
        self.msp_group_serial.place(x=510, y=345)
        lbl_msp_command.place(x=0, y=10)
        lbl_msp_bt1.place(x=0, y=45)
        self.send_msp_text.place(x=100, y=10)
        send_msp_button.place(x=100, y=45)
        lbl_msp_recv.place(x=0, y=80)
        self.Text_msp_recieved.place(x=100, y=80)
        ########

        #####################
        # TAB 3 CONTROL
        self.displaydatamemory = tk.LabelFrame(self.tab4, text="機械パラメータ", labelanchor="nw", width=500, height=500)
        self.displaydatamemory.propagate(False)
        self.displaydatamemory.place(x=0, y=0)
        self.displayprint = tk.LabelFrame(self.tab4, text="Serial設定", labelanchor="nw", width=500, height=500)
        self.displayprint.propagate(False)
        self.displayprint.place(x=500, y=0)
        self.displayprintstatus = tk.LabelFrame(self.displayprint, text="Serial通信", labelanchor="nw", width=100, height=185)
        self.displayprintstatus.propagate(False)
        self.displayprintstatus.place(x=380, y=0)
        self.display_riruhikae = tk.LabelFrame(self.displayprint, text="NC", labelanchor="nw", width=100, height=85)
        self.display_riruhikae.propagate(False)
        self.display_riruhikae.place(x=210, y=0)
        self.display_hakohikae = tk.LabelFrame(self.displayprint, text="NC", labelanchor="nw", width=100, height=85)
        self.display_hakohikae.propagate(False)
        self.display_hakohikae.place(x=210, y=100)
        self.displaynolabel = tk.LabelFrame(self.displayprint, text="NC", labelanchor="nw", width=100, height=200)
        self.displaynolabel.propagate(False)
        self.displaynolabel.place(x=380, y=270)

        print_all_clear_data = tk.Button(self.displayprint, height=5, width=13, text="BUTTON4", command=lambda: self.Socket_send("ST R9002"), bg="orange")
        print_ryru = tk.Button(self.displayprint, height=5, width=13, wraplength=80, text="BUTTON0", bg="green")
        print_hako = tk.Button(self.displayprint, height=5, width=13, wraplength=80, text="BUTTON2", bg="green")
        self.print_communicate = tk.Button(self.displayprintstatus, height=3, width=8, text="通信開始", bg="gray", command=lambda: self.Socket_send("ST R9000"))
        self.print_communicate_success = tk.Button(self.displayprintstatus, height=3, width=8, wraplength=80, text="BUTTON3", bg="gray", command=lambda: self.Socket_send("ST R9001"))
        print_riruumu = tk.Button(self.displaynolabel, height=4, width=9, wraplength=80, text="BUTTON", bg="gray")
        print_hakoumu = tk.Button(self.displaynolabel, height=4, width=9, wraplength=80, text="BUTTON1", bg="gray")
        self.lbl_datariru = tk.Label(self.display_riruhikae, textvariable=self.text_var1, font=("Arial", 38, "bold"))
        self.lbl_datahako = tk.Label(self.display_hakohikae, textvariable=self.text_var2, font=("Arial", 38, "bold"))
        print_all_clear_data.place(x=30, y=230)
        print_ryru.place(x=30, y=0)
        print_hako.place(x=30, y=100)
        self.print_communicate.place(x=15, y=15)
        self.print_communicate_success.place(x=15, y=90)
        print_riruumu.place(x=11, y=13)
        print_hakoumu.place(x=11, y=103)
        self.lbl_datariru.place(x=0, y=0)
        self.lbl_datahako.place(x=0, y=0)

        print_ryru.bind("<ButtonPress-1>", self.print_ryru_on_click)
        print_ryru.bind("<ButtonRelease-1>", self.print_ryru_off_click)
        print_hako.bind("<ButtonPress-1>", self.print_hako_on_click)
        print_hako.bind("<ButtonRelease-1>", self.print_hako_off_click)

        lbl_dm1000 = Label(self.displaydatamemory, text="SPEED DM1000")
        scale_DM1000 = tk.Scale(self.displaydatamemory, variable=self.DM1000_var, command=self.slider_DM1000, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=0, to=1000, resolution=1, tickinterval=0)
        lbl_dm1002 = Label(self.displaydatamemory, text="SPEED DM1002")
        scale_DM1002 = tk.Scale(self.displaydatamemory, variable=self.DM1002_var, command=self.slider_DM1002, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=0, to=1000, resolution=1, tickinterval=0)
        lbl_dm1004 = Label(self.displaydatamemory, text="SPEED DM1004")
        scale_DM1004 = tk.Scale(self.displaydatamemory, variable=self.DM1004_var, command=self.slider_DM1004, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=0, to=1000, resolution=1, tickinterval=0)
        lbl_dm1006 = Label(self.displaydatamemory, text="SPEED DM1006")
        scale_DM1006 = tk.Scale(self.displaydatamemory, variable=self.DM1006_var, command=self.slider_DM1006, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=0, to=1000, resolution=1, tickinterval=0)
        lbl_dm1000.place(x=0, y=5)
        scale_DM1000.place(x=120, y=5)
        lbl_dm1002.place(x=0, y=40)
        scale_DM1002.place(x=120, y=40)
        lbl_dm1004.place(x=0, y=75)
        scale_DM1004.place(x=120, y=75)
        lbl_dm1006.place(x=0, y=110)
        scale_DM1006.place(x=120, y=110)

        self.labelhonsu = tk.LabelFrame(self.tab5, text="現在本数変更", labelanchor="nw", width=200, height=300)
        self.labelhonsu.propagate(False)
        self.kasan_button = tk.Button(self.labelhonsu, wraplength=90, font=("Arial", 20), text="＋", height=1, width=3, bd=6, bg='#c0c0c0', fg="green", command=lambda: self.Socket_send("ST R4108"))
        self.gensan_button = tk.Button(self.labelhonsu, wraplength=90, font=("Arial", 20), text="ー", height=1, width=3, bd=6, bg='#c0c0c0', fg="#ef4123", command=lambda: self.Socket_send("ST R4109"))
        spinbox_DM2000 = ttk.Spinbox(self.labelhonsu, textvariable=self.riru_henko, from_=0, to=5, increment=1, wrap=True, state="normal", width=1, font=("Arial", 110))

        self.labelhonsu.place(x=430, y=10)
        self.kasan_button.place(x=10, y=200)
        self.gensan_button.place(x=120, y=200)
        spinbox_DM2000.place(x=50, y=10)

    def create_log_page(self):
        Elogframe = Frame(self.tab5_rb)
        Elogframe.place(x=40, y=15)
        scrollbar = Scrollbar(Elogframe)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.tab5_rb.ElogView = Listbox(Elogframe, width=100, height=40, yscrollcommand=scrollbar.set)
        try:
            Elog = pickle.load(open("ErrorLog", "rb"))
        except:
            Elog = ['##BEGINNING OF LOG##']
            pickle.dump(Elog, open("ErrorLog", "wb"))
        time.sleep(.1)
        for item in Elog:
            self.tab5_rb.ElogView.insert(END, item)
        self.tab5_rb.ElogView.pack()
        scrollbar.config(command=self.tab5_rb.ElogView.yview)

        clearLogBut = Button(self.tab5_rb, text="Clear Log", width=20, command=self.clearLog)
        clearLogBut.place(x=800, y=500)

    def clearLog(self):
        self.tab5_rb.ElogView.delete(1, END)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def robot_variable(self):

        self.JogStepsStat = IntVar()
        self.J1OpenLoopStat = IntVar()
        self.J2OpenLoopStat = IntVar()
        self.J3OpenLoopStat = IntVar()
        self.J4OpenLoopStat = IntVar()
        self.J5OpenLoopStat = IntVar()
        self.J6OpenLoopStat = IntVar()
        self.DisableWristRot = IntVar()
        self.xboxUse = 0
        global curTheme

        self.J1CalStat = IntVar()
        self.J2CalStat = IntVar()
        self.J3CalStat = IntVar()
        self.J4CalStat = IntVar()
        self.J5CalStat = IntVar()
        self.J6CalStat = IntVar()
        self.J7CalStat = IntVar()
        self.J8CalStat = IntVar()
        self.J9CalStat = IntVar()
        self.J1CalStat2 = IntVar()
        self.J2CalStat2 = IntVar()
        self.J3CalStat2 = IntVar()
        self.J4CalStat2 = IntVar()
        self.J5CalStat2 = IntVar()
        self.J6CalStat2 = IntVar()
        self.J7CalStat2 = IntVar()
        self.J8CalStat2 = IntVar()
        self.J9CalStat2 = IntVar()
        self.IncJogStat = IntVar()
        self.fullRot = IntVar()
        self.pick180 = IntVar()
        self.pickClosest = IntVar()
        self.autoBG = IntVar()
        self.estopActive = False
        self.posOutreach = False
        self.SplineTrue = False
        self.gcodeSpeed = "10"
        self.inchTrue = False
        self.moveInProc = 0

        # define axis limits in degrees
        self.J1PosLim = 170;
        self.J1NegLim = 170;
        self.J2PosLim = 90;
        self.J2NegLim = 42;
        self.J3PosLim = 52;
        self.J3NegLim = 89;
        self.J4PosLim = 165;
        self.J4NegLim = 165;
        self.J5PosLim = 105;
        self.J5NegLim = 105;
        self.J6PosLim = 155;
        self.J6NegLim = 155;
        self.J7PosLim = 500;
        self.J7NegLim = 0;
        self.J8PosLim = 500;
        self.J8NegLim = 0;
        self.J9PosLim = 500;
        self.J9NegLim = 0;

        self.cam_on = False
        self.cap = None

    def StopJog(self, a):
        command = "S\n"
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 0):
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

    def create_robot_program(self):
        self.label_robot_para = tk.LabelFrame(self.tab10, text="PARAMETER", labelanchor="nw", width=500, height=200)
        self.label_robot_para.propagate(False)
        self.label_robot_para.place(x=0, y=10)

        curRowLab = Label(self.label_robot_para, text="Current Row:")
        curRowLab.place(x=98, y=120)
        self.almStatusLab = Label(self.label_robot_para, text="SYSTEM READY - NO ACTIVE ALARMS", style="OK.TLabel")
        self.almStatusLab.place(x=25, y=12)
        xbcStatusLab = Label(self.label_robot_para, text="Xbox OFF")
        xbcStatusLab.place(x=1270, y=80)
        runStatusLab = Label(self.label_robot_para, text="PROGRAM STOPPED")
        runStatusLab.place(x=20, y=150)
        ProgLab = Label(self.label_robot_para, text="Program:")
        ProgLab.place(x=10, y=45)
        jogIncrementLab = Label(self.label_robot_para, text="Increment Value:")
        # jogIncrementLab.place(x=370, y=45)
        speedLab = Label(self.label_robot_para, text="self.Speed")
        speedLab.place(x=300, y=83)
        ACCLab = Label(self.label_robot_para, text="Acceleration               %")
        ACCLab.place(x=300, y=103)
        DECLab = Label(self.label_robot_para, text="Deceleration               %")
        DECLab.place(x=300, y=123)
        DECLab = Label(self.label_robot_para, text="Ramp                           %")
        DECLab.place(x=300, y=143)

        RoundLab = Label(self.label_robot_para, text="Rounding               mm")
        RoundLab.place(x=525, y=82)
        self.incrementEntryField = Entry(self.label_robot_para, width=4, justify="center")
        self.incrementEntryField.place(x=380, y=45)
        self.curRowEntryField = Entry(self.label_robot_para, width=4, justify="center")
        self.curRowEntryField.place(x=174, y=120)
        self.manEntryField = Entry(self.label_robot_para, width=105)
        self.manEntryField.place(x=10, y=700)
        self.ProgEntryField = Entry(self.label_robot_para, width=20, justify="center")
        self.ProgEntryField.place(x=70, y=45)
        self.speedEntryField = Entry(self.label_robot_para, width=4, justify="center")
        self.speedEntryField.place(x=380, y=80)
        self.ACCspeedField = Entry(self.label_robot_para, width=4, justify="center")
        self.ACCspeedField.place(x=380, y=100)
        self.DECspeedField = Entry(self.label_robot_para, width=4, justify="center")
        self.DECspeedField.place(x=380, y=120)
        self.ACCrampField = Entry(self.label_robot_para, width=4, justify="center")
        self.ACCrampField.place(x=380, y=140)
        self.roundEntryField = Entry(self.label_robot_para, width=4, justify="center")
        self.roundEntryField.place(x=590, y=80)
        self.speedOption = StringVar()
        self.speedMenu = ttk.OptionMenu(self.label_robot_para, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
        self.speedMenu.place(x=412, y=76)

        ProgBut = Button(self.label_robot_para, text="Load", width=10, command=self.loadProg)
        ProgBut.place(x=202, y=42)

        CreateBut = Button(self.label_robot_para, text="New self.Progr", width=10, command=self.CreateProg)
        CreateBut.place(x=285, y=42)

        runProgBut = Button(self.label_robot_para, command=self.runProg)
        playPhoto = PhotoImage(file="play-icon.gif")
        runProgBut.config(image=playPhoto)
        runProgBut.place(x=20, y=80)

        xboxBut = Button(self.label_robot_para, command=self.xbox)
        xboxPhoto = PhotoImage(file="xbox.gif")
        xboxBut.config(image=xboxPhoto)
        xboxBut.place(x=700, y=80)

        stopProgBut = Button(self.label_robot_para, command=self.stopProg)
        stopPhoto = PhotoImage(file="stop-icon.gif")
        stopProgBut.config(image=stopPhoto)
        stopProgBut.place(x=220, y=80)

        revBut = Button(self.label_robot_para, text="REV ", width=5, command=self.stepRev)
        revBut.place(x=105, y=80)

        fwdBut = Button(self.label_robot_para, text="FWD", width=5, command=self.stepFwd)
        fwdBut.place(x=160, y=80)

        IncJogCbut = Checkbutton(self.label_robot_para, text="Incremental Jog", variable=self.IncJogStat)
        IncJogCbut.place(x=412, y=46)

        #################################################################################

        ##
        self.label_robot_pg = tk.LabelFrame(self.tab10, text="ROBOT SLIDER", labelanchor="nw", width=480, height=150)
        self.label_robot_pg.propagate(False)
        self.label_robot_pg.place(x=500, y=10)

        ##################################################################
        J1Lab = Label(self.label_robot_pg, font=("Arial", 14), text="J1")
        J1Lab.place(x=5, y=5)
        self.J1curAngEntryField = Entry(self.label_robot_pg, width=5, justify="center")
        self.J1curAngEntryField.place(x=35, y=9)
        self.J1jogNegBut = Button(self.label_robot_pg, text="-", width=2)
        self.J1jogNegBut.bind("<ButtonPress>", self.SelJ1jogNeg)
        self.J1jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J1jogNegBut.place(x=77, y=7, width=30, height=25)
        self.J1jogslide = Scale(self.label_robot_pg, from_=-self.J1NegLim, to=self.J1PosLim, length=90, orient=HORIZONTAL, command=self.J1sliderUpdate)
        self.J1jogslide.bind("<ButtonRelease-1>", self.J1sliderExecute)
        self.J1jogslide.place(x=115, y=7)
        self.J1jogPosBut = Button(self.label_robot_pg, text="+", width=2)
        self.J1jogPosBut.bind("<ButtonPress>", self.SelJ1jogPos)
        self.J1jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J1jogPosBut.place(x=200, y=7, width=30, height=25)
        self.J1negLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(-self.J1NegLim), style="Jointlim.TLabel")
        self.J1negLimLab.place(x=115, y=25)
        self.J1posLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(self.J1PosLim), style="Jointlim.TLabel")
        self.J1posLimLab.place(x=170, y=25)
        self.J1slidelabel = Label(self.label_robot_pg)
        self.J1slidelabel.place(x=190, y=25)
        #####################################################################
        J2Lab = Label(self.label_robot_pg, font=("Arial", 14), text="J2")
        J2Lab.place(x=5, y=5 + 40)
        self.J2curAngEntryField = Entry(self.label_robot_pg, width=5, justify="center")
        self.J2curAngEntryField.place(x=35, y=9 + 40)
        self.J2jogNegBut = Button(self.label_robot_pg, text="-", width=3)
        self.J2jogNegBut.bind("<ButtonPress>", self.SelJ2jogNeg)
        self.J2jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J2jogNegBut.place(x=77, y=7 + 40, width=30, height=25)
        self.J2jogPosBut = Button(self.label_robot_pg, text="+", width=3)
        self.J2jogPosBut.bind("<ButtonPress>", self.SelJ2jogPos)
        self.J2jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J2jogPosBut.place(x=200, y=7 + 40, width=30, height=25)
        self.J2negLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(-self.J2NegLim), style="Jointlim.TLabel")
        self.J2negLimLab.place(x=115, y=25 + 40)
        self.J2posLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(self.J2PosLim), style="Jointlim.TLabel")
        self.J2posLimLab.place(x=170, y=25 + 40)
        self.J2slidelabel = Label(self.label_robot_pg)
        self.J2slidelabel.place(x=190, y=25 + 40)
        self.J2jogslide = Scale(self.label_robot_pg, from_=-self.J2NegLim, to=self.J2PosLim, length=90, orient=HORIZONTAL, command=self.J2sliderUpdate)
        self.J2jogslide.bind("<ButtonRelease-1>", self.J2sliderExecute)
        self.J2jogslide.place(x=115, y=7 + 40)
        #####################################################################
        self.J3Lab = Label(self.label_robot_pg, font=("Arial", 14), text="J3")
        self.J3Lab.place(x=5, y=5 + 80)
        self.J3curAngEntryField = Entry(self.label_robot_pg, width=5, justify="center")
        self.J3curAngEntryField.place(x=35, y=9 + 80)
        self.J3jogNegBut = Button(self.label_robot_pg, text="-", width=3)
        self.J3jogNegBut.bind("<ButtonPress>", self.SelJ3jogNeg)
        self.J3jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J3jogNegBut.place(x=77, y=7 + 80, width=30, height=25)
        self.J3jogPosBut = Button(self.label_robot_pg, text="+", width=3)
        self.J3jogPosBut.bind("<ButtonPress>", self.SelJ3jogPos)
        self.J3jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J3jogPosBut.place(x=200, y=7 + 80, width=30, height=25)
        self.J3negLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(-self.J3NegLim), style="Jointlim.TLabel")
        self.J3negLimLab.place(x=115, y=25 + 80)
        self.J3posLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(self.J3PosLim), style="Jointlim.TLabel")
        self.J3posLimLab.place(x=170, y=25 + 80)
        self.J3slidelabel = Label(self.label_robot_pg)
        self.J3slidelabel.place(x=190, y=25 + 80)
        self.J3jogslide = Scale(self.label_robot_pg, from_=-self.J3NegLim, to=self.J3PosLim, length=90, orient=HORIZONTAL, command=self.J3sliderUpdate)
        self.J3jogslide.bind("<ButtonRelease-1>", self.J3sliderExecute)
        self.J3jogslide.place(x=115, y=7 + 80)
        #####################################################################
        J4Lab = Label(self.label_robot_pg, font=("Arial", 14), text="J4")
        J4Lab.place(x=5 + 230, y=5)
        self.J4curAngEntryField = Entry(self.label_robot_pg, width=5, justify="center")
        self.J4curAngEntryField.place(x=35 + 230, y=9)
        self.J4jogNegBut = Button(self.label_robot_pg, text="-", width=3)
        self.J4jogNegBut.bind("<ButtonPress>", self.SelJ4jogNeg)
        self.J4jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J4jogNegBut.place(x=77 + 230, y=7, width=30, height=25)
        self.J4jogPosBut = Button(self.label_robot_pg, text="+", width=3)
        self.J4jogPosBut.bind("<ButtonPress>", self.SelJ4jogPos)
        self.J4jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J4jogPosBut.place(x=200 + 230, y=7, width=30, height=25)
        self.J4negLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(-self.J4NegLim), style="Jointlim.TLabel")
        self.J4negLimLab.place(x=115 + 230, y=25)
        self.J4posLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(self.J4PosLim), style="Jointlim.TLabel")
        self.J4posLimLab.place(x=170 + 230, y=25)
        self.J4slidelabel = Label(self.label_robot_pg)
        self.J4slidelabel.place(x=190 + 230, y=25)
        self.J4jogslide = Scale(self.label_robot_pg, from_=-self.J4NegLim, to=self.J4PosLim, length=90, orient=HORIZONTAL, command=self.J4sliderUpdate)
        self.J4jogslide.bind("<ButtonRelease-1>", self.J4sliderExecute)
        self.J4jogslide.place(x=115 + 230, y=7)
        ###########################################################################
        J5Lab = Label(self.label_robot_pg, font=("Arial", 14), text="J5")
        J5Lab.place(x=5 + 230, y=5 + 40)
        self.J5curAngEntryField = Entry(self.label_robot_pg, width=5, justify="center")
        self.J5curAngEntryField.place(x=35 + 230, y=9 + 40)
        self.J5jogNegBut = Button(self.label_robot_pg, text="-", width=3)
        self.J5jogNegBut.bind("<ButtonPress>", self.SelJ5jogNeg)
        self.J5jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J5jogNegBut.place(x=77 + 230, y=7 + 40, width=30, height=25)
        self.J5jogPosBut = Button(self.label_robot_pg, text="+", width=3)
        self.J5jogPosBut.bind("<ButtonPress>", self.SelJ5jogPos)
        self.J5jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J5jogPosBut.place(x=200 + 230, y=7 + 40, width=30, height=25)
        self.J5negLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(-self.J5NegLim), style="Jointlim.TLabel")
        self.J5negLimLab.place(x=115 + 230, y=25 + 40)
        self.J5posLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(self.J5PosLim), style="Jointlim.TLabel")
        self.J5posLimLab.place(x=170 + 230, y=25 + 40)
        self.J5slidelabel = Label(self.label_robot_pg)
        self.J5slidelabel.place(x=190 + 230, y=25 + 40)
        self.J5jogslide = Scale(self.label_robot_pg, from_=-self.J5NegLim, to=self.J5PosLim, length=90, orient=HORIZONTAL, command=self.J5sliderUpdate)
        self.J5jogslide.bind("<ButtonRelease-1>", self.J5sliderExecute)
        self.J5jogslide.place(x=115 + 230, y=7 + 40)
        ############################################################################
        J6Lab = Label(self.label_robot_pg, font=("Arial", 14), text="J6")
        J6Lab.place(x=5 + 230, y=5 + 80)
        self.J6curAngEntryField = Entry(self.label_robot_pg, width=5, justify="center")
        self.J6curAngEntryField.place(x=35 + 230, y=9 + 80)
        self.J6jogNegBut = Button(self.label_robot_pg, text="-", width=3)
        self.J6jogNegBut.bind("<ButtonPress>", self.SelJ6jogNeg)
        self.J6jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J6jogNegBut.place(x=77 + 230, y=7 + 80, width=30, height=25)
        self.J6jogPosBut = Button(self.label_robot_pg, text="+", width=3)
        self.J6jogPosBut.bind("<ButtonPress>", self.SelJ6jogPos)
        self.J6jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J6jogPosBut.place(x=200 + 230, y=7 + 80, width=30, height=25)
        self.J6negLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(-self.J6NegLim), style="Jointlim.TLabel")
        self.J6negLimLab.place(x=115 + 230, y=25 + 80)
        self.J6posLimLab = Label(self.label_robot_pg, font=("Arial", 8), text=str(self.J6PosLim), style="Jointlim.TLabel")
        self.J6posLimLab.place(x=170 + 230, y=25 + 80)
        self.J6slidelabel = Label(self.label_robot_pg)
        self.J6slidelabel.place(x=190 + 230, y=25 + 80)
        self.J6jogslide = Scale(self.label_robot_pg, from_=-self.J6NegLim, to=self.J6PosLim, length=90, orient=HORIZONTAL, command=self.J6sliderUpdate)
        self.J6jogslide.bind("<ButtonRelease-1>", self.J6sliderExecute)
        self.J6jogslide.place(x=115 + 230, y=7 + 80)
        #################################################################################
        self.label_robot_pg_add = tk.LabelFrame(self.tab10, text="ROBOT SLIDER ADD", labelanchor="nw", width=480, height=120)
        self.label_robot_pg_add.propagate(False)
        self.label_robot_pg_add.place(x=500, y=165)

        J7Lab = Label(self.label_robot_pg_add, font=("Arial", 14), text="7th Axis")
        J7Lab.place(x=15, y=5)
        self.J7curAngEntryField = Entry(self.label_robot_pg_add, width=5, justify="center")
        self.J7curAngEntryField.place(x=95, y=9)
        self.J7jogNegBut = Button(self.label_robot_pg_add, text="-", width=3)
        self.J7jogNegBut.bind("<ButtonPress>", self.SelJ7jogNeg)
        self.J7jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J7jogNegBut.place(x=10, y=65, width=30, height=25)
        self.J7jogPosBut = Button(self.label_robot_pg_add, text="+", width=3)
        self.J7jogPosBut.bind("<ButtonPress>", self.SelJ7jogPos)
        self.J7jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J7jogPosBut.place(x=105, y=65, width=30, height=25)
        self.J7negLimLab = Label(self.label_robot_pg_add, font=("Arial", 8), text=str(-self.J7NegLim), style="Jointlim.TLabel")
        self.J7negLimLab.place(x=10, y=30)
        self.J7posLimLab = Label(self.label_robot_pg_add, font=("Arial", 8), text=str(self.J7PosLim), style="Jointlim.TLabel")
        self.J7posLimLab.place(x=110, y=30)
        self.J7slideLimLab = Label(self.label_robot_pg_add)
        self.J7slideLimLab.place(x=60, y=70)
        self.J7jogslide = Scale(self.label_robot_pg_add, from_=-self.J7NegLim, to=self.J7PosLim, length=125, orient=HORIZONTAL, command=self.J7sliderUpdate)
        self.J7jogslide.bind("<ButtonRelease-1>", self.J7sliderExecute)
        self.J7jogslide.place(x=10, y=43)
        ################################################################################

        J8Lab = Label(self.label_robot_pg_add, font=("Arial", 14), text="8th Axis")
        J8Lab.place(x=15 + 170, y=5)
        self.J8curAngEntryField = Entry(self.label_robot_pg_add, width=5, justify="center")
        self.J8curAngEntryField.place(x=95 + 170, y=9)
        self.J8jogNegBut = Button(self.label_robot_pg_add, text="-", width=3)
        self.J8jogNegBut.bind("<ButtonPress>", self.SelJ8jogNeg)
        self.J8jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J8jogNegBut.place(x=10 + 170, y=65, width=30, height=25)
        self.J8jogPosBut = Button(self.label_robot_pg_add, text="+", width=3)
        self.J8jogPosBut.bind("<ButtonPress>", self.SelJ8jogPos)
        self.J8jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J8jogPosBut.place(x=105 + 170, y=65, width=30, height=25)
        self.J8negLimLab = Label(self.label_robot_pg_add, font=("Arial", 8), text=str(-self.J8NegLim), style="Jointlim.TLabel")
        self.J8negLimLab.place(x=10 + 170, y=30)
        self.J8posLimLab = Label(self.label_robot_pg_add, font=("Arial", 8), text=str(self.J8PosLim), style="Jointlim.TLabel")
        self.J8posLimLab.place(x=110 + 170, y=30)
        self.J8slideLimLab = Label(self.label_robot_pg_add)
        self.J8slideLimLab.place(x=60 + 170, y=70)
        self.J8jogslide = Scale(self.label_robot_pg_add, from_=-self.J8NegLim, to=self.J8PosLim, length=125, orient=HORIZONTAL, command=self.J8sliderUpdate)
        self.J8jogslide.bind("<ButtonRelease-1>", self.J8sliderExecute)
        self.J8jogslide.place(x=10 + 170, y=43)
        ######################################################################################

        J9Lab = Label(self.label_robot_pg_add, font=("Arial", 14), text="9th Axis")
        J9Lab.place(x=15 + 340, y=5)
        self.J9curAngEntryField = Entry(self.label_robot_pg_add, width=5, justify="center")
        self.J9curAngEntryField.place(x=95 + 340, y=9)
        self.J9jogNegBut = Button(self.label_robot_pg_add, text="-", width=3)
        self.J9jogNegBut.bind("<ButtonPress>", self.SelJ9jogNeg)
        self.J9jogNegBut.bind("<ButtonRelease>", self.StopJog)
        self.J9jogNegBut.place(x=10 + 340, y=65, width=30, height=25)
        self.J9jogPosBut = Button(self.label_robot_pg_add, text="+", width=3)
        self.J9jogPosBut.bind("<ButtonPress>", self.SelJ9jogPos)
        self.J9jogPosBut.bind("<ButtonRelease>", self.StopJog)
        self.J9jogPosBut.place(x=105 + 340, y=65, width=30, height=25)
        self.J9negLimLab = Label(self.label_robot_pg_add, font=("Arial", 8), text=str(-self.J9NegLim), style="Jointlim.TLabel")
        self.J9negLimLab.place(x=10 + 340, y=30)
        self.J9posLimLab = Label(self.label_robot_pg_add, font=("Arial", 8), text=str(self.J9PosLim), style="Jointlim.TLabel")
        self.J9posLimLab.place(x=110 + 340, y=30)
        self.J9slideLimLab = Label(self.label_robot_pg_add)
        self.J9slideLimLab.place(x=60 + 340, y=70)
        self.J9jogslide = Scale(self.label_robot_pg_add, from_=-self.J9NegLim, to=self.J9PosLim, length=125, orient=HORIZONTAL, command=self.J9sliderUpdate)
        self.J9jogslide.bind("<ButtonRelease-1>", self.J9sliderExecute)
        self.J9jogslide.place(x=10 + 340, y=43)

        ### X ###
        self.label_robot_pg2 = tk.LabelFrame(self.tab10, text="ROBOT SLIDER 2", labelanchor="nw", width=90, height=630)
        self.label_robot_pg2.propagate(False)
        self.label_robot_pg2.place(x=980, y=10)

        self.XcurEntryField = Entry(self.label_robot_pg2, width=5, justify="center")
        self.XcurEntryField.place(x=25, y=5)
        ### Y ###
        self.YcurEntryField = Entry(self.label_robot_pg2, width=5, justify="center")
        self.YcurEntryField.place(x=25, y=55)
        ### Z ###
        self.ZcurEntryField = Entry(self.label_robot_pg2, width=5, justify="center")
        self.ZcurEntryField.place(x=25, y=105)
        ### Rz ###
        self.RzcurEntryField = Entry(self.label_robot_pg2, width=5, justify="center")
        self.RzcurEntryField.place(x=25, y=155)
        ### Ry ###
        self.RycurEntryField = Entry(self.label_robot_pg2, width=5, justify="center")
        self.RycurEntryField.place(x=25, y=205)
        ### Rx ###
        self.RxcurEntryField = Entry(self.label_robot_pg2, width=5, justify="center")
        self.RxcurEntryField.place(x=25, y=255)
        ##
        XLab = Label(self.label_robot_pg2, font=("Arial", 13), text=" X")
        XLab.place(x=0, y=5)
        YLab = Label(self.label_robot_pg2, font=("Arial", 13), text=" Y")
        YLab.place(x=0, y=55)
        ZLab = Label(self.label_robot_pg2, font=("Arial", 13), text=" Z")
        ZLab.place(x=0, y=105)
        yLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Rz")
        yLab.place(x=0, y=155)
        pLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Ry")
        pLab.place(x=0, y=205)
        rLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Rx")
        rLab.place(x=0, y=255)
        TXLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Tx")
        TXLab.place(x=0, y=305)
        TYLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Ty")
        TYLab.place(x=0, y=355)
        self.TZLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Tz")
        self.TZLab.place(x=0, y=405)
        self.TyLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Trz")
        self.TyLab.place(x=0, y=455)
        self.TpLab = Label(self.label_robot_pg2, font=("Arial", 13), text="Try")
        self.TpLab.place(x=0, y=505)
        self.J7Lab = Label(self.label_robot_pg2, font=("Arial", 13), text="Trx")
        self.J7Lab.place(x=0, y=555)
        #
        self.XjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        self.XjogNegBut.place(x=10, y=28, width=30, height=25)
        XjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        XjogPosBut.place(x=45, y=28, width=30, height=25)
        #
        YjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        YjogNegBut.place(x=10, y=78, width=30, height=25)
        YjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        YjogPosBut.place(x=45, y=78, width=30, height=25)
        #
        ZjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        ZjogNegBut.place(x=10, y=128, width=30, height=25)
        ZjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        ZjogPosBut.place(x=45, y=128, width=30, height=25)
        #
        RzjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        RzjogNegBut.place(x=10, y=178, width=30, height=25)
        RzjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        RzjogPosBut.place(x=45, y=178, width=30, height=25)
        #
        RyjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        RyjogNegBut.place(x=10, y=228, width=30, height=25)
        RyjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        RyjogPosBut.place(x=45, y=228, width=30, height=25)
        #
        RxjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        RxjogNegBut.place(x=10, y=278, width=30, height=25)
        RxjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        RxjogPosBut.place(x=45, y=278, width=30, height=25)
        #
        TXjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        TXjogNegBut.place(x=10, y=328, width=30, height=25)
        TXjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        TXjogPosBut.place(x=45, y=328, width=30, height=25)
        #
        TYjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        TYjogNegBut.place(x=10, y=378, width=30, height=25)
        TYjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        TYjogPosBut.place(x=45, y=378, width=30, height=25)
        #
        TZjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        TZjogNegBut.place(x=10, y=428, width=30, height=25)
        TZjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        TZjogPosBut.place(x=45, y=428, width=30, height=25)
        #
        TRzjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        TRzjogNegBut.place(x=10, y=478, width=30, height=25)
        TRzjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        TRzjogPosBut.place(x=45, y=478, width=30, height=25)
        #
        TRyjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        TRyjogNegBut.place(x=10, y=528, width=30, height=25)
        TRyjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        TRyjogPosBut.place(x=45, y=528, width=30, height=25)
        #
        TRxjogNegBut = Button(self.label_robot_pg2, text="-", width=3)

        TRxjogNegBut.place(x=10, y=578, width=30, height=25)
        TRxjogPosBut = Button(self.label_robot_pg2, text="+", width=3)

        TRxjogPosBut.place(x=45, y=578, width=30, height=25)
        ##########
        self.XjogNegBut.bind("<ButtonPress>", self.SelXjogNeg)  # SelXjogNeg
        self.XjogNegBut.bind("<ButtonRelease>", self.StopJog)
        XjogPosBut.bind("<ButtonPress>", self.SelXjogPos)
        XjogPosBut.bind("<ButtonRelease>", self.StopJog)
        YjogNegBut.bind("<ButtonPress>", self.SelYjogNeg)
        YjogNegBut.bind("<ButtonRelease>", self.StopJog)
        YjogPosBut.bind("<ButtonPress>", self.SelYjogPos)
        YjogPosBut.bind("<ButtonRelease>", self.StopJog)
        ZjogNegBut.bind("<ButtonPress>", self.SelZjogNeg)
        ZjogNegBut.bind("<ButtonRelease>", self.StopJog)
        ZjogPosBut.bind("<ButtonPress>", self.SelZjogPos)
        ZjogPosBut.bind("<ButtonRelease>", self.StopJog)
        RzjogNegBut.bind("<ButtonPress>", self.SelRzjogNeg)
        RzjogNegBut.bind("<ButtonRelease>", self.StopJog)
        RzjogPosBut.bind("<ButtonPress>", self.SelRzjogPos)
        RzjogPosBut.bind("<ButtonRelease>", self.StopJog)
        RyjogNegBut.bind("<ButtonPress>", self.SelRyjogNeg)
        RyjogNegBut.bind("<ButtonRelease>", self.StopJog)
        RyjogPosBut.bind("<ButtonPress>", self.SelRyjogPos)
        RyjogPosBut.bind("<ButtonRelease>", self.StopJog)
        RxjogNegBut.bind("<ButtonPress>", self.SelRxjogNeg)
        RxjogNegBut.bind("<ButtonRelease>", self.StopJog)
        RxjogPosBut.bind("<ButtonPress>", self.SelRxjogPos)
        RxjogPosBut.bind("<ButtonRelease>", self.StopJog)
        TXjogNegBut.bind("<ButtonPress>", self.SelTxjogNeg)
        TXjogNegBut.bind("<ButtonRelease>", self.StopJog)
        TXjogPosBut.bind("<ButtonPress>", self.SelTxjogPos)
        TXjogPosBut.bind("<ButtonRelease>", self.StopJog)
        TYjogNegBut.bind("<ButtonPress>", self.SelTyjogNeg)
        TYjogNegBut.bind("<ButtonRelease>", self.StopJog)
        TYjogPosBut.bind("<ButtonPress>", self.SelTyjogPos)
        TYjogPosBut.bind("<ButtonRelease>", self.StopJog)
        TZjogNegBut.bind("<ButtonPress>", self.SelTzjogNeg)
        TZjogNegBut.bind("<ButtonRelease>", self.StopJog)
        TZjogPosBut.bind("<ButtonPress>", self.SelTzjogPos)
        TZjogPosBut.bind("<ButtonRelease>", self.StopJog)
        TRzjogNegBut.bind("<ButtonPress>", self.SelTRzjogNeg)
        TRzjogNegBut.bind("<ButtonRelease>", self.StopJog)
        TRzjogPosBut.bind("<ButtonPress>", self.SelTRzjogPos)
        TRzjogPosBut.bind("<ButtonRelease>", self.StopJog)
        TRyjogNegBut.bind("<ButtonPress>", self.SelTRyjogNeg)
        TRyjogNegBut.bind("<ButtonRelease>", self.StopJog)
        TRyjogPosBut.bind("<ButtonPress>", self.SelTRyjogPos)
        TRyjogPosBut.bind("<ButtonRelease>", self.StopJog)
        TRxjogNegBut.bind("<ButtonPress>", self.SelTRxjogNeg)
        TRxjogNegBut.bind("<ButtonRelease>", self.StopJog)
        TRxjogPosBut.bind("<ButtonPress>", self.SelTRxjogPos)
        TRxjogPosBut.bind("<ButtonRelease>", self.StopJog)

        ######################################################################
        progframe = Frame(self.tab10)
        progframe.place(x=7, y=210)
        self.scrollbar = Scrollbar(progframe)
        self.scrollbar.pack(side=RIGHT, fill=Y)
        self.tab10.progView = Listbox(progframe, exportselection=0, width=77, height=22, yscrollcommand=self.scrollbar.set)
        self.tab10.progView.bind('<<ListboxSelect>>', self.progViewselect)
        self.tab10.progView.pack()
        self.scrollbar.config(command=self.tab10.progView.yview)
        self.label_command_1 = tk.LabelFrame(self.tab10, text="Manual entry code", labelanchor="nw", width=500, height=60)
        self.label_command_1.propagate(False)
        self.label_command_1.place(x=0, y=570)

        getSelBut = Button(self.label_command_1, text="Get Selected", width=12, command=self.getSel)
        getSelBut.place(x=10, y=10)

        manInsBut = Button(self.label_command_1, text="Insert", width=12, command=self.manInsItem)
        manInsBut.place(x=105, y=10)

        manRepBut = Button(self.label_command_1, text="Replace", width=12, command=self.manReplItem)
        manRepBut.place(x=200, y=10)

        openTextBut = Button(self.label_command_1, text="Open Text", width=12, command=self.openText)
        openTextBut.place(x=295, y=10)

        reloadProgBut = Button(self.label_command_1, text="Reload", width=12, command=self.reloadProg)
        reloadProgBut.place(x=390, y=10)

        self.label_program = tk.LabelFrame(self.tab10, text="Progarm command", labelanchor="nw", width=480, height=370)
        self.label_program.propagate(False)
        self.label_program.place(x=500, y=290)
        self.options = StringVar()
        # style = ttk.Style()
        # style.configure("BW.TLabel", foreground="black", background="white")
        self.menu = OptionMenu(self.label_program, self.options, "Move J", "Move J", "OFF J", "Move L",
                               "Move R", "Move A Mid", "Move A End", "Move C Center", "Move C Start",
                               "Move C Plane", "Start Spline", "End Spline", "Move PR", "OFF PR ", "Teach PR", "Move Vis",
                               command=self.posRegFieldVisible)  # ,style="BW.TLabel")
        self.menu.grid(row=2, column=2)
        self.menu.config(width=18)
        self.menu.place(x=0, y=0)

        self.SavePosEntryField = Entry(self.label_program, width=5, justify="center")
        # self.SavePosEntryField.place(x=800, y=183)

        teachInsBut = Button(self.label_program, text="Teach New Position", width=22, command=self.teachInsertBelSelected)
        teachInsBut.place(x=0, y=25)

        teachReplaceBut = Button(self.label_program, text="Modify Position", width=22, command=self.teachReplaceSelected)
        teachReplaceBut.place(x=0, y=50)

        deleteBut = Button(self.label_program, text="Delete", width=22, command=self.deleteitem)
        deleteBut.place(x=0, y=75)

        CalibrateBut = Button(self.label_program, text="Auto Calibrate CMD", width=22, command=self.insCalibrate)
        CalibrateBut.place(x=0, y=100)

        camOnBut = Button(self.label_program, text="Camera On", width=22, command=self.cameraOn)
        camOnBut.place(x=0, y=125)

        camOffBut = Button(self.label_program, text="Camera Off", width=22, command=self.cameraOff)
        camOffBut.place(x=0, y=150)

        # buttons with 1 entry

        waitTimeBut = Button(self.label_program, text="Wait Time (seconds)", width=22, command=self.waitTime)
        waitTimeBut.place(x=0, y=175)

        waitInputOnBut = Button(self.label_program, text="Wait Input ON", width=22, command=self.waitInputOn)
        waitInputOnBut.place(x=0, y=200)

        waitInputOffBut = Button(self.label_program, text="Wait Input OFF", width=22, command=self.waitInputOff)
        waitInputOffBut.place(x=0, y=225)

        setOutputOnBut = Button(self.label_program, text="Set Output On", width=22, command=self.setOutputOn)
        setOutputOnBut.place(x=0, y=250)

        setOutputOffBut = Button(self.label_program, text="Set Output OFF", width=22, command=self.setOutputOff)
        setOutputOffBut.place(x=0, y=275)

        tabNumBut = Button(self.label_program, text="Create Tab", width=22, command=self.tabNumber)
        tabNumBut.place(x=0, y=300)

        jumpTabBut = Button(self.label_program, text="Jump to Tab", width=22, command=self.jumpTab)
        jumpTabBut.place(x=0, y=325)

        self.waitTimeEntryField = Entry(self.label_program, width=5, justify="center")
        self.waitTimeEntryField.place(x=145, y=175)

        self.waitInputEntryField = Entry(self.label_program, width=5, justify="center")
        self.waitInputEntryField.place(x=145, y=200)

        self.waitInputOffEntryField = Entry(self.label_program, width=5, justify="center")
        self.waitInputOffEntryField.place(x=145, y=225)

        self.outputOnEntryField = Entry(self.label_program, width=5, justify="center")
        self.outputOnEntryField.place(x=145, y=250)

        self.outputOffEntryField = Entry(self.label_program, width=5, justify="center")
        self.outputOffEntryField.place(x=145, y=275)

        self.tabNumEntryField = Entry(self.label_program, width=5, justify="center")
        self.tabNumEntryField.place(x=145, y=300)

        self.jumpTabEntryField = Entry(self.label_program, width=5, justify="center")
        self.jumpTabEntryField.place(x=145, y=325)

        # buttons with multiple entry

        # IfOnjumpTabBut = Button(self.label_program,  text="If On Jump",  width=22,   command = IfOnjumpTab)
        # IfOnjumpTabBut.place(x=950, y=360)

        ifSelLab = Label(self.label_program, font=("Arial 10 bold"), text="IF")
        ifSelLab.place(x=150, y=0)

        self.iFoption = StringVar(self.label_program)
        iFmenu = OptionMenu(self.label_program, self.iFoption, "Input", "Input", "Register", "COM Device")
        # iFmenu.grid(row=2, column=2)
        iFmenu.config(width=5)
        iFmenu.place(x=170, y=0)

        self.IfVarEntryField = Entry(self.label_program, width=5, justify="center")
        self.IfVarEntryField.place(x=220, y=0)

        ifEqualLab = Label(self.label_program, font=("Arial 10 bold"), text="=")
        ifEqualLab.place(x=250, y=0)

        self.IfInputEntryField = Entry(self.label_program, width=5, justify="center")
        self.IfInputEntryField.place(x=265, y=0)

        self.iFselection = StringVar(self.label_program)
        iFSelmenu = OptionMenu(self.label_program, self.iFselection, "Call self.Progr", "Call self.Progr", "Jump Tab")
        iFSelmenu.grid(row=2, column=2)
        iFSelmenu.config(width=5)
        iFSelmenu.place(x=300, y=0)

        self.IfDestEntryField = Entry(self.label_program, width=9, justify="center")
        self.IfDestEntryField.place(x=360, y=0)

        ifEqualLab = Label(self.label_program, font=("Arial 10 bold"), text="•")
        ifEqualLab.place(x=390, y=0)

        insertIfCMDBut = Button(self.label_program, text="Insert IF CMD", width=10, command=self.IfCMDInsert)
        insertIfCMDBut.place(x=400, y=0)

        GCplayBut = Button(self.label_program, text="Play Gcode", width=22, command=self.insertGCprog)
        GCplayBut.place(x=200, y=30)

        readAuxComBut = Button(self.label_program, text="Read COM Device", width=22, command=self.ReadAuxCom)
        readAuxComBut.place(x=200, y=70)

        servoBut = Button(self.label_program, text="Servo", width=22, command=self.Servo)
        servoBut.place(x=200, y=110)

        RegNumBut = Button(self.label_program, text="Register", width=22, command=self.insertRegister)
        RegNumBut.place(x=200, y=150)

        StorPosBut = Button(self.label_program, text="Position Register", width=22, command=self.storPos)
        StorPosBut.place(x=200, y=190)

        callBut = Button(self.label_program, text="Call Program", width=22, command=self.insertCallProg)
        callBut.place(x=200, y=230)

        returnBut = Button(self.label_program, text="Return", width=22, command=self.insertReturn)
        returnBut.place(x=200, y=270)

        visFindBut = Button(self.label_program, text="Vision Find", width=22, command=self.insertvisFind)
        visFindBut.place(x=200, y=310)

        self.PlayGCEntryField = Entry(self.label_program, width=18, justify="center")
        self.PlayGCEntryField.place(x=300 + 50, y=30)

        self.auxPortEntryField = Entry(self.label_program, width=5, justify="center")
        self.auxPortEntryField.place(x=300 + 50, y=70)

        self.auxCharEntryField = Entry(self.label_program, width=5, justify="center")
        self.auxCharEntryField.place(x=350 + 40, y=70)

        self.servoNumEntryField = Entry(self.label_program, width=5, justify="center")
        self.servoNumEntryField.place(x=300 + 50, y=110)

        self.servoPosEntryField = Entry(self.label_program, width=5, justify="center")
        self.servoPosEntryField.place(x=350 + 40, y=110)

        self.regNumEntryField = Entry(self.label_program, width=5, justify="center")
        self.regNumEntryField.place(x=300 + 50, y=150)

        self.regEqEntryField = Entry(self.label_program, width=5, justify="center")
        self.regEqEntryField.place(x=350 + 40, y=150)

        self.storPosNumEntryField = Entry(self.label_program, width=5, justify="center")
        self.storPosNumEntryField.place(x=300 + 50, y=190)

        self.storPosElEntryField = Entry(self.label_program, width=5, justify="center")
        self.storPosElEntryField.place(x=350 + 40, y=190)

        self.storPosValEntryField = Entry(self.label_program, width=5, justify="center")
        self.storPosValEntryField.place(x=400 + 30, y=190)

        self.changeProgEntryField = Entry(self.label_program, width=18, justify="center")
        self.changeProgEntryField.place(x=300 + 50, y=230)

        self.visPassEntryField = Entry(self.label_program, width=5, justify="center")
        self.visPassEntryField.place(x=300 + 50, y=310)

        self.visFailEntryField = Entry(self.label_program, width=5, justify="center")
        self.visFailEntryField.place(x=350 + 40, y=310)

        manEntLab = Label(self.label_program, font=("Arial", 6), text="Manual Program Entry")
        manEntLab.place(x=10, y=685)

        auxComLab = Label(self.label_program, font=("Arial", 6), text="Port             Char")
        auxComLab.place(x=1107, y=429)

        servoLab = Label(self.label_program, font=("Arial", 6), text="Number      Position")
        servoLab.place(x=1107, y=469)

        regEqLab = Label(self.label_program, font=("Arial", 6), text="Register       (++/--)")
        regEqLab.place(x=1107, y=509)

        auxComLab = Label(self.label_program, font=("Arial", 6), text="Port      Characters")
        auxComLab.place(x=1107, y=549)

        storPosEqLab = Label(self.label_program, font=("Arial", 6), text=" Pos Reg      Element       (++/--)")
        storPosEqLab.place(x=1107, y=549)

        visPassLab = Label(self.label_program, font=("Arial", 6), text="Pass Tab     Fail Tab")
        visPassLab.place(x=1107, y=670)

    def caieo(self):
        print("co nhe")

    ########################################################################################################################program
    def ErrorHandler(self, response):
        global estopActive
        global posOutreach
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        ##AXIS LIMIT ERROR
        if (response[1:2] == 'L'):
            if (response[2:3] == '1'):
                message = "J1 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[3:4] == '1'):
                message = "J2 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[4:5] == '1'):
                message = "J3 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[5:6] == '1'):
                message = "J4 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[6:7] == '1'):
                message = "J5 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[7:8] == '1'):
                message = "J6 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[8:9] == '1'):
                message = "J7 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[9:10] == '1'):
                message = "J8 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[10:11] == '1'):
                message = "J9 Axis Limit"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            self.cmdRecEntryField.delete(0, 'end')
            self.cmdRecEntryField.insert(0, response)
            message = "Axis Limit Error - See Log"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")
            # self.stopProg()
        ##COLLISION ERROR
        elif (response[1:2] == 'C'):
            if (response[2:3] == '1'):
                message = "J1 Collision or Motor Error"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
                self.correctPos()
                self.stopProg()
                message = "Collision or Motor Error - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")
            if (response[3:4] == '1'):
                message = "J2 Collision or Motor Error"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
                self.correctPos()
                self.stopProg()
                message = "Collision or Motor Error - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")
            if (response[4:5] == '1'):
                message = "J3 Collision or Motor Error"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
                self.correctPos()
                self.stopProg()
                message = "Collision or Motor Error - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")
            if (response[5:6] == '1'):
                message = "J4 Collision or Motor Error"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
                self.correctPos()
                self.stopProg()
                message = "Collision or Motor Error - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")
            if (response[6:7] == '1'):
                message = "J5 Collision or Motor Error"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
                self.correctPos()
                self.stopProg()
                message = "Collision or Motor Error - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")
            if (response[7:8] == '1'):
                message = "J6 Collision or Motor Error"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
                self.correctPos()
                self.stopProg()
                message = "Collision or Motor Error - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")

        ##REACH ERROR
        elif (response[1:2] == 'R'):
            posOutreach = TRUE
            self.stopProg()
            message = "Position Out of Reach"
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")

        ##SPLINE ERROR
        elif (response[1:2] == 'S'):
            self.stopProg()
            message = "Spline Can Only Have Move L Types"
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")

        ##GCODE ERROR
        elif (response[1:2] == 'G'):
            self.stopProg()
            message = "Gcode file not found"
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")

        ##ESTOP BUTTON
        elif (response[1:2] == 'B'):
            estopActive = TRUE
            self.stopProg()
            message = "Estop Button was Pressed"
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")

            ##CALIBRATION ERROR
        elif (response[1:2] == 'A'):
            if (response[2:3] == '1'):
                message = "J1 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '2'):
                message = "J2 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '3'):
                message = "J3 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '4'):
                message = "J4 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '5'):
                message = "J5 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '6'):
                message = "J6 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '7'):
                message = "J7 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '8'):
                message = "J8 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            if (response[2:3] == '9'):
                message = "J9 CALIBRATION ERROR"
                self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))

        else:
            self.stopProg()
            message = "Unknown Error"
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.GCalmStatusLab.config(text=message, style="Alarm.TLabel")

    def runProg(self):
        def threadProg():
            global rowinproc

            global splineActive
            global estopActive
            estopActive = False
            global posOutreach
            posOutreach = False
            self.stopQueue = "0"
            splineActive = "0"
            try:
                curRow = self.tab10.progView.curselection()[0]
                if (curRow == 0):
                    curRow = 1
            except:
                curRow = 1
                self.tab10.progView.selection_clear(0, END)
                self.tab10.progView.select_set(curRow)
            self.tab10.runTrue = 1
            while self.tab10.runTrue == 1:
                if (self.tab10.runTrue == 0):
                    if (estopActive == TRUE):
                        self.almStatusLab.config(text="Estop Button was Pressed", style="Alarm.TLabel")
                        self.almStatusLab2.config(text="Estop Button was Pressed", style="Alarm.TLabel")
                    elif (posOutreach == TRUE):
                        self.almStatusLab.config(text="Position Out of Reach", style="Alarm.TLabel")
                        self.almStatusLab2.config(text="Position Out of Reach", style="Alarm.TLabel")
                    else:
                        self.almStatusLab.config(text="PROGRAM STOPPED", style="Alarm.TLabel")
                        self.almStatusLab2.config(text="PROGRAM STOPPED", style="Alarm.TLabel")
                else:
                    self.almStatusLab.config(text="PROGRAM RUNNING", style="OK.TLabel")
                    self.almStatusLab2.config(text="PROGRAM RUNNING", style="OK.TLabel")
                rowinproc = 1
                self.executeRow()
                while rowinproc == 1:
                    time.sleep(.1)
                selRow = self.tab10.progView.curselection()[0]
                last = self.tab10.progView.index('end')
                # for row in range (0,selRow):
                # self.tab10.progView.itemconfig(row, {'fg': 'dodger blue'})
                # self.tab10.progView.itemconfig(selRow, {'fg': 'blue2'})
                # for row in range (selRow+1,last):
                # self.tab10.progView.itemconfig(row, {'fg': 'black'})
                self.tab10.progView.selection_clear(0, END)
                selRow += 1
                self.tab10.progView.select_set(selRow)
                curRow += 1
                time.sleep(.1)
                try:
                    selRow = self.tab10.progView.curselection()[0]
                    self.curRowEntryField.delete(0, 'end')
                    self.curRowEntryField.insert(0, selRow)
                except:
                    self.curRowEntryField.delete(0, 'end')
                    self.curRowEntryField.insert(0, "---")
                    self.tab10.runTrue = 0
                    if (estopActive == TRUE):
                        self.almStatusLab.config(text="Estop Button was Pressed", style="Alarm.TLabel")
                        self.almStatusLab2.config(text="Estop Button was Pressed", style="Alarm.TLabel")
                    elif (posOutreach == TRUE):
                        self.almStatusLab.config(text="Position Out of Reach", style="Alarm.TLabel")
                        self.almStatusLab2.config(text="Position Out of Reach", style="Alarm.TLabel")
                    else:
                        self.almStatusLab.config(text="PROGRAM STOPPED", style="Alarm.TLabel")
                        self.almStatusLab2.config(text="PROGRAM STOPPED", style="Alarm.TLabel")

        t = threading.Thread(target=threadProg)
        t.start()

    def stepFwd(self):
        global estopActive
        estopActive = False
        global posOutreach
        posOutreach = False
        self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
        self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        self.executeRow()
        selRow = self.tab10.progView.curselection()[0]
        last = self.tab10.progView.index('end')
        for row in range(0, selRow):
            self.tab10.progView.itemconfig(row, {'fg': 'dodger blue'})
        self.tab10.progView.itemconfig(selRow, {'fg': 'blue2'})
        for row in range(selRow + 1, last):
            self.tab10.progView.itemconfig(row, {'fg': 'black'})
        self.tab10.progView.selection_clear(0, END)
        selRow += 1
        self.tab10.progView.select_set(selRow)
        try:
            selRow = self.tab10.progView.curselection()[0]
            self.curRowEntryField.delete(0, 'end')
            self.curRowEntryField.insert(0, selRow)
        except:
            self.curRowEntryField.delete(0, 'end')
            self.curRowEntryField.insert(0, "---")

    def stepRev(self):
        global estopActive
        estopActive = False
        global posOutreach
        posOutreach = False
        self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
        self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        self.executeRow()
        selRow = self.tab10.progView.curselection()[0]
        last = self.tab10.progView.index('end')
        for row in range(0, selRow):
            self.tab10.progView.itemconfig(row, {'fg': 'black'})
        self.tab10.progView.itemconfig(selRow, {'fg': 'red'})
        for row in range(selRow + 1, last):
            self.tab10.progView.itemconfig(row, {'fg': 'tomato2'})
        self.tab10.progView.selection_clear(0, END)
        selRow -= 1
        self.tab10.progView.select_set(selRow)
        try:
            selRow = self.tab10.progView.curselection()[0]
            self.curRowEntryField.delete(0, 'end')
            self.curRowEntryField.insert(0, selRow)
        except:
            self.curRowEntryField.delete(0, 'end')
            self.curRowEntryField.insert(0, "---")

    def stopProg(self):
        global cmdType
        global splineActive
        global estopActive
        global posOutreach

        lastProg = ""
        self.tab10.runTrue = 0
        if (estopActive == TRUE):
            self.almStatusLab.config(text="Estop Button was Pressed", style="Alarm.TLabel")
            self.almStatusLab2.config(text="Estop Button was Pressed", style="Alarm.TLabel")
        elif (posOutreach == TRUE):
            self.almStatusLab.config(text="Position Out of Reach", style="Alarm.TLabel")
            self.almStatusLab2.config(text="Position Out of Reach", style="Alarm.TLabel")
        else:
            self.almStatusLab.config(text="PROGRAM STOPPED", style="Alarm.TLabel")
            self.almStatusLab2.config(text="PROGRAM STOPPED", style="Alarm.TLabel")

    def executeRow(self):

        global calStat
        global rowinproc
        global LineDist
        global Xv
        global Yv
        global Zv
        global commandCalc

        global splineActive

        selRow = self.tab10.progView.curselection()[0]
        self.tab10.progView.see(selRow + 2)
        data = list(map(int, self.tab10.progView.curselection()))
        command = self.tab10.progView.get(data[0]).decode().strip()
        cmdType = command[:6]

        ##Call Program##
        if (cmdType == "Call P"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            self.tab10.lastRow = self.tab10.progView.curselection()[0]
            self.tab10.lastProg = self.ProgEntryField.get()
            programIndex = command.find("Program -")
            progNum = str(command[programIndex + 10:])
            self.ProgEntryField.delete(0, 'end')
            self.ProgEntryField.insert(0, progNum)
            self.callProg(progNum)
            time.sleep(.4)
            index = 0
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(index)

            ##Run Gcode Program##
        if (cmdType == "Run Gc"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            self.tab10.lastRow = self.tab10.progView.curselection()[0]
            self.tab10.lastProg = self.ProgEntryField.get()
            programIndex = command.find("Program -")
            filename = str(command[programIndex + 10:])
            self.manEntryField.delete(0, 'end')
            self.manEntryField.insert(0, filename)
            self.GCplayProg(filename)
            time.sleep(.4)
            index = 0
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(index)

            ##Return Program##
        if (cmdType == "Return"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            lastRow = self.tab10.lastRow
            lastProg = self.tab10.lastProg
            self.ProgEntryField.delete(0, 'end')
            self.ProgEntryField.insert(0, lastProg)
            self.callProg(lastProg)
            time.sleep(.4)
            index = 0
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(lastRow)
            ##Test Limit Switches
        if (cmdType == "Test L"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            command = "TL\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.05)
            response = str(self.ser.readline().strip(), 'utf-8')
            self.manEntryField.delete(0, 'end')
            self.manEntryField.insert(0, response)
        ##Set Encoders 1000
        if (cmdType == "Set En"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            command = "SE\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.05)
            self.ser.read()
            ##Read Encoders
        if (cmdType == "Read E"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            command = "RE\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.05)
            response = str(self.ser.readline().strip(), 'utf-8')
            self.manEntryField.delete(0, 'end')
            self.manEntryField.insert(0, response)
            ##Servo Command##
        if (cmdType == "Servo "):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            servoIndex = command.find("number ")
            posIndex = command.find("position: ")
            servoNum = str(command[servoIndex + 7:posIndex - 4])
            servoPos = str(command[posIndex + 10:])
            command = "SV" + servoNum + "P" + servoPos + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser_IO.write(command.encode())
            self.ser_IO.flushInput()
            time.sleep(.1)
            self.ser_IO.read()

            ##If Input##
        if (cmdType == "If Inp"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("# ")
            valIndex = command.find("= ")
            actionIndex = command.find(": ")
            inputNum = str(command[inputIndex + 2:valIndex])
            valNum = int(command[valIndex + 2:actionIndex - 1])
            action = str(command[actionIndex + 2:actionIndex + 6])
            ##querry board for IO value
            cmd = "JFX" + inputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, cmd)
            self.ser_IO.write(cmd.encode())
            self.ser_IO.flushInput()
            time.sleep(.1)
            response = str(self.ser_IO.readline().strip(), 'utf-8')
            if (response == "T"):
                querry = 1
            elif (response == "F"):
                querry = 0
            if (querry == valNum):
                if (action == "Call"):
                    self.tab10.lastRow = self.tab10.progView.curselection()[0]
                    self.tab10.lastProg = self.ProgEntryField.get()
                    progIndex = command.find("self.Progr")
                    progName = str(command[progIndex + 5:]) + ".ar4"
                    self.callProg(progName)
                    time.sleep(.4)
                    index = 0
                    self.tab10.progView.selection_clear(0, END)
                    self.tab10.progView.select_set(index)
                elif (action == "Jump"):
                    tabIndex = command.find("Tab")
                    tabNum = str(command[tabIndex + 4:])
                    tabNum = ("Tab Number " + tabNum + "\r\n").encode('utf-8')
                    index = self.tab10.progView.get(0, "end").index(tabNum)
                    index = index - 1
                    self.tab10.progView.selection_clear(0, END)
                    self.tab10.progView.select_set(index)

        ##If Read Com##
        if (cmdType == "Read C"):
            comIndex = command.find("# ")
            charIndex = command.find("Char: ")
            actionIndex = command.find(": ")
            comNum = str(command[comIndex + 2:charIndex - 1])
            charNum = int(command[charIndex + 6:])
            try:
                global ser3
                port = "COM" + comNum
                baud = 115200
                ser3 = serial.Serial(port, baud, timeout=10)
            except:
                Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
                self.tab5_rb.ElogView.insert(END, Curtime + " - UNABLE TO ESTABLISH COMMUNICATIONS WITH SERIAL DEVICE")
                value = self.tab5_rb.ElogView.get(0, END)
                pickle.dump(value, open("ErrorLog", "wb"))
            ser3.flushInput()
            response = str(ser3.read(charNum).strip(), 'utf-8')
            self.com3outPortEntryField.delete(0, 'end')
            self.com3outPortEntryField.insert(0, response)
            self.manEntryField.delete(0, 'end')
            self.manEntryField.insert(0, response)

        ##If Register##
        if (cmdType == "If Reg"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("# ")
            valIndex = command.find("= ")
            actionIndex = command.find(": ")
            inputNum = str(command[inputIndex + 2:valIndex - 1])
            valNum = int(command[valIndex + 2:actionIndex - 1])
            action = str(command[actionIndex + 2:actionIndex + 6])
            ##querry board for IO value
            regEntry = "R" + inputNum + "EntryField"
            curRegVal = eval(regEntry).get()
            if (int(curRegVal) == valNum):
                if (action == "Call"):
                    self.tab10.lastRow = self.tab10.progView.curselection()[0]
                    self.tab10.lastProg = self.ProgEntryField.get()
                    progIndex = command.find("self.Progr")
                    progName = str(command[progIndex + 5:]) + ".ar4"
                    self.callProg(progName)
                    time.sleep(.4)
                    index = 0
                    self.tab10.progView.selection_clear(0, END)
                    self.tab10.progView.select_set(index)
                elif (action == "Jump"):
                    tabIndex = command.find("Tab")
                    tabNum = str(command[tabIndex + 4:])
                    tabNum = ("Tab Number " + tabNum + "\r\n").encode('utf-8')
                    index = self.tab10.progView.get(0, "end").index(tabNum)
                    index = index - 1
                    self.tab10.progView.selection_clear(0, END)
                    self.tab10.progView.select_set(index)

        ##If COM device##
        if (cmdType == "If COM"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("# ")
            valIndex = command.find("= ")
            actionIndex = command.find(": ")
            inputNum = str(command[inputIndex + 2:valIndex - 1])
            valNum = str(command[valIndex + 2:actionIndex - 1])
            action = str(command[actionIndex + 2:actionIndex + 6])
            curCOMVal = self.com3outPortEntryField.get()
            if (curCOMVal == valNum):
                if (action == "Call"):
                    self.tab10.lastRow = self.tab10.progView.curselection()[0]
                    self.tab10.lastProg = self.ProgEntryField.get()
                    progIndex = command.find("self.Progr")
                    progName = str(command[actionIndex + 12:]) + ".ar4"
                    self.callProg(progName)
                    time.sleep(.4)
                    index = 0
                    self.tab10.progView.selection_clear(0, END)
                    self.tab10.progView.select_set(index)
                elif (action == "Jump"):
                    tabIndex = command.find("Tab")
                    tabNum = str(command[tabIndex + 4:])
                    tabNum = ("Tab Number " + tabNum + "\r\n").encode('utf-8')
                    index = self.tab10.progView.get(0, "end").index(tabNum)
                    index = index - 1
                    self.tab10.progView.selection_clear(0, END)
                    self.tab10.progView.select_set(index)

                    ##If Input On Jump to Tab Teensy##
        if (cmdType == "TifOn "):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("Input-")
            tabIndex = command.find("Tab-")
            inputNum = str(command[inputIndex + 6:tabIndex - 9])
            tabNum = str(command[tabIndex + 4:])
            command = "JFX" + inputNum + "T" + tabNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response == "T"):
                tabNum = ("Tab Number " + str(command[tabIndex + 4:])).encode('utf-8')
                index = self.tab10.progView.get(0, "end").index(tabNum)
                index = index - 1
                self.tab10.progView.selection_clear(0, END)
                self.tab10.progView.select_set(index)
        ##If Input Off Jump to Tab Teensy##
        if (cmdType == "TifOff"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("Input-")
            tabIndex = command.find("Tab-")
            inputNum = str(command[inputIndex + 6:tabIndex - 9])
            tabNum = str(command[tabIndex + 4:])
            command = "JFX" + inputNum + "T" + tabNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response == "F"):
                tabNum = ("Tab Number " + str(command[tabIndex + 4:])).encode('utf-8')
                index = self.tab10.progView.get(0, "end").index(tabNum)
                index = index - 1
                self.tab10.progView.selection_clear(0, END)
                self.tab10.progView.select_set(index)

        ##Jump to Row##
        if (cmdType == "Jump T"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            tabIndex = command.find("Tab-")
            tabNum = ("Tab Number " + str(command[tabIndex + 4:]) + "\r\n").encode('utf-8')
            index = self.tab10.progView.get(0, "end").index(tabNum)
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(index)

            ##Set Output ON Command IO Board##
        if (cmdType == "Out On"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            outputIndex = command.find("Out On = ")
            outputNum = str(command[outputIndex + 9:])
            command = "ONX" + outputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser_IO.write(command.encode())
            self.ser_IO.flushInput()
            time.sleep(.1)
            self.ser_IO.read()
            ##Set Output OFF Command IO Board##
        if (cmdType == "Out Of"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            outputIndex = command.find("Out Off = ")
            outputNum = str(command[outputIndex + 10:])
            command = "OFX" + outputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser_IO.write(command.encode())
            self.ser_IO.flushInput()
            time.sleep(.1)
            self.ser_IO.read()

            ##Set Output ON Command Teensy##
        if (cmdType == "ToutOn"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            outputIndex = command.find("outOn = ")
            outputNum = str(command[outputIndex + 8:])
            command = "ONX" + outputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()
            ##Set Output OFF Command Teensy##
        if (cmdType == "ToutOf"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            outputIndex = command.find("outOff = ")
            outputNum = str(command[outputIndex + 9:])
            command = "OFX" + outputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()

            ##Wait Input ON Command IO Board##
        if (cmdType == "Wait I"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("Wait Input On = ")
            inputNum = str(command[inputIndex + 16:])
            command = "WIN" + inputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser_IO.write(command.encode())
            self.ser_IO.flushInput()
            time.sleep(.1)
            self.ser_IO.read()
            ##Wait Input OFF Command IO Board##
        if (cmdType == "Wait O"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("Wait Off Input = ")
            inputNum = str(command[inputIndex + 17:])
            command = "WON" + inputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser_IO.write(command.encode())
            self.ser_IO.flushInput()
            time.sleep(.1)
            self.ser_IO.read()

            ##Wait Input ON Command Teensy##
        if (cmdType == "TwaitI"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("TwaitInput On = ")
            inputNum = str(command[inputIndex + 16:])
            command = "WIN" + inputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()
            ##Wait Input OFF Command Teensy##
        if (cmdType == "TwaitO"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            inputIndex = command.find("TwaitOff Input = ")
            inputNum = str(command[inputIndex + 16:])
            command = "WON" + inputNum + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()

            ##Wait Time Command##
        if (cmdType == "Wait T"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            timeIndex = command.find("Wait Time = ")
            timeSeconds = str(command[timeIndex + 12:])
            command = "WTS" + timeSeconds + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()

            ##Set Register##
        if (cmdType == "Regist"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            regNumIndex = command.find("Register ")
            regEqIndex = command.find(" = ")
            regNumVal = str(command[regNumIndex + 9:regEqIndex])
            regEntry = "R" + regNumVal + "EntryField"
            testOper = str(command[regEqIndex + 3:regEqIndex + 5])
            if (testOper == "++"):
                regCEqVal = str(command[regEqIndex + 5:])
                curRegVal = eval(regEntry).get()
                regEqVal = str(int(regCEqVal) + int(curRegVal))
            elif (testOper == "--"):
                regCEqVal = str(command[regEqIndex + 5:])
                curRegVal = eval(regEntry).get()
                regEqVal = str(int(curRegVal) - int(regCEqVal))
            else:
                regEqVal = str(command[regEqIndex + 3:])
            eval(regEntry).delete(0, 'end')
            eval(regEntry).insert(0, regEqVal)
        ##Set Position Register##
        if (cmdType == "Positi"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            regNumIndex = command.find("Position Register ")
            regElIndex = command.find("Element")
            regEqIndex = command.find(" = ")
            regNumVal = str(command[regNumIndex + 18:regElIndex - 1])
            regNumEl = str(command[regElIndex + 8:regEqIndex])
            regEntry = "SP_" + regNumVal + "_E" + regNumEl + "_EntryField"
            testOper = str(command[regEqIndex + 3:regEqIndex + 5])
            if (testOper == "++"):
                regCEqVal = str(command[regEqIndex + 4:])
                curRegVal = eval(regEntry).get()
                regEqVal = str(float(regCEqVal) + float(curRegVal))
            elif (testOper == "--"):
                regCEqVal = str(command[regEqIndex + 5:])
                curRegVal = eval(regEntry).get()
                regEqVal = str(float(curRegVal) - float(regCEqVal))
            else:
                regEqVal = str(command[regEqIndex + 3:])
            eval(regEntry).delete(0, 'end')
            eval(regEntry).insert(0, regEqVal)

        ##Calibrate Command##
        if (cmdType == "Calibr"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            self.calRobotAll()
            if (calStat == 0):
                self.stopProg()

        ##Set tool##
        if (cmdType == "Tool S"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
            xIndex = command.find(" X ")
            yIndex = command.find(" Y ")
            zIndex = command.find(" Z ")
            rzIndex = command.find(" Rz ")
            ryIndex = command.find(" Ry ")
            rxIndex = command.find(" Rx ")
            xVal = command[xIndex + 3:yIndex]
            yVal = command[yIndex + 3:zIndex]
            zVal = command[zIndex + 3:rzIndex]
            rzVal = command[rzIndex + 4:ryIndex]
            ryVal = command[ryIndex + 4:rxIndex]
            rxVal = command[rxIndex + 4:]
            self.TFxEntryField.delete(0, 'end')
            self.TFyEntryField.delete(0, 'end')
            self.TFzEntryField.delete(0, 'end')
            self.TFrzEntryField.delete(0, 'end')
            self.TFryEntryField.delete(0, 'end')
            self.TFrxEntryField.delete(0, 'end')
            self.TFxEntryField.insert(0, str(xVal))
            self.TFyEntryField.insert(0, str(yVal))
            self.TFzEntryField.insert(0, str(zVal))
            self.TFrzEntryField.insert(0, str(rzVal))
            self.TFryEntryField.insert(0, str(ryVal))
            self.TFrxEntryField.insert(0, str(rxVal))
            command = "TF" + "A" + xVal + "B" + yVal + "C" + zVal + "D" + rzVal + "E" + ryVal + "F" + rxVal + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()

        ##Move J Command##
        if (cmdType == "Move J"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            xIndex = command.find(" X ")
            yIndex = command.find(" Y ")
            zIndex = command.find(" Z ")
            rzIndex = command.find(" Rz ")
            ryIndex = command.find(" Ry ")
            rxIndex = command.find(" Rx ")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            WristConfIndex = command.find(" $")
            xVal = command[xIndex + 3:yIndex]
            yVal = command[yIndex + 3:zIndex]
            zVal = command[zIndex + 3:rzIndex]
            rzVal = command[rzIndex + 4:ryIndex]
            ryVal = command[ryIndex + 4:rxIndex]
            rxVal = command[rxIndex + 4:J7Index]
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            # self.ser.read()
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##Offs J Command##
        if (cmdType == "OFF J "):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            SPnewInex = command.find("[ PR: ")
            SPendInex = command.find(" ] [")
            xIndex = command.find(" X ")
            yIndex = command.find(" Y ")
            zIndex = command.find(" Z ")
            rzIndex = command.find(" Rz ")
            ryIndex = command.find(" Ry ")
            rxIndex = command.find(" Rx ")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            WristConfIndex = command.find(" $")
            SP = str(command[SPnewInex + 6:SPendInex])
            cx = eval("SP_" + SP + "_E1_EntryField").get()
            cy = eval("SP_" + SP + "_E2_EntryField").get()
            cz = eval("SP_" + SP + "_E3_EntryField").get()
            crz = eval("SP_" + SP + "_E4_EntryField").get()
            cry = eval("SP_" + SP + "_E5_EntryField").get()
            crx = eval("SP_" + SP + "_E6_EntryField").get()
            xVal = str(float(cx) + float(command[xIndex + 3:yIndex]))
            yVal = str(float(cy) + float(command[yIndex + 3:zIndex]))
            zVal = str(float(cz) + float(command[zIndex + 3:rzIndex]))
            rzVal = str(float(crz) + float(command[rzIndex + 4:ryIndex]))
            ryVal = str(float(cry) + float(command[ryIndex + 4:rxIndex]))
            rxVal = str(float(crx) + float(command[rxIndex + 4:J7Index]))
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##Move Vis Command##
        if (cmdType == "Move V"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            SPnewInex = command.find("[ PR: ")
            SPendInex = command.find(" ] [")
            xIndex = command.find(" X ")
            yIndex = command.find(" Y ")
            zIndex = command.find(" Z ")
            rzIndex = command.find(" Rz ")
            ryIndex = command.find(" Ry ")
            rxIndex = command.find(" Rx ")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            WristConfIndex = command.find(" $")
            SP = str(command[SPnewInex + 6:SPendInex])
            cx = eval("SP_" + SP + "_E1_EntryField").get()
            cy = eval("SP_" + SP + "_E2_EntryField").get()
            cz = eval("SP_" + SP + "_E3_EntryField").get()
            crz = eval("SP_" + SP + "_E4_EntryField").get()
            cry = eval("SP_" + SP + "_E5_EntryField").get()
            crx = eval("SP_" + SP + "_E6_EntryField").get()
            xVal = str(float(cx) + float(self.VisRetXrobEntryField.get()))
            yVal = str(float(cy) + float(self.VisRetYrobEntryField.get()))
            zVal = str(float(cz) + float(command[zIndex + 3:rzIndex]))
            rzVal = str(float(crz) + float(command[rzIndex + 4:ryIndex]))
            ryVal = str(float(cry) + float(command[ryIndex + 4:rxIndex]))
            rxVal = str(float(crx) + float(command[rxIndex + 4:J7Index]))
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            visRot = self.VisRetAngleEntryField.get()
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            command = "MV" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Vr" + visRot + "Lm" + self.LoopMode + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##Move PR Command##
        if (cmdType == "Move P"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            SPnewInex = command.find("[ PR: ")
            SPendInex = command.find(" ] [")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            WristConfIndex = command.find(" $")
            SP = str(command[SPnewInex + 6:SPendInex])
            cx = eval("SP_" + SP + "_E1_EntryField").get()
            cy = eval("SP_" + SP + "_E2_EntryField").get()
            cz = eval("SP_" + SP + "_E3_EntryField").get()
            crz = eval("SP_" + SP + "_E4_EntryField").get()
            cry = eval("SP_" + SP + "_E5_EntryField").get()
            crx = eval("SP_" + SP + "_E6_EntryField").get()
            xVal = str(float(cx))
            yVal = str(float(cy))
            zVal = str(float(cz))
            rzVal = str(float(crz))
            ryVal = str(float(cry))
            rxVal = str(float(crx))
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##OFFS PR Command##
        if (cmdType == "OFF PR"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            SPnewInex = command.find("[ PR: ")
            SPendInex = command.find(" ] offs")
            SP2newInex = command.find("[ *PR: ")
            SP2endInex = command.find(" ]  [")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            WristConfIndex = command.find(" $")
            SP = str(command[SPnewInex + 6:SPendInex])
            SP2 = str(command[SP2newInex + 7:SP2endInex])
            xVal = str(float(eval("SP_" + SP + "_E1_EntryField").get()) + float(eval("SP_" + SP2 + "_E1_EntryField").get()))
            yVal = str(float(eval("SP_" + SP + "_E2_EntryField").get()) + float(eval("SP_" + SP2 + "_E2_EntryField").get()))
            zVal = str(float(eval("SP_" + SP + "_E3_EntryField").get()) + float(eval("SP_" + SP2 + "_E3_EntryField").get()))
            rzVal = str(float(eval("SP_" + SP + "_E4_EntryField").get()) + float(eval("SP_" + SP2 + "_E4_EntryField").get()))
            ryVal = str(float(eval("SP_" + SP + "_E5_EntryField").get()) + float(eval("SP_" + SP2 + "_E5_EntryField").get()))
            rxVal = str(float(eval("SP_" + SP + "_E6_EntryField").get()) + float(eval("SP_" + SP2 + "_E6_EntryField").get()))
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##Move L Command##
        if (cmdType == "Move L"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            xIndex = command.find(" X ")
            yIndex = command.find(" Y ")
            zIndex = command.find(" Z ")
            rzIndex = command.find(" Rz ")
            ryIndex = command.find(" Ry ")
            rxIndex = command.find(" Rx ")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            RoundingIndex = command.find(" Rnd ")
            WristConfIndex = command.find(" $")
            xVal = command[xIndex + 3:yIndex]
            yVal = command[yIndex + 3:zIndex]
            zVal = command[zIndex + 3:rzIndex]
            rzVal = command[rzIndex + 4:ryIndex]
            if (np.sign(float(rzVal)) != np.sign(float(self.RzcurPos))):
                rzVal = str(float(rzVal) * -1)
            ryVal = command[ryIndex + 4:rxIndex]
            rxVal = command[rxIndex + 4:J7Index]
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:RoundingIndex]
            Rounding = command[RoundingIndex + 5:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            DisWrist = str(self.DisableWristRot.get())
            command = "ML" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "Rnd" + Rounding + "W" + self.WC + "Lm" + self.LoopMode + "Q" + DisWrist + "\n"

            # self.tab5_rb.ElogView.insert(END, command)
            # value=self.tab5_rb.ElogView.get(0,END)
            # pickle.dump(value,open("ErrorLog","wb"))
            start = time.time()
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            end = time.time()
            # manEntryField.delete(0, 'end')
            # manEntryField.insert(0,end-start)

            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

        ##Move R Command##
        if (cmdType == "Move R"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            J1Index = command.find(" J1 ")
            J2Index = command.find(" J2 ")
            J3Index = command.find(" J3 ")
            J4Index = command.find(" J4 ")
            J5Index = command.find(" J5 ")
            J6Index = command.find(" J6 ")
            J7Index = command.find(" J7 ")
            J8Index = command.find(" J8 ")
            J9Index = command.find(" J9 ")
            SpeedIndex = command.find(" S")
            ACCspdIndex = command.find(" Ac ")
            DECspdIndex = command.find(" Dc ")
            ACCrampIndex = command.find(" Rm ")
            WristConfIndex = command.find(" $")
            J1Val = command[J1Index + 4:J2Index]
            J2Val = command[J2Index + 4:J3Index]
            J3Val = command[J3Index + 4:J4Index]
            J4Val = command[J4Index + 4:J5Index]
            J5Val = command[J5Index + 4:J6Index]
            J6Val = command[J6Index + 4:J7Index]
            J7Val = command[J7Index + 4:J8Index]
            J8Val = command[J8Index + 4:J9Index]
            J9Val = command[J9Index + 4:SpeedIndex]
            self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
            self.Speed = command[SpeedIndex + 4:ACCspdIndex]
            self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
            self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
            self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
            self.WC = command[WristConfIndex + 3:]
            self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
            command = "RJ" + "A" + J1Val + "B" + J2Val + "C" + J3Val + "D" + J4Val + "E" + J5Val + "F" + J6Val + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##Move A Command##
        if (cmdType == "Move A"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            subCmd = command[:10]
            if (subCmd == "Move A End"):
                self.almStatusLab.config(text="Move A must start with a Mid followed by End", style="Alarm.TLabel")
                self.almStatusLab2.config(text="Move A must start with a Mid followed by End", style="Alarm.TLabel")
            else:
                xIndex = command.find(" X ")
                yIndex = command.find(" Y ")
                zIndex = command.find(" Z ")
                rzIndex = command.find(" Rz ")
                ryIndex = command.find(" Ry ")
                rxIndex = command.find(" Rx ")
                trIndex = command.find(" Tr ")
                SpeedIndex = command.find(" S")
                ACCspdIndex = command.find(" Ac ")
                DECspdIndex = command.find(" Dc ")
                ACCrampIndex = command.find(" Rm ")
                WristConfIndex = command.find(" $")
                xVal = command[xIndex + 3:yIndex]
                yVal = command[yIndex + 3:zIndex]
                zVal = command[zIndex + 3:rzIndex]
                rzVal = command[rzIndex + 4:ryIndex]
                ryVal = command[ryIndex + 4:rxIndex]
                rxVal = command[rxIndex + 4:trIndex]
                trVal = command[trIndex + 4:SpeedIndex]
                self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
                self.Speed = command[SpeedIndex + 4:ACCspdIndex]
                self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
                self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
                self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
                self.WC = command[WristConfIndex + 3:]
                TCX = 0
                TCY = 0
                TCZ = 0
                TCRx = 0
                TCRy = 0
                TCRz = 0
                ##read next row for End position
                curRow = self.tab10.progView.curselection()[0]
                selRow = self.tab10.progView.curselection()[0]
                last = self.tab10.progView.index('end')
                for row in range(0, selRow):
                    self.tab10.progView.itemconfig(row, {'fg': 'dodger blue'})
                self.tab1.progView.itemconfig(selRow, {'fg': 'blue2'})
                for row in range(selRow + 1, last):
                    self.tab10.progView.itemconfig(row, {'fg': 'black'})
                self.tab10.progView.selection_clear(0, END)
                selRow += 1
                self.tab10.progView.select_set(selRow)
                curRow += 1
                selRow = self.tab10.progView.curselection()[0]
                self.tab10.progView.see(selRow + 2)
                data = list(map(int, self.tab10.progView.curselection()))
                command = self.tab10.progView.get(data[0]).decode()
                xIndex = command.find(" X ")
                yIndex = command.find(" Y ")
                zIndex = command.find(" Z ")
                rzIndex = command.find(" Rz ")
                ryIndex = command.find(" Ry ")
                rxIndex = command.find(" Rx ")
                trIndex = command.find(" Tr ")
                SpeedIndex = command.find(" S")
                ACCspdIndex = command.find(" Ac ")
                DECspdIndex = command.find(" Dc ")
                ACCrampIndex = command.find(" Rm ")
                WristConfIndex = command.find(" $")
                Xend = command[xIndex + 3:yIndex]
                Yend = command[yIndex + 3:zIndex]
                Zend = command[zIndex + 3:rzIndex]
                rzVal = command[rzIndex + 4:ryIndex]
                ryVal = command[ryIndex + 4:rxIndex]
                rxVal = command[rxIndex + 4:trIndex]
                trVal = command[trIndex + 4:SpeedIndex]
                self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
                self.Speed = command[SpeedIndex + 4:ACCspdIndex]
                self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
                self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
                self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
                self.WC = command[WristConfIndex + 3:]
                TCX = 0
                TCY = 0
                TCZ = 0
                TCRx = 0
                TCRy = 0
                TCRz = 0
                # move arc command
                self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
                command = "MA" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "Ex" + Xend + "Ey" + Yend + "Ez" + Zend + "Tr" + trVal + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
                self.cmdSentEntryField.delete(0, 'end')
                self.cmdSentEntryField.insert(0, command)
                self.ser.write(command.encode())
                self.ser.flushInput()
                time.sleep(.1)
                response = str(self.ser.readline().strip(), 'utf-8')
                if (response[:1] == 'E'):
                    self.ErrorHandler(response)
                else:
                    self.displayPosition(response)

                    ##Move C Command##
        if (cmdType == "Move C"):
            if (self.moveInProc == 0):
                self.moveInProc == 1
            subCmd = command[:10]
            if (subCmd == "Move C Sta" or subCmd == "Move C Pla"):
                self.almStatusLab.config(text="Move C must start with a Center followed by Start & Plane", style="Alarm.TLabel")
                self.almStatusLab2.config(text="Move C must start with a Center followed by Start & Plane", style="Alarm.TLabel")
            else:
                xIndex = command.find(" X ")
                yIndex = command.find(" Y ")
                zIndex = command.find(" Z ")
                rzIndex = command.find(" Rz ")
                ryIndex = command.find(" Ry ")
                rxIndex = command.find(" Rx ")
                trIndex = command.find(" Tr ")
                SpeedIndex = command.find(" S")
                ACCspdIndex = command.find(" Ac ")
                DECspdIndex = command.find(" Dc ")
                ACCrampIndex = command.find(" Rm ")
                WristConfIndex = command.find(" $")
                xVal = command[xIndex + 3:yIndex]
                yVal = command[yIndex + 3:zIndex]
                zVal = command[zIndex + 3:rzIndex]
                rzVal = command[rzIndex + 4:ryIndex]
                ryVal = command[ryIndex + 4:rxIndex]
                rxVal = command[rxIndex + 4:trIndex]
                trVal = command[trIndex + 4:SpeedIndex]
                self.speedPrefix = command[SpeedIndex + 1:SpeedIndex + 3]
                self.Speed = command[SpeedIndex + 4:ACCspdIndex]
                self.ACCspd = command[ACCspdIndex + 4:DECspdIndex]
                self.DECspd = command[DECspdIndex + 4:ACCrampIndex]
                self.ACCramp = command[ACCrampIndex + 4:WristConfIndex]
                self.WC = command[WristConfIndex + 3:]
                TCX = 0
                TCY = 0
                TCZ = 0
                TCRx = 0
                TCRy = 0
                TCRz = 0
                ##read next row for Mid position
                curRow = self.tab10.progView.curselection()[0]
                selRow = self.tab10.progView.curselection()[0]
                last = self.tab10.progView.index('end')
                for row in range(0, selRow):
                    self.tab10.progView.itemconfig(row, {'fg': 'dodger blue'})
                self.tab10.progView.itemconfig(selRow, {'fg': 'blue2'})
                for row in range(selRow + 1, last):
                    self.tab10.progView.itemconfig(row, {'fg': 'black'})
                self.tab10.progView.selection_clear(0, END)
                selRow += 1
                self.tab10.progView.select_set(selRow)
                curRow += 1
                selRow = self.tab1.progView.curselection()[0]
                self.tab10.progView.see(selRow + 2)
                data = list(map(int, self.tab10.progView.curselection()))
                command = self.tab10.progView.get(data[0]).decode()
                xIndex = command.find(" X ")
                yIndex = command.find(" Y ")
                zIndex = command.find(" Z ")
                Xmid = command[xIndex + 3:yIndex]
                Ymid = command[yIndex + 3:zIndex]
                Zmid = command[zIndex + 3:rzIndex]
                ##read next row for End position
                curRow = self.tab10.progView.curselection()[0]
                selRow = self.tab10.progView.curselection()[0]
                last = self.tab10.progView.index('end')
                for row in range(0, selRow):
                    self.tab10.progView.itemconfig(row, {'fg': 'dodger blue'})
                self.tab10.progView.itemconfig(selRow, {'fg': 'blue2'})
                for row in range(selRow + 1, last):
                    self.tab10.progView.itemconfig(row, {'fg': 'black'})
                self.tab10.progView.selection_clear(0, END)
                selRow += 1
                self.tab10.progView.select_set(selRow)
                curRow += 1
                selRow = self.tab10.progView.curselection()[0]
                self.tab10.progView.see(selRow + 2)
                data = list(map(int, self.tab10.progView.curselection()))
                command = self.tab10.progView.get(data[0]).decode()
                xIndex = command.find(" X ")
                yIndex = command.find(" Y ")
                zIndex = command.find(" Z ")
                Xend = command[xIndex + 3:yIndex]
                Yend = command[yIndex + 3:zIndex]
                Zend = command[zIndex + 3:rzIndex]
                # move j to the beginning (second or mid point is start of circle)
                self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
                command = "MJ" + "X" + Xmid + "Y" + Ymid + "Z" + Zmid + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "Tr" + trVal + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
                self.ser.write(command.encode())
                self.ser.flushInput()
                time.sleep(.1)
                response = str(self.ser.readline().strip(), 'utf-8')
                # move circle command
                start = time.time()
                self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
                command = "MC" + "Cx" + xVal + "Cy" + yVal + "Cz" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "Bx" + Xmid + "By" + Ymid + "Bz" + Zmid + "Px" + Xend + "Py" + Yend + "Pz" + Zend + "Tr" + trVal + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
                self.cmdSentEntryField.delete(0, 'end')
                self.cmdSentEntryField.insert(0, command)
                self.ser.write(command.encode())
                self.ser.flushInput()
                time.sleep(.1)
                response = str(self.ser.readline().strip(), 'utf-8')
                end = time.time()
                # manEntryField.delete(0, 'end')
                # manEntryField.insert(0,end-start)
                if (response[:1] == 'E'):
                    self.ErrorHandler(response)
                else:
                    self.displayPosition(response)

                    ##Start Spline
        if (cmdType == "Start "):
            splineActive = "1"
            if (self.moveInProc == 1):
                self.moveInProc == 2
            command = "SL\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            self.ser.read()

            ##End Spline
        if (cmdType == "End Sp"):
            splineActive = "0"
            if (self.stopQueue == "1"):
                self.stopQueue = "0"
                stop()
            if (self.moveInProc == 1):
                self.moveInProc == 2
            command = "SS\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

                ##Camera On
        if (cmdType == "Cam On"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            self.start_vid()

        ##Camera Off
        if (cmdType == "Cam Of"):
            if (self.moveInProc == 1):
                self.moveInProc == 2
            self.stop_vid()

            ##Vision Find
        if (cmdType == "Vis Fi"):
            # if (self.moveInProc == 1):
            # self.moveInProc == 2
            templateIndex = command.find("Vis Find - ")
            bgColorIndex = command.find(" - BGcolor ")
            scoreIndex = command.find(" Score ")
            passIndex = command.find(" Pass ")
            failIndex = command.find(" Fail ")
            template = command[templateIndex + 11:bgColorIndex]
            checkBG = command[bgColorIndex + 11:scoreIndex]
            if (checkBG == "(Auto)"):
                background = "Auto"
            else:
                background = eval(command[bgColorIndex + 11:scoreIndex])
            min_score = float(command[scoreIndex + 7:passIndex]) * .01
            self.take_pic()
            status = self.visFind(template, min_score, background)
            if (status == "pass"):
                tabNum = ("Tab Number " + str(command[passIndex + 6:failIndex]) + "\r\n").encode('utf-8')
                index = self.tab10.progView.get(0, "end").index(tabNum)
                self.tab10.progView.selection_clear(0, END)
                self.tab10.progView.select_set(index)
            elif (status == "fail"):
                tabNum = ("Tab Number " + str(command[failIndex + 6:]) + "\r\n").encode('utf-8')
                index = self.tab10.progView.get(0, "end").index(tabNum)
                self.tab10.progView.selection_clear(0, END)
                self.tab10.progView.select_set(index)

        rowinproc = 0

    ##############################################################################################################################################################
    ### BUTTON JOGGING DEFS ############################################################################################################## BUTTON JOGGING DEFS ###
    ##############################################################################################################################################################
    def xbox(self):
        def threadxbox():
            from inputs import get_gamepad

            jogMode = 1
            if self.xboxUse == 0:
                self.xboxUse = 1
                mainMode = 1
                jogMode = 1
                grip = 0
                self.almStatusLab.config(text='JOGGING JOINTS 1 & 2', style="Warn.TLabel")
                self.almStatusLab2.config(text='JOGGING JOINTS 1 & 2', style="Warn.TLabel")
                self.xbcStatusLab.config(text='Xbox ON', )
                self.ChgDis(2)
            else:
                self.xboxUse = 0
                self.almStatusLab.config(text='XBOX CONTROLLER OFF', style="Warn.TLabel")
                self.almStatusLab2.config(text='XBOX CONTROLLER OFF', style="Warn.TLabel")
                self.xbcStatusLab.config(text='Xbox OFF', )
            while self.xboxUse == 1:
                try:
                    # if (TRUE):
                    events = get_gamepad()
                    for event in events:
                        ##DISTANCE
                        if (event.code == 'ABS_RZ' and event.state >= 100):
                            self.ChgDis(0)
                        elif (event.code == 'ABS_Z' and event.state >= 100):
                            self.ChgDis(1)
                        ##SPEED
                        elif (event.code == 'BTN_TR' and event.state == 1):
                            self.ChgSpd(0)
                        elif (event.code == 'BTN_TL' and event.state == 1):
                            self.ChgSpd(1)
                        ##JOINT MODE
                        elif (event.code == 'BTN_WEST' and event.state == 1):
                            if mainMode != 1:
                                mainMode = 1
                                jogMode = 1
                                self.almStatusLab.config(text='JOGGING JOINTS 1 & 2', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING JOINTS 1 & 2', style="Warn.TLabel")
                            else:
                                jogMode += 1
                            if jogMode == 2:
                                self.almStatusLab.config(text='JOGGING JOINTS 3 & 4', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING JOINTS 3 & 4', style="Warn.TLabel")
                            elif jogMode == 3:
                                self.almStatusLab.config(text='JOGGING JOINTS 5 & 6', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING JOINTS 5 & 6', style="Warn.TLabel")
                            elif jogMode == 4:
                                jogMode = 1
                                self.almStatusLab.config(text='JOGGING JOINTS 1 & 2', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING JOINTS 1 & 2', style="Warn.TLabel")
                        ##JOINT JOG
                        elif (mainMode == 1 and event.code == 'ABS_HAT0X' and event.state == 1 and jogMode == 1):
                            self.J1jogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0X' and event.state == -1 and jogMode == 1):
                            self.J1jogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0Y' and event.state == -1 and jogMode == 1):
                            self.J2jogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0Y' and event.state == 1 and jogMode == 1):
                            self.J2jogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0Y' and event.state == -1 and jogMode == 2):
                            self.J3jogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0Y' and event.state == 1 and jogMode == 2):
                            self.J3jogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0X' and event.state == 1 and jogMode == 2):
                            self.J4jogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0X' and event.state == -1 and jogMode == 2):
                            self.J4jogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0Y' and event.state == -1 and jogMode == 3):
                            self.J5jogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0Y' and event.state == 1 and jogMode == 3):
                            self.J5jogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0X' and event.state == 1 and jogMode == 3):
                            self.J6jogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 1 and event.code == 'ABS_HAT0X' and event.state == -1 and jogMode == 3):
                            self.J6jogPos(float(self.incrementEntryField.get()))
                            ##CARTESIAN DIR MODE
                        elif (event.code == 'BTN_SOUTH' and event.state == 1):
                            if mainMode != 2:
                                mainMode = 2
                                jogMode = 1
                                self.almStatusLab.config(text='JOGGING X & Y AXIS', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING X & Y AXIS', style="Warn.TLabel")
                            else:
                                jogMode += 1
                            if jogMode == 2:
                                self.almStatusLab.config(text='JOGGING Z AXIS', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING Z AXIS', style="Warn.TLabel")
                            elif jogMode == 3:
                                jogMode = 1
                                self.almStatusLab.config(text='JOGGING X & Y AXIS', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING X & Y AXIS', style="Warn.TLabel")
                        ##CARTESIAN DIR JOG
                        elif (mainMode == 2 and event.code == 'ABS_HAT0Y' and event.state == -1 and jogMode == 1):
                            self.XjogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 2 and event.code == 'ABS_HAT0Y' and event.state == 1 and jogMode == 1):
                            self.XjogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 2 and event.code == 'ABS_HAT0X' and event.state == 1 and jogMode == 1):
                            self.YjogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 2 and event.code == 'ABS_HAT0X' and event.state == -1 and jogMode == 1):
                            self.YjogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 2 and event.code == 'ABS_HAT0Y' and event.state == 1 and jogMode == 2):
                            self.ZjogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 2 and event.code == 'ABS_HAT0Y' and event.state == -1 and jogMode == 2):
                            self.ZjogPos(float(self.incrementEntryField.get()))
                            ##CARTESIAN ORIENTATION MODE
                        elif (event.code == 'BTN_EAST' and event.state == 1):
                            if mainMode != 3:
                                mainMode = 3
                                jogMode = 1
                                self.almStatusLab.config(text='JOGGING Rx & Ry AXIS', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING Rx & Ry AXIS', style="Warn.TLabel")
                            else:
                                jogMode += 1
                            if jogMode == 2:
                                self.almStatusLab.config(text='JOGGING Rz AXIS', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING Rz AXIS', style="Warn.TLabel")
                            elif jogMode == 3:
                                jogMode = 1
                                self.almStatusLab.config(text='JOGGING Rx & Ry AXIS', style="Warn.TLabel")
                                self.almStatusLab2.config(text='JOGGING Rx & Ry AXIS', style="Warn.TLabel")
                        ##CARTESIAN ORIENTATION JOG
                        elif (mainMode == 3 and event.code == 'ABS_HAT0X' and event.state == -1 and jogMode == 1):
                            self.RxjogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 3 and event.code == 'ABS_HAT0X' and event.state == 1 and jogMode == 1):
                            self.RxjogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 3 and event.code == 'ABS_HAT0Y' and event.state == 1 and jogMode == 1):
                            self.RyjogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 3 and event.code == 'ABS_HAT0Y' and event.state == -1 and jogMode == 1):
                            self.RyjogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 3 and event.code == 'ABS_HAT0X' and event.state == 1 and jogMode == 2):
                            self.RzjogNeg(float(self.incrementEntryField.get()))
                        elif (mainMode == 3 and event.code == 'ABS_HAT0X' and event.state == -1 and jogMode == 2):
                            self.RzjogPos(float(self.incrementEntryField.get()))
                        ##J7 MODE
                        elif (event.code == 'BTN_START' and event.state == 1):
                            mainMode = 4
                            self.almStatusLab.config(text='JOGGING TRACK', style="Warn.TLabel")
                            self.almStatusLab2.config(text='JOGGING TRACK', style="Warn.TLabel")
                        ##TRACK JOG
                        elif (mainMode == 4 and event.code == 'ABS_HAT0X' and event.state == 1):
                            self.J7jogPos(float(self.incrementEntryField.get()))
                        elif (mainMode == 4 and event.code == 'ABS_HAT0X' and event.state == -1):
                            self.J7jogNeg(float(self.incrementEntryField.get()))
                            ##TEACH POS
                        elif (event.code == 'BTN_NORTH' and event.state == 1):
                            self.teachInsertBelSelected()
                        ##GRIPPER
                        elif (event.code == 'BTN_SELECT' and event.state == 1):
                            if grip == 0:
                                grip = 1
                                outputNum = self.DO1offEntryField.get()
                                command = "OFX" + outputNum + "\n"
                                self.ser_IO.write(command.encode())
                                self.ser_IO.flushInput()
                                time.sleep(.1)
                                self.ser_IO.read()
                            else:
                                grip = 0
                                outputNum = self.DO1onEntryField.get()
                                command = "ONX" + outputNum + "\n"
                                self.ser_IO.write(command.encode())
                                self.ser_IO.flushInput()
                                time.sleep(.1)
                                self.ser_IO.read()
                                time.sleep(.1)
                        else:
                            pass
                except:
                    # else:
                    self.almStatusLab.config(text='XBOX CONTROLLER NOT RESPONDING', style="Alarm.TLabel")
                    self.almStatusLab2.config(text='XBOX CONTROLLER NOT RESPONDING', style="Alarm.TLabel")

        t = threading.Thread(target=threadxbox)
        t.start()

    def ChgDis(self, val):
        curSpd = int(self.incrementEntryField.get())
        if curSpd >= 100 and val == 0:
            curSpd = 100
        elif curSpd < 5 and val == 0:
            curSpd += 1
        elif val == 0:
            curSpd += 5
        if curSpd <= 1 and val == 1:
            curSpd = 1
        elif curSpd <= 5 and val == 1:
            curSpd -= 1
        elif val == 1:
            curSpd -= 5
        elif val == 2:
            curSpd = 5
        self.incrementEntryField.delete(0, 'end')
        self.incrementEntryField.insert(0, str(curSpd))

        time.sleep(.3)

    def ChgSpd(self, val):
        curSpd = int(self.speedEntryField.get())
        if curSpd >= 100 and val == 0:
            curSpd = 100
        elif curSpd < 5 and val == 0:
            curSpd += 1
        elif val == 0:
            curSpd += 5
        if curSpd <= 1 and val == 1:
            curSpd = 1
        elif curSpd <= 5 and val == 1:
            curSpd -= 1
        elif val == 1:
            curSpd -= 5
        elif val == 2:
            curSpd = 5
        self.speedEntryField.delete(0, 'end')
        self.speedEntryField.insert(0, str(curSpd))

    def LiveCarJog(self, value):

        self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
        self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        self.checkSpeedVals()
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "LC" + "V" + str(value) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def LiveToolJog(self, value):

        self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
        self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        self.checkSpeedVals()
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "LT" + "V" + str(value) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def teachInsertBelSelected(self):

        self.checkSpeedVals()
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        self.Speed = self.speedEntryField.get()
        speedtype = self.speedOption.get()
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        Rounding = self.roundEntryField.get()
        movetype = self.options.get()
        if (movetype == "OFF J"):
            movetype = movetype + " [ PR: " + str(self.SavePosEntryField.get()) + " ]"
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        if (movetype == "Move Vis"):
            movetype = movetype + " [ PR: " + str(self.SavePosEntryField.get()) + " ]"
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move PR"):
            movetype = movetype + " [ PR: " + str(self.SavePosEntryField.get()) + " ]"
            newPos = movetype + " [*]" + " J7 " + str(self.J7PosCur) + " J8 " + str(self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "OFF PR "):
            movetype = movetype + " [ PR: " + str(self.SavePosEntryField.get()) + " ] offs [ *PR: " + str(int(self.SavePosEntryField.get()) + 1) + " ] "
            newPos = movetype + " [*]" + " J7 " + str(self.J7PosCur) + " J8 " + str(self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move J"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move L"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " Rnd " + Rounding + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move R"):
            newPos = movetype + " [*] J1 " + self.J1AngCur + " J2 " + self.J2AngCur + " J3 " + self.J3AngCur + " J4 " + self.J4AngCur + " J5 " + self.J5AngCur + " J6 " + self.J6AngCur + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move A Mid"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab1.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move A End"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move C Center"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos + " Rz " + self.RzcurPos + " Ry " + self.RycurPos + " Rx " + self.RxcurPos + " J7 " + str(self.J7PosCur) + " J8 " + str(
                self.J8PosCur) + " J9 " + str(
                self.J9PosCur) + " " + self.speedPrefix + " " + self.Speed + " Ac " + self.ACCspd + " Dc " + self.DECspd + " Rm " + self.ACCramp + " $ " + self.WC
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move C Start"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Move C Plane"):
            newPos = movetype + " [*] X " + self.XcurPos + " Y " + self.YcurPos + " Z " + self.ZcurPos
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Start Spline" or movetype == "End Spline"):
            newPos = movetype
            self.tab10.progView.insert(selRow, bytes(newPos + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()
        elif (movetype == "Teach PR"):
            PR = str(self.SavePosEntryField.get())
            SPE6 = "Position Register " + PR + " Element 6 = " + self.RxcurPos
            self.tab10.progView.insert(selRow, bytes(SPE6 + '\n', 'utf-8'))
            SPE5 = "Position Register " + PR + " Element 5 = " + self.RycurPos
            self.tab10.progView.insert(selRow, bytes(SPE5 + '\n', 'utf-8'))
            SPE4 = "Position Register " + PR + " Element 4 = " + self.RzcurPos
            self.tab10.progView.insert(selRow, bytes(SPE4 + '\n', 'utf-8'))
            SPE3 = "Position Register " + PR + " Element 3 = " + self.ZcurPos
            self.tab10.progView.insert(selRow, bytes(SPE3 + '\n', 'utf-8'))
            SPE2 = "Position Register " + PR + " Element 2 = " + self.YcurPos
            self.tab10.progView.insert(selRow, bytes(SPE2 + '\n', 'utf-8'))
            SPE1 = "Position Register " + PR + " Element 1 = " + self.XcurPos
            self.tab10.progView.insert(selRow, bytes(SPE1 + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab1.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()

    def teachReplaceSelected(self):
        try:
            self.deleteitem()
            selRow = self.tab10.progView.curselection()[0]
            self.tab10.progView.select_set(selRow - 1)
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        self.teachInsertBelSelected()

    def deleteitem(self):
        selRow = self.tab10.progView.curselection()[0]
        selection = self.tab10.progView.curselection()
        self.tab10.progView.delete(selection[0])
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def waitTime(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        seconds = self.waitTimeEntryField.get()
        newTime = "Wait Time = " + seconds
        self.tab10.progView.insert(selRow, bytes(newTime + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def waitInputOn(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        input = self.waitInputEntryField.get()
        newInput = "Wait Input On = " + input
        self.tab10.progView.insert(selRow, bytes(newInput + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def waitInputOff(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        input = self.waitInputOffEntryField.get()
        newInput = "Wait Off Input = " + input
        self.tab10.progView.insert(selRow, bytes(newInput + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def setOutputOn(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        output = self.outputOnEntryField.get()
        newOutput = "Out On = " + output
        self.tab10.progView.insert(selRow, bytes(newOutput + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def setOutputOff(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        output = self.outputOffEntryField.get()
        newOutput = "Out Off = " + output
        self.tab10.progView.insert(selRow, bytes(newOutput + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def tabNumber(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        tabNum = self.tabNumEntryField.get()
        tabins = "Tab Number " + tabNum
        self.tab10.progView.insert(selRow, bytes(tabins + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def jumpTab(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        tabNum = self.jumpTabEntryField.get()
        tabjmp = "Jump Tab-" + tabNum
        self.tab10.progView.insert(selRow, bytes(tabjmp + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def cameraOn(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        value = "Cam On"
        self.tab10.progView.insert(selRow, bytes(value + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def cameraOff(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        value = "Cam Off"
        self.tab10.progView.insert(selRow, bytes(value + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def IfCMDInsert(self):
        localErrorFlag = False
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)

        option = self.iFoption.get()
        selection = self.iFselection.get()
        variable = self.IfVarEntryField.get()
        if (variable == ""):
            localErrorFlag = True
            message = "Please enter an input, register number or COM Port"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
        inputVal = self.IfInputEntryField.get()
        destVal = self.IfDestEntryField.get()
        if (option == "Input"):
            if (inputVal == "1"):
                prefix = "If Input # " + variable + " = " + inputVal + " :"
            elif (inputVal == "0"):
                prefix = "If Input # " + variable + " = " + inputVal + " :"
            else:
                localErrorFlag = True
                message = "Please enter a 1 or 0 for the = value"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")

        elif (option == "Register"):
            if (inputVal == ""):
                localErrorFlag = True
                message = "Please enter a register number"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
            prefix = "If Register # " + variable + " = " + inputVal + " :"

        elif (option == "COM Device"):
            if (inputVal == ""):
                localErrorFlag = True
                message = "Please enter expected COM device input"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
            prefix = "If COM Device # " + variable + " = " + inputVal + " :"

        if (selection == "Call self.Progr"):
            if (destVal == ""):
                localErrorFlag = True
                message = "Please enter a program name"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
            value = prefix + " Call self.Progr " + destVal
        elif (selection == "Jump Tab"):
            if (destVal == ""):
                localErrorFlag = True
                message = "Please enter a destination tab"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
            value = prefix + " Jump to Tab " + destVal

        if (localErrorFlag == False):
            self.tab10.progView.insert(selRow, bytes(value + '\n', 'utf-8'))
            self.tab10.progView.selection_clear(0, END)
            self.tab10.progView.select_set(selRow)
            items = self.tab10.progView.get(0, END)
            file_path = path.relpath(self.ProgEntryField.get())
            with open(file_path, 'w', encoding='utf-8') as f:
                for item in items:
                    f.write(str(item.strip(), encoding='utf-8'))
                    f.write('\n')
                f.close()

    def ReadAuxCom(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        comNum = self.auxPortEntryField.get()
        comChar = self.auxCharEntryField.get()
        servoins = "Read COM # " + comNum + " Char: " + comChar
        self.tab10.progView.insert(selRow, bytes(servoins + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def TestAuxCom(self):
        try:
            global ser3
            port = "COM" + self.com3PortEntryField.get()
            baud = 115200
            self.ser3 = serial.Serial(port, baud, timeout=5)
        except:
            Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
            self.tab5_rb.ElogView.insert(END, Curtime + " - UNABLE TO ESTABLISH COMMUNICATIONS WITH SERIAL DEVICE")
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
        self.ser3.flushInput()
        numChar = int(self.com3charPortEntryField.get())
        response = str(self.ser3.read(numChar).strip(), 'utf-8')
        self.com3outPortEntryField.delete(0, 'end')
        self.com3outPortEntryField.insert(0, response)

    def Servo(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        servoNum = self.servoNumEntryField.get()
        servoPos = self.servoPosEntryField.get()
        servoins = "Servo number " + servoNum + " to position: " + servoPos
        self.tab10.progView.insert(selRow, bytes(servoins + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def loadProg(self):
        if getattr(sys, 'frozen', False):
            folder = os.path.dirname(sys.executable)
        elif __file__:
            folder = os.path.dirname(os.path.realpath(__file__))
        # folder = os.path.dirname(os.path.realpath(__file__))
        filetypes = (('robot program', '*.ar4'), ("all files", "*.*"))
        filename = fd.askopenfilename(title='Open files', initialdir=folder, filetypes=filetypes)
        name = os.path.basename(filename)
        self.ProgEntryField.delete(0, 'end')
        self.ProgEntryField.insert(0, name)
        self.tab10.progView.delete(0, END)
        self.Progr = open(filename, "rb")
        time.sleep(.1)
        for item in self.Progr:
            self.tab10.progView.insert(END, item)
        self.tab10.progView.pack()
        self.scrollbar.config(command=self.tab10.progView.yview)
        self.savePosData()

    def callProg(self, name):
        self.ProgEntryField.delete(0, 'end')
        self.ProgEntryField.insert(0, name)
        self.tab10.progView.delete(0, END)
        self.Progr = open(name, "rb")
        time.sleep(.1)
        for item in self.Progr:
            self.tab10.progView.insert(END, item)
        self.tab10.progView.pack()
        self.scrollbar.config(command=self.tab10.progView.yview)

    def CreateProg(self):
        user_input = simpledialog.askstring(title="New Program", prompt="New Program Name:")
        file_path = user_input + ".ar4"
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("##BEGINNING OF PROGRAM##")
            f.write('\n')
        f.close()
        self.ProgEntryField.delete(0, 'end')
        self.ProgEntryField.insert(0, file_path)
        self.tab10.progView.delete(0, END)
        self.Progr = open(file_path, "rb")
        time.sleep(.1)
        for item in self.Progr:
            self.tab10.progView.insert(END, item)
        self.tab10.progView.pack()
        self.scrollbar.config(command=self.tab10.progView.yview)
        self.savePosData()

    def insertCallProg(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        newProg = self.changeProgEntryField.get()
        changeProg = "Call Program - " + newProg
        if str(changeProg[-4:]) != ".ar4":
            changeProg = changeProg + ".ar4"
        self.tab10.progView.insert(selRow, bytes(changeProg + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def insertGCprog(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        newProg = self.PlayGCEntryField.get()
        GCProg = "Run Gcode Program - " + newProg
        self.tab10.progView.insert(selRow, bytes(GCProg + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def insertReturn(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        value = "Return"
        self.tab10.progView.insert(selRow, bytes(value + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    ################################################################################################################################# end
    def posRegFieldVisible(self,a):
        # print("covo")
        curCmdtype = self.options.get()
        if (curCmdtype == "Move PR" or curCmdtype == "OFF PR " or curCmdtype == "Teach PR"):
            self.SavePosEntryField.place(x=780, y=183)
        else:
            self.SavePosEntryField.place_forget()

    def displayPosition(self, response):

        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        J1AngIndex = response.find('A');
        J2AngIndex = response.find('B');
        J3AngIndex = response.find('C');
        J4AngIndex = response.find('D');
        J5AngIndex = response.find('E');
        J6AngIndex = response.find('F');
        XposIndex = response.find('G');
        YposIndex = response.find('H');
        ZposIndex = response.find('I');
        RzposIndex = response.find('J');
        RyposIndex = response.find('K');
        RxposIndex = response.find('L');
        SpeedVioIndex = response.find('M');
        DebugIndex = response.find('N');
        FlagIndex = response.find('O');
        J7PosIndex = response.find('P');
        J8PosIndex = response.find('Q');
        J9PosIndex = response.find('R');
        self.J1AngCur = response[J1AngIndex + 1:J2AngIndex].strip();
        self.J2AngCur = response[J2AngIndex + 1:J3AngIndex].strip();
        self.J3AngCur = response[J3AngIndex + 1:J4AngIndex].strip();
        self.J4AngCur = response[J4AngIndex + 1:J5AngIndex].strip();
        self.J5AngCur = response[J5AngIndex + 1:J6AngIndex].strip();
        self.J6AngCur = response[J6AngIndex + 1:XposIndex].strip();

        if (float(self.J5AngCur) > 0):
            self.WC = "F"
        else:
            self.WC = "N"
        self.XcurPos = response[XposIndex + 1:YposIndex].strip();
        self.YcurPos = response[YposIndex + 1:ZposIndex].strip();
        self.ZcurPos = response[ZposIndex + 1:RzposIndex].strip();
        self.RzcurPos = response[RzposIndex + 1:RyposIndex].strip();
        self.RycurPos = response[RyposIndex + 1:RxposIndex].strip();
        self.RxcurPos = response[RxposIndex + 1:SpeedVioIndex].strip();
        SpeedVioation = response[SpeedVioIndex + 1:DebugIndex].strip();
        Debug = response[DebugIndex + 1:FlagIndex].strip();
        Flag = response[FlagIndex + 1:J7PosIndex].strip();
        self.J7PosCur = float(response[J7PosIndex + 1:J8PosIndex].strip());
        self.J8PosCur = float(response[J8PosIndex + 1:J9PosIndex].strip());
        self.J9PosCur = float(response[J9PosIndex + 1:].strip());

        self.J1curAngEntryField.delete(0, 'end')
        self.J1curAngEntryField.insert(0, self.J1AngCur)
        self.J2curAngEntryField.delete(0, 'end')
        self.J2curAngEntryField.insert(0, self.J2AngCur)
        self.J3curAngEntryField.delete(0, 'end')
        self.J3curAngEntryField.insert(0, self.J3AngCur)
        self.J4curAngEntryField.delete(0, 'end')
        self.J4curAngEntryField.insert(0, self.J4AngCur)
        self.J5curAngEntryField.delete(0, 'end')
        self.J5curAngEntryField.insert(0, self.J5AngCur)
        self.J6curAngEntryField.delete(0, 'end')
        self.J6curAngEntryField.insert(0, self.J6AngCur)
        self.XcurEntryField.delete(0, 'end')
        self.XcurEntryField.insert(0, self.XcurPos)
        self.YcurEntryField.delete(0, 'end')
        self.YcurEntryField.insert(0, self.YcurPos)
        self.ZcurEntryField.delete(0, 'end')
        self.ZcurEntryField.insert(0, self.ZcurPos)
        self.RzcurEntryField.delete(0, 'end')
        self.RzcurEntryField.insert(0, self.RzcurPos)
        self.RycurEntryField.delete(0, 'end')
        self.RycurEntryField.insert(0, self.RycurPos)
        self.RxcurEntryField.delete(0, 'end')
        self.RxcurEntryField.insert(0, self.RxcurPos)
        self.J7curAngEntryField.delete(0, 'end')
        self.J7curAngEntryField.insert(0, self.J7PosCur)
        self.J8curAngEntryField.delete(0, 'end')
        self.J8curAngEntryField.insert(0, self.J8PosCur)
        self.J9curAngEntryField.delete(0, 'end')
        self.J9curAngEntryField.insert(0, self.J9PosCur)
        self.J1jogslide.set(self.J1AngCur)
        self.J2jogslide.set(self.J2AngCur)
        self.J3jogslide.set(self.J3AngCur)
        self.J4jogslide.set(self.J4AngCur)
        self.J5jogslide.set(self.J5AngCur)
        self.J6jogslide.set(self.J6AngCur)
        self.J7jogslide.set(self.J7PosCur)
        self.J8jogslide.set(self.J8PosCur)
        self.J9jogslide.set(self.J9PosCur)
        self.manEntryField.delete(0, 'end')
        self.manEntryField.insert(0, Debug)

        # self.tab5_rb.ElogView.insert(END, Debug)
        # value=self.tab5_rb.ElogView.get(0,END)
        # pickle.dump(value,open("ErrorLog","wb"))

        self.savePosData()
        if (Flag != ""):
            self.ErrorHandler(Flag)
        if (SpeedVioation == '1'):
            Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
            message = "Max self.Speed Violation - Reduce self.Speed Setpoint or Travel Distance"
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))
            self.almStatusLab.config(text=message, style="Warn.TLabel")
            self.almStatusLab2.config(text=message, style="Warn.TLabel")

    def SelXjogNeg(self, a):
        # print(a)
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.XjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(10)

    def SelXjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.XjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(11)

    def SelYjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.YjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(20)

    def SelYjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.YjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(21)

    def SelZjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.ZjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(30)

    def SelZjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.ZjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(31)

    def SelRzjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.RzjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(40)

    def SelRzjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.RzjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(41)

    def SelRyjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.RyjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(50)

    def SelRyjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.RyjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(51)

    def SelRxjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.RxjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(60)

    def SelRxjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.RxjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveCarJog(61)

    def SelTxjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TXjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(10)

    def SelTxjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TXjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(11)

    def SelTyjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TYjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(20)

    def SelTyjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TYjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(21)

    def SelTzjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TZjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(30)

    def SelTzjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TZjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(31)

    def SelTRzjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TRzjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(40)

    def SelTRzjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TRzjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(41)

    def SelTRyjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TRyjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(50)

    def SelTRyjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TRyjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(51)

    def SelTRxjogNeg(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TRxjogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(60)

    def SelTRxjogPos(self, a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.TRxjogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveToolJog(61)

    def reloadProg(self):
        file_path = path.relpath(self.ProgEntryField.get())
        self.ProgEntryField.delete(0, 'end')
        self.ProgEntryField.insert(0, file_path)
        self.tab10.progView.delete(0, END)
        self.Progr = open(file_path, "rb")
        time.sleep(.1)
        for item in self.Progr:
            self.tab10.progView.insert(END, item)
        self.tab10.progView.pack()
        self.scrollbar.config(command=self.tab10.progView.yview)
        self.savePosData()

    def openText(self):
        file_path = path.relpath(self.ProgEntryField.get())
        os.startfile(file_path)

    def manReplItem(self):
        # selRow = self.curRowEntryField.get()
        selRow = self.tab10.progView.curselection()[0]
        self.tab10.progView.delete(selRow)
        self.tab10.progView.insert(selRow, bytes(self.manEntryField.get() + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        self.tab10.progView.itemconfig(selRow, {'fg': 'darkgreen'})
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def checkSpeedVals(self):
        speedtype = self.speedOption.get()
        self.Speed = float(self.speedEntryField.get())
        if (speedtype == "mm per Sec"):
            if (self.Speed <= .01):
                self.speedEntryField.delete(0, 'end')
                self.speedEntryField.insert(0, "5")
        if (speedtype == "Seconds"):
            if (self.Speed <= .001):
                self.speedEntryField.delete(0, 'end')
                self.speedEntryField.insert(0, "1")
        if (speedtype == "Percent"):
            if (self.Speed <= .01 or self.Speed > 100):
                self.speedEntryField.delete(0, 'end')
                self.speedEntryField.insert(0, "10")
        self.ACCspd = float(self.ACCspeedField.get())
        if (self.ACCspd <= .01 or self.ACCspd > 100):
            self.ACCspeedField.delete(0, 'end')
            self.ACCspeedField.insert(0, "10")
        self.DECspd = float(self.DECspeedField.get())
        if (self.DECspd <= .01 or self.DECspd >= 100):
            self.DECspeedField.delete(0, 'end')
            self.DECspeedField.insert(0, "10")
        self.ACCramp = float(self.ACCrampField.get())
        if (self.ACCramp <= .01 or self.ACCramp > 100):
            self.ACCrampField.delete(0, 'end')
            self.ACCrampField.insert(0, "50")

    def manInsItem(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        self.tab10.progView.insert(selRow, bytes(self.manEntryField.get() + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        selRow = self.tab10.progView.curselection()[0]
        self.curRowEntryField.delete(0, 'end')
        self.curRowEntryField.insert(0, selRow)
        self.tab10.progView.itemconfig(selRow, {'fg': 'darkgreen'})
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def getSel(self):
        selRow = self.tab10.progView.curselection()[0]
        self.tab10.progView.see(selRow + 2)
        data = list(map(int, self.tab10.progView.curselection()))
        command = self.tab10.progView.get(data[0]).decode()
        self.manEntryField.delete(0, 'end')
        self.manEntryField.insert(0, command)

    def progViewselect(self, e):
        selRow = self.tab10.progView.curselection()[0]
        self.curRowEntryField.delete(0, 'end')
        self.curRowEntryField.insert(0, selRow)

    def SelJ1jogPos(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J1jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(11)

    def SelJ1jogNeg(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J1jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(10)

    def J1sliderUpdate(self, foo):
        self.J1slidelabel.config(text=round(float(self.J1jogslide.get()), 2))

    def J1sliderExecute(self, foo):
        J1delta = float(self.J1jogslide.get()) - float(self.J1curAngEntryField.get())
        # print(self.J1curAngEntryField.get())
        if (J1delta < 0):
            self.J1jogNeg(abs(J1delta))
        else:
            self.J1jogPos(abs(J1delta))

    #
    def SelJ2jogNeg(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J2jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(20)

    def SelJ2jogPos(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J2jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(21)

    def J2sliderUpdate(self, foo):
        self.J2slidelabel.config(text=round(float(self.J2jogslide.get()), 2))

    def J2sliderExecute(self, foo):
        J2delta = float(self.J2jogslide.get()) - float(self.J2curAngEntryField.get())
        if (J2delta < 0):
            self.J2jogNeg(abs(J2delta))
        else:
            self.J2jogPos(abs(J2delta))

    def SelJ3jogNeg(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J3jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(30)

    def SelJ3jogPos(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J3jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(31)

    def J3sliderUpdate(self, foo):
        self.J3slidelabel.config(text=round(float(self.J3jogslide.get()), 2))

    def J3sliderExecute(self, foo):
        J3delta = float(self.J3jogslide.get()) - float(self.J3curAngEntryField.get())
        if (J3delta < 0):
            self.J3jogNeg(abs(J3delta))
        else:
            self.J3jogPos(abs(J3delta))

    def SelJ4jogNeg(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J4jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(40)

    def SelJ4jogPos(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J4jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(41)

    def J4sliderUpdate(self, foo):
        self.J4slidelabel.config(text=round(float(self.J4jogslide.get()), 2))

    def J4sliderExecute(self, foo):
        J4delta = float(self.J4jogslide.get()) - float(self.J4curAngEntryField.get())
        if (J4delta < 0):
            self.J4jogNeg(abs(J4delta))
        else:
            self.J4jogPos(abs(J4delta))

    def SelJ5jogNeg(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J5jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(50)

    def SelJ5jogPos(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J5jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(51)

    def J5sliderUpdate(self, foo):
        self.J5slidelabel.config(text=round(float(self.J5jogslide.get()), 2))

    def J5sliderExecute(self, foo):
        J5delta = float(self.J5jogslide.get()) - float(self.J5curAngEntryField.get())
        if (J5delta < 0):
            self.J5jogNeg(abs(J5delta))
        else:
            self.J5jogPos(abs(J5delta))

    def SelJ6jogNeg(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J6jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(60)

    def SelJ6jogPos(self):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J6jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(61)

    def J6sliderUpdate(self, foo):
        self.J6slidelabel.config(text=round(float(self.J6jogslide.get()), 2))

    def J6sliderExecute(self, foo):
        J6delta = float(self.J6jogslide.get()) - float(self.J6curAngEntryField.get())
        if (J6delta < 0):
            self.J6jogNeg(abs(J6delta))
        else:
            self.J6jogPos(abs(J6delta))

    def SelJ7jogNeg(self,a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J7jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(70)

    def SelJ7jogPos(self,a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J7jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(71)

    def J7sliderUpdate(self, foo):
        self.J7slideLimLab.config(text=round(float(self.J7jogslide.get()), 2))

    def J7sliderExecute(self, foo):
        J7delta = float(self.J7jogslide.get()) - float(self.J7curAngEntryField.get())
        if (J7delta < 0):
            self.J7jogNeg(abs(J7delta))
        else:
            self.J7jogPos(abs(J7delta))

    def SelJ8jogNeg(self,a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J8jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(80)

    def SelJ8jogPos(self,a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J8jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(81)

    def J8sliderUpdate(self, foo):
        self.J8slideLimLab.config(text=round(float(self.J8jogslide.get()), 2))

    def J8sliderExecute(self, foo):
        J8delta = float(self.J8jogslide.get()) - float(self.J8curAngEntryField.get())
        if (J8delta < 0):
            self.J8jogNeg(abs(J8delta))
        else:
            self.J8jogPos(abs(J8delta))

    def SelJ9jogNeg(self,a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J9jogNeg(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(90)

    def SelJ9jogPos(self,a):
        IncJogStatVal = int(self.IncJogStat.get())
        if (IncJogStatVal == 1):
            self.J9jogPos(float(self.incrementEntryField.get()))
        else:
            self.LiveJointJog(91)

    def J9sliderUpdate(self, foo):
        self.J9slideLimLab.config(text=round(float(self.J9jogslide.get()), 2))

    def J9sliderExecute(self, foo):
        J9delta = float(self.J9jogslide.get()) - float(self.J9curAngEntryField.get())
        if (J9delta < 0):
            self.J9jogNeg(abs(J9delta))
        else:
            self.J9jogPos(abs(J9delta))

    def LiveJointJog(self, value):

        self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
        self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        self.checkSpeedVals()
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "LJ" + "V" + str(value) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def XjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = str(float(self.XcurPos) - value)
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def YjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = str(float(self.YcurPos) - value)
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def ZjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = str(float(self.ZcurPos) - value)
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def RxjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = str(float(self.RxcurPos) - value)
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def RyjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = str(float(self.RycurPos) - value)
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def RzjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = str(float(self.RzcurPos) - value)
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def XjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = str(float(self.XcurPos) + value)
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def YjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = str(float(self.YcurPos) + value)
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def ZjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = str(float(self.ZcurPos) + value)
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def RxjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = self.RycurPos
        rxVal = str(float(self.RxcurPos) + value)
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def RyjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = self.RzcurPos
        ryVal = str(float(self.RycurPos) + value)
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def RzjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # mm/sec
        if (speedtype == "mm per Sec"):
            self.speedPrefix = "Sm"
            # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        xVal = self.XcurPos
        yVal = self.YcurPos
        zVal = self.ZcurPos
        rzVal = str(float(self.RzcurPos) + value)
        ryVal = self.RycurPos
        rxVal = self.RxcurPos
        j7Val = str(self.J7PosCur)
        j8Val = str(self.J8PosCur)
        j9Val = str(self.J9PosCur)
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + j7Val + "J8" + j8Val + "J9" + j9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TXjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTX1" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TYjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTY1" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TZjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTZ1" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TRxjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTW1" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TRyjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTP1" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TRzjogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTR1" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TXjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTX0" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TYjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTY0" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TZjogPos(self, value):
        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTZ0" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TRxjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTW0" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TRyjogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTP0" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def TRzjogPos(self, value):
        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to sec
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Ss"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "JTR0" + str(value) + self.speedPrefix + self.Speed + "G" + self.ACCspd + "H" + self.DECspd + "I" + self.ACCramp + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def create_robot_widget(self):
        # self.wireframe = GraphicsOptions.set_draw_wireframe(self,True)
        style_rb = ttk.Style()
        style_rb.configure("Custom.TNotebook", tabposition="wn")
        self.tab_RB_control = ttk.Notebook(self.tab2, style="Custom.TNotebook", height=768, width=1080)

        self.tab0_rb = ttk.Frame(self.tab_RB_control)
        self.tab1_rb = ttk.Frame(self.tab_RB_control)
        self.tab2_rb = ttk.Frame(self.tab_RB_control)
        self.tab3_rb = ttk.Frame(self.tab_RB_control)
        self.tab4_rb = ttk.Frame(self.tab_RB_control)
        self.tab5_rb = ttk.Frame(self.tab_RB_control)

        self.tab_RB_control.add(self.tab0_rb, state='normal', text='CONFIG_RB')
        self.tab_RB_control.add(self.tab1_rb, state='normal', text='KINEMATIC')
        self.tab_RB_control.add(self.tab2_rb, state='normal', text='IN_OUT')
        self.tab_RB_control.add(self.tab3_rb, state='normal', text='REGITER')
        self.tab_RB_control.add(self.tab4_rb, state='normal', text='G-CODE')
        self.tab_RB_control.add(self.tab5_rb, state='normal', text='LOG')

        self.tab_RB_control.place(x=0, y=0)

        self.label_robotset = tk.LabelFrame(self.tab1_rb, text="KINEMATIC_forward", labelanchor="nw", width=220, height=280)
        self.label_robotset.propagate(False)
        self.label_robotset.place(x=750, y=10)
        self.label_robot_check = tk.LabelFrame(self.tab1_rb, text="KINEMATIC_forward", labelanchor="nw", width=220, height=280)
        self.label_robot_check.propagate(False)
        self.label_robot_check.place(x=530, y=10)

        self.bool_draw_wireframe = tk.BooleanVar()
        self.bool_draw_wireframe.set(False)  # 最初はチェックなし
        self.bool_draw_motion_vectors = tk.BooleanVar()
        self.bool_draw_motion_vectors.set(False)
        self.bool_use_alpha = tk.BooleanVar()
        self.bool_use_alpha.set(False)
        self.bool_draw_joint_frame = tk.BooleanVar()
        self.bool_draw_joint_frame.set(False)
        self.bool_draw_all_joint_frames = tk.BooleanVar()
        self.bool_draw_all_joint_frames.set(False)

        check_option_rb = tk.Checkbutton(self.label_robot_check, text="draw_wireframe", variable=self.bool_draw_wireframe)
        check_option_rb1 = tk.Checkbutton(self.label_robot_check, text="draw_motion_vectors", variable=self.bool_draw_motion_vectors)
        check_option_rb2 = tk.Checkbutton(self.label_robot_check, text="draw_use_alpha", variable=self.bool_use_alpha)
        check_option_rb3 = tk.Checkbutton(self.label_robot_check, text="draw_joint_frame", variable=self.bool_draw_joint_frame)
        check_option_rb4 = tk.Checkbutton(self.label_robot_check, text="draw_all_joint_frames", variable=self.bool_draw_all_joint_frames)

        check_option_rb.place(x=0, y=10)
        check_option_rb1.place(x=0, y=30)
        check_option_rb2.place(x=0, y=50)
        check_option_rb3.place(x=0, y=70)
        check_option_rb4.place(x=0, y=90)

        #
        scale_rb1 = tk.Scale(self.label_robotset, variable=self.val1, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)
        scale_rb2 = tk.Scale(self.label_robotset, variable=self.val2, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)
        scale_rb3 = tk.Scale(self.label_robotset, variable=self.val3, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)
        scale_rb4 = tk.Scale(self.label_robotset, variable=self.val4, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)
        scale_rb5 = tk.Scale(self.label_robotset, variable=self.val5, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)
        scale_rb6 = tk.Scale(self.label_robotset, variable=self.val6, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)
        scale_rb7 = tk.Scale(self.label_robotset, variable=self.val7, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=20, from_=-100, to=100, resolution=1, tickinterval=0)

        scale_rb1.place(x=0, y=10)
        scale_rb2.place(x=0, y=40)
        scale_rb3.place(x=0, y=70)
        scale_rb4.place(x=0, y=100)
        scale_rb5.place(x=0, y=130)
        scale_rb6.place(x=0, y=170)
        scale_rb7.place(x=0, y=200)
        #

        self.frm.bind("<Motion>", self.rb_mouse_move)  # MouseMove
        self.frm.bind("<B1-Motion>", self.rb_mouse_move_left)  # MouseMove（左ボタンを押しながら移動）
        self.frm.bind("<Button-1>", self.rb_mouse_down_left)  # MouseDown（左ボタン）
        self.frm.bind("<B2-Motion>", self.rb_mouse_move_mid)  #
        self.frm.bind("<Double-Button-1>", self.rb_mouse_double_click_left)  # MouseDoubleClick（左ボタン）
        self.frm.bind("<MouseWheel>", self.rb_mouse_wheel)  # MouseWheel
        self.frm.bind("<ButtonRelease-1>", self.rb_mouse_up_left)
        self.frm.bind("<Button-2>", self.rb_mouse_down_mid)  # MouseMove（中ボタンを押しながら移動）
        self.frm.bind("<B3-Motion>", self.rb_mouse_move_right)  # MouseMove（左ボタンを押しながら移動）
        self.frm.bind("<Button-3>", self.rb_mouse_down_right)  # MouseDown（左ボタン）
        self.frm.bind("<ButtonRelease-3>", self.rb_mouse_up_right)
        self.mouse_down_position_x = 0
        self.mouse_down_position_y = 0
        self.mouse_x_old = 0
        self.mouse_y_old = 0

    def robot_tab_0_widget(self):
        self.calib_auto_group = tk.LabelFrame(self.tab0_rb, text="RB calibration", labelanchor="nw", width=120, height=300)
        self.calib_auto_group.propagate(False)
        self.calib_auto_group.place(x=830, y=10)
        self.almStatusLab2 = Label(self.tab0_rb, text="SYSTEM READY - NO ACTIVE ALARMS", style="OK.TLabel")
        self.almStatusLab2.place(x=25, y=20)

        autoCalBut = Button(self.calib_auto_group, text="  Auto Calibrate  ", command=self.calRobotAll)
        #
        J1calCbut = Checkbutton(self.calib_auto_group, text="J1", variable=self.J1CalStat)
        J2calCbut = Checkbutton(self.calib_auto_group, text="J2", variable=self.J2CalStat)
        J3calCbut = Checkbutton(self.calib_auto_group, text="J3", variable=self.J3CalStat)
        J4calCbut = Checkbutton(self.calib_auto_group, text="J4", variable=self.J4CalStat)
        J5calCbut = Checkbutton(self.calib_auto_group, text="J5", variable=self.J5CalStat)
        J6calCbut = Checkbutton(self.calib_auto_group, text="J6", variable=self.J6CalStat)
        J7calCbut = Checkbutton(self.calib_auto_group, text="J7", variable=self.J7CalStat)
        J8calCbut = Checkbutton(self.calib_auto_group, text="J8", variable=self.J8CalStat)
        J9calCbut = Checkbutton(self.calib_auto_group, text="J9", variable=self.J9CalStat)
        #
        J1calCbut2 = Checkbutton(self.calib_auto_group, text="J1", variable=self.J1CalStat2)
        J2calCbut2 = Checkbutton(self.calib_auto_group, text="J2", variable=self.J2CalStat2)
        J3calCbut2 = Checkbutton(self.calib_auto_group, text="J3", variable=self.J3CalStat2)
        J4calCbut2 = Checkbutton(self.calib_auto_group, text="J4", variable=self.J4CalStat2)
        J5calCbut2 = Checkbutton(self.calib_auto_group, text="J5", variable=self.J5CalStat2)
        J6calCbut2 = Checkbutton(self.calib_auto_group, text="J6", variable=self.J6CalStat2)
        J7calCbut2 = Checkbutton(self.calib_auto_group, text="J7", variable=self.J7CalStat2)
        J8calCbut2 = Checkbutton(self.calib_auto_group, text="J8", variable=self.J8CalStat2)
        J9calCbut2 = Checkbutton(self.calib_auto_group, text="J9", variable=self.J9CalStat2)
        #
        autoCalBut.place(x=10, y=0)
        J1calCbut.place(x=0, y=35)
        J2calCbut.place(x=35, y=35)
        J3calCbut.place(x=70, y=35)
        J4calCbut.place(x=0, y=55)
        J5calCbut.place(x=35, y=55)
        J6calCbut.place(x=70, y=55)
        J7calCbut.place(x=0, y=75)
        J8calCbut.place(x=35, y=75)
        J9calCbut.place(x=70, y=75)
        J1calCbut2.place(x=0, y=95)
        J2calCbut2.place(x=35, y=95)
        J3calCbut2.place(x=70, y=95)
        J4calCbut2.place(x=0, y=115)
        J5calCbut2.place(x=35, y=115)
        J6calCbut2.place(x=70, y=115)
        J7calCbut2.place(x=0, y=135)
        J8calCbut2.place(x=35, y=135)
        J9calCbut2.place(x=70, y=135)

        self.calib_ofset_group = tk.LabelFrame(self.tab0_rb, text="Calibration Offsets", labelanchor="nw", width=120, height=300)
        self.calib_ofset_group.propagate(False)
        self.calib_ofset_group.place(x=705, y=10)
        J1calLab = Label(self.calib_ofset_group, text="J1 Offset")
        J2calLab = Label(self.calib_ofset_group, text="J2 Offset")
        J3calLab = Label(self.calib_ofset_group, text="J3 Offset")
        J4calLab = Label(self.calib_ofset_group, text="J4 Offset")
        J5calLab = Label(self.calib_ofset_group, text="J5 Offset")
        J6calLab = Label(self.calib_ofset_group, text="J6 Offset")
        J7calLab = Label(self.calib_ofset_group, text="J7 Offset")
        J8calLab = Label(self.calib_ofset_group, text="J8 Offset")
        J9calLab = Label(self.calib_ofset_group, text="J9 Offset")

        self.J1calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J2calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J3calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J4calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J5calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J6calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J7calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J8calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")
        self.J9calOffEntryField = Entry(self.calib_ofset_group, width=8, justify="center")

        self.J1calOffEntryField.place(x=55, y=0)
        self.J2calOffEntryField.place(x=55, y=30)
        self.J3calOffEntryField.place(x=55, y=60)
        self.J4calOffEntryField.place(x=55, y=90)
        self.J5calOffEntryField.place(x=55, y=120)
        self.J6calOffEntryField.place(x=55, y=150)
        self.J7calOffEntryField.place(x=55, y=180)
        self.J8calOffEntryField.place(x=55, y=210)
        self.J9calOffEntryField.place(x=55, y=240)

        J1calLab.place(x=0, y=0)
        J2calLab.place(x=0, y=30)
        J3calLab.place(x=0, y=60)
        J4calLab.place(x=0, y=90)
        J5calLab.place(x=0, y=120)
        J6calLab.place(x=0, y=150)
        J7calLab.place(x=0, y=180)
        J8calLab.place(x=0, y=210)
        J9calLab.place(x=0, y=240)
        self.calib_only_group = tk.LabelFrame(self.tab0_rb, text="Calibration Axis", labelanchor="nw", width=900, height=55)
        self.calib_only_group.propagate(False)
        self.calib_only_group.place(x=0, y=580)
        #
        CalJ1But = Button(self.calib_only_group, text="Calibrate J1 Only", command=self.calRobotJ1)
        CalJ1But.place(x=0, y=0)
        CalJ2But = Button(self.calib_only_group, text="Calibrate J2 Only", command=self.calRobotJ2)
        CalJ2But.place(x=100, y=0)
        CalJ3But = Button(self.calib_only_group, text="Calibrate J3 Only", command=self.calRobotJ3)
        CalJ3But.place(x=200, y=0)
        CalJ4But = Button(self.calib_only_group, text="Calibrate J4 Only", command=self.calRobotJ4)
        CalJ4But.place(x=300, y=0)
        CalJ5But = Button(self.calib_only_group, text="Calibrate J5 Only", command=self.calRobotJ5)
        CalJ5But.place(x=400, y=0)
        CalJ6But = Button(self.calib_only_group, text="Calibrate J6 Only", command=self.calRobotJ6)
        CalJ6But.place(x=500, y=0)
        CalZeroBut = Button(self.calib_only_group, text="Force CaL to Home", command=self.CalZeroPos)
        CalZeroBut.place(x=600, y=0)
        CalRestBut = Button(self.calib_only_group, text="Force Cal to Rest", command=self.CalRestPos)
        CalRestBut.place(x=715, y=0)
        self.encoder_en_group = tk.LabelFrame(self.tab0_rb, text="Encoder control", labelanchor="nw", width=200, height=300)
        self.encoder_en_group.propagate(False)
        self.encoder_en_group.place(x=500, y=10)

        J1OpenLoopCbut = Checkbutton(self.encoder_en_group, text="J1 Open Loop (disable encoder)", variable=self.J1OpenLoopStat)
        J1OpenLoopCbut.place(x=0, y=0)
        J2OpenLoopCbut = Checkbutton(self.encoder_en_group, text="J2 Open Loop (disable encoder)", variable=self.J2OpenLoopStat)
        J2OpenLoopCbut.place(x=0, y=30)
        J3OpenLoopCbut = Checkbutton(self.encoder_en_group, text="J3 Open Loop (disable encoder)", variable=self.J3OpenLoopStat)
        J3OpenLoopCbut.place(x=0, y=60)
        J4OpenLoopCbut = Checkbutton(self.encoder_en_group, text="J4 Open Loop (disable encoder)", variable=self.J4OpenLoopStat)
        J4OpenLoopCbut.place(x=0, y=90)
        J5OpenLoopCbut = Checkbutton(self.encoder_en_group, text="J5 Open Loop (disable encoder)", variable=self.J5OpenLoopStat)
        J5OpenLoopCbut.place(x=0, y=120)
        J6OpenLoopCbut = Checkbutton(self.encoder_en_group, text="J6 Open Loop (disable encoder)", variable=self.J6OpenLoopStat)
        J6OpenLoopCbut.place(x=0, y=150)

        self.th7_calib_group = tk.LabelFrame(self.tab0_rb, text="7th Axis Calibration", labelanchor="nw", width=200, height=200)
        self.th7_calib_group.propagate(False)
        self.th7_calib_group.place(x=000, y=310)
        axis7lengthLab = Label(self.th7_calib_group, text="7th Axis Length:")
        axis7lengthLab.place(x=0, y=10)
        axis7rotLab = Label(self.th7_calib_group, text="MM per Rotation:")
        axis7rotLab.place(x=0, y=40)
        axis7stepsLab = Label(self.th7_calib_group, text="Drive Steps:")
        axis7stepsLab.place(x=0, y=70)
        axis7pinsetLab = Label(self.th7_calib_group, font=("Arial", 8), text="StepPin = 12 / DirPin = 13 / CalPin = 36")
        axis7pinsetLab.place(x=0, y=160)
        J7zerobut = Button(self.th7_calib_group, text="Set Axis 7 Calibration to Zero", width=26, command=self.zeroAxis7)
        J7zerobut.place(x=0, y=105)
        J7calbut = Button(self.th7_calib_group, text="Autocalibrate Axis 7", width=26, command=self.calRobotJ7)
        J7calbut.place(x=0, y=135)
        self.axis7lengthEntryField = Entry(self.th7_calib_group, width=6, justify="center")
        self.axis7lengthEntryField.place(x=100, y=10)
        self.axis7rotEntryField = Entry(self.th7_calib_group, width=6, justify="center")
        self.axis7rotEntryField.place(x=100, y=40)
        self.axis7stepsEntryField = Entry(self.th7_calib_group, width=6, justify="center")
        self.axis7stepsEntryField.place(x=100, y=70)
        #
        self.th8_calib_group = tk.LabelFrame(self.tab0_rb, text="8th Axis Calibration", labelanchor="nw", width=200, height=200)
        self.th8_calib_group.propagate(False)
        self.th8_calib_group.place(x=210, y=310)
        axis8lengthLab = Label(self.th8_calib_group, text="8th Axis Length:")
        axis8lengthLab.place(x=0, y=10)
        axis8rotLab = Label(self.th8_calib_group, text="MM per Rotation:")
        axis8rotLab.place(x=0, y=40)
        axis8stepsLab = Label(self.th8_calib_group, text="Drive Steps:")
        axis8stepsLab.place(x=0, y=70)
        axis8pinsetLab = Label(self.th8_calib_group, font=("Arial", 8), text="StepPin = 32 / DirPin = 33 / CalPin = 37")
        axis8pinsetLab.place(x=0, y=160)
        J8zerobut = Button(self.th8_calib_group, text="Set Axis 8 Calibration to Zero", width=26, command=self.zeroAxis8)
        J8zerobut.place(x=0, y=105)
        J8calbut = Button(self.th8_calib_group, text="Autocalibrate Axis 8", width=26, command=self.calRobotJ8)
        J8calbut.place(x=0, y=135)
        self.axis8lengthEntryField = Entry(self.th8_calib_group, width=6, justify="center")
        self.axis8lengthEntryField.place(x=100, y=10)
        self.axis8rotEntryField = Entry(self.th8_calib_group, width=6, justify="center")
        self.axis8rotEntryField.place(x=100, y=40)
        self.axis8stepsEntryField = Entry(self.th8_calib_group, width=6, justify="center")
        self.axis8stepsEntryField.place(x=100, y=70)
        #
        self.th9_calib_group = tk.LabelFrame(self.tab0_rb, text="9th Axis Calibration", labelanchor="nw", width=200, height=200)
        self.th9_calib_group.propagate(False)
        self.th9_calib_group.place(x=420, y=310)
        axis9lengthLab = Label(self.th9_calib_group, text="9th Axis Length:")
        axis9lengthLab.place(x=0, y=10)
        axis9rotLab = Label(self.th9_calib_group, text="MM per Rotation:")
        axis9rotLab.place(x=0, y=40)
        axis9stepsLab = Label(self.th9_calib_group, text="Drive Steps:")
        axis9stepsLab.place(x=0, y=70)
        axis9pinsetLab = Label(self.th9_calib_group, font=("Arial", 8), text="StepPin = 34 / DirPin = 35 / CalPin = 38")
        axis9pinsetLab.place(x=0, y=160)
        J9zerobut = Button(self.th9_calib_group, text="Set Axis 9 Calibration to Zero", width=26, command=self.zeroAxis9)
        J9zerobut.place(x=0, y=105)
        J9calbut = Button(self.th9_calib_group, text="Autocalibrate Axis 9", width=26, command=self.calRobotJ9)
        J9calbut.place(x=0, y=135)
        self.axis9lengthEntryField = Entry(self.th9_calib_group, width=6, justify="center")
        self.axis9lengthEntryField.place(x=100, y=10)
        self.axis9rotEntryField = Entry(self.th9_calib_group, width=6, justify="center")
        self.axis9rotEntryField.place(x=100, y=40)
        self.axis9stepsEntryField = Entry(self.th9_calib_group, width=6, justify="center")
        self.axis9stepsEntryField.place(x=100, y=70)
        ###########
        cmdSentLab = Label(self.tab0_rb, text="Last Command Sent to Controller")
        cmdSentLab.place(x=10, y=80)
        cmdRecLab = Label(self.tab0_rb, text="Last Response From Controller")
        cmdRecLab.place(x=10, y=120)
        self.cmdSentEntryField = Entry(self.tab0_rb, width=70, justify="center")
        self.cmdSentEntryField.place(x=10, y=100)
        self.cmdRecEntryField = Entry(self.tab0_rb, width=70, justify="center")
        self.cmdRecEntryField.place(x=10, y=140)
        saveCalBut = Button(self.tab0_rb, text="    SAVE    ", width=26, command=self.SaveAndApplyCalibration)
        saveCalBut.place(x=750, y=520)

    def calRobotJ1(self):
        command = "LLA1B0C0D0E0F0G0H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J1 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J1 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ2(self):
        command = "LLA0B1C0D0E0F0G0H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J2 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J2 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ3(self):
        command = "LLA0B0C1D0E0F0G0H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J3 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J3 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ4(self):
        command = "LLA0B0C0D1E0F0G0H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J4 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J4 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ5(self):
        command = "LLA0B0C0D0E1F0G0H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J5 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J5 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ6(self):
        command = "LLA0B0C0D0E0F1G0H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J6 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J6 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ7(self):
        command = "LLA0B0C0D0E0F0G1H0I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J7 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J7 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ8(self):
        command = "LLA0B0C0D0E0F0G0H1I0" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J8 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J8 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotJ9(self):
        command = "LLA0B0C0D0E0F0G0H0I1" + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(
            self.J7calOff) + "Q" + str(self.J8calOff) + "R" + str(
            self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "J9 Calibrated Successfully"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "J9 Calibrated Failed"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotMid(self):
        print("foo")
        # add mid command

    def CalZeroPos(self):
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        command = "SPA0B0C0D0E90F0\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = self.ser.read()
        self.requestPos()
        self.almStatusLab.config(text="Calibration Forced to Home", style="Warn.TLabel")
        self.almStatusLab2.config(text="Calibration Forced to Home", style="Warn.TLabel")
        message = "Calibration Forced to Home - this is for commissioning and testing - be careful!"
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def CalRestPos(self):
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        command = "SPA0B0C-89D0E0F0\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = self.ser.read()
        self.requestPos()
        self.almStatusLab.config(text="Calibration Forced to Vertical Rest Pos", style="Warn.TLabel")
        self.almStatusLab2.config(text="Calibration Forced to Vertical Rest Pos", style="Warn.TLabel")
        message = "Calibration Forced to Vertical - this is for commissioning and testing - be careful!"
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))

    def calRobotAll(self):
        ##### STAGE 1 ########
        command = "LL" + "A" + str(self.J1CalStatVal) + "B" + str(self.J2CalStatVal) + "C" + str(self.J3CalStatVal) + "D" + str(self.J4CalStatVal) + "E" + str(self.J5CalStatVal) + "F" + str(self.J6CalStatVal) + "G" + str(
            self.J7CalStatVal) + "H" + str(
            self.J8CalStatVal) + "I" + str(
            self.J9CalStatVal) + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(self.J6calOff) + "P" + str(self.J7calOff) + "Q" + str(
            self.J8calOff) + "R" + str(self.J9calOff) + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        response = str(self.ser.readline().strip(), 'utf-8')
        self.cmdRecEntryField.delete(0, 'end')
        self.cmdRecEntryField.insert(0, response)
        if (response[:1] == 'A'):
            self.displayPosition(response)
            message = "Auto Calibration Stage 1 Successful"
            self.almStatusLab.config(text=message, style="OK.TLabel")
            self.almStatusLab2.config(text=message, style="OK.TLabel")
        else:
            message = "Auto Calibration Stage 1 Failed - See Log"
            self.almStatusLab.config(text=message, style="Alarm.TLabel")
            self.almStatusLab2.config(text=message, style="Alarm.TLabel")
            self.ErrorHandler(response)
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))
        ##### STAGE 2 ########
        CalStatVal2 = int(self.J1CalStatVal2) + int(self.J2CalStatVal2) + int(self.J3CalStatVal2) + int(self.J4CalStatVal2) + int(self.J5CalStatVal2) + int(self.J6CalStatVal2)
        if (CalStatVal2 > 0):
            command = "LL" + "A" + str(self.J1CalStatVal2) + "B" + str(self.J2CalStatVal2) + "C" + str(self.J3CalStatVal2) + "D" + str(self.J4CalStatVal2) + "E" + str(self.J5CalStatVal2) + "F" + str(self.J6CalStatVal2) + "G" + str(
                self.J7CalStatVal2) + "H" + str(
                self.J8CalStatVal2) + "I" + str(self.J9CalStatVal2) + "J" + str(self.J1calOff) + "K" + str(self.J2calOff) + "L" + str(self.J3calOff) + "M" + str(self.J4calOff) + "N" + str(self.J5calOff) + "O" + str(
                self.J6calOff) + "P" + str(self.J7calOff) + "Q" + str(
                self.J8calOff) + "R" + str(self.J9calOff) + "\n"
            self.ser.write(command.encode())
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.flushInput()
            response = str(self.ser.readline().strip(), 'utf-8')
            self.cmdRecEntryField.delete(0, 'end')
            self.cmdRecEntryField.insert(0, response)
            if (response[:1] == 'A'):
                self.displayPosition(response)
                message = "Auto Calibration Stage 2 Successful"
                self.almStatusLab.config(text=message, style="OK.TLabel")
                self.almStatusLab2.config(text=message, style="OK.TLabel")
            else:
                message = "Auto Calibration Stage 2 Failed - See Log"
                self.almStatusLab.config(text=message, style="Alarm.TLabel")
                self.almStatusLab2.config(text=message, style="Alarm.TLabel")
                self.ErrorHandler(response)
            Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
            self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
            value = self.tab5_rb.ElogView.get(0, END)
            pickle.dump(value, open("ErrorLog", "wb"))

    def zeroAxis7(self):
        command = "Z7" + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.almStatusLab.config(text="J7 Calibration Forced to Zero", style="Warn.TLabel")
        self.almStatusLab2.config(text="J7 Calibration Forced to Zero", style="Warn.TLabel")
        message = "J7 Calibration Forced to Zero - this is for commissioning and testing - be careful!"
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))
        response = str(self.ser.readline().strip(), 'utf-8')
        self.displayPosition(response)

    def zeroAxis8(self):
        command = "Z8" + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.almStatusLab.config(text="J8 Calibration Forced to Zero", style="Warn.TLabel")
        self.almStatusLab2.config(text="J8 Calibration Forced to Zero", style="Warn.TLabel")
        message = "J8 Calibration Forced to Zero - this is for commissioning and testing - be careful!"
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))
        response = str(self.ser.readline().strip(), 'utf-8')
        self.displayPosition(response)

    def zeroAxis9(self):
        command = "Z9" + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.almStatusLab.config(text="J9 Calibration Forced to Zero", style="Warn.TLabel")
        self.almStatusLab2.config(text="J9 Calibration Forced to Zero", style="Warn.TLabel")
        message = "J9 Calibration Forced to Zero - this is for commissioning and testing - be careful!"
        Curtime = datetime.datetime.now().strftime("%B %d %Y - %I:%M%p")
        self.tab5_rb.ElogView.insert(END, Curtime + " - " + message)
        value = self.tab5_rb.ElogView.get(0, END)
        pickle.dump(value, open("ErrorLog", "wb"))
        response = str(self.ser.readline().strip(), 'utf-8')
        self.displayPosition(response)

    def SaveAndApplyCalibration(self):

        self.J7PosCur = self.J7curAngEntryField.get()
        self.J8PosCur = self.J8curAngEntryField.get()
        self.J9PosCur = self.J9curAngEntryField.get()
        # self.VisFileLoc = VisFileLocEntryField.get()
        self.VisProg = self.Jvisoptions.get()
        # self.VisOrigXpix = float(VisPicOxPEntryField.get())
        # self.VisOrigXmm  = float(VisPicOxMEntryField.get())
        # self.VisOrigYpix = float(VisPicOyPEntryField.get())
        # self.VisOrigYmm  = float(VisPicOyMEntryField.get())
        # self.VisEndXpix  = float(VisPicXPEntryField.get())
        # self.VisEndXmm   = float(VisPicXMEntryField.get())
        # self.VisEndYpix  = float(VisPicYPEntryField.get())
        # self.VisEndYmm   = float(VisPicYMEntryField.get())
        self.J1calOff = float(self.J1calOffEntryField.get())
        self.J2calOff = float(self.J2calOffEntryField.get())
        self.J3calOff = float(self.J3calOffEntryField.get())
        self.J4calOff = float(self.J4calOffEntryField.get())
        self.J5calOff = float(self.J5calOffEntryField.get())
        self.J6calOff = float(self.J6calOffEntryField.get())
        self.J7calOff = float(self.J7calOffEntryField.get())
        self.J8calOff = float(self.J8calOffEntryField.get())
        self.J9calOff = float(self.J9calOffEntryField.get())
        self.J1OpenLoopVal = int(self.J1OpenLoopStat.get())
        self.J2OpenLoopVal = int(self.J2OpenLoopStat.get())
        self.J3OpenLoopVal = int(self.J3OpenLoopStat.get())
        self.J4OpenLoopVal = int(self.J4OpenLoopStat.get())
        self.J5OpenLoopVal = int(self.J5OpenLoopStat.get())
        self.J6OpenLoopVal = int(self.J6OpenLoopStat.get())
        self.DisableWristRotVal = int(self.DisableWristRot.get())
        self.J1CalStatVal = int(self.J1CalStat.get())
        self.J2CalStatVal = int(self.J2CalStat.get())
        self.J3CalStatVal = int(self.J3CalStat.get())
        self.J4CalStatVal = int(self.J4CalStat.get())
        self.J5CalStatVal = int(self.J5CalStat.get())
        self.J6CalStatVal = int(self.J6CalStat.get())
        self.J7CalStatVal = int(self.J7CalStat.get())
        self.J8CalStatVal = int(self.J8CalStat.get())
        self.J9CalStatVal = int(self.J9CalStat.get())
        self.J1CalStatVal2 = int(self.J1CalStat2.get())
        self.J2CalStatVal2 = int(self.J2CalStat2.get())
        self.J3CalStatVal2 = int(self.J3CalStat2.get())
        self.J4CalStatVal2 = int(self.J4CalStat2.get())
        self.J5CalStatVal2 = int(self.J5CalStat2.get())
        self.J6CalStatVal2 = int(self.J6CalStat2.get())
        self.J7CalStatVal2 = int(self.J7CalStat2.get())
        self.J8CalStatVal2 = int(self.J8CalStat2.get())
        self.J9CalStatVal2 = int(self.J9CalStat2.get())
        self.J7PosLim = float(self.axis7lengthEntryField.get())
        self.J7rotation = float(self.axis7rotEntryField.get())
        self.J7steps = float(self.axis7stepsEntryField.get())
        self.J8length = float(self.axis8lengthEntryField.get())
        self.J8rotation = float(self.axis8rotEntryField.get())
        self.J8steps = float(self.axis8stepsEntryField.get())
        self.J9length = float(self.axis9lengthEntryField.get())
        self.J9rotation = float(self.axis9rotEntryField.get())
        self.J9steps = float(self.axis9stepsEntryField.get())
        try:
            self.updateParams()
            time.sleep(.1)
            self.calExtAxis()
        except:
            print("no serial connection with Teensy board")
        self.savePosData()

    def savePosData(self):



        self.calibration.delete(0, END)
        self.calibration.insert(END, self.J1AngCur)
        self.calibration.insert(END, self.J2AngCur)
        self.calibration.insert(END, self.J3AngCur)
        self.calibration.insert(END, self.J4AngCur)
        self.calibration.insert(END, self.J5AngCur)
        self.calibration.insert(END, self.J6AngCur)
        self.calibration.insert(END, self.XcurPos)
        self.calibration.insert(END, self.YcurPos)
        self.calibration.insert(END, self.ZcurPos)
        self.calibration.insert(END, self.RzcurPos)
        self.calibration.insert(END, self.RycurPos)
        self.calibration.insert(END, self.RxcurPos)

        self.calibration.insert(END, self.ProgEntryField.get())
        self.calibration.insert(END, self.servo0onEntryField.get())
        self.calibration.insert(END, self.servo0offEntryField.get())
        self.calibration.insert(END, self.servo1onEntryField.get())
        self.calibration.insert(END, self.servo1offEntryField.get())
        self.calibration.insert(END, self.DO1onEntryField.get())
        self.calibration.insert(END, self.DO1offEntryField.get())
        self.calibration.insert(END, self.DO2onEntryField.get())
        self.calibration.insert(END, self.DO2offEntryField.get())
        self.calibration.insert(END, self.TFxEntryField.get())
        self.calibration.insert(END, self.TFyEntryField.get())
        self.calibration.insert(END, self.TFzEntryField.get())
        self.calibration.insert(END, self.TFrxEntryField.get())
        self.calibration.insert(END, self.TFryEntryField.get())
        self.calibration.insert(END, self.TFrzEntryField.get())
        self.calibration.insert(END, self.J7curAngEntryField.get())
        self.calibration.insert(END, self.J8curAngEntryField.get())
        self.calibration.insert(END, self.J9curAngEntryField.get())
        self.calibration.insert(END, "VisFileLocEntryField")
        self.calibration.insert(END, self.Jvisoptions.get())
        self.calibration.insert(END, "VisPicOxPEntryField")
        self.calibration.insert(END, "VisPicOxMEntryField")
        self.calibration.insert(END, "VisPicOyPEntryField")
        self.calibration.insert(END, "VisPicOyMEntryField")
        self.calibration.insert(END, "VisPicXPEntryField")
        self.calibration.insert(END, "VisPicXMEntryField")
        self.calibration.insert(END, "VisPicYPEntryField")
        self.calibration.insert(END, "VisPicYMEntryField")
        self.calibration.insert(END, self.J1calOffEntryField.get())
        self.calibration.insert(END, self.J2calOffEntryField.get())
        self.calibration.insert(END, self.J3calOffEntryField.get())
        self.calibration.insert(END, self.J4calOffEntryField.get())
        self.calibration.insert(END, self.J5calOffEntryField.get())
        self.calibration.insert(END, self.J6calOffEntryField.get())
        self.calibration.insert(END, self.J1OpenLoopVal)
        self.calibration.insert(END, self.J2OpenLoopVal)
        self.calibration.insert(END, self.J3OpenLoopVal)
        self.calibration.insert(END, self.J4OpenLoopVal)
        self.calibration.insert(END, self.J5OpenLoopVal)
        self.calibration.insert(END, self.J6OpenLoopVal)

        self.calibration.insert(END, self.J1CalStatVal)
        self.calibration.insert(END, self.J2CalStatVal)
        self.calibration.insert(END, self.J3CalStatVal)
        self.calibration.insert(END, self.J4CalStatVal)
        self.calibration.insert(END, self.J5CalStatVal)
        self.calibration.insert(END, self.J6CalStatVal)
        self.calibration.insert(END, self.J7PosLim)
        self.calibration.insert(END, self.J7rotation)
        self.calibration.insert(END, self.J7steps)
        self.calibration.insert(END, self.J7StepCur)  # is this used?
        self.calibration.insert(END, self.J1CalStatVal2)
        self.calibration.insert(END, self.J2CalStatVal2)
        self.calibration.insert(END, self.J3CalStatVal2)
        self.calibration.insert(END, self.J4CalStatVal2)
        self.calibration.insert(END, self.J5CalStatVal2)
        self.calibration.insert(END, self.J6CalStatVal2)
        self.calibration.insert(END, self.VisBrightSlide.get())
        self.calibration.insert(END, self.VisContrastSlide.get())
        self.calibration.insert(END, self.VisBacColorEntryField.get())
        self.calibration.insert(END, self.VisScoreEntryField.get())
        self.calibration.insert(END, self.VisX1PixEntryField.get())
        self.calibration.insert(END, self.VisY1PixEntryField.get())
        self.calibration.insert(END, self.VisX2PixEntryField.get())
        self.calibration.insert(END, self.VisY2PixEntryField.get())
        self.calibration.insert(END, self.VisX1RobEntryField.get())
        self.calibration.insert(END, self.VisY1RobEntryField.get())
        self.calibration.insert(END, self.VisX2RobEntryField.get())
        self.calibration.insert(END, self.VisY2RobEntryField.get())
        self.calibration.insert(END, self.VisZoomSlide.get())
        self.calibration.insert(END, self.pick180.get())
        self.calibration.insert(END, self.pickClosest.get())
        self.calibration.insert(END, self.Jvisoptions.get())
        self.calibration.insert(END, self.fullRot.get())
        self.calibration.insert(END, self.autoBG.get())
        self.calibration.insert(END, self.mX1)
        self.calibration.insert(END, self.mY1)
        self.calibration.insert(END, self.mX2)
        self.calibration.insert(END, self.mY2)
        self.calibration.insert(END, self.J8length)
        self.calibration.insert(END, self.J8rotation)
        self.calibration.insert(END, self.J8steps)
        self.calibration.insert(END, self.J9length)
        self.calibration.insert(END, self.J9rotation)
        self.calibration.insert(END, self.J9steps)
        self.calibration.insert(END, self.J7calOffEntryField.get())
        self.calibration.insert(END, self.J8calOffEntryField.get())
        self.calibration.insert(END, self.J9calOffEntryField.get())
        self.calibration.insert(END, self.GC_ST_E1_EntryField.get())
        self.calibration.insert(END, self.GC_ST_E2_EntryField.get())
        self.calibration.insert(END, self.GC_ST_E3_EntryField.get())
        self.calibration.insert(END, self.GC_ST_E4_EntryField.get())
        self.calibration.insert(END, self.GC_ST_E5_EntryField.get())
        self.calibration.insert(END, self.GC_ST_E6_EntryField.get())
        self.calibration.insert(END, self.GC_SToff_E1_EntryField.get())
        self.calibration.insert(END, self.GC_SToff_E2_EntryField.get())
        self.calibration.insert(END, self.GC_SToff_E3_EntryField.get())
        self.calibration.insert(END, self.GC_SToff_E4_EntryField.get())
        self.calibration.insert(END, self.GC_SToff_E5_EntryField.get())
        self.calibration.insert(END, self.GC_SToff_E6_EntryField.get())
        self.calibration.insert(END, self.DisableWristRotVal)
        self.calibration.insert(END, self.J1MotDirEntryField.get())
        self.calibration.insert(END, self.J2MotDirEntryField.get())
        self.calibration.insert(END, self.J3MotDirEntryField.get())
        self.calibration.insert(END, self.J4MotDirEntryField.get())
        self.calibration.insert(END, self.J5MotDirEntryField.get())
        self.calibration.insert(END, self.J6MotDirEntryField.get())
        self.calibration.insert(END, self.J7MotDirEntryField.get())
        self.calibration.insert(END, self.J8MotDirEntryField.get())
        self.calibration.insert(END, self.J9MotDirEntryField.get())
        self.calibration.insert(END, self.J1CalDirEntryField.get())
        self.calibration.insert(END, self.J2CalDirEntryField.get())
        self.calibration.insert(END, self.J3CalDirEntryField.get())
        self.calibration.insert(END, self.J4CalDirEntryField.get())
        self.calibration.insert(END, self.J5CalDirEntryField.get())
        self.calibration.insert(END, self.J6CalDirEntryField.get())
        self.calibration.insert(END, self.J7CalDirEntryField.get())
        self.calibration.insert(END, self.J8CalDirEntryField.get())
        self.calibration.insert(END, self.J9CalDirEntryField.get())
        self.calibration.insert(END, self.J1PosLimEntryField.get())
        self.calibration.insert(END, self.J1NegLimEntryField.get())
        self.calibration.insert(END, self.J2PosLimEntryField.get())
        self.calibration.insert(END, self.J2NegLimEntryField.get())
        self.calibration.insert(END, self.J3PosLimEntryField.get())
        self.calibration.insert(END, self.J3NegLimEntryField.get())
        self.calibration.insert(END, self.J4PosLimEntryField.get())
        self.calibration.insert(END, self.J4NegLimEntryField.get())
        self.calibration.insert(END, self.J5PosLimEntryField.get())
        self.calibration.insert(END, self.J5NegLimEntryField.get())
        self.calibration.insert(END, self.J6PosLimEntryField.get())
        self.calibration.insert(END, self.J6NegLimEntryField.get())
        self.calibration.insert(END, self.J1StepDegEntryField.get())
        self.calibration.insert(END, self.J2StepDegEntryField.get())
        self.calibration.insert(END, self.J3StepDegEntryField.get())
        self.calibration.insert(END, self.J4StepDegEntryField.get())
        self.calibration.insert(END, self.J5StepDegEntryField.get())
        self.calibration.insert(END, self.J6StepDegEntryField.get())
        self.calibration.insert(END, self.J1DriveMSEntryField.get())
        self.calibration.insert(END, self.J2DriveMSEntryField.get())
        self.calibration.insert(END, self.J3DriveMSEntryField.get())
        self.calibration.insert(END, self.J4DriveMSEntryField.get())
        self.calibration.insert(END, self.J5DriveMSEntryField.get())
        self.calibration.insert(END, self.J6DriveMSEntryField.get())
        self.calibration.insert(END, self.J1EncCPREntryField.get())
        self.calibration.insert(END, self.J2EncCPREntryField.get())
        self.calibration.insert(END, self.J3EncCPREntryField.get())
        self.calibration.insert(END, self.J4EncCPREntryField.get())
        self.calibration.insert(END, self.J5EncCPREntryField.get())
        self.calibration.insert(END, self.J6EncCPREntryField.get())
        self.calibration.insert(END, self.J1ΘEntryField.get())
        self.calibration.insert(END, self.J2ΘEntryField.get())
        self.calibration.insert(END, self.J3ΘEntryField.get())
        self.calibration.insert(END, self.J4ΘEntryField.get())
        self.calibration.insert(END, self.J5ΘEntryField.get())
        self.calibration.insert(END, self.J6ΘEntryField.get())
        self.calibration.insert(END, self.J1αEntryField.get())
        self.calibration.insert(END, self.J2αEntryField.get())
        self.calibration.insert(END, self.J3αEntryField.get())
        self.calibration.insert(END, self.J4αEntryField.get())
        self.calibration.insert(END, self.J5αEntryField.get())
        self.calibration.insert(END, self.J6αEntryField.get())
        self.calibration.insert(END, self.J1dEntryField.get())
        self.calibration.insert(END, self.J2dEntryField.get())
        self.calibration.insert(END, self.J3dEntryField.get())
        self.calibration.insert(END, self.J4dEntryField.get())
        self.calibration.insert(END, self.J5dEntryField.get())
        self.calibration.insert(END, self.J6dEntryField.get())
        self.calibration.insert(END, self.J1aEntryField.get())
        self.calibration.insert(END, self.J2aEntryField.get())
        self.calibration.insert(END, self.J3aEntryField.get())
        self.calibration.insert(END, self.J4aEntryField.get())
        self.calibration.insert(END, self.J5aEntryField.get())
        self.calibration.insert(END, self.J6aEntryField.get())
        self.calibration.insert(END, self.GC_ST_WC_EntryField.get())
        self.calibration.insert(END, self.J7CalStatVal)
        self.calibration.insert(END, self.J8CalStatVal)
        self.calibration.insert(END, self.J9CalStatVal)
        self.calibration.insert(END, self.J7CalStatVal2)
        self.calibration.insert(END, self.J8CalStatVal2)
        self.calibration.insert(END, self.J9CalStatVal2)

        ###########
        value = self.calibration.get(0, END)
        pickle.dump(value, open("ARbot.cal", "wb"))

    def robot_tab_1_widget(self):
        self.Frame_Offset_group = tk.LabelFrame(self.tab1_rb, text="Tool Frame Offset", labelanchor="nw", width=240, height=150)
        self.Frame_Offset_group.propagate(False)
        self.Frame_Offset_group.place(x=720, y=480)

        UFxLab = Label(self.Frame_Offset_group, font=("Arial", 11), text="X")
        UFxLab.place(x=10, y=0)
        UFyLab = Label(self.Frame_Offset_group, font=("Arial", 11), text="Y")
        UFyLab.place(x=50, y=0)
        UFzLab = Label(self.Frame_Offset_group, font=("Arial", 11), text="Z")
        UFzLab.place(x=90, y=0)
        UFRxLab = Label(self.Frame_Offset_group, font=("Arial", 11), text="Rz")
        UFRxLab.place(x=130, y=0)
        UFRyLab = Label(self.Frame_Offset_group, font=("Arial", 11), text="Ry")
        UFRyLab.place(x=170, y=0)
        UFRzLab = Label(self.Frame_Offset_group, font=("Arial", 11), text="Rx")
        UFRzLab.place(x=210, y=0)

        self.TFxEntryField = Entry(self.Frame_Offset_group, width=5, justify="center")
        self.TFxEntryField.place(x=0, y=25)
        self.TFyEntryField = Entry(self.Frame_Offset_group, width=5, justify="center")
        self.TFyEntryField.place(x=40, y=25)
        self.TFzEntryField = Entry(self.Frame_Offset_group, width=5, justify="center")
        self.TFzEntryField.place(x=80, y=25)
        self.TFrzEntryField = Entry(self.Frame_Offset_group, width=5, justify="center")
        self.TFrzEntryField.place(x=120, y=25)
        self.TFryEntryField = Entry(self.Frame_Offset_group, width=5, justify="center")
        self.TFryEntryField.place(x=160, y=25)
        self.TFrxEntryField = Entry(self.Frame_Offset_group, width=5, justify="center")
        self.TFrxEntryField.place(x=200, y=25)

        DisableWristCbut = Checkbutton(self.Frame_Offset_group, text="Disable Wrist Rotation - Linear Moves", variable=self.DisableWristRot)
        DisableWristCbut.place(x=0, y=60)

        self.MOTOR_DIRECTIONS_group = tk.LabelFrame(self.tab1_rb, text="MOTOR DIRECTIONS", labelanchor="nw", width=170, height=250)
        self.MOTOR_DIRECTIONS_group.propagate(False)
        self.MOTOR_DIRECTIONS_group.place(x=0, y=0)
        ####  MOTOR DIRECTIONS ####

        J1MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J1 Motor Direction")
        J1MotDirLab.place(x=0, y=20 - 10)
        J2MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J2 Motor Direction")
        J2MotDirLab.place(x=0, y=45 - 10)
        J3MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J3 Motor Direction")
        J3MotDirLab.place(x=0, y=70 - 10)
        J4MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J4 Motor Direction")
        J4MotDirLab.place(x=0, y=95 - 10)
        J5MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J5 Motor Direction")
        J5MotDirLab.place(x=0, y=120 - 10)
        J6MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J6 Motor Direction")
        J6MotDirLab.place(x=0, y=145 - 10)
        J7MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J7 Motor Direction")
        J7MotDirLab.place(x=0, y=170 - 10)
        J8MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J8 Motor Direction")
        J8MotDirLab.place(x=0, y=195 - 10)
        J9MotDirLab = Label(self.MOTOR_DIRECTIONS_group, font=("Arial", 8), text="J9 Motor Direction")
        J9MotDirLab.place(x=0, y=220 - 10)

        self.J1MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J1MotDirEntryField.place(x=100, y=20 - 10)
        self.J2MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J2MotDirEntryField.place(x=100, y=45 - 10)
        self.J3MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J3MotDirEntryField.place(x=100, y=70 - 10)
        self.J4MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J4MotDirEntryField.place(x=100, y=95 - 10)
        self.J5MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J5MotDirEntryField.place(x=100, y=120 - 10)
        self.J6MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J6MotDirEntryField.place(x=100, y=145 - 10)
        self.J7MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J7MotDirEntryField.place(x=100, y=170 - 10)
        self.J8MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J8MotDirEntryField.place(x=100, y=195 - 10)
        self.J9MotDirEntryField = Entry(self.MOTOR_DIRECTIONS_group, width=8, justify="center")
        self.J9MotDirEntryField.place(x=100, y=220 - 10)

        self.CALIBRATION_DIRECTIONS_group = tk.LabelFrame(self.tab1_rb, text="CALIBRATION DIRECTIONS", labelanchor="nw", width=170, height=250)
        self.CALIBRATION_DIRECTIONS_group.propagate(False)
        self.CALIBRATION_DIRECTIONS_group.place(x=175, y=0)
        ####  CALIBRATION DIRECTIONS ####

        J1CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J1 Calibration Dir.")
        J1CalDirLab.place(x=0, y=10)
        J2CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J2 Calibration Dir.")
        J2CalDirLab.place(x=0, y=35)
        J3CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J3 Calibration Dir.")
        J3CalDirLab.place(x=0, y=60)
        J4CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J4 Calibration Dir.")
        J4CalDirLab.place(x=0, y=85)
        J5CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J5 Calibration Dir.")
        J5CalDirLab.place(x=0, y=110)
        J6CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J6 Calibration Dir.")
        J6CalDirLab.place(x=0, y=135)
        J7CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J7 Calibration Dir.")
        J7CalDirLab.place(x=0, y=160)
        J8CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J8 Calibration Dir.")
        J8CalDirLab.place(x=0, y=185)
        J9CalDirLab = Label(self.CALIBRATION_DIRECTIONS_group, font=("Arial", 8), text="J9 Calibration Dir.")
        J9CalDirLab.place(x=0, y=210)

        self.J1CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J1CalDirEntryField.place(x=100, y=10)
        self.J2CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J2CalDirEntryField.place(x=100, y=35)
        self.J3CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J3CalDirEntryField.place(x=100, y=60)
        self.J4CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J4CalDirEntryField.place(x=100, y=85)
        self.J5CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J5CalDirEntryField.place(x=100, y=110)
        self.J6CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J6CalDirEntryField.place(x=100, y=135)
        self.J7CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J7CalDirEntryField.place(x=100, y=160)
        self.J8CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J8CalDirEntryField.place(x=100, y=185)
        self.J9CalDirEntryField = Entry(self.CALIBRATION_DIRECTIONS_group, width=8, justify="center")
        self.J9CalDirEntryField.place(x=100, y=210)

        self.Axis_limit_group = tk.LabelFrame(self.tab1_rb, text="Axis limit", labelanchor="nw", width=170, height=350)
        self.Axis_limit_group.propagate(False)
        self.Axis_limit_group.place(x=0, y=255)
        #
        ### axis limits
        self.J1PosLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J1 Pos Limit")
        self.J1PosLimLab.place(x=0, y=20 - 10)
        self.J1NegLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J1 Neg Limit")
        self.J1NegLimLab.place(x=0, y=45 - 10)
        self.J2PosLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J2 Pos Limit")
        self.J2PosLimLab.place(x=0, y=70 - 10)
        self.J2NegLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J2 Neg Limit")
        self.J2NegLimLab.place(x=0, y=95 - 10)
        self.J3PosLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J3 Pos Limit")
        self.J3PosLimLab.place(x=0, y=120 - 10)
        self.J3NegLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J3 Neg Limit")
        self.J3NegLimLab.place(x=0, y=145 - 10)
        self.J4PosLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J4 Pos Limit")
        self.J4PosLimLab.place(x=0, y=170 - 10)
        self.J4NegLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J4 Neg Limit")
        self.J4NegLimLab.place(x=0, y=195 - 10)
        self.J5PosLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J5 Pos Limit")
        self.J5PosLimLab.place(x=0, y=220 - 10)
        self.J5NegLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J5 Neg Limit")
        self.J5NegLimLab.place(x=0, y=245 - 10)
        self.J6PosLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J6 Pos Limit")
        self.J6PosLimLab.place(x=0, y=270 - 10)
        self.J6NegLimLab = Label(self.Axis_limit_group, font=("Arial", 8), text="J6 Neg Limit")
        self.J6NegLimLab.place(x=0, y=295 - 10)

        self.J1PosLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J1PosLimEntryField.place(x=80, y=20 - 10)
        self.J1NegLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J1NegLimEntryField.place(x=80, y=45 - 10)
        self.J2PosLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J2PosLimEntryField.place(x=80, y=70 - 10)
        self.J2NegLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J2NegLimEntryField.place(x=80, y=95 - 10)
        self.J3PosLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J3PosLimEntryField.place(x=80, y=120 - 10)
        self.J3NegLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J3NegLimEntryField.place(x=80, y=145 - 10)
        self.J4PosLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J4PosLimEntryField.place(x=80, y=170 - 10)
        self.J4NegLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J4NegLimEntryField.place(x=80, y=195 - 10)
        self.J5PosLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J5PosLimEntryField.place(x=80, y=220 - 10)
        self.J5NegLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J5NegLimEntryField.place(x=80, y=245 - 10)
        self.J6PosLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J6PosLimEntryField.place(x=80, y=270 - 10)
        self.J6NegLimEntryField = Entry(self.Axis_limit_group, width=8, justify="center")
        self.J6NegLimEntryField.place(x=80, y=295 - 10)

        self.steps_per_degress_group = tk.LabelFrame(self.tab1_rb, text="steps per degress", labelanchor="nw", width=170, height=200)
        self.steps_per_degress_group.propagate(False)
        self.steps_per_degress_group.place(x=175, y=255)
        ### steps per degress
        J1StepDegLab = Label(self.steps_per_degress_group, font=("Arial", 8), text="J1 Step/Deg")
        J1StepDegLab.place(x=0, y=10)
        J2StepDegLab = Label(self.steps_per_degress_group, font=("Arial", 8), text="J2 Step/Deg")
        J2StepDegLab.place(x=0, y=35)
        J3StepDegLab = Label(self.steps_per_degress_group, font=("Arial", 8), text="J3 Step/Deg")
        J3StepDegLab.place(x=0, y=60)
        J4StepDegLab = Label(self.steps_per_degress_group, font=("Arial", 8), text="J4 Step/Deg")
        J4StepDegLab.place(x=0, y=85)
        J5StepDegLab = Label(self.steps_per_degress_group, font=("Arial", 8), text="J5 Step/Deg")
        J5StepDegLab.place(x=0, y=110)
        J6StepDegLab = Label(self.steps_per_degress_group, font=("Arial", 8), text="J6 Step/Deg")
        J6StepDegLab.place(x=0, y=135)

        self.J1StepDegEntryField = Entry(self.steps_per_degress_group, width=8, justify="center")
        self.J1StepDegEntryField.place(x=80, y=10)
        self.J2StepDegEntryField = Entry(self.steps_per_degress_group, width=8, justify="center")
        self.J2StepDegEntryField.place(x=80, y=35)
        self.J3StepDegEntryField = Entry(self.steps_per_degress_group, width=8, justify="center")
        self.J3StepDegEntryField.place(x=80, y=60)
        self.J4StepDegEntryField = Entry(self.steps_per_degress_group, width=8, justify="center")
        self.J4StepDegEntryField.place(x=80, y=85)
        self.J5StepDegEntryField = Entry(self.steps_per_degress_group, width=8, justify="center")
        self.J5StepDegEntryField.place(x=80, y=110)
        self.J6StepDegEntryField = Entry(self.steps_per_degress_group, width=8, justify="center")
        self.J6StepDegEntryField.place(x=80, y=135)
        #
        self.DRIVER_STEPS_group = tk.LabelFrame(self.tab1_rb, text="DRIVER STEPS", labelanchor="nw", width=170, height=200)
        self.DRIVER_STEPS_group.propagate(False)
        self.DRIVER_STEPS_group.place(x=350, y=0)
        ### DRIVER STEPS
        J1DriveMSLab = Label(self.DRIVER_STEPS_group, font=("Arial", 8), text="J1 Drive Microstep")
        J1DriveMSLab.place(x=0, y=20 - 10)
        J2DriveMSLab = Label(self.DRIVER_STEPS_group, font=("Arial", 8), text="J2 Drive Microstep")
        J2DriveMSLab.place(x=0, y=45 - 10)
        J3DriveMSLab = Label(self.DRIVER_STEPS_group, font=("Arial", 8), text="J3 Drive Microstep")
        J3DriveMSLab.place(x=0, y=70 - 10)
        J4DriveMSLab = Label(self.DRIVER_STEPS_group, font=("Arial", 8), text="J4 Drive Microstep")
        J4DriveMSLab.place(x=0, y=95 - 10)
        J5DriveMSLab = Label(self.DRIVER_STEPS_group, font=("Arial", 8), text="J5 Drive Microstep")
        J5DriveMSLab.place(x=0, y=120 - 10)
        J6DriveMSLab = Label(self.DRIVER_STEPS_group, font=("Arial", 8), text="J6 Drive Microstep")
        J6DriveMSLab.place(x=0, y=145 - 10)

        self.J1DriveMSEntryField = Entry(self.DRIVER_STEPS_group, width=8, justify="center")
        self.J1DriveMSEntryField.place(x=110, y=20 - 10)
        self.J2DriveMSEntryField = Entry(self.DRIVER_STEPS_group, width=8, justify="center")
        self.J2DriveMSEntryField.place(x=110, y=45 - 10)
        self.J3DriveMSEntryField = Entry(self.DRIVER_STEPS_group, width=8, justify="center")
        self.J3DriveMSEntryField.place(x=110, y=70 - 10)
        self.J4DriveMSEntryField = Entry(self.DRIVER_STEPS_group, width=8, justify="center")
        self.J4DriveMSEntryField.place(x=110, y=95 - 10)
        self.J5DriveMSEntryField = Entry(self.DRIVER_STEPS_group, width=8, justify="center")
        self.J5DriveMSEntryField.place(x=110, y=120 - 10)
        self.J6DriveMSEntryField = Entry(self.DRIVER_STEPS_group, width=8, justify="center")
        self.J6DriveMSEntryField.place(x=110, y=145 - 10)
        #
        self.ENCODER_CPR_group = tk.LabelFrame(self.tab1_rb, text="ENCODER CPR", labelanchor="nw", width=170, height=200)
        self.ENCODER_CPR_group.propagate(False)
        self.ENCODER_CPR_group.place(x=350, y=175)

        ###ENCODER CPR
        J1EncCPRLab = Label(self.ENCODER_CPR_group, font=("Arial", 8), text="J1 Encoder CPR")
        J1EncCPRLab.place(x=0, y=10)
        J2EncCPRLab = Label(self.ENCODER_CPR_group, font=("Arial", 8), text="J2 Encoder CPR")
        J2EncCPRLab.place(x=0, y=35)
        J3EncCPRLab = Label(self.ENCODER_CPR_group, font=("Arial", 8), text="J3 Encoder CPR")
        J3EncCPRLab.place(x=0, y=60)
        J4EncCPRLab = Label(self.ENCODER_CPR_group, font=("Arial", 8), text="J4 Encoder CPR")
        J4EncCPRLab.place(x=0, y=85)
        J5EncCPRLab = Label(self.ENCODER_CPR_group, font=("Arial", 8), text="J5 Encoder CPR")
        J5EncCPRLab.place(x=0, y=110)
        J6EncCPRLab = Label(self.ENCODER_CPR_group, font=("Arial", 8), text="J6 Encoder CPR")
        J6EncCPRLab.place(x=0, y=135)

        self.J1EncCPREntryField = Entry(self.ENCODER_CPR_group, width=8, justify="center")
        self.J1EncCPREntryField.place(x=110, y=10)
        self.J2EncCPREntryField = Entry(self.ENCODER_CPR_group, width=8, justify="center")
        self.J2EncCPREntryField.place(x=110, y=35)
        self.J3EncCPREntryField = Entry(self.ENCODER_CPR_group, width=8, justify="center")
        self.J3EncCPREntryField.place(x=110, y=60)
        self.J4EncCPREntryField = Entry(self.ENCODER_CPR_group, width=8, justify="center")
        self.J4EncCPREntryField.place(x=110, y=85)
        self.J5EncCPREntryField = Entry(self.ENCODER_CPR_group, width=8, justify="center")
        self.J5EncCPREntryField.place(x=110, y=110)
        self.J6EncCPREntryField = Entry(self.ENCODER_CPR_group, width=8, justify="center")
        self.J6EncCPREntryField.place(x=110, y=135)

        ###
        self.KINEMATIC_TABLE_group = tk.LabelFrame(self.tab1_rb, text="KINEMATIC TABLE", labelanchor="nw", width=300, height=200)
        self.KINEMATIC_TABLE_group.propagate(False)
        self.KINEMATIC_TABLE_group.place(x=350, y=400)

        J1DHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="J1")
        J1DHparamLab.place(x=600 - 600, y=45 - 20)
        J1DHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="J2")
        J1DHparamLab.place(x=600 - 600, y=70 - 20)
        J1DHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="J3")
        J1DHparamLab.place(x=600 - 600, y=95 - 20)
        J1DHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="J4")
        J1DHparamLab.place(x=600 - 600, y=120 - 20)
        J1DHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="J5")
        J1DHparamLab.place(x=600 - 600, y=145 - 20)
        J1DHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="J6")
        J1DHparamLab.place(x=600 - 600, y=170 - 20)

        ΘDHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="DH-Θ")
        ΘDHparamLab.place(x=645 - 600, y=20 - 20)
        αDHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="DH-α")
        αDHparamLab.place(x=700 - 600, y=20 - 20)
        dDHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="DH-d")
        dDHparamLab.place(x=755 - 600, y=20 - 20)
        aDHparamLab = Label(self.KINEMATIC_TABLE_group, font=("Arial", 8), text="DH-a")
        aDHparamLab.place(x=810 - 600, y=20 - 20)

        self.J1ΘEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J1ΘEntryField.place(x=630 - 600, y=45 - 20)
        self.J2ΘEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J2ΘEntryField.place(x=630 - 600, y=70 - 20)
        self.J3ΘEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J3ΘEntryField.place(x=630 - 600, y=95 - 20)
        self.J4ΘEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J4ΘEntryField.place(x=630 - 600, y=120 - 20)
        self.J5ΘEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J5ΘEntryField.place(x=630 - 600, y=145 - 20)
        self.J6ΘEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J6ΘEntryField.place(x=630 - 600, y=170 - 20)

        self.J1αEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J1αEntryField.place(x=685 - 600, y=45 - 20)
        self.J2αEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J2αEntryField.place(x=685 - 600, y=70 - 20)
        self.J3αEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J3αEntryField.place(x=685 - 600, y=95 - 20)
        self.J4αEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J4αEntryField.place(x=685 - 600, y=120 - 20)
        self.J5αEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J5αEntryField.place(x=685 - 600, y=145 - 20)
        self.J6αEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J6αEntryField.place(x=685 - 600, y=170 - 20)

        self.J1dEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J1dEntryField.place(x=740 - 600, y=45 - 20)
        self.J2dEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J2dEntryField.place(x=740 - 600, y=70 - 20)
        self.J3dEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J3dEntryField.place(x=740 - 600, y=95 - 20)
        self.J4dEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J4dEntryField.place(x=740 - 600, y=120 - 20)
        self.J5dEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J5dEntryField.place(x=740 - 600, y=145 - 20)
        self.J6dEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J6dEntryField.place(x=740 - 600, y=170 - 20)

        self.J1aEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J1aEntryField.place(x=795 - 600, y=45 - 20)
        self.J2aEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J2aEntryField.place(x=795 - 600, y=70 - 20)
        self.J3aEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J3aEntryField.place(x=795 - 600, y=95 - 20)
        self.J4aEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J4aEntryField.place(x=795 - 600, y=120 - 20)
        self.J5aEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J5aEntryField.place(x=795 - 600, y=145 - 20)
        self.J6aEntryField = Entry(self.KINEMATIC_TABLE_group, width=8, justify="center")
        self.J6aEntryField.place(x=795 - 600, y=170 - 20)

        saveCalBut = Button(self.tab1_rb, text="SAVE", width=26, command=self.SaveAndApplyCalibration)
        saveCalBut.place(x=750, y=400)

    def robot_tab_2_widget(self):
        self.IO_TABLE_group = tk.LabelFrame(self.tab2_rb, text="IO TABLE", labelanchor="nw", width=350, height=500)
        self.IO_TABLE_group.propagate(False)
        self.IO_TABLE_group.place(x=20, y=0)
        #
        servo0onequalsLab = Label(self.IO_TABLE_group, text="=")
        servo0onequalsLab.place(x=70, y=12)

        servo0offequalsLab = Label(self.IO_TABLE_group, text="=")
        servo0offequalsLab.place(x=70, y=52)

        servo1onequalsLab = Label(self.IO_TABLE_group, text="=")
        servo1onequalsLab.place(x=70, y=92)

        servo1offequalsLab = Label(self.IO_TABLE_group, text="=")
        servo1offequalsLab.place(x=70, y=132)

        servo2onequalsLab = Label(self.IO_TABLE_group, text="=")
        servo2onequalsLab.place(x=70, y=172)

        servo2offequalsLab = Label(self.IO_TABLE_group, text="=")
        servo2offequalsLab.place(x=70, y=212)

        servo3onequalsLab = Label(self.IO_TABLE_group, text="=")
        servo3onequalsLab.place(x=70, y=252)

        servo3offequalsLab = Label(self.IO_TABLE_group, text="=")
        servo3offequalsLab.place(x=70, y=292)

        Do1onequalsLab = Label(self.IO_TABLE_group, text="=")
        Do1onequalsLab.place(x=210, y=12)

        Do1offequalsLab = Label(self.IO_TABLE_group, text="=")
        Do1offequalsLab.place(x=210, y=52)

        Do2onequalsLab = Label(self.IO_TABLE_group, text="=")
        Do2onequalsLab.place(x=210, y=92)

        Do2offequalsLab = Label(self.IO_TABLE_group, text="=")
        Do2offequalsLab.place(x=210, y=132)

        Do3onequalsLab = Label(self.IO_TABLE_group, text="=")
        Do3onequalsLab.place(x=210, y=172)

        Do3offequalsLab = Label(self.IO_TABLE_group, text="=")
        Do3offequalsLab.place(x=210, y=212)

        Do4onequalsLab = Label(self.IO_TABLE_group, text="=")
        Do4onequalsLab.place(x=210, y=252)

        Do4offequalsLab = Label(self.IO_TABLE_group, text="=")
        Do4offequalsLab.place(x=210, y=292)

        Do5onequalsLab = Label(self.IO_TABLE_group, text="=")
        Do5onequalsLab.place(x=210, y=332)

        Do5offequalsLab = Label(self.IO_TABLE_group, text="=")
        Do5offequalsLab.place(x=210, y=372)

        Do6onequalsLab = Label(self.IO_TABLE_group, text="=")
        Do6onequalsLab.place(x=210, y=412)

        Do6offequalsLab = Label(self.IO_TABLE_group, text="=")
        Do6offequalsLab.place(x=210, y=452)

        inoutavailLab = Label(self.IO_TABLE_group, text="NOTE: the following are available when using the default Nano board for IO:   Inputs = 2-7  /  Outputs = 8-13  /  Servos = A0-A7")
        inoutavailLab.place(x=10, y=640)

        inoutavailLab = Label(self.IO_TABLE_group, text="If using IO on Teensy board:  Inputs = 32-36  /  Outputs = 37-41 - if using IO on Teensy you must manually change the command from 'Out On =' to 'ToutOn ='")
        inoutavailLab.place(x=10, y=655)

        ### 4 BUTTONS################################################################
        #############################################################################

        servo0onBut = Button(self.IO_TABLE_group, text="Servo 0", command=self.Servo0on)
        servo0onBut.place(x=10, y=10)

        servo0offBut = Button(self.IO_TABLE_group, text="Servo 0", command=self.Servo0off)
        servo0offBut.place(x=10, y=50)

        servo1onBut = Button(self.IO_TABLE_group, text="Servo 1", command=self.Servo1on)
        servo1onBut.place(x=10, y=90)

        servo1offBut = Button(self.IO_TABLE_group, text="Servo 1", command=self.Servo1off)
        servo1offBut.place(x=10, y=130)

        servo2onBut = Button(self.IO_TABLE_group, text="Servo 2", command=self.Servo2on)
        servo2onBut.place(x=10, y=170)

        servo2offBut = Button(self.IO_TABLE_group, text="Servo 2", command=self.Servo2off)
        servo2offBut.place(x=10, y=210)

        servo3onBut = Button(self.IO_TABLE_group, text="Servo 3", command=self.Servo3on)
        servo3onBut.place(x=10, y=250)

        servo3offBut = Button(self.IO_TABLE_group, text="Servo 3", command=self.Servo3off)
        servo3offBut.place(x=10, y=290)

        DO1onBut = Button(self.IO_TABLE_group, text="DO on", command=self.DO1on)
        DO1onBut.place(x=150, y=10)

        DO1offBut = Button(self.IO_TABLE_group, text="DO off", command=self.DO1off)
        DO1offBut.place(x=150, y=50)

        DO2onBut = Button(self.IO_TABLE_group, text="DO on", command=self.DO2on)
        DO2onBut.place(x=150, y=90)

        DO2offBut = Button(self.IO_TABLE_group, text="DO off", command=self.DO2off)
        DO2offBut.place(x=150, y=130)

        DO3onBut = Button(self.IO_TABLE_group, text="DO on", command=self.DO3on)
        DO3onBut.place(x=150, y=170)

        DO3offBut = Button(self.IO_TABLE_group, text="DO off", command=self.DO3off)
        DO3offBut.place(x=150, y=210)

        DO4onBut = Button(self.IO_TABLE_group, text="DO on", command=self.DO4on)
        DO4onBut.place(x=150, y=250)

        DO4offBut = Button(self.IO_TABLE_group, text="DO off", command=self.DO4off)
        DO4offBut.place(x=150, y=290)

        DO5onBut = Button(self.IO_TABLE_group, text="DO on", command=self.DO5on)
        DO5onBut.place(x=150, y=330)

        DO5offBut = Button(self.IO_TABLE_group, text="DO off", command=self.DO5off)
        DO5offBut.place(x=150, y=370)

        DO6onBut = Button(self.IO_TABLE_group, text="DO on", command=self.DO6on)
        DO6onBut.place(x=150, y=410)

        DO6offBut = Button(self.IO_TABLE_group, text="DO off", command=self.DO6off)
        DO6offBut.place(x=150, y=450)

        #### 4 ENTRY FIELDS##########################################################
        #############################################################################

        self.servo0onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo0onEntryField.place(x=90, y=15)

        self.servo0offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo0offEntryField.place(x=90, y=55)

        self.servo1onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo1onEntryField.place(x=90, y=95)

        self.servo1offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo1offEntryField.place(x=90, y=135)

        self.servo2onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo2onEntryField.place(x=90, y=175)

        self.servo2offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo2offEntryField.place(x=90, y=215)

        self.servo3onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo3onEntryField.place(x=90, y=255)

        self.servo3offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.servo3offEntryField.place(x=90, y=295)

        self.DO1onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO1onEntryField.place(x=230, y=15)

        self.DO1offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO1offEntryField.place(x=230, y=55)

        self.DO2onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO2onEntryField.place(x=230, y=95)

        self.DO2offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO2offEntryField.place(x=230, y=135)

        self.DO3onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO3onEntryField.place(x=230, y=175)

        self.DO3offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO3offEntryField.place(x=230, y=215)

        self.DO4onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO4onEntryField.place(x=230, y=255)

        self.DO4offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO4offEntryField.place(x=230, y=295)

        self.DO5onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO5onEntryField.place(x=230, y=335)

        self.DO5offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO5offEntryField.place(x=230, y=375)

        self.DO6onEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO6onEntryField.place(x=230, y=415)

        self.DO6offEntryField = Entry(self.IO_TABLE_group, width=5, justify="center")
        self.DO6offEntryField.place(x=230, y=455)

    def Servo0on(self):
        self.savePosData()
        servoPos = self.servo0onEntryField.get()
        command = "SV0P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo0off(self):
        self.savePosData()
        servoPos = self.servo0offEntryField.get()
        command = "SV0P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo1on(self):
        self.savePosData()
        servoPos = self.servo1onEntryField.get()
        command = "SV1P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo1off(self):
        self.savePosData()
        servoPos = self.servo1offEntryField.get()
        command = "SV1P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo2on(self):
        self.savePosData()
        servoPos = self.servo2onEntryField.get()
        command = "SV2P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo2off(self):
        self.savePosData()
        servoPos = self.servo2offEntryField.get()
        command = "SV2P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo3on(self):
        self.savePosData()
        servoPos = self.servo3onEntryField.get()
        command = "SV3P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def Servo3off(self):
        self.savePosData()
        servoPos = self.servo3offEntryField.get()
        command = "SV3P" + servoPos + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO1on(self):
        outputNum = self.DO1onEntryField.get()
        command = "ONX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO1off(self):
        outputNum = self.DO1offEntryField.get()
        command = "OFX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO2on(self):
        outputNum = self.DO2onEntryField.get()
        command = "ONX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO2off(self):
        outputNum = self.DO2offEntryField.get()
        command = "OFX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO3on(self):
        outputNum = self.DO3onEntryField.get()
        command = "ONX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO3off(self):
        outputNum = self.DO3offEntryField.get()
        command = "OFX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO4on(self):
        outputNum = self.DO4onEntryField.get()
        command = "ONX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO4off(self):
        outputNum = self.DO4offEntryField.get()
        command = "OFX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO5on(self):
        outputNum = self.DO5onEntryField.get()
        command = "ONX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO5off(self):
        outputNum = self.DO5offEntryField.get()
        command = "OFX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO6on(self):
        outputNum = self.DO6onEntryField.get()
        command = "ONX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def DO6off(self):
        outputNum = self.DO6offEntryField.get()
        command = "OFX" + outputNum + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        self.ser.read()

    def robot_tab_3_widget(self):
        self.REGISTER_TABLE_group = tk.LabelFrame(self.tab3_rb, text="REGISTER TABLE", labelanchor="nw", width=700, height=600)
        self.REGISTER_TABLE_group.propagate(False)
        self.REGISTER_TABLE_group.place(x=20, y=0)
        #
        R1Lab = Label(self.REGISTER_TABLE_group, text="R1")
        R1Lab.place(x=70, y=30)
        R2Lab = Label(self.REGISTER_TABLE_group, text="R2")
        R2Lab.place(x=70, y=60)
        R3Lab = Label(self.REGISTER_TABLE_group, text="R3")
        R3Lab.place(x=70, y=90)
        R4Lab = Label(self.REGISTER_TABLE_group, text="R4")
        R4Lab.place(x=70, y=120)
        R5Lab = Label(self.REGISTER_TABLE_group, text="R5")
        R5Lab.place(x=70, y=150)
        R6Lab = Label(self.REGISTER_TABLE_group, text="R6")
        R6Lab.place(x=70, y=180)
        R7Lab = Label(self.REGISTER_TABLE_group, text="R7")
        R7Lab.place(x=70, y=210)
        R8Lab = Label(self.REGISTER_TABLE_group, text="R8")
        R8Lab.place(x=70, y=240)
        R9Lab = Label(self.REGISTER_TABLE_group, text="R9")
        R9Lab.place(x=70, y=270)
        R10Lab = Label(self.REGISTER_TABLE_group, text="R10")
        R10Lab.place(x=70, y=300)
        R11Lab = Label(self.REGISTER_TABLE_group, text="R11")
        R11Lab.place(x=70, y=330)
        R12Lab = Label(self.REGISTER_TABLE_group, text="R12")
        R12Lab.place(x=70, y=360)
        R13Lab = Label(self.REGISTER_TABLE_group, text="R14")
        R13Lab.place(x=70, y=390)
        R14Lab = Label(self.REGISTER_TABLE_group, text="R14")
        R14Lab.place(x=70, y=420)
        R15Lab = Label(self.REGISTER_TABLE_group, text="R15")
        R15Lab.place(x=70, y=450)
        R16Lab = Label(self.REGISTER_TABLE_group, text="R16")
        R16Lab.place(x=70, y=480)
        SP1Lab = Label(self.REGISTER_TABLE_group, text="PR1")
        SP1Lab.place(x=640, y=30)
        SP2Lab = Label(self.REGISTER_TABLE_group, text="PR2")
        SP2Lab.place(x=640, y=60)
        SP3Lab = Label(self.REGISTER_TABLE_group, text="PR3")
        SP3Lab.place(x=640, y=90)
        SP4Lab = Label(self.REGISTER_TABLE_group, text="PR4")
        SP4Lab.place(x=640, y=120)
        SP5Lab = Label(self.REGISTER_TABLE_group, text="PR5")
        SP5Lab.place(x=640, y=150)
        SP6Lab = Label(self.REGISTER_TABLE_group, text="PR6")
        SP6Lab.place(x=640, y=180)
        SP7Lab = Label(self.REGISTER_TABLE_group, text="PR7")
        SP7Lab.place(x=640, y=210)
        SP8Lab = Label(self.REGISTER_TABLE_group, text="PR8")
        SP8Lab.place(x=640, y=240)
        SP9Lab = Label(self.REGISTER_TABLE_group, text="PR9")
        SP9Lab.place(x=640, y=270)
        SP10Lab = Label(self.REGISTER_TABLE_group, text="PR10")
        SP10Lab.place(x=640, y=300)
        SP11Lab = Label(self.REGISTER_TABLE_group, text="PR11")
        SP11Lab.place(x=640, y=330)
        SP12Lab = Label(self.REGISTER_TABLE_group, text="PR12")
        SP12Lab.place(x=640, y=360)
        SP13Lab = Label(self.REGISTER_TABLE_group, text="PR14")
        SP13Lab.place(x=640, y=390)
        SP14Lab = Label(self.REGISTER_TABLE_group, text="PR14")
        SP14Lab.place(x=640, y=420)
        SP15Lab = Label(self.REGISTER_TABLE_group, text="PR15")
        SP15Lab.place(x=640, y=450)
        SP16Lab = Label(self.REGISTER_TABLE_group, text="PR16")
        SP16Lab.place(x=640, y=480)
        SP_E1_Lab = Label(self.REGISTER_TABLE_group, text="X")
        SP_E1_Lab.place(x=410, y=10)
        SP_E2_Lab = Label(self.REGISTER_TABLE_group, text="Y")
        SP_E2_Lab.place(x=450, y=10)
        SP_E3_Lab = Label(self.REGISTER_TABLE_group, text="Z")
        SP_E3_Lab.place(x=490, y=10)
        SP_E4_Lab = Label(self.REGISTER_TABLE_group, text="Rz")
        SP_E4_Lab.place(x=530, y=10)
        SP_E5_Lab = Label(self.REGISTER_TABLE_group, text="Ry")
        SP_E5_Lab.place(x=570, y=10)
        SP_E6_Lab = Label(self.REGISTER_TABLE_group, text="Rx")
        SP_E6_Lab.place(x=610, y=10)

        #############################################################################

        self.R1EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R1EntryField.place(x=30, y=30)

        self.R2EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R2EntryField.place(x=30, y=60)

        self.R3EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R3EntryField.place(x=30, y=90)

        self.R4EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R4EntryField.place(x=30, y=120)

        self.R5EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R5EntryField.place(x=30, y=150)

        self.R6EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R6EntryField.place(x=30, y=180)

        self.R7EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R7EntryField.place(x=30, y=210)

        self.R8EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R8EntryField.place(x=30, y=240)

        self.R9EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R9EntryField.place(x=30, y=270)

        self.R10EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R10EntryField.place(x=30, y=300)

        self.R11EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R11EntryField.place(x=30, y=330)

        self.R12EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R12EntryField.place(x=30, y=360)

        self.R13EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R13EntryField.place(x=30, y=390)

        self.R14EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R14EntryField.place(x=30, y=420)

        self.R15EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R15EntryField.place(x=30, y=450)

        self.R16EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.R16EntryField.place(x=30, y=480)

        self.SP_1_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_1_E1_EntryField.place(x=400, y=30)

        self.SP_2_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_2_E1_EntryField.place(x=400, y=60)

        self.SP_3_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_3_E1_EntryField.place(x=400, y=90)

        self.SP_4_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_4_E1_EntryField.place(x=400, y=120)

        self.SP_5_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_5_E1_EntryField.place(x=400, y=150)

        self.SP_6_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_6_E1_EntryField.place(x=400, y=180)

        self.SP_7_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_7_E1_EntryField.place(x=400, y=210)

        self.SP_8_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_8_E1_EntryField.place(x=400, y=240)

        self.SP_9_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_9_E1_EntryField.place(x=400, y=270)

        self.SP_10_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_10_E1_EntryField.place(x=400, y=300)

        self.SP_11_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_11_E1_EntryField.place(x=400, y=330)

        self.SP_12_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_12_E1_EntryField.place(x=400, y=360)

        self.SP_13_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_13_E1_EntryField.place(x=400, y=390)

        self.SP_14_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_14_E1_EntryField.place(x=400, y=420)

        self.SP_15_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_15_E1_EntryField.place(x=400, y=450)

        self.SP_16_E1_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_16_E1_EntryField.place(x=400, y=480)

        self.SP_1_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_1_E2_EntryField.place(x=440, y=30)

        self.SP_2_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_2_E2_EntryField.place(x=440, y=60)

        self.SP_3_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_3_E2_EntryField.place(x=440, y=90)

        self.SP_4_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_4_E2_EntryField.place(x=440, y=120)

        self.SP_5_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_5_E2_EntryField.place(x=440, y=150)

        self.SP_6_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_6_E2_EntryField.place(x=440, y=180)

        self.SP_7_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_7_E2_EntryField.place(x=440, y=210)

        self.SP_8_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_8_E2_EntryField.place(x=440, y=240)

        self.SP_9_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_9_E2_EntryField.place(x=440, y=270)

        self.SP_10_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_10_E2_EntryField.place(x=440, y=300)

        self.SP_11_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_11_E2_EntryField.place(x=440, y=330)

        self.SP_12_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_12_E2_EntryField.place(x=440, y=360)

        self.SP_13_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_13_E2_EntryField.place(x=440, y=390)

        self.SP_14_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_14_E2_EntryField.place(x=440, y=420)

        self.SP_15_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_15_E2_EntryField.place(x=440, y=450)

        self.SP_16_E2_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_16_E2_EntryField.place(x=440, y=480)

        self.SP_1_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_1_E3_EntryField.place(x=480, y=30)

        self.SP_2_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_2_E3_EntryField.place(x=480, y=60)

        self.SP_3_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_3_E3_EntryField.place(x=480, y=90)

        self.SP_4_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_4_E3_EntryField.place(x=480, y=120)

        self.SP_5_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_5_E3_EntryField.place(x=480, y=150)

        self.SP_6_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_6_E3_EntryField.place(x=480, y=180)

        self.SP_7_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_7_E3_EntryField.place(x=480, y=210)

        self.SP_8_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_8_E3_EntryField.place(x=480, y=240)

        self.SP_9_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_9_E3_EntryField.place(x=480, y=270)

        self.SP_10_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_10_E3_EntryField.place(x=480, y=300)

        self.SP_11_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_11_E3_EntryField.place(x=480, y=330)

        self.SP_12_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_12_E3_EntryField.place(x=480, y=360)

        self.SP_13_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_13_E3_EntryField.place(x=480, y=390)

        self.SP_14_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_14_E3_EntryField.place(x=480, y=420)

        self.SP_15_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_15_E3_EntryField.place(x=480, y=450)

        self.SP_16_E3_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_16_E3_EntryField.place(x=480, y=480)

        self.SP_1_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_1_E4_EntryField.place(x=520, y=30)

        self.SP_2_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_2_E4_EntryField.place(x=520, y=60)

        self.SP_3_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_3_E4_EntryField.place(x=520, y=90)

        self.SP_4_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_4_E4_EntryField.place(x=520, y=120)

        self.SP_5_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_5_E4_EntryField.place(x=520, y=150)

        self.SP_6_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_6_E4_EntryField.place(x=520, y=180)

        self.SP_7_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_7_E4_EntryField.place(x=520, y=210)

        self.SP_8_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_8_E4_EntryField.place(x=520, y=240)

        self.SP_9_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_9_E4_EntryField.place(x=520, y=270)

        self.SP_10_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_10_E4_EntryField.place(x=520, y=300)

        self.SP_11_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_11_E4_EntryField.place(x=520, y=330)

        self.SP_12_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_12_E4_EntryField.place(x=520, y=360)

        self.SP_13_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_13_E4_EntryField.place(x=520, y=390)

        self.SP_14_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_14_E4_EntryField.place(x=520, y=420)

        self.SP_15_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_15_E4_EntryField.place(x=520, y=450)

        self.SP_16_E4_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_16_E4_EntryField.place(x=520, y=480)

        self.SP_1_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_1_E5_EntryField.place(x=560, y=30)

        self.SP_2_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_2_E5_EntryField.place(x=560, y=60)

        self.SP_3_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_3_E5_EntryField.place(x=560, y=90)

        self.SP_4_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_4_E5_EntryField.place(x=560, y=120)

        self.SP_5_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_5_E5_EntryField.place(x=560, y=150)

        self.SP_6_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_6_E5_EntryField.place(x=560, y=180)

        self.SP_7_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_7_E5_EntryField.place(x=560, y=210)

        self.SP_8_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_8_E5_EntryField.place(x=560, y=240)

        self.SP_9_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_9_E5_EntryField.place(x=560, y=270)

        self.SP_10_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_10_E5_EntryField.place(x=560, y=300)

        self.SP_11_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_11_E5_EntryField.place(x=560, y=330)

        self.SP_12_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_12_E5_EntryField.place(x=560, y=360)

        self.SP_13_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_13_E5_EntryField.place(x=560, y=390)

        self.SP_14_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_14_E5_EntryField.place(x=560, y=420)

        self.SP_15_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_15_E5_EntryField.place(x=560, y=450)

        self.SP_16_E5_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_16_E5_EntryField.place(x=560, y=480)

        self.SP_1_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_1_E6_EntryField.place(x=600, y=30)

        self.SP_2_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_2_E6_EntryField.place(x=600, y=60)

        self.SP_3_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_3_E6_EntryField.place(x=600, y=90)

        self.SP_4_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_4_E6_EntryField.place(x=600, y=120)

        self.SP_5_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_5_E6_EntryField.place(x=600, y=150)

        self.SP_6_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_6_E6_EntryField.place(x=600, y=180)

        self.SP_7_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_7_E6_EntryField.place(x=600, y=210)

        self.SP_8_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_8_E6_EntryField.place(x=600, y=240)

        self.SP_9_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_9_E6_EntryField.place(x=600, y=270)

        self.SP_10_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_10_E6_EntryField.place(x=600, y=300)

        self.SP_11_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_11_E6_EntryField.place(x=600, y=330)

        self.SP_12_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_12_E6_EntryField.place(x=600, y=360)

        self.SP_13_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_13_E6_EntryField.place(x=600, y=390)

        self.SP_14_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_14_E6_EntryField.place(x=600, y=420)

        self.SP_15_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_15_E6_EntryField.place(x=600, y=450)

        self.SP_16_E6_EntryField = Entry(self.REGISTER_TABLE_group, width=5, justify="center")
        self.SP_16_E6_EntryField.place(x=600, y=480)

    def robot_tab_4_widget(self):
        self.GcodeProgEntryField = Entry(self.tab4_rb, width=60, justify="center")
        self.GcodeProgEntryField.place(x=20, y=55)

        self.GcodCurRowEntryField = Entry(self.tab4_rb, width=10, justify="center")
        self.GcodCurRowEntryField.place(x=1175, y=20)

        self.GC_ST_E1_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_ST_E1_EntryField.place(x=20, y=140)

        self.GC_ST_E2_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_ST_E2_EntryField.place(x=75, y=140)

        self.GC_ST_E3_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_ST_E3_EntryField.place(x=130, y=140)

        self.GC_ST_E4_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_ST_E4_EntryField.place(x=185, y=140)

        self.GC_ST_E5_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_ST_E5_EntryField.place(x=240, y=140)

        self.GC_ST_E6_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_ST_E6_EntryField.place(x=295, y=140)

        self.GC_ST_WC_EntryField = Entry(self.tab4_rb, width=3, justify="center")
        self.GC_ST_WC_EntryField.place(x=350, y=140)

        self.GC_SToff_E1_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_SToff_E1_EntryField.place(x=20, y=205)

        self.GC_SToff_E2_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_SToff_E2_EntryField.place(x=75, y=205)

        self.GC_SToff_E3_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_SToff_E3_EntryField.place(x=130, y=205)

        self.GC_SToff_E4_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_SToff_E4_EntryField.place(x=185, y=205)

        self.GC_SToff_E5_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_SToff_E5_EntryField.place(x=240, y=205)

        self.GC_SToff_E6_EntryField = Entry(self.tab4_rb, width=8, justify="center")
        self.GC_SToff_E6_EntryField.place(x=295, y=205)

        self.GcodeFilenameField = Entry(self.tab4_rb, width=40, justify="center")
        self.GcodeFilenameField.place(x=20, y=340)

        self.GCalmStatusLab = Label(self.tab4_rb, text="GCODE IDLE", style="OK.TLabel")
        self.GCalmStatusLab.place(x=400, y=20)

        gcodeframe = Frame(self.tab4_rb)
        gcodeframe.place(x=400, y=53)
        self.gcodescrollbar = Scrollbar(gcodeframe)
        self.gcodescrollbar.pack(side=RIGHT, fill=Y)
        self.tab4_rb.gcodeView = Listbox(gcodeframe, exportselection=0, width=90, height=35, yscrollcommand=self.gcodescrollbar.set)
        self.tab4_rb.gcodeView.bind('<<ListboxSelect>>', self.gcodeViewselect)
        self.tab4_rb.gcodeView.pack()
        self.gcodescrollbar.config(command=self.tab4_rb.gcodeView.yview)

        self.tab4_rb.gcodeView.bind("<<ListboxSelect>>", self.GCcallback)

        LoadGcodeBut = Button(self.tab4_rb, text="Load Program", width=25, command=self.loadGcodeProg)
        LoadGcodeBut.place(x=20, y=20)

        GcodeStartPosBut = Button(self.tab4_rb, text="Set Start Position", width=25, command=self.SetGcodeStartPos)
        GcodeStartPosBut.place(x=20, y=100)

        GcodeMoveStartPosBut = Button(self.tab4_rb, text="Move to Start Offset", width=25, command=self.MoveGcodeStartPos)
        GcodeMoveStartPosBut.place(x=20, y=240)

        runGcodeBut = Button(self.tab4_rb, text="Convert & Upload to SD", width=25, command=self.GCconvertProg)
        # playGPhoto=PhotoImage(file="play-icon.gif")
        # runGcodeBut.config(image=playGPhoto)
        runGcodeBut.place(x=20, y=375)

        stopGcodeBut = Button(self.tab4_rb, text="Stop Conversion & Upload", width=25, command=self.GCstopProg)
        # stopGPhoto=PhotoImage(file="stop-icon.gif")
        # stopGcodeBut.config(image=stopGPhoto)
        stopGcodeBut.place(x=190, y=375)

        delGcodeBut = Button(self.tab4_rb, text="Delete File from SD", width=25, command=self.GCdelete)
        delGcodeBut.place(x=20, y=415)

        readGcodeBut = Button(self.tab4_rb, text="Read Files from SD", width=25, command=partial(self.GCread, "yes"))
        readGcodeBut.place(x=20, y=455)

        playGPhoto = PhotoImage(file="play-icon.gif")
        readGcodeBut = Button(self.tab4_rb, text="Play Gcode File", width=20, command=self.GCplay, image=playGPhoto, compound=LEFT)
        readGcodeBut.place(x=20, y=495)

        # revGcodeBut = Button(self.tab4_rb,  text="REV ",  command = stepRev)
        # revGcodeBut.place(x=180, y=290)

        # fwdGcodeBut = Button(self.tab4_rb,  text="FWD", command = GCstepFwd)
        # fwdGcodeBut.place(x=230, y=290)

        saveGCBut = Button(self.tab4_rb, text="SAVE DATA", width=26, command=self.SaveAndApplyCalibration)
        saveGCBut.place(x=20, y=600)

        gcodeCurRowLab = Label(self.tab4_rb, text="Current Row: ")
        gcodeCurRowLab.place(x=800, y=21)

        gcodeStartPosOFfLab = Label(self.tab4_rb, text="Start Position Offset")
        gcodeStartPosOFfLab.place(x=20, y=180)

        gcodeFilenameLab = Label(self.tab4_rb, text="Filename:")
        gcodeFilenameLab.place(x=20, y=320)

    def gcodeFrame(self):
        gcodeframe = Frame(self.tab4_rb)
        gcodeframe.place(x=300, y=10)
        # progframe.pack(side=RIGHT, fill=Y)
        scrollbar = Scrollbar(gcodeframe)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.tab4_rb.gcodeView = Listbox(gcodeframe, width=105, height=46, yscrollcommand=scrollbar.set)
        self.tab4_rb.gcodeView.bind('<<ListboxSelect>>', self.gcodeViewselect)
        time.sleep(.1)
        self.tab4_rb.gcodeView.pack()
        scrollbar.config(command=self.tab4_rb.gcodeView.yview)

    def gcodeViewselect(self, e):
        gcodeRow = self.tab4_rb.gcodeView.curselection()[0]
        self.GcodCurRowEntryField.delete(0, 'end')
        self.GcodCurRowEntryField.insert(0, gcodeRow)

    def GCcallback(self, event):
        selection = event.widget.curselection()
        try:
            if selection:
                index = selection[0]
                data = event.widget.get(index)
                data = data.replace('.txt', '')
                self.GcodeFilenameField.delete(0, 'end')
                self.GcodeFilenameField.insert(0, data)
                self.PlayGCEntryField.delete(0, 'end')
                self.PlayGCEntryField.insert(0, data)
            else:
                self.GcodeFilenameField.insert(0, "")
        except:
            print("not an SD file")

    def loadGcodeProg(self):
        filetypes = (('gcode files', '*.gcode *.nc *.ngc *.cnc *.tap'), ('text files', '*.txt'))
        filename = fd.askopenfilename(title='Open files', initialdir='/', filetypes=filetypes)
        self.GcodeProgEntryField.delete(0, 'end')
        self.GcodeProgEntryField.insert(0, filename)
        gcodeProg = open(self.GcodeProgEntryField.get(), "rb")
        self.tab4_rb.gcodeView.delete(0, END)
        previtem = ""
        for item in gcodeProg:
            try:
                commentIndex = item.find(b";")
                item = item[:commentIndex]
            except:
                pass
            item = item + b" "
            if (item != previtem):
                self.tab4_rb.gcodeView.insert(END, item)
            previtem = item
        self.tab4_rb.gcodeView.pack()
        self.gcodescrollbar.config(command=self.tab4_rb.gcodeView.yview)

    def SetGcodeStartPos(self):
        self.GC_ST_E1_EntryField.delete(0, 'end')
        self.GC_ST_E1_EntryField.insert(0, str(self.XcurPos))
        self.GC_ST_E2_EntryField.delete(0, 'end')
        self.GC_ST_E2_EntryField.insert(0, str(self.YcurPos))
        self.GC_ST_E3_EntryField.delete(0, 'end')
        self.GC_ST_E3_EntryField.insert(0, str(self.ZcurPos))
        self.GC_ST_E4_EntryField.delete(0, 'end')
        self.GC_ST_E4_EntryField.insert(0, str(self.RzcurPos))
        self.GC_ST_E5_EntryField.delete(0, 'end')
        self.GC_ST_E5_EntryField.insert(0, str(self.RycurPos))
        self.GC_ST_E6_EntryField.delete(0, 'end')
        self.GC_ST_E6_EntryField.insert(0, str(self.RxcurPos))
        self.GC_ST_WC_EntryField.delete(0, 'end')
        self.GC_ST_WC_EntryField.insert(0, str(self.WC))

    def MoveGcodeStartPos(self):
        xVal = str(float(self.GC_ST_E1_EntryField.get()) + float(self.GC_SToff_E1_EntryField.get()))
        yVal = str(float(self.GC_ST_E2_EntryField.get()) + float(self.GC_SToff_E2_EntryField.get()))
        zVal = str(float(self.GC_ST_E3_EntryField.get()) + float(self.GC_SToff_E3_EntryField.get()))
        rzVal = str(float(self.GC_ST_E4_EntryField.get()) + float(self.GC_SToff_E4_EntryField.get()))
        ryVal = str(float(self.GC_ST_E5_EntryField.get()) + float(self.GC_SToff_E5_EntryField.get()))
        rxVal = str(float(self.GC_ST_E6_EntryField.get()) + float(self.GC_SToff_E6_EntryField.get()))
        J7Val = str(self.J7PosCur)
        J8Val = str(self.J8PosCur)
        J9Val = str(self.J9PosCur)
        self.speedPrefix = "Sm"
        self.Speed = "25"
        self.ACCspd = "10"
        self.DECspd = "10"
        self.ACCramp = "100"
        self.WC = self.GC_ST_WC_EntryField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "MJ" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def GCstepFwd(self):
        self.GCalmStatusLab.config(text="GCODE READY", style="OK.TLabel")
        self.GCexecuteRow()
        GCselRow = self.tab4_rb.gcodeView.curselection()[0]
        last = self.tab4_rb.gcodeView.index('end')
        for row in range(0, GCselRow):
            self.tab4_rb.gcodeView.itemconfig(row, {'fg': 'dodger blue'})
        self.tab4_rb.gcodeView.itemconfig(GCselRow, {'fg': 'blue2'})
        for row in range(GCselRow + 1, last):
            self.tab4_rb.gcodeView.itemconfig(row, {'fg': 'black'})
        self.tab4_rb.gcodeView.selection_clear(0, END)
        GCselRow += 1
        self.tab4_rb.gcodeView.select_set(GCselRow)
        try:
            GCselRow = self.tab4_rb.gcodeView.curselection()[0]
            self.GcodCurRowEntryField.delete(0, 'end')
            self.GcodCurRowEntryField.insert(0, GCselRow)
        except:
            self.GcodCurRowEntryField.delete(0, 'end')
            self.GcodCurRowEntryField.insert(0, "---")

    def GCdelete(self):
        if (self.GcodeFilenameField.get() != ""):
            Filename = self.GcodeFilenameField.get() + ".txt"
            command = "DG" + "Fn" + Filename + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                if (response == "P"):
                    text = Filename + " has been deleted"
                    self.GCalmStatusLab.config(text=text, style="OK.TLabel")
                    status = "no"
                    self.GCread(status)
                elif (response == "F"):
                    text = Filename + " was not found"
                    self.GCalmStatusLab.config(text=text, style="Alarm.TLabel")
        else:
            messagebox.showwarning("warning", "Please Enter a Filename")

    def GCread(self, status):
        command = "RG" + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            if (status == "yes"):
                self.GCalmStatusLab.config(text="FILES FOUND ON SD CARD:", style="OK.TLabel")
            self.GcodeProgEntryField.delete(0, 'end')
            self.tab4_rb.gcodeView.delete(0, END)
            for value in response.split(","):
                self.tab4_rb.gcodeView.insert(END, value)
            self.tab4_rb.gcodeView.pack()
            self.gcodescrollbar.config(command=self.tab4_rb.gcodeView.yview)

    def GCplay(self):
        Filename = self.GcodeFilenameField.get()
        self.GCplayProg(Filename)

    def GCplayProg(self, Filename):
        self.GCalmStatusLab.config(text="GCODE FILE RUNNING", style="OK.TLabel")

        def GCthreadPlay():
            global estopActive
            Fn = Filename + ".txt"
            command = "PG" + "Fn" + Fn + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)
                if (estopActive == TRUE):
                    self.GCalmStatusLab.config(text="Estop Button was Pressed", style="Alarm.TLabel")
                else:
                    self.GCalmStatusLab.config(text="GCODE FILE COMPLETE", style="Warn.TLabel")

        GCplay = threading.Thread(target=GCthreadPlay)
        GCplay.start()

    def GCconvertProg(self):
        if (self.GcodeProgEntryField.get() == ""):
            messagebox.showwarning("warning", "Please Load a Gcode Program")
        elif (self.GcodeFilenameField.get() == ""):
            messagebox.showwarning("warning", "Please Enter a Filename")
        else:
            Filename = self.GcodeFilenameField.get() + ".txt"
            command = "DG" + "Fn" + Filename + "\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            time.sleep(.1)
            response = str(self.ser.readline().strip(), 'utf-8')
            last = self.tab4_rb.gcodeView.index('end')
            for row in range(0, last):
                self.tab4_rb.gcodeView.itemconfig(row, {'fg': 'black'})

            def GCthreadProg():
                global GCrowinproc
                global GCstopQueue
                global splineActive
                global prevxVal
                global prevyVal
                global prevzVal
                prevxVal = 0
                prevyVal = 0
                prevzVal = 0
                GCstopQueue = "0"
                splineActive = "0"
                try:
                    GCselRow = self.tab4_rb.gcodeView.curselection()[0]
                    if (GCselRow == 0):
                        GCselRow = 1
                except:
                    GCselRow = 1
                    self.tab4_rb.gcodeView.selection_clear(0, END)
                    self.tab4_rb.gcodeView.select_set(GCselRow)
                self.tab4_rb.GCrunTrue = 1
                while self.tab4_rb.GCrunTrue == 1:
                    if (self.tab4_rb.GCrunTrue == 0):
                        self.GCalmStatusLab.config(text="GCODE CONVERSION STOPPED", style="Alarm.TLabel")
                    else:
                        self.GCalmStatusLab.config(text="GCODE CONVERSION RUNNING", style="OK.TLabel")
                    GCrowinproc = 1
                    self.GCexecuteRow()
                    while GCrowinproc == 1:
                        time.sleep(.1)
                    GCselRow = self.tab4_rb.gcodeView.curselection()[0]
                    # last = self.tab4_rb.gcodeView.index('end')
                    # for row in range (0,GCselRow):
                    #  self.tab4_rb.gcodeView.itemconfig(row, {'fg': 'dodger blue'})
                    self.tab4_rb.gcodeView.itemconfig(GCselRow, {'fg': 'blue2'})
                    # for row in range (GCselRow+1,last):
                    #  self.tab4_rb.gcodeView.itemconfig(row, {'fg': 'black'})
                    self.tab4_rb.gcodeView.selection_clear(0, END)
                    GCselRow += 1
                    self.tab4_rb.gcodeView.select_set(GCselRow)
                    # gcodeRow += 1
                    # self.GcodCurRowEntryField.delete(0, 'end')
                    # self.GcodCurRowEntryField.insert(0,GCselRow)
                    # time.sleep(.1)
                    try:
                        GCselRow = self.tab4_rb.gcodeView.curselection()[0]
                        self.GcodCurRowEntryField.delete(0, 'end')
                        self.GcodCurRowEntryField.insert(0, GCselRow)
                    except:
                        self.GcodCurRowEntryField.delete(0, 'end')
                        self.GcodCurRowEntryField.insert(0, "---")
                        self.tab4_rb.GCrunTrue = 0
                        self.GCalmStatusLab.config(text="GCODE CONVERSION STOPPED", style="Alarm.TLabel")

            GCt = threading.Thread(target=GCthreadProg)
            GCt.start()

    def GCstopProg(self):
        global cmdType
        global splineActive
        global GCstopQueue
        lastProg = ""
        self.tab4_rb.GCrunTrue = 0
        self.GCalmStatusLab.config(text="GCODE CONVERSION STOPPED", style="Alarm.TLabel")
        if (splineActive == 1):
            splineActive = "0"
            if (self.stopQueue == "1"):
                self.stopQueue = "0"
                stop()
            if (self.moveInProc == 1):
                self.moveInProc == 2
            command = "SS\n"
            self.cmdSentEntryField.delete(0, 'end')
            self.cmdSentEntryField.insert(0, command)
            self.ser.write(command.encode())
            self.ser.flushInput()
            response = str(self.ser.readline().strip(), 'utf-8')
            if (response[:1] == 'E'):
                self.ErrorHandler(response)
            else:
                self.displayPosition(response)

    def GCexecuteRow(self):

        global calStat
        global GCrowinproc
        global LineDist
        global Xv
        global Yv
        global Zv
        global commandCalc

        global splineActive

        global gcodeSpeed
        global inchTrue
        global prevxVal
        global prevyVal
        global prevzVal
        global xVal
        global yVal
        global zVal
        GCstartTime = time.time()
        GCselRow = self.tab4_rb.gcodeView.curselection()[0]
        self.tab4_rb.gcodeView.see(GCselRow + 2)
        data = list(map(int, self.tab4_rb.gcodeView.curselection()))
        command = self.tab4_rb.gcodeView.get(data[0]).decode()
        cmdType = command[:1]
        subCmd = command[1:command.find(" ")].rstrip()

        ## F ##
        if (cmdType == "F"):
            gcodeSpeed = command[command.find("F") + 1:]

        ## G ##
        if (cmdType == "G"):

            # IMPERIAL
            if (subCmd == "20"):
                inchTrue = True;

                # METRIC
            if (subCmd == "21"):
                inchTrue = False;

            # ABSOLUTE / INCREMENTAL - HOME (absolute is forced and moves to start position offset)
            if (subCmd == "90" or subCmd == "91" or subCmd == "28"):

                xVal = str(float(self.GC_ST_E1_EntryField.get()) + float(self.GC_SToff_E1_EntryField.get()))
                yVal = str(float(self.GC_ST_E2_EntryField.get()) + float(self.GC_SToff_E2_EntryField.get()))
                zVal = str(float(self.GC_ST_E3_EntryField.get()) + float(self.GC_SToff_E3_EntryField.get()))
                rzVal = str(float(self.GC_ST_E4_EntryField.get()) + float(self.GC_SToff_E4_EntryField.get()))
                ryVal = str(float(self.GC_ST_E5_EntryField.get()) + float(self.GC_SToff_E5_EntryField.get()))
                rxVal = str(float(self.GC_ST_E6_EntryField.get()) + float(self.GC_SToff_E6_EntryField.get()))
                J7Val = str(self.J7PosCur)
                J8Val = str(self.J8PosCur)
                J9Val = str(self.J9PosCur)
                self.speedPrefix = "Sm"
                self.Speed = "25"
                self.ACCspd = "10"
                self.DECspd = "10"
                self.ACCramp = "100"
                self.WC = self.GC_ST_WC_EntryField.get()
                self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
                Filename = self.GcodeFilenameField.get() + ".txt"
                command = "self.WC" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "Fn" + Filename + "\n"
                self.cmdSentEntryField.delete(0, 'end')

                self.cmdSentEntryField.insert(0, command)
                self.ser.write(command.encode())
                self.ser.flushInput()
                time.sleep(.1)
                response = str(self.ser.readline().strip(), 'utf-8')
                if (response[:1] == 'E'):
                    self.ErrorHandler(response)
                    self.GCstopProg()
                    self.tab4_rb.GCrunTrue = 0
                    self.GCalmStatusLab.config(text="UNABLE TO WRITE TO SD CARD", style="Alarm.TLabel")
                else:
                    self.displayPosition(response)

                    # LINEAR MOVE
            if (subCmd == "0" or subCmd == "1"):

                if ("X" in command):
                    xtemp = command[command.find("X") + 1:]
                    xVal = xtemp[:xtemp.find(" ")]
                    xVal = str(round(float(xVal), 3))
                else:
                    xVal = ""
                if ("Y" in command):
                    ytemp = command[command.find("Y") + 1:]
                    yVal = ytemp[:ytemp.find(" ")]
                    yVal = str(round(float(yVal), 3))
                else:
                    yVal = ""
                if ("Z" in command):
                    ztemp = command[command.find("Z") + 1:]
                    zVal = ztemp[:ztemp.find(" ")]
                    zVal = str(round(float(zVal), 3))
                else:
                    zVal = ""
                if ("A" in command):
                    atemp = command[command.find("A") + 1:]
                    aVal = atemp[:atemp.find(" ")]
                    aVal = str(round(float(aVal), 3))
                else:
                    aVal = ""
                if ("B" in command):
                    btemp = command[command.find("B") + 1:]
                    bVal = btemp[:btemp.find(" ")]
                    bVal = str(round(float(bVal), 3))
                else:
                    bVal = ""
                if ("C" in command):
                    ctemp = command[command.find("C") + 1:]
                    cVal = ctemp[:ctemp.find(" ")]
                    cVal = str(round(float(cVal), 3))
                else:
                    cVal = ""
                if ("E" in command):
                    etemp = command[command.find("E") + 1:]
                    eVal = etemp[:etemp.find(" ")]
                    eVal = str(round(float(eVal), 3))
                else:
                    eVal = ""
                if ("F" in command):
                    ftemp = command[command.find("F") + 1:]
                    fVal = ftemp[:ftemp.find(" ")]
                    fVal = str(round(float(fVal), 3))
                else:
                    fVal = ""

                if (xVal != ""):
                    if (inchTrue == True):
                        xVal = str(float(xVal) * 25.4)
                    xVal = str(round((float(self.GC_ST_E1_EntryField.get()) + float(xVal)), 3))
                else:
                    try:
                        if (prevxVal != 0):
                            xVal = prevxVal
                        else:
                            xVal = str(self.XcurPos)
                    except:
                        xVal = str(self.XcurPos)

                if (yVal != ""):
                    if (inchTrue == True):
                        yVal = str(float(yVal) * 25.4)
                    yVal = str(round((float(self.GC_ST_E2_EntryField.get()) + float(yVal)), 3))
                else:
                    try:
                        if (prevyVal != 0):
                            yVal = prevyVal
                        else:
                            yVal = str(self.YcurPos)
                    except:
                        yVal = str(self.YcurPos)

                if (zVal != ""):
                    if (inchTrue == True):
                        zVal = str(float(zVal) * 25.4)
                    zVal = str(round((float(self.GC_ST_E3_EntryField.get()) + float(zVal)), 3))
                else:
                    try:
                        if (prevzVal != 0):
                            zVal = prevzVal
                        else:
                            zVal = str(self.ZcurPos)
                    except:
                        zVal = str(self.ZcurPos)

                if (aVal != ""):
                    rzVal = str(float(self.GC_ST_E4_EntryField.get()) + float(aVal))
                    if (np.sign(float(rzVal)) != np.sign(float(self.RzcurPos))):
                        rzVal = str(round((float(rzVal) * -1), 3))
                else:
                    rzVal = str(self.RzcurPos)

                if (bVal != ""):
                    ryVal = str(round((float(self.GC_ST_E5_EntryField.get()) + float(bVal))), 3)
                else:
                    ryVal = str(self.RycurPos)

                if (cVal != ""):
                    rxVal = str(round((float(self.GC_ST_E6_EntryField.get()) + float(cVal)), 3))
                else:
                    rxVal = str(self.RxcurPos)

                if (eVal != ""):
                    J7Val = eVal
                else:
                    J7Val = str(self.J7PosCur)

                J8Val = str(self.J8PosCur)
                J9Val = str(self.J9PosCur)

                if (fVal != ""):
                    if (inchTrue == True):
                        gcodeSpeed = str(round((float(fVal) / 25.4), 2))
                    else:
                        gcodeSpeed = str(round((float(fVal) / 60), 2))
                self.speedPrefix = "Sm"
                self.Speed = gcodeSpeed

                if (subCmd == "0"):
                    self.Speed = self.speedEntryField.get()

                # FORCE ROTATIONS TO BASE VALUE FOR NOW
                rzVal = self.GC_ST_E4_EntryField.get()
                ryVal = self.GC_ST_E5_EntryField.get()
                rxVal = self.GC_ST_E6_EntryField.get()

                # self.ACCspd = self.ACCspeedField.get()
                # self.DECspd = self.DECspeedField.get()
                # self.ACCramp = self.ACCrampField.get()

                self.ACCspd = ".1"
                self.DECspd = ".1"
                self.ACCramp = "100"

                Rounding = "0"
                self.WC = self.GC_ST_WC_EntryField.get()
                # self.LoopMode = str(self.J1OpenLoopStat.get())+str(self.J2OpenLoopStat.get())+str(self.J3OpenLoopStat.get())+str(self.J4OpenLoopStat.get())+str(self.J5OpenLoopStat.get())+str(self.J6OpenLoopStat.get())
                self.LoopMode = "111111"
                # DisWrist = str(self.DisableWristRot.get())
                Filename = self.GcodeFilenameField.get() + ".txt"

                command = "self.WC" + "X" + xVal + "Y" + yVal + "Z" + zVal + "Rz" + rzVal + "Ry" + ryVal + "Rx" + rxVal + "J7" + J7Val + "J8" + J8Val + "J9" + J9Val + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "Rnd" + Rounding + "W" + self.WC + "Lm" + self.LoopMode + "Fn" + Filename + "\n"
                prevxVal = xVal
                prevyVal = yVal
                prevzVal = zVal
                self.cmdSentEntryField.delete(0, 'end')
                self.cmdSentEntryField.insert(0, command)

                # self.tab5_rb.ElogView.insert(END, command)
                # value=self.tab5_rb.ElogView.get(0,END)
                # pickle.dump(value,open("ErrorLog","wb"))

                self.ser.write(command.encode())
                self.ser.flushInput()
                time.sleep(.05)
                # self.ser.read()
                response = str(self.ser.readline().strip(), 'utf-8')
                if (response[:1] == 'E'):
                    self.tab4_rb.GCrunTrue = 0
                    self.GCalmStatusLab.config(text="UNABLE TO WRITE TO SD CARD", style="Alarm.TLabel")
                    self.ErrorHandler(response)
                else:
                    self.displayPosition(response)

        GCrowinproc = 0

    ##########################################CAMERA VISION ROBOT###########################################################
    def robot_Cam_Search_widget(self):

        VisBackdropImg = ImageTk.PhotoImage(Image.open('VisBackdrop.png'))
        VisBackdromLbl = Label(self.tab9, image=VisBackdropImg)
        VisBackdromLbl.place(x=15, y=215)

        # self.cap= cv2.VideoCapture(0)
        video_frame = Frame(self.tab9, width=640, height=480)
        video_frame.place(x=50, y=250)

        self.vid_lbl = Label(video_frame)
        self.vid_lbl.place(x=0, y=0)

        self.vid_lbl.bind('<Button-1>', self.motion_search)

        LiveLab = Label(self.tab9, text="LIVE VIDEO FEED")
        LiveLab.place(x=750, y=390)
        liveCanvas = Canvas(self.tab9, width=490, height=330)
        liveCanvas.place(x=750, y=410)
        live_frame = Frame(self.tab9, width=480, height=320)
        live_frame.place(x=757, y=417)
        self.live_lbl = Label(live_frame)
        self.live_lbl.place(x=0, y=0)

        template_frame = Frame(self.tab9, width=150, height=150)
        template_frame.place(x=575, y=50)

        self.template_lbl = Label(template_frame)
        self.template_lbl.place(x=0, y=0)

        FoundValuesLab = Label(self.tab9, text="FOUND VALUES")
        FoundValuesLab.place(x=750, y=30)

        CalValuesLab = Label(self.tab9, text="CALIBRATION VALUES")
        CalValuesLab.place(x=900, y=30)

        ### 6 BUTTONS################################################################
        #############################################################################

        graph = FilterGraph()
        try:
            self.camList = graph.get_input_devices()
        except:
            self.camList = ["Select a Camera"]
        self.Jvisoptions = StringVar(self.tab9)
        self.Jvisoptions.set("Select a Camera")
        try:
            vismenu = OptionMenu(self.tab9, self.Jvisoptions, self.camList[0], *self.camList)
            vismenu.config(width=20)
            vismenu.place(x=10, y=10)
        except:
            print("no camera")

        StartCamBut = Button(self.tab9, text="Start Camera", width=15, command=self.start_vid_search)
        StartCamBut.place(x=200, y=10)

        StopCamBut = Button(self.tab9, text="Stop Camera", width=15, command=self.stop_vid_search)
        StopCamBut.place(x=315, y=10)

        CapImgBut = Button(self.tab9, text="Snap Image", width=15, command=self.take_pic_search)
        CapImgBut.place(x=10, y=50)

        TeachImgBut = Button(self.tab9, text="Teach Object", width=15, command=self.selectTemplate_search)
        TeachImgBut.place(x=140, y=50)

        FindVisBut = Button(self.tab9, text="Snap & Find", width=15, command=self.snapFind_search)
        FindVisBut.place(x=270, y=50)

        ZeroBrCnBut = Button(self.tab9, text="Zero", width=5, command=self.zeroBrCn_search)
        ZeroBrCnBut.place(x=10, y=110)

        maskBut = Button(self.tab9, text="Mask", width=5, command=self.selectMask_search)
        maskBut.place(x=10, y=150)

        self.VisZoomSlide = Scale(self.tab9, from_=50, to=1, length=250, orient=HORIZONTAL)
        self.VisZoomSlide.bind("<ButtonRelease-1>", self.VisUpdateBriCon)
        self.VisZoomSlide.place(x=75, y=95)
        self.VisZoomSlide.set(50)

        VisZoomLab = Label(self.tab9, text="Zoom")
        VisZoomLab.place(x=75, y=115)

        self.VisBrightSlide = Scale(self.tab9, from_=-127, to=127, length=250, orient=HORIZONTAL)
        self.VisBrightSlide.bind("<ButtonRelease-1>", self.VisUpdateBriCon)
        self.VisBrightSlide.place(x=75, y=130)

        VisBrightLab = Label(self.tab9, text="Brightness")
        VisBrightLab.place(x=75, y=150)

        self.VisContrastSlide = Scale(self.tab9, from_=-127, to=127, length=250, orient=HORIZONTAL)
        self.VisContrastSlide.bind("<ButtonRelease-1>", self.VisUpdateBriCon)
        self.VisContrastSlide.place(x=75, y=165)

        VisContrastLab = Label(self.tab9, text="Contrast")
        VisContrastLab.place(x=75, y=185)

        fullRotCbut = Checkbutton(self.tab9, text="Full Rotation Search", variable=self.fullRot)
        fullRotCbut.place(x=900, y=255)

        pick180Cbut = Checkbutton(self.tab9, text="Pick Closest 180°", variable=self.pick180)
        pick180Cbut.place(x=900, y=275)

        pickClosestCbut = Checkbutton(self.tab9, text="Try Closest When Out of Range", variable=self.pickClosest)
        pickClosestCbut.place(x=900, y=295)

        saveCalBut = Button(self.tab9, text="SAVE VISION DATA", width=26, command=self.SaveAndApplyCalibration)
        saveCalBut.place(x=915, y=340)

        #### 6 ENTRY FIELDS##########################################################
        #############################################################################

        self.VisBacColorEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisBacColorEntryField.place(x=390, y=100)
        VisBacColorLab = Label(self.tab9, text="Background Color")
        VisBacColorLab.place(x=390, y=120)

        bgAutoCbut = Checkbutton(self.tab9, command=self.checkAutoBG, text="Auto", variable=self.autoBG)
        bgAutoCbut.place(x=490, y=101)

        self.VisScoreEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisScoreEntryField.place(x=390, y=150)
        VisScoreLab = Label(self.tab9, text="Score Threshold")
        VisScoreLab.place(x=390, y=170)

        self.VisRetScoreEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisRetScoreEntryField.place(x=750, y=55)
        VisRetScoreLab = Label(self.tab9, text="Scored Value")
        VisRetScoreLab.place(x=750, y=75)

        self.VisRetAngleEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisRetAngleEntryField.place(x=750, y=105)
        VisRetAngleLab = Label(self.tab9, text="Found Angle")
        VisRetAngleLab.place(x=750, y=125)

        self.VisRetXpixEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisRetXpixEntryField.place(x=750, y=155)
        VisRetXpixLab = Label(self.tab9, text="Pixel X Position")
        VisRetXpixLab.place(x=750, y=175)

        self.VisRetYpixEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisRetYpixEntryField.place(x=750, y=205)
        VisRetYpixLab = Label(self.tab9, text="Pixel Y Position")
        VisRetYpixLab.place(x=750, y=225)

        self.VisRetXrobEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisRetXrobEntryField.place(x=750, y=255)
        VisRetXrobLab = Label(self.tab9, text="Robot X Position")
        VisRetXrobLab.place(x=750, y=275)

        self.VisRetYrobEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisRetYrobEntryField.place(x=750, y=305)
        VisRetYrobLab = Label(self.tab9, text="Robot Y Position")
        VisRetYrobLab.place(x=750, y=325)

        self.VisX1PixEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisX1PixEntryField.place(x=900, y=55)
        VisX1PixLab = Label(self.tab9, text="X1 Pixel Pos")
        VisX1PixLab.place(x=900, y=75)

        self.VisY1PixEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisY1PixEntryField.place(x=900, y=105)
        VisY1PixLab = Label(self.tab9, text="Y1 Pixel Pos")
        VisY1PixLab.place(x=900, y=125)

        self.VisX2PixEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisX2PixEntryField.place(x=900, y=155)
        VisX2PixLab = Label(self.tab9, text="X2 Pixel Pos")
        VisX2PixLab.place(x=900, y=175)

        self.VisY2PixEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisY2PixEntryField.place(x=900, y=205)
        VisY2PixLab = Label(self.tab9, text="Y2 Pixel Pos")
        VisY2PixLab.place(x=900, y=225)

        self.VisX1RobEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisX1RobEntryField.place(x=1010, y=55)
        VisX1RobLab = Label(self.tab9, text="X1 Robot Pos")
        VisX1RobLab.place(x=1010, y=75)

        self.VisY1RobEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisY1RobEntryField.place(x=1010, y=105)
        VisY1RobLab = Label(self.tab9, text="Y1 Robot Pos")
        VisY1RobLab.place(x=1010, y=125)

        self.VisX2RobEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisX2RobEntryField.place(x=1010, y=155)
        VisX2RobLab = Label(self.tab9, text="X2 Robot Pos")
        VisX2RobLab.place(x=1010, y=175)

        self.VisY2RobEntryField = Entry(self.tab9, width=15, justify="center")
        self.VisY2RobEntryField.place(x=1010, y=205)
        VisY2RobLab = Label(self.tab9, text="Y2 Robot Pos")
        VisY2RobLab.place(x=1010, y=225)

    def motion_search(self, event):
        y = event.x
        x = event.y

        if (x <= 240 and y <= 320):
            self.VisX1PixEntryField.delete(0, 'end')
            self.VisX1PixEntryField.insert(0, x)
            self.VisY1PixEntryField.delete(0, 'end')
            self.VisY1PixEntryField.insert(0, y)
        elif (x > 240):
            self.VisX2PixEntryField.delete(0, 'end')
            self.VisX2PixEntryField.insert(0, x)
        elif (y > 320):
            self.VisY2PixEntryField.delete(0, 'end')
            self.VisY2PixEntryField.insert(0, y)

        # print(str(x) +","+str(y))

    def start_vid_search(self):

        self.stop_vid()
        self.cam_on = True
        curVisStingSel = self.Jvisoptions.get()
        l = len(self.camList)
        for i in range(l):
            if (self.Jvisoptions.get() == self.camList[i]):
                self.selectedCam = i
        self.cap = cv2.VideoCapture(self.selectedCam)
        self.show_frame()

    def stop_vid(self):

        self.cam_on = False

        if self.cap:
            self.cap.release()

    def stop_vid_search(self):

        self.cam_on = False

        if self.cap:
            self.cap.release()

    def selectTemplate_search(self):
        global oriImage
        global button_down
        button_down = False
        x_start, y_start, x_end, y_end = 0, 0, 0, 0
        image = cv2.imread('curImage.jpg')
        oriImage = image.copy()

        cv2.namedWindow("image")
        cv2.setMouseCallback("image", self.mouse_crop)
        cv2.imshow("image", image)

    def snapFind_search(self):

        global BGavg
        self.take_pic()
        template = self.selectedTemplate.get()
        min_score = float(self.VisScoreEntryField.get()) * .01
        self.autoBGVal = int(self.autoBG.get())
        if (self.autoBGVal == 1):
            background = BGavg
            self.VisBacColorEntryField.configure(state='enabled')
            self.VisBacColorEntryField.delete(0, 'end')
            self.VisBacColorEntryField.insert(0, str(BGavg))
            self.VisBacColorEntryField.configure(state='disabled')
        else:
            background = eval(self.VisBacColorEntryField.get())
        self.visFind(template, min_score, background)

    def take_pic_search(self):

        global BGavg

        if (self.cam_on == True):
            ret, frame = self.cap.read()
        else:
            curVisStingSel = self.Jvisoptions.get()
            l = len(self.camList)
            for i in range(l):
                if (self.Jvisoptions.get() == self.camList[i]):
                    self.selectedCam = i
                    # print(self.selectedCam)
            self.cap = cv2.VideoCapture(self.selectedCam)
            ret, frame = self.cap.read()

        brightness = int(self.VisBrightSlide.get())
        contrast = int(self.VisContrastSlide.get())
        self.zoom = int(self.VisZoomSlide.get())

        frame = np.int16(frame)
        frame = frame * (contrast / 127 + 1) - contrast + brightness
        frame = np.clip(frame, 0, 255)
        frame = np.uint8(frame)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # get the webcam size
        height, width = cv2image.shape

        # prepare the crop
        centerX, centerY = int(height / 2), int(width / 2)
        radiusX, radiusY = int(self.zoom * height / 100), int(self.zoom * width / 100)

        minX, maxX = centerX - radiusX, centerX + radiusX
        minY, maxY = centerY - radiusY, centerY + radiusY

        cropped = cv2image[minX:maxX, minY:maxY]
        cv2image = cv2.resize(cropped, (width, height))

        self.autoBGVal = int(self.autoBG.get())
        if (self.autoBGVal == 1):
            BG1 = cv2image[int(self.VisX1PixEntryField.get())][int(self.VisY1PixEntryField.get())]
            BG2 = cv2image[int(self.VisX1PixEntryField.get())][int(self.VisY2PixEntryField.get())]
            BG3 = cv2image[int(self.VisX2PixEntryField.get())][int(self.VisY2PixEntryField.get())]
            avg = int(mean([BG1, BG2, BG3]))
            BGavg = (avg, avg, avg)
            background = avg
            self.VisBacColorEntryField.configure(state='enabled')
            self.VisBacColorEntryField.delete(0, 'end')
            self.VisBacColorEntryField.insert(0, str(BGavg))
            self.VisBacColorEntryField.configure(state='disabled')
        else:
            temp = self.VisBacColorEntryField.get()
            startIndex = temp.find("(")
            endIndex = temp.find(",")
            background = int(temp[startIndex + 1:endIndex])
            # background = eval(self.VisBacColorEntryField.get())

        h = cv2image.shape[0]
        w = cv2image.shape[1]
        # loop over the image
        for y in range(0, h):
            for x in range(0, w):
                # change the pixel
                cv2image[y, x] = background if x >= self.mX2 or x <= self.mX1 or y <= self.mY1 or y >= self.mY2 else cv2image[y, x]

        img = Image.fromarray(cv2image).resize((640, 480))

        imgtk = ImageTk.PhotoImage(image=img)
        self.vid_lbl.imgtk = imgtk
        self.vid_lbl.configure(image=imgtk)
        filename = 'curImage.jpg'
        cv2.imwrite(filename, cv2image)

    def selectMask_search(self):
        global oriImage
        global button_down
        button_down = False
        x_start, y_start, x_end, y_end = 0, 0, 0, 0
        self.mask_pic()

        image = cv2.imread('curImage.jpg')
        oriImage = image.copy()

        cv2.namedWindow("image")
        cv2.setMouseCallback("image", self.mask_crop)
        cv2.imshow("image", image)

    def zeroBrCn_search(self):

        self.mX1 = 0
        self.mY1 = 0
        self.mX2 = 640
        self.mY2 = 480
        self.VisBrightSlide.set(0)
        self.VisContrastSlide.set(0)
        # self.VisZoomSlide.set(50)
        self.take_pic()

    def VisUpdateBriCon(self, foo):
        self.take_pic()

    def checkAutoBG(self):
        self.autoBGVal = int(self.autoBG.get())
        if (self.autoBGVal == 1):
            self.VisBacColorEntryField.configure(state='disabled')
        else:
            self.VisBacColorEntryField.configure(state='enabled')

        ### GCODE DEFS ###################################################################

    def take_pic(self):

        global BGavg

        if (self.cam_on == True):
            ret, frame = self.cap.read()
        else:
            curVisStingSel = self.Jvisoptions.get()
            l = len(self.camList)
            for i in range(l):
                if (self.Jvisoptions.get() == self.camList[i]):
                    self.selectedCam = i
                    # print(self.selectedCam)
            cap = cv2.VideoCapture(self.selectedCam)
            ret, frame = cap.read()

        brightness = int(self.VisBrightSlide.get())
        contrast = int(self.VisContrastSlide.get())
        zoom = int(self.VisZoomSlide.get())

        frame = np.int16(frame)
        frame = frame * (contrast / 127 + 1) - contrast + brightness
        frame = np.clip(frame, 0, 255)
        frame = np.uint8(frame)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # get the webcam size
        height, width = cv2image.shape

        # prepare the crop
        centerX, centerY = int(height / 2), int(width / 2)
        radiusX, radiusY = int(zoom * height / 100), int(zoom * width / 100)

        minX, maxX = centerX - radiusX, centerX + radiusX
        minY, maxY = centerY - radiusY, centerY + radiusY

        cropped = cv2image[minX:maxX, minY:maxY]
        cv2image = cv2.resize(cropped, (width, height))

        autoBGVal = int(self.autoBG.get())
        if (autoBGVal == 1):
            BG1 = cv2image[int(self.VisX1PixEntryField.get())][int(self.VisY1PixEntryField.get())]
            BG2 = cv2image[int(self.VisX1PixEntryField.get())][int(self.VisY2PixEntryField.get())]
            BG3 = cv2image[int(self.VisX2PixEntryField.get())][int(self.VisY2PixEntryField.get())]
            avg = int(mean([BG1, BG2, BG3]))
            BGavg = (avg, avg, avg)
            background = avg
            self.VisBacColorEntryField.configure(state='enabled')
            self.VisBacColorEntryField.delete(0, 'end')
            self.VisBacColorEntryField.insert(0, str(BGavg))
            self.VisBacColorEntryField.configure(state='disabled')
        else:
            temp = self.VisBacColorEntryField.get()
            startIndex = temp.find("(")
            endIndex = temp.find(",")
            background = int(temp[startIndex + 1:endIndex])
            # background = eval(VisBacColorEntryField.get())

        h = cv2image.shape[0]
        w = cv2image.shape[1]
        # loop over the image
        for y in range(0, h):
            for x in range(0, w):
                # change the pixel
                cv2image[y, x] = background if x >= self.mX2 or x <= self.mX1 or y <= self.mY1 or y >= self.mY2 else cv2image[y, x]

        img = Image.fromarray(cv2image).resize((640, 480))

        imgtk = ImageTk.PhotoImage(image=img)
        self.vid_lbl.imgtk = imgtk
        self.vid_lbl.configure(image=imgtk)
        filename = 'curImage.jpg'
        cv2.imwrite(filename, cv2image)

    def testvis(self):
        visprog = self.Jvisoptions.get()
        if (visprog[:] == "Openvision"):
            self.openvision()
        if (visprog[:] == "Roborealm 1.7.5"):
            self.roborealm175()
        if (visprog[:] == "x,y,r"):
            self.xyr()

    def openvision(self):
        global Xpos
        global Ypos

        visfail = 1
        while (visfail == 1):
            value = 0
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
            while (value == 0):
                try:
                    f = open(self.VisFileLoc, "r")
                    value = f.readlines()[-1]  # .decode()
                except:
                    value = 0
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
            x = int(value[110:122])
            y = int(value[130:142])
            self.viscalc(x, y)
            if (Ypos > self.VisEndYmm):
                visfail = 1
                time.sleep(.1)
            else:
                visfail = 0
        open(self.VisFileLoc, "w").close()
        self.VisXfindEntryField.delete(0, 'end')
        self.VisXfindEntryField.insert(0, Xpos)
        self.VisYfindEntryField.delete(0, 'end')
        self.VisYfindEntryField.insert(0, Ypos)
        self.VisRZfindEntryField.delete(0, 'end')
        self.VisRZfindEntryField.insert(0, 0)
        ##
        self.VisXpixfindEntryField.delete(0, 'end')
        self.VisXpixfindEntryField.insert(0, x)
        self.VisYpixfindEntryField.delete(0, 'end')
        self.VisYpixfindEntryField.insert(0, y)
        ##
        self.SP_1_E1_EntryField.delete(0, 'end')
        self.SP_1_E1_EntryField.insert(0, Xpos)
        self.SP_1_E2_EntryField.delete(0, 'end')
        self.SP_1_E2_EntryField.insert(0, Ypos)

    def roborealm175(self):
        global Xpos
        global Ypos

        visfail = 1
        while (visfail == 1):
            value = 0
            self.almStatusLab.config(text="WAITING FOR CAMERA", style="Alarm.TLabel")
            self.almStatusLab2.config(text="WAITING FOR CAMERA", style="Alarm.TLabel")
            while (value == 0):
                try:
                    f = open(self.VisFileLoc, "r")
                    value = f.readlines()[-1]  # .decode()
                except:
                    value = 0
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
            Index = value.index(',')
            x = float(value[:Index])
            y = float(value[Index + 1:])
            self.viscalc(x, y)
            if (float(Ypos) > float(self.VisEndYmm)):
                visfail = 1
                time.sleep(.1)
            else:
                visfail = 0
        open(self.VisFileLoc, "w").close()
        self.VisXfindEntryField.delete(0, 'end')
        self.VisXfindEntryField.insert(0, Xpos)
        self.VisYfindEntryField.delete(0, 'end')
        self.VisYfindEntryField.insert(0, Ypos)
        self.VisRZfindEntryField.delete(0, 'end')
        self.VisRZfindEntryField.insert(0, 0)
        ##
        self.VisXpixfindEntryField.delete(0, 'end')
        self.VisXpixfindEntryField.insert(0, x)
        self.VisYpixfindEntryField.delete(0, 'end')
        self.VisYpixfindEntryField.insert(0, y)
        ##
        self.SP_1_E1_EntryField.delete(0, 'end')
        self.SP_1_E1_EntryField.insert(0, Xpos)
        self.SP_1_E2_EntryField.delete(0, 'end')
        self.SP_1_E2_EntryField.insert(0, Ypos)

    def xyr(self):
        global Xpos
        global Ypos

        visfail = 1
        while (visfail == 1):
            value = 0
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
            while (value == 0):
                try:
                    f = open(self.VisFileLoc, "r")
                    value = f.readlines()[-1]  # .decode()
                except:
                    value = 0
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
            Index = value.index(',')
            x = float(value[:Index])
            value2 = value[Index + 1:]
            Index2 = value2.index(',')
            y = float(value2[:Index2])
            r = float(value2[Index2 + 1:])
            self.viscalc(x, y)
            if (Ypos > float(self.VisEndYmm)):
                visfail = 1
                time.sleep(.1)
            else:
                visfail = 0
        open(self.VisFileLoc, "w").close()
        self.VisXfindEntryField.delete(0, 'end')
        self.VisXfindEntryField.insert(0, Xpos)
        self.VisYfindEntryField.delete(0, 'end')
        self.VisYfindEntryField.insert(0, Ypos)
        self.VisRZfindEntryField.delete(0, 'end')
        self.VisRZfindEntryField.insert(0, r)
        ##
        self.VisXpixfindEntryField.delete(0, 'end')
        self.VisXpixfindEntryField.insert(0, x)
        self.VisYpixfindEntryField.delete(0, 'end')
        self.VisYpixfindEntryField.insert(0, y)
        ##
        self.SP_1_E1_EntryField.delete(0, 'end')
        self.SP_1_E1_EntryField.insert(0, str(Xpos))
        self.SP_1_E2_EntryField.delete(0, 'end')
        self.SP_1_E2_EntryField.insert(0, str(Ypos))
        self.SP_1_E3_EntryField.delete(0, 'end')
        self.SP_1_E3_EntryField.insert(0, r)

    def viscalc(self):
        global xMMpos
        global yMMpos
        # origin x1 y1
        self.VisOrigXpix = float(self.VisX1PixEntryField.get())
        self.VisOrigXmm = float(self.VisX1RobEntryField.get())
        self.VisOrigYpix = float(self.VisY1PixEntryField.get())
        self.VisOrigYmm = float(self.VisY1RobEntryField.get())
        # x2 y2
        self.VisEndXpix = float(self.VisX2PixEntryField.get())
        self.VisEndXmm = float(self.VisX2RobEntryField.get())
        self.VisEndYpix = float(self.VisY2PixEntryField.get())
        self.VisEndYmm = float(self.VisY2RobEntryField.get())

        x = float(self.VisRetXpixEntryField.get())
        y = float(self.VisRetYpixEntryField.get())

        XPrange = float(self.VisEndXpix) - float(self.VisOrigXpix)
        XPratio = (x - float(self.VisOrigXpix)) / XPrange
        XMrange = float(self.VisEndXmm) - float(self.VisOrigXmm)
        XMpos = float(XMrange) * float(XPratio)
        xMMpos = float(self.VisOrigXmm) + XMpos
        ##
        YPrange = float(self.VisEndYpix) - float(self.VisOrigYpix)
        YPratio = (y - float(self.VisOrigYpix)) / YPrange
        YMrange = float(self.VisEndYmm) - float(self.VisOrigYmm)
        YMpos = float(YMrange) * float(YPratio)
        yMMpos = float(self.VisOrigYmm) + YMpos
        return (xMMpos, yMMpos)

    # Define function to show frame
    def show_frame(self):
        if self.cam_on:

            ret, frame = self.cap.read()

            if ret:
                cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(cv2image).resize((480, 320))
                imgtk = ImageTk.PhotoImage(image=img)
                self.live_lbl.imgtk = imgtk
                self.live_lbl.configure(image=imgtk)

            self.live_lbl.after(10, self.show_frame)

    def mask_pic(self):

        global BGavg

        if (self.cam_on == True):
            ret, frame = self.cap.read()
        else:
            curVisStingSel = self.Jvisoptions.get()
            l = len(self.camList)
            for i in range(l):
                if (self.Jvisoptions.get() == self.camList[i]):
                    self.selectedCam = i
                    # print(self.self.selectedCam)
            self.cap = cv2.VideoCapture(self.selectedCam)
            ret, frame = self.cap.read()
        brightness = int(self.VisBrightSlide.get())
        contrast = int(self.VisContrastSlide.get())
        self.zoom = int(self.VisZoomSlide.get())
        frame = np.int16(frame)
        frame = frame * (contrast / 127 + 1) - contrast + brightness
        frame = np.clip(frame, 0, 255)
        frame = np.uint8(frame)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        # get the webcam size
        height, width = cv2image.shape
        # prepare the crop
        centerX, centerY = int(height / 2), int(width / 2)
        radiusX, radiusY = int(self.zoom * height / 100), int(self.zoom * width / 100)
        minX, maxX = centerX - radiusX, centerX + radiusX
        minY, maxY = centerY - radiusY, centerY + radiusY
        cropped = cv2image[minX:maxX, minY:maxY]
        cv2image = cv2.resize(cropped, (width, height))
        # img = Image.fromarray(cv2image).resize((640,480))
        # imgtk = ImageTk.PhotoImage(image=img)
        # self.vid_lbl.imgtk = imgtk
        # self.vid_lbl.configure(image=imgtk)
        filename = 'curImage.jpg'
        cv2.imwrite(filename, cv2image)

    def mask_crop(self, event, x, y, flags, param):
        global x_start, y_start, x_end, y_end, cropping
        global oriImage
        global box_points
        global button_down

        cropDone = False

        if (button_down == False) and (event == cv2.EVENT_LBUTTONDOWN):
            x_start, y_start, x_end, y_end = x, y, x, y
            cropping = True
            button_down = True
            box_points = [(x, y)]

        # Mouse is Moving
        elif (button_down == True) and (event == cv2.EVENT_MOUSEMOVE):
            if cropping == True:
                image_copy = oriImage.copy()
                x_end, y_end = x, y
                point = (x, y)
                cv2.rectangle(image_copy, box_points[0], point, (0, 255, 0), 2)
                cv2.imshow("image", image_copy)

        # if the left mouse button was released
        elif event == cv2.EVENT_LBUTTONUP:
            button_down = False
            box_points.append((x, y))
            cv2.rectangle(oriImage, box_points[0], box_points[1], (0, 255, 0), 2)
            cv2.imshow("image", oriImage)
            # record the ending (x, y) coordinates
            x_end, y_end = x, y
            cropping = False  # cropping is finished

            self.mX1 = x_start + 3
            self.mY1 = y_start + 3
            self.mX2 = x_end - 3
            self.mY2 = y_end - 3

            self.autoBGVal = int(self.autoBG.get())
            if (self.autoBGVal == 1):
                BG1 = oriImage[int(self.VisX1PixEntryField.get())][int(self.VisY1PixEntryField.get())]
                BG2 = oriImage[int(self.VisX1PixEntryField.get())][int(self.VisY2PixEntryField.get())]
                BG3 = oriImage[int(self.VisX2PixEntryField.get())][int(self.VisY2PixEntryField.get())]
                avg = int(mean([BG1, BG2, BG3]))
                BGavg = (avg, avg, avg)
                background = avg
                self.VisBacColorEntryField.configure(state='enabled')
                self.VisBacColorEntryField.delete(0, 'end')
                self.VisBacColorEntryField.insert(0, str(BGavg))
                self.VisBacColorEntryField.configure(state='disabled')
            else:
                background = eval(self.VisBacColorEntryField.get())

            h = oriImage.shape[0]
            w = oriImage.shape[1]
            # loop over the image
            for y in range(0, h):
                for x in range(0, w):
                    # change the pixel
                    oriImage[y, x] = background if x >= self.mX2 or x <= self.mX1 or y <= self.mY1 or y >= self.mY2 else oriImage[y, x]

            img = Image.fromarray(oriImage)
            imgtk = ImageTk.PhotoImage(image=img)
            self.vid_lbl.imgtk = imgtk
            self.vid_lbl.configure(image=imgtk)
            filename = 'curImage.jpg'
            cv2.imwrite(filename, oriImage)
            cv2.destroyAllWindows()

    def mouse_crop(self, event, x, y, flags, param):
        global x_start, y_start, x_end, y_end, cropping
        global oriImage
        global box_points
        global button_down

        cropDone = False

        if (button_down == False) and (event == cv2.EVENT_LBUTTONDOWN):
            x_start, y_start, x_end, y_end = x, y, x, y
            cropping = True
            button_down = True
            box_points = [(x, y)]

        # Mouse is Moving
        elif (button_down == True) and (event == cv2.EVENT_MOUSEMOVE):
            if cropping == True:
                image_copy = oriImage.copy()
                x_end, y_end = x, y
                point = (x, y)
                cv2.rectangle(image_copy, box_points[0], point, (0, 255, 0), 2)
                cv2.imshow("image", image_copy)

        # if the left mouse button was released
        elif event == cv2.EVENT_LBUTTONUP:
            button_down = False
            box_points.append((x, y))
            cv2.rectangle(oriImage, box_points[0], box_points[1], (0, 255, 0), 2)
            cv2.imshow("image", oriImage)
            # record the ending (x, y) coordinates
            x_end, y_end = x, y
            cropping = False  # cropping is finished

            refPoint = [(x_start + 3, y_start + 3), (x_end - 3, y_end - 3)]

            if len(refPoint) == 2:  # when two points were found
                roi = oriImage[refPoint[0][1]:refPoint[1][1], refPoint[0][0]:refPoint[1][0]]

                cv2.imshow("Cropped", roi)
                USER_INP = simpledialog.askstring(title="Teach Vision Object",
                                                  prompt="Save Object As:")
                templateName = USER_INP + ".jpg"
                cv2.imwrite(templateName, roi)
                cv2.destroyAllWindows()
                self.updateVisOp()

    def rotate_image(self, img, angle, background):
        image_center = tuple(np.array(img.shape[1::-1]) / 2)
        rot_mat = cv2.getRotationMatrix2D(image_center, -angle, 1.0)
        result = cv2.warpAffine(img, rot_mat, img.shape[1::-1], borderMode=cv2.BORDER_CONSTANT, borderValue=background, flags=cv2.INTER_LINEAR)
        return result

    def visFind(self, template, min_score, background):
        global xMMpos
        global yMMpos
        global autoBG

        if (background == "Auto"):
            background = BGavg
            self.VisBacColorEntryField.configure(state='enabled')
            self.VisBacColorEntryField.delete(0, 'end')
            self.VisBacColorEntryField.insert(0, str(BGavg))
            self.VisBacColorEntryField.configure(state='disabled')

        green = (0, 255, 0)
        red = (255, 0, 0)
        blue = (0, 0, 255)
        dkgreen = (0, 128, 0)
        status = "fail"
        highscore = 0
        img1 = cv2.imread('curImage.jpg')  # target Image
        img2 = cv2.imread(template)  # target Image

        # method = cv2.TM_CCOEFF_NORMED
        # method = cv2.TM_CCORR_NORMED

        img = img1.copy()

        self.fullRotVal = int(self.fullRot.get())

        for i in range(1):
            if (i == 0):
                method = cv2.TM_CCOEFF_NORMED
            else:
                # method = cv2.TM_CCOEFF_NORMED
                method = cv2.TM_CCORR_NORMED

                # USE 1/3 - EACH SIDE SEARCH
            if (self.fullRotVal == 0):
                ## fist pass 1/3rds
                curangle = 0
                highangle = 0
                highscore = 0
                highmax_loc = 0
                for x in range(3):
                    template = img2
                    template = self.rotate_image(img2, curangle, background)
                    w, h = template.shape[1::-1]
                    res = cv2.matchTemplate(img, template, method)
                    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
                    if (max_val > highscore):
                        highscore = max_val
                        highangle = curangle
                        highmax_loc = max_loc
                        highw, highh = w, h
                    curangle += 120

                # check each side and narrow in
                while True:
                    curangle = curangle / 2
                    if (curangle < .9):
                        break
                    nextangle1 = highangle + curangle
                    nextangle2 = highangle - curangle
                    template = img2
                    template = self.rotate_image(img2, nextangle1, background)
                    w, h = template.shape[1::-1]
                    res = cv2.matchTemplate(img, template, method)
                    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
                    if (max_val > highscore):
                        highscore = max_val
                        highangle = nextangle1
                        highmax_loc = max_loc
                        highw, highh = w, h
                    template = img2
                    template = self.rotate_image(img2, nextangle2, background)
                    w, h = template.shape[1::-1]
                    res = cv2.matchTemplate(img, template, method)
                    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
                    if (max_val > highscore):
                        highscore = max_val
                        highangle = nextangle2
                        highmax_loc = max_loc
                        highw, highh = w, h

                        # USE FULL 360 SEARCh
            else:
                for i in range(720):
                    template = self.rotate_image(img2, i, background)
                    w, h = template.shape[1::-1]

                    img = img1.copy()
                    # Apply template Matching
                    res = cv2.matchTemplate(img, template, method)
                    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
                    highscore = max_val
                    highangle = i
                    highmax_loc = max_loc
                    highw, highh = w, h
                    if highscore >= min_score:
                        break
            if (i == 1):
                highscore = highscore * .5
            if highscore >= min_score:
                break

        if highscore >= min_score:
            status = "pass"
            # normalize angle to increment of +180 to -180
            if (highangle > 180):
                highangle = -360 + highangle
            # pick closest 180
            self.pick180Val = int(self.pick180.get())
            if (self.pick180Val == 1):
                if (highangle > 90):
                    highangle = -180 + highangle
                elif (highangle < -90):
                    highangle = 180 + highangle
            # try closest
            self.pickClosestVal = int(self.pickClosest.get())
            if (self.pickClosestVal == highangle and highangle > int(self.J6PosLim)):
                highangle = self.J6PosLim
            elif (self.pickClosestVal == 0 and highangle > int(self.J6PosLim)):
                status = "fail"
            if (self.pickClosestVal == 1 and highangle < (int(self.J6NegLim) * -1)):
                highangle = self.J6NegLim * -1
            elif (self.pickClosestVal == 0 and highangle < (int(self.J6NegLim) * -1)):
                status = "fail"

            top_left = highmax_loc
            bottom_right = (top_left[0] + highw, top_left[1] + highh)
            # find center
            center = (top_left[0] + highw / 2, top_left[1] + highh / 2)
            xPos = int(center[1])
            yPos = int(center[0])

            imgxPos = int(center[0])
            imgyPos = int(center[1])

            # find line 1 end
            line1x = int(imgxPos + 60 * math.cos(math.radians(highangle - 90)))
            line1y = int(imgyPos + 60 * math.sin(math.radians(highangle - 90)))
            cv2.line(img, (imgxPos, imgyPos), (line1x, line1y), green, 3)

            # find line 2 end
            line2x = int(imgxPos + 60 * math.cos(math.radians(highangle + 90)))
            line2y = int(imgyPos + 60 * math.sin(math.radians(highangle + 90)))
            cv2.line(img, (imgxPos, imgyPos), (line2x, line2y), green, 3)

            # find line 3 end
            line3x = int(imgxPos + 30 * math.cos(math.radians(highangle)))
            line3y = int(imgyPos + 30 * math.sin(math.radians(highangle)))
            cv2.line(img, (imgxPos, imgyPos), (line3x, line3y), green, 3)

            # find line 4 end
            line4x = int(imgxPos + 30 * math.cos(math.radians(highangle + 180)))
            line4y = int(imgyPos + 30 * math.sin(math.radians(highangle + 180)))
            cv2.line(img, (imgxPos, imgyPos), (line4x, line4y), green, 3)

            # find tip start
            lineTx = int(imgxPos + 56 * math.cos(math.radians(highangle - 90)))
            lineTy = int(imgyPos + 56 * math.sin(math.radians(highangle - 90)))
            cv2.line(img, (lineTx, lineTy), (line1x, line1y), dkgreen, 2)

            cv2.circle(img, (imgxPos, imgyPos), 20, green, 1)
            # cv2.rectangle(img,top_left, bottom_right, green, 2)
            cv2.imwrite('temp.jpg', img)
            img = Image.fromarray(img).resize((640, 480))
            imgtk = ImageTk.PhotoImage(image=img)
            self.vid_lbl.imgtk = imgtk
            self.vid_lbl.configure(image=imgtk)
            self.VisRetScoreEntryField.delete(0, 'end')
            self.VisRetScoreEntryField.insert(0, str(round((highscore * 100), 2)))
            self.VisRetAngleEntryField.delete(0, 'end')
            self.VisRetAngleEntryField.insert(0, str(highangle))
            self.VisRetXpixEntryField.delete(0, 'end')
            self.VisRetXpixEntryField.insert(0, str(xPos))
            self.VisRetYpixEntryField.delete(0, 'end')
            self.VisRetYpixEntryField.insert(0, str(yPos))
            self.viscalc()
            self.VisRetXrobEntryField.delete(0, 'end')
            self.VisRetXrobEntryField.insert(0, str(round(xMMpos, 2)))
            self.VisRetYrobEntryField.delete(0, 'end')
            self.VisRetYrobEntryField.insert(0, str(round(yMMpos, 2)))

            # break
            # if (score > highscore):
            # highscore=score

        if status == "fail":
            cv2.rectangle(img, (5, 5), (635, 475), red, 5)
            cv2.imwrite('temp.jpg', img)
            img = Image.fromarray(img).resize((640, 480))
            imgtk = ImageTk.PhotoImage(image=img)
            self.vid_lbl.imgtk = imgtk
            self.vid_lbl.configure(image=imgtk)
            self.VisRetScoreEntryField.delete(0, 'end')
            self.VisRetScoreEntryField.insert(0, str(round((highscore * 100), 2)))
            self.VisRetAngleEntryField.delete(0, 'end')
            self.VisRetAngleEntryField.insert(0, "NA")
            self.VisRetXpixEntryField.delete(0, 'end')
            self.VisRetXpixEntryField.insert(0, "NA")
            self.VisRetYpixEntryField.delete(0, 'end')
            self.VisRetYpixEntryField.insert(0, "NA")

        return (status)

    def updateVisOp(self):

        self.selectedTemplate = StringVar()
        if getattr(sys, 'frozen', False):
            folder = os.path.dirname(sys.executable)
        elif __file__:
            folder = os.path.dirname(os.path.realpath(__file__))
        # folder = os.path.dirname(os.path.realpath(__file__))
        filelist = [fname for fname in os.listdir(folder) if fname.endswith('.jpg')]
        Visoptmenu = ttk.Combobox(self.tab9, textvariable=self.selectedTemplate, values=filelist, state='readonly')
        Visoptmenu.place(x=390, y=52)
        Visoptmenu.bind("<<ComboboxSelected>>", self.VisOpUpdate)

    def VisOpUpdate(self, foo):

        file = self.selectedTemplate.get()
        print(file)
        img = cv2.imread(file, cv2.IMREAD_COLOR)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        TARGET_PIXEL_AREA = 22500

        ratio = float(img.shape[1]) / float(img.shape[0])
        new_h = int(math.sqrt(TARGET_PIXEL_AREA / ratio) + 0.5)
        new_w = int((new_h * ratio) + 0.5)

        img = cv2.resize(img, (new_w, new_h))

        img = Image.fromarray(img)
        imgtk = ImageTk.PhotoImage(image=img)
        self.template_lbl.imgtk = imgtk
        self.template_lbl.configure(image=imgtk)

    #################################################### MAIN ROBOT #########################################################

    def startup(self):
        # global self.moveInProc
        self.moveInProc = 0
        self.updateParams()
        time.sleep(.1)
        self.calExtAxis()
        time.sleep(.1)
        self.sendPos()
        time.sleep(.1)
        self.requestPos()

    def updateParams(self):
        self.TFx = self.TFxEntryField.get()
        self.TFy = self.TFyEntryField.get()
        self.TFz = self.TFzEntryField.get()
        self.TFrz = self.TFrzEntryField.get()
        self.TFry = self.TFryEntryField.get()
        self.TFrx = self.TFrxEntryField.get()
        self.J1motDir = self.J1MotDirEntryField.get()
        self.J2motDir = self.J2MotDirEntryField.get()
        self.J3motDir = self.J3MotDirEntryField.get()
        self.J4motDir = self.J4MotDirEntryField.get()
        self.J5motDir = self.J5MotDirEntryField.get()
        self.J6motDir = self.J6MotDirEntryField.get()
        self.J7motDir = self.J7MotDirEntryField.get()
        self.J8motDir = self.J8MotDirEntryField.get()
        self.J9motDir = self.J9MotDirEntryField.get()
        self.J1calDir = self.J1CalDirEntryField.get()
        self.J2calDir = self.J2CalDirEntryField.get()
        self.J3calDir = self.J3CalDirEntryField.get()
        self.J4calDir = self.J4CalDirEntryField.get()
        self.J5calDir = self.J5CalDirEntryField.get()
        self.J6calDir = self.J6CalDirEntryField.get()
        self.J7calDir = self.J7CalDirEntryField.get()
        self.J8calDir = self.J8CalDirEntryField.get()
        self.J9calDir = self.J9CalDirEntryField.get()
        self.J1PosLim = self.J1PosLimEntryField.get()
        self.J1NegLim = self.J1NegLimEntryField.get()
        self.J2PosLim = self.J2PosLimEntryField.get()
        self.J2NegLim = self.J2NegLimEntryField.get()
        self.J3PosLim = self.J3PosLimEntryField.get()
        self.J3NegLim = self.J3NegLimEntryField.get()
        self.J4PosLim = self.J4PosLimEntryField.get()
        self.J4NegLim = self.J4NegLimEntryField.get()
        self.J5PosLim = self.J5PosLimEntryField.get()
        self.J5NegLim = self.J5NegLimEntryField.get()
        self.J6PosLim = self.J6PosLimEntryField.get()
        self.J6NegLim = self.J6NegLimEntryField.get()
        self.J1StepDeg = self.J1StepDegEntryField.get()
        self.J2StepDeg = self.J2StepDegEntryField.get()
        self.J3StepDeg = self.J3StepDegEntryField.get()
        self.J4StepDeg = self.J4StepDegEntryField.get()
        self.J5StepDeg = self.J5StepDegEntryField.get()
        self.J6StepDeg = self.J6StepDegEntryField.get()
        self.J1EncMult = str(float(self.J1EncCPREntryField.get()) / float(self.J1DriveMSEntryField.get()))
        self.J2EncMult = str(float(self.J2EncCPREntryField.get()) / float(self.J2DriveMSEntryField.get()))
        self.J3EncMult = str(float(self.J3EncCPREntryField.get()) / float(self.J3DriveMSEntryField.get()))
        self.J4EncMult = str(float(self.J4EncCPREntryField.get()) / float(self.J4DriveMSEntryField.get()))
        self.J5EncMult = str(float(self.J5EncCPREntryField.get()) / float(self.J5DriveMSEntryField.get()))
        self.J6EncMult = str(float(self.J6EncCPREntryField.get()) / float(self.J6DriveMSEntryField.get()))
        self.J1ΘDHpar = self.J1ΘEntryField.get()
        self.J2ΘDHpar = self.J2ΘEntryField.get()
        self.J3ΘDHpar = self.J3ΘEntryField.get()
        self.J4ΘDHpar = self.J4ΘEntryField.get()
        self.J5ΘDHpar = self.J5ΘEntryField.get()
        self.J6ΘDHpar = self.J6ΘEntryField.get()
        self.J1αDHpar = self.J1αEntryField.get()
        self.J2αDHpar = self.J2αEntryField.get()
        self.J3αDHpar = self.J3αEntryField.get()
        self.J4αDHpar = self.J4αEntryField.get()
        self.J5αDHpar = self.J5αEntryField.get()
        self.J6αDHpar = self.J6αEntryField.get()
        self.J1dDHpar = self.J1dEntryField.get()
        self.J2dDHpar = self.J2dEntryField.get()
        self.J3dDHpar = self.J3dEntryField.get()
        self.J4dDHpar = self.J4dEntryField.get()
        self.J5dDHpar = self.J5dEntryField.get()
        self.J6dDHpar = self.J6dEntryField.get()
        self.J1aDHpar = self.J1aEntryField.get()
        self.J2aDHpar = self.J2aEntryField.get()
        self.J3aDHpar = self.J3aEntryField.get()
        self.J4aDHpar = self.J4aEntryField.get()
        self.J5aDHpar = self.J5aEntryField.get()
        self.J6aDHpar = self.J6aEntryField.get()

        self.J1negLimLab.config(text="-" + self.J1NegLim, style="Jointlim.TLabel")
        self.J1posLimLab.config(text=self.J1PosLim, style="Jointlim.TLabel")
        self.J1jogslide.config(from_=float("-" + self.J1NegLim), to=float(self.J1PosLim), length=180, orient=HORIZONTAL, command=self.J1sliderUpdate)
        self.J2negLimLab.config(text="-" + self.J2NegLim, style="Jointlim.TLabel")
        self.J2posLimLab.config(text=self.J2PosLim, style="Jointlim.TLabel")
        self.J2jogslide.config(from_=float("-" + self.J2NegLim), to=float(self.J2PosLim), length=180, orient=HORIZONTAL, command=self.J2sliderUpdate)
        self.J3negLimLab.config(text="-" + self.J3NegLim, style="Jointlim.TLabel")
        self.J3posLimLab.config(text=self.J3PosLim, style="Jointlim.TLabel")
        self.J3jogslide.config(from_=float("-" + self.J3NegLim), to=float(self.J3PosLim), length=180, orient=HORIZONTAL, command=self.J3sliderUpdate)
        self.J4negLimLab.config(text="-" + self.J4NegLim, style="Jointlim.TLabel")
        self.J4posLimLab.config(text=self.J4PosLim, style="Jointlim.TLabel")
        self.J4jogslide.config(from_=float("-" + self.J4NegLim), to=float(self.J4PosLim), length=180, orient=HORIZONTAL, command=self.J4sliderUpdate)
        self.J5negLimLab.config(text="-" + self.J5NegLim, style="Jointlim.TLabel")
        self.J5posLimLab.config(text=self.J5PosLim, style="Jointlim.TLabel")
        self.J5jogslide.config(from_=float("-" + self.J5NegLim), to=float(self.J5PosLim), length=180, orient=HORIZONTAL, command=self.J5sliderUpdate)
        self.J6negLimLab.config(text="-" + self.J6NegLim, style="Jointlim.TLabel")
        self.J6posLimLab.config(text=self.J6PosLim, style="Jointlim.TLabel")
        self.J6jogslide.config(from_=float("-" + self.J6NegLim), to=float(self.J6PosLim), length=180, orient=HORIZONTAL, command=self.J6sliderUpdate)

        command = "UP" + "A" + self.TFx + "B" + self.TFy + "C" + self.TFz + "D" + self.TFrz + "E" + self.TFry + "F" + self.TFrx + \
                  "G" + self.J1motDir + "H" + self.J2motDir + "I" + self.J3motDir + "J" + self.J4motDir + "K" + self.J5motDir + "L" + self.J6motDir + "M" + self.J7motDir + "N" + self.J8motDir + "O" + self.J9motDir + \
                  "P" + self.J1calDir + "Q" + self.J2calDir + "R" + self.J3calDir + "S" + self.J4calDir + "T" + self.J5calDir + "U" + self.J6calDir + "V" + self.J7calDir + "W" + self.J8calDir + "X" + self.J9calDir + \
                  "Y" + self.J1PosLim + "Z" + self.J1NegLim + "a" + self.J2PosLim + "b" + self.J2NegLim + "c" + self.J3PosLim + "d" + self.J3NegLim + "e" + self.J4PosLim + "f" + self.J4NegLim + "g" + self.J5PosLim + "h" + self.J5NegLim + "i" + self.J6PosLim + "j" + self.J6NegLim + \
                  "k" + self.J1StepDeg + "l" + self.J2StepDeg + "m" + self.J3StepDeg + "n" + self.J4StepDeg + "o" + self.J5StepDeg + "p" + self.J6StepDeg + \
                  "q" + self.J1EncMult + "r" + self.J2EncMult + "s" + self.J3EncMult + "t" + self.J4EncMult + "u" + self.J5EncMult + "v" + self.J6EncMult + \
                  "w" + self.J1ΘDHpar + "x" + self.J2ΘDHpar + "y" + self.J3ΘDHpar + "z" + self.J4ΘDHpar + "!" + self.J5ΘDHpar + "@" + self.J6ΘDHpar + \
                  "#" + self.J1αDHpar + "$" + self.J2αDHpar + "%" + self.J3αDHpar + "^" + self.J4αDHpar + "&" + self.J5αDHpar + "*" + self.J6αDHpar + \
                  "(" + self.J1dDHpar + ")" + self.J2dDHpar + "+" + self.J3dDHpar + "=" + self.J4dDHpar + "," + self.J5dDHpar + "_" + self.J6dDHpar + \
                  "<" + self.J1aDHpar + ">" + self.J2aDHpar + "?" + self.J3aDHpar + "{" + self.J4aDHpar + "}" + self.J5aDHpar + "~" + self.J6aDHpar + \
                  "\n"
        self.ser.write(command.encode())
        self.ser.flush()
        time.sleep(.1)
        self.ser.flushInput()
        time.sleep(.1)
        response = self.ser.read_all()

    def sendPos(self):
        command = "SP" + "A" + str(self.J1AngCur) + "B" + str(self.J2AngCur) + "C" + str(self.J3AngCur) + "D" + str(self.J4AngCur) + "E" + str(self.J5AngCur) + "F" + str(self.J6AngCur) + "G" + str(self.J7PosCur) + "H" + str(
            self.J8PosCur) + "I" + str(self.J9PosCur) + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = self.ser.read()

    def calExtAxis(self):
        self.J7NegLim = 0
        self.J8NegLim = 0
        self.J9NegLim = 0

        self.J7PosLim = float(self.axis7lengthEntryField.get())
        self.J8PosLim = float(self.axis8lengthEntryField.get())
        self.J9PosLim = float(self.axis9lengthEntryField.get())

        self.J7negLimLab.config(text=str(-self.J7NegLim), style="Jointlim.TLabel")
        self.J8negLimLab.config(text=str(-self.J8NegLim), style="Jointlim.TLabel")
        self.J9negLimLab.config(text=str(-self.J9NegLim), style="Jointlim.TLabel")

        self.J7posLimLab.config(text=str(self.J7PosLim), style="Jointlim.TLabel")
        self.J8posLimLab.config(text=str(self.J8PosLim), style="Jointlim.TLabel")
        self.J9posLimLab.config(text=str(self.J9PosLim), style="Jointlim.TLabel")

        self.J7jogslide.config(from_=-self.J7NegLim, to=self.J7PosLim, length=125, orient=HORIZONTAL, command=self.J7sliderUpdate)
        self.J8jogslide.config(from_=-self.J8NegLim, to=self.J8PosLim, length=125, orient=HORIZONTAL, command=self.J8sliderUpdate)
        self.J9jogslide.config(from_=-self.J9NegLim, to=self.J9PosLim, length=125, orient=HORIZONTAL, command=self.J9sliderUpdate)

        command = "CE" + "A" + str(self.J7PosLim) + "B" + str(self.J7rotation) + "C" + str(self.J7steps) + "D" + str(self.J8PosLim) + "E" + str(self.J8rotation) + "F" + str(self.J8steps) + "G" + str(self.J9PosLim) + "H" + str(
            self.J9rotation) + "I" + str(self.J9steps) + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = self.ser.read()

    def correctPos(self):
        command = "CP\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        self.displayPosition(response)

    def requestPos(self):
        command = "RP\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        self.displayPosition(response)

    def TestString(self):
        message = self.testSendEntryField.get()
        command = "TM" + message + "\n"
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(0)
        echo = self.ser.readline()
        self.testRecEntryField.delete(0, 'end')
        self.testRecEntryField.insert(0, echo)

    def ClearTestString(self):
        self.testRecEntryField.delete(0, 'end')

    def CalcLinDist(self, X2, Y2, Z2):
        global LineDist
        X1 = self.XcurPos
        Y1 = self.YcurPos
        Z1 = self.ZcurPos
        LineDist = (((X2 - X1) ** 2) + ((Y2 - Y1) ** 2) + ((Z2 - Z1) ** 2)) ** .5
        return (LineDist)

    def CalcLinVect(self, X2, Y2, Z2):

        global Xv
        global Yv
        global Zv
        X1 = self.XcurPos
        Y1 = self.YcurPos
        Z1 = self.ZcurPos
        Xv = X2 - X1
        Yv = Y2 - Y1
        Zv = Z2 - Z1
        return (Xv, Yv, Zv)

    def insertvisFind(self):

        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        template = self.selectedTemplate.get()
        if (template == ""):
            template = "None_Selected.jpg"
        self.autoBGVal = int(self.autoBG.get())
        if (self.autoBGVal == 1):
            BGcolor = "(Auto)"
        else:
            BGcolor = self.VisBacColorEntryField.get()
        score = self.VisScoreEntryField.get()
        passTab = self.visPassEntryField.get()
        failTab = self.visFailEntryField.get()
        value = "Vis Find - " + template + " - BGcolor " + BGcolor + " Score " + score + " Pass " + passTab + " Fail " + failTab
        self.tab10.progView.insert(selRow, bytes(value + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def IfRegjumpTab(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        regNum = self.regNumJmpEntryField.get()
        regEqNum = self.regEqJmpEntryField.get()
        tabNum = self.regTabJmpEntryField.get()
        tabjmp = "If Register " + regNum + " = " + regEqNum + " Jump to Tab " + tabNum
        self.tab10.progView.insert(selRow, bytes(tabjmp + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def insertRegister(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        regNum = self.regNumEntryField.get()
        regCmd = self.regEqEntryField.get()
        regIns = "Register " + regNum + " = " + regCmd
        self.tab10.progView.insert(selRow, bytes(regIns + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def storPos(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        regNum = self.storPosNumEntryField.get()
        regElmnt = self.storPosElEntryField.get()
        regCmd = self.storPosValEntryField.get()
        regIns = "Position Register " + regNum + " Element " + regElmnt + " = " + regCmd
        self.tab10.progView.insert(selRow, bytes(regIns + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def insCalibrate(self):
        try:
            selRow = self.tab10.progView.curselection()[0]
            selRow += 1
        except:
            last = self.tab10.progView.index('end')
            selRow = last
            self.tab10.progView.select_set(selRow)
        insCal = "Calibrate Robot"
        self.tab10.progView.insert(selRow, bytes(insCal + '\n', 'utf-8'))
        self.tab10.progView.selection_clear(0, END)
        self.tab10.progView.select_set(selRow)
        items = self.tab10.progView.get(0, END)
        file_path = path.relpath(self.ProgEntryField.get())
        with open(file_path, 'w', encoding='utf-8') as f:
            for item in items:
                f.write(str(item.strip(), encoding='utf-8'))
                f.write('\n')
            f.close()

    def J1jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + str(float(self.J1AngCur) - value) + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J1jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + str(float(self.J1AngCur) + value) + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J2jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + str(float(self.J2AngCur) - value) + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J2jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + str(float(self.J2AngCur) + value) + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J3jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + str(float(self.J3AngCur) - value) + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J3jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + str(float(self.J3AngCur) + value) + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J4jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + str(float(self.J4AngCur) - value) + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J4jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + str(float(self.J4AngCur) + value) + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J5jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + str(float(self.J5AngCur) - value) + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.write(command.encode())
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J5jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + str(float(self.J5AngCur) + value) + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J6jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + str(float(self.J6AngCur) - value) + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J6jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + str(float(self.J6AngCur) + value) + "J7" + str(self.J7PosCur) + "J8" + str(
            self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J7jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(float(self.J7PosCur) - value) + "J8" + str(self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J7jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(float(self.J7PosCur) + value) + "J8" + str(self.J8PosCur) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J8jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(float(self.J8PosCur) - value) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J8jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(float(self.J8PosCur) + value) + "J9" + str(
            self.J9PosCur) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J9jogNeg(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(self.J8PosCur) + "J9" + str(
            float(self.J9PosCur) - value) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    def J9jogPos(self, value):

        self.checkSpeedVals()
        if self.xboxUse != 1:
            self.almStatusLab.config(text="SYSTEM READY", style="OK.TLabel")
            self.almStatusLab2.config(text="SYSTEM READY", style="OK.TLabel")
        speedtype = self.speedOption.get()
        # dont allow mm/sec - switch to percent
        if (speedtype == "mm per Sec"):
            speedMenu = OptionMenu(self.tab10, self.speedOption, "Percent", "Percent", "Seconds", "mm per Sec")
            self.speedPrefix = "Sp"
            self.speedEntryField.delete(0, 'end')
            self.speedEntryField.insert(0, "50")
        # seconds
        if (speedtype == "Seconds"):
            self.speedPrefix = "Ss"
        # percent
        if (speedtype == "Percent"):
            self.speedPrefix = "Sp"
        self.Speed = self.speedEntryField.get()
        self.ACCspd = self.ACCspeedField.get()
        self.DECspd = self.DECspeedField.get()
        self.ACCramp = self.ACCrampField.get()
        self.LoopMode = str(self.J1OpenLoopStat.get()) + str(self.J2OpenLoopStat.get()) + str(self.J3OpenLoopStat.get()) + str(self.J4OpenLoopStat.get()) + str(self.J5OpenLoopStat.get()) + str(self.J6OpenLoopStat.get())
        command = "RJ" + "A" + self.J1AngCur + "B" + self.J2AngCur + "C" + self.J3AngCur + "D" + self.J4AngCur + "E" + self.J5AngCur + "F" + self.J6AngCur + "J7" + str(self.J7PosCur) + "J8" + str(self.J8PosCur) + "J9" + str(
            float(self.J9PosCur) + value) + self.speedPrefix + self.Speed + "Ac" + self.ACCspd + "Dc" + self.DECspd + "Rm" + self.ACCramp + "W" + self.WC + "Lm" + self.LoopMode + "\n"
        self.ser.write(command.encode())
        self.cmdSentEntryField.delete(0, 'end')
        self.cmdSentEntryField.insert(0, command)
        self.ser.flushInput()
        time.sleep(.1)
        response = str(self.ser.readline().strip(), 'utf-8')
        if (response[:1] == 'E'):
            self.ErrorHandler(response)
        else:
            self.displayPosition(response)

    #####################################################################################################################
    def create_camera_processing(self):
        ############TAB 7  #############CAMERA#####################

        self.canvas_cam = tk.Canvas(self.tab7, background="#CBD3C0", width=500, height=550)
        self.canvas_cam.place(x=0, y=0)
        # マウスイベント
        self.canvas_cam.bind("<Motion>", self.mouse_move)  # MouseMove
        self.canvas_cam.bind("<B1-Motion>", self.mouse_move_left)  # MouseMove（左ボタンを押しながら移動）
        self.canvas_cam.bind("<Button-1>", self.mouse_down_left)  # MouseDown（左ボタン）
        self.canvas_cam.bind("<B2-Motion>", self.mouse_move_mid)  #
        self.canvas_cam.bind("<Double-Button-1>", self.mouse_double_click_left)  # MouseDoubleClick（左ボタン）
        self.canvas_cam.bind("<MouseWheel>", self.mouse_wheel)  # MouseWheel
        self.canvas_cam.bind("<ButtonRelease-1>", self.mouse_up_left)
        self.canvas_cam.bind("<Button-2>", self.mouse_down_mid)  # MouseMove（中ボタンを押しながら移動）
        self.canvas_cam.bind("<B3-Motion>", self.mouse_move_right)  # MouseMove（左ボタンを押しながら移動）
        self.canvas_cam.bind("<Button-3>", self.mouse_down_right)  # MouseDown（左ボタン）
        self.canvas_cam.bind("<ButtonRelease-3>", self.mouse_up_right)

        self.cam_capture = tk.Button(self.tab7, command=self.camera_callback, text="capture", bg="green", font=self.small_font)
        self.video_capture = tk.Button(self.tab7, command=self.capture_switch, text="Continue_capture", font=self.small_font, bg='#FF4C00')  # , wraplength=60
        self.reset_data_cam = tk.Button(self.tab7, command=self.reset_data, text="Reset_all_setting", font=self.small_font, bg="gray")  # , wraplength=60
        self.cam_capture.place(x=0, y=570)
        self.video_capture.place(x=150, y=570)
        self.reset_data_cam.place(x=300, y=570)
        self.reset_data_cam["state"] = "disable"
        #
        self.tabcam_control = ttk.Notebook(self.tab7, height=650, width=510)
        self.tabcam0 = ttk.Frame(self.tabcam_control)
        self.tabcam2 = ttk.Frame(self.tabcam_control)
        self.tabcam3 = ttk.Frame(self.tabcam_control)

        self.tabcam_control.add(self.tabcam0, state=self.status, text='カメラ設定')
        self.tabcam_control.add(self.tabcam2, state=self.status, text='カメラFilter')
        self.tabcam_control.add(self.tabcam3, state=self.status, text='検査設定')

        # tab camerasetting
        self.labelcamset = tk.LabelFrame(self.tabcam0, text="カメラ設定", labelanchor="nw", width=480, height=400)
        self.labelcamset.propagate(False)
        self.labelcamset.place(x=20, y=20)
        self.labelsetting = tk.LabelFrame(self.tabcam3, text="Area＿現在価値", labelanchor="nw", width=130, height=90)
        self.labelsetting.propagate(False)
        self.labelsetting.place(x=370, y=270)
        self.area_test_check = tk.Label(self.labelsetting, textvariable=self.now_area_check, font=("Arial", 30, "bold"), )
        self.area_test_check.place(x=5, y=5)
        self.chuvi_test_check = tk.Label(self.labelsetting, textvariable=self.now_chuvi_check, font=("Arial", 30, "bold"), )
        FRAME_WIDTH_lb = Label(self.labelcamset, text="FRAME_WIDTH")
        FRAME_WIDTH = tk.Scale(self.labelcamset, variable=self.FRAME_WIDTH_var, command=self.camera_setting, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=2048, resolution=1, tickinterval=0)
        FRAME_HEIGHT_lb = Label(self.labelcamset, text="FRAME_HEIGHT")
        FRAME_HEIGHT = tk.Scale(self.labelcamset, variable=self.FRAME_HEIGHT_var, command=self.camera_setting1, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=1296, resolution=1, tickinterval=0)
        FPS_lb = Label(self.labelcamset, text="FPS")
        FPS = tk.Scale(self.labelcamset, variable=self.FPS_var, command=self.camera_setting2, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=200, resolution=1, tickinterval=0)
        BRIGHTNESS_lb = Label(self.labelcamset, text="BRIGHTNESS")
        BRIGHTNESS = tk.Scale(self.labelcamset, variable=self.BRIGHTNESS_var, command=self.camera_setting3, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-100, to=100, resolution=1, tickinterval=0)
        CONTRAST_lb = Label(self.labelcamset, text="CONTRAST")
        CONTRAST = tk.Scale(self.labelcamset, variable=self.CONTRAST_var, command=self.camera_setting4, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=200, resolution=1, tickinterval=0)
        SATURATION_lb = Label(self.labelcamset, text="SATURATION")
        SATURATION = tk.Scale(self.labelcamset, variable=self.SATURATION_var, command=self.camera_setting5, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=200, resolution=1, tickinterval=0)
        HUE_lb = Label(self.labelcamset, text="HUE")
        HUE = tk.Scale(self.labelcamset, variable=self.HUE_var, command=self.camera_setting6, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-100, to=100, resolution=1, tickinterval=0)
        GAIN_lb = Label(self.labelcamset, text="GAIN")
        GAIN = tk.Scale(self.labelcamset, variable=self.GAIN_var, command=self.camera_setting7, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-100, to=100, resolution=1, tickinterval=0)
        expose_lb = Label(self.labelcamset, text="EXPOSE")
        EXPOSE_slide = tk.Scale(self.labelcamset, variable=self.EXPOSE_var, command=self.camera_setting10, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-10, to=10, resolution=0.1, tickinterval=0)
        auto_expose_enable = tk.Checkbutton(self.labelcamset, text='Auto EXPOSURE', onvalue=1, offvalue=0, variable=self.camera_setting8)
        auto_focus_enable = tk.Checkbutton(self.labelcamset, text='Auto FOCUS', onvalue=1, offvalue=0, variable=self.camera_setting9)

        FRAME_WIDTH_lb.place(x=0, y=15)
        FRAME_WIDTH.place(x=110, y=0)
        FRAME_HEIGHT_lb.place(x=0, y=45)
        FRAME_HEIGHT.place(x=110, y=30)
        FPS_lb.place(x=0, y=75)
        FPS.place(x=110, y=60)
        BRIGHTNESS_lb.place(x=0, y=105)
        BRIGHTNESS.place(x=110, y=90)
        CONTRAST_lb.place(x=0, y=135)
        CONTRAST.place(x=110, y=120)
        SATURATION_lb.place(x=0, y=165)
        SATURATION.place(x=110, y=150)
        HUE_lb.place(x=0, y=195)
        HUE.place(x=110, y=180)
        GAIN_lb.place(x=0, y=225)
        GAIN.place(x=110, y=210)
        expose_lb.place(x=0, y=255)
        EXPOSE_slide.place(x=110, y=240)
        auto_expose_enable.place(x=0, y=280)
        auto_focus_enable.place(x=0, y=305)

        for child in self.labelcamset.winfo_children():
            child.configure(state='disable')

        self.tabcam_control.place(x=500, y=0)
        # ボタンの作成
        btn_modeless = tk.Button(self.tab6, text="Unlock", command=self.create_anso_dialog_off)
        # btn_modeless.place(x=800, y=0)
        btn_modal = tk.Button(self.tab7, text="Click to Unlock", command=self.create_anso_dialog_on)
        btn_modal.place(x=900, y=0)

        ########### CAM FITER tab 2#######################
        self.labelfilter = tk.LabelFrame(self.tabcam2, text="EDGE_発見", labelanchor="nw", width=350, height=300)
        self.labelfilter.propagate(False)
        self.labelfilter.place(x=150, y=0)

        ###########CAM TAB hani #######################
        hani_enable = tk.Checkbutton(self.tabcam2, text='Check_範囲', onvalue=1, offvalue=0, variable=self.check_hani_ok)
        hani_enable.place(x=0, y=305)
        hani_enable1 = tk.Checkbutton(self.tabcam2, text='玉', onvalue=1, offvalue=0, variable=self.check_patan1)
        hani_enable1.place(x=0, y=325)
        hani_enable2 = tk.Checkbutton(self.tabcam2, text='3角', onvalue=1, offvalue=0, variable=self.check_patan2)
        hani_enable2.place(x=0, y=345)
        hani_enable3 = tk.Checkbutton(self.tabcam2, text='4角', onvalue=1, offvalue=0, variable=self.check_patan3)
        hani_enable3.place(x=0, y=365)
        hani_enable4 = tk.Checkbutton(self.tabcam2, text='5角', onvalue=1, offvalue=0, variable=self.check_patan4)
        hani_enable4.place(x=0, y=385)
        hani_enable5 = tk.Checkbutton(self.tabcam2, text='6角', onvalue=1, offvalue=0, variable=self.check_patan5)
        hani_enable5.place(x=0, y=405)
        hani_enable6 = tk.Checkbutton(self.tabcam2, text='Rectangular', onvalue=1, offvalue=0, variable=self.check_patan6)
        hani_enable6.place(x=0, y=425)

        colormap_lb = Label(self.labelfilter, text="COLOR MAP")
        colormap_lb.place(x=0, y=135)
        thresmap1_lb = Label(self.labelfilter, text="THRES MAP")
        thresmap1_lb.place(x=0, y=165)
        thresmap2_lb = Label(self.labelfilter, text="THRES ADV MAP")
        thresmap2_lb.place(x=0, y=195)

        self.colorchoosen = ttk.Combobox(self.labelfilter, width=20, textvariable=self.colormap)
        self.colorchoosen['values'] = (
            'NORMAL_NO_FILTER', 'GRAY_SCALE', 'COLORMAP_AUTUMN', 'COLORMAP_JET', 'COLORMAP_WINTER', 'COLORMAP_RAINBOW', 'COLORMAP_OCEAN', 'COLORMAP_SUMMER', 'COLORMAP_SPRING', 'COLORMAP_COOL', 'COLORMAP_HSV', 'COLORMAP_PINK',
            'COLORMAP_HOT')
        self.colorchoosen.place(x=200, y=135)
        self.colorchoosen.current(0)

        self.boder_filter_map = ttk.Combobox(self.labelfilter, width=15, textvariable=self.boder_filter_map_sl)
        self.boder_filter_map['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_filter_map.place(x=200, y=165)
        self.boder_filter_map.current(0)

        self.boder_filter_map1 = ttk.Combobox(self.labelfilter, width=15, textvariable=self.boder_filter_map_sl1)
        self.boder_filter_map1['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_filter_map1.place(x=200, y=195)
        self.boder_filter_map1.current(0)

        thresholdmap_lb = Label(self.labelfilter, text="threshold")
        thresholdmap_lb.place(x=0, y=10)

        scalethreshold = tk.Scale(self.labelfilter, variable=self.scale_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=0.1, to=255, resolution=0.1, tickinterval=0)
        scalethreshold.place(x=70, y=-5)
        spinbox_Threshold = ttk.Spinbox(self.labelfilter, from_=0.01, to=255, increment=0.1, wrap=True, state="normal", width=15, textvariable=self.scale_var)
        spinbox_Threshold.place(x=220, y=10)

        adaptiveThreshold_lb = Label(self.labelfilter, text="adapt_thres")
        adaptiveThreshold_lb.place(x=0, y=40)
        adaptiveThreshold = tk.Scale(self.labelfilter, variable=self.adaptiveThreshold_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=0.01, to=100, resolution=0.01, tickinterval=0)
        adaptiveThreshold.place(x=70, y=25)
        spinbox_adaptiveThreshold = ttk.Spinbox(self.labelfilter, from_=0.01, to=100, increment=0.01, wrap=True, state="normal", width=15, textvariable=self.adaptiveThreshold_var)
        spinbox_adaptiveThreshold.place(x=220, y=40)
        #
        canny1_lb = Label(self.labelfilter, text="canny_1")
        canny1_lb.place(x=0, y=70)
        canny1 = tk.Scale(self.labelfilter, variable=self.canny1_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=1, to=600, resolution=1, tickinterval=0)
        canny1.place(x=70, y=55)
        spinbox_canny1 = ttk.Spinbox(self.labelfilter, from_=1, to=600, increment=1, wrap=True, state="normal", width=15, textvariable=self.canny1_var)
        spinbox_canny1.place(x=220, y=70)

        canny2_lb = Label(self.labelfilter, text="canny_2")
        canny2_lb.place(x=0, y=100)
        canny2 = tk.Scale(self.labelfilter, variable=self.canny2_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=1, to=600, resolution=1, tickinterval=0)
        canny2.place(x=70, y=85)
        spinbox_canny2 = ttk.Spinbox(self.labelfilter, from_=1, to=600, increment=1, wrap=True, state="normal", width=15, textvariable=self.canny2_var)
        spinbox_canny2.place(x=220, y=100)
        #
        self.labelfilter1 = tk.LabelFrame(self.tabcam2, text="Smooth", labelanchor="nw", width=350, height=290)
        self.labelfilter1.propagate(False)
        self.labelfilter1.place(x=150, y=300)

        gaussion_lb = Label(self.labelfilter1, text="gaussion")
        gaussion_lb.place(x=0, y=25)
        gaussion = tk.Scale(self.labelfilter1, variable=self.gaussion_var, orient=tk.HORIZONTAL, length=100, width=9, sliderlength=15, from_=0.1, to=300, resolution=0.1, tickinterval=0)
        gaussion.place(x=70, y=10)
        spinbox_gaussion = ttk.Spinbox(self.labelfilter1, from_=0.1, to=300, increment=0.1, wrap=True, state="normal", width=15, textvariable=self.gaussion_var)
        spinbox_gaussion.place(x=220, y=25)
        #
        Blur_lb = Label(self.labelfilter1, text="Blur")
        Blur_lb.place(x=0, y=55)
        Blur = tk.Scale(self.labelfilter1, variable=self.Blur_var, orient=tk.HORIZONTAL, length=100, width=9, sliderlength=15, from_=1, to=100, resolution=1, tickinterval=0)
        Blur.place(x=70, y=40)
        spinbox_Blur = ttk.Spinbox(self.labelfilter1, from_=1, to=100, increment=1, wrap=True, state="normal", width=15, textvariable=self.Blur_var)
        spinbox_Blur.place(x=220, y=55)
        #
        boldx_lb = Label(self.labelfilter1, text="Bold_x")
        boldx_lb.place(x=0, y=85)
        boldy_lb = Label(self.labelfilter1, text="Bold_y")
        boldy_lb.place(x=0, y=115)
        deboldx_lb = Label(self.labelfilter1, text="Debold_x")
        deboldx_lb.place(x=0, y=145)
        deboldy_lb = Label(self.labelfilter1, text="Debold_y")
        deboldy_lb.place(x=0, y=175)
        spinbox_bold = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.bold_var)
        spinbox_bold.place(x=220, y=85)
        spinbox_debold = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.debold_var)
        spinbox_debold.place(x=220, y=115)
        spinbox_bold1 = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.bold_var1)
        spinbox_bold1.place(x=220, y=145)
        spinbox_debold1 = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.debold_var1)
        spinbox_debold1.place(x=220, y=175)
        self.labelfilter2 = tk.LabelFrame(self.tabcam2, text="Process", labelanchor="nw", width=150, height=300)
        self.labelfilter2.propagate(False)
        self.labelfilter2.place(x=0, y=0)
        self.filterchoosen = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap)
        self.filterchoosen['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen.place(x=5, y=0)
        self.filterchoosen.current(0)
        self.filterchoosen1 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap1)
        self.filterchoosen1['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen1.current(0)
        self.filterchoosen2 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap2)
        self.filterchoosen2['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen2.current(0)
        self.filterchoosen3 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap3)
        self.filterchoosen3['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen3.current(0)
        self.filterchoosen4 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap4)
        self.filterchoosen4['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen4.current(0)

        self.add_fiter = tk.Button(self.labelfilter2, command=self.add_fiter_to, text="+", bg="green")
        self.add_fiter.place(x=30, y=220)
        self.minus_fiter = tk.Button(self.labelfilter2, command=self.minus_fiter_to, text="-", bg="red")
        self.minus_fiter.place(x=50, y=220)

        for child in self.labelfilter.winfo_children():
            child.configure(state='disable')
        for child in self.labelfilter1.winfo_children():
            child.configure(state='disable')
        for child in self.labelfilter2.winfo_children():
            child.configure(state='disable')

        ############TAB TEST  ##################################
        self.bar_tenken = tk.LabelFrame(self.tab8, text="TEST", labelanchor="nw", width=550, height=150, font=self.small_font)
        self.bar_tenken.propagate(False)
        self.bar_tenken.place(x=20, y=10)
        self.test_start = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.test_start.place(x=10, y=2)
        self.test_setok = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.test_setok.place(x=150, y=2)
        self.scan_start = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.scan_start.place(x=290, y=2)
        self.toridasi_ok = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.toridasi_ok.place(x=430, y=2)
        self.bar_tenken_recieved = tk.LabelFrame(self.tab8, text="バーコード表記内容", labelanchor="nw", width=300, height=300, font=self.small_font)
        self.bar_tenken_recieved.propagate(False)
        self.bar_tenken_recieved.place(x=600, y=10)

        self.barreceived_text = tk.Text(self.bar_tenken_recieved, height=20, width=40, background='#856ff8')
        self.barreceived_text.place(x=5, y=5)
        self.bar_ketka = tk.LabelFrame(self.tab8, text="NC", labelanchor="nw", width=300, height=130, font=self.small_font)
        self.bar_ketka.propagate(False)
        self.bar_ketka.place(x=270, y=180)
        self.BAR_test_check = tk.Label(self.bar_ketka, textvariable=self.BAR_check, font=("Arial", 30, "bold"), )
        self.BAR_test_check.place(x=5, y=5)

        self.test_start.bind("<ButtonPress-1>", self.test_start_on_click)
        self.test_start.bind("<ButtonRelease-1>", self.test_start_off_click)
        self.test_setok.bind("<ButtonPress-1>", self.test_setok_on_click)
        self.test_setok.bind("<ButtonRelease-1>", self.test_setok_off_click)
        self.scan_start.bind("<ButtonPress-1>", self.scan_start_on_click)
        self.scan_start.bind("<ButtonRelease-1>", self.scan_start_off_click)
        self.toridasi_ok.bind("<ButtonPress-1>", self.toridasi_ok_on_click)
        self.toridasi_ok.bind("<ButtonRelease-1>", self.toridasi_ok_off_click)

        # ===============================================================================================================================================
        self.fil_by_Area = tk.IntVar()
        self.fil_by_Area.set(1)
        self.filByCircularity = tk.IntVar()
        self.filByCircularity.set(1)
        self.filByConvexity = tk.IntVar()
        self.filByConvexity.set(1)
        self.filByInertia = tk.IntVar()
        self.filByInertia.set(1)
        self.minArea_var = tk.IntVar()
        self.minArea_var.set(50)
        self.minCircularity_var = tk.DoubleVar()
        self.minCircularity_var.set(0.1)
        self.minConvexity_var = tk.DoubleVar()
        self.minConvexity_var.set(0.1)
        self.minInertia_var = tk.DoubleVar()
        self.minInertia_var.set(0.01)
        self.labeldetect = tk.LabelFrame(self.tabcam3, text="Circle_detect", labelanchor="nw", width=170, height=110)
        self.labeldetect.propagate(False)
        self.fil_select1 = tk.Checkbutton(self.labeldetect, text='Fil by Area', onvalue=1, offvalue=0, variable=self.fil_by_Area)
        self.fil_select2 = tk.Checkbutton(self.labeldetect, text='Fil By Circularity', onvalue=1, offvalue=0, variable=self.filByCircularity)
        self.fil_select3 = tk.Checkbutton(self.labeldetect, text='fil By Convexity ', onvalue=1, offvalue=0, variable=self.filByConvexity)
        self.fil_select4 = tk.Checkbutton(self.labeldetect, text='fil By Inertia  ', onvalue=1, offvalue=0, variable=self.filByInertia)
        minArea = ttk.Spinbox(self.labeldetect, from_=1, to=800, increment=1, width=6, textvariable=self.minArea_var)
        minCircularity = ttk.Spinbox(self.labeldetect, from_=0.01, to=300, increment=0.01, width=6, textvariable=self.minCircularity_var)
        minConvexity = ttk.Spinbox(self.labeldetect, from_=0.01, to=300, increment=0.01, width=6, textvariable=self.minConvexity_var)
        minInertia = ttk.Spinbox(self.labeldetect, from_=0.01, to=300, increment=0.01, width=6, textvariable=self.minInertia_var)
        self.labeldetect.place(x=0, y=0)
        self.fil_select1.place(x=0, y=0)
        self.fil_select2.place(x=0, y=20)
        self.fil_select3.place(x=0, y=40)
        self.fil_select4.place(x=0, y=60)
        minArea.place(x=110, y=0)
        minCircularity.place(x=110, y=20)
        minConvexity.place(x=110, y=40)
        minInertia.place(x=110, y=60)

        # ===============================================================================================================================================
        self.Kernel_var = tk.IntVar()
        self.Kernel_var.set(3)
        self.param1_var = tk.IntVar()
        self.param1_var.set(50)
        self.param2_var = tk.IntVar()
        self.param2_var.set(30)
        self.minRadius_var = tk.IntVar()
        self.minRadius_var.set(1)
        self.maxRadius_var = tk.IntVar()
        self.maxRadius_var.set(40)
        self.labeldetect1 = tk.LabelFrame(self.tabcam3, text="Circle_detect_adv", labelanchor="nw", width=170, height=130)
        self.labeldetect1.propagate(False)
        Kernel_cir_lb = Label(self.labeldetect1, text="Kernel")
        param1_cir = Label(self.labeldetect1, text="param1")
        param2_cir = Label(self.labeldetect1, text="param2")
        minRadius_lb = Label(self.labeldetect1, text="minRadius")
        maxRadius_lb = Label(self.labeldetect1, text="maxRadius")
        Kernel_ = ttk.Spinbox(self.labeldetect1, from_=1, to=100, increment=1, width=6, textvariable=self.Kernel_var)
        param1_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.param1_var)
        param2_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.param2_var)
        minRadius_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.minRadius_var)
        maxRadius_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.maxRadius_var)
        self.labeldetect1.place(x=0, y=110)
        Kernel_cir_lb.place(x=0, y=0)
        param1_cir.place(x=0, y=20)
        param2_cir.place(x=0, y=40)
        minRadius_lb.place(x=0, y=60)
        maxRadius_lb.place(x=0, y=80)
        Kernel_.place(x=110, y=0)
        param1_.place(x=110, y=20)
        param2_.place(x=110, y=40)
        minRadius_.place(x=110, y=60)
        maxRadius_.place(x=110, y=80)
        # ===============================================================================================================================================
        self.var_canny_a = tk.IntVar()
        self.var_canny_a.set(250)
        self.var_canny_b = tk.IntVar()
        self.var_canny_b.set(150)
        self.var_thr = tk.IntVar()
        self.var_thr.set(50)
        self.len_size2 = tk.DoubleVar()
        self.len_size2.set(50)
        self.len_size3 = tk.DoubleVar()
        self.len_size3.set(15)
        self.len_size1 = IntVar()
        self.len_size1.set(1)
        self.labeldetect2 = tk.LabelFrame(self.tabcam3, text="line_detect", labelanchor="nw", width=170, height=150)
        self.labeldetect2.propagate(False)
        var_canny_a_lb = Label(self.labeldetect2, text="canny-")
        var_canny_b_lb = Label(self.labeldetect2, text="canny+")
        var_thrlb = Label(self.labeldetect2, text="threshole")
        minLineLength_lb = Label(self.labeldetect2, text="minLineLength")
        maxLineGap_lb = Label(self.labeldetect2, text="maxLineGap")
        canny_a = ttk.Spinbox(self.labeldetect2, from_=1, to=600, increment=1, width=6, textvariable=self.var_canny_a)
        canny_b = ttk.Spinbox(self.labeldetect2, from_=1, to=600, increment=1, width=6, textvariable=self.var_canny_b)
        var_thrxx = ttk.Spinbox(self.labeldetect2, from_=1, to=300, increment=1, width=6, textvariable=self.var_thr)
        minLineLength = ttk.Spinbox(self.labeldetect2, from_=1, to=300, increment=1, width=6, textvariable=self.len_size2)
        maxLineGap = ttk.Spinbox(self.labeldetect2, from_=1, to=300, increment=1, width=6, textvariable=self.len_size3)
        spin_length = ttk.Spinbox(self.labeldetect2, from_=1, to=100, width=5, textvariable=self.len_size1)
        lbl_size = Label(self.labeldetect2, text="Distance pixels")
        self.labeldetect2.place(x=0, y=240)
        var_canny_a_lb.place(x=0, y=0)
        var_canny_b_lb.place(x=0, y=20)
        var_thrlb.place(x=0, y=40)
        minLineLength_lb.place(x=0, y=60)
        maxLineGap_lb.place(x=0, y=80)
        canny_a.place(x=110, y=0)
        canny_b.place(x=110, y=20)
        var_thrxx.place(x=110, y=40)
        minLineLength.place(x=110, y=60)
        maxLineGap.place(x=110, y=80)
        spin_length.place(x=110, y=100)
        lbl_size.place(x=0, y=100)

        # ===============================================================================================================================================
        self.labeselectdetect = tk.LabelFrame(self.tabcam3, text="Select_detect", labelanchor="nw", width=170, height=130)
        self.labeselectdetect.propagate(False)
        self.labeselectdetect.place(x=330, y=440)

        self.circle_select = tk.IntVar()
        self.circle_select.set(0)
        self.circle_select_ = tk.Checkbutton(self.labeselectdetect, text='Circle_dt1', onvalue=1, offvalue=0, variable=self.circle_select)
        self.circle_select_.place(x=0, y=0)

        self.circle_select_adv = tk.IntVar()
        self.circle_select_adv.set(0)
        self.circle_select_adv_ = tk.Checkbutton(self.labeselectdetect, text='Circle_dt2', onvalue=1, offvalue=0, variable=self.circle_select_adv)
        self.circle_select_adv_.place(x=0, y=20)

        self.triang_select = tk.IntVar()
        self.triang_select.set(0)
        self.triang_select_ = tk.Checkbutton(self.labeselectdetect, text='triangle_dt ', onvalue=1, offvalue=0, variable=self.triang_select)
        self.triang_select_.place(x=0, y=40)

        self.rect_select = tk.IntVar()
        self.rect_select.set(0)
        self.rect_select_ = tk.Checkbutton(self.labeselectdetect, text='Rect_dt ', onvalue=1, offvalue=0, variable=self.rect_select)
        self.rect_select_.place(x=0, y=60)
        #
        self.edge_select = tk.IntVar()
        self.edge_select.set(0)
        self.edge_select_ = tk.Checkbutton(self.labeselectdetect, text='Line_dt', onvalue=1, offvalue=0, variable=self.edge_select)
        self.edge_select_.place(x=0, y=80)
        #
        self.draw_contour_select = tk.IntVar()
        self.draw_contour_select.set(0)
        self.draw_contour_select_ = tk.Checkbutton(self.labeselectdetect, text='Draw_Ct', onvalue=1, offvalue=0, variable=self.draw_contour_select)
        self.draw_contour_select_.place(x=85, y=0)
        #
        self.connected_select = tk.IntVar()
        self.connected_select.set(0)
        self.connected_select_ = tk.Checkbutton(self.labeselectdetect, text='Conect_dt', onvalue=1, offvalue=0, variable=self.connected_select)
        self.connected_select_.place(x=85, y=20)
        #
        self.rectangge_component_select = tk.IntVar()
        self.rectangge_component_select.set(0)
        self.rectangge_component_select_ = tk.Checkbutton(self.labeselectdetect, text='■形状_dt', onvalue=1, offvalue=0, variable=self.rectangge_component_select)
        self.rectangge_component_select_.place(x=85, y=40)

        # ===============================================================================================================================================
        self.labeldetect3 = tk.LabelFrame(self.tabcam3, text="rect_detect", labelanchor="nw", width=170, height=180)
        self.labeldetect3.propagate(False)
        self.labeldetect3.place(x=0, y=390)
        self.border_sl = tk.StringVar()
        self.boder_map = ttk.Combobox(self.labeldetect3, width=20, textvariable=self.border_sl)
        # Adding combobox drop down list
        self.boder_map['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_map.place(x=0, y=0)
        self.boder_map.current(0)
        self.contour_sl = tk.StringVar()
        self.contua_map = ttk.Combobox(self.labeldetect3, width=20, textvariable=self.contour_sl)
        # Adding combobox drop down list
        self.contua_map['values'] = ('RETR_TREE', 'RETR_CCOMP', 'RETR_LIST', 'RETR_EXTERNAL')
        self.contua_map.place(x=0, y=25)
        self.contua_map.current(0)
        self.contour_sl1 = tk.StringVar()
        self.contua_map1 = ttk.Combobox(self.labeldetect3, width=20, textvariable=self.contour_sl1)
        # Adding combobox drop down list
        self.contua_map1['values'] = ('CHAIN_APPROX_NONE', 'CHAIN_APPROX_SIMPLE', 'CHAIN_APPROX_TC89_L1')
        self.contua_map1.place(x=0, y=50)
        self.contua_map1.current(1)
        rect_para1_lb = Label(self.labeldetect3, text="Approx")
        rect_para1_lb.place(x=0, y=75)
        area_rect_lb = Label(self.labeldetect3, text="Area")
        area_rect_lb.place(x=0, y=95)
        chuvi_rect_lb = Label(self.labeldetect3, text="Perimeter")
        chuvi_rect_lb.place(x=0, y=115)
        thres_rect_lb = Label(self.labeldetect3, text="Threshole")
        thres_rect_lb.place(x=0, y=135)

        # ===============================================================================================================================================
        self.rect_para1 = tk.DoubleVar()
        self.rect_para1.set(0.04)
        rect_para1_ = ttk.Spinbox(self.labeldetect3, from_=0.01, to=10, increment=0.01, width=6, textvariable=self.rect_para1)
        rect_para1_.place(x=110, y=75)
        self.area_rect = tk.IntVar()
        self.area_rect.set(150)
        area_rect_ = ttk.Spinbox(self.labeldetect3, from_=1, to=600, increment=1, width=6, textvariable=self.area_rect)
        area_rect_.place(x=110, y=95)
        self.chuvi_rect = tk.IntVar()
        self.chuvi_rect.set(50)
        chuvi_rect_ = ttk.Spinbox(self.labeldetect3, from_=1, to=300, increment=1, width=6, textvariable=self.chuvi_rect)
        chuvi_rect_.place(x=110, y=115)
        self.rect_thres = tk.IntVar()
        self.rect_thres.set(150)
        rect_thres_rect_ = ttk.Spinbox(self.labeldetect3, from_=1, to=300, increment=1, width=6, textvariable=self.rect_thres)
        rect_thres_rect_.place(x=110, y=135)
        #############################################################
        self.labeldetect4 = tk.LabelFrame(self.tabcam3, text="Triangle_detect", labelanchor="nw", width=170, height=180)
        self.labeldetect4.propagate(False)
        self.labeldetect4.place(x=172, y=0)
        self.border_sl2 = tk.StringVar()
        self.boder_map2 = ttk.Combobox(self.labeldetect4, width=20, textvariable=self.border_sl2)
        # Adding combobox drop down list
        self.boder_map2['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_map2.place(x=0, y=0)
        self.boder_map2.current(0)
        self.contour_sl2 = tk.StringVar()
        self.contua_map2 = ttk.Combobox(self.labeldetect4, width=20, textvariable=self.contour_sl2)
        # Adding combobox drop down list
        self.contua_map2['values'] = ('RETR_TREE', 'RETR_CCOMP', 'RETR_LIST', 'RETR_EXTERNAL')
        self.contua_map2.place(x=0, y=25)
        self.contua_map2.current(0)
        self.contour_sl21 = tk.StringVar()
        self.contua_map21 = ttk.Combobox(self.labeldetect4, width=20, textvariable=self.contour_sl21)
        # Adding combobox drop down list
        self.contua_map21['values'] = ('CHAIN_APPROX_NONE', 'CHAIN_APPROX_SIMPLE', 'CHAIN_APPROX_TC89_L1')
        self.contua_map21.place(x=0, y=50)
        self.contua_map21.current(1)
        rect_trian1_lb = Label(self.labeldetect4, text="Approx")
        rect_trian1_lb.place(x=0, y=75)
        area_trian_lb = Label(self.labeldetect4, text="Area")
        area_trian_lb.place(x=0, y=95)
        chuvi_trian_lb = Label(self.labeldetect4, text="Perimeter")
        chuvi_trian_lb.place(x=0, y=115)
        thres_trian_lb = Label(self.labeldetect4, text="Threshole")
        thres_trian_lb.place(x=0, y=135)
        # ===============================================================================================================================================
        self.trian_para1 = tk.DoubleVar()
        self.trian_para1.set(0.04)
        trian_para1_ = ttk.Spinbox(self.labeldetect4, from_=0.01, to=10, increment=0.01, width=6, textvariable=self.trian_para1)
        trian_para1_.place(x=110, y=75)
        self.area_trian = tk.IntVar()
        self.area_trian.set(150)
        area_trian_ = ttk.Spinbox(self.labeldetect4, from_=1, to=600, increment=1, width=6, textvariable=self.area_trian)
        area_trian_.place(x=110, y=95)
        self.chuvi_trian = tk.IntVar()
        self.chuvi_trian.set(50)
        chuvi_trian_ = ttk.Spinbox(self.labeldetect4, from_=1, to=300, increment=1, width=6, textvariable=self.chuvi_trian)
        chuvi_trian_.place(x=110, y=115)
        self.trian_thres = tk.IntVar()
        self.trian_thres.set(150)
        trian_thres_rect_ = ttk.Spinbox(self.labeldetect4, from_=1, to=300, increment=1, width=6, textvariable=self.trian_thres)
        trian_thres_rect_.place(x=110, y=135)
        self.labeldetect5 = tk.LabelFrame(self.tabcam3, text="Connected_detect", labelanchor="nw", width=170, height=180)
        self.labeldetect5.propagate(False)
        self.labeldetect5.place(x=172, y=180)
        self.border_connect_sl = tk.StringVar()
        self.border_connect = ttk.Combobox(self.labeldetect5, width=20, textvariable=self.border_connect_sl)
        self.border_connect['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.border_connect.place(x=0, y=0)
        self.border_connect.current(0)
        rect_trian1_lb = Label(self.labeldetect5, text="kernel")
        rect_trian1_lb.place(x=0, y=30)
        iterations_lb = Label(self.labeldetect5, text="iterations")
        iterations_lb.place(x=0, y=50)
        Area_lb = Label(self.labeldetect5, text="Area_detect")
        Area_lb.place(x=0, y=70)
        thres_connect_lb = Label(self.labeldetect5, text="Threshole")
        thres_connect_lb.place(x=0, y=90)
        # ===============================================================================================================================================
        self.kernel_connect = tk.IntVar()
        self.kernel_connect.set(3)
        kernel_connect_ = ttk.Spinbox(self.labeldetect5, from_=1, to=10, increment=1, width=6, textvariable=self.kernel_connect)
        kernel_connect_.place(x=110, y=30)
        self.iterations_var = tk.IntVar()
        self.iterations_var.set(15)
        iterations_ = ttk.Spinbox(self.labeldetect5, from_=1, to=100, increment=1, width=6, textvariable=self.iterations_var)
        iterations_.place(x=110, y=50)
        self.area_detect_connect = tk.IntVar()
        self.area_detect_connect.set(50)
        area_detect_connect_ = ttk.Spinbox(self.labeldetect5, from_=1, to=300, increment=1, width=6, textvariable=self.area_detect_connect)
        area_detect_connect_.place(x=110, y=70)
        self.thres_connect = tk.IntVar()
        self.thres_connect.set(150)
        thres_connect_ = ttk.Spinbox(self.labeldetect5, from_=1, to=300, increment=1, width=6, textvariable=self.thres_connect)
        thres_connect_.place(x=110, y=90)
        # ===============================================================================================================================================

        self.objectdetect = tk.LabelFrame(self.tabcam3, text="■形状_detect", labelanchor="nw", width=170, height=100)
        self.objectdetect.propagate(False)
        self.objectdetect.place(x=344, y=0)
        fill_change_var_lb = Label(self.objectdetect, text="Fill_%")
        fill_change_var_lb.place(x=0, y=0)
        area_patan_lb = Label(self.objectdetect, text="Area")
        area_patan_lb.place(x=0, y=20)
        chuvi_patan_lb = Label(self.objectdetect, text="Perimeter")
        chuvi_patan_lb.place(x=0, y=40)
        thres_patan_lb = Label(self.objectdetect, text="Threshole")
        thres_patan_lb.place(x=0, y=60)
        self.fill_change_var = tk.IntVar()
        self.fill_change_var.set(50)
        fill_change_var_ = ttk.Spinbox(self.objectdetect, from_=1, to=100, increment=0.01, width=6, textvariable=self.fill_change_var)
        fill_change_var_.place(x=110, y=0)
        self.area_patan = tk.IntVar()
        self.area_patan.set(150)
        area_patan_ = ttk.Spinbox(self.objectdetect, from_=1, to=10000, increment=1, width=6, textvariable=self.area_patan)
        area_patan_.place(x=110, y=20)
        self.chuvi_patan = tk.IntVar()
        self.chuvi_patan.set(500)
        chuvi_patan_ = ttk.Spinbox(self.objectdetect, from_=1, to=300, increment=1, width=6, textvariable=self.chuvi_patan)
        chuvi_patan_.place(x=110, y=40)
        self.thres_patan = tk.IntVar()
        self.thres_patan.set(150)
        thres_patan_ = ttk.Spinbox(self.objectdetect, from_=1, to=255, increment=1, width=6, textvariable=self.thres_patan)
        thres_patan_.place(x=110, y=60)

        # ===============================================================================================================================================
        self.objectdetect1 = tk.LabelFrame(self.tabcam3, text="Contuar_detect", labelanchor="nw", width=170, height=100)
        self.objectdetect1.propagate(False)
        self.objectdetect1.place(x=344, y=110)
        Threshole_var_lb = Label(self.objectdetect1, text="Threshole")
        Threshole_var_lb.place(x=0, y=0)
        area_patan_lb = Label(self.objectdetect1, text="approx")
        area_patan_lb.place(x=0, y=20)
        chuvi_contua_lb = Label(self.objectdetect1, text="Perimeter")
        chuvi_contua_lb.place(x=0, y=40)
        thres_patan_lb = Label(self.objectdetect1, text="Area")
        thres_patan_lb.place(x=0, y=60)
        self.Threshole_contua_var = tk.DoubleVar()
        self.Threshole_contua_var.set(50)
        Threshole_contua_var_ = ttk.Spinbox(self.objectdetect1, from_=1, to=255, increment=1, width=6, textvariable=self.Threshole_contua_var)
        Threshole_contua_var_.place(x=110, y=0)
        self.contua_aprrox = tk.DoubleVar()
        self.contua_aprrox.set(0.01)
        contua_aprrox_ = ttk.Spinbox(self.objectdetect1, from_=0.01, to=10, increment=0.01, width=6, textvariable=self.contua_aprrox)
        contua_aprrox_.place(x=110, y=20)
        self.chuvi_contua = tk.IntVar()
        self.chuvi_contua.set(50)
        chuvi_contua_ = ttk.Spinbox(self.objectdetect1, from_=10, to=30000, increment=1, width=6, textvariable=self.chuvi_contua)
        chuvi_contua_.place(x=110, y=40)
        self.dientich_contua = tk.IntVar()
        self.dientich_contua.set(150)
        dientich_contua_ = ttk.Spinbox(self.objectdetect1, from_=1, to=50000, increment=10, width=6, textvariable=self.dientich_contua)
        dientich_contua_.place(x=110, y=60)

        # ===============================================================================================================================================
        for child in self.labeldetect.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect1.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect2.winfo_children():
            child.configure(state='disable')
        for child in self.labeselectdetect.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect3.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect4.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect5.winfo_children():
            child.configure(state='disable')
        for child in self.objectdetect.winfo_children():
            child.configure(state='disable')
        for child in self.objectdetect1.winfo_children():
            child.configure(state='disable')

    #####################################################################################################################
    def rb_mouse_move(self, event):
        # print("dang move")
        ''''''

    def rb_mouse_move_left(self, event):
        self.mouse_now_position_x = event.x
        self.mouse_now_position_y = event.y
        if (self.mouse_now_position_x > self.mouse_down_position_x and self.mouse_x_old < self.mouse_now_position_x):
            self.frm.eye_phi -= 0.03
            self.frm.update_view()
            self.mouse_x_old = self.mouse_now_position_x

        if (self.mouse_x_old > self.mouse_now_position_x):
            self.frm.eye_phi += 0.03
            self.frm.update_view()
            self.mouse_x_old = self.mouse_now_position_x

        if (self.mouse_now_position_y > self.mouse_down_position_y and self.mouse_y_old < self.mouse_now_position_y):
            self.frm.eye_th -= 0.01
            self.frm.update_view()
            self.mouse_y_old = self.mouse_now_position_y
        if (self.mouse_y_old > self.mouse_now_position_y):
            self.frm.eye_th += 0.01
            self.frm.update_view()
            self.mouse_y_old = self.mouse_now_position_y

    def rb_mouse_down_left(self, event):
        self.mouse_down_position_x = event.x
        self.mouse_down_position_y = event.y
        self.mouse_x_old = event.x
        self.mouse_y_old = event.y

    def rb_mouse_move_mid(self, event):
        ''''''

    def rb_mouse_double_click_left(self, event):
        ''''''

    def rb_mouse_wheel(self, event):
        if (event.delta < 0):
            self.frm.eye_r -= 0.1
            self.frm.update_view()
        else:
            self.frm.eye_r += 0.1
            self.frm.update_view()

    def rb_mouse_up_left(self, event):
        ''''''

    def rb_mouse_down_mid(self, event):
        mouse_position = np.array([event.x, event.y, 1])
        # print(mouse_position)

    def rb_mouse_move_right(self, event):
        ''''''

    def rb_mouse_down_right(self, event):
        ''''''

    def rb_mouse_up_right(self, event):
        ''''''

    def operation_main(self):
        amp = 0.5
        freq = 0.5
        for i in range(self.robot_.n_dofs):
            self.robot_state_.joint_state_des[i].th = amp * np.sin(2.0 * np.pi * freq * self.frm.sim_time)
            self.robot_state_.joint_state_des[i].thd = 2.0 * np.pi * freq * amp * np.cos(2.0 * np.pi * freq * self.frm.sim_time)
            self.robot_state_.joint_state_des[i].thdd = -2.0 * np.pi * freq * 2.0 * np.pi * freq * amp * np.sin(2.0 * np.pi * freq * self.frm.sim_time)
            # print(self.robot_state_.joint_state_des[i].th)

    def key_handler(self, event):

        # print(event.char, event.keysym, event.keycode)

        if event.keysym == 'w':
            self.frm.eye_r -= 0.1
            self.frm.update_view()


        elif event.keysym == 's':
            self.frm.eye_r += 0.1
            self.frm.update_view()

        elif event.keysym == 'q':
            self.frm.eye_th += 0.01
            self.frm.update_view()

        elif event.keysym == 'e':
            self.frm.eye_th -= 0.01
            self.frm.update_view()

        elif event.keysym == 'a':
            self.frm.eye_phi -= 0.05
            self.frm.update_view()

        elif event.keysym == 'd':
            self.frm.eye_phi += 0.05
            self.frm.update_view()

        elif event.keysym == 't':
            self.bool_draw_wireframe.set(True)
            self.frm.graphics_options.toggle_draw_wireframe()

        elif event.keysym == 'v':
            self.frm.graphics_options.toggle_draw_motion_vectors()

        elif event.keysym == '0':
            self.frm.graphics_options.toggle_draw_all_joint_frames()

        elif event.keysym == '1':
            self.frm.graphics_options.toggle_draw_joint_frame(1)

        elif event.keysym == '2':
            self.frm.graphics_options.toggle_draw_joint_frame(2)

        elif event.keysym == '3':
            self.frm.graphics_options.toggle_draw_joint_frame(3)

        elif event.keysym == '4':
            self.frm.graphics_options.toggle_draw_joint_frame(4)

        elif event.keysym == '5':
            self.frm.graphics_options.toggle_draw_joint_frame(5)

        elif event.keysym == '6':
            self.frm.graphics_options.toggle_draw_joint_frame(6)

        elif event.keysym == '7':
            self.frm.graphics_options.toggle_draw_joint_frame(7)

        elif event.keysym == '8':
            self.frm.graphics_options.toggle_draw_joint_frame(8)

        elif event.keysym == '9':
            self.frm.graphics_options.toggle_draw_joint_frame(9)

        elif event.keysym == 'k':
            self.frm.controller_type = robot_defs.CONTROLLER_TYPE_PD

    #####################################################################################################################
    def edit_1_switch(self):
        if self.para_edit_1_is_on:
            self.edit_para_m1_2.config(bg='#AD8EE3')
            self.para_edit_1_is_on = False
        else:
            self.edit_para_m1_2.config(bg='gray')
            self.para_edit_1_is_on = True

    def edit_2_switch(self):
        if self.para_edit_2_is_on:
            self.edit_para_m3_4.config(bg='#AD8EE3')
            self.para_edit_2_is_on = False
        else:
            self.edit_para_m3_4.config(bg='gray')
            self.para_edit_2_is_on = True

    def edit_3_switch(self):
        if self.para_edit_3_is_on:
            self.edit_para_m5_6.config(bg='#AD8EE3')
            self.para_edit_3_is_on = False
        else:
            self.edit_para_m5_6.config(bg='gray')
            self.para_edit_3_is_on = True

    def save_para_1(self):

        payload = []
        float32_value = np.float32(self.pid_pp_1.get())
        nn0 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_1.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_1.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_1.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_1.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_1.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE1_SET, payload))
        payload = [3]
        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 1")

    def save_para_2(self):

        payload = []
        float32_value = np.float32(self.pid_pp_2.get())
        nn0 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_2.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_2.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_2.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_2.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_2.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE2_SET, payload))
        payload = [4]
        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 2")

    def save_para_3(self):

        payload = []
        float32_value = np.float32(self.pid_pp_3.get())
        nn0 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_3.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_3.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_3.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_3.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_3.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE1_SET, payload))
        payload = [3]
        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 3")

    def save_para_4(self):

        payload = []
        float32_value = np.float32(self.pid_pp_4.get())
        nn0 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_4.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_4.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_4.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_4.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_4.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE2_SET, payload))
        payload = [4]
        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 4")

    def save_para_5(self):

        payload = []
        float32_value = np.float32(self.pid_pp_5.get())
        nn0 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_5.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_5.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_5.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_5.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_5.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE1_SET, payload))
        payload = [3]
        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 5")

    def save_para_6(self):

        payload = []
        float32_value = np.float32(self.pid_pp_6.get())
        nn0 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_6.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_6.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_6.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_6.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_6.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE2_SET, payload))
        payload = [4]
        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 6")

    ################################################# MSP Protocol MOTOR 1 2 #####################################
    def refresh_serial_M1_2(self):
        global port_var_msp
        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var_msp.set('')  # 初期値を空に設定
        self.port_menu_msp["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu_msp["menu"].add_command(label=port_name, command=tk._setit(self.port_var_msp, port_name))

    def toggle_serial_connect_M1_2(self):

        if self.ser_msp.is_open:
            try:
                self.ser_msp.close()
                self.baudrate_menu_msp.config(state="normal")
                self.delimiter_menu_msp.config(state="normal")
                self.toggle_button_msp.config(text="接続", background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp.get())
            try:
                #######################################
                self.ser_msp = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp.config(state="disabled")
                self.delimiter_menu_msp.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp = threading.Thread(target=self.Read_data_Serial_M1_2)
                self.com_thread_msp.daemon = True
                self.com_thread_msp.start()
                self.toggle_button_msp.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    def READ32_M1_2(self):
        global p
        x = (inBuf[p] & 0xff) + ((inBuf[p + 1] & 0xff) << 8) + ((inBuf[p + 2] & 0xff) << 16) + ((inBuf[p + 3] & 0xff) << 24)
        p = p + 4
        return x

    def READ16_M1_2(self):
        global p
        y = (inBuf[p] & 0xff) + ((inBuf[p + 1]) << 8)
        p = p + 2
        return y

    def READ8_M1_2(self):
        global p
        z = inBuf[p] & 0xff
        p = p + 1
        return z

    def request_ask_msp_M1_2(self, msp: int) -> Optional[List[int]]:
        return self.ask_msp_payload_M1_2(msp, None)

    def request_msp_multiple(self, msps: List[int]) -> List[int]:
        s = []
        for m in msps:
            s.extend(self.ask_msp_payload_M1_2(m, None))
        return s

    def ask_msp_payload_M1_2(self, msp: int, payload: Optional[List[chr]]) -> Optional[List[int]]:
        if msp < 0:
            return None
        bf = [ord(c) for c in MSP_HEADER]

        checksum = 0
        pl_size = (len(payload) if payload is not None else 0) & 0xFF
        bf.append(pl_size)
        checksum ^= pl_size

        bf.append(msp & 0xFF)
        checksum ^= (msp & 0xFF)

        if payload is not None:
            for c in payload:
                bf.append(c & 0xFF)
                checksum ^= (c & 0xFF)
        bf.append(checksum)
        return bf

    def ask_msp_M1_2(self, msp: List[int]):
        arr = bytearray(msp)
        # print(arr)
        self.ser_msp.write(arr)  # send the complete byte sequence in one go

    def Operation_MSP_M1_2(self, cmd, data_size):
        icmd = cmd & 0xFF

        if icmd == MSP_MOTOR_MODE:
            error = self.READ8_M1_2()
            run_mode_val = self.READ8_M1_2()
            cur_motor_number = self.READ8_M1_2()

            motor1_resolution = self.READ8_M1_2()
            motor2_resolution = self.READ8_M1_2()
            motor1_pole = self.READ8_M1_2()
            motor2_pole = self.READ8_M1_2()

            data_motor1 = self.READ8_M1_2()
            data_motor2 = self.READ8_M1_2()
            control_motion_1 = self.READ8_M1_2()
            control_motion_2 = self.READ8_M1_2()
            # print(data_motor1)
            if run_mode_val == 1:
                self.mode_run_var.set(MODE_RUN[0])
                self.mode_menu.config(background="#F5B639")
            if run_mode_val == 2:
                self.mode_run_var.set(MODE_RUN[1])
                self.mode_menu.config(background="#F5B639")
            if run_mode_val == 3:
                self.mode_run_var.set(MODE_RUN[2])
                self.mode_menu.config(background="#4DEA42")
            #
            if cur_motor_number == 1:
                self.num_motor_var.set(NUM_MOTOR[0])
                self.num_motor.config(background="#F5B639")
            if cur_motor_number == 2:
                self.num_motor_var.set(NUM_MOTOR[1])
                self.num_motor.config(background="#4DEA42")
            # print(data_motor2)
            if (data_motor1 == 5 or data_motor1 == 14 or error == 3):
                self.toggle_servo_1.config(text="SERVO_1(Off)", background="#FA3559")
            if (data_motor1 == 4):
                self.toggle_servo_1.config(text="SERVO_1(On)", background="#17FA2F")
            ###
            if (data_motor2 == 5 or data_motor2 == 14 or error == 4):
                self.toggle_servo_2.config(text="SERVO_2(Off)", background="#FA3559")
            if (data_motor2 == 4):
                self.toggle_servo_2.config(text="SERVO_2(On)", background="#17FA2F")

        elif icmd == MSP_MOTOR_POS:
            current_pos_11 = self.READ32_M1_2()
            current_pos_12 = self.READ32_M1_2()
            current_pos_21 = self.READ32_M1_2()
            current_pos_22 = self.READ32_M1_2()
            # print(self.uint_to_float(current_pos_11))
            # print(self.uint_to_float(current_pos_12))
            # print(self.uint_to_float(current_pos_21))
            # print(self.uint_to_float(current_pos_22))
        elif icmd == MSP_MOTOR_VEL:
            current_vel11 = self.READ16_M1_2()
            current_vel12 = self.READ16_M1_2()
            current_vel21 = self.READ16_M1_2()
            current_vel22 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_ACCEL_DECCEL:
            current_accel1 = self.READ16_M1_2()
            current_accel2 = self.READ16_M1_2()
            current_deccel1 = self.READ16_M1_2()
            current_deccel2 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_VOL:
            if self.para_edit_1_is_on == False:
                current_vol_in1 = self.READ32_M1_2()
                current_vol_in2 = self.READ32_M1_2()
                self.spinbox_VOL_IN1.set(self.uint_to_float(current_vol_in1))
                self.spinbox_VOL_IN2.set(self.uint_to_float(current_vol_in2))

        elif icmd == MSP_MOTOR_VOLLIMIT:
            if self.para_edit_1_is_on == False:
                current_vol_limit1 = self.READ32_M1_2()
                current_vol_limit2 = self.READ32_M1_2()
                self.spinbox_VOL_LIMIT1.set(self.uint_to_float(current_vol_limit1))
                self.spinbox_VOL_LIMIT2.set(self.uint_to_float(current_vol_limit2))

        elif icmd == MSP_MOTOR_PID_P:
            if self.para_edit_1_is_on == False:
                pid_pos_p1 = self.READ32_M1_2()
                pid_pos_p2 = self.READ32_M1_2()
                self.spinbox_PID_PP_1.set(self.uint_to_float(pid_pos_p1))
                self.spinbox_PID_PP_2.set(self.uint_to_float(pid_pos_p2))

        elif icmd == MSP_MOTOR_PID_VEL_P:
            if self.para_edit_1_is_on == False:
                pid_vel_p1 = self.READ32_M1_2()
                pid_vel_p2 = self.READ32_M1_2()
                self.spinbox_PID_VP_1.set(self.uint_to_float(pid_vel_p1))
                self.spinbox_PID_VP_2.set(self.uint_to_float(pid_vel_p2))

        elif icmd == MSP_MOTOR_PID_VEL_I:
            if self.para_edit_1_is_on == False:
                pid_vel_i1 = self.READ32_M1_2()
                pid_vel_i2 = self.READ32_M1_2()
                self.spinbox_PID_VI_1.set(self.uint_to_float(pid_vel_i1))
                self.spinbox_PID_VI_2.set(self.uint_to_float(pid_vel_i2))

        elif icmd == MSP_MOTOR_PID_VEL_D:
            if self.para_edit_1_is_on == False:
                pid_vel_d1 = self.READ32_M1_2()
                pid_vel_d2 = self.READ32_M1_2()
                self.spinbox_PID_VD_1.set(self.uint_to_float(pid_vel_d1))
                self.spinbox_PID_VD_2.set(self.uint_to_float(pid_vel_d2))

        elif icmd == MSP_MOTOR_FB_ICHI:
            position_now_motor1 = self.READ32_M1_2()
            position_now_motor2 = self.READ32_M1_2()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M1:
            now_velocity_1 = self.READ16_M1_2()
            now_acceleration_1 = self.READ16_M1_2()
            now_decceleration_1 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M2:
            now_velocity_2 = self.READ16_M1_2()
            now_acceleration_2 = self.READ16_M1_2()
            now_decceleration_2 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_RUN_SET:
            pass

        else:
            print("Don't know how to handle reply " + str(icmd))

    def Operation_ASK_MSP_M1_2(self):
        global time2
        time0 = time.time() * 1000.0
        if ((time0 - time2) > 50):
            time2 = time0
            if (self.i_couter == 0):
                requests = {MSP_MOTOR_POS, MSP_MOTOR_VOL, MSP_MOTOR_PID_VEL_P}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 1):
                requests = {MSP_MOTOR_VEL, MSP_MOTOR_VOLLIMIT, MSP_MOTOR_PID_P}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 2):
                requests = {MSP_MOTOR_PID_VEL_D, MSP_MOTOR_MODE, MSP_MOTOR_ACCEL_DECCEL}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 3):
                requests = {MSP_MOTOR_PID_VEL_I}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 4):
                requests = {MSP_MOTOR_FB_ICHI}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 5):
                requests = {MSP_MOTOR_FB_VEL_ACC_DEC_M1, MSP_MOTOR_FB_VEL_ACC_DEC_M2}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            self.i_couter += 1
            if (self.i_couter > 5):
                self.i_couter = 0
            if self.communicate_on == True:
                self.MSP_id = self.after(100, self.Operation_ASK_MSP_M1_2)
            else:
                self.MSP_id = None

    def send_data_M1_2(self):
        data = self.send_msp_text.get()
        requests = {MSP_MOTOR_POS}  # MSP_MOTOR_MEOBIET,MSP_MOTOR_POS
        # print(self.request_msp_multiple(requests))
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ask_msp_M1_2(self.request_msp_multiple(requests))
            # self.ser.write((data + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%M:%S')
            self.Text_msp_recieved.insert(tk.END, f"[S>{timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.Text_msp_recieved.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))

    def Read_data_Serial_M1_2(self):
        # 設定したデリミタを取得
        global c_state, err_rcvd, dataSize, p, offset, checksum, cmd, inBuf
        delimiter = DELIMITERS[self.delimiter_var_msp.get()]
        buf = ''
        while self.ser_msp.is_open:
            try:

                data = self.ser_msp.read(1).decode("latin1")  # .decode("CP932")//not UTF-8
                c = int.from_bytes(data.encode('latin1'), byteorder='big', signed=False)
                # print(c)
                ###########
                if c_state == IDLE:
                    if (c == 36):
                        c_state = HEADER_START
                    else:
                        c_state = IDLE
                elif c_state == HEADER_START:
                    c_state = HEADER_M if c == 77 else IDLE

                elif c_state == HEADER_M:
                    if c == 62:
                        c_state = HEADER_ARROW
                    elif c == '33':  # not use
                        c_state = HEADER_ERR
                    else:
                        c_state = IDLE
                elif c_state in (HEADER_ARROW, HEADER_ERR):
                    err_rcvd = (c_state == HEADER_ERR)  # now we are expecting the payload size
                    dataSize = c & 0xFF
                    # reset index variables
                    p = 0
                    offset = 0
                    checksum = 0
                    checksum ^= (c & 0xFF)

                    # the command is to follow
                    c_state = HEADER_SIZE
                elif c_state == HEADER_SIZE:
                    cmd = c & 0xFF
                    checksum ^= (c & 0xFF)
                    c_state = HEADER_CMD
                elif c_state == HEADER_CMD and offset < dataSize:
                    checksum ^= (c & 0xFF)
                    inBuf[offset] = c & 0xFF
                    offset += 1

                elif c_state == HEADER_CMD and offset >= dataSize:
                    # compare calculated and transferred checksum
                    if (checksum & 0xFF) == (c & 0xFF):
                        if err_rcvd:
                            print(f"Copter did not understand request type {c}")
                        else:
                            # we got a valid response packet, evaluate it
                            # print(cmd)
                            self.Operation_MSP_M1_2(cmd, int(dataSize))
                    else:
                        print(f"invalid checksum for command {cmd & 0xFF}: {checksum & 0xFF} expected, got {c & 0xFF}")
                        print(f"<{cmd & 0xFF} {dataSize & 0xFF}> {{", end=' ')
                        for i in range(dataSize):
                            if i != 0:
                                print(' ', end='')
                            print(inBuf[i] & 0xFF, end='')
                        print("} [", c, "]")
                        print(''.join(map(chr, inBuf[:dataSize])))
                    c_state = IDLE

                #######
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.Text_msp_recieved.see("end")
                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))

    def autoconnect_msp_comport_M1_2(self):
        self.port_var_msp.set('COM9')
        if self.ser_msp.is_open:
            try:
                self.ser_msp.close()
                self.baudrate_menu_msp.config(state="normal")
                self.delimiter_menu_msp.config(state="normal")
                self.toggle_button_msp.config(text="接続")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp.get()
            if not port:
                messagebox.showerror("接続エラー", "MSPシリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp.get())
            try:
                #######################################
                self.ser_msp = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp.config(state="disabled")
                self.delimiter_menu_msp.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp = threading.Thread(target=self.Read_data_Serial_M1_2)
                self.com_thread_msp.daemon = True
                self.com_thread_msp.start()
                self.toggle_button_msp.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    ################################################# MSP Protocol MOTOR 3 4 #####################################
    def refresh_serial_M3_4(self):

        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var_msp_1.set('')  # 初期値を空に設定
        self.port_menu_msp_1["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu_msp_1["menu"].add_command(label=port_name, command=tk._setit(self.port_var_msp_1, port_name))

    def toggle_serial_connect_M3_4(self):

        if self.ser_msp_1.is_open:
            try:
                self.ser_msp_1.close()
                self.baudrate_menu_msp_1.config(state="normal")
                self.delimiter_menu_msp_1.config(state="normal")
                self.toggle_button_msp_1.config(text="接続", background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_1.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_1.get())
            try:
                #######################################
                self.ser_msp_1 = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_1.config(state="disabled")
                self.delimiter_menu_msp_1.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_1 = threading.Thread(target=self.Read_data_Serial_M3_4)
                self.com_thread_msp_1.daemon = True
                self.com_thread_msp_1.start()
                self.toggle_button_msp_1.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    def READ32_M3_4(self):
        global p
        x = (inBuf[p] & 0xff) + ((inBuf[p + 1] & 0xff) << 8) + ((inBuf[p + 2] & 0xff) << 16) + ((inBuf[p + 3] & 0xff) << 24)
        p = p + 4
        return x

    def READ16_M3_4(self):
        global p
        y = (inBuf[p] & 0xff) + ((inBuf[p + 1]) << 8)
        p = p + 2
        return y

    def READ8_M3_4(self):
        global p
        z = inBuf[p] & 0xff
        p = p + 1
        return z

    def request_ask_msp_M3_4(self, msp: int) -> Optional[List[int]]:
        return self.ask_msp_payload_M3_4(msp, None)

    def request_multiple_M3_4(self, msps: List[int]) -> List[int]:
        s = []
        for m in msps:
            s.extend(self.ask_msp_payload_M3_4(m, None))
        return s

    def ask_msp_payload_M3_4(self, msp: int, payload: Optional[List[chr]]) -> Optional[List[int]]:
        if msp < 0:
            return None
        bf = [ord(c) for c in MSP_HEADER]

        checksum = 0
        pl_size = (len(payload) if payload is not None else 0) & 0xFF
        bf.append(pl_size)
        checksum ^= pl_size

        bf.append(msp & 0xFF)
        checksum ^= (msp & 0xFF)

        if payload is not None:
            for c in payload:
                bf.append(c & 0xFF)
                checksum ^= (c & 0xFF)
        bf.append(checksum)
        return bf

    def ask_msp_M3_4(self, msp: List[int]):
        arr = bytearray(msp)
        # print(arr)
        self.ser_msp_1.write(arr)  # send the complete byte sequence in one go

    def Operation_MSP_M3_4(self, cmd, data_size):
        icmd = cmd & 0xFF

        if icmd == MSP_MOTOR_MODE:
            error = self.READ8_M3_4()
            run_mode_val = self.READ8_M3_4()
            cur_motor_number = self.READ8_M3_4()

            motor1_resolution = self.READ8_M3_4()
            motor2_resolution = self.READ8_M3_4()
            motor1_pole = self.READ8_M3_4()
            motor2_pole = self.READ8_M3_4()

            data_motor1 = self.READ8_M3_4()
            data_motor2 = self.READ8_M3_4()
            control_motion_1 = self.READ8_M3_4()
            control_motion_2 = self.READ8_M3_4()
            # print(data_motor1)
            if run_mode_val == 1:
                self.mode_run_var_1.set(MODE_RUN[0])
                self.mode_menu_1.config(background="#F5B639")
            if run_mode_val == 2:
                self.mode_run_var_1.set(MODE_RUN[1])
                self.mode_menu_1.config(background="#F5B639")
            if run_mode_val == 3:
                self.mode_run_var_1.set(MODE_RUN[2])
                self.mode_menu_1.config(background="#4DEA42")
            #
            if cur_motor_number == 1:
                self.num_motor_var_1.set(NUM_MOTOR[0])
                self.num_motor_1.config(background="#F5B639")
            if cur_motor_number == 2:
                self.num_motor_var_1.set(NUM_MOTOR[1])
                self.num_motor_1.config(background="#4DEA42")
            # print(data_motor2)
            if (data_motor1 == 5 or data_motor1 == 14 or error == 3):
                self.toggle_servo_3.config(text="SERVO_3(Off)", background="#FA3559")
            if (data_motor1 == 4):
                self.toggle_servo_3.config(text="SERVO_3(On)", background="#17FA2F")
            ###
            if (data_motor2 == 5 or data_motor2 == 14 or error == 4):
                self.toggle_servo_4.config(text="SERVO_4(Off)", background="#FA3559")
            if (data_motor2 == 4):
                self.toggle_servo_4.config(text="SERVO_4(On)", background="#17FA2F")

        elif icmd == MSP_MOTOR_POS:
            current_pos_11 = self.READ32_M3_4()
            current_pos_12 = self.READ32_M3_4()
            current_pos_21 = self.READ32_M3_4()
            current_pos_22 = self.READ32_M3_4()
            # print(self.uint_to_float(current_pos_11))
            # print(self.uint_to_float(current_pos_12))
            # print(self.uint_to_float(current_pos_21))
            # print(self.uint_to_float(current_pos_22))
        elif icmd == MSP_MOTOR_VEL:
            current_vel11 = self.READ16_M3_4()
            current_vel12 = self.READ16_M3_4()
            current_vel21 = self.READ16_M3_4()
            current_vel22 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_ACCEL_DECCEL:
            current_accel1 = self.READ16_M3_4()
            current_accel2 = self.READ16_M3_4()
            current_deccel1 = self.READ16_M3_4()
            current_deccel2 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_VOL:
            if self.para_edit_2_is_on == False:
                current_vol_in1 = self.READ32_M3_4()
                current_vol_in2 = self.READ32_M3_4()
                self.spinbox_VOL_IN3.set(self.uint_to_float(current_vol_in1))
                self.spinbox_VOL_IN4.set(self.uint_to_float(current_vol_in2))

        elif icmd == MSP_MOTOR_VOLLIMIT:
            if self.para_edit_2_is_on == False:
                current_vol_limit1 = self.READ32_M3_4()
                current_vol_limit2 = self.READ32_M3_4()
                self.spinbox_VOL_LIMIT3.set(self.uint_to_float(current_vol_limit1))
                self.spinbox_VOL_LIMIT4.set(self.uint_to_float(current_vol_limit2))

        elif icmd == MSP_MOTOR_PID_P:
            if self.para_edit_2_is_on == False:
                pid_pos_p1 = self.READ32_M3_4()
                pid_pos_p2 = self.READ32_M3_4()
                self.spinbox_PID_PP_3.set(self.uint_to_float(pid_pos_p1))
                self.spinbox_PID_PP_4.set(self.uint_to_float(pid_pos_p2))

        elif icmd == MSP_MOTOR_PID_VEL_P:
            if self.para_edit_2_is_on == False:
                pid_vel_p1 = self.READ32_M3_4()
                pid_vel_p2 = self.READ32_M3_4()
                self.spinbox_PID_VP_3.set(self.uint_to_float(pid_vel_p1))
                self.spinbox_PID_VP_4.set(self.uint_to_float(pid_vel_p2))

        elif icmd == MSP_MOTOR_PID_VEL_I:
            if self.para_edit_2_is_on == False:
                pid_vel_i1 = self.READ32_M3_4()
                pid_vel_i2 = self.READ32_M3_4()
                self.spinbox_PID_VI_3.set(self.uint_to_float(pid_vel_i1))
                self.spinbox_PID_VI_4.set(self.uint_to_float(pid_vel_i2))

        elif icmd == MSP_MOTOR_PID_VEL_D:
            if self.para_edit_2_is_on == False:
                pid_vel_d1 = self.READ32_M3_4()
                pid_vel_d2 = self.READ32_M3_4()
                self.spinbox_PID_VD_3.set(self.uint_to_float(pid_vel_d1))
                self.spinbox_PID_VD_4.set(self.uint_to_float(pid_vel_d2))

        elif icmd == MSP_MOTOR_FB_ICHI:
            position_now_motor1 = self.READ32_M3_4()
            position_now_motor2 = self.READ32_M3_4()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M1:
            now_velocity_1 = self.READ16_M3_4()
            now_acceleration_1 = self.READ16_M3_4()
            now_decceleration_1 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M2:
            now_velocity_2 = self.READ16_M3_4()
            now_acceleration_2 = self.READ16_M3_4()
            now_decceleration_2 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_RUN_SET:
            pass

        else:
            print("Don't know how to handle reply " + str(icmd))

    def Operation_ASK_MSP_M3_4(self):
        global time3
        time1 = time.time() * 1000.0
        if ((time1 - time3) > 50):
            time3 = time1
            if (self.i_couter_1 == 0):
                requests = {MSP_MOTOR_POS, MSP_MOTOR_VOL, MSP_MOTOR_PID_VEL_P}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 1):
                requests = {MSP_MOTOR_VEL, MSP_MOTOR_VOLLIMIT, MSP_MOTOR_PID_P}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 2):
                requests = {MSP_MOTOR_PID_VEL_D, MSP_MOTOR_MODE, MSP_MOTOR_ACCEL_DECCEL}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 3):
                requests = {MSP_MOTOR_PID_VEL_I}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 4):
                requests = {MSP_MOTOR_FB_ICHI}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 5):
                requests = {MSP_MOTOR_FB_VEL_ACC_DEC_M1, MSP_MOTOR_FB_VEL_ACC_DEC_M2}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            self.i_couter_1 += 1
            if (self.i_couter_1 > 5):
                self.i_couter_1 = 0
            if self.communicate_on == True:
                self.MSP_id_1 = self.after(100, self.Operation_ASK_MSP_M3_4)
            else:
                self.MSP_id_1 = None

    def send_data_M3_4(self):
        data = self.send_msp_text.get()
        requests = {MSP_MOTOR_POS}  # MSP_MOTOR_MEOBIET,MSP_MOTOR_POS
        # print(self.request_msp_multiple(requests))
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ask_msp_M3_4(self.request_multiple_M3_4(requests))
            # self.ser.write((data + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%M:%S')
            self.Text_msp_recieved.insert(tk.END, f"[S>{timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.Text_msp_recieved.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))

    def Read_data_Serial_M3_4(self):
        # 設定したデリミタを取得
        global c_state, err_rcvd, dataSize, p, offset, checksum, cmd, inBuf
        delimiter = DELIMITERS[self.delimiter_var_msp.get()]
        buf = ''
        while self.ser_msp_1.is_open:
            try:

                data = self.ser_msp_1.read(1).decode("latin1")  # .decode("CP932")//not UTF-8
                c = int.from_bytes(data.encode('latin1'), byteorder='big', signed=False)
                # print(c)
                ###########
                if c_state == IDLE:
                    if (c == 36):
                        c_state = HEADER_START
                    else:
                        c_state = IDLE
                elif c_state == HEADER_START:
                    c_state = HEADER_M if c == 77 else IDLE

                elif c_state == HEADER_M:
                    if c == 62:
                        c_state = HEADER_ARROW
                    elif c == '33':  # not use
                        c_state = HEADER_ERR
                    else:
                        c_state = IDLE
                elif c_state in (HEADER_ARROW, HEADER_ERR):
                    err_rcvd = (c_state == HEADER_ERR)  # now we are expecting the payload size
                    dataSize = c & 0xFF
                    # reset index variables
                    p = 0
                    offset = 0
                    checksum = 0
                    checksum ^= (c & 0xFF)

                    # the command is to follow
                    c_state = HEADER_SIZE
                elif c_state == HEADER_SIZE:
                    cmd = c & 0xFF
                    checksum ^= (c & 0xFF)
                    c_state = HEADER_CMD
                elif c_state == HEADER_CMD and offset < dataSize:
                    checksum ^= (c & 0xFF)
                    inBuf[offset] = c & 0xFF
                    offset += 1

                elif c_state == HEADER_CMD and offset >= dataSize:
                    # compare calculated and transferred checksum
                    if (checksum & 0xFF) == (c & 0xFF):
                        if err_rcvd:
                            print(f"Copter did not understand request type {c}")
                        else:
                            # we got a valid response packet, evaluate it
                            # print(cmd)
                            self.Operation_MSP_M3_4(cmd, int(dataSize))
                    else:
                        print(f"invalid checksum for command {cmd & 0xFF}: {checksum & 0xFF} expected, got {c & 0xFF}")
                        print(f"<{cmd & 0xFF} {dataSize & 0xFF}> {{", end=' ')
                        for i in range(dataSize):
                            if i != 0:
                                print(' ', end='')
                            print(inBuf[i] & 0xFF, end='')
                        print("} [", c, "]")
                        print(''.join(map(chr, inBuf[:dataSize])))
                    c_state = IDLE

                #######
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.Text_msp_recieved.see("end")
                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))

    def autoconnect_msp_comport_M3_4(self):
        self.port_var_msp_1.set('COM2')
        if self.ser_msp_1.is_open:
            try:
                self.ser_msp_1.close()
                self.baudrate_menu_msp_1.config(state="normal")
                self.delimiter_menu_msp_1.config(state="normal")
                self.toggle_button_msp_1.config(text="接続")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_1.get()
            if not port:
                messagebox.showerror("接続エラー", "MSPシリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_1.get())
            try:
                #######################################
                self.ser_msp_1 = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_1.config(state="disabled")
                self.delimiter_menu_msp_1.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_1 = threading.Thread(target=self.Read_data_Serial_M3_4)
                self.com_thread_msp_1.daemon = True
                self.com_thread_msp_1.start()
                self.toggle_button_msp_1.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
        #################################################MSP Protocol MOTOR 1 2 #####################################

        ################################################# MSP Protocol MOTOR 5 6 #####################################

    ################################################# MSP Protocol MOTOR 5 6 #####################################
    def refresh_serial_M5_6(self):

        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var_msp_2.set('')  # 初期値を空に設定
        self.port_menu_msp_2["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu_msp_2["menu"].add_command(label=port_name, command=tk._setit(self.port_var_msp_2, port_name))

    def toggle_serial_connect_M5_6(self):

        if self.ser_msp_2.is_open:
            try:
                self.ser_msp_2.close()
                self.baudrate_menu_msp_2.config(state="normal")
                self.delimiter_menu_msp_2.config(state="normal")
                self.toggle_button_msp_2.config(text="接続", background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_2.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_2.get())
            try:
                #######################################
                self.ser_msp_2 = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_2.config(state="disabled")
                self.delimiter_menu_msp_2.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_2 = threading.Thread(target=self.Read_data_Serial_M5_6)
                self.com_thread_msp_2.daemon = True
                self.com_thread_msp_2.start()
                self.toggle_button_msp_2.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    def READ32_M5_6(self):
        global p
        x = (inBuf[p] & 0xff) + ((inBuf[p + 1] & 0xff) << 8) + ((inBuf[p + 2] & 0xff) << 16) + ((inBuf[p + 3] & 0xff) << 24)
        p = p + 4
        return x

    def READ16_M5_6(self):
        global p
        y = (inBuf[p] & 0xff) + ((inBuf[p + 1]) << 8)
        p = p + 2
        return y

    def READ8_M5_6(self):
        global p
        z = inBuf[p] & 0xff
        p = p + 1
        return z

    def request_ask_msp_M5_6(self, msp: int) -> Optional[List[int]]:
        return self.ask_msp_payload_M5_6(msp, None)

    def request_multiple_M5_6(self, msps: List[int]) -> List[int]:
        s = []
        for m in msps:
            s.extend(self.ask_msp_payload_M5_6(m, None))
        return s

    def ask_msp_payload_M5_6(self, msp: int, payload: Optional[List[chr]]) -> Optional[List[int]]:
        if msp < 0:
            return None
        bf = [ord(c) for c in MSP_HEADER]

        checksum = 0
        pl_size = (len(payload) if payload is not None else 0) & 0xFF
        bf.append(pl_size)
        checksum ^= pl_size

        bf.append(msp & 0xFF)
        checksum ^= (msp & 0xFF)

        if payload is not None:
            for c in payload:
                bf.append(c & 0xFF)
                checksum ^= (c & 0xFF)
        bf.append(checksum)
        return bf

    def ask_msp_M5_6(self, msp: List[int]):
        arr = bytearray(msp)
        # print(arr)
        self.ser_msp_2.write(arr)  # send the complete byte sequence in one go

    def Operation_MSP_M5_6(self, cmd, data_size):
        icmd = cmd & 0xFF

        if icmd == MSP_MOTOR_MODE:
            error = self.READ8_M5_6()
            run_mode_val = self.READ8_M5_6()
            cur_motor_number = self.READ8_M5_6()

            motor1_resolution = self.READ8_M5_6()
            motor2_resolution = self.READ8_M5_6()
            motor1_pole = self.READ8_M5_6()
            motor2_pole = self.READ8_M5_6()

            data_motor1 = self.READ8_M5_6()
            data_motor2 = self.READ8_M5_6()
            control_motion_1 = self.READ8_M5_6()
            control_motion_2 = self.READ8_M5_6()
            # print(data_motor1)
            if run_mode_val == 1:
                self.mode_run_var_2.set(MODE_RUN[0])
                self.mode_menu_2.config(background="#F5B639")
            if run_mode_val == 2:
                self.mode_run_var_2.set(MODE_RUN[1])
                self.mode_menu_2.config(background="#F5B639")
            if run_mode_val == 3:
                self.mode_run_var_2.set(MODE_RUN[2])
                self.mode_menu_2.config(background="#4DEA42")
            #
            if cur_motor_number == 1:
                self.num_motor_var_2.set(NUM_MOTOR[0])
                self.num_motor_2.config(background="#F5B639")
            if cur_motor_number == 2:
                self.num_motor_var_2.set(NUM_MOTOR[1])
                self.num_motor_2.config(background="#4DEA42")
            # print(data_motor2)
            if (data_motor1 == 5 or data_motor1 == 14 or error == 3):
                self.toggle_servo_5.config(text="SERVO_5(Off)", background="#FA3559")
            if (data_motor1 == 4):
                self.toggle_servo_5.config(text="SERVO_5(On)", background="#17FA2F")
            ###
            if (data_motor2 == 5 or data_motor2 == 14 or error == 4):
                self.toggle_servo_6.config(text="SERVO_6(Off)", background="#FA3559")
            if (data_motor2 == 4):
                self.toggle_servo_6.config(text="SERVO_6(On)", background="#17FA2F")

        elif icmd == MSP_MOTOR_POS:
            current_pos_11 = self.READ32_M5_6()
            current_pos_12 = self.READ32_M5_6()
            current_pos_21 = self.READ32_M5_6()
            current_pos_22 = self.READ32_M5_6()
            # print(self.uint_to_float(current_pos_11))
            # print(self.uint_to_float(current_pos_12))
            # print(self.uint_to_float(current_pos_21))
            # print(self.uint_to_float(current_pos_22))
        elif icmd == MSP_MOTOR_VEL:
            current_vel11 = self.READ16_M5_6()
            current_vel12 = self.READ16_M5_6()
            current_vel21 = self.READ16_M5_6()
            current_vel22 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_ACCEL_DECCEL:
            current_accel1 = self.READ16_M5_6()
            current_accel2 = self.READ16_M5_6()
            current_deccel1 = self.READ16_M5_6()
            current_deccel2 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_VOL:
            if self.para_edit_3_is_on == False:
                current_vol_in1 = self.READ32_M5_6()
                current_vol_in2 = self.READ32_M5_6()
                self.spinbox_VOL_IN5.set(self.uint_to_float(current_vol_in1))
                self.spinbox_VOL_IN6.set(self.uint_to_float(current_vol_in2))

        elif icmd == MSP_MOTOR_VOLLIMIT:
            if self.para_edit_3_is_on == False:
                current_vol_limit1 = self.READ32_M5_6()
                current_vol_limit2 = self.READ32_M5_6()
                self.spinbox_VOL_LIMIT5.set(self.uint_to_float(current_vol_limit1))
                self.spinbox_VOL_LIMIT6.set(self.uint_to_float(current_vol_limit2))

        elif icmd == MSP_MOTOR_PID_P:
            if self.para_edit_3_is_on == False:
                pid_pos_p1 = self.READ32_M5_6()
                pid_pos_p2 = self.READ32_M5_6()
                self.spinbox_PID_PP_5.set(self.uint_to_float(pid_pos_p1))
                self.spinbox_PID_PP_6.set(self.uint_to_float(pid_pos_p2))

        elif icmd == MSP_MOTOR_PID_VEL_P:
            if self.para_edit_3_is_on == False:
                pid_vel_p1 = self.READ32_M5_6()
                pid_vel_p2 = self.READ32_M5_6()
                self.spinbox_PID_VP_5.set(self.uint_to_float(pid_vel_p1))
                self.spinbox_PID_VP_6.set(self.uint_to_float(pid_vel_p2))

        elif icmd == MSP_MOTOR_PID_VEL_I:
            if self.para_edit_3_is_on == False:
                pid_vel_i1 = self.READ32_M5_6()
                pid_vel_i2 = self.READ32_M5_6()
                self.spinbox_PID_VI_5.set(self.uint_to_float(pid_vel_i1))
                self.spinbox_PID_VI_6.set(self.uint_to_float(pid_vel_i2))

        elif icmd == MSP_MOTOR_PID_VEL_D:
            if self.para_edit_3_is_on == False:
                pid_vel_d1 = self.READ32_M5_6()
                pid_vel_d2 = self.READ32_M5_6()
                self.spinbox_PID_VD_5.set(self.uint_to_float(pid_vel_d1))
                self.spinbox_PID_VD_6.set(self.uint_to_float(pid_vel_d2))

        elif icmd == MSP_MOTOR_FB_ICHI:
            position_now_motor1 = self.READ32_M5_6()
            position_now_motor2 = self.READ32_M5_6()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M1:
            now_velocity_1 = self.READ16_M5_6()
            now_acceleration_1 = self.READ16_M5_6()
            now_decceleration_1 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M2:
            now_velocity_2 = self.READ16_M5_6()
            now_acceleration_2 = self.READ16_M5_6()
            now_decceleration_2 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_RUN_SET:
            pass

        else:
            print("Don't know how to handle reply " + str(icmd))

    def Operation_ASK_MSP_M5_6(self):
        global time4
        time5 = time.time() * 1000.0
        if ((time5 - time4) > 50):
            time4 = time5
            if (self.i_couter_2 == 0):
                requests = {MSP_MOTOR_POS, MSP_MOTOR_VOL, MSP_MOTOR_PID_VEL_P}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 1):
                requests = {MSP_MOTOR_VEL, MSP_MOTOR_VOLLIMIT, MSP_MOTOR_PID_P}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 2):
                requests = {MSP_MOTOR_PID_VEL_D, MSP_MOTOR_MODE, MSP_MOTOR_ACCEL_DECCEL}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 3):
                requests = {MSP_MOTOR_PID_VEL_I}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 4):
                requests = {MSP_MOTOR_FB_ICHI}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 5):
                requests = {MSP_MOTOR_FB_VEL_ACC_DEC_M1, MSP_MOTOR_FB_VEL_ACC_DEC_M2}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            self.i_couter_2 += 1
            if (self.i_couter_2 > 5):
                self.i_couter_2 = 0
            if self.communicate_on == True:
                self.MSP_id_2 = self.after(100, self.Operation_ASK_MSP_M5_6)
            else:
                self.MSP_id_2 = None

    def send_data_M5_6(self):
        data = self.send_msp_text.get()
        requests = {MSP_MOTOR_POS}  # MSP_MOTOR_MEOBIET,MSP_MOTOR_POS
        # print(self.request_msp_multiple(requests))
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ask_msp_M5_6(self.request_multiple_M5_6(requests))
            # self.ser.write((data + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%M:%S')
            self.Text_msp_recieved.insert(tk.END, f"[S>{timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.Text_msp_recieved.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))

    def Read_data_Serial_M5_6(self):
        # 設定したデリミタを取得
        global c_state, err_rcvd, dataSize, p, offset, checksum, cmd, inBuf
        delimiter = DELIMITERS[self.delimiter_var_msp.get()]
        buf = ''
        while self.ser_msp_2.is_open:
            try:

                data = self.ser_msp_2.read(1).decode("latin1")  # .decode("CP932")//not UTF-8
                c = int.from_bytes(data.encode('latin1'), byteorder='big', signed=False)
                # print(c)
                ###########
                if c_state == IDLE:
                    if (c == 36):
                        c_state = HEADER_START
                    else:
                        c_state = IDLE
                elif c_state == HEADER_START:
                    c_state = HEADER_M if c == 77 else IDLE

                elif c_state == HEADER_M:
                    if c == 62:
                        c_state = HEADER_ARROW
                    elif c == '33':  # not use
                        c_state = HEADER_ERR
                    else:
                        c_state = IDLE
                elif c_state in (HEADER_ARROW, HEADER_ERR):
                    err_rcvd = (c_state == HEADER_ERR)  # now we are expecting the payload size
                    dataSize = c & 0xFF
                    # reset index variables
                    p = 0
                    offset = 0
                    checksum = 0
                    checksum ^= (c & 0xFF)

                    # the command is to follow
                    c_state = HEADER_SIZE
                elif c_state == HEADER_SIZE:
                    cmd = c & 0xFF
                    checksum ^= (c & 0xFF)
                    c_state = HEADER_CMD
                elif c_state == HEADER_CMD and offset < dataSize:
                    checksum ^= (c & 0xFF)
                    inBuf[offset] = c & 0xFF
                    offset += 1

                elif c_state == HEADER_CMD and offset >= dataSize:
                    # compare calculated and transferred checksum
                    if (checksum & 0xFF) == (c & 0xFF):
                        if err_rcvd:
                            print(f"Copter did not understand request type {c}")
                        else:
                            # we got a valid response packet, evaluate it
                            # print(cmd)
                            self.Operation_MSP_M5_6(cmd, int(dataSize))
                    else:
                        print(f"invalid checksum for command {cmd & 0xFF}: {checksum & 0xFF} expected, got {c & 0xFF}")
                        print(f"<{cmd & 0xFF} {dataSize & 0xFF}> {{", end=' ')
                        for i in range(dataSize):
                            if i != 0:
                                print(' ', end='')
                            print(inBuf[i] & 0xFF, end='')
                        print("} [", c, "]")
                        print(''.join(map(chr, inBuf[:dataSize])))
                    c_state = IDLE

                #######
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.Text_msp_recieved.see("end")
                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))

    def autoconnect_msp_comport_M5_6(self):
        self.port_var_msp_2.set('COM1')
        if self.ser_msp_2.is_open:
            try:
                self.ser_msp_2.close()
                self.baudrate_menu_msp_2.config(state="normal")
                self.delimiter_menu_msp_2.config(state="normal")
                self.toggle_button_msp_2.config(text="接続")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_2.get()
            if not port:
                messagebox.showerror("接続エラー", "MSPシリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_2.get())
            try:
                #######################################
                self.ser_msp_2 = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_2.config(state="disabled")
                self.delimiter_menu_msp_2.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_2 = threading.Thread(target=self.Read_data_Serial_M5_6)
                self.com_thread_msp_2.daemon = True
                self.com_thread_msp_2.start()
                self.toggle_button_msp_2.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
        #################################################MSP Protocol MOTOR 1 2 #####################################

    #################################################Main#####################################
    def Communication_serial_run(self):
        self.communicate_on = True
        self.MSP_id = self.after(100, self.Operation_ASK_MSP_M1_2)

    def Communication_serial_stop(self):
        self.communicate_on = False
        self.MSP_id = None

    ############################################################################################
    def riru_number_change(self):
        data = str(self.riru_henko.get())
        self.Socket_send("WR DM600 " + data)

    def reset_kikai(self):
        self.data_old = ''
        self.data_old1 = ''
        self.data_old2 = ''
        self.data_old_a = ''
        self.data_old_b = ''
        self.data_old_c = ''

    def add_fiter_to(self):
        self.so_fiter = self.so_fiter + 1
        if self.so_fiter == 2:
            self.filterchoosen1.place(x=5, y=40)
        if self.so_fiter == 3:
            self.filterchoosen1.place(x=5, y=40)
            self.filterchoosen2.place(x=5, y=80)
        if self.so_fiter == 4:
            self.filterchoosen1.place(x=5, y=40)
            self.filterchoosen2.place(x=5, y=80)
            self.filterchoosen3.place(x=5, y=120)
        if self.so_fiter == 5:
            self.filterchoosen1.place(x=5, y=40)
            self.filterchoosen2.place(x=5, y=80)
            self.filterchoosen3.place(x=5, y=120)
            self.filterchoosen4.place(x=5, y=160)
        if self.so_fiter >= 5:
            self.so_fiter = 5

    def minus_fiter_to(self):
        self.so_fiter = self.so_fiter - 1

        if self.so_fiter == 2:
            self.filterchoosen2.place(x=-100, y=-100)
        if self.so_fiter == 3:
            self.filterchoosen3.place(x=-100, y=-100)
        if self.so_fiter == 4:
            self.filterchoosen4.place(x=-100, y=-100)
        if self.so_fiter <= 1:
            self.filterchoosen1.place(x=-100, y=-100)
            self.so_fiter = 1

    def reset_button_on_click(self, event):
        self.Socket_send("ST R3003")

    def reset_button_off_click(self, event):
        self.Socket_send("RS R3003")

    def print_ryru_on_click(self, event):
        self.Socket_send("ST R9004")

    def print_ryru_off_click(self, event):
        self.Socket_send("RS R9004")

    def print_hako_on_click(self, event):
        self.Socket_send("ST R9005")

    def print_hako_off_click(self, event):
        self.Socket_send("RS R9005")

    def test_start_on_click(self, event):
        self.Socket_send("ST R4203")
        self.BAR_check.set('検査開始')
        self.barreceived_text.delete(1.0, tk.END)

    def test_start_off_click(self, event):
        self.Socket_send("RS R4203")

    def test_setok_on_click(self, event):
        self.Socket_send("ST R4204")

    def test_setok_off_click(self, event):
        self.Socket_send("RS R4204")

    def scan_start_on_click(self, event):
        self.Socket_send("ST R4205")

    def scan_start_off_click(self, event):
        self.Socket_send("RS R4205")

    def toridasi_ok_on_click(self, event):
        self.Socket_send("ST R4206")
        self.BAR_check.set('検査終わり')

    def toridasi_ok_off_click(self, event):
        self.Socket_send("RS R4206")

    def syudou_kirikae_switch(self):

        # Determine is on or off
        if self.syudou_kirikaeis_on:
            # self.syudou_kirikae.config(wraplength=90,text="手動モード",bg='yellow')
            self.Socket_send("ST MR3313")
            self.syudou_kirikaeis_on = False
        else:
            # self.syudou_kirikae.config(wraplength=90, text="手動モード", bg='yellow')
            # self.syudou_kirikae.config(wraplength=90, text="自動モード", height=2, width=9, bg='lightgreen',)
            self.Socket_send("RS MR3313")
            self.syudou_kirikaeis_on = True

    #####################
    def slider_DM1000(self, a):
        data = str(a)
        self.Socket_send("WR DM1000 " + data)

    def slider_DM1002(self, a):
        data = str(a)
        self.Socket_send("WR DM1002 " + data)

    def slider_DM1004(self, a):
        data = str(a)
        self.Socket_send("WR DM1004 " + data)

    def slider_DM1006(self, a):
        data = str(a)
        self.Socket_send("WR DM1006 " + data)

    def slider_DM1008(self, a):
        data = str(a)
        self.Socket_send("WR DM1008 " + data)

    def slider_DM1010(self, a):
        data = str(a)
        self.Socket_send("WR DM1010 " + data)

    def slider_DM1012(self, a):
        data = str(a)
        self.Socket_send("WR DM1012 " + data)

    def slider_DM1014(self, a):
        data = str(a)
        self.Socket_send("WR DM1014 " + data)

    def slider_DM1016(self, a):
        data = str(a)
        self.Socket_send("WR DM1016 " + data)

    def slider_DM1018(self, a):
        data = str(a)
        self.Socket_send("WR DM1018 " + data)

    def slider_DM1020(self, a):
        data = str(a)
        self.Socket_send("WR DM1020 " + data)

    #########################

    def gettab_select(self, event):
        note = event.widget
        if note.tab(note.select(), "text") == "メイン":
            self.tab_select = 0
        if note.tab(note.select(), "text") == "手動":
            self.tab_select = 1
        if note.tab(note.select(), "text") == "保守１":
            self.tab_select = 2
        if note.tab(note.select(), "text") == "保守２":
            self.tab_select = 3
        if note.tab(note.select(), "text") == "設定１":
            self.tab_select = 4
        if note.tab(note.select(), "text") == "設定２":
            self.tab_select = 5
        if note.tab(note.select(), "text") == "通信":
            self.tab_select = 6
        if note.tab(note.select(), "text") == "カメラ":
            self.tab_select = 7

    def gettab1(self):
        self.tab_select = 1
        ''''''

    def gettab2(self):
        self.tab_select = 2
        ''''''

    def gettab3(self):
        self.tab_select = 3
        ''''''

    def gettab4(self):
        self.tab_select = 4
        ''''''

    def gettab5(self):
        self.tab_select = 5
        ''''''

    def gettab6(self):
        self.tab_select = 6
        ''''''

    def gettab7(self):
        self.tab_select = 7
        ''''''

    def seihin_ok(self):
        ''''''

    def clear_error_text(self):
        self.PLC_error_text.delete(1.0, tk.END)
        self.PLC_error_annai.delete(1.0, tk.END)

    ###########################################################
    def set_image_sita(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img_sita = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_sita(self.pil_img_sita.width, self.pil_img_sita.height)
        # 画像の表示
        self.draw_image_sita(self.pil_img_sita)

    def reset_transform_sita(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_sita = np.eye(3)  # 3x3の単位行列

    def translate_sita(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_sita = np.dot(mat, self.mat_affine_sita)

    def scale_sita(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_sita = np.dot(mat, self.mat_affine_sita)

    def scale_at_sita(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_sita(-cx, -cy)
        # 拡大縮小
        self.scale_sita(scale)
        # 元に戻す
        self.translate_sita(cx, cy)

    def zoom_fit_sita(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_sita.winfo_width()
        canvas_height = self.canvas_sita.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_sita()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_sita(scale)
        # あまり部分を中央に寄せる
        self.translate_sita(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------

    def draw_image_sita(self, pil_image):

        if pil_image == None:
            return

        self.canvas_sita.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_sita.winfo_width()
        canvas_height = self.canvas_sita.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_sita)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_sita = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_sita.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_sita  # 表示画像データ
        )

    ###########################################################
    def set_image_main(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_main(self.pil_img.width, self.pil_img.height)
        # 画像の表示
        self.draw_image_main(self.pil_img)

    def reset_transform_main(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_main = np.eye(3)  # 3x3の単位行列

    def translate_main(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_main = np.dot(mat, self.mat_affine_main)

    def scale_main(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_main = np.dot(mat, self.mat_affine_main)

    def scale_at_main(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_main(-cx, -cy)
        # 拡大縮小
        self.scale_main(scale)
        # 元に戻す
        self.translate_main(cx, cy)

    def zoom_fit_main(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_main()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_main(scale)
        # あまり部分を中央に寄せる
        self.translate_main(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------

    def draw_image_main(self, pil_image):

        if pil_image == None:
            return

        self.canvas.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_main)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_main = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_main  # 表示画像データ
        )

    ###########################################################
    def set_image_hoshu1(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img_hoshu1 = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_hoshu1(self.pil_img_hoshu1.width, self.pil_img_hoshu1.height)
        # 画像の表示
        self.draw_image_hoshu1(self.pil_img_hoshu1)

    def reset_transform_hoshu1(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_hoshu1 = np.eye(3)  # 3x3の単位行列

    def translate_hoshu1(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_hoshu1 = np.dot(mat, self.mat_affine_hoshu1)

    def scale_hoshu1(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_hoshu1 = np.dot(mat, self.mat_affine_hoshu1)

    def scale_at_hoshu1(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_hoshu1(-cx, -cy)
        # 拡大縮小
        self.scale_hoshu1(scale)
        # 元に戻す
        self.translate_hoshu1(cx, cy)

    def zoom_fit_hoshu1(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu1.winfo_width()
        canvas_height = self.canvas_hoshu1.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_hoshu1()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_hoshu1(scale)
        # あまり部分を中央に寄せる
        self.translate_hoshu1(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------

    def draw_image_hoshu1(self, pil_image):

        if pil_image == None:
            return

        self.canvas_hoshu1.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu1.winfo_width()
        canvas_height = self.canvas_hoshu1.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_hoshu1)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_hoshu1 = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_hoshu1.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_hoshu1  # 表示画像データ
        )

    ######################################################################
    def set_image_hoshu2(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img_hoshu2 = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_hoshu2(self.pil_img_hoshu2.width, self.pil_img_hoshu2.height)
        # 画像の表示
        self.draw_image_hoshu2(self.pil_img_hoshu2)

    def reset_transform_hoshu2(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_hoshu2 = np.eye(3)  # 3x3の単位行列

    def translate_hoshu2(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_hoshu2 = np.dot(mat, self.mat_affine_hoshu2)

    def scale_hoshu2(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_hoshu2 = np.dot(mat, self.mat_affine_hoshu2)

    def scale_at_hoshu2(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_hoshu2(-cx, -cy)
        # 拡大縮小
        self.scale_hoshu2(scale)
        # 元に戻す
        self.translate_hoshu2(cx, cy)

    def zoom_fit_hoshu2(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu2.winfo_width()
        canvas_height = self.canvas_hoshu2.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_hoshu2()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_hoshu2(scale)
        # あまり部分を中央に寄せる
        self.translate_hoshu2(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------

    def draw_image_hoshu2(self, pil_image):

        if pil_image == None:
            return

        self.canvas_hoshu2.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu2.winfo_width()
        canvas_height = self.canvas_hoshu2.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_hoshu2)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_hoshu2 = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_hoshu2.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_hoshu2  # 表示画像データ
        )

    ######################### SSEERRIIAALL##########################
    def refresh_serial_ports(self):
        global port_var
        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var.set('')  # 初期値を空に設定
        self.port_menu["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu["menu"].add_command(label=port_name, command=tk._setit(self.port_var, port_name))

    def autoconnect_comport(self):
        self.port_var.set('COM6')
        if self.ser.is_open:
            try:
                self.ser.close()
                self.baudrate_menu.config(state="normal")
                self.delimiter_menu.config(state="normal")
                self.toggle_button.config(text="接続", background="#E53333")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var.get())
            try:
                #######################################
                self.ser = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu.config(state="disabled")
                self.delimiter_menu.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread = threading.Thread(target=self.read_comport_data)
                self.com_thread.daemon = True
                self.com_thread.start()
                self.toggle_button.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    def toggle_serial_port(self):

        if self.ser.is_open:
            try:
                self.ser.close()
                self.baudrate_menu.config(state="normal")
                self.delimiter_menu.config(state="normal")
                self.toggle_button.config(text="接続", background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var.get())
            try:
                #######################################
                self.ser = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu.config(state="disabled")
                self.delimiter_menu.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread = threading.Thread(target=self.read_comport_data)
                self.com_thread.daemon = True
                self.com_thread.start()
                self.toggle_button.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    def send_data(self):
        # 送信ボックスの入力値を取得
        data = self.entry_text.get()
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ser.write((data + delimiter).encode())

            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%H:%M:%S')
            self.received_text.insert(tk.END, f"[Send -> {timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.received_text.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))

    def Serial_send_button(self, data_comsend):
        # 送信ボックスの入力値を取得

        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信

            self.ser.write((data_comsend + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%H:%M:%S')
            self.received_text.insert(tk.END, f"[Send -> {timestamp}] {data_comsend}{REV_DELIMITER[delimiter]}\n")
            self.received_text.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))

    def read_comport_data(self):
        # 設定したデリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        buf = ''
        while self.ser.is_open:
            try:
                data = self.ser.read(1).decode()
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.received_text.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.received_text.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.received_text.see("end")
                        self.barreceived_text.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.barreceived_text.see("end")
                        self.barcoder_data = buf

                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))

    def data_caseread(self):
        ''''''

    def motor_para(self):
        ''''''

    # 履歴をクリア
    def clear_received_text(self):
        self.received_text.delete(1.0, tk.END)

    ################################################################ SOCKET ##########################
    def socket_connect(self):

        if self.PLC_CN_OK == 0:

            while True:
                try:
                    # 01. Preparing Socket : socket()
                    self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    self.sock.connect((self.IP_PLC.get(), self.PORT_PLC.get()))
                    mess = "ST MR0" + '\r'
                    self.sock.send(mess.encode('utf-8'))
                    rec_mess = self.sock.recv(MAX_MESSAGE).decode('utf-8')
                    # print(rec_mess)
                    self.sock.close()
                    time.sleep(0.01)

                    if rec_mess == 'OK' + '\r\n':
                        self.PLC_stt.delete(1.0, tk.END)
                        print('PLC connected')
                        self.PLC_stt.insert(tk.END, "OK")
                        self.PLC_CN_OK = 1
                        for i in range(1000, 1022, 2):
                            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                            self.sock.connect((self.IP_PLC.get(), self.PORT_PLC.get()))
                            # print(i)
                            mess = "RD DM" + str(i) + '\r'
                            self.sock.send(mess.encode('utf-8'))
                            # self.sock.settimeout(0.1)
                            rec_mess = self.sock.recv(MAX_MESSAGE).decode('utf-8')
                            int_mess = int(rec_mess)
                            # print(int_mess)
                            if i == 1000:
                                self.DM1000_var.set(int_mess)
                            if i == 1002:
                                self.DM1002_var.set(int_mess)
                            if i == 1004:
                                self.DM1004_var.set(int_mess)
                            if i == 1006:
                                self.DM1006_var.set(int_mess)
                            if i == 1008:
                                self.DM1008_var.set(int_mess)
                            if i == 1010:
                                self.DM1010_var.set(int_mess)
                            if i == 1012:
                                self.DM1012_var.set(int_mess)
                            if i == 1014:
                                self.DM1014_var.set(int_mess)
                            if i == 1016:
                                self.DM1016_var.set(int_mess)
                            if i == 1018:
                                self.DM1018_var.set(int_mess)
                            if i == 1020:
                                self.DM1020_var.set(int_mess)

                            self.sock.close()
                            time.sleep(0.01)

                    self.Main_update()
                    break
                except Exception as e:
                    self.PLC_stt.insert(tk.END, "NG")
                    messagebox.showerror("PLC 送信エラー", str(e))
                    self.sock.close()
                    break

    def Socket_send(self, mess):
        self.send_flag = 1
        while True:
            try:
                # 通信の確立
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.connect((self.IP_PLC.get(), self.PORT_PLC.get()))
                mess = mess + '\r'
                # sock.send(mess)

                # メッセージ送信
                sock.send(mess.encode('utf-8'))  # ASCII#'utf-8'
                timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                self.Text_sock_recieved.insert(tk.END, f"⏎{timestamp} {mess}\n")
                self.Text_sock_recieved.see("end")  # 自動スクロール
                ###
                rec_mess = sock.recv(MAX_MESSAGE).decode('utf-8')
                timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                self.Text_sock_recieved.insert(tk.END, f"▶{timestamp} {rec_mess}\n")
                self.Text_sock_recieved.see("end")  # 自動スクロール

                # 通信の終了
                sock.close()
                self.send_flag = 0
                break

            except Exception as e:
                self.draw_id = None
                messagebox.showerror("PLC 送信エラー", str(e))
                sock.close()

                break

    def Socket_receive(self):

        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((HOST, PORT))
        # mess = ""
        mess = sock.recv(MAX_MESSAGE)
        mess = mess.decode('utf-8')
        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        self.Text_sock_recieved.insert(tk.END, f"▶{timestamp} {mess}\n")
        self.Text_sock_recieved.see("end")  # 自動スクロール

        self.sock.close()
        # self.recv_id = self.after(10, self.Socket_receive)

    ####################### CAMERA SET ##########################
    def camera_callback(self):
        # if self.read_defaut == False:
        #    self.camera_defaut()
        #    self.read_defaut = True

        ret, frame = self.capture.read()
        self.cv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        self.cv_image_base = self.cv_image
        # self.cv_background = self.cv_image.copy()
        if self.check_hani_ok.get() == 1:
            if self.check_patan6.get() == 1:
                # self.cv_image = self.area_select_special(self.cv_image)
                self.cv_image = self.area_select(self.cv_image)
                self.area_roi = self.color_map_select(self.area_roi)
                width, height = self.area_roi.shape[:2]
                self.area_roi = self.cam_filter(self.area_roi)
                if self.circle_select.get() == 1:
                    self.area_roi = self.cam_circle_elipse_detect(self.area_roi)
                if self.circle_select_adv.get() == 1:
                    self.area_roi = self.circle_detect(self.area_roi)
                if self.triang_select.get() == 1:
                    self.area_roi = self.triangle_detect(self.area_roi)
                if self.rect_select.get() == 1:
                    self.area_roi = self.rectange_detect(self.area_roi)
                if self.edge_select.get() == 1:
                    self.area_roi = self.edge_detect(self.area_roi)
                if self.draw_contour_select.get() == 1:
                    self.area_roi = self.draw_contour_object(self.area_roi)
                if self.connected_select.get() == 1:
                    self.area_roi = self.connected_component(self.area_roi)
                if self.rectangge_component_select.get() == 1:
                    self.area_roi = self.rectangge_component(self.area_roi)

                self.cv_image[self.pt1[1]:self.pt1[1] + width, self.pt1[0]:self.pt1[0] + height] = self.area_roi

            if self.check_patan2.get() == 1:
                self.cv_image = self.color_map_select(self.cv_image)
                w_3c = np.full_like(self.cv_image, fill_value=(255, 255, 255))  # Create white image
                #
                roi_corners = np.array([[self.pt1_specal, self.pt2_specal, self.pt3_specal]], dtype=np.int32)
                d_3c = np.zeros_like(self.cv_image[:, :], dtype='uint8')  # Create mask image
                cv2.fillPoly(d_3c, roi_corners, [255] * 3)
                d_1c = d_3c[:, :, 0]
                masked = cv2.bitwise_and(self.cv_image, w_3c, mask=d_1c)  # black mask
                # Add white background
                d_3c_i = ~d_3c
                final = cv2.bitwise_or(self.cv_image, d_3c_i)

                masked_image = self.cam_filter(final)

                if self.circle_select.get() == 1:
                    masked_image = self.cam_circle_elipse_detect(masked_image)
                if self.circle_select_adv.get() == 1:
                    masked_image = self.circle_detect(masked_image)
                if self.triang_select.get() == 1:
                    masked_image = self.triangle_detect(masked_image)
                if self.rect_select.get() == 1:
                    masked_image = self.rectange_detect(masked_image)
                if self.edge_select.get() == 1:
                    masked_image = self.edge_detect(masked_image)
                if self.draw_contour_select.get() == 1:
                    masked_image = self.draw_contour_object(masked_image)
                if self.connected_select.get() == 1:
                    masked_image = self.connected_component(masked_image)
                if self.rectangge_component_select.get() == 1:
                    masked_image = self.rectangge_component(masked_image)

                cv_image_sel = self.area_select_3kaku(masked_image)
                final_select = cv2.bitwise_and(cv_image_sel, self.cv_image_base)

                self.cv_image = final_select

            if self.check_patan3.get() == 1:
                self.cv_image = self.color_map_select(self.cv_image)
                w_3c = np.full_like(self.cv_image, fill_value=(255, 255, 255))  # Create white image
                #
                roi_corners = np.array([[self.pt1_kaku, self.pt2_kaku, self.pt3_kaku, self.pt4_kaku]], dtype=np.int32)
                d_3c = np.zeros_like(self.cv_image[:, :], dtype='uint8')  # Create mask image
                cv2.fillPoly(d_3c, roi_corners, [255] * 3)
                d_1c = d_3c[:, :, 0]
                masked = cv2.bitwise_and(self.cv_image, w_3c, mask=d_1c)  # black mask
                # Add white background
                d_3c_i = ~d_3c
                final = cv2.bitwise_or(self.cv_image, d_3c_i)

                masked_image = self.cam_filter(final)
                if self.circle_select.get() == 1:
                    masked_image = self.cam_circle_elipse_detect(masked_image)
                if self.circle_select_adv.get() == 1:
                    masked_image = self.circle_detect(masked_image)
                if self.triang_select.get() == 1:
                    masked_image = self.triangle_detect(masked_image)
                if self.rect_select.get() == 1:
                    masked_image = self.rectange_detect(masked_image)
                if self.edge_select.get() == 1:
                    masked_image = self.edge_detect(masked_image)
                if self.draw_contour_select.get() == 1:
                    masked_image = self.draw_contour_object(masked_image)
                if self.connected_select.get() == 1:
                    masked_image = self.connected_component(masked_image)
                if self.rectangge_component_select.get() == 1:
                    masked_image = self.rectangge_component(masked_image)
                cv_image_sel = self.area_select_4kaku(masked_image)
                # self.cv_image  = cv2.bitwise_and(self.cv_image, masked_image)
                # self.cv_image = masked_image
                # self.cv_image[self.pt1[1]:self.pt1[1] + width, self.pt1[0]:self.pt1[0] + height] = masked_image
                final_select = cv2.bitwise_and(cv_image_sel, self.cv_image_base)

                self.cv_image = final_select

            if self.check_patan4.get() == 1:
                self.cv_image = self.color_map_select(self.cv_image)
                w_3c = np.full_like(self.cv_image, fill_value=(255, 255, 255))  # Create white image
                #
                roi_corners = np.array([[self.pt1_polygol5, self.pt2_polygol5, self.pt3_polygol5, self.pt4_polygol5, self.pt5_polygol5]], dtype=np.int32)
                d_3c = np.zeros_like(self.cv_image[:, :], dtype='uint8')  # Create mask image
                cv2.fillPoly(d_3c, roi_corners, [255] * 3)
                d_1c = d_3c[:, :, 0]
                masked = cv2.bitwise_and(self.cv_image, w_3c, mask=d_1c)  # black mask
                # Add white background
                d_3c_i = ~d_3c
                final = cv2.bitwise_or(self.cv_image, d_3c_i)

                masked_image = self.cam_filter(final)
                if self.circle_select.get() == 1:
                    masked_image = self.cam_circle_elipse_detect(masked_image)
                if self.circle_select_adv.get() == 1:
                    masked_image = self.circle_detect(masked_image)
                if self.triang_select.get() == 1:
                    masked_image = self.triangle_detect(masked_image)
                if self.rect_select.get() == 1:
                    masked_image = self.rectange_detect(masked_image)
                if self.edge_select.get() == 1:
                    masked_image = self.edge_detect(masked_image)
                if self.draw_contour_select.get() == 1:
                    masked_image = self.draw_contour_object(masked_image)
                if self.connected_select.get() == 1:
                    masked_image = self.connected_component(masked_image)
                if self.rectangge_component_select.get() == 1:
                    masked_image = self.rectangge_component(masked_image)
                cv_image_sel = self.area_select_5kaku(masked_image)
                # self.cv_image  = cv2.bitwise_and(self.cv_image, masked_image)
                # self.cv_image = masked_image
                # self.cv_image[self.pt1[1]:self.pt1[1] + width, self.pt1[0]:self.pt1[0] + height] = masked_image
                final_select = cv2.bitwise_and(cv_image_sel, self.cv_image_base)

                self.cv_image = final_select

        self.pil_image = Image.fromarray(self.cv_image)
        if self.zoom_load_defaut == False:
            self.zoom_fit(self.pil_image.width, self.pil_image.height)
            self.zoom_load_defaut = True

        self.redraw_image()
        # print(self.Camera_check)
        # self.disp_image(frame)
        if self.capture_switchis_on == True:
            self.disp_id = self.after(10, self.camera_callback)
        else:
            self.disp_id = None

    def camera_defaut(self):
        a = self.capture.get(cv2.CAP_PROP_FRAME_WIDTH)
        self.FRAME_WIDTH_var.set(a)
        b = self.capture.get(cv2.CAP_PROP_FRAME_HEIGHT)
        self.FRAME_HEIGHT_var.set(b)
        c = self.capture.get(cv2.CAP_PROP_FPS)
        self.FPS_var.set(30)
        d = self.capture.get(cv2.CAP_PROP_BRIGHTNESS)
        self.BRIGHTNESS_var.set(0)
        e = self.capture.get(cv2.CAP_PROP_CONTRAST)
        self.CONTRAST_var.set(0)
        f = self.capture.get(cv2.CAP_PROP_SATURATION)
        self.SATURATION_var.set(0)
        g = self.capture.get(cv2.CAP_PROP_HUE)
        self.HUE_var.set(0)
        h = self.capture.get(cv2.CAP_PROP_GAIN)
        self.GAIN_var.set(0)
        i = self.capture.get(cv2.CAP_PROP_EXPOSURE)
        self.GAIN_var.set(-2)
        j = self.capture.get(cv2.CAP_PROP_AUTO_EXPOSURE)
        self.GAIN_var.set(1)
        k = self.capture.get(cv2.CAP_PROP_AUTOFOCUS)
        self.GAIN_var.set(1)
        #
        # print(a,b,c,d,e,f,g,h)

    def camera_setting(self, a):
        self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, self.FRAME_WIDTH_var.get())

    def camera_setting1(self, a):
        self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, self.FRAME_HEIGHT_var.get())

    def camera_setting2(self, a):
        self.capture.set(cv2.CAP_PROP_FPS, self.FPS_var.get())

    def camera_setting3(self, a):
        self.capture.set(cv2.CAP_PROP_BRIGHTNESS, self.BRIGHTNESS_var.get())

    def camera_setting4(self, a):
        self.capture.set(cv2.CAP_PROP_CONTRAST, self.CONTRAST_var.get())

    def camera_setting5(self, a):
        self.capture.set(cv2.CAP_PROP_SATURATION, self.SATURATION_var.get())

    def camera_setting6(self, a):
        self.capture.set(cv2.CAP_PROP_HUE, self.HUE_var.get())

    def camera_setting7(self, a):
        self.capture.set(cv2.CAP_PROP_GAIN, self.GAIN_var.get())

    def camera_setting8(self, a):
        self.capture.set(cv2.CAP_PROP_AUTO_EXPOSURE, self.auto_exposure_var.get())

    def camera_setting9(self, a):
        self.capture.set(cv2.CAP_PROP_AUTOFOCUS, self.auto_focus_var.get())

    def camera_setting10(self, a):
        self.capture.set(cv2.CAP_PROP_EXPOSURE, self.EXPOSE_var.get())

    def reset_data(self):
        self.capture.set(cv2.CAP_PROP_BRIGHTNESS, 0)
        self.capture.set(cv2.CAP_PROP_CONTRAST, 32)
        self.capture.set(cv2.CAP_PROP_SATURATION, 55)
        self.capture.set(cv2.CAP_PROP_HUE, 0)
        self.capture.set(cv2.CAP_PROP_GAIN, 0)
        #
        self.colorchoosen.current(0)
        self.boder_filter_map.current(0)
        self.boder_filter_map1.current(0)
        self.scale_var.set(50)
        self.adaptiveThreshold_var.set(0.01)
        self.canny1_var.set(100)
        self.canny2_var.set(500)
        self.gaussion_var.set(0)
        self.Blur_var.set(0)
        self.bold_var.set(0)
        self.debold_var.set(0)
        self.bold_var1.set(0)
        self.debold_var1.set(0)
        self.filterchoosen.current(0)
        self.filterchoosen1.current(0)
        self.filterchoosen2.current(0)
        self.filterchoosen3.current(0)
        self.filterchoosen4.current(0)
        self.so_fiter = 1
        self.fil_by_Area.set(1)
        self.filByCircularity.set(1)
        self.filByConvexity.set(1)
        self.filByInertia.set(1)
        self.minArea_var.set(50)
        self.minCircularity_var.set(0.1)
        self.minConvexity_var.set(0.1)
        self.minInertia_var.set(0.01)
        self.Kernel_var.set(3)
        self.param1_var.set(50)
        self.param2_var.set(30)
        self.minRadius_var.set(1)
        self.maxRadius_var.set(40)
        self.var_canny_a.set(250)
        self.var_canny_b.set(150)
        self.var_thr.set(50)
        self.len_size2.set(50)
        self.len_size3.set(15)
        self.len_size1.set(1)
        self.circle_select.set(0)
        self.circle_select_adv.set(0)
        self.triang_select.set(0)
        self.rect_select.set(0)
        self.edge_select.set(0)
        self.boder_map.current(0)
        self.contua_map.current(0)
        self.contua_map1.current(1)
        self.rect_para1.set(0.04)
        self.area_rect.set(150)
        self.chuvi_rect.set(50)
        self.rect_thres.set(150)
        #
        self.boder_map2.current(0)
        self.contua_map2.current(0)
        self.contua_map21.current(1)
        self.trian_para1.set(0.04)
        self.area_trian.set(150)
        self.chuvi_trian.set(50)
        self.trian_thres.set(150)
        self.add_fiter_to()
        self.minus_fiter_to()

        #

    def capture_switch(self):
        if self.capture_switchis_on:
            self.video_capture.config(bg='#FF4C00')
            self.video_capture.config(text='Continue_enable')
            self.camera_callback()
            self.capture_switchis_on = False
        else:

            self.video_capture.config(bg='gray')
            self.video_capture.config(text='Continue_disable')
            self.capture_switchis_on = True

    # $#################### CAMERA ZOOM ####################
    def mouse_move(self, event):
        ''' マウスの移動時
        # マウス座標
        self.mouse_position["text"] = f"mouse(x, y) = ({event.x: 4d}, {event.y: 4d})"

        if self.pil_image == None:
            return

        # 画像座標
        mouse_posi = np.array([event.x, event.y, 1])  # マウス座標(numpyのベクトル)
        mat_inv = np.linalg.inv(self.mat_affine)  # 逆行列（画像→Cancasの変換からCanvas→画像の変換へ）
        image_posi = np.dot(mat_inv, mouse_posi)  # 座標のアフィン変換
        x = int(np.floor(image_posi[0]))
        y = int(np.floor(image_posi[1]))
        if x >= 0 and x < self.pil_image.width and y >= 0 and y < self.pil_image.height:
            # 輝度値の取得
            value = self.pil_image.getpixel((x, y))
            self.image_position["text"] = f"image({x: 4d}, {y: 4d}) = {value}"
        else:
            self.image_position["text"] = "-------------------------" '''

    def mouse_move_mid(self, event):
        if self.pil_image == None:
            return
        self.translate(event.x - self.__old_event.x, event.y - self.__old_event.y)
        self.redraw_image()  # 再描画
        self.__old_event = event

    def mouse_down_mid(self, event):
        ''' マウスの左ボタンをドラッグ '''
        self.__old_event = event

    def mouse_move_right(self, event):
        ''''''

    def mouse_down_right(self, event):
        ''''''

    def mouse_up_right(self, event):
        ''''''

    def mouse_down_left(self, event):
        ''' マウスの左ボタンを押した '''
        if self.pil_image == None:
            return
        self.__old_event = event
        mouse_posi = np.array([event.x, event.y, 1])  # マウス座標(numpyのベクトル)
        mat_inv = np.linalg.inv(self.mat_affine)  # 逆行列（画像→Cancasの変換からCanvas→画像の変換へ）
        image_posi = np.dot(mat_inv, mouse_posi)  # 座標のアフィン変換
        x = int(np.floor(image_posi[0]))
        y = int(np.floor(image_posi[1]))
        if x >= 0 and x < self.pil_image.width and y >= 0 and y < self.pil_image.height and self.cam_edit_hani == True:
            # 輝度値の取得
            if self.check_patan6.get() == 1:
                if abs(x - self.pt1[0]) < 16 and abs(y - self.pt1[1]) < 16:
                    self.edit_pt1 = True
                if abs(x - self.pt2[0]) < 16 and abs(y - self.pt2[1]) < 16:
                    self.edit_pt2 = True
            ####
            if self.check_patan2.get() == 1:
                if abs(x - self.pt1_specal[0]) < 16 and abs(y - self.pt1_specal[1]) < 16:
                    self.edit_pt1_specal = True
                if abs(x - self.pt2_specal[0]) < 16 and abs(y - self.pt2_specal[1]) < 16:
                    self.edit_pt2_specal = True
                if abs(x - self.pt3_specal[0]) < 16 and abs(y - self.pt3_specal[1]) < 16:
                    self.edit_pt3_specal = True
            ########
            if self.check_patan3.get() == 1:
                if abs(x - self.pt1_kaku[0]) < 16 and abs(y - self.pt1_kaku[1]) < 16:
                    self.edit_pt1_kaku = True
                if abs(x - self.pt2_kaku[0]) < 16 and abs(y - self.pt2_kaku[1]) < 16:
                    self.edit_pt2_kaku = True
                if abs(x - self.pt3_kaku[0]) < 16 and abs(y - self.pt3_kaku[1]) < 16:
                    self.edit_pt3_kaku = True
                if abs(x - self.pt4_kaku[0]) < 16 and abs(y - self.pt4_kaku[1]) < 16:
                    self.edit_pt4_kaku = True
            if self.check_patan4.get() == 1:
                if abs(x - self.pt1_polygol5[0]) < 16 and abs(y - self.pt1_polygol5[1]) < 16:
                    self.edit_pt1_polygol5 = True
                if abs(x - self.pt2_polygol5[0]) < 16 and abs(y - self.pt2_polygol5[1]) < 16:
                    self.edit_pt2_polygol5 = True
                if abs(x - self.pt3_polygol5[0]) < 16 and abs(y - self.pt3_polygol5[1]) < 16:
                    self.edit_pt3_polygol5 = True
                if abs(x - self.pt4_polygol5[0]) < 16 and abs(y - self.pt4_polygol5[1]) < 16:
                    self.edit_pt4_polygol5 = True
                if abs(x - self.pt5_polygol5[0]) < 16 and abs(y - self.pt5_polygol5[1]) < 16:
                    self.edit_pt5_polygol5 = True

    def mouse_up_left(self, event):
        self.edit_pt1 = False
        self.edit_pt2 = False
        self.edit_pt1_specal = False
        self.edit_pt2_specal = False
        self.edit_pt3_specal = False
        self.edit_pt1_kaku = False
        self.edit_pt2_kaku = False
        self.edit_pt3_kaku = False
        self.edit_pt4_kaku = False

        self.edit_pt1_polygol5 = False
        self.edit_pt2_polygol5 = False
        self.edit_pt3_polygol5 = False
        self.edit_pt4_polygol5 = False
        self.edit_pt5_polygol5 = False

    def mouse_move_left(self, event):
        if self.pil_image == None:
            return
        mouse_posi = np.array([event.x, event.y, 1])  # マウス座標(numpyのベクトル)
        mat_inv = np.linalg.inv(self.mat_affine)  # 逆行列（画像→Cancasの変換からCanvas→画像の変換へ）
        image_posi = np.dot(mat_inv, mouse_posi)  # 座標のアフィン変換
        x = int(np.floor(image_posi[0]))
        y = int(np.floor(image_posi[1]))
        if x >= 0 and x < self.pil_image.width and y >= 0 and y < self.pil_image.height:
            if self.edit_pt1 == True:
                # print("ohshit")
                self.pt1 = (x, y)
            if self.edit_pt2 == True:
                # print("ohshit")
                self.pt2 = (x, y)
            ###############
            if self.edit_pt1_specal == True:
                # print("ohshit")
                self.pt1_specal = (x, y)
            if self.edit_pt2_specal == True:
                # print("ohshit")
                self.pt2_specal = (x, y)
            if self.edit_pt3_specal == True:
                # print("ohshit")
                self.pt3_specal = (x, y)
            ###########
            if self.edit_pt1_kaku == True:
                self.pt1_kaku = (x, y)
            if self.edit_pt2_kaku == True:
                self.pt2_kaku = (x, y)
            if self.edit_pt3_kaku == True:
                self.pt3_kaku = (x, y)
            if self.edit_pt4_kaku == True:
                self.pt4_kaku = (x, y)
            ##########
            if self.edit_pt1_polygol5 == True:
                self.pt1_polygol5 = (x, y)
            if self.edit_pt2_polygol5 == True:
                self.pt2_polygol5 = (x, y)
            if self.edit_pt3_polygol5 == True:
                self.pt3_polygol5 = (x, y)
            if self.edit_pt4_polygol5 == True:
                self.pt4_polygol5 = (x, y)
            if self.edit_pt5_polygol5 == True:
                self.pt5_polygol5 = (x, y)

    def mouse_double_click_left(self, event):
        ''' マウスの左ボタンをダブルクリック '''
        if self.pil_image == None:
            return
        self.zoom_fit(self.pil_image.width, self.pil_image.height)
        self.redraw_image()  # 再描画

    def mouse_wheel(self, event):
        ''' マウスホイールを回した '''
        if self.pil_image == None:
            return

        if (event.delta < 0):
            # 上に回転の場合、縮小
            self.scale_at(0.8, event.x, event.y)
        else:
            # 下に回転の場合、拡大
            self.scale_at(1.25, event.x, event.y)

        self.redraw_image()  # 再描画

    def reset_transform(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine = np.eye(3)  # 3x3の単位行列

    def translate(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine = np.dot(mat, self.mat_affine)

    def scale(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine = np.dot(mat, self.mat_affine)

    def scale_at(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate(-cx, -cy)
        # 拡大縮小
        self.scale(scale)
        # 元に戻す
        self.translate(cx, cy)

    def zoom_fit(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_cam.winfo_width()
        canvas_height = self.canvas_cam.winfo_height()

        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale(scale)
        # あまり部分を中央に寄せる
        self.translate(offsetx, offsety)

    def draw_image(self, pil_image):

        if pil_image == None:
            return

        self.canvas_cam.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_cam.winfo_width()
        canvas_height = self.canvas_cam.winfo_height()

        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor="#CBD3C0"
        )

        # 表示用画像を保持
        self.image = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_cam.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image  # 表示画像データ
        )

    def redraw_image(self):
        ''' 画像の再描画 '''
        if self.pil_image == None:
            return
        self.draw_image(self.pil_image)

    ############## AREA SELECT ##############
    def area_select(self, cv_image):
        rect_x0 = self.pt1[0] - 5  # -3
        if rect_x0 < 0:
            rect_x0 = 0
        rect_y0 = self.pt1[1] - 5  # -3
        if rect_y0 < 0:
            rect_y0 = 0
        rect_ofset1 = [rect_x0, rect_y0]
        ###########3
        rect_x1 = self.pt2[0] + 5  # +3
        # if rect_x1 < 0:
        #    rect_x1 = 0
        rect_y1 = self.pt2[1] + 5  # +3
        # if rect_y1 > self.:
        #    rect_y1 = 0
        rect_ofset2 = [rect_x1, rect_y1]

        cv_image = cv2.rectangle(cv_image, rect_ofset1, rect_ofset2, (0, 255, 0), 2)
        # cv_image = cv2.rectangle(cv_image, self.pt1, self.pt2, (255, 255, 255), 1)
        cv_image = cv2.circle(cv_image, (self.pt1[0] - 5, self.pt1[1] - 5), 5, (255, 0, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2[0] + 5, self.pt2[1] + 5), 5, (255, 0, 0), -1)

        roi = (self.pt1[0] + 1, self.pt1[1] + 1, self.pt2[0] - 1, self.pt2[1] - 1)
        self.area_roi = cv_image[roi[1]: roi[3], roi[0]: roi[2]]

        return cv_image

    def area_select_3kaku(self, cv_image):

        cv_image = cv2.circle(cv_image, (self.pt1_specal[0], self.pt1_specal[1]), 5, (0, 255, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2_specal[0], self.pt2_specal[1]), 5, (0, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt3_specal[0], self.pt3_specal[1]), 5, (255, 0, 0), -1)
        cv_image = cv2.line(cv_image, self.pt1_specal, self.pt2_specal, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt2_specal, self.pt3_specal, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt3_specal, self.pt1_specal, (0, 255, 255), 2)
        return cv_image

    def area_select_4kaku(self, cv_image):

        cv_image = cv2.circle(cv_image, (self.pt1_kaku[0], self.pt1_kaku[1]), 5, (0, 255, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2_kaku[0], self.pt2_kaku[1]), 5, (0, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt3_kaku[0], self.pt3_kaku[1]), 5, (255, 0, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt4_kaku[0], self.pt4_kaku[1]), 5, (255, 0, 255), -1)

        cv_image = cv2.line(cv_image, self.pt1_kaku, self.pt2_kaku, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt2_kaku, self.pt3_kaku, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt3_kaku, self.pt4_kaku, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt4_kaku, self.pt1_kaku, (0, 255, 255), 2)
        return cv_image

    def area_select_5kaku(self, cv_image):

        cv_image = cv2.circle(cv_image, (self.pt1_polygol5[0], self.pt1_polygol5[1]), 5, (0, 255, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2_polygol5[0], self.pt2_polygol5[1]), 5, (0, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt3_polygol5[0], self.pt3_polygol5[1]), 5, (255, 0, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt4_polygol5[0], self.pt4_polygol5[1]), 5, (255, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt5_polygol5[0], self.pt5_polygol5[1]), 5, (255, 255, 0), -1)

        cv_image = cv2.line(cv_image, self.pt1_polygol5, self.pt2_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt2_polygol5, self.pt3_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt3_polygol5, self.pt4_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt4_polygol5, self.pt5_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt5_polygol5, self.pt1_polygol5, (0, 255, 255), 2)
        return cv_image

    def color_map_select(self, cv_image):
        grayscale_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)

        if self.colormap.get() == "":
            return cv_image
        if self.colormap.get() == "NORMAL_NO_FILTER":
            return cv_image
        if self.colormap.get() == "GRAY_SCALE":
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.colormap.get() == "COLORMAP_AUTUMN":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_AUTUMN)
            # cv2.putText(self.cv_image, "COLORMAP_AUTUMN", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_JET":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_JET)
            # cv2.putText(self.cv_image, "COLORMAP_JET", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_WINTER":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_WINTER)
            # cv2.putText(self.cv_image, "COLORMAP_WINTER", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_RAINBOW":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_RAINBOW)
            # cv2.putText(self.cv_image, "COLORMAP_RAINBOW", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_OCEAN":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_OCEAN)
            # cv2.putText(self.cv_image, "COLORMAP_OCEAN", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_SUMMER":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_SUMMER)
            # cv2.putText(self.cv_image, "COLORMAP_SUMMER", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_SPRING":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_SPRING)
            # cv2.putText(self.cv_image, "COLORMAP_SPRING", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_COOL":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_COOL)
            # cv2.putText(self.cv_image, "COLORMAP_COOL", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_HSV":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_HSV)
            # cv2.putText(self.cv_image, "COLORMAP_HSV", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_PINK":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_PINK)
            # cv2.putText(self.cv_image, "COLORMAP_PINK", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_HOT":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_HOT)
            # cv2.putText(self.cv_image, "COLORMAP_HOT", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        return cv_image

    def cam_filter(self, cv_image):

        if self.filtermap.get() == 'None':
            cv_image = cv_image
        if self.filtermap.get() == 'Threshold':
            if self.boder_filter_map_sl.get() == 'THRESH_BINARY':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
            if self.boder_filter_map_sl.get() == 'THRESH_BINARY_INV':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY_INV)
            if self.boder_filter_map_sl.get() == 'THRESH_TRUNC':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_TRUNC)
            if self.boder_filter_map_sl.get() == 'THRESH_TOZERO':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_TOZERO)
            if self.boder_filter_map_sl.get() == 'THRESH_TOZERO_INV':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_TOZERO_INV)

        if self.filtermap.get() == 'Adapt_Thre':
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)

            if self.boder_filter_map_sl1.get() == 'THRESH_BINARY':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 51,
                                                 self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_BINARY_INV':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV,
                                                 51, self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_TRUNC':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_TRUNC, 51,
                                                 self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_TOZERO':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_TOZERO, 51,
                                                 self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_TOZERO_INV':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_TOZERO_INV,
                                                 51, self.adaptiveThreshold_var.get())

            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.filtermap.get() == 'Canny_1':
            cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.filtermap.get() == 'Canny_2':
            cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.filtermap.get() == 'Gaussion':
            ksize = int(self.gaussion_var.get())
            ksize = int(ksize / 2) * 2 + 1
            cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
        if self.filtermap.get() == 'Blur':
            ksize = int(self.Blur_var.get())
            cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
        if self.filtermap.get() == 'Bold':
            kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
            cv_image = cv2.erode(cv_image, kernel, iterations=1)
            # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
        if self.filtermap.get() == 'Debold':
            kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
            # cv_image = cv2.erode(cv_image, kernel, iterations=1)
            cv_image = cv2.dilate(cv_image, kernel, iterations=1)
        #########################################################
        if self.so_fiter >= 2:
            if self.filtermap1.get() == 'None':
                cv_image = cv_image
            if self.filtermap1.get() == 'Threshold':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
            if self.filtermap1.get() == 'Adapt_Thre':
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 51,
                                                 self.adaptiveThreshold_var.get())
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
            if self.filtermap1.get() == 'Canny_1':
                cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
            if self.filtermap1.get() == 'Canny_2':
                cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
            if self.filtermap1.get() == 'Gaussion':
                ksize = int(self.gaussion_var.get())
                ksize = int(ksize / 2) * 2 + 1
                cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
            if self.filtermap1.get() == 'Blur':
                ksize = int(self.Blur_var.get())
                cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
            if self.filtermap1.get() == 'Bold':
                kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                cv_image = cv2.erode(cv_image, kernel, iterations=1)
                # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
            if self.filtermap1.get() == 'Debold':
                kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                ################################################
            if self.so_fiter >= 3:
                if self.filtermap2.get() == 'None':
                    cv_image = cv_image
                if self.filtermap2.get() == 'Threshold':
                    _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
                if self.filtermap2.get() == 'Adapt_Thre':
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                    cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
                                                     51,
                                                     self.adaptiveThreshold_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap2.get() == 'Canny_1':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap2.get() == 'Canny_2':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap2.get() == 'Gaussion':
                    ksize = int(self.gaussion_var.get())
                    ksize = int(ksize / 2) * 2 + 1
                    cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
                if self.filtermap2.get() == 'Blur':
                    ksize = int(self.Blur_var.get())
                    cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
                if self.filtermap2.get() == 'Bold':
                    kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                    cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                if self.filtermap2.get() == 'Debold':
                    kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                    # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    cv_image = cv2.dilate(cv_image, kernel, iterations=1)
            ############################################################
            if self.so_fiter >= 4:
                if self.filtermap3.get() == 'None':
                    cv_image = cv_image
                if self.filtermap3.get() == 'Threshold':
                    _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
                if self.filtermap3.get() == 'Adapt_Thre':
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                    cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
                                                     51,
                                                     self.adaptiveThreshold_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap3.get() == 'Canny_1':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap3.get() == 'Canny_2':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap3.get() == 'Gaussion':
                    ksize = int(self.gaussion_var.get())
                    ksize = int(ksize / 2) * 2 + 1
                    cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
                if self.filtermap3.get() == 'Blur':
                    ksize = int(self.Blur_var.get())
                    cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
                if self.filtermap3.get() == 'Bold':
                    kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                    cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                if self.filtermap3.get() == 'Debold':
                    kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                    # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    cv_image = cv2.dilate(cv_image, kernel, iterations=1)
            ####################################
            if self.so_fiter >= 5:
                if self.filtermap4.get() == 'None':
                    cv_image = cv_image
                if self.filtermap4.get() == 'Threshold':
                    _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
                if self.filtermap4.get() == 'Adapt_Thre':
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                    cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
                                                     51,
                                                     self.adaptiveThreshold_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap4.get() == 'Canny_1':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap4.get() == 'Canny_2':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap4.get() == 'Gaussion':
                    ksize = int(self.gaussion_var.get())
                    ksize = int(ksize / 2) * 2 + 1
                    cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
                if self.filtermap4.get() == 'Blur':
                    ksize = int(self.Blur_var.get())
                    cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
                if self.filtermap4.get() == 'Bold':
                    kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                    cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                if self.filtermap4.get() == 'Debold':
                    kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                    # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    cv_image = cv2.dilate(cv_image, kernel, iterations=1)

        return cv_image

    def cam_circle_elipse_detect(self, cv_image):
        params = cv2.SimpleBlobDetector_Params()

        # Set Area filtering parameters
        if self.fil_by_Area.get() == 1:
            params.filterByArea = True
            params.minArea = self.minArea_var.get()
        if self.filByCircularity.get() == 1:
            # Set Circularity filtering parameters
            params.filterByCircularity = True
            params.minCircularity = self.minCircularity_var.get()
        if self.filByConvexity.get() == 1:
            # Set Convexity filtering parameters
            params.filterByConvexity = True
            params.minConvexity = self.minConvexity_var.get()
        if self.filByInertia.get() == 1:
            # Set inertia filtering parameters
            params.filterByInertia = True
            params.minInertiaRatio = self.minInertia_var.get()

        # Create a detector with the parameters
        detector = cv2.SimpleBlobDetector_create(params)

        # Detect blobs
        keypoints = detector.detect(cv_image)

        # Draw blobs on our image as red circles
        blank = np.zeros((1, 1))
        blobs = cv2.drawKeypoints(cv_image, keypoints, blank, (0, 0, 255),
                                  cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)

        number_of_blobs = len(keypoints)
        text = "Number of Circular Blobs: " + str(len(keypoints))
        cv2.putText(blobs, text, (20, 550),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 100, 255), 2)
        # print(text)
        return blobs

    def circle_detect(self, cv_image):
        # Convert to grayscale.
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)

        # Blur using 3 * 3 kernel.
        gray_blurred = cv2.blur(gray, (self.Kernel_var.get(), self.Kernel_var.get()))

        # Apply Hough transform on the blurred image.
        detected_circles = cv2.HoughCircles(gray_blurred,
                                            cv2.HOUGH_GRADIENT, 1, 20, param1=self.param1_var.get(),
                                            param2=self.param2_var.get(), minRadius=self.minRadius_var.get(),
                                            maxRadius=self.maxRadius_var.get())

        # Draw circles that are detected.
        if detected_circles is not None:
            self.Camera_check = 1
            # Convert the circle parameters a, b and r to integers.
            detected_circles = np.uint16(np.around(detected_circles))

            for pt in detected_circles[0, :]:
                a, b, r = pt[0], pt[1], pt[2]

                # Draw the circumference of the circle.
                cv2.circle(cv_image, (a, b), r, (0, 255, 0), 2)

                # Draw a small circle (of radius 1) to show the center.
                cv2.circle(cv_image, (a, b), 1, (0, 0, 255), 3)
        else:
            self.Camera_check = 0
        return cv_image

    def rectange_detect(self, cv_image):
        # converting image into grayscale image
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        # setting threshold of gray image

        if self.border_sl.get() == 'THRESH_BINARY':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_BINARY)
        if self.border_sl.get() == 'THRESH_BINARY_INV':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_BINARY_INV)
        if self.border_sl.get() == 'THRESH_TRUNC':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_TRUNC)
        if self.border_sl.get() == 'THRESH_TOZERO':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_TOZERO)
        if self.border_sl.get() == 'THRESH_TOZERO_INV':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_TOZERO_INV)

        # _, threshold = cv2.threshold(gray, self.rect_thres, 255, cv2.THRESH_BINARY)

        # using a findContours() function
        if self.contour_sl.get() == 'RETR_TREE':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl.get() == 'RETR_CCOMP':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl.get() == 'RETR_LIST':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl.get() == 'RETR_EXTERNAL':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_TC89_L1)

        i = 0

        # list for storing names of shapes
        for contour in contours:

            # here we are ignoring first counter because
            # findcontour function detects whole image as shape
            if i == 0:
                i = 1
                continue

            # cv2.approxPloyDP() function to approximate the shape
            approx = cv2.approxPolyDP(
                contour, self.rect_para1.get() * cv2.arcLength(contour, True), True)

            # using drawContours() function
            # cv2.drawContours(cv_image, [contour], 0, (0, 0, 255), 3)
            area = cv2.contourArea(contour)
            chuvi = cv2.arcLength(contour, True)
            # finding center point of shape
            M = cv2.moments(contour)
            if M['m00'] != 0.0:
                x = int(M['m10'] / M['m00'])
                y = int(M['m01'] / M['m00'])

                # putting shape name at center of each shape
            '''if len(approx) == 3:
                cv2.putText(cv_image, 'Triangle', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''

            if len(approx) == 4 and area >= self.area_rect.get() and chuvi >= self.chuvi_rect.get():
                x, y, w, h = cv2.boundingRect(approx)
                cv2.rectangle(cv_image, (x, y), (x + w, y + h), (0, 255, 0), 2)
                # print('Rectangle Detect')
                self.Camera_check = 1
                # cv2.putText(cv_image, 'Quadrilateral', (x, y),
                # cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
            else:
                self.Camera_check = 0
            '''elif len(approx) == 5:
                cv2.putText(cv_image, 'Pentagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            elif len(approx) == 6:
                cv2.putText(cv_image, 'Hexagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            else:
                print('co cai eo')
                #cv2.putText(cv_image, 'circle', (x, y),
                            #cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''
        return cv_image

    def triangle_detect(self, cv_image):
        self.Camera_check = 0
        # converting image into grayscale image
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        # setting threshold of gray image

        if self.border_sl2.get() == 'THRESH_BINARY':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_BINARY)
        if self.border_sl2.get() == 'THRESH_BINARY_INV':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_BINARY_INV)
        if self.border_sl2.get() == 'THRESH_TRUNC':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_TRUNC)
        if self.border_sl2.get() == 'THRESH_TOZERO':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_TOZERO)
        if self.border_sl2.get() == 'THRESH_TOZERO_INV':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_TOZERO_INV)

        # _, threshold = cv2.threshold(gray, self.rect_thres, 255, cv2.THRESH_BINARY)

        # using a findContours() function
        if self.contour_sl2.get() == 'RETR_TREE':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl2.get() == 'RETR_CCOMP':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl2.get() == 'RETR_LIST':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl2.get() == 'RETR_EXTERNAL':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_TC89_L1)

        i = 0

        # list for storing names of shapes
        for contour in contours:
            if contour is None:
                self.Camera_check = 0
                return cv_image
            # here we are ignoring first counter because
            # findcontour function detects whole image as shape
            if i == 0:
                i = 1
                continue

            # cv2.approxPloyDP() function to approximate the shape
            approx = cv2.approxPolyDP(
                contour, self.rect_para1.get() * cv2.arcLength(contour, True), True)

            # using drawContours() function
            # cv2.drawContours(cv_image, [contour], 0, (0, 0, 255), 3)
            area = cv2.contourArea(contour)
            chuvi = cv2.arcLength(contour, True)
            # finding center point of shape
            M = cv2.moments(contour)
            if M['m00'] != 0.0:
                x = int(M['m10'] / M['m00'])
                y = int(M['m01'] / M['m00'])

                # putting shape name at center of each shape
            '''if len(approx) == 3:
                cv2.putText(cv_image, 'Triangle', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''

            if len(approx) == 3 and area >= self.area_trian.get() and chuvi >= self.chuvi_trian.get():
                x, y, w, h = cv2.boundingRect(approx)
                cv2.rectangle(cv_image, (x, y), (x + w, y + h), (0, 255, 0), 2)
                # print('TriAngle Detect')
                self.Camera_check = 1
                # cv2.putText(cv_image, 'Quadrilateral', (x, y),
                # cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
                return cv_image
            '''elif len(approx) == 5:
                cv2.putText(cv_image, 'Pentagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            elif len(approx) == 6:
                cv2.putText(cv_image, 'Hexagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            else:
                print('co cai eo')
                #cv2.putText(cv_image, 'circle', (x, y),
                            #cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''
        return cv_image

    def draw_contour_object(self, cv_image):
        self.Camera_check = 0
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        # setting threshold of gray image
        self.now_area = 0
        self.now_chuvi = 0
        ret, threshold = cv2.threshold(gray, self.Threshole_contua_var.get(), 255, cv2.THRESH_BINARY)
        contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

        i = 0

        # list for storing names of shapes
        for contour in contours:
            if contour is None:
                self.Camera_check = 0
                return cv_image
            # here we are ignoring first counter because
            # findcontour function detects whole image as shape
            if i == 0:
                i = 1
                continue
            area = cv2.contourArea(contour)
            chuvi = cv2.arcLength(contour, True)
            # cv2.approxPloyDP() function to approximate the shape
            approx = cv2.approxPolyDP(
                contour, self.contua_aprrox.get() * cv2.arcLength(contour, True), True)
            if chuvi >= self.chuvi_contua.get() and area > self.dientich_contua.get():
                # using drawContours() function
                self.Camera_check = 1
                self.now_area = cv2.contourArea(contour)
                self.now_chuvi = cv2.arcLength(contour, True)
                self.now_area_check.set(str(self.now_area))
                cv2.drawContours(cv_image, [contour], 0, (255, 0, 0), 2)
                return cv_image
        self.now_area_check.set(str(self.now_area))
        return cv_image

    def edge_detect(self, image):
        self.Camera_check = 0
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Use canny edge detection
        edges = cv2.Canny(gray, self.var_canny_a.get(), self.var_canny_b.get(), apertureSize=3)
        # Apply HoughLinesP method to
        # to directly obtain line end points
        lines_list = []
        lines = cv2.HoughLinesP(
            edges,  # Input edge image
            self.len_size1.get(),  # Distance resolution in pixels
            np.pi / 180,  # Angle resolution in radians
            self.var_thr.get(),  # Min number of votes for valid line
            minLineLength=self.len_size2.get(),  # Min allowed length of line
            maxLineGap=self.len_size3.get()  # Max allowed gap between line for joining them
        )

        # Iterate over points
        if lines is None:
            self.Camera_check = 0
            return image
        # print(lines)
        self.Camera_check = 1
        for points in lines:
            # Extracted points nested in the list
            x1, y1, x2, y2 = points[0]
            # Draw the lines joing the points
            # On the original image
            cv2.line(image, (x1, y1), (x2, y2), (0, 0, 255), 3, cv2.LINE_8)
            # Maintain a simples lookup list for points
            lines_list.append([(x1, y1), (x2, y2)])

        return image

    def connected_component(self, cv_image):
        # 画像表示用に入力画像をカラーデータに変換する
        image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        cv_image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
        if self.border_connect_sl.get() == 'THRESH_BINARY':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_BINARY)
        if self.border_connect_sl.get() == 'THRESH_BINARY_INV':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_BINARY_INV)
        if self.border_connect_sl.get() == 'THRESH_TRUNC':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_TRUNC)
        if self.border_connect_sl.get() == 'THRESH_TOZERO':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_TOZERO)
        if self.border_connect_sl.get() == 'THRESH_TOZERO_INV':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_TOZERO_INV)
        # 二値化(二値化して白黒反転)
        # retval, cv_image = cv2.threshold(cv_image, self.thres_connect.get(), 255, cv2.THRESH_BINARY_INV)

        # クロージング処理（細切れ状態を防ぐため）
        kernel = np.ones((self.kernel_connect.get(), self.kernel_connect.get()), np.uint8)
        image = cv2.morphologyEx(image, cv2.MORPH_CLOSE, kernel, iterations=self.iterations_var.get())

        # ラベリング
        retval, labels, stats, centroids = cv2.connectedComponentsWithStats(image)

        # 結果表示
        for i in range(1, retval):
            x, y, width, height, area = stats[i]  # x座標, y座標, 幅, 高さ, 面積

            if area > self.area_detect_connect.get():  # 面積が１０画素以上の部分
                cv2.rectangle(cv_image, (x, y), (x + width, y + height), (0, 0, 255), 2)
                cv2.putText(cv_image, f"[{i}]:{area}", (x, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (255, 0, 0), 1,
                            cv2.LINE_AA)
                # print(i, area)
        # cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        return cv_image

    def rectangge_component(self, img):
        self.Camera_check = 0
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # dientich = self.var_Height_tab3.get() * self.var_Width_tab3.get()
        fill_change = self.fill_change_var.get() * 0.01

        # 白黒反転して二値化
        ret, img = cv2.threshold(img, self.thres_patan.get(), 255, cv2.THRESH_BINARY_INV)

        # 一番外側の輪郭のみを取得
        contours, hierarchy = cv2.findContours(img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # 画像表示用に入力画像をカラーデータに変換する
        img_disp = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

        # 全ての輪郭を描画
        cv2.drawContours(img_disp, contours, -1, (255, 0, 255), 1)

        # 輪郭の点の描画
        for i, contour in enumerate(contours):  ### for contour in contours:

            area = cv2.contourArea(contours[i])
            if area > self.area_patan.get():
                self.have_contour = 1
            else:
                self.have_contour = 0
                self.Camera_check = 0

                return img_disp
            if self.have_contour == 1:
                # print(area)
                # 傾いた外接する矩形領域の描画
                rect = cv2.minAreaRect(contour)
                box = cv2.boxPoints(rect)
                box = np.intp(box)
                cv2.drawContours(img_disp, [box], 0, (0, 255, 255), 1)
                # 矩形度の計算
                val = self.rectangularity(contour)
                # 輪郭の矩形領域
                x, y, w, h = cv2.boundingRect(contour)
                # 矩形度の描画
                if val is not None:

                    cv2.putText(img_disp, f"{val:.3f}", (x, y + 15), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0), 1,
                                cv2.LINE_AA)
                    # 円らしい領域（円形度が0.85以上）を囲う
                    if val > fill_change + fill_change * 0.1:
                        self.Camera_check = 1
                        cv2.rectangle(img_disp, (x - 0, y - 0), (x + w + 0, y + h + 0), (0, 255, 0), 2)  # 少し外側を囲う
                    if val >= fill_change - fill_change * 0.1 and val <= fill_change + fill_change * 0.1:
                        self.Camera_check = 1
                        cv2.rectangle(img_disp, (x - 0, y - 0), (x + w + 0, y + h + 0), (255, 255, 0), 2)  # 少し外側を囲う
                    if val < fill_change - fill_change * 0.1:
                        self.Camera_check = 0
                        cv2.rectangle(img_disp, (x - 0, y - 0), (x + w + 0, y + h + 0), (255, 0, 0), 2)  # 少し外側を囲う

        return img_disp

    def rectangularity(self, contour):
        '''
        矩形度を求める

        Parameters
        ----------
        contour : ndarray
            輪郭の(x,y)座標の配列

        Returns
        -------
            矩形度

        '''
        # 面積
        area = cv2.contourArea(contour)
        # 傾いた外接する矩形領域
        _, (width, height), _ = cv2.minAreaRect(contour)

        # 矩形度を返す
        if width == 0 or height == 0:
            return

        return area / width / height

    def hakotori_kakunin(self):
        '''モードレスダイアログボックスの作成'''
        self.hakotori_dlg = tk.Toplevel(self)
        self.hakotori_dlg.title("箱取り確認")  # ウィンドウタイトル
        self.hakotori_dlg.geometry("700x200")  # ウィンドウサイズ(幅x高さ)
        x = root.winfo_x()
        y = root.winfo_y()
        self.hakotori_dlg.geometry("+%d+%d" % (x + 220, y + 250))

        self.hakotori_ok = tk.Button(self.hakotori_dlg, font=("Arial", 15), wraplength=150, height=3, width=20, text="箱取り完了", command=self.hakotori_ok_clicked)
        self.hakotori_ok.place(x=50, y=70)
        self.hakomodosi_ok = tk.Button(self.hakotori_dlg, font=("Arial", 15), wraplength=150, height=3, width=20, text="箱入れ直し完了", command=self.close_hakotori)
        self.hakomodosi_ok.place(x=400, y=70)

        if self.hako_tori_kakunin == 0:
            self.lbl_hakotori = Label(self.hakotori_dlg, font=("Arial", 15),
                                      text="落とした空箱を一旦回収し、完了を押してください。")
            self.lbl_hakotori.place(x=0, y=0)
            self.hakotori_ok["state"] = "normal"  # normal
            self.hakomodosi_ok["state"] = "disable"  # disable
        # time.sleep(1)
        self.hakotori_dlg.grab_set()  # モーダルにする
        self.hakotori_dlg.focus_set()  # フォーカスを新しいウィンドウをへ移す
        self.hakotori_dlg.transient(self.master)  # タスクバーに表示しない
        app.wait_window(self.hakotori_dlg)

    def hakotori_ok_clicked(self):
        self.Socket_send("ST 4114")
        self.hako_tori_kakunin = 1
        self.lbl_hakotori = Label(self.hakotori_dlg, font=("Arial", 15),
                                  text="取り除いた箱の向きを合わせて入れ直し、入れ直し完了を押してください。")
        self.lbl_hakotori.place(x=0, y=0)
        self.hakotori_ok["state"] = "disable"
        self.hakomodosi_ok["state"] = "normal"

    ###########################DIALOG####################
    def close_hakotori(self):
        self.Socket_send("ST R4113")
        self.hako_tori_kakunin = 0
        self.hakotori_dlg.destroy()

    def create_anso_dialog_off(self):
        '''モードレスダイアログボックスの作成'''
        self.dlg_modeless = tk.Toplevel(self)
        self.dlg_modeless.title("リール表裏直し")  # ウィンドウタイトル
        self.dlg_modeless.geometry("600x200")  # ウィンドウサイズ(幅x高さ)
        # back_color
        x = root.winfo_x()
        y = root.winfo_y()
        self.dlg_modeless.geometry("+%d+%d" % (x + 220, y + 250))

        self.lbl_lock = Label(self.dlg_modeless, font=("Arial", 20), text="リールの表裏を修正し、OKを押してください。")
        self.lbl_lock.place(x=0, y=0)
        self.btnkakunin_ok = tk.Button(self.dlg_modeless, font=("Arial", 15), wraplength=150, height=3, width=20, text="リール表裏修正OK", command=self.close_messenger)
        self.btnkakunin_ok.place(x=190, y=70)
        self.skip_bt = tk.Button(self.dlg_modeless, font=("Arial", 10), wraplength=150, height=2, width=8,
                                 text="スキップ", command=self.skip_messenger)
        self.skip_bt.place(x=520, y=150)
        # time.sleep(1)
        self.dlg_modeless.grab_set()  # モーダルにする
        self.dlg_modeless.focus_set()  # フォーカスを新しいウィンドウをへ移す
        self.dlg_modeless.transient(self.master)  # タスクバーに表示しない

        # ダイアログが閉じられるまで待つ
        app.wait_window(self.dlg_modeless)

    def close_messenger(self):
        self.Socket_send("ST R4115")
        self.dlg_modeless.destroy()

    def skip_messenger(self):
        self.Socket_send("ST R4115")
        self.Socket_send("ST MR3502")
        self.dlg_modeless.destroy()

    def create_anso_dialog_on(self):
        '''モーダルダイアログボックスの作成'''
        self.dlg_modal = tk.Toplevel(self)
        self.dlg_modal.title("パスコードロック")  # ウィンドウタイトル
        self.dlg_modal.geometry("400x300")  # ウィンドウサイズ(幅x高さ)
        x = root.winfo_x()
        y = root.winfo_y()
        self.dlg_modal.geometry("+%d+%d" % (x + 450, y + 200))

        #####
        lbl_lock = Label(self.dlg_modal, font=("Arial", 15), text="パスコードを入力してください")
        lbl_lock.place(x=60, y=10)
        self.password = tk.StringVar()
        self.password.set("")
        self.txt_password = tk.Entry(self.dlg_modal, justify=tk.RIGHT, width=15, font=("Arial", 20), textvariable=self.password)
        self.txt_password.place(x=90, y=50)
        self.btn_modal = tk.Button(self.dlg_modal, font=("Arial", 15), text="ログイン", command=self.passcheck)
        self.btn_modal.place(x=160, y=100)
        self.lock_stt = tk.StringVar(value="")
        self.lbl_lock_stt = Label(self.dlg_modal, font=("Arial", 15), textvariable=self.lock_stt)
        self.lbl_lock_stt.place(x=95, y=170)
        # モーダルにする設定
        self.dlg_modal.grab_set()  # モーダルにする
        self.dlg_modal.focus_set()  # フォーカスを新しいウィンドウをへ移す
        self.dlg_modal.transient(self.master)  # タスクバーに表示しない

        # ダイアログが閉じられるまで待つ
        app.wait_window(self.dlg_modal)
        print("ダイアログが閉じられた")

    def passcheck(self):
        if self.password.get() != "1122":
            self.reset_data_cam["state"] = "disable"
            self.cam_edit_hani = False;
            # self.vitri_test["state"] = "disable"

            self.lock_stt.set("パスコードが違います。")
            for child in self.labelcamset.winfo_children():
                child.configure(state='disable')
            for child in self.labelfilter.winfo_children():
                child.configure(state='disable')
            for child in self.labelfilter1.winfo_children():
                child.configure(state='disable')
            for child in self.labelfilter2.winfo_children():
                child.configure(state='disable')
            #
            for child in self.labeldetect.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect1.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect2.winfo_children():
                child.configure(state='disable')
            for child in self.labeselectdetect.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect3.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect4.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect5.winfo_children():
                child.configure(state='disable')
            for child in self.objectdetect.winfo_children():
                child.configure(state='disable')
            for child in self.objectdetect1.winfo_children():
                child.configure(state='disable')
                self.setting_lock = 1
            return
        if self.password.get() == "1122":
            self.lock_stt.set("Ok")
            self.cam_edit_hani = True;
            self.reset_data_cam["state"] = "normal"

            for child in self.labelcamset.winfo_children():
                child.configure(state='normal')
            for child in self.labelfilter.winfo_children():
                child.configure(state='normal')
            for child in self.labelfilter1.winfo_children():
                child.configure(state='normal')
            for child in self.labelfilter2.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect1.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect2.winfo_children():
                child.configure(state='normal')
            for child in self.labeselectdetect.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect3.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect4.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect5.winfo_children():
                child.configure(state='normal')
            for child in self.objectdetect.winfo_children():
                child.configure(state='normal')
            for child in self.objectdetect1.winfo_children():
                child.configure(state='normal')
            self.setting_lock = 0
            self.dlg_modal.destroy()

    def enable(childList):
        for child in childList:
            child.configure(state='enable')

    def ending_program(self):
        self.Socket_send("ST R4015")
        sys.exit()
        root.destroy()

    def end_thread(self):
        sys.exit()
        root.destroy()

    ###########################SAVE DATA###############
    def save_data(self):
        # colum1
        self.write_data(1, 1, self.IP_PLC.get())
        self.write_data(2, 1, self.PORT_PLC.get())
        self.write_data(3, 1, self.port_var.get())
        self.write_data(4, 1, self.baudrate_var.get())
        self.write_data(5, 1, self.delimiter_var.get())
        self.write_data(6, 1, self.port_var_msp.get())
        self.write_data(7, 1, self.baudrate_var_msp.get())
        self.write_data(8, 1, self.delimiter_var_msp.get())
        # colum2
        self.write_data(1, 2, self.FRAME_WIDTH_var.get())
        self.write_data(2, 2, self.FRAME_HEIGHT_var.get())
        self.write_data(3, 2, self.FPS_var.get())
        self.write_data(4, 2, self.BRIGHTNESS_var.get())
        self.write_data(5, 2, self.CONTRAST_var.get())
        self.write_data(6, 2, self.SATURATION_var.get())
        self.write_data(7, 2, self.HUE_var.get())
        self.write_data(8, 2, self.GAIN_var.get())
        self.write_data(9, 2, self.check_hani_ok.get())
        self.write_data(10, 2, self.colorchoosen.current())

        self.write_data(11, 2, self.EXPOSE_var.get())
        self.write_data(12, 2, self.auto_exposure_var.get())
        self.write_data(13, 2, self.auto_focus_var.get())

        # colum3
        self.write_data(1, 3, self.boder_filter_map.current())
        self.write_data(2, 3, self.boder_filter_map1.current())
        self.write_data(3, 3, self.scale_var.get())
        self.write_data(4, 3, self.adaptiveThreshold_var.get())
        self.write_data(5, 3, self.canny1_var.get())
        self.write_data(6, 3, self.canny2_var.get())
        self.write_data(7, 3, self.gaussion_var.get())
        self.write_data(8, 3, self.Blur_var.get())
        self.write_data(9, 3, self.bold_var.get())
        self.write_data(10, 3, self.debold_var.get())
        self.write_data(11, 3, self.bold_var1.get())
        self.write_data(12, 3, self.debold_var1.get())
        # colum4
        self.write_data(1, 4, self.so_fiter)
        self.write_data(2, 4, self.filterchoosen.current())
        self.write_data(3, 4, self.filterchoosen1.current())
        self.write_data(4, 4, self.filterchoosen2.current())
        self.write_data(5, 4, self.filterchoosen3.current())
        self.write_data(6, 4, self.filterchoosen4.current())
        self.write_data(7, 4, self.pt1[0])
        self.write_data(8, 4, self.pt1[1])
        self.write_data(9, 4, self.pt2[0])
        self.write_data(10, 4, self.pt2[1])
        #
        self.write_data(1, 14, self.pt1_specal[0])
        self.write_data(2, 14, self.pt1_specal[1])
        self.write_data(3, 14, self.pt2_specal[0])
        self.write_data(4, 14, self.pt2_specal[1])
        self.write_data(5, 14, self.pt3_specal[0])
        self.write_data(6, 14, self.pt3_specal[1])
        #
        self.write_data(7, 14, self.pt1_kaku[0])
        self.write_data(8, 14, self.pt1_kaku[1])
        self.write_data(9, 14, self.pt2_kaku[0])
        self.write_data(10, 14, self.pt2_kaku[1])
        self.write_data(11, 14, self.pt3_kaku[0])
        self.write_data(12, 14, self.pt3_kaku[1])
        self.write_data(13, 14, self.pt4_kaku[0])
        self.write_data(14, 14, self.pt4_kaku[1])
        #
        self.write_data(15, 14, self.pt1_polygol5[0])
        self.write_data(16, 14, self.pt1_polygol5[1])
        self.write_data(17, 14, self.pt2_polygol5[0])
        self.write_data(18, 14, self.pt2_polygol5[1])
        self.write_data(19, 14, self.pt3_polygol5[0])
        self.write_data(20, 14, self.pt3_polygol5[1])
        self.write_data(21, 14, self.pt4_polygol5[0])
        self.write_data(22, 14, self.pt4_polygol5[1])
        self.write_data(23, 14, self.pt5_polygol5[0])
        self.write_data(24, 14, self.pt5_polygol5[1])

        # colum5

        self.write_data(1, 5, self.fil_by_Area.get())
        self.write_data(2, 5, self.filByCircularity.get())
        self.write_data(3, 5, self.filByConvexity.get())
        self.write_data(4, 5, self.filByInertia.get())
        self.write_data(5, 5, self.minArea_var.get())
        self.write_data(6, 5, self.minCircularity_var.get())
        self.write_data(7, 5, self.minConvexity_var.get())
        self.write_data(8, 5, self.minInertia_var.get())
        # colum6
        self.write_data(1, 6, self.Kernel_var.get())
        self.write_data(2, 6, self.param1_var.get())
        self.write_data(3, 6, self.param2_var.get())
        self.write_data(4, 6, self.minRadius_var.get())
        self.write_data(5, 6, self.maxRadius_var.get())
        self.write_data(6, 6, self.var_canny_a.get())
        self.write_data(7, 6, self.var_canny_b.get())
        self.write_data(8, 6, self.var_thr.get())
        self.write_data(9, 6, self.len_size2.get())
        self.write_data(10, 6, self.len_size3.get())
        self.write_data(11, 6, self.len_size1.get())
        # colum7

        self.write_data(1, 7, self.circle_select.get())
        self.write_data(2, 7, self.circle_select_adv.get())
        self.write_data(3, 7, self.triang_select.get())
        self.write_data(4, 7, self.rect_select.get())
        self.write_data(5, 7, self.edge_select.get())
        self.write_data(6, 7, self.draw_contour_select.get())
        self.write_data(7, 7, self.connected_select.get())
        self.write_data(8, 7, self.rectangge_component_select.get())

        # colum8
        self.write_data(1, 8, self.boder_map.current())
        self.write_data(2, 8, self.contua_map.current())
        self.write_data(3, 8, self.contua_map1.current())
        self.write_data(4, 8, self.rect_para1.get())
        self.write_data(5, 8, self.area_rect.get())
        self.write_data(6, 8, self.chuvi_rect.get())
        self.write_data(7, 8, self.rect_thres.get())
        # colum9
        self.write_data(1, 9, self.boder_map2.current())
        self.write_data(2, 9, self.contua_map2.current())
        self.write_data(3, 9, self.contua_map21.current())
        self.write_data(4, 9, self.trian_para1.get())
        self.write_data(5, 9, self.area_trian.get())
        self.write_data(6, 9, self.chuvi_trian.get())
        self.write_data(7, 9, self.trian_thres.get())
        ##colum 10

        self.write_data(1, 10, self.Threshole_contua_var.get())
        self.write_data(2, 10, self.contua_aprrox.get())
        self.write_data(3, 10, self.chuvi_contua.get())
        self.write_data(4, 10, self.dientich_contua.get())
        # colum 11
        self.write_data(1, 11, self.border_connect.current())
        self.write_data(2, 11, self.kernel_connect.get())
        self.write_data(3, 11, self.iterations_var.get())
        self.write_data(4, 11, self.area_detect_connect.get())
        self.write_data(5, 11, self.thres_connect.get())
        # colum 12
        self.write_data(1, 12, self.fill_change_var.get())
        self.write_data(2, 12, self.area_patan.get())
        self.write_data(3, 12, self.chuvi_patan.get())
        self.write_data(4, 12, self.thres_patan.get())

        self.write_data(1, 15, self.check_patan1.get())
        self.write_data(2, 15, self.check_patan2.get())
        self.write_data(3, 15, self.check_patan3.get())
        self.write_data(4, 15, self.check_patan4.get())
        self.write_data(5, 15, self.check_patan5.get())
        self.write_data(6, 15, self.check_patan6.get())
        # self.write_data(2, 2, self.AEC_Metering_Mode.current())

    def read_data(self, row, colum):
        cell = self.ws.cell(row=row, column=colum).value
        return cell

    def write_data(self, row, colum, value):
        self.ws.cell(row=row, column=colum, value=value)  #
        # 保存
        self.workbook.save('data.xlsx')

    def auto_load(self):
        self.excel_path = 'data.xlsx'
        self.workbook = load_workbook(filename=self.excel_path, read_only=False)
        self.ws = self.workbook.active
        # シートのロード
        self.sheet = self.workbook['Sheet1']
        # print(self.read_data(1,1))
        ###  LOAD DATA  TAB 1 ####################
        # self.write_data(1, 1, self.IP_PLC.get())
        self.IP_PLC.set(self.read_data(1, 1))
        # self.write_data(2, 1, self.PORT_PLC.get())
        self.PORT_PLC.set(self.read_data(2, 1))
        # self.write_data(3, 1, self.port_var.get())
        self.port_var.set(self.read_data(3, 1))
        # self.write_data(4, 1, self.baudrate_var.get())
        self.baudrate_var.set(self.read_data(4, 1))
        # self.write_data(5, 1, self.delimiter_var.get())
        self.delimiter_var.set(self.read_data(5, 1))
        self.port_var_msp.set(self.read_data(6, 1))
        # self.write_data(4, 1, self.baudrate_var.get())
        self.baudrate_var_msp.set(self.read_data(7, 1))
        # self.write_data(5, 1, self.delimiter_var.get())
        self.delimiter_var_msp.set(self.read_data(8, 1))

        # colum2
        # self.write_data(1, 2, self.FRAME_WIDTH_var.get())
        self.FRAME_WIDTH_var.set(self.read_data(1, 2))
        # self.write_data(2, 2, self.FRAME_HEIGHT_var.get())
        self.FRAME_HEIGHT_var.set(self.read_data(2, 2))
        # self.write_data(3, 2, self.FPS_var.get())
        self.FPS_var.set(self.read_data(3, 2))
        # self.write_data(4, 2, self.BRIGHTNESS_var.get())
        self.BRIGHTNESS_var.set(self.read_data(4, 2))
        # self.write_data(5, 2, self.CONTRAST_var.get())
        self.CONTRAST_var.set(self.read_data(5, 2))
        # self.write_data(6, 2, self.SATURATION_var.get())
        self.SATURATION_var.set(self.read_data(6, 2))
        # self.write_data(7, 2, self.HUE_var.get())
        self.HUE_var.set(self.read_data(7, 2))
        # self.write_data(8, 2, self.GAIN_var.get())
        self.GAIN_var.set(self.read_data(8, 2))
        # self.write_data(9, 2, self.check_hani_ok.get())
        self.check_hani_ok.set(self.read_data(9, 2))
        # self.write_data(10, 2, self.colorchoosen.current())
        self.colorchoosen.current(self.read_data(10, 2))
        self.EXPOSE_var.set(self.read_data(11, 2))
        self.auto_exposure_var.set(self.read_data(12, 2))
        self.auto_focus_var.set(self.read_data(13, 2))

        # colum3
        # self.write_data(1, 3, self.boder_filter_map.current())
        self.boder_filter_map.current(self.read_data(1, 3))
        # self.write_data(2, 3, self.boder_filter_map1.current())
        self.boder_filter_map1.current(self.read_data(2, 3))
        # self.write_data(3, 3, self.scale_var.get())
        self.scale_var.set(self.read_data(3, 3))
        # self.write_data(4, 3, self.adaptiveThreshold_var.get())
        self.adaptiveThreshold_var.set(self.read_data(4, 3))
        # self.write_data(5, 3, self.canny1_var.get())
        self.canny1_var.set(self.read_data(5, 3))
        # self.write_data(6, 3, self.canny2_var.get())
        self.canny2_var.set(self.read_data(6, 3))
        # self.write_data(7, 3, self.gaussion_var.get())
        self.gaussion_var.set(self.read_data(7, 3))
        # self.write_data(8, 3, self.Blur_var.get())
        self.Blur_var.set(self.read_data(8, 3))
        # self.write_data(9, 3, self.bold_var.get())
        self.bold_var.set(self.read_data(9, 3))
        # self.write_data(10, 3, self.debold_var.get())
        self.debold_var.set(self.read_data(10, 3))
        # self.write_data(11, 3, self.bold_var1.get())
        self.bold_var1.set(self.read_data(11, 3))
        # self.write_data(12, 3, self.debold_var1.get())
        self.debold_var1.set(self.read_data(12, 3))
        # colum4
        # self.write_data(1, 4, self.so_fiter)
        self.so_fiter = self.read_data(1, 4)
        # self.write_data(2, 4, self.filterchoosen.current())
        self.filterchoosen.current(self.read_data(2, 4))
        # self.write_data(3, 4, self.filterchoosen1.current())
        self.filterchoosen1.current(self.read_data(3, 4))
        # self.write_data(4, 4, self.filterchoosen2.current())
        self.filterchoosen2.current(self.read_data(4, 4))
        # self.write_data(5, 4, self.filterchoosen3.current())
        self.filterchoosen3.current(self.read_data(5, 4))
        # self.write_data(6, 4, self.filterchoosen4.current())
        self.filterchoosen4.current(self.read_data(6, 4))
        # self.pt1[0], self.pt1[1], self.pt2[0], self.pt2[1]
        self.pt1 = [self.read_data(7, 4), self.read_data(8, 4)]
        # self.pt1[1] = self.read_data(8, 4)
        self.pt2 = [self.read_data(9, 4), self.read_data(10, 4)]
        # self.pt2[1] = self.read_data(10, 4)

        # self..set(self.read_data(7, 3))
        # colum5

        # self.write_data(1, 5, self.fil_by_Area.get())
        self.fil_by_Area.set(self.read_data(1, 5))
        # self.write_data(2, 5, self.filByCircularity.get())
        self.filByCircularity.set(self.read_data(2, 5))
        # self.write_data(3, 5, self.filByConvexity.get())
        self.filByConvexity.set(self.read_data(3, 5))
        # self.write_data(4, 5, self.filByInertia.get())
        self.filByInertia.set(self.read_data(4, 5))
        # self.write_data(5, 5, self.minArea_var.get())
        self.minArea_var.set(self.read_data(5, 5))
        # self.write_data(6, 5, self.minCircularity_var.get())
        self.minCircularity_var.set(self.read_data(6, 5))
        # self.write_data(7, 5, self.minConvexity_var.get())
        self.minConvexity_var.set(self.read_data(7, 5))
        # self.write_data(8, 5, self.minInertia_var.get())
        self.minInertia_var.set(self.read_data(7, 5))

        # # colum6
        # self.write_data(1, 6, self.Kernel_var.get())
        self.Kernel_var.set(self.read_data(1, 6))
        # self.write_data(2, 6, self.param1_var.get())
        self.param1_var.set(self.read_data(2, 6))
        # self.write_data(3, 6, self.param2_var.get())
        self.param2_var.set(self.read_data(3, 6))
        # self.write_data(4, 6, self.minRadius_var.get())
        self.minRadius_var.set(self.read_data(4, 6))
        # self.write_data(5, 6, self.maxRadius_var.get())
        self.maxRadius_var.set(self.read_data(5, 6))
        # self.write_data(6, 6, self.var_canny_a.get())
        self.var_canny_a.set(self.read_data(6, 6))
        # self.write_data(7, 6, self.var_canny_b.get())
        self.var_canny_b.set(self.read_data(7, 6))
        # self.write_data(8, 6, self.var_thr.get())
        self.var_thr.set(self.read_data(8, 6))
        # self.write_data(9, 6, self.len_size2.get())
        self.len_size2.set(self.read_data(9, 6))
        # self.write_data(10, 6, self.len_size3.get())
        self.len_size3.set(self.read_data(10, 6))
        # self.write_data(11, 6, self.len_size1.get())
        self.len_size1.set(self.read_data(11, 6))
        # # colum7
        #
        # self.write_data(1, 7, self.circle_select.get())
        self.circle_select.set(self.read_data(1, 7))
        # self.write_data(2, 7, self.circle_select_adv.get())
        self.circle_select_adv.set(self.read_data(2, 7))
        # self.write_data(3, 7, self.triang_select.get())
        self.triang_select.set(self.read_data(3, 7))
        # self.write_data(4, 7, self.rect_select.get())
        self.rect_select.set(self.read_data(4, 7))
        # self.write_data(5, 7, self.edge_select.get())
        self.edge_select.set(self.read_data(5, 7))
        # self.draw_contour_select.set(0)
        self.draw_contour_select.set(self.read_data(6, 7))
        #
        self.connected_select.set(self.read_data(7, 7))
        #
        self.rectangge_component_select.set(self.read_data(8, 7))
        # # colum8
        # self.write_data(1, 8, self.boder_map.current())
        self.boder_map.current(self.read_data(1, 8))
        # self.write_data(2, 8, self.contua_map.current())
        self.contua_map.current(self.read_data(2, 8))
        # self.write_data(3, 8, self.contua_map1.current())
        self.contua_map1.current(self.read_data(3, 8))
        # self.write_data(4, 8, self.rect_para1.get())
        self.rect_para1.set(self.read_data(4, 8))
        # self.write_data(5, 8, self.area_rect.get())
        self.area_rect.set(self.read_data(5, 8))
        # self.write_data(6, 8, self.chuvi_rect.get())
        self.chuvi_rect.set(self.read_data(6, 8))
        # self.write_data(7, 8, self.rect_thres.get())
        self.rect_thres.set(self.read_data(7, 8))
        # colum9
        # self.write_data(1, 8, self.boder_map.current())
        self.boder_map2.current(self.read_data(1, 9))
        # self.write_data(2, 8, self.contua_map.current())
        self.contua_map2.current(self.read_data(2, 9))
        # self.write_data(3, 8, self.contua_map1.current())
        self.contua_map21.current(self.read_data(3, 9))
        # self.write_data(4, 8, self.rect_para1.get())
        self.trian_para1.set(self.read_data(4, 9))
        # self.write_data(5, 8, self.area_rect.get())
        self.area_trian.set(self.read_data(5, 9))
        # self.write_data(6, 8, self.chuvi_rect.get())
        self.chuvi_trian.set(self.read_data(6, 9))
        # self.write_data(7, 8, self.rect_thres.get())
        self.trian_thres.set(self.read_data(7, 9))
        ##########
        # self.write_data(1, 10, self.Threshole_contua_var.get())
        self.Threshole_contua_var.set(self.read_data(1, 10))
        # self.write_data(2, 10, self.contua_aprrox.get())
        self.contua_aprrox.set(self.read_data(2, 10))
        # self.write_data(3, 10, self.chuvi_contua.get())
        self.chuvi_contua.set(self.read_data(3, 10))
        # self.write_data(4, 10, self.dientich_contua.get())
        self.dientich_contua.set(self.read_data(4, 10))
        # colum 11
        # self.write_data(1, 11, self.border_connect.current())
        self.border_connect.current(self.read_data(1, 11))
        # self.write_data(2, 11, self.kernel_connect.get())
        self.kernel_connect.set(self.read_data(2, 11))
        # self.write_data(3, 11, self.iterations_var.get())
        self.iterations_var.set(self.read_data(3, 11))
        # self.write_data(4, 11, self.area_detect_connect.get())
        self.area_detect_connect.set(self.read_data(4, 11))
        # self.write_data(5, 11, self.thres_connect.get())
        self.thres_connect.set(self.read_data(5, 11))
        # colum 12
        # self.write_data(1, 12, self.fill_change_var.get())
        self.fill_change_var.set(self.read_data(1, 12))
        # self.write_data(2, 12, self.area_patan.get())
        self.area_patan.set(self.read_data(2, 12))
        # self.write_data(3, 12, self.chuvi_patan.get())
        self.chuvi_patan.set(self.read_data(3, 12))
        # self.write_data(4, 12, self.thres_patan.get())
        self.thres_patan.set(self.read_data(4, 12))

        self.stt = self.read_data(1, 13)
        self.last_day_on = self.read_data(2, 13)

        # self.write_data(11, 4, self.pt1_specal[0])
        self.pt1_specal = [self.read_data(1, 14), self.read_data(2, 14)]
        self.pt2_specal = [self.read_data(3, 14), self.read_data(4, 14)]
        self.pt3_specal = [self.read_data(5, 14), self.read_data(6, 14)]

        #
        self.pt1_kaku = [self.read_data(7, 14), self.read_data(8, 14)]
        self.pt2_kaku = [self.read_data(9, 14), self.read_data(10, 14)]
        self.pt3_kaku = [self.read_data(11, 14), self.read_data(12, 14)]
        self.pt4_kaku = [self.read_data(13, 14), self.read_data(14, 14)]
        #
        #
        self.pt1_polygol5 = [self.read_data(15, 14), self.read_data(16, 14)]
        self.pt2_polygol5 = [self.read_data(17, 14), self.read_data(18, 14)]
        self.pt3_polygol5 = [self.read_data(19, 14), self.read_data(20, 14)]
        self.pt4_polygol5 = [self.read_data(21, 14), self.read_data(22, 14)]
        self.pt5_polygol5 = [self.read_data(23, 14), self.read_data(24, 14)]

        self.check_patan1.set(self.read_data(1, 15))
        self.check_patan2.set(self.read_data(2, 15))
        self.check_patan3.set(self.read_data(3, 15))
        self.check_patan4.set(self.read_data(4, 15))
        self.check_patan5.set(self.read_data(5, 15))
        self.check_patan6.set(self.read_data(6, 15))

        self.add_fiter_to()
        self.minus_fiter_to()

        self.capture.set(cv2.CAP_PROP_BRIGHTNESS, self.BRIGHTNESS_var.get())
        self.capture.set(cv2.CAP_PROP_CONTRAST, self.CONTRAST_var.get())
        self.capture.set(cv2.CAP_PROP_SATURATION, self.SATURATION_var.get())
        self.capture.set(cv2.CAP_PROP_HUE, self.HUE_var.get())
        self.capture.set(cv2.CAP_PROP_GAIN, self.GAIN_var.get())
        self.capture.set(cv2.CAP_PROP_AUTO_EXPOSURE, self.auto_exposure_var.get())
        self.capture.set(cv2.CAP_PROP_AUTOFOCUS, self.auto_focus_var.get())
        self.capture.set(cv2.CAP_PROP_EXPOSURE, self.EXPOSE_var.get())

        self.day_log()

    def on_closing(self):
        if messagebox.askokcancel("Close Program", "Do you want to quit?"):
            try:
                command = "CL"
                self.ser.write(command.encode())
            except:
                print("foo")
            self.ser.close()
            root.destroy()

    def calib_data_robot_on_starting(self):
        self.calibration = Listbox(self.tab1_rb, height=60)

        try:
            Cal = pickle.load(open("ARbot.cal", "rb"))
        except:
            Cal = "0"
            pickle.dump(Cal, open("ARbot.cal", "wb"))
        for item in Cal:
            self.calibration.insert(END, item)

        self.J1AngCur = self.calibration.get("0")
        self.J2AngCur = self.calibration.get("1")
        self.J3AngCur = self.calibration.get("2")
        self.J4AngCur = self.calibration.get("3")
        self.J5AngCur = self.calibration.get("4")
        self.J6AngCur = self.calibration.get("5")
        self.XcurPos = self.calibration.get("6")
        self.YcurPos = self.calibration.get("7")
        self.ZcurPos = self.calibration.get("8")
        self.RxcurPos = self.calibration.get("9")
        self.RycurPos = self.calibration.get("10")
        self.RzcurPos = self.calibration.get("11")
        self.comPort = self.calibration.get("12")
        self.Progr = self.calibration.get("13")
        self.Servo0on = self.calibration.get("14")
        self.Servo0off = self.calibration.get("15")
        self.Servo1on = self.calibration.get("16")
        self.Servo1off = self.calibration.get("17")
        self.DO1on = self.calibration.get("18")
        self.DO1off = self.calibration.get("19")
        self.DO2on = self.calibration.get("20")
        self.DO2off = self.calibration.get("21")
        self.TFx = self.calibration.get("22")
        self.TFy = self.calibration.get("23")
        self.TFz = self.calibration.get("24")
        self.TFrx = self.calibration.get("25")
        self.TFry = self.calibration.get("26")
        self.TFrz = self.calibration.get("27")
        self.J7PosCur = self.calibration.get("28")
        self.J8PosCur = self.calibration.get("29")
        self.J9PosCur = self.calibration.get("30")
        self.VisFileLoc = self.calibration.get("31")
        self.VisProg = self.calibration.get("32")
        self.VisOrigXpix = self.calibration.get("33")
        self.VisOrigXmm = self.calibration.get("34")
        self.VisOrigYpix = self.calibration.get("35")
        self.VisOrigYmm = self.calibration.get("36")
        self.VisEndXpix = self.calibration.get("37")
        self.VisEndXmm = self.calibration.get("38")
        self.VisEndYpix = self.calibration.get("39")
        self.VisEndYmm = self.calibration.get("40")
        self.J1calOff = self.calibration.get("41")
        self.J2calOff = self.calibration.get("42")
        self.J3calOff = self.calibration.get("43")
        self.J4calOff = self.calibration.get("44")
        self.J5calOff = self.calibration.get("45")
        self.J6calOff = self.calibration.get("46")
        self.J1OpenLoopVal = self.calibration.get("47")
        self.J2OpenLoopVal = self.calibration.get("48")
        self.J3OpenLoopVal = self.calibration.get("49")
        self.J4OpenLoopVal = self.calibration.get("50")
        self.J5OpenLoopVal = self.calibration.get("51")
        self.J6OpenLoopVal = self.calibration.get("52")
        self.com2Port = self.calibration.get("53")
        self.curTheme = self.calibration.get("54")

        self.J1CalStatVal = self.calibration.get("55")
        self.J2CalStatVal = self.calibration.get("56")
        self.J3CalStatVal = self.calibration.get("57")
        self.J4CalStatVal = self.calibration.get("58")
        self.J5CalStatVal = self.calibration.get("59")
        self.J6CalStatVal = self.calibration.get("60")
        self.J7PosLim = self.calibration.get("61")
        self.J7rotation = self.calibration.get("62")
        self.J7steps = self.calibration.get("63")
        self.J7StepCur = self.calibration.get("64")  # is this used
        self.J1CalStatVal2 = self.calibration.get("65")
        self.J2CalStatVal2 = self.calibration.get("66")
        self.J3CalStatVal2 = self.calibration.get("67")
        self.J4CalStatVal2 = self.calibration.get("68")
        self.J5CalStatVal2 = self.calibration.get("69")
        self.J6CalStatVal2 = self.calibration.get("70")
        self.VisBrightVal = self.calibration.get("71")
        self.VisContVal = self.calibration.get("72")
        self.VisBacColor = self.calibration.get("73")
        self.VisScore = self.calibration.get("74")
        self.VisX1Val = self.calibration.get("75")
        self.VisY1Val = self.calibration.get("76")
        self.VisX2Val = self.calibration.get("77")
        self.VisY2Val = self.calibration.get("78")
        self.VisRobX1Val = self.calibration.get("79")
        self.VisRobY1Val = self.calibration.get("80")
        self.VisRobX2Val = self.calibration.get("81")
        self.VisRobY2Val = self.calibration.get("82")
        self.zoom = self.calibration.get("83")
        self.pick180Val = self.calibration.get("84")
        self.pickClosestVal = self.calibration.get("85")
        self.curCam = self.calibration.get("86")
        self.fullRotVal = self.calibration.get("87")
        self.autoBGVal = self.calibration.get("88")
        mX1val = self.calibration.get("89")
        mY1val = self.calibration.get("90")
        mX2val = self.calibration.get("91")
        mY2val = self.calibration.get("92")
        self.J8length = self.calibration.get("93")
        self.J8rotation = self.calibration.get("94")
        self.J8steps = self.calibration.get("95")
        self.J9length = self.calibration.get("96")
        self.J9rotation = self.calibration.get("97")
        self.J9steps = self.calibration.get("98")
        self.J7calOff = self.calibration.get("99")
        self.J8calOff = self.calibration.get("100")
        self.J9calOff = self.calibration.get("101")
        self.GC_ST_E1 = self.calibration.get("102")
        self.GC_ST_E2 = self.calibration.get("103")
        self.GC_ST_E3 = self.calibration.get("104")
        self.GC_ST_E4 = self.calibration.get("105")
        self.GC_ST_E5 = self.calibration.get("106")
        self.GC_ST_E6 = self.calibration.get("107")
        self.GC_SToff_E1 = self.calibration.get("108")
        self.GC_SToff_E2 = self.calibration.get("109")
        self.GC_SToff_E3 = self.calibration.get("110")
        self.GC_SToff_E4 = self.calibration.get("111")
        self.GC_SToff_E5 = self.calibration.get("112")
        self.GC_SToff_E6 = self.calibration.get("113")
        self.DisableWristRotVal = self.calibration.get("114")
        self.J1MotDir = self.calibration.get("115")
        self.J2MotDir = self.calibration.get("116")
        self.J3MotDir = self.calibration.get("117")
        self.J4MotDir = self.calibration.get("118")
        self.J5MotDir = self.calibration.get("119")
        self.J6MotDir = self.calibration.get("120")
        self.J7MotDir = self.calibration.get("121")
        self.J8MotDir = self.calibration.get("122")
        self.J9MotDir = self.calibration.get("123")
        self.J1CalDir = self.calibration.get("124")
        self.J2CalDir = self.calibration.get("125")
        self.J3CalDir = self.calibration.get("126")
        self.J4CalDir = self.calibration.get("127")
        self.J5CalDir = self.calibration.get("128")
        self.J6CalDir = self.calibration.get("129")
        self.J7CalDir = self.calibration.get("130")
        self.J8CalDir = self.calibration.get("131")
        self.J9CalDir = self.calibration.get("132")
        self.J1PosLim = self.calibration.get("133")
        self.J1NegLim = self.calibration.get("134")
        self.J2PosLim = self.calibration.get("135")
        self.J2NegLim = self.calibration.get("136")
        self.J3PosLim = self.calibration.get("137")
        self.J3NegLim = self.calibration.get("138")
        self.J4PosLim = self.calibration.get("139")
        self.J4NegLim = self.calibration.get("140")
        self.J5PosLim = self.calibration.get("141")
        self.J5NegLim = self.calibration.get("142")
        self.J6PosLim = self.calibration.get("143")
        self.J6NegLim = self.calibration.get("144")
        self.J1StepDeg = self.calibration.get("145")
        self.J2StepDeg = self.calibration.get("146")
        self.J3StepDeg = self.calibration.get("147")
        self.J4StepDeg = self.calibration.get("148")
        self.J5StepDeg = self.calibration.get("149")
        self.J6StepDeg = self.calibration.get("150")
        self.J1DriveMS = self.calibration.get("151")
        self.J2DriveMS = self.calibration.get("152")
        self.J3DriveMS = self.calibration.get("153")
        self.J4DriveMS = self.calibration.get("154")
        self.J5DriveMS = self.calibration.get("155")
        self.J6DriveMS = self.calibration.get("156")
        self.J1EncCPR = self.calibration.get("157")
        self.J2EncCPR = self.calibration.get("158")
        self.J3EncCPR = self.calibration.get("159")
        self.J4EncCPR = self.calibration.get("160")
        self.J5EncCPR = self.calibration.get("161")
        self.J6EncCPR = self.calibration.get("162")
        self.J1ΘDHpar = self.calibration.get("163")
        self.J2ΘDHpar = self.calibration.get("164")
        self.J3ΘDHpar = self.calibration.get("165")
        self.J4ΘDHpar = self.calibration.get("166")
        self.J5ΘDHpar = self.calibration.get("167")
        self.J6ΘDHpar = self.calibration.get("168")
        self.J1αDHpar = self.calibration.get("169")
        self.J2αDHpar = self.calibration.get("170")
        self.J3αDHpar = self.calibration.get("171")
        self.J4αDHpar = self.calibration.get("172")
        self.J5αDHpar = self.calibration.get("173")
        self.J6αDHpar = self.calibration.get("174")
        self.J1dDHpar = self.calibration.get("175")
        self.J2dDHpar = self.calibration.get("176")
        self.J3dDHpar = self.calibration.get("177")
        self.J4dDHpar = self.calibration.get("178")
        self.J5dDHpar = self.calibration.get("179")
        self.J6dDHpar = self.calibration.get("180")
        self.J1aDHpar = self.calibration.get("181")
        self.J2aDHpar = self.calibration.get("182")
        self.J3aDHpar = self.calibration.get("183")
        self.J4aDHpar = self.calibration.get("184")
        self.J5aDHpar = self.calibration.get("185")
        self.J6aDHpar = self.calibration.get("186")
        self.GC_ST_WC = self.calibration.get("187")
        self.J7CalStatVal = self.calibration.get("188")
        self.J8CalStatVal = self.calibration.get("189")
        self.J9CalStatVal = self.calibration.get("190")
        self.J7CalStatVal2 = self.calibration.get("191")
        self.J8CalStatVal2 = self.calibration.get("192")
        self.J9CalStatVal2 = self.calibration.get("193")

        ####

        self.incrementEntryField.insert(0, "10")
        self.speedEntryField.insert(0, "25")
        self.ACCspeedField.insert(0, "20")
        self.DECspeedField.insert(0, "20")
        self.ACCrampField.insert(0, "100")
        self.roundEntryField.insert(0, "0")

        self.SavePosEntryField.insert(0, "1")
        self.R1EntryField.insert(0, "0")
        self.R2EntryField.insert(0, "0")
        self.R3EntryField.insert(0, "0")
        self.R4EntryField.insert(0, "0")
        self.R5EntryField.insert(0, "0")
        self.R6EntryField.insert(0, "0")
        self.R7EntryField.insert(0, "0")
        self.R8EntryField.insert(0, "0")
        self.R9EntryField.insert(0, "0")
        self.R10EntryField.insert(0, "0")
        self.R11EntryField.insert(0, "0")
        self.R12EntryField.insert(0, "0")
        self.R13EntryField.insert(0, "0")
        self.R14EntryField.insert(0, "0")
        self.R15EntryField.insert(0, "0")
        self.R16EntryField.insert(0, "0")
        self.SP_1_E1_EntryField.insert(0, "0")
        self.SP_2_E1_EntryField.insert(0, "0")
        self.SP_3_E1_EntryField.insert(0, "0")
        self.SP_4_E1_EntryField.insert(0, "0")
        self.SP_5_E1_EntryField.insert(0, "0")
        self.SP_6_E1_EntryField.insert(0, "0")
        self.SP_7_E1_EntryField.insert(0, "0")
        self.SP_8_E1_EntryField.insert(0, "0")
        self.SP_9_E1_EntryField.insert(0, "0")
        self.SP_10_E1_EntryField.insert(0, "0")
        self.SP_11_E1_EntryField.insert(0, "0")
        self.SP_12_E1_EntryField.insert(0, "0")
        self.SP_13_E1_EntryField.insert(0, "0")
        self.SP_14_E1_EntryField.insert(0, "0")
        self.SP_15_E1_EntryField.insert(0, "0")
        self.SP_16_E1_EntryField.insert(0, "0")
        self.SP_1_E2_EntryField.insert(0, "0")
        self.SP_2_E2_EntryField.insert(0, "0")
        self.SP_3_E2_EntryField.insert(0, "0")
        self.SP_4_E2_EntryField.insert(0, "0")
        self.SP_5_E2_EntryField.insert(0, "0")
        self.SP_6_E2_EntryField.insert(0, "0")
        self.SP_7_E2_EntryField.insert(0, "0")
        self.SP_8_E2_EntryField.insert(0, "0")
        self.SP_9_E2_EntryField.insert(0, "0")
        self.SP_10_E2_EntryField.insert(0, "0")
        self.SP_11_E2_EntryField.insert(0, "0")
        self.SP_12_E2_EntryField.insert(0, "0")
        self.SP_13_E2_EntryField.insert(0, "0")
        self.SP_14_E2_EntryField.insert(0, "0")
        self.SP_15_E2_EntryField.insert(0, "0")
        self.SP_16_E2_EntryField.insert(0, "0")
        self.SP_1_E3_EntryField.insert(0, "0")
        self.SP_2_E3_EntryField.insert(0, "0")
        self.SP_3_E3_EntryField.insert(0, "0")
        self.SP_4_E3_EntryField.insert(0, "0")
        self.SP_5_E3_EntryField.insert(0, "0")
        self.SP_6_E3_EntryField.insert(0, "0")
        self.SP_7_E3_EntryField.insert(0, "0")
        self.SP_8_E3_EntryField.insert(0, "0")
        self.SP_9_E3_EntryField.insert(0, "0")
        self.SP_10_E3_EntryField.insert(0, "0")
        self.SP_11_E3_EntryField.insert(0, "0")
        self.SP_12_E3_EntryField.insert(0, "0")
        self.SP_13_E3_EntryField.insert(0, "0")
        self.SP_14_E3_EntryField.insert(0, "0")
        self.SP_15_E3_EntryField.insert(0, "0")
        self.SP_16_E3_EntryField.insert(0, "0")
        self.SP_1_E4_EntryField.insert(0, "0")
        self.SP_2_E4_EntryField.insert(0, "0")
        self.SP_3_E4_EntryField.insert(0, "0")
        self.SP_4_E4_EntryField.insert(0, "0")
        self.SP_5_E4_EntryField.insert(0, "0")
        self.SP_6_E4_EntryField.insert(0, "0")
        self.SP_7_E4_EntryField.insert(0, "0")
        self.SP_8_E4_EntryField.insert(0, "0")
        self.SP_9_E4_EntryField.insert(0, "0")
        self.SP_10_E4_EntryField.insert(0, "0")
        self.SP_11_E4_EntryField.insert(0, "0")
        self.SP_12_E4_EntryField.insert(0, "0")
        self.SP_13_E4_EntryField.insert(0, "0")
        self.SP_14_E4_EntryField.insert(0, "0")
        self.SP_15_E4_EntryField.insert(0, "0")
        self.SP_16_E4_EntryField.insert(0, "0")
        self.SP_1_E5_EntryField.insert(0, "0")
        self.SP_2_E5_EntryField.insert(0, "0")
        self.SP_3_E5_EntryField.insert(0, "0")
        self.SP_4_E5_EntryField.insert(0, "0")
        self.SP_5_E5_EntryField.insert(0, "0")
        self.SP_6_E5_EntryField.insert(0, "0")
        self.SP_7_E5_EntryField.insert(0, "0")
        self.SP_8_E5_EntryField.insert(0, "0")
        self.SP_9_E5_EntryField.insert(0, "0")
        self.SP_10_E5_EntryField.insert(0, "0")
        self.SP_11_E5_EntryField.insert(0, "0")
        self.SP_12_E5_EntryField.insert(0, "0")
        self.SP_13_E5_EntryField.insert(0, "0")
        self.SP_14_E5_EntryField.insert(0, "0")
        self.SP_15_E5_EntryField.insert(0, "0")
        self.SP_16_E5_EntryField.insert(0, "0")
        self.SP_1_E6_EntryField.insert(0, "0")
        self.SP_2_E6_EntryField.insert(0, "0")
        self.SP_3_E6_EntryField.insert(0, "0")
        self.SP_4_E6_EntryField.insert(0, "0")
        self.SP_5_E6_EntryField.insert(0, "0")
        self.SP_6_E6_EntryField.insert(0, "0")
        self.SP_7_E6_EntryField.insert(0, "0")
        self.SP_8_E6_EntryField.insert(0, "0")
        self.SP_9_E6_EntryField.insert(0, "0")
        self.SP_10_E6_EntryField.insert(0, "0")
        self.SP_11_E6_EntryField.insert(0, "0")
        self.SP_12_E6_EntryField.insert(0, "0")
        self.SP_13_E6_EntryField.insert(0, "0")
        self.SP_14_E6_EntryField.insert(0, "0")
        self.SP_15_E6_EntryField.insert(0, "0")
        self.SP_16_E6_EntryField.insert(0, "0")
        self.servo0onEntryField.insert(0, str(self.Servo0on))
        self.servo0offEntryField.insert(0, str(self.Servo0off))
        self.servo1onEntryField.insert(0, str(self.Servo1on))
        self.servo1offEntryField.insert(0, str(self.Servo1off))
        self.DO1onEntryField.insert(0, str(self.DO1on))
        self.DO1offEntryField.insert(0, str(self.DO1off))
        self.DO2onEntryField.insert(0, str(self.DO2on))
        self.DO2offEntryField.insert(0, str(self.DO2off))
        self.TFxEntryField.insert(0, str(self.TFx))
        self.TFyEntryField.insert(0, str(self.TFy))
        self.TFzEntryField.insert(0, str(self.TFz))
        self.TFrxEntryField.insert(0, str(self.TFrx))
        self.TFryEntryField.insert(0, str(self.TFry))
        self.TFrzEntryField.insert(0, str(self.TFrz))
        self.J7curAngEntryField.insert(0, str(self.J7PosCur))
        self.J8curAngEntryField.insert(0, str(self.J8PosCur))
        self.J9curAngEntryField.insert(0, str(self.J9PosCur))
        self.J1calOffEntryField.insert(0, str(self.J1calOff))
        self.J2calOffEntryField.insert(0, str(self.J2calOff))
        self.J3calOffEntryField.insert(0, str(self.J3calOff))
        self.J4calOffEntryField.insert(0, str(self.J4calOff))
        self.J5calOffEntryField.insert(0, str(self.J5calOff))
        self.J6calOffEntryField.insert(0, str(self.J6calOff))
        self.J7calOffEntryField.insert(0, str(self.J7calOff))
        self.J8calOffEntryField.insert(0, str(self.J8calOff))
        self.J9calOffEntryField.insert(0, str(self.J9calOff))
        if (self.J1OpenLoopVal == 1):
            self.J1OpenLoopStat.set(True)
        if (self.J2OpenLoopVal == 1):
            self.J2OpenLoopStat.set(True)
        if (self.J3OpenLoopVal == 1):
            self.J3OpenLoopStat.set(True)
        if (self.J4OpenLoopVal == 1):
            self.J4OpenLoopStat.set(True)
        if (self.J5OpenLoopVal == 1):
            self.J5OpenLoopStat.set(True)
        if (self.J6OpenLoopVal == 1):
            self.J6OpenLoopStat.set(True)
        if (self.DisableWristRotVal == 1):
            self.DisableWristRot.set(True)

        if (self.J1CalStatVal == 1):
            self.J1CalStat.set(True)
        if (self.J2CalStatVal == 1):
            self.J2CalStat.set(True)
        if (self.J3CalStatVal == 1):
            self.J3CalStat.set(True)
        if (self.J4CalStatVal == 1):
            self.J4CalStat.set(True)
        if (self.J5CalStatVal == 1):
            self.J5CalStat.set(True)
        if (self.J6CalStatVal == 1):
            self.J6CalStat.set(True)
        if (self.J7CalStatVal == 1):
            self.J7CalStat.set(True)
        if (self.J8CalStatVal == 1):
            self.J8CalStat.set(True)
        if (self.J9CalStatVal == 1):
            self.J9CalStat.set(True)
        if (self.J1CalStatVal2 == 1):
            self.J1CalStat2.set(True)
        if (self.J2CalStatVal2 == 1):
            self.J2CalStat2.set(True)
        if (self.J3CalStatVal2 == 1):
            self.J3CalStat2.set(True)
        if (self.J4CalStatVal2 == 1):
            self.J4CalStat2.set(True)
        if (self.J5CalStatVal2 == 1):
            self.J5CalStat2.set(True)
        if (self.J6CalStatVal2 == 1):
            self.J6CalStat2.set(True)
        if (self.J7CalStatVal2 == 1):
            self.J7CalStat2.set(True)
        if (self.J8CalStatVal2 == 1):
            self.J8CalStat2.set(True)
        if (self.J9CalStatVal2 == 1):
            self.J9CalStat2.set(True)
        self.axis7lengthEntryField.insert(0, str(self.J7PosLim))
        self.axis7rotEntryField.insert(0, str(self.J7rotation))
        self.axis7stepsEntryField.insert(0, str(self.J7steps))
        self.VisBrightSlide.set(self.VisBrightVal)
        self.VisContrastSlide.set(self.VisContVal)
        self.VisBacColorEntryField.insert(0, str(self.VisBacColor))
        self.VisScoreEntryField.insert(0, str(self.VisScore))
        self.VisX1PixEntryField.insert(0, str(self.VisX1Val))
        self.VisY1PixEntryField.insert(0, str(self.VisY1Val))
        self.VisX2PixEntryField.insert(0, str(self.VisX2Val))
        self.VisY2PixEntryField.insert(0, str(self.VisY2Val))
        self.VisX1RobEntryField.insert(0, str(self.VisRobX1Val))
        self.VisY1RobEntryField.insert(0, str(self.VisRobY1Val))
        self.VisX2RobEntryField.insert(0, str(self.VisRobX2Val))
        self.VisY2RobEntryField.insert(0, str(self.VisRobY2Val))
        self.VisZoomSlide.set(self.zoom)
        if (self.pickClosestVal == 1):
            self.pickClosest.set(True)
        if (self.pick180Val == 1):
            self.pick180.set(True)
        self.Jvisoptions.set(self.curCam)
        if (self.fullRotVal == 1):
            self.fullRot.set(True)
        if (self.autoBGVal == 1):
            self.autoBG.set(True)
        self.mX1 = mX1val
        self.mY1 = mY1val
        self.mX2 = mX2val
        self.mY2 = mY2val
        self.axis8lengthEntryField.insert(0, str(self.J8length))
        self.axis8rotEntryField.insert(0, str(self.J8rotation))
        self.axis8stepsEntryField.insert(0, str(self.J8steps))
        self.axis9lengthEntryField.insert(0, str(self.J9length))
        self.axis9rotEntryField.insert(0, str(self.J9rotation))
        self.axis9stepsEntryField.insert(0, str(self.J9steps))
        self.GC_ST_E1_EntryField.insert(0, str(self.GC_ST_E1))
        self.GC_ST_E2_EntryField.insert(0, str(self.GC_ST_E2))
        self.GC_ST_E3_EntryField.insert(0, str(self.GC_ST_E3))
        self.GC_ST_E4_EntryField.insert(0, str(self.GC_ST_E4))
        self.GC_ST_E5_EntryField.insert(0, str(self.GC_ST_E5))
        self.GC_ST_E6_EntryField.insert(0, str(self.GC_ST_E6))
        self.GC_ST_WC_EntryField.insert(0, str(self.GC_ST_WC))
        self.GC_SToff_E1_EntryField.insert(0, str(self.GC_SToff_E1))
        self.GC_SToff_E2_EntryField.insert(0, str(self.GC_SToff_E2))
        self.GC_SToff_E3_EntryField.insert(0, str(self.GC_SToff_E3))
        self.GC_SToff_E4_EntryField.insert(0, str(self.GC_SToff_E4))
        self.GC_SToff_E5_EntryField.insert(0, str(self.GC_SToff_E5))
        self.GC_SToff_E6_EntryField.insert(0, str(self.GC_SToff_E6))
        self.J1MotDirEntryField.insert(0, str(self.J1MotDir))
        self.J2MotDirEntryField.insert(0, str(self.J2MotDir))
        self.J3MotDirEntryField.insert(0, str(self.J3MotDir))
        self.J4MotDirEntryField.insert(0, str(self.J4MotDir))
        self.J5MotDirEntryField.insert(0, str(self.J5MotDir))
        self.J6MotDirEntryField.insert(0, str(self.J6MotDir))
        self.J7MotDirEntryField.insert(0, str(self.J7MotDir))
        self.J8MotDirEntryField.insert(0, str(self.J8MotDir))
        self.J9MotDirEntryField.insert(0, str(self.J9MotDir))
        self.J1CalDirEntryField.insert(0, str(self.J1CalDir))
        self.J2CalDirEntryField.insert(0, str(self.J2CalDir))
        self.J3CalDirEntryField.insert(0, str(self.J3CalDir))
        self.J4CalDirEntryField.insert(0, str(self.J4CalDir))
        self.J5CalDirEntryField.insert(0, str(self.J5CalDir))
        self.J6CalDirEntryField.insert(0, str(self.J6CalDir))
        self.J7CalDirEntryField.insert(0, str(self.J7CalDir))
        self.J8CalDirEntryField.insert(0, str(self.J8CalDir))
        self.J9CalDirEntryField.insert(0, str(self.J9CalDir))
        self.J1PosLimEntryField.insert(0, str(self.J1PosLim))
        self.J1NegLimEntryField.insert(0, str(self.J1NegLim))
        self.J2PosLimEntryField.insert(0, str(self.J2PosLim))
        self.J2NegLimEntryField.insert(0, str(self.J2NegLim))
        self.J3PosLimEntryField.insert(0, str(self.J3PosLim))
        self.J3NegLimEntryField.insert(0, str(self.J3NegLim))
        self.J4PosLimEntryField.insert(0, str(self.J4PosLim))
        self.J4NegLimEntryField.insert(0, str(self.J4NegLim))
        self.J5PosLimEntryField.insert(0, str(self.J5PosLim))
        self.J5NegLimEntryField.insert(0, str(self.J5NegLim))
        self.J6PosLimEntryField.insert(0, str(self.J6PosLim))
        self.J6NegLimEntryField.insert(0, str(self.J6NegLim))
        self.J1StepDegEntryField.insert(0, str(self.J1StepDeg))
        self.J2StepDegEntryField.insert(0, str(self.J2StepDeg))
        self.J3StepDegEntryField.insert(0, str(self.J3StepDeg))
        self.J4StepDegEntryField.insert(0, str(self.J4StepDeg))
        self.J5StepDegEntryField.insert(0, str(self.J5StepDeg))
        self.J6StepDegEntryField.insert(0, str(self.J6StepDeg))
        self.J1DriveMSEntryField.insert(0, str(self.J1DriveMS))
        self.J2DriveMSEntryField.insert(0, str(self.J2DriveMS))
        self.J3DriveMSEntryField.insert(0, str(self.J3DriveMS))
        self.J4DriveMSEntryField.insert(0, str(self.J4DriveMS))
        self.J5DriveMSEntryField.insert(0, str(self.J5DriveMS))
        self.J6DriveMSEntryField.insert(0, str(self.J6DriveMS))
        self.J1EncCPREntryField.insert(0, str(self.J1EncCPR))
        self.J2EncCPREntryField.insert(0, str(self.J2EncCPR))
        self.J3EncCPREntryField.insert(0, str(self.J3EncCPR))
        self.J4EncCPREntryField.insert(0, str(self.J4EncCPR))
        self.J5EncCPREntryField.insert(0, str(self.J5EncCPR))
        self.J6EncCPREntryField.insert(0, str(self.J6EncCPR))
        self.J1ΘEntryField.insert(0, str(self.J1ΘDHpar))
        self.J2ΘEntryField.insert(0, str(self.J2ΘDHpar))
        self.J3ΘEntryField.insert(0, str(self.J3ΘDHpar))
        self.J4ΘEntryField.insert(0, str(self.J4ΘDHpar))
        self.J5ΘEntryField.insert(0, str(self.J5ΘDHpar))
        self.J6ΘEntryField.insert(0, str(self.J6ΘDHpar))
        self.J1αEntryField.insert(0, str(self.J1αDHpar))
        self.J2αEntryField.insert(0, str(self.J2αDHpar))
        self.J3αEntryField.insert(0, str(self.J3αDHpar))
        self.J4αEntryField.insert(0, str(self.J4αDHpar))
        self.J5αEntryField.insert(0, str(self.J5αDHpar))
        self.J6αEntryField.insert(0, str(self.J6αDHpar))
        self.J1dEntryField.insert(0, str(self.J1dDHpar))
        self.J2dEntryField.insert(0, str(self.J2dDHpar))
        self.J3dEntryField.insert(0, str(self.J3dDHpar))
        self.J4dEntryField.insert(0, str(self.J4dDHpar))
        self.J5dEntryField.insert(0, str(self.J5dDHpar))
        self.J6dEntryField.insert(0, str(self.J6dDHpar))
        self.J1aEntryField.insert(0, str(self.J1aDHpar))
        self.J2aEntryField.insert(0, str(self.J2aDHpar))
        self.J3aEntryField.insert(0, str(self.J3aDHpar))
        self.J4aEntryField.insert(0, str(self.J4aDHpar))
        self.J5aEntryField.insert(0, str(self.J5aDHpar))
        self.J6aEntryField.insert(0, str(self.J6aDHpar))


if __name__ == "__main__":
    root = tk.Tk()
    app = Application(master=root)
    # app.Disp_draw()
    # draw_thread = threading.Thread(target=app.Disp_draw)
    # draw_thread.start()
    app.auto_load()
    # app.camera_callback()
    app.calib_data_robot_on_starting()
    app.updateVisOp()
    # app.autoconnect_comport()
    # app.autoconnect_msp_comport_M1_2()
    # app.autoconnect_msp_comport_M3_4()
    # app.autoconnect_msp_comport_M5_6()
    # app.hakotori_kakunin()
    # app.camera_callback()
    # app.socket_connect()
    # time.sleep(10)
    # app.redraw_image()
    app.mainloop()
