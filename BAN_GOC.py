import multiprocessing as mp
import tkinter.simpledialog as simpledialog
import tkinter.messagebox as messagebox
import socket
# from socket import socket, AF_INET, SOCK_STREAM
# socket
from typing import List, Optional
import serial.tools.list_ports
from serial import Serial
# serial
import tkinter as tk  # ウィンドウ作成用
from tkinter import filedialog  # ファイルを開くダイアログ用
from tkinter import *
from tkinter import ttk
import tkinter.font as ft
from tkinter import scrolledtext
from openpyxl import load_workbook
from tkinter import scrolledtext
from PIL import Image, ImageTk, ImageOps
import numpy as np  # アフィン変換行列演算用
import os  # ディレクトリ操作用
import threading
import time
import cv2
import math
import datetime
import sys
import struct
from copy import deepcopy
from random import randint
import display_error

import BAN_GOC
HOST = '192.168.0.10'
PORT = 8501
HOST1 = '192.168.0.11'
PORT1 = 8501

MAX_MESSAGE = 2048
NUM_THREAD = 1

CHR_CAN = '\18'
CHR_EOT = '\04'

# ボーレートのリスト
BAUDRATES = [2400, 4800, 9600, 19200, 38400, 56000, 57600, 115200]
# デリミタのリスト
DELIMITERS = {'[CR]': '\r', '[LF]': '\n', '[CRLF]': '\r\n', '[None]': ''}

# デリミタのリスト(履歴表示用)
REV_DELIMITER = {'\r': '[CR]', '\n': '[LF]', '\r\n': '[CRLF]', '': ''}
MODE_RUN = ["MODE_IO", "MODE_PULSE", "MODE_PC"]
NUM_MOTOR = ["ONE_MOTOR", "TWO_MOTOR"]
################################# MSP ###############################
MSP_HEADER = "$M>"
MSP_MOTOR_MEOBIET = 199
MSP_MOTOR_MODE = 100
MSP_MOTOR_POS = 101
MSP_MOTOR_VEL = 102
MSP_MOTOR_ACCEL_DECCEL = 103
MSP_MOTOR_VOL = 104
MSP_MOTOR_VOLLIMIT = 105
MSP_MOTOR_PID_P = 106
MSP_MOTOR_PID_VEL_P = 107
MSP_MOTOR_PID_VEL_I = 108
MSP_MOTOR_PID_VEL_D = 109

MSP_MOTOR_MOTOR_SET = 1
MSP_MOTOR_MODE_SET = 2
MSP_MOTOR_SAVE_SET = 3
MSP_MOTOR_RESET_SET = 4
MSP_MOTOR_RUN_SET = 5
MSP_MOTOR_ERROR_SET = 6  # new
MSP_MOTOR_NC_SET = 7  # new
MSP_MOTOR_DATAPOS1_SET = 8
MSP_MOTOR_DATAPOS2_SET = 9
MSP_MOTOR_PARASAVE1_SET = 34
MSP_MOTOR_PARASAVE2_SET = 35

MSP_MOTOR_POS11_SET = 10
MSP_MOTOR_POS12_SET = 11
MSP_MOTOR_POS21_SET = 12
MSP_MOTOR_POS22_SET = 13

MSP_MOTOR_VEL11_SET = 14
MSP_MOTOR_VEL12_SET = 15
MSP_MOTOR_VEL21_SET = 16
MSP_MOTOR_VEL22_SET = 17

MSP_MOTOR_ACCEL1_SET = 18
MSP_MOTOR_ACCEL2_SET = 19
MSP_MOTOR_DECCEL1_SET = 20
MSP_MOTOR_DECCEL2_SET = 21

MSP_MOTOR_VOL_M1_SET = 22
MSP_MOTOR_VOL_M2_SET = 23

MSP_MOTOR_VOLLIMIT_M1_SET = 24
MSP_MOTOR_VOLLIMIT_M2_SET = 25
MSP_MOTOR_PID_P_M1_SET = 26
MSP_MOTOR_PID_P_M2_SET = 27
MSP_MOTOR_PID_VEL_P_M1_SET = 28
MSP_MOTOR_PID_VEL_P_M2_SET = 29
MSP_MOTOR_PID_VEL_I_M1_SET = 30
MSP_MOTOR_PID_VEL_I_M2_SET = 31
MSP_MOTOR_PID_VEL_D_M1_SET = 32
MSP_MOTOR_PID_VEL_D_M2_SET = 33
MSP_MOTOR_FB_ICHI = 50
MSP_MOTOR_FB_VEL_ACC_DEC_M1 = 51
MSP_MOTOR_FB_VEL_ACC_DEC_M2 = 52

MSP_MOTOR_IDOU_1_SET = 60
MSP_MOTOR_IDOU_2_SET = 61
MSP_MOTOR_ACC_DEC_1_SET = 62
MSP_MOTOR_ACC_DEC_2_SET = 63
MSP_MOTOR_SPEED_1_SET = 64
MSP_MOTOR_SPEED_2_SET = 65

MSP_MOTOR_DISABLE_MOTOR_SET = 66
MSP_MOTOR_RESO1_SET = 67
MSP_MOTOR_RESO2_SET = 68
MSP_MOTOR_POLE_1_SET = 69
MSP_MOTOR_POLE_2_SET = 70
MSP_MOTOR_MODE_RUN = 71
MSP_MOTOR_MOTION_CONTROL1 = 72
MSP_MOTOR_MOTION_CONTROL2 = 73

IDLE = 0
HEADER_START = 1
HEADER_M = 2
HEADER_ARROW = 3
HEADER_SIZE = 4
HEADER_CMD = 5
HEADER_ERR = 6

c_state = IDLE
err_rcvd = False

checksum = 0
cmd = 0
offset = 0
dataSize = 0
inBuf = bytearray(256)

mode = 0
p = 0
payload = []
present = 0

time2 = 0
time3 = 0
time4 = 0
time5 = 0
time6 = 0

import decimal
df = decimal.Decimal('0.00')
global error, run_mode_val, cur_motor_number
global motor1_resolution, motor2_resolution, motor1_pole, motor2_pole
global data_motor1, data_motor2, control_motion_1, control_motion_2
global current_pos_11, current_pos_12, current_pos_21, current_pos_22
global current_vel11, current_vel12, current_vel21, current_vel22
global current_accel1, current_accel2, current_deccel1, current_deccel2
global current_vol_in1, current_vol_in2
global current_vol_limit1, current_vol_limit2
global pid_pos_p1, pid_pos_p2
global pid_vel_p1, pid_vel_p2
global pid_vel_i1, pid_vel_i2
global pid_vel_d1, pid_vel_d2
global position_now_motor1, position_now_motor2
global now_velocity_1, now_acceleration_1, now_decceleration_1
global now_velocity_2, now_acceleration_2, now_decceleration_2


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)


        self.IP_PLC = '192.168.0.10'
        self.PORT_PLC = 8501
        self.pack()
        # self.member1 = Winzoom()
        # self.root.iconbitmap('c:/gui/codemy.ico')
        #font
        self.my_font = ft.Font(self.master, family="System", size=14, weight="bold")
        self.sys_font = ft.Font(self.master, family="System", size=30, weight="bold")
        self.sysbol_font = ft.Font(self.master, family="System", size=45)
        self.small_font = ft.Font(self.master, family="System", size=6)
        #
        self.my_title = "ROBOT_ARM"  # タイトル
        self.back_color = "#DBFFFF"  # 背景色
        # self.pil_img = Image.open("label4.PNG")
        self.status = str("normal")
        # ウィンドウの設定
        self.master.title(self.my_title)  # タイトル
        self.master.geometry("1024x700")  # サイズ
        #self.master.geometry("%dx%d" % (root.winfo_screenwidth(), root.winfo_screenheight()))
        # self.slave.geometry("400x400")
        self.point =0
        self.i_couter =0
        self.i_couter_1 = 0
        self.i_couter_2 = 0
        # Create workbook EXCEL instance
        # エクセルファイルのロード
        excel_path = 'data.xlsx'
        self.workbook = load_workbook(filename=excel_path, read_only=False)
        self.ws = self.workbook.active
        # シートのロード
        self.sheet = self.workbook['Sheet1']
        self.cam_edit_hani =False
        self.capture = cv2.VideoCapture(0)
        self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
        self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        # Scale（オプションをいくつか設定）
        self.loop_display = 0
        self.pil_image = None  # 表示する画像データ
        self.filename = None  # 最後に開いた画像ファイル名
        self.DM1000_var = tk.DoubleVar()
        self.DM1002_var = tk.DoubleVar()
        self.DM1004_var = tk.DoubleVar()
        self.DM1006_var = tk.DoubleVar()
        self.DM1008_var = tk.DoubleVar()
        self.DM1010_var = tk.DoubleVar()
        self.DM1012_var = tk.DoubleVar()
        self.DM1014_var = tk.DoubleVar()
        self.DM1016_var = tk.DoubleVar()
        self.DM1018_var = tk.DoubleVar()
        self.DM1020_var = tk.DoubleVar()
        self.now_area_check = tk.StringVar(value="00")
        self.now_chuvi_check = tk.StringVar(value="00")
        self.prty_val = 'N'
        self.create_menu()  # メニューの作成
        # self.create_treeview()
        self.creat_textbox()
        self.create_widget()  # ウィジェットの作成
        self.create_arm_setting()
        self.create_camera_processing()
        self.count = 0
        self.record = 0
        self.Camera_check = 0
        self.disp_id = None
        self.MSP_id  = None
        self.MSP_id_1 = None
        self.MSP_id_2 = None
        self.comport = None
        self.communicate_on = False
        self.have_contour = 0
        self.ser = Serial()
        self.ser.close()
        self.ser_msp = Serial()
        self.ser_msp.close()
        self.ser_msp_1 = Serial()
        self.ser_msp_1.close()
        self.ser_msp_2 = Serial()
        self.ser_msp_2.close()
        self.refresh_serial_ports()
        self.refresh_serial_M1_2()
        self.refresh_serial_M3_4()
        self.refresh_serial_M5_6()
        self.recv_id = None
        self.draw_id = None
        self.status = 0
        self.error_status = 0
        self.toggle_blink = 0
        self.syudou_kirikaeis_on = True
        self.ryru_hansoubelt3is_on = True
        self.ryru_hansoubelt4is_on = True
        self.ryru_checkis_on = True
        self.ryru_check1is_on = True
        self.ryru_check2is_on = True
        self.ryru_check9is_on = True
        self.ryru_check10is_on = True
        self.ryru_hariis_on = True
        self.ryru_hari1is_on = True
        self.ryru_hari2is_on = True
        self.ryru_hari3is_on = True
        self.ryru_hari4is_on = True
        self.ryru_hari5is_on = True
        self.ryru_hari6is_on = True
        self.ryru_hari7is_on = True
        self.ryru_hari8is_on = True
        #
        self.ryru_check_bar1is_on = True
        self.ryru_check_bar2is_on = True
        self.ryru_check_bar3is_on = True
        self.ryru_check_bar4is_on = True
        self.ryru_check_bar5is_on = True
        self.ryru_check_bar6is_on = True
        self.ryru_check_bar7is_on = True
        self.ryru_check_bar9is_on = True
        self.hakokyokyu2is_on = True
        self.hakokyokyu3is_on = True
        self.hakokyokyu4is_on = True
        self.hakokyokyu5is_on = True
        self.hakokyokyu6is_on = True
        self.hakokyokyu7is_on = True
        self.hakokyokyu8is_on = True
        self.hakokyokyu9is_on = True
        self.hakokyokyu10is_on = True
        self.hakokyokyu11_1is_on = True
        self.hakokyokyu11_2is_on = True
        self.hakokyokyu11_3is_on = True
        self.hakokyokyu12is_on = True
        self.hakokyokyu13is_on = True
        self.hakokyokyu17is_on = True
        self.hakokyokyu18is_on = True
        self.hakokyokyu19is_on = True
        self.hakokyokyu20is_on = True
        self.hakokyokyu21is_on = True
        self.hakokyokyu22is_on = True
        self.hakokyokyu23is_on = True
        self.hakokyokyu24is_on = True
        self.hakokyokyu27is_on = True
        self.hakokyokyu29is_on = True
        self.hakokyokyu30is_on = True
        self.hakokyokyu31is_on = True
        self.hakokyokyu33is_on = True
        self.hakokyokyu34is_on = True
        self.hakokyokyu35is_on = True
        self.print_riruumuis_on = True
        self.print_hakoumuis_on = True
        self.capture_switchis_on = False
        #
        self.time_lock = 0
        self.update_main_id = None
        self.RIRU_NUMBER = tk.DoubleVar()  ##ERROR STATUS
        self.KIKAI_var = tk.DoubleVar()  ##KIKAI STATUS
        self.FLAG_Barcode = tk.DoubleVar()  ##FLAG STATUS
        self.FLAG_Camera = tk.DoubleVar()  ##SENSOR STATUS
        self.YOBI_var = tk.DoubleVar()  ##YOBI STATUS
        self.print_data1_var = tk.DoubleVar()  ##print1 STATUS
        self.print_data2_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data3_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data4_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data5_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data6_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data7_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data8_var = tk.DoubleVar()  ##print2 STATUS
        self.print_data9_var = tk.DoubleVar()  ##print2 STATUS
        self.send_flag = 0
        self.data_old = ""
        self.data_old1 = ""
        self.data_old2 = ""
        self.data_old_a = ""
        self.data_old_b = ""
        self.data_old_c = ""
        self.scanned = False
        self.barcoder_data = ""
        self.ONE_scan = False
        self.cv_image = None
        self.read_defaut = False
        self.zoom_load_defaut = False
        # CAMERA
        self.Camera_kido = 0
        self.mat_affine = np.eye(3)
        self.mat_affine_main = np.eye(3)
        self.mat_affine_sita = np.eye(3)
        self.mat_affine_hoshu1 = np.eye(3)
        self.mat_affine_hoshu2 = np.eye(3)
        self.pt1 = (20, 20)
        self.pt2 = (600, 400)
        self.edit_pt1 = False
        self.edit_pt2 = False

        self.pt1_specal = (100, 200)
        self.pt2_specal = (100, 300)
        self.pt3_specal = (200, 250)
        self.edit_pt1_specal = False
        self.edit_pt2_specal = False
        self.edit_pt3_specal = False
        self.pt1_kaku = (200,200)
        self.pt2_kaku = (400, 200)
        self.pt3_kaku = (400, 400)
        self.pt4_kaku = (200, 400)
        self.edit_pt1_kaku = False
        self.edit_pt2_kaku = False
        self.edit_pt3_kaku = False
        self.edit_pt4_kaku = False
        self.pt1_polygol5 =(200,200)
        self.pt2_polygol5 = (300, 100)
        self.pt3_polygol5 = (400, 200)
        self.pt4_polygol5 = (300, 300)
        self.pt5_polygol5 = (200, 400)
        self.edit_pt1_polygol5 = False
        self.edit_pt2_polygol5 = False
        self.edit_pt3_polygol5 = False
        self.edit_pt4_polygol5 = False
        self.edit_pt5_polygol5 = False

        self.so_fiter = 1
        self.setting_lock = 1
        self.tab_select = 0
        self.PLC_CN_OK = 0
        self.stt = 0
        self.last_day_on ='2024_07_25'
        self.hako_tori_kakunin = 0
        self.now_area = 0
        self.now_chuvi = 0
    # -------------------------------------------------------------------------------
    def day_log(self):

        self.day_month_year = datetime.datetime.now().strftime('%Y_%m_%d')
        if self.day_month_year != self.last_day_on:
            self.stt = self.stt +1
            self.last_day_on = self.day_month_year
            self.write_data(1, 13, self.stt)
            self.write_data(2, 13, self.last_day_on)

        self.name_log = self.day_month_year + '_' + str(self.stt) + '.txt'
        print(self.name_log)
    def testDevice(source):
        cap = cv2.VideoCapture(source)
        if cap is None or not cap.isOpened():
            print('Warning: unable to open video source: ', source)
    def uint_to_float(self,uint32_value):
        packed = struct.pack('I', uint32_value)  # 'I' is the format for unsigned 32-bit integer
        float_value = struct.unpack('f', packed)[0]  # 'f' is the format for a 32-bit float
        return float_value
    # ------------------------------------SCREEN HAICHI-----------------------------------
    def menu_open_clicked(self, event=None):
        print("「ファイルを開く」が選択された")
        self.filename = filedialog.askopenfilename(
            filetypes=[("file", ".xlsx .xlsm .xltx .xltm")],  # ファイルフィルタ
            title="ファイルを開く",
            initialdir="./"  # 自分自身のディレクトリ
        )
        self.excel_path = self.filename
        self.workbook = load_workbook(filename=self.excel_path, read_only=False)
        self.ws = self.workbook.active
        # シートのロード
        self.sheet = self.workbook['Sheet1']

        print(self.filename)
        # print(self.read_data(1,1))
        ###  LOAD DATA  TAB 1 ####################
        '''self.resolution_h_var.set(self.read_data(1, 1))
        self.resolution_w_var.set(self.read_data(2, 1))
        self.preview_check_var.set(self.read_data(3, 1))
        self.hdr_var.set(self.read_data(4, 1))
        self.saturation_var.set(self.read_data(5, 1))
        self.contrast_var.set(self.read_data(6, 1))
        self.sharpness_var.set(self.read_data(7, 1))
        self.brightness_var.set(self.read_data(8, 1))
        # hdr
        self.num_hdr_var.set(self.read_data(9, 1, ))
        self.stops_hdr_above_var.set(self.read_data(10, 1))
        self.stops_hdr_below_var.set(self.read_data(11, 1))
        self.hdr_gamma_var.set(self.read_data(12, 1))'''

        #
    def menu_reload_clicked(self, event=None):
        self.status = str("disabled")

        ''''''
    def menu_quit_clicked(self):
        # ウィンドウを閉じる
        self.master.destroy()

        # -------------------------------------------------------------------------------
    def creat_textbox(self):
        '''lbl_sock = tk.Label(self.master, text="PLC socket")
        lbl_sock.place(x=5, y=250)
        lbl_comp = tk.Label(self.master, text="Serial")
        lbl_comp.place(x=5, y=420)
        self.Text_sock_recieved = Text(self.master, height=11, width=28, background='#856ff8')
        self.Text_sock_recieved.place(x=0, y=270)
        self.received_text = tk.Text(self.master, height=11, width=28, background='#856ff8')
        self.received_text.place(x=0, y=440)
        clear_button = tk.Button(self.master, text="履歴クリア", command=self.clear_received_text)
        clear_button.place(x=137, y=560)'''
    def create_treeview(self):
        left_frame = tk.Frame(self.master, relief=tk.SUNKEN, bd=2, width=200, height=500)
        left_frame.pack(side=tk.LEFT, anchor=NW, fill=tk.X)
        # Creating treeview window
        '''self.treeview = ttk.Treeview(left_frame)

        # Calling pack method on the treeview
        self.treeview.pack(side=tk.RIGHT)
        self.treeview.insert('', '0', 'item1',
                             text='Motor parameter', values=1)

        # Inserting child
        self.treeview.insert('', '1', 'item2',
                             text='Motor_asix1', values=2)
        self.treeview.insert('', '2', 'item3',
                             text='Motor_asix2', values=3)
        self.treeview.insert('', 'end', 'item4',
                             text='Motor_asix3', values=4)

        # Inserting more than one attribute of an item
        self.treeview.insert('item2', 'end', 'Orignal',
                             text='Vol input asix1', values=5)
        self.treeview.insert('item2', 'end', 'fix_pixel',
                             text='基板_Area', values=6)
        self.treeview.insert('item2', 'end', 'fix_point',
                             text='記号_Area', values=7)
        self.treeview.insert('item2', 'end', 'fix_grey_pixel',
                             text='C_面_Area', values=8)
        self.treeview.insert('item3', 'end', 'point check1',
                             text='Vol input asix2', values=9)
        self.treeview.insert('item3', 'end', 'point check2',
                             text='point check　●', values=10)
        self.treeview.insert('item3', 'end', 'point check3',
                             text='point check　▲', values=11)
        self.treeview.insert('item3', 'end', 'point check4',
                             text='point check　C面', values=12)
        self.treeview.insert('item4', 'end', 'Check all Raw_Colum',
                             text='Check all Raw Colum', values=13)
        self.treeview.insert('item4', 'end', 'Check center',
                             text='Check all center', values=14)
        self.treeview.insert('item4', 'end', 'Check パタン',
                             text='Check パタン', values=15)

        # Placing each child items in parent widget
        self.treeview.move('item2', 'item1', 'end')
        self.treeview.move('item3', 'item1', 'end')
        self.treeview.move('item4', 'item1', 'end')
        self.treeview.bind('<<TreeviewSelect>>', self.item_selected)'''
    def item_selected(self, event):
        for selected_item in self.treeview.selection():
            item = self.treeview.item(selected_item)
            self.record = item['values']
            # show a message
            # print(self.record)
    def create_menu(self):
        self.menu_bar = tk.Menu(self)  # Menuクラスからmenu_barインスタンスを生成

        self.file_menu = tk.Menu(self.menu_bar, tearoff=tk.OFF)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        self.file_menu.add_command(label="Open", command=self.menu_open_clicked, accelerator="Ctrl+O")
        self.file_menu.add_command(label="ReLoad_data", command=self.auto_load)
        self.file_menu.add_command(label="Save_data", command=self.save_data)
        self.file_menu.add_separator()  # セパレーターを追加
        self.file_menu.add_command(label="Exit", command=self.menu_quit_clicked)

        self.menu_bar.bind_all("<Control-o>", self.menu_open_clicked)  # ファイルを開くのショートカット(Ctrol-Oボタン)

        self.master.config(menu=self.menu_bar)  # メニューバーの配置
    def create_widget(self):
        #Variable
        font = ('Yu Gothic UI', 11)
        style = ttk.Style()
        style.configure('TNotebook.Tab', font=font, padding=(8, 12))
        self.text_var = tk.StringVar(value="A")
        self.error_code = tk.IntVar()
        self.error_code.set(0)
        self.riru_ok_var = tk.StringVar(value="E")
        self.riru_ng_var = tk.StringVar(value="D")
        self.hako_ok_var = tk.StringVar(value="C")
        self.lot_number = tk.StringVar(value="B")
        self.BAR_check = tk.StringVar(value="F")
        self.IP_PLC = tk.StringVar()
        self.IP_PLC.set("192.168.0.10")
        self.PORT_PLC = tk.IntVar()
        self.PORT_PLC.set(8501)
        self.port_var = tk.StringVar(root)
        self.baudrate_var = tk.StringVar(root)
        self.baudrate_var.set(BAUDRATES[4])
        self.delimiter_var = tk.StringVar(root)
        self.delimiter_var.set(REV_DELIMITER['\r'])
        ##msp
        self.port_var_msp = tk.StringVar(root)
        self.baudrate_var_msp = tk.StringVar(root)
        self.baudrate_var_msp.set(BAUDRATES[4])
        self.delimiter_var_msp = tk.StringVar(root)
        self.delimiter_var_msp.set(REV_DELIMITER['\r'])
        ##msp1
        self.port_var_msp_1 = tk.StringVar(root)
        self.baudrate_var_msp_1 = tk.StringVar(root)
        self.baudrate_var_msp_1.set(BAUDRATES[4])
        self.delimiter_var_msp_1 = tk.StringVar(root)
        self.delimiter_var_msp_1.set(REV_DELIMITER['\r'])
        ##msp2
        self.port_var_msp_2 = tk.StringVar(root)
        self.baudrate_var_msp_2 = tk.StringVar(root)
        self.baudrate_var_msp_2.set(BAUDRATES[4])
        self.delimiter_var_msp_2 = tk.StringVar(root)
        self.delimiter_var_msp_2.set(REV_DELIMITER['\r'])
        #
        self.mode_run_var =tk.StringVar(root)
        self.mode_run_var.set(MODE_RUN[2])
        self.num_motor_var =tk.StringVar(root)
        self.num_motor_var.set(NUM_MOTOR[1])
        #
        self.mode_run_var_1 = tk.StringVar(root)
        self.mode_run_var_1.set(MODE_RUN[2])
        self.num_motor_var_1 = tk.StringVar(root)
        self.num_motor_var_1.set(NUM_MOTOR[1])
        #
        self.mode_run_var_2 = tk.StringVar(root)
        self.mode_run_var_2.set(MODE_RUN[2])
        self.num_motor_var_2 = tk.StringVar(root)
        self.num_motor_var_2.set(NUM_MOTOR[1])
        #
        self.text_var1 = tk.StringVar(value="000")
        self.text_var2 = tk.StringVar(value="000")
        self.DM1000_var.set(0)
        self.DM1002_var.set(0)
        self.riru_henko = tk.IntVar()
        self.FRAME_WIDTH_var = tk.DoubleVar()
        self.FRAME_HEIGHT_var = tk.DoubleVar()
        self.FPS_var = tk.DoubleVar()
        self.BRIGHTNESS_var = tk.DoubleVar()
        self.CONTRAST_var = tk.DoubleVar()
        self.SATURATION_var = tk.DoubleVar()
        self.HUE_var = tk.DoubleVar()
        self.GAIN_var = tk.DoubleVar()
        self.EXPOSE_var = tk.DoubleVar()
        self.auto_exposure_var = tk.IntVar()
        self.auto_exposure_var.set(0)
        self.auto_focus_var = tk.IntVar()
        self.auto_focus_var.set(0)
        self.canny1_var = tk.DoubleVar()
        self.canny1_var.set(100)
        self.canny2_var = tk.DoubleVar()
        self.canny2_var.set(500)
        self.gaussion_var = tk.DoubleVar()
        self.Blur_var = tk.DoubleVar()
        self.bold_var = tk.IntVar()
        self.debold_var = tk.IntVar()
        self.bold_var1 = tk.IntVar()
        self.debold_var1 = tk.IntVar()
        self.filtermap = tk.StringVar()
        self.filtermap1 = tk.StringVar()
        self.filtermap2 = tk.StringVar()
        self.filtermap3 = tk.StringVar()
        self.filtermap4 = tk.StringVar()
        self.check_hani_ok = tk.IntVar()

        self.check_patan1 = tk.IntVar()
        self.check_patan2 = tk.IntVar()
        self.check_patan3 = tk.IntVar()
        self.check_patan4 = tk.IntVar()
        self.check_patan5 = tk.IntVar()
        self.check_patan6 = tk.IntVar()
        self.check_hani_ok.set(1)
        self.colormap = tk.StringVar()
        self.boder_filter_map_sl = tk.StringVar()
        self.scale_var = tk.DoubleVar()
        self.scale_var.set(50)
        self.adaptiveThreshold_var = tk.DoubleVar()
        self.boder_filter_map_sl1 = tk.StringVar()
        #####################################################
        # ステータスバー相当(親に追加)
        self.statusbar = tk.Frame(self.master)
        self.mouse_position = tk.Label(self.statusbar, relief=tk.SUNKEN, text="mouse position")  # マウスの座標
        self.image_position = tk.Label(self.statusbar, relief=tk.SUNKEN, text="image position")  # 画像の座標
        self.label_space = tk.Label(self.statusbar, relief=tk.SUNKEN)  # 隙間を埋めるだけ
        self.image_info = tk.Label(self.statusbar, relief=tk.SUNKEN, text="info")  # 画像情報
        self.mouse_position.pack(side=tk.LEFT)
        self.image_position.pack(side=tk.LEFT)
        self.label_space.pack(side=tk.LEFT, expand=True, fill=tk.X)
        self.image_info.pack(side=tk.RIGHT)
        self.statusbar.pack(side=tk.BOTTOM, fill=tk.X)
        #####################WINDOW2##################################
        # ButtonWINDOW2

        #####################################################
        # 右側フレーム（画像処理用ボタン配置用）
        right_frame = tk.Frame(self.master, relief=tk.SUNKEN, bd=2, width=1020)
        right_frame.pack(side=tk.RIGHT, fill=tk.Y)

        self.tab_control = ttk.Notebook(right_frame, height=768, width=1020)
        self.tab0 = ttk.Frame(self.tab_control)
        self.tab1 = ttk.Frame(self.tab_control)
        self.tab2 = ttk.Frame(self.tab_control)
        self.tab3 = ttk.Frame(self.tab_control)
        self.tab4 = ttk.Frame(self.tab_control)
        self.tab5 = ttk.Frame(self.tab_control)
        self.tab6 = ttk.Frame(self.tab_control)
        self.tab7 = ttk.Frame(self.tab_control)
        self.tab8 = ttk.Frame(self.tab_control)

        self.tab_control.add(self.tab0, state=self.status, text='メイン')
        self.tab_control.add(self.tab1, state=self.status, text='アーム設定')
        self.tab_control.add(self.tab2, state=self.status, text='保守１')
        self.tab_control.add(self.tab3, state=self.status, text='保守２')
        self.tab_control.add(self.tab4, state=self.status, text='設定１')
        self.tab_control.add(self.tab5, state=self.status, text='設定２')
        self.tab_control.add(self.tab6, state=self.status, text='通信')
        self.tab_control.add(self.tab7, state=self.status, text='カメラ')
        self.tab_control.add(self.tab8, state=self.status, text='点検')
        self.tab_control.place(x=0, y=0)
        self.tab_control.bind("<<NotebookTabChanged>>", self.gettab_select)

        # TAB 0======================================================================


        lbl_error = tk.Label(self.tab0, text="EEROR")
        lbl_error.place(x=505, y=0)

        self.syudou_kirikae = tk.Button(right_frame, command=self.syudou_kirikae_switch, height=2, width=11,text="BT_3", bg="gray")
        self.karabako_okuri = tk.Button(right_frame, command=lambda: self.Socket_send("ST R3011"), height=2, width=11,text="BT_2", bg="#fdd2b9")
        self.riru_haraidasi = tk.Button(right_frame, command=lambda: self.Socket_send("ST R3012"), height=2, width=11,text="BT_1", bg="#faa755")
        self.syudou_kirikae.place(x=916, y=4)
        self.karabako_okuri.place(x=810, y=4)
        self.riru_haraidasi.place(x=702, y=4)

        self.canvas =        tk.Canvas(self.tab0, background="#C3C2C4", width=500, height=340)  #
        self.canvas_sita =   tk.Canvas(self.tab0, background="#C3C2C4", width=500, height=330)  #
        self.canvas_hoshu1 = tk.Canvas(self.tab2, background="#C3C2C4", width=1020, height=630)  #
        self.canvas_hoshu2 = tk.Canvas(self.tab3, background="#C3C2C4", width=1020, height=630)  #
        self.canvas.place(x=0, y=0)
        self.canvas_sita.place(x=0, y=340)
        self.canvas_hoshu1.place(x=0, y=0)
        self.canvas_hoshu2.place(x=0, y=0)

        self.riru_number_couter_lb = tk.LabelFrame(self.tab0, text="GROUP0", labelanchor="nw", width=100,height=130,font=self.small_font)
        self.riru_number_couter_lb.propagate(False)
        self.PLC_error_text=scrolledtext.ScrolledText(self.tab0,wrap=tk.WORD,width=27,height=8,font=("Arial", 19))
        self.PLC_error_annai = scrolledtext.ScrolledText(self.tab0, wrap=tk.WORD, width=27, height=7, font=("Arial", 19))
        self.lbl_riru_number = tk.Label(self.riru_number_couter_lb, textvariable=self.text_var,font=("Arial", 60, "bold"), bg="yellow")
        self.clear_error = tk.Button(self.tab0, text="履歴クリア", command=self.clear_error_text, width=16, height=2,bg="gray")
        lbl_code = Label(self.tab0, text="エラーCODE")
        txt_address = tk.Entry(self.tab0, justify=tk.RIGHT, width=15, textvariable=self.error_code)
        self.start_button = tk.Button(self.tab0, text="START", height=3, width=6, bd=6, bg='gray', font=self.sys_font,fg="lightgreen",command=lambda: self.Socket_send("ST R3000"))
        self.stop_button = tk.Button(self.tab0, text="STOP", height=3, width=6, bd=6, bg='gray', font=self.sys_font,fg="red", command=lambda: self.Socket_send("ST R3001"))
        self.buzz_button = tk.Button(self.tab0,wraplength = 90, text="LOT RESET", height=3, width=6, bd=6, bg='gray', font=self.sys_font,fg="blue", command=lambda: self.Socket_send("ST R3005"))
        self.reset_button = tk.Button(self.tab0, text="RESET", height=3, width=6, bd=6, bg='gray', font=self.sys_font,fg="yellow",command=self.reset_kikai)
        #position
        self.riru_number_couter_lb.place(x=910, y=320)
        self.PLC_error_text.place(x=505, y=20)
        self.PLC_error_annai.place(x=505, y=250)
        self.lbl_riru_number.place(x=16, y=3)
        self.clear_error.place(x=872, y=450)
        lbl_code.place(x=580, y=460)
        txt_address.place(x=650, y=460)
        self.start_button.place(x=510, y=490)
        self.stop_button.place(x=630, y=490)
        self.buzz_button.place(x=750, y=490)
        self.reset_button.place(x=870, y=490)

        self.labelhonsu1 = tk.LabelFrame(self.tab0, text="GROUP3", labelanchor="nw", width=100, height=80)
        self.labelhonsu2 = tk.LabelFrame(self.tab0, text="GROUP5", labelanchor="nw", width=100, height=80)
        self.labelhonsu3 = tk.LabelFrame(self.tab0, text="GROUP4", labelanchor="nw", width=100, height=80)
        self.labelhonsu4 = tk.LabelFrame(self.tab0, text="GROUP6", labelanchor="nw", width=100, height=80)
        self.ARM1_2 = tk.LabelFrame(self.tab1,      text="SETTING_ARM_1_2", labelanchor="nw", width=1000, height=200, font=self.small_font)
        self.ARM3_4 = tk.LabelFrame(self.tab1,      text="SETTING_ARM_3_4", labelanchor="nw", width=1000, height=200, font=self.small_font)
        self.ARM5_6 = tk.LabelFrame(self.tab1,      text="SETTING_ARM_5_6", labelanchor="nw", width=1000, height=200, font=self.small_font)
        self.labelhonsu1.propagate(False)
        self.labelhonsu2.propagate(False)
        self.labelhonsu3.propagate(False)
        self.labelhonsu4.propagate(False)
        self.ARM1_2.propagate(False)
        self.ARM3_4.propagate(False)
        self.ARM5_6.propagate(False)
        #########################################
        self.para_edit_1_is_on =False
        self.para_edit_2_is_on = False
        self.para_edit_3_is_on = False
        self.pid_vp_1 = tk.DoubleVar()
        self.pid_vp_1.set(0.00)
        self.pid_pp_1 = tk.DoubleVar()
        self.pid_pp_1.set(0.00)
        self.pid_vi_1 = tk.DoubleVar()
        self.pid_vi_1.set(0.00)
        self.pid_vd_1 = tk.DoubleVar()
        self.pid_vd_1.set(0.00)
        self.volin_1 = tk.DoubleVar()
        self.volin_1.set(0.00)
        self.vollimit_1 = tk.DoubleVar()
        self.vollimit_1.set(0.00)
        self.pid_vp_2 = tk.DoubleVar()
        self.pid_vp_2.set(0.00)
        self.pid_pp_2 = tk.DoubleVar()
        self.pid_pp_2.set(0.00)
        self.pid_vi_2 = tk.DoubleVar()
        self.pid_vi_2.set(0.00)
        self.pid_vd_2 = tk.DoubleVar()
        self.pid_vd_2.set(0.00)
        self.volin_2 = tk.DoubleVar()
        self.volin_2.set(0.00)
        self.vollimit_2 = tk.DoubleVar()
        self.vollimit_2.set(0.00)
        self.pid_vp_3 = tk.DoubleVar()
        self.pid_vp_3.set(0.00)
        self.pid_pp_3 = tk.DoubleVar()
        self.pid_pp_3.set(0.00)
        self.pid_vi_3 = tk.DoubleVar()
        self.pid_vi_3.set(0.00)
        self.pid_vd_3 = tk.DoubleVar()
        self.pid_vd_3.set(0.00)
        self.volin_3 = tk.DoubleVar()
        self.volin_3.set(0.00)
        self.vollimit_3 = tk.DoubleVar()
        self.vollimit_3.set(0.00)
        self.pid_vp_4 = tk.DoubleVar()
        self.pid_vp_4.set(0.00)
        self.pid_pp_4 = tk.DoubleVar()
        self.pid_pp_4.set(0.00)
        self.pid_vi_4 = tk.DoubleVar()
        self.pid_vi_4.set(0.00)
        self.pid_vd_4 = tk.DoubleVar()
        self.pid_vd_4.set(0.00)
        self.volin_4 = tk.DoubleVar()
        self.volin_4.set(0.00)
        self.vollimit_4 = tk.DoubleVar()
        self.vollimit_4.set(0.00)
        self.pid_vp_5 = tk.DoubleVar()
        self.pid_vp_5.set(0.00)
        self.pid_pp_5 = tk.DoubleVar()
        self.pid_pp_5.set(0.00)
        self.pid_vi_5 = tk.DoubleVar()
        self.pid_vi_5.set(0.00)
        self.pid_vd_5 = tk.DoubleVar()
        self.pid_vd_5.set(0.00)
        self.volin_5 = tk.DoubleVar()
        self.volin_5.set(0.00)
        self.vollimit_5 = tk.DoubleVar()
        self.vollimit_5.set(0.00)
        self.pid_vp_6 = tk.DoubleVar()
        self.pid_vp_6.set(0.00)
        self.pid_pp_6 = tk.DoubleVar()
        self.pid_pp_6.set(0.00)
        self.pid_vi_6 = tk.DoubleVar()
        self.pid_vi_6.set(0.00)
        self.pid_vd_6 = tk.DoubleVar()
        self.pid_vd_6.set(0.00)
        self.volin_6 = tk.DoubleVar()
        self.volin_6.set(0.00)
        self.vollimit_6 = tk.DoubleVar()
        self.vollimit_6.set(0.00)
        ############################### SETTING_ARM_1_2 ###############################
        lbl_mode = Label(self.ARM1_2, text="MODE_選択")
        self.mode_menu = tk.OptionMenu(self.ARM1_2, self.mode_run_var, *MODE_RUN)#self.mode_run_var/self.num_motor_var
        lbl_num_motor = Label(self.ARM1_2, text="NUM_MOTOR_選択")
        self.num_motor = tk.OptionMenu(self.ARM1_2, self.num_motor_var, *NUM_MOTOR)
        self.toggle_servo_1 = tk.Button(self.ARM1_2, height=1, width=10, text="SERVO_1(Off)",background="#FFC652")
        self.toggle_servo_2 = tk.Button(self.ARM1_2, height=1, width=10, text="SERVO_2(Off)",background="#FFC652")
        self.group_pid_1 = tk.LabelFrame(self.ARM1_2, text="arm_1_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_1.propagate(False)
        lbl_pid_p1 = Label(self.group_pid_1, text="PID_position:")
        self.spinbox_PID_PP_1 = ttk.Spinbox(self.group_pid_1,from_=0, to=100, increment=0.01,width=8,textvariable=self.pid_pp_1)
        lbl_pid_vp1 = Label(self.group_pid_1, text="PID_(P)velocity:")
        self.spinbox_PID_VP_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vp_1)
        lbl_vol_in1 = Label(self.group_pid_1, text="VOL_input:")
        self.spinbox_VOL_IN1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8,textvariable=self.volin_1)
        lbl_pid_vi1 = Label(self.group_pid_1, text="PID_(I)velocity:")
        self.spinbox_PID_VI_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vi_1)
        lbl_vol_limit1 = Label(self.group_pid_1, text="VOL_limit:")
        self.spinbox_VOL_LIMIT1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.01, width=8,textvariable=self.vollimit_1)
        lbl_pid_vd1 = Label(self.group_pid_1, text="PID_(D)velocity:")
        self.spinbox_PID_VD_1 = ttk.Spinbox(self.group_pid_1, from_=0, to=100, increment=0.001, width=8,textvariable=self.pid_vd_1)
        #
        self.group_pid_2 = tk.LabelFrame(self.ARM1_2, text="arm_2_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_2.propagate(False)
        lbl_pid_p2 = Label(self.group_pid_2, text="PID_position:")
        self.spinbox_PID_PP_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_pp_2)
        lbl_pid_vp2 = Label(self.group_pid_2, text="PID_(P)velocity:")
        self.spinbox_PID_VP_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vp_2)
        lbl_vol_in2 = Label(self.group_pid_2, text="VOL_input:")
        self.spinbox_VOL_IN2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8,textvariable=self.volin_2)
        lbl_pid_vi2 = Label(self.group_pid_2, text="PID_(I)velocity:")
        self.spinbox_PID_VI_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vi_2)
        lbl_vol_limit2 = Label(self.group_pid_2, text="VOL_limit:")
        self.spinbox_VOL_LIMIT2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.01, width=8,textvariable=self.vollimit_2)
        lbl_pid_vd2 = Label(self.group_pid_2, text="PID_(D)velocity:")
        self.spinbox_PID_VD_2 = ttk.Spinbox(self.group_pid_2, from_=0, to=100, increment=0.001, width=8,textvariable=self.pid_vd_2)

        self.save_para_m1 = tk.Button(self.ARM1_2, height=1, width=6, text="Save M1", background="#AD8EE3",command=self.save_para_1)
        self.save_para_m2 = tk.Button(self.ARM1_2, height=1, width=6, text="Save M2", background="#AD8EE3", command=self.save_para_2)
        self.edit_para_m1_2 = tk.Button(self.ARM1_2, height=1, width=4, text="Edit", background="#AD8EE3",command=self.edit_1_switch)
        #1
        lbl_mode.place(x=560, y=-5)
        self.mode_menu.place(x=640, y=-10)
        lbl_num_motor.place(x=760, y=-5)
        self.num_motor.place(x=870, y=-10)
        self.toggle_servo_1.place(x=150, y=-5)
        self.toggle_servo_2.place(x=250, y=-5)
        self.group_pid_1.place(x=450, y=20)
        lbl_pid_p1.place(x=0, y=0)
        self.spinbox_PID_PP_1.place(x=85, y=0)
        lbl_pid_vp1.place(x=0, y=30)
        self.spinbox_PID_VP_1.place(x=85, y=30)
        lbl_vol_in1.place(x=150, y=0)
        self.spinbox_VOL_IN1.place(x=235, y=0)
        lbl_pid_vi1.place(x=150, y=30)
        self.spinbox_PID_VI_1.place(x=235, y=30)
        lbl_vol_limit1.place(x=300, y=0)
        self.spinbox_VOL_LIMIT1.place(x=385, y=0)
        lbl_pid_vd1.place(x=300, y=30)
        self.spinbox_PID_VD_1.place(x=385, y=30)
        #2
        self.group_pid_2.place(x=450, y=100)
        lbl_pid_p2.place(x=0, y=0)
        self.spinbox_PID_PP_2.place(x=85, y=0)
        lbl_pid_vp2.place(x=0, y=30)
        self.spinbox_PID_VP_2.place(x=85, y=30)
        lbl_vol_in2.place(x=150, y=0)
        self.spinbox_VOL_IN2.place(x=235, y=0)
        lbl_pid_vi2.place(x=150, y=30)
        self.spinbox_PID_VI_2.place(x=235, y=30)
        lbl_vol_limit2.place(x=300, y=0)
        self.spinbox_VOL_LIMIT2.place(x=385, y=0)
        lbl_pid_vd2.place(x=300, y=30)
        self.spinbox_PID_VD_2.place(x=385, y=30)

        self.save_para_m1.place(x=941, y=120)
        self.save_para_m2.place(x=941, y=148)
        self.edit_para_m1_2.place(x=945, y=90)
        #################################################################################
        ############################### SETTING_ARM_3_4 ###############################
        lbl_mode_1 = Label(self.ARM3_4, text="MODE_選択")
        self.mode_menu_1 = tk.OptionMenu(self.ARM3_4, self.mode_run_var_1, *MODE_RUN)  # self.mode_run_var/self.num_motor_var
        lbl_num_motor_1 = Label(self.ARM3_4, text="NUM_MOTOR_選択")
        self.num_motor_1 = tk.OptionMenu(self.ARM3_4, self.num_motor_var_1, *NUM_MOTOR)
        self.toggle_servo_3 = tk.Button(self.ARM3_4, height=1, width=10, text="SERVO_3(Off)", background="#FFC652")
        self.toggle_servo_4 = tk.Button(self.ARM3_4, height=1, width=10, text="SERVO_4(Off)", background="#FFC652")
        self.group_pid_3 = tk.LabelFrame(self.ARM3_4, text="arm_3_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_3.propagate(False)
        lbl_pid_p3 = Label(self.group_pid_3, text="PID_position:")
        self.spinbox_PID_PP_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_pp_3)
        lbl_pid_vp3 = Label(self.group_pid_3, text="PID_(P)velocity:")
        self.spinbox_PID_VP_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vp_3)
        lbl_vol_in3 = Label(self.group_pid_3, text="VOL_input:")
        self.spinbox_VOL_IN3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8,textvariable=self.volin_3)
        lbl_pid_vi3 = Label(self.group_pid_3, text="PID_(I)velocity:")
        self.spinbox_PID_VI_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vi_3)
        lbl_vol_limit3 = Label(self.group_pid_3, text="VOL_limit:")
        self.spinbox_VOL_LIMIT3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.01, width=8,textvariable=self.vollimit_3)
        lbl_pid_vd3 = Label(self.group_pid_3, text="PID_(D)velocity:")
        self.spinbox_PID_VD_3 = ttk.Spinbox(self.group_pid_3, from_=0, to=100, increment=0.001, width=8,textvariable=self.pid_vd_3)
        #
        self.group_pid_4 = tk.LabelFrame(self.ARM3_4, text="arm_4_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_4.propagate(False)
        lbl_pid_p4 = Label(self.group_pid_4, text="PID_position:")
        self.spinbox_PID_PP_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_pp_4)
        lbl_pid_vp4 = Label(self.group_pid_4, text="PID_(P)velocity:")
        self.spinbox_PID_VP_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vp_4)
        lbl_vol_in4 = Label(self.group_pid_4, text="VOL_input:")
        self.spinbox_VOL_IN4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8,textvariable=self.volin_4)
        lbl_pid_vi4 = Label(self.group_pid_4, text="PID_(I)velocity:")
        self.spinbox_PID_VI_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vi_4)
        lbl_vol_limit4 = Label(self.group_pid_4, text="VOL_limit:")
        self.spinbox_VOL_LIMIT4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.01, width=8,textvariable=self.vollimit_4)
        lbl_pid_vd4 = Label(self.group_pid_4, text="PID_(D)velocity:")
        self.spinbox_PID_VD_4 = ttk.Spinbox(self.group_pid_4, from_=0, to=100, increment=0.001, width=8,textvariable=self.pid_vd_4)

        self.save_para_m3 = tk.Button(self.ARM3_4, height=1, width=6, text="Save M3", background="#AD8EE3",command=self.save_para_3)
        self.save_para_m4 = tk.Button(self.ARM3_4, height=1, width=6, text="Save M4", background="#AD8EE3", command=self.save_para_4)
        self.edit_para_m3_4 = tk.Button(self.ARM3_4, height=1, width=4, text="Edit", background="#AD8EE3",command=self.edit_2_switch)
        #3
        lbl_mode_1.place(x=560, y=-5)
        self.mode_menu_1.place(x=640, y=-10)
        lbl_num_motor_1.place(x=760, y=-5)
        self.num_motor_1.place(x=870, y=-10)
        self.toggle_servo_3.place(x=150, y=-5)
        self.toggle_servo_4.place(x=250, y=-5)
        self.group_pid_3.place(x=450, y=20)
        lbl_pid_p3.place(x=0, y=0)
        self.spinbox_PID_PP_3.place(x=85, y=0)
        lbl_pid_vp3.place(x=0, y=30)
        self.spinbox_PID_VP_3.place(x=85, y=30)
        lbl_vol_in3.place(x=150, y=0)
        self.spinbox_VOL_IN3.place(x=235, y=0)
        lbl_pid_vi3.place(x=150, y=30)
        self.spinbox_PID_VI_3.place(x=235, y=30)
        lbl_vol_limit3.place(x=300, y=0)
        self.spinbox_VOL_LIMIT3.place(x=385, y=0)
        lbl_pid_vd3.place(x=300, y=30)
        self.spinbox_PID_VD_3.place(x=385, y=30)
        # 4
        self.group_pid_4.place(x=450, y=100)
        lbl_pid_p4.place(x=0, y=0)
        self.spinbox_PID_PP_4.place(x=85, y=0)
        lbl_pid_vp4.place(x=0, y=30)
        self.spinbox_PID_VP_4.place(x=85, y=30)
        lbl_vol_in4.place(x=150, y=0)
        self.spinbox_VOL_IN4.place(x=235, y=0)
        lbl_pid_vi4.place(x=150, y=30)
        self.spinbox_PID_VI_4.place(x=235, y=30)
        lbl_vol_limit4.place(x=300, y=0)
        self.spinbox_VOL_LIMIT4.place(x=385, y=0)
        lbl_pid_vd4.place(x=300, y=30)
        self.spinbox_PID_VD_4.place(x=385, y=30)

        self.save_para_m3.place(x=941, y=120)
        self.save_para_m4.place(x=941, y=148)
        self.edit_para_m3_4.place(x=945, y=90)
        #################################################################################
        ############################### SETTING_ARM_5_6 #################################
        lbl_mode_2 = Label(self.ARM5_6, text="MODE_選択")
        self.mode_menu_2 = tk.OptionMenu(self.ARM5_6, self.mode_run_var_2, *MODE_RUN)  # self.mode_run_var/self.num_motor_var
        lbl_num_motor_2 = Label(self.ARM5_6, text="NUM_MOTOR_選択")
        self.num_motor_2 = tk.OptionMenu(self.ARM5_6, self.num_motor_var_2, *NUM_MOTOR)
        self.toggle_servo_5 = tk.Button(self.ARM5_6, height=1, width=10, text="SERVO_5(Off)", background="#FFC652")
        self.toggle_servo_6 = tk.Button(self.ARM5_6, height=1, width=10, text="SERVO_6(Off)", background="#FFC652")
        self.group_pid_5 = tk.LabelFrame(self.ARM5_6, text="arm_5_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_5.propagate(False)
        lbl_pid_p5 = Label(self.group_pid_5, text="PID_position:")
        self.spinbox_PID_PP_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_pp_5)
        lbl_pid_vp5 = Label(self.group_pid_5, text="PID_(P)velocity:")
        self.spinbox_PID_VP_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vp_5)
        lbl_vol_in5 = Label(self.group_pid_5, text="VOL_input:")
        self.spinbox_VOL_IN5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8,textvariable=self.volin_5)
        lbl_pid_vi5 = Label(self.group_pid_5, text="PID_(I)velocity:")
        self.spinbox_PID_VI_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vi_5)
        lbl_vol_limit5 = Label(self.group_pid_5, text="VOL_limit:")
        self.spinbox_VOL_LIMIT5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.01, width=8,textvariable=self.vollimit_5)
        lbl_pid_vd5 = Label(self.group_pid_5, text="PID_(D)velocity:")
        self.spinbox_PID_VD_5 = ttk.Spinbox(self.group_pid_5, from_=0, to=100, increment=0.001, width=8,textvariable=self.pid_vd_5)
        #
        self.group_pid_6 = tk.LabelFrame(self.ARM5_6, text="arm_6_parameter", labelanchor="nw", width=490, height=75)
        self.group_pid_6.propagate(False)
        lbl_pid_p6 = Label(self.group_pid_6, text="PID_position:")
        self.spinbox_PID_PP_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_pp_6)
        lbl_pid_vp6 = Label(self.group_pid_6, text="PID_(P)velocity:")
        self.spinbox_PID_VP_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vp_6)
        lbl_vol_in6 = Label(self.group_pid_6, text="VOL_input:")
        self.spinbox_VOL_IN6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8,textvariable=self.volin_6)
        lbl_pid_vi6 = Label(self.group_pid_6, text="PID_(I)velocity:")
        self.spinbox_PID_VI_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8,textvariable=self.pid_vi_6)
        lbl_vol_limit6 = Label(self.group_pid_6, text="VOL_limit:")
        self.spinbox_VOL_LIMIT6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.01, width=8,textvariable=self.vollimit_6)
        lbl_pid_vd6 = Label(self.group_pid_6, text="PID_(D)velocity:")
        self.spinbox_PID_VD_6 = ttk.Spinbox(self.group_pid_6, from_=0, to=100, increment=0.001, width=8,textvariable=self.pid_vd_6)

        self.save_para_m5 = tk.Button(self.ARM5_6, height=1, width=6, text="Save M5", background="#AD8EE3",command=self.save_para_5)
        self.save_para_m6 = tk.Button(self.ARM5_6, height=1, width=6, text="Save M6", background="#AD8EE3", command=self.save_para_6)
        self.edit_para_m5_6 = tk.Button(self.ARM5_6, height=1, width=4, text="Edit", background="#AD8EE3",command=self.edit_3_switch)
        #6
        lbl_mode_2.place(x=560, y=-5)
        self.mode_menu_2.place(x=640, y=-10)
        lbl_num_motor_2.place(x=760, y=-5)
        self.num_motor_2.place(x=870, y=-10)
        self.toggle_servo_5.place(x=150, y=-5)
        self.toggle_servo_6.place(x=250, y=-5)
        self.group_pid_6.place(x=450, y=100)
        lbl_pid_p6.place(x=0, y=0)
        self.spinbox_PID_PP_6.place(x=85, y=0)
        lbl_pid_vp6.place(x=0, y=30)
        self.spinbox_PID_VP_6.place(x=85, y=30)
        lbl_vol_in6.place(x=150, y=0)
        self.spinbox_VOL_IN6.place(x=235, y=0)
        lbl_pid_vi6.place(x=150, y=30)
        self.spinbox_PID_VI_6.place(x=235, y=30)
        lbl_vol_limit6.place(x=300, y=0)
        self.spinbox_VOL_LIMIT6.place(x=385, y=0)
        lbl_pid_vd6.place(x=300, y=30)
        self.spinbox_PID_VD_6.place(x=385, y=30)
        # 5
        self.group_pid_5.place(x=450, y=20)
        lbl_pid_p5.place(x=0, y=0)
        self.spinbox_PID_PP_5.place(x=85, y=0)
        lbl_pid_vp5.place(x=0, y=30)
        self.spinbox_PID_VP_5.place(x=85, y=30)
        lbl_vol_in5.place(x=150, y=0)
        self.spinbox_VOL_IN5.place(x=235, y=0)
        lbl_pid_vi5.place(x=150, y=30)
        self.spinbox_PID_VI_5.place(x=235, y=30)
        lbl_vol_limit5.place(x=300, y=0)
        self.spinbox_VOL_LIMIT5.place(x=385, y=0)
        lbl_pid_vd5.place(x=300, y=30)
        self.spinbox_PID_VD_5.place(x=385, y=30)

        self.save_para_m5.place(x=941, y=120)
        self.save_para_m6.place(x=941, y=148)
        self.edit_para_m5_6.place(x=945, y=90)
        #################################################################################
        self.labelhonsu1_counter = tk.Label(self.labelhonsu1, textvariable=self.riru_ok_var,font=("Arial", 30, "bold"), )
        self.labelhonsu2_counter = tk.Label(self.labelhonsu2, textvariable=self.hako_ok_var,font=("Arial", 30, "bold"), )
        self.labelhonsu3_counter = tk.Label(self.labelhonsu3, textvariable=self.riru_ng_var,font=("Arial", 30, "bold"), )
        self.labelhonsu4_counter = tk.Label(self.labelhonsu4, textvariable=self.lot_number,font=("Arial", 30, "bold"), )
        self.labelhonsu1.place(x=910, y=0)
        self.labelhonsu1_counter.place(x=0, y=3)
        self.labelhonsu2.place(x=910, y=160)
        self.labelhonsu2_counter.place(x=0, y=3)
        self.labelhonsu3.place(x=910, y=80)
        self.labelhonsu3_counter.place(x=0, y=3)
        self.labelhonsu4.place(x=910, y=240)
        self.labelhonsu4_counter.place(x=0, y=3)
        self.ARM1_2.place(x=0, y=0)
        self.ARM3_4.place(x=0, y=200)
        self.ARM5_6.place(x=0, y=400)

        self.reset_button.bind("<ButtonPress-1>", self.reset_button_on_click)
        self.reset_button.bind("<ButtonRelease-1>", self.reset_button_off_click)
        # TAB 6======================================================================
        self.labelconnect = tk.LabelFrame(self.tab6, text="PLC address", labelanchor="nw", width=300, height=170)
        self.labelconnect.propagate(False)
        lbl_box = Label(self.labelconnect, text="Local address IP")
        txt_address = tk.Entry(self.labelconnect, justify=tk.RIGHT, width=15, textvariable=self.IP_PLC)
        lbl_box1 = Label(self.labelconnect, text="Port")
        txt_port = tk.Entry(self.labelconnect, justify=tk.RIGHT, width=15, textvariable=self.PORT_PLC)
        Sock_connect_button = tk.Button(self.labelconnect, text="PLC_Connect", command=self.socket_connect)
        self.PLC_stt = tk.Text(self.labelconnect, height=2, width=6, font=("Arial", 15), background='white')
        #
        self.labelconnect.place(x=0, y=0)
        lbl_box.place(x=0, y=5)
        txt_address.place(x=100, y=5)
        lbl_box1.place(x=0, y=30)
        txt_port.place(x=100, y=30)
        Sock_connect_button.place(x=0, y=70)
        self.PLC_stt.place(x=200, y=90)
        # シリアルポート選択メニュー

        self.labelserial = tk.LabelFrame(self.tab6, text="Serial_setting", labelanchor="nw", width=250, height=170)
        self.plc_mode = tk.LabelFrame(self.tab6, text="GROUP7", labelanchor="nw", width=340, height=170)
        self.dengen_lb = tk.LabelFrame(self.tab6, text="GROUP8", labelanchor="nw", width=170, height=170)
        self.end_lb = tk.LabelFrame(self.tab6, text="プログラム終了", labelanchor="nw", width=170, height=170)
        self.labelserial.propagate(False)
        self.plc_mode.propagate(False)
        self.dengen_lb.propagate(False)
        self.end_lb.propagate(False)

        ##################### SERIAL ######################################
        lbl_box31 = Label(self.labelserial, text="COM 選択")
        self.port_menu = tk.OptionMenu(self.labelserial, self.port_var, 'None')
        self.port_menu.config(height=1, width=5, bg='gray')
        self.refresh_port_bt = tk.Button(self.labelserial, height=1, width=7, text="Refresh",command=self.refresh_serial_ports)
        lbl_box3 = Label(self.labelserial, text="ボーレート選択")
        self.baudrate_menu = tk.OptionMenu(self.labelserial, self.baudrate_var, *BAUDRATES)
        lbl_box4 = Label(self.labelserial, text="デリミタ選択")
        self.delimiter_menu = tk.OptionMenu(self.labelserial, self.delimiter_var, *DELIMITERS.keys())
        self.delimiter_menu.config(height=1, width=5)
        lbl_box5 = Label(self.labelserial, text="接続")
        self.toggle_button = tk.Button(self.labelserial, height=1, width=9, text="Connect",command=self.toggle_serial_port)
        self.entry_text = tk.Entry(self.labelserial, width=25)
        Serial_send_button = tk.Button(self.labelserial, height=1, width=5, text="送信", command=self.send_data)

        self.labelserial.place(x=0, y=170)
        self.plc_mode.place(x=650, y=0)
        self.dengen_lb.place(x=300, y=0)
        self.end_lb.place(x=470, y=0)
        lbl_box31.place(x=0, y=0)
        self.port_menu.place(x=100, y=0)
        self.refresh_port_bt.place(x=180, y=0)
        lbl_box3.place(x=0, y=30)
        self.baudrate_menu.place(x=100, y=30)
        lbl_box4.place(x=0, y=60)
        self.delimiter_menu.place(x=100, y=60)
        lbl_box5.place(x=0, y=90)
        self.toggle_button.place(x=100, y=90)
        self.entry_text.place(x=0, y=120)
        Serial_send_button.place(x=170, y=120)

        self.displayserial = tk.LabelFrame(self.tab6, text="IN<>OUT", labelanchor="nw", width=500, height=250)
        self.displayserial.propagate(False)
        lbl_sock = tk.Label(self.displayserial, text="PLC socket")
        lbl_comp = tk.Label(self.displayserial, text="Serial")
        self.Text_sock_recieved = Text(self.displayserial, height=13, width=28, background='#C2DEFA')
        self.received_text = tk.Text(self.displayserial, height=13, width=28, background='#C2DEFA')
        clear_button = tk.Button(self.displayserial, text="履歴クリア", command=self.clear_received_text)
        self.connect_mode = tk.Button(self.plc_mode, wraplength=90, text="PROGRAM RUN", height=5,width=12, bd=6, bg='#019a66', font=self.small_font,fg="black",command=self.Communication_serial_run)
        self.program_mode = tk.Button(self.plc_mode, wraplength=90, text="プログラム  停止", height=5,width=12, bd=6, bg='#ef4123',font=self.small_font,fg="black",command=self.Communication_serial_stop)
        self.dengen_bt = tk.Button(self.dengen_lb, wraplength=90, text="電源OFF", height=5,width=12, bd=6, bg='#ef4123',font=self.small_font,fg="black",command=self.ending_program)
        self.end_bt = tk.Button(self.end_lb, wraplength=90, text="プログラム   終了", height=5,width=12, bd=6, bg='pink',font=self.small_font,fg="black",command=self.end_thread)
        #
        self.displayserial.place(x=0, y=345)
        lbl_sock.place(x=0, y=0)
        lbl_comp.place(x=250, y=0)
        self.Text_sock_recieved.place(x=0, y=20)
        self.received_text.place(x=250, y=20)
        clear_button.place(x=250, y=200)
        self.connect_mode.place(x=20, y=20)
        self.program_mode.place(x=200, y=20)
        self.dengen_bt.place(x=30, y=20)
        self.end_bt.place(x=30, y=20)
        ################################################MSP_arm1_2 SERIAL Group
        self.labelserial_msp = tk.LabelFrame(self.tab6, text="Serial_ARM_1-2_setting", labelanchor="nw", width=250, height=170)
        self.labelserial_msp.propagate(False)
        lbl_box_msp = Label(self.labelserial_msp, text="COM 選択")
        self.port_menu_msp = tk.OptionMenu(self.labelserial_msp, self.port_var_msp, 'None')
        self.port_menu_msp.config(height=1, width=5, bg='gray')
        self.refresh_port_bt_msp = tk.Button(self.labelserial_msp, height=1, width=7, text="Refresh", command=self.refresh_serial_M1_2)
        lbl_box_msp3 = Label(self.labelserial_msp, text="ボーレート選択")
        self.baudrate_menu_msp = tk.OptionMenu(self.labelserial_msp, self.baudrate_var_msp, *BAUDRATES)
        lbl_box_msp4 = Label(self.labelserial_msp, text="デリミタ選択")
        self.delimiter_menu_msp = tk.OptionMenu(self.labelserial_msp, self.delimiter_var_msp, *DELIMITERS.keys())
        self.delimiter_menu_msp.config(height=1, width=5)
        lbl_box_msp5 = Label(self.labelserial_msp, text="接続")
        self.toggle_button_msp = tk.Button(self.labelserial_msp, height=1, width=9, text="Connect", command=self.toggle_serial_connect_M1_2)
        self.entry_text_msp = tk.Entry(self.labelserial_msp, width=25)
        Serial_send_button_msp = tk.Button(self.labelserial_msp, height=1, width=5, text="送信", command=self.send_data_M1_2)

        self.labelserial_msp.place(x=255, y=170)
        lbl_box_msp.place(x=0, y=0)
        self.port_menu_msp.place(x=100, y=0)
        self.refresh_port_bt_msp.place(x=180, y=0)
        lbl_box_msp3.place(x=0, y=30)
        self.baudrate_menu_msp.place(x=100, y=30)
        lbl_box_msp4.place(x=0, y=60)
        self.delimiter_menu_msp.place(x=100, y=60)
        lbl_box_msp5.place(x=0, y=90)
        self.toggle_button_msp.place(x=100, y=90)
        self.entry_text_msp.place(x=0, y=120)
        Serial_send_button_msp.place(x=170, y=120)
        ##
        ################################################MSP_arm3_4 SERIAL Group
        self.labelserial_msp_1 = tk.LabelFrame(self.tab6, text="Serial_ARM_3-4_setting", labelanchor="nw", width=250, height=170)
        self.labelserial_msp_1.propagate(False)
        lbl_box_msp_1 = Label(self.labelserial_msp_1, text="COM 選択")
        self.port_menu_msp_1 = tk.OptionMenu(self.labelserial_msp_1, self.port_var_msp_1, 'None')
        self.port_menu_msp_1.config(height=1, width=5, bg='gray')
        self.refresh_port_bt_msp_1 = tk.Button(self.labelserial_msp_1, height=1, width=7, text="Refresh", command=self.refresh_serial_M3_4)
        lbl_box_msp3_1 = Label(self.labelserial_msp_1, text="ボーレート選択")
        self.baudrate_menu_msp_1 = tk.OptionMenu(self.labelserial_msp_1, self.baudrate_var_msp_1, *BAUDRATES)
        lbl_box_msp4_1 = Label(self.labelserial_msp_1, text="デリミタ選択")
        self.delimiter_menu_msp_1 = tk.OptionMenu(self.labelserial_msp_1, self.delimiter_var_msp_1, *DELIMITERS.keys())
        self.delimiter_menu_msp_1.config(height=1, width=5)
        lbl_box_msp5_1 = Label(self.labelserial_msp_1, text="接続")
        self.toggle_button_msp_1 = tk.Button(self.labelserial_msp_1, height=1, width=9, text="Connect", command=self.toggle_serial_connect_M3_4)
        self.entry_text_msp_1 = tk.Entry(self.labelserial_msp_1, width=25)
        Serial_send_button_msp_1 = tk.Button(self.labelserial_msp_1, height=1, width=5, text="送信", command=self.send_data_M3_4)

        self.labelserial_msp_1.place(x=505, y=170)
        lbl_box_msp_1.place(x=0, y=0)
        self.port_menu_msp_1.place(x=100, y=0)
        self.refresh_port_bt_msp_1.place(x=180, y=0)
        lbl_box_msp3_1.place(x=0, y=30)
        self.baudrate_menu_msp_1.place(x=100, y=30)
        lbl_box_msp4_1.place(x=0, y=60)
        self.delimiter_menu_msp_1.place(x=100, y=60)
        lbl_box_msp5_1.place(x=0, y=90)
        self.toggle_button_msp_1.place(x=100, y=90)
        self.entry_text_msp_1.place(x=0, y=120)
        Serial_send_button_msp_1.place(x=170, y=120)
        ###
        ################################################MSP_arm5_6 SERIAL Group
        self.labelserial_msp_2 = tk.LabelFrame(self.tab6, text="Serial_ARM_5-6_setting", labelanchor="nw", width=250, height=170)
        self.labelserial_msp_2.propagate(False)
        lbl_box_msp_2 = Label(self.labelserial_msp_2, text="COM 選択")
        self.port_menu_msp_2 = tk.OptionMenu(self.labelserial_msp_2, self.port_var_msp_2, 'None')
        self.port_menu_msp_2.config(height=1, width=5, bg='gray')
        self.refresh_port_bt_msp_2 = tk.Button(self.labelserial_msp_2, height=1, width=7, text="Refresh", command=self.refresh_serial_M5_6)
        lbl_box_msp3_2 = Label(self.labelserial_msp_2, text="ボーレート選択")
        self.baudrate_menu_msp_2 = tk.OptionMenu(self.labelserial_msp_2, self.baudrate_var_msp_2, *BAUDRATES)
        lbl_box_msp4_2 = Label(self.labelserial_msp_2, text="デリミタ選択")
        self.delimiter_menu_msp_2 = tk.OptionMenu(self.labelserial_msp_2, self.delimiter_var_msp_2, *DELIMITERS.keys())
        self.delimiter_menu_msp_2.config(height=1, width=5)
        lbl_box_msp5_2 = Label(self.labelserial_msp_2, text="接続")
        self.toggle_button_msp_2 = tk.Button(self.labelserial_msp_2, height=1, width=9, text="Connect", command=self.toggle_serial_connect_M5_6)
        self.entry_text_msp_2 = tk.Entry(self.labelserial_msp_2, width=25)
        Serial_send_button_msp_2 = tk.Button(self.labelserial_msp_2, height=1, width=5, text="送信", command=self.send_data_M5_6)

        self.labelserial_msp_2.place(x=755, y=170)
        lbl_box_msp_2.place(x=0, y=0)
        self.port_menu_msp_2.place(x=100, y=0)
        self.refresh_port_bt_msp_2.place(x=180, y=0)
        lbl_box_msp3_2.place(x=0, y=30)
        self.baudrate_menu_msp_2.place(x=100, y=30)
        lbl_box_msp4_2.place(x=0, y=60)
        self.delimiter_menu_msp_2.place(x=100, y=60)
        lbl_box_msp5_2.place(x=0, y=90)
        self.toggle_button_msp_2.place(x=100, y=90)
        self.entry_text_msp_2.place(x=0, y=120)
        Serial_send_button_msp_2.place(x=170, y=120)
        ##
        self.msp_group_serial = tk.LabelFrame(self.tab6, text="MSP_protocol", labelanchor="nw", width=250, height=250)
        self.msp_group_serial.propagate(False)
        lbl_msp_bt1 = tk.Label(self.msp_group_serial, text="MSP send request")
        lbl_msp_recv = tk.Label(self.msp_group_serial, text="MSP recever")
        lbl_msp_command = tk.Label(self.msp_group_serial, text="MSP Command")
        self.send_msp_text = tk.Entry(self.msp_group_serial, width=20, background='#C2DEFA')
        send_msp_button = tk.Button(self.msp_group_serial, text="MSP Cmd")
        self.Text_msp_recieved = Text(self.msp_group_serial, height=8, width=20, background='#C2DEFA')
        self.msp_group_serial.place(x=510, y=345)
        lbl_msp_command.place(x=0, y=10)
        lbl_msp_bt1.place(x=0, y=45)
        self.send_msp_text.place(x=100, y=10)
        send_msp_button.place(x=100, y=45)
        lbl_msp_recv.place(x=0, y=80)
        self.Text_msp_recieved.place(x=100, y=80)
        ########

        #####################
        # TAB 3 CONTROL
        self.displaydatamemory = tk.LabelFrame(self.tab4, text="機械パラメータ", labelanchor="nw", width=500,height=500)
        self.displaydatamemory.propagate(False)
        self.displaydatamemory.place(x=0, y=0)
        self.displayprint = tk.LabelFrame(self.tab4, text="Serial設定", labelanchor="nw", width=500, height=500)
        self.displayprint.propagate(False)
        self.displayprint.place(x=500, y=0)
        self.displayprintstatus = tk.LabelFrame(self.displayprint, text="Serial通信", labelanchor="nw", width=100,height=185)
        self.displayprintstatus.propagate(False)
        self.displayprintstatus.place(x=380, y=0)
        self.display_riruhikae = tk.LabelFrame(self.displayprint, text="NC", labelanchor="nw", width=100,height=85)
        self.display_riruhikae.propagate(False)
        self.display_riruhikae.place(x=210, y=0)
        self.display_hakohikae = tk.LabelFrame(self.displayprint, text="NC", labelanchor="nw",width=100,height=85)
        self.display_hakohikae.propagate(False)
        self.display_hakohikae.place(x=210, y=100)
        self.displaynolabel = tk.LabelFrame(self.displayprint, text="NC", labelanchor="nw", width=100,height=200)
        self.displaynolabel.propagate(False)
        self.displaynolabel.place(x=380, y=270)

        print_all_clear_data = tk.Button(self.displayprint, height=5, width=13, text="BUTTON4",command=lambda: self.Socket_send("ST R9002"), bg="orange")
        print_ryru = tk.Button(self.displayprint, height=5, width=13, wraplength=80, text="BUTTON0",bg="green")
        print_hako = tk.Button(self.displayprint, height=5, width=13, wraplength=80, text="BUTTON2",bg="green")
        self.print_communicate = tk.Button(self.displayprintstatus, height=3, width=8, text="通信開始", bg="gray",command=lambda: self.Socket_send("ST R9000"))
        self.print_communicate_success = tk.Button(self.displayprintstatus, height=3, width=8, wraplength=80,text="BUTTON3", bg="gray",command=lambda: self.Socket_send("ST R9001"))
        print_riruumu = tk.Button(self.displaynolabel, height=4, width=9, wraplength=80, text="BUTTON",bg="gray")
        print_hakoumu = tk.Button(self.displaynolabel, height=4, width=9, wraplength=80, text="BUTTON1",bg="gray")
        self.lbl_datariru = tk.Label(self.display_riruhikae, textvariable=self.text_var1, font=("Arial", 38, "bold"))
        self.lbl_datahako = tk.Label(self.display_hakohikae, textvariable=self.text_var2, font=("Arial", 38, "bold"))
        print_all_clear_data.place(x=30, y=230)
        print_ryru.place(x=30, y=0)
        print_hako.place(x=30, y=100)
        self.print_communicate.place(x=15, y=15)
        self.print_communicate_success.place(x=15, y=90)
        print_riruumu.place(x=11, y=13)
        print_hakoumu.place(x=11, y=103)
        self.lbl_datariru.place(x=0, y=0)
        self.lbl_datahako.place(x=0, y=0)

        print_ryru.bind("<ButtonPress-1>", self.print_ryru_on_click)
        print_ryru.bind("<ButtonRelease-1>", self.print_ryru_off_click)
        print_hako.bind("<ButtonPress-1>", self.print_hako_on_click)
        print_hako.bind("<ButtonRelease-1>", self.print_hako_off_click)

        lbl_dm1000 = Label(self.displaydatamemory, text="SPEED DM1000")
        scale_DM1000 = tk.Scale(self.displaydatamemory,variable=self.DM1000_var,command=self.slider_DM1000,orient=tk.HORIZONTAL,length=200,width=10,sliderlength=20,from_=0,to=1000,resolution=1,tickinterval=0)
        lbl_dm1002 = Label(self.displaydatamemory, text="SPEED DM1002")
        scale_DM1002 = tk.Scale(self.displaydatamemory,variable=self.DM1002_var,command=self.slider_DM1002,orient=tk.HORIZONTAL,length=200,width=10,sliderlength=20,from_=0,to=1000,resolution=1,tickinterval=0 )
        lbl_dm1004 = Label(self.displaydatamemory, text="SPEED DM1004")
        scale_DM1004 = tk.Scale(self.displaydatamemory,variable=self.DM1004_var,command=self.slider_DM1004,orient=tk.HORIZONTAL,length=200,width=10,sliderlength=20,from_=0,to=1000,resolution=1,tickinterval=0)
        lbl_dm1006 = Label(self.displaydatamemory, text="SPEED DM1006")
        scale_DM1006 = tk.Scale(self.displaydatamemory,variable=self.DM1006_var,command=self.slider_DM1006,orient=tk.HORIZONTAL,length=200,width=10,sliderlength=20,from_=0,to=1000,resolution=1,tickinterval=0 )
        lbl_dm1000.place(x=0, y=5)
        scale_DM1000.place(x=120, y=5)
        lbl_dm1002.place(x=0, y=40)
        scale_DM1002.place(x=120, y=40)
        lbl_dm1004.place(x=0, y=75)
        scale_DM1004.place(x=120, y=75)
        lbl_dm1006.place(x=0, y=110)
        scale_DM1006.place(x=120, y=110)

        self.labelhonsu = tk.LabelFrame(self.tab5, text="現在本数変更", labelanchor="nw", width=200, height=300)
        self.labelhonsu.propagate(False)
        self.kasan_button = tk.Button(self.labelhonsu, wraplength=90,font=("Arial", 20), text="＋", height=1,width=3, bd=6, bg='#c0c0c0',fg="green",command=lambda: self.Socket_send("ST R4108"))
        self.gensan_button = tk.Button(self.labelhonsu, wraplength=90, font=("Arial", 20), text="ー", height=1,width=3, bd=6, bg='#c0c0c0',fg="#ef4123",command=lambda: self.Socket_send("ST R4109"))
        spinbox_DM2000 = ttk.Spinbox(self.labelhonsu, textvariable=self.riru_henko, from_=0, to=5, increment=1,wrap=True, state="normal", width=1, font=("Arial", 110))

        self.labelhonsu.place(x=430, y=10)
        self.kasan_button.place(x=10, y=200)
        self.gensan_button.place(x=120, y=200)
        spinbox_DM2000.place(x=50, y=10)
    def create_arm_setting(self):
        ''''''
    #####################################################################################################################
    def create_camera_processing(self):
        ############TAB 7  #############CAMERA#####################

        self.canvas_cam = tk.Canvas(self.tab7, background="#CBD3C0", width=500, height=550)
        self.canvas_cam.place(x=0, y=0)
        # マウスイベント
        self.canvas_cam.bind("<Motion>", self.mouse_move)  # MouseMove
        self.canvas_cam.bind("<B1-Motion>", self.mouse_move_left)  # MouseMove（左ボタンを押しながら移動）
        self.canvas_cam.bind("<Button-1>", self.mouse_down_left)  # MouseDown（左ボタン）
        self.canvas_cam.bind("<B2-Motion>", self.mouse_move_mid)  #
        self.canvas_cam.bind("<Double-Button-1>", self.mouse_double_click_left)  # MouseDoubleClick（左ボタン）
        self.canvas_cam.bind("<MouseWheel>", self.mouse_wheel)  # MouseWheel
        self.canvas_cam.bind("<ButtonRelease-1>", self.mouse_up_left)
        self.canvas_cam.bind("<Button-2>", self.mouse_down_mid)  # MouseMove（中ボタンを押しながら移動）
        self.canvas_cam.bind("<B3-Motion>", self.mouse_move_right)  # MouseMove（左ボタンを押しながら移動）
        self.canvas_cam.bind("<Button-3>", self.mouse_down_right)  # MouseDown（左ボタン）
        self.canvas_cam.bind("<ButtonRelease-3>", self.mouse_up_right)

        self.cam_capture = tk.Button(self.tab7, command=self.camera_callback, text="capture", bg="green", font=self.small_font)
        self.video_capture = tk.Button(self.tab7, command=self.capture_switch, text="Continue_capture", font=self.small_font, bg='#FF4C00')  # , wraplength=60
        self.reset_data_cam = tk.Button(self.tab7, command=self.reset_data, text="Reset_all_setting", font=self.small_font, bg="gray")  # , wraplength=60
        self.cam_capture.place(x=0, y=570)
        self.video_capture.place(x=150, y=570)
        self.reset_data_cam.place(x=300, y=570)
        self.reset_data_cam["state"] = "disable"
        #
        self.tabcam_control = ttk.Notebook(self.tab7, height=650, width=510)
        self.tabcam0 = ttk.Frame(self.tabcam_control)
        self.tabcam2 = ttk.Frame(self.tabcam_control)
        self.tabcam3 = ttk.Frame(self.tabcam_control)

        self.tabcam_control.add(self.tabcam0, state=self.status, text='カメラ設定')
        self.tabcam_control.add(self.tabcam2, state=self.status, text='カメラFilter')
        self.tabcam_control.add(self.tabcam3, state=self.status, text='検査設定')

        # tab camerasetting
        self.labelcamset = tk.LabelFrame(self.tabcam0, text="カメラ設定", labelanchor="nw", width=480, height=400)
        self.labelcamset.propagate(False)
        self.labelcamset.place(x=20, y=20)
        self.labelsetting = tk.LabelFrame(self.tabcam3, text="Area＿現在価値", labelanchor="nw", width=130, height=90)
        self.labelsetting.propagate(False)
        self.labelsetting.place(x=370, y=270)
        self.area_test_check = tk.Label(self.labelsetting, textvariable=self.now_area_check, font=("Arial", 30, "bold"), )
        self.area_test_check.place(x=5, y=5)
        self.chuvi_test_check = tk.Label(self.labelsetting, textvariable=self.now_chuvi_check, font=("Arial", 30, "bold"), )
        FRAME_WIDTH_lb = Label(self.labelcamset, text="FRAME_WIDTH")
        FRAME_WIDTH = tk.Scale(self.labelcamset, variable=self.FRAME_WIDTH_var, command=self.camera_setting, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=2048, resolution=1, tickinterval=0)
        FRAME_HEIGHT_lb = Label(self.labelcamset, text="FRAME_HEIGHT")
        FRAME_HEIGHT = tk.Scale(self.labelcamset, variable=self.FRAME_HEIGHT_var, command=self.camera_setting1, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=1296, resolution=1, tickinterval=0)
        FPS_lb = Label(self.labelcamset, text="FPS")
        FPS = tk.Scale(self.labelcamset, variable=self.FPS_var, command=self.camera_setting2, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=200, resolution=1, tickinterval=0)
        BRIGHTNESS_lb = Label(self.labelcamset, text="BRIGHTNESS")
        BRIGHTNESS = tk.Scale(self.labelcamset, variable=self.BRIGHTNESS_var, command=self.camera_setting3, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-100, to=100, resolution=1, tickinterval=0)
        CONTRAST_lb = Label(self.labelcamset, text="CONTRAST")
        CONTRAST = tk.Scale(self.labelcamset, variable=self.CONTRAST_var, command=self.camera_setting4, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=200, resolution=1, tickinterval=0)
        SATURATION_lb = Label(self.labelcamset, text="SATURATION")
        SATURATION = tk.Scale(self.labelcamset, variable=self.SATURATION_var, command=self.camera_setting5, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=1, to=200, resolution=1, tickinterval=0)
        HUE_lb = Label(self.labelcamset, text="HUE")
        HUE = tk.Scale(self.labelcamset, variable=self.HUE_var, command=self.camera_setting6, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-100, to=100, resolution=1, tickinterval=0)
        GAIN_lb = Label(self.labelcamset, text="GAIN")
        GAIN = tk.Scale(self.labelcamset, variable=self.GAIN_var, command=self.camera_setting7, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-100, to=100, resolution=1, tickinterval=0)
        expose_lb = Label(self.labelcamset, text="EXPOSE")
        EXPOSE_slide = tk.Scale(self.labelcamset, variable=self.EXPOSE_var, command=self.camera_setting10, orient=tk.HORIZONTAL, length=200, width=10, sliderlength=30, from_=-10, to=10, resolution=0.1, tickinterval=0)
        auto_expose_enable = tk.Checkbutton(self.labelcamset, text='Auto EXPOSURE', onvalue=1, offvalue=0, variable=self.camera_setting8)
        auto_focus_enable = tk.Checkbutton(self.labelcamset, text='Auto FOCUS', onvalue=1, offvalue=0, variable=self.camera_setting9)

        FRAME_WIDTH_lb.place(x=0, y=15)
        FRAME_WIDTH.place(x=110, y=0)
        FRAME_HEIGHT_lb.place(x=0, y=45)
        FRAME_HEIGHT.place(x=110, y=30)
        FPS_lb.place(x=0, y=75)
        FPS.place(x=110, y=60)
        BRIGHTNESS_lb.place(x=0, y=105)
        BRIGHTNESS.place(x=110, y=90)
        CONTRAST_lb.place(x=0, y=135)
        CONTRAST.place(x=110, y=120)
        SATURATION_lb.place(x=0, y=165)
        SATURATION.place(x=110, y=150)
        HUE_lb.place(x=0, y=195)
        HUE.place(x=110, y=180)
        GAIN_lb.place(x=0, y=225)
        GAIN.place(x=110, y=210)
        expose_lb.place(x=0, y=255)
        EXPOSE_slide.place(x=110, y=240)
        auto_expose_enable.place(x=0, y=280)
        auto_focus_enable.place(x=0, y=305)

        for child in self.labelcamset.winfo_children():
            child.configure(state='disable')

        self.tabcam_control.place(x=500, y=0)
        # ボタンの作成
        btn_modeless = tk.Button(self.tab6, text="Unlock", command=self.create_anso_dialog_off)
        # btn_modeless.place(x=800, y=0)
        btn_modal = tk.Button(self.tab7, text="Click to Unlock", command=self.create_anso_dialog_on)
        btn_modal.place(x=900, y=0)

        ########### CAM FITER tab 2#######################
        self.labelfilter = tk.LabelFrame(self.tabcam2, text="EDGE_発見", labelanchor="nw", width=350, height=300)
        self.labelfilter.propagate(False)
        self.labelfilter.place(x=150, y=0)

        ###########CAM TAB hani #######################
        hani_enable = tk.Checkbutton(self.tabcam2, text='Check_範囲', onvalue=1, offvalue=0, variable=self.check_hani_ok)
        hani_enable.place(x=0, y=305)
        hani_enable1 = tk.Checkbutton(self.tabcam2, text='玉', onvalue=1, offvalue=0, variable=self.check_patan1)
        hani_enable1.place(x=0, y=325)
        hani_enable2 = tk.Checkbutton(self.tabcam2, text='3角', onvalue=1, offvalue=0, variable=self.check_patan2)
        hani_enable2.place(x=0, y=345)
        hani_enable3 = tk.Checkbutton(self.tabcam2, text='4角', onvalue=1, offvalue=0, variable=self.check_patan3)
        hani_enable3.place(x=0, y=365)
        hani_enable4 = tk.Checkbutton(self.tabcam2, text='5角', onvalue=1, offvalue=0, variable=self.check_patan4)
        hani_enable4.place(x=0, y=385)
        hani_enable5 = tk.Checkbutton(self.tabcam2, text='6角', onvalue=1, offvalue=0, variable=self.check_patan5)
        hani_enable5.place(x=0, y=405)
        hani_enable6 = tk.Checkbutton(self.tabcam2, text='Rectangular', onvalue=1, offvalue=0, variable=self.check_patan6)
        hani_enable6.place(x=0, y=425)

        colormap_lb = Label(self.labelfilter, text="COLOR MAP")
        colormap_lb.place(x=0, y=135)
        thresmap1_lb = Label(self.labelfilter, text="THRES MAP")
        thresmap1_lb.place(x=0, y=165)
        thresmap2_lb = Label(self.labelfilter, text="THRES ADV MAP")
        thresmap2_lb.place(x=0, y=195)

        self.colorchoosen = ttk.Combobox(self.labelfilter, width=20, textvariable=self.colormap)
        self.colorchoosen['values'] = (
        'NORMAL_NO_FILTER', 'GRAY_SCALE', 'COLORMAP_AUTUMN', 'COLORMAP_JET', 'COLORMAP_WINTER', 'COLORMAP_RAINBOW', 'COLORMAP_OCEAN', 'COLORMAP_SUMMER', 'COLORMAP_SPRING', 'COLORMAP_COOL', 'COLORMAP_HSV', 'COLORMAP_PINK', 'COLORMAP_HOT')
        self.colorchoosen.place(x=200, y=135)
        self.colorchoosen.current(0)

        self.boder_filter_map = ttk.Combobox(self.labelfilter, width=15, textvariable=self.boder_filter_map_sl)
        self.boder_filter_map['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_filter_map.place(x=200, y=165)
        self.boder_filter_map.current(0)

        self.boder_filter_map1 = ttk.Combobox(self.labelfilter, width=15, textvariable=self.boder_filter_map_sl1)
        self.boder_filter_map1['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_filter_map1.place(x=200, y=195)
        self.boder_filter_map1.current(0)

        thresholdmap_lb = Label(self.labelfilter, text="threshold")
        thresholdmap_lb.place(x=0, y=10)

        scalethreshold = tk.Scale(self.labelfilter, variable=self.scale_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=0.1, to=255, resolution=0.1, tickinterval=0)
        scalethreshold.place(x=70, y=-5)
        spinbox_Threshold = ttk.Spinbox(self.labelfilter, from_=0.01, to=255, increment=0.1, wrap=True, state="normal", width=15, textvariable=self.scale_var)
        spinbox_Threshold.place(x=220, y=10)

        adaptiveThreshold_lb = Label(self.labelfilter, text="adapt_thres")
        adaptiveThreshold_lb.place(x=0, y=40)
        adaptiveThreshold = tk.Scale(self.labelfilter, variable=self.adaptiveThreshold_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=0.01, to=100, resolution=0.01, tickinterval=0)
        adaptiveThreshold.place(x=70, y=25)
        spinbox_adaptiveThreshold = ttk.Spinbox(self.labelfilter, from_=0.01, to=100, increment=0.01, wrap=True, state="normal", width=15, textvariable=self.adaptiveThreshold_var)
        spinbox_adaptiveThreshold.place(x=220, y=40)
        #
        canny1_lb = Label(self.labelfilter, text="canny_1")
        canny1_lb.place(x=0, y=70)
        canny1 = tk.Scale(self.labelfilter, variable=self.canny1_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=1, to=600, resolution=1, tickinterval=0)
        canny1.place(x=70, y=55)
        spinbox_canny1 = ttk.Spinbox(self.labelfilter, from_=1, to=600, increment=1, wrap=True, state="normal", width=15, textvariable=self.canny1_var)
        spinbox_canny1.place(x=220, y=70)

        canny2_lb = Label(self.labelfilter, text="canny_2")
        canny2_lb.place(x=0, y=100)
        canny2 = tk.Scale(self.labelfilter, variable=self.canny2_var, orient=tk.HORIZONTAL, length=100, width=7, sliderlength=15, from_=1, to=600, resolution=1, tickinterval=0)
        canny2.place(x=70, y=85)
        spinbox_canny2 = ttk.Spinbox(self.labelfilter, from_=1, to=600, increment=1, wrap=True, state="normal", width=15, textvariable=self.canny2_var)
        spinbox_canny2.place(x=220, y=100)
        #
        self.labelfilter1 = tk.LabelFrame(self.tabcam2, text="Smooth", labelanchor="nw", width=350, height=290)
        self.labelfilter1.propagate(False)
        self.labelfilter1.place(x=150, y=300)

        gaussion_lb = Label(self.labelfilter1, text="gaussion")
        gaussion_lb.place(x=0, y=25)
        gaussion = tk.Scale(self.labelfilter1, variable=self.gaussion_var, orient=tk.HORIZONTAL, length=100, width=9, sliderlength=15, from_=0.1, to=300, resolution=0.1, tickinterval=0)
        gaussion.place(x=70, y=10)
        spinbox_gaussion = ttk.Spinbox(self.labelfilter1, from_=0.1, to=300, increment=0.1, wrap=True, state="normal", width=15, textvariable=self.gaussion_var)
        spinbox_gaussion.place(x=220, y=25)
        #
        Blur_lb = Label(self.labelfilter1, text="Blur")
        Blur_lb.place(x=0, y=55)
        Blur = tk.Scale(self.labelfilter1, variable=self.Blur_var, orient=tk.HORIZONTAL, length=100, width=9, sliderlength=15, from_=1, to=100, resolution=1, tickinterval=0)
        Blur.place(x=70, y=40)
        spinbox_Blur = ttk.Spinbox(self.labelfilter1, from_=1, to=100, increment=1, wrap=True, state="normal", width=15, textvariable=self.Blur_var)
        spinbox_Blur.place(x=220, y=55)
        #
        boldx_lb = Label(self.labelfilter1, text="Bold_x")
        boldx_lb.place(x=0, y=85)
        boldy_lb = Label(self.labelfilter1, text="Bold_y")
        boldy_lb.place(x=0, y=115)
        deboldx_lb = Label(self.labelfilter1, text="Debold_x")
        deboldx_lb.place(x=0, y=145)
        deboldy_lb = Label(self.labelfilter1, text="Debold_y")
        deboldy_lb.place(x=0, y=175)
        spinbox_bold = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.bold_var)
        spinbox_bold.place(x=220, y=85)
        spinbox_debold = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.debold_var)
        spinbox_debold.place(x=220, y=115)
        spinbox_bold1 = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.bold_var1)
        spinbox_bold1.place(x=220, y=145)
        spinbox_debold1 = ttk.Spinbox(self.labelfilter1, from_=1, to=200, increment=1, wrap=True, state="normal", width=15, textvariable=self.debold_var1)
        spinbox_debold1.place(x=220, y=175)
        self.labelfilter2 = tk.LabelFrame(self.tabcam2, text="Process", labelanchor="nw", width=150, height=300)
        self.labelfilter2.propagate(False)
        self.labelfilter2.place(x=0, y=0)
        self.filterchoosen = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap)
        self.filterchoosen['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen.place(x=5, y=0)
        self.filterchoosen.current(0)
        self.filterchoosen1 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap1)
        self.filterchoosen1['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen1.current(0)
        self.filterchoosen2 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap2)
        self.filterchoosen2['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen2.current(0)
        self.filterchoosen3 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap3)
        self.filterchoosen3['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen3.current(0)
        self.filterchoosen4 = ttk.Combobox(self.labelfilter2, width=10, textvariable=self.filtermap4)
        self.filterchoosen4['values'] = ('None', 'Threshold', 'Adapt_Thre', 'Canny_1', 'Canny_2', 'Gaussion', 'Blur', 'Bold', 'Debold')
        self.filterchoosen4.current(0)

        self.add_fiter = tk.Button(self.labelfilter2, command=self.add_fiter_to, text="+", bg="green")
        self.add_fiter.place(x=30, y=220)
        self.minus_fiter = tk.Button(self.labelfilter2, command=self.minus_fiter_to, text="-", bg="red")
        self.minus_fiter.place(x=50, y=220)

        for child in self.labelfilter.winfo_children():
            child.configure(state='disable')
        for child in self.labelfilter1.winfo_children():
            child.configure(state='disable')
        for child in self.labelfilter2.winfo_children():
            child.configure(state='disable')

        ############TAB TEST  ##################################
        self.bar_tenken = tk.LabelFrame(self.tab8, text="TEST", labelanchor="nw", width=550, height=150, font=self.small_font)
        self.bar_tenken.propagate(False)
        self.bar_tenken.place(x=20, y=10)
        self.test_start = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.test_start.place(x=10, y=2)
        self.test_setok = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.test_setok.place(x=150, y=2)
        self.scan_start = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.scan_start.place(x=290, y=2)
        self.toridasi_ok = tk.Button(self.bar_tenken, wraplength=90, font=("Arial", 18), text="NC", height=3, width=6, bd=6, bg='#c0c0c0')
        self.toridasi_ok.place(x=430, y=2)
        self.bar_tenken_recieved = tk.LabelFrame(self.tab8, text="バーコード表記内容", labelanchor="nw", width=300, height=300, font=self.small_font)
        self.bar_tenken_recieved.propagate(False)
        self.bar_tenken_recieved.place(x=600, y=10)

        self.barreceived_text = tk.Text(self.bar_tenken_recieved, height=20, width=40, background='#856ff8')
        self.barreceived_text.place(x=5, y=5)
        self.bar_ketka = tk.LabelFrame(self.tab8, text="NC", labelanchor="nw", width=300, height=130, font=self.small_font)
        self.bar_ketka.propagate(False)
        self.bar_ketka.place(x=270, y=180)
        self.BAR_test_check = tk.Label(self.bar_ketka, textvariable=self.BAR_check, font=("Arial", 30, "bold"), )
        self.BAR_test_check.place(x=5, y=5)

        self.test_start.bind("<ButtonPress-1>", self.test_start_on_click)
        self.test_start.bind("<ButtonRelease-1>", self.test_start_off_click)
        self.test_setok.bind("<ButtonPress-1>", self.test_setok_on_click)
        self.test_setok.bind("<ButtonRelease-1>", self.test_setok_off_click)
        self.scan_start.bind("<ButtonPress-1>", self.scan_start_on_click)
        self.scan_start.bind("<ButtonRelease-1>", self.scan_start_off_click)
        self.toridasi_ok.bind("<ButtonPress-1>", self.toridasi_ok_on_click)
        self.toridasi_ok.bind("<ButtonRelease-1>", self.toridasi_ok_off_click)

        # ===============================================================================================================================================
        self.fil_by_Area = tk.IntVar()
        self.fil_by_Area.set(1)
        self.filByCircularity = tk.IntVar()
        self.filByCircularity.set(1)
        self.filByConvexity = tk.IntVar()
        self.filByConvexity.set(1)
        self.filByInertia = tk.IntVar()
        self.filByInertia.set(1)
        self.minArea_var = tk.IntVar()
        self.minArea_var.set(50)
        self.minCircularity_var = tk.DoubleVar()
        self.minCircularity_var.set(0.1)
        self.minConvexity_var = tk.DoubleVar()
        self.minConvexity_var.set(0.1)
        self.minInertia_var = tk.DoubleVar()
        self.minInertia_var.set(0.01)
        self.labeldetect = tk.LabelFrame(self.tabcam3, text="Circle_detect", labelanchor="nw", width=170, height=110)
        self.labeldetect.propagate(False)
        self.fil_select1 = tk.Checkbutton(self.labeldetect, text='Fil by Area', onvalue=1, offvalue=0, variable=self.fil_by_Area)
        self.fil_select2 = tk.Checkbutton(self.labeldetect, text='Fil By Circularity', onvalue=1, offvalue=0, variable=self.filByCircularity)
        self.fil_select3 = tk.Checkbutton(self.labeldetect, text='fil By Convexity ', onvalue=1, offvalue=0, variable=self.filByConvexity)
        self.fil_select4 = tk.Checkbutton(self.labeldetect, text='fil By Inertia  ', onvalue=1, offvalue=0, variable=self.filByInertia)
        minArea = ttk.Spinbox(self.labeldetect, from_=1, to=800, increment=1, width=6, textvariable=self.minArea_var)
        minCircularity = ttk.Spinbox(self.labeldetect, from_=0.01, to=300, increment=0.01, width=6, textvariable=self.minCircularity_var)
        minConvexity = ttk.Spinbox(self.labeldetect, from_=0.01, to=300, increment=0.01, width=6, textvariable=self.minConvexity_var)
        minInertia = ttk.Spinbox(self.labeldetect, from_=0.01, to=300, increment=0.01, width=6, textvariable=self.minInertia_var)
        self.labeldetect.place(x=0, y=0)
        self.fil_select1.place(x=0, y=0)
        self.fil_select2.place(x=0, y=20)
        self.fil_select3.place(x=0, y=40)
        self.fil_select4.place(x=0, y=60)
        minArea.place(x=110, y=0)
        minCircularity.place(x=110, y=20)
        minConvexity.place(x=110, y=40)
        minInertia.place(x=110, y=60)

        # ===============================================================================================================================================
        self.Kernel_var = tk.IntVar()
        self.Kernel_var.set(3)
        self.param1_var = tk.IntVar()
        self.param1_var.set(50)
        self.param2_var = tk.IntVar()
        self.param2_var.set(30)
        self.minRadius_var = tk.IntVar()
        self.minRadius_var.set(1)
        self.maxRadius_var = tk.IntVar()
        self.maxRadius_var.set(40)
        self.labeldetect1 = tk.LabelFrame(self.tabcam3, text="Circle_detect_adv", labelanchor="nw", width=170, height=130)
        self.labeldetect1.propagate(False)
        Kernel_cir_lb = Label(self.labeldetect1, text="Kernel")
        param1_cir = Label(self.labeldetect1, text="param1")
        param2_cir = Label(self.labeldetect1, text="param2")
        minRadius_lb = Label(self.labeldetect1, text="minRadius")
        maxRadius_lb = Label(self.labeldetect1, text="maxRadius")
        Kernel_ = ttk.Spinbox(self.labeldetect1, from_=1, to=100, increment=1, width=6, textvariable=self.Kernel_var)
        param1_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.param1_var)
        param2_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.param2_var)
        minRadius_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.minRadius_var)
        maxRadius_ = ttk.Spinbox(self.labeldetect1, from_=1, to=300, increment=1, width=6, textvariable=self.maxRadius_var)
        self.labeldetect1.place(x=0, y=110)
        Kernel_cir_lb.place(x=0, y=0)
        param1_cir.place(x=0, y=20)
        param2_cir.place(x=0, y=40)
        minRadius_lb.place(x=0, y=60)
        maxRadius_lb.place(x=0, y=80)
        Kernel_.place(x=110, y=0)
        param1_.place(x=110, y=20)
        param2_.place(x=110, y=40)
        minRadius_.place(x=110, y=60)
        maxRadius_.place(x=110, y=80)
        # ===============================================================================================================================================
        self.var_canny_a = tk.IntVar()
        self.var_canny_a.set(250)
        self.var_canny_b = tk.IntVar()
        self.var_canny_b.set(150)
        self.var_thr = tk.IntVar()
        self.var_thr.set(50)
        self.len_size2 = tk.DoubleVar()
        self.len_size2.set(50)
        self.len_size3 = tk.DoubleVar()
        self.len_size3.set(15)
        self.len_size1 = IntVar()
        self.len_size1.set(1)
        self.labeldetect2 = tk.LabelFrame(self.tabcam3, text="line_detect", labelanchor="nw", width=170, height=150)
        self.labeldetect2.propagate(False)
        var_canny_a_lb = Label(self.labeldetect2, text="canny-")
        var_canny_b_lb = Label(self.labeldetect2, text="canny+")
        var_thrlb = Label(self.labeldetect2, text="threshole")
        minLineLength_lb = Label(self.labeldetect2, text="minLineLength")
        maxLineGap_lb = Label(self.labeldetect2, text="maxLineGap")
        canny_a = ttk.Spinbox(self.labeldetect2, from_=1, to=600, increment=1, width=6, textvariable=self.var_canny_a)
        canny_b = ttk.Spinbox(self.labeldetect2, from_=1, to=600, increment=1, width=6, textvariable=self.var_canny_b)
        var_thrxx = ttk.Spinbox(self.labeldetect2, from_=1, to=300, increment=1, width=6, textvariable=self.var_thr)
        minLineLength = ttk.Spinbox(self.labeldetect2, from_=1, to=300, increment=1, width=6, textvariable=self.len_size2)
        maxLineGap = ttk.Spinbox(self.labeldetect2, from_=1, to=300, increment=1, width=6, textvariable=self.len_size3)
        spin_length = ttk.Spinbox(self.labeldetect2, from_=1, to=100, width=5, textvariable=self.len_size1)
        lbl_size = Label(self.labeldetect2, text="Distance pixels")
        self.labeldetect2.place(x=0, y=240)
        var_canny_a_lb.place(x=0, y=0)
        var_canny_b_lb.place(x=0, y=20)
        var_thrlb.place(x=0, y=40)
        minLineLength_lb.place(x=0, y=60)
        maxLineGap_lb.place(x=0, y=80)
        canny_a.place(x=110, y=0)
        canny_b.place(x=110, y=20)
        var_thrxx.place(x=110, y=40)
        minLineLength.place(x=110, y=60)
        maxLineGap.place(x=110, y=80)
        spin_length.place(x=110, y=100)
        lbl_size.place(x=0, y=100)

        # ===============================================================================================================================================
        self.labeselectdetect = tk.LabelFrame(self.tabcam3, text="Select_detect", labelanchor="nw", width=170, height=130)
        self.labeselectdetect.propagate(False)
        self.labeselectdetect.place(x=330, y=440)

        self.circle_select = tk.IntVar()
        self.circle_select.set(0)
        self.circle_select_ = tk.Checkbutton(self.labeselectdetect, text='Circle_dt1', onvalue=1, offvalue=0, variable=self.circle_select)
        self.circle_select_.place(x=0, y=0)

        self.circle_select_adv = tk.IntVar()
        self.circle_select_adv.set(0)
        self.circle_select_adv_ = tk.Checkbutton(self.labeselectdetect, text='Circle_dt2', onvalue=1, offvalue=0, variable=self.circle_select_adv)
        self.circle_select_adv_.place(x=0, y=20)

        self.triang_select = tk.IntVar()
        self.triang_select.set(0)
        self.triang_select_ = tk.Checkbutton(self.labeselectdetect, text='triangle_dt ', onvalue=1, offvalue=0, variable=self.triang_select)
        self.triang_select_.place(x=0, y=40)

        self.rect_select = tk.IntVar()
        self.rect_select.set(0)
        self.rect_select_ = tk.Checkbutton(self.labeselectdetect, text='Rect_dt ', onvalue=1, offvalue=0, variable=self.rect_select)
        self.rect_select_.place(x=0, y=60)
        #
        self.edge_select = tk.IntVar()
        self.edge_select.set(0)
        self.edge_select_ = tk.Checkbutton(self.labeselectdetect, text='Line_dt', onvalue=1, offvalue=0, variable=self.edge_select)
        self.edge_select_.place(x=0, y=80)
        #
        self.draw_contour_select = tk.IntVar()
        self.draw_contour_select.set(0)
        self.draw_contour_select_ = tk.Checkbutton(self.labeselectdetect, text='Draw_Ct', onvalue=1, offvalue=0, variable=self.draw_contour_select)
        self.draw_contour_select_.place(x=85, y=0)
        #
        self.connected_select = tk.IntVar()
        self.connected_select.set(0)
        self.connected_select_ = tk.Checkbutton(self.labeselectdetect, text='Conect_dt', onvalue=1, offvalue=0, variable=self.connected_select)
        self.connected_select_.place(x=85, y=20)
        #
        self.rectangge_component_select = tk.IntVar()
        self.rectangge_component_select.set(0)
        self.rectangge_component_select_ = tk.Checkbutton(self.labeselectdetect, text='■形状_dt', onvalue=1, offvalue=0, variable=self.rectangge_component_select)
        self.rectangge_component_select_.place(x=85, y=40)

        # ===============================================================================================================================================
        self.labeldetect3 = tk.LabelFrame(self.tabcam3, text="rect_detect", labelanchor="nw", width=170, height=180)
        self.labeldetect3.propagate(False)
        self.labeldetect3.place(x=0, y=390)
        self.border_sl = tk.StringVar()
        self.boder_map = ttk.Combobox(self.labeldetect3, width=20, textvariable=self.border_sl)
        # Adding combobox drop down list
        self.boder_map['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_map.place(x=0, y=0)
        self.boder_map.current(0)
        self.contour_sl = tk.StringVar()
        self.contua_map = ttk.Combobox(self.labeldetect3, width=20, textvariable=self.contour_sl)
        # Adding combobox drop down list
        self.contua_map['values'] = ('RETR_TREE', 'RETR_CCOMP', 'RETR_LIST', 'RETR_EXTERNAL')
        self.contua_map.place(x=0, y=25)
        self.contua_map.current(0)
        self.contour_sl1 = tk.StringVar()
        self.contua_map1 = ttk.Combobox(self.labeldetect3, width=20, textvariable=self.contour_sl1)
        # Adding combobox drop down list
        self.contua_map1['values'] = ('CHAIN_APPROX_NONE', 'CHAIN_APPROX_SIMPLE', 'CHAIN_APPROX_TC89_L1')
        self.contua_map1.place(x=0, y=50)
        self.contua_map1.current(1)
        rect_para1_lb = Label(self.labeldetect3, text="Approx")
        rect_para1_lb.place(x=0, y=75)
        area_rect_lb = Label(self.labeldetect3, text="Area")
        area_rect_lb.place(x=0, y=95)
        chuvi_rect_lb = Label(self.labeldetect3, text="Perimeter")
        chuvi_rect_lb.place(x=0, y=115)
        thres_rect_lb = Label(self.labeldetect3, text="Threshole")
        thres_rect_lb.place(x=0, y=135)

        # ===============================================================================================================================================
        self.rect_para1 = tk.DoubleVar()
        self.rect_para1.set(0.04)
        rect_para1_ = ttk.Spinbox(self.labeldetect3, from_=0.01, to=10, increment=0.01, width=6, textvariable=self.rect_para1)
        rect_para1_.place(x=110, y=75)
        self.area_rect = tk.IntVar()
        self.area_rect.set(150)
        area_rect_ = ttk.Spinbox(self.labeldetect3, from_=1, to=600, increment=1, width=6, textvariable=self.area_rect)
        area_rect_.place(x=110, y=95)
        self.chuvi_rect = tk.IntVar()
        self.chuvi_rect.set(50)
        chuvi_rect_ = ttk.Spinbox(self.labeldetect3, from_=1, to=300, increment=1, width=6, textvariable=self.chuvi_rect)
        chuvi_rect_.place(x=110, y=115)
        self.rect_thres = tk.IntVar()
        self.rect_thres.set(150)
        rect_thres_rect_ = ttk.Spinbox(self.labeldetect3, from_=1, to=300, increment=1, width=6, textvariable=self.rect_thres)
        rect_thres_rect_.place(x=110, y=135)
        #############################################################
        self.labeldetect4 = tk.LabelFrame(self.tabcam3, text="Triangle_detect", labelanchor="nw", width=170, height=180)
        self.labeldetect4.propagate(False)
        self.labeldetect4.place(x=172, y=0)
        self.border_sl2 = tk.StringVar()
        self.boder_map2 = ttk.Combobox(self.labeldetect4, width=20, textvariable=self.border_sl2)
        # Adding combobox drop down list
        self.boder_map2['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.boder_map2.place(x=0, y=0)
        self.boder_map2.current(0)
        self.contour_sl2 = tk.StringVar()
        self.contua_map2 = ttk.Combobox(self.labeldetect4, width=20, textvariable=self.contour_sl2)
        # Adding combobox drop down list
        self.contua_map2['values'] = ('RETR_TREE', 'RETR_CCOMP', 'RETR_LIST', 'RETR_EXTERNAL')
        self.contua_map2.place(x=0, y=25)
        self.contua_map2.current(0)
        self.contour_sl21 = tk.StringVar()
        self.contua_map21 = ttk.Combobox(self.labeldetect4, width=20, textvariable=self.contour_sl21)
        # Adding combobox drop down list
        self.contua_map21['values'] = ('CHAIN_APPROX_NONE', 'CHAIN_APPROX_SIMPLE', 'CHAIN_APPROX_TC89_L1')
        self.contua_map21.place(x=0, y=50)
        self.contua_map21.current(1)
        rect_trian1_lb = Label(self.labeldetect4, text="Approx")
        rect_trian1_lb.place(x=0, y=75)
        area_trian_lb = Label(self.labeldetect4, text="Area")
        area_trian_lb.place(x=0, y=95)
        chuvi_trian_lb = Label(self.labeldetect4, text="Perimeter")
        chuvi_trian_lb.place(x=0, y=115)
        thres_trian_lb = Label(self.labeldetect4, text="Threshole")
        thres_trian_lb.place(x=0, y=135)
        # ===============================================================================================================================================
        self.trian_para1 = tk.DoubleVar()
        self.trian_para1.set(0.04)
        trian_para1_ = ttk.Spinbox(self.labeldetect4, from_=0.01, to=10, increment=0.01, width=6, textvariable=self.trian_para1)
        trian_para1_.place(x=110, y=75)
        self.area_trian = tk.IntVar()
        self.area_trian.set(150)
        area_trian_ = ttk.Spinbox(self.labeldetect4, from_=1, to=600, increment=1, width=6, textvariable=self.area_trian)
        area_trian_.place(x=110, y=95)
        self.chuvi_trian = tk.IntVar()
        self.chuvi_trian.set(50)
        chuvi_trian_ = ttk.Spinbox(self.labeldetect4, from_=1, to=300, increment=1, width=6, textvariable=self.chuvi_trian)
        chuvi_trian_.place(x=110, y=115)
        self.trian_thres = tk.IntVar()
        self.trian_thres.set(150)
        trian_thres_rect_ = ttk.Spinbox(self.labeldetect4, from_=1, to=300, increment=1, width=6, textvariable=self.trian_thres)
        trian_thres_rect_.place(x=110, y=135)
        self.labeldetect5 = tk.LabelFrame(self.tabcam3, text="Connected_detect", labelanchor="nw", width=170, height=180)
        self.labeldetect5.propagate(False)
        self.labeldetect5.place(x=172, y=180)
        self.border_connect_sl = tk.StringVar()
        self.border_connect = ttk.Combobox(self.labeldetect5, width=20, textvariable=self.border_connect_sl)
        self.border_connect['values'] = ('THRESH_BINARY', 'THRESH_BINARY_INV', 'THRESH_TRUNC', 'THRESH_TOZERO', 'THRESH_TOZERO_INV')
        self.border_connect.place(x=0, y=0)
        self.border_connect.current(0)
        rect_trian1_lb = Label(self.labeldetect5, text="kernel")
        rect_trian1_lb.place(x=0, y=30)
        iterations_lb = Label(self.labeldetect5, text="iterations")
        iterations_lb.place(x=0, y=50)
        Area_lb = Label(self.labeldetect5, text="Area_detect")
        Area_lb.place(x=0, y=70)
        thres_connect_lb = Label(self.labeldetect5, text="Threshole")
        thres_connect_lb.place(x=0, y=90)
        # ===============================================================================================================================================
        self.kernel_connect = tk.IntVar()
        self.kernel_connect.set(3)
        kernel_connect_ = ttk.Spinbox(self.labeldetect5, from_=1, to=10, increment=1, width=6, textvariable=self.kernel_connect)
        kernel_connect_.place(x=110, y=30)
        self.iterations_var = tk.IntVar()
        self.iterations_var.set(15)
        iterations_ = ttk.Spinbox(self.labeldetect5, from_=1, to=100, increment=1, width=6, textvariable=self.iterations_var)
        iterations_.place(x=110, y=50)
        self.area_detect_connect = tk.IntVar()
        self.area_detect_connect.set(50)
        area_detect_connect_ = ttk.Spinbox(self.labeldetect5, from_=1, to=300, increment=1, width=6, textvariable=self.area_detect_connect)
        area_detect_connect_.place(x=110, y=70)
        self.thres_connect = tk.IntVar()
        self.thres_connect.set(150)
        thres_connect_ = ttk.Spinbox(self.labeldetect5, from_=1, to=300, increment=1, width=6, textvariable=self.thres_connect)
        thres_connect_.place(x=110, y=90)
        # ===============================================================================================================================================

        self.objectdetect = tk.LabelFrame(self.tabcam3, text="■形状_detect", labelanchor="nw", width=170, height=100)
        self.objectdetect.propagate(False)
        self.objectdetect.place(x=344, y=0)
        fill_change_var_lb = Label(self.objectdetect, text="Fill_%")
        fill_change_var_lb.place(x=0, y=0)
        area_patan_lb = Label(self.objectdetect, text="Area")
        area_patan_lb.place(x=0, y=20)
        chuvi_patan_lb = Label(self.objectdetect, text="Perimeter")
        chuvi_patan_lb.place(x=0, y=40)
        thres_patan_lb = Label(self.objectdetect, text="Threshole")
        thres_patan_lb.place(x=0, y=60)
        self.fill_change_var = tk.IntVar()
        self.fill_change_var.set(50)
        fill_change_var_ = ttk.Spinbox(self.objectdetect, from_=1, to=100, increment=0.01, width=6, textvariable=self.fill_change_var)
        fill_change_var_.place(x=110, y=0)
        self.area_patan = tk.IntVar()
        self.area_patan.set(150)
        area_patan_ = ttk.Spinbox(self.objectdetect, from_=1, to=10000, increment=1, width=6, textvariable=self.area_patan)
        area_patan_.place(x=110, y=20)
        self.chuvi_patan = tk.IntVar()
        self.chuvi_patan.set(500)
        chuvi_patan_ = ttk.Spinbox(self.objectdetect, from_=1, to=300, increment=1, width=6, textvariable=self.chuvi_patan)
        chuvi_patan_.place(x=110, y=40)
        self.thres_patan = tk.IntVar()
        self.thres_patan.set(150)
        thres_patan_ = ttk.Spinbox(self.objectdetect, from_=1, to=255, increment=1, width=6, textvariable=self.thres_patan)
        thres_patan_.place(x=110, y=60)

        # ===============================================================================================================================================
        self.objectdetect1 = tk.LabelFrame(self.tabcam3, text="Contuar_detect", labelanchor="nw", width=170, height=100)
        self.objectdetect1.propagate(False)
        self.objectdetect1.place(x=344, y=110)
        Threshole_var_lb = Label(self.objectdetect1, text="Threshole")
        Threshole_var_lb.place(x=0, y=0)
        area_patan_lb = Label(self.objectdetect1, text="approx")
        area_patan_lb.place(x=0, y=20)
        chuvi_contua_lb = Label(self.objectdetect1, text="Perimeter")
        chuvi_contua_lb.place(x=0, y=40)
        thres_patan_lb = Label(self.objectdetect1, text="Area")
        thres_patan_lb.place(x=0, y=60)
        self.Threshole_contua_var = tk.DoubleVar()
        self.Threshole_contua_var.set(50)
        Threshole_contua_var_ = ttk.Spinbox(self.objectdetect1, from_=1, to=255, increment=1, width=6, textvariable=self.Threshole_contua_var)
        Threshole_contua_var_.place(x=110, y=0)
        self.contua_aprrox = tk.DoubleVar()
        self.contua_aprrox.set(0.01)
        contua_aprrox_ = ttk.Spinbox(self.objectdetect1, from_=0.01, to=10, increment=0.01, width=6, textvariable=self.contua_aprrox)
        contua_aprrox_.place(x=110, y=20)
        self.chuvi_contua = tk.IntVar()
        self.chuvi_contua.set(50)
        chuvi_contua_ = ttk.Spinbox(self.objectdetect1, from_=10, to=30000, increment=1, width=6, textvariable=self.chuvi_contua)
        chuvi_contua_.place(x=110, y=40)
        self.dientich_contua = tk.IntVar()
        self.dientich_contua.set(150)
        dientich_contua_ = ttk.Spinbox(self.objectdetect1, from_=1, to=50000, increment=10, width=6, textvariable=self.dientich_contua)
        dientich_contua_.place(x=110, y=60)

        # ===============================================================================================================================================
        for child in self.labeldetect.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect1.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect2.winfo_children():
            child.configure(state='disable')
        for child in self.labeselectdetect.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect3.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect4.winfo_children():
            child.configure(state='disable')
        for child in self.labeldetect5.winfo_children():
            child.configure(state='disable')
        for child in self.objectdetect.winfo_children():
            child.configure(state='disable')
        for child in self.objectdetect1.winfo_children():
            child.configure(state='disable')

    def edit_1_switch(self):
        if self.para_edit_1_is_on:
            self.edit_para_m1_2.config(bg='#AD8EE3')
            self.para_edit_1_is_on = False
        else:
            self.edit_para_m1_2.config(bg='gray')
            self.para_edit_1_is_on = True
    def edit_2_switch(self):
        if self.para_edit_2_is_on:
            self.edit_para_m3_4.config(bg='#AD8EE3')
            self.para_edit_2_is_on = False
        else:
            self.edit_para_m3_4.config(bg='gray')
            self.para_edit_2_is_on = True
    def edit_3_switch(self):
        if self.para_edit_3_is_on:
            self.edit_para_m5_6.config(bg='#AD8EE3')
            self.para_edit_3_is_on = False
        else:
            self.edit_para_m5_6.config(bg='gray')
            self.para_edit_3_is_on = True
    def save_para_1(self):

        payload = []
        float32_value = np.float32(self.pid_pp_1.get())
        nn0  = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_1.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_1.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_1.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_1.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_1.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE1_SET, payload))
        payload = [3]
        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 1")
    def save_para_2(self):

        payload = []
        float32_value = np.float32(self.pid_pp_2.get())
        nn0  = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_2.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_2.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_2.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_2.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_2.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE2_SET, payload))
        payload = [4]
        if self.ser_msp.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 2")
    def save_para_3(self):

        payload = []
        float32_value = np.float32(self.pid_pp_3.get())
        nn0  = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_3.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_3.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_3.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_3.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_3.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE1_SET, payload))
        payload = [3]
        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 3")
    def save_para_4(self):

        payload = []
        float32_value = np.float32(self.pid_pp_4.get())
        nn0  = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_4.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_4.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_4.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_4.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_4.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE2_SET, payload))
        payload = [4]
        if self.ser_msp_1.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 4")
    def save_para_5(self):

        payload = []
        float32_value = np.float32(self.pid_pp_5.get())
        nn0  = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_5.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_5.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_5.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_5.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_5.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE1_SET, payload))
        payload = [3]
        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 5")
    def save_para_6(self):

        payload = []
        float32_value = np.float32(self.pid_pp_6.get())
        nn0  = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn0 & 0xff)
        payload.append((nn0 >> 8) & 0xff)
        payload.append((nn0 >> 16) & 0xFF)
        payload.append((nn0 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vp_6.get())
        nn1 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn1 & 0xff)
        payload.append((nn1 >> 8) & 0xff)
        payload.append((nn1 >> 16) & 0xFF)
        payload.append((nn1 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vi_6.get())
        nn2 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn2 & 0xff)
        payload.append((nn2 >> 8) & 0xff)
        payload.append((nn2 >> 16) & 0xFF)
        payload.append((nn2 >> 24) & 0xFF)

        float32_value = np.float32(self.pid_vd_6.get())
        nn3 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn3 & 0xff)
        payload.append((nn3 >> 8) & 0xff)
        payload.append((nn3 >> 16) & 0xFF)
        payload.append((nn3 >> 24) & 0xFF)

        float32_value = np.float32(self.volin_6.get())
        nn4 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn4 & 0xff)
        payload.append((nn4 >> 8) & 0xff)
        payload.append((nn4 >> 16) & 0xFF)
        payload.append((nn4 >> 24) & 0xFF)

        float32_value = np.float32(self.vollimit_6.get())
        nn5 = np.frombuffer(float32_value.tobytes(), dtype=np.uint32)[0]
        payload.append(nn5 & 0xff)
        payload.append((nn5 >> 8) & 0xff)
        payload.append((nn5 >> 16) & 0xFF)
        payload.append((nn5 >> 24) & 0xFF)

        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_PARASAVE2_SET, payload))
        payload = [4]
        if self.ser_msp_2.is_open and nn4 != 0 and nn5 != 0:
            self.ask_msp_M1_2(self.ask_msp_payload_M1_2(MSP_MOTOR_SAVE_SET, payload))
            print("saved M 6")
    ################################################# MSP Protocol MOTOR 1 2 #####################################
    def refresh_serial_M1_2(self):
        global port_var_msp
        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var_msp.set('')  # 初期値を空に設定
        self.port_menu_msp["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu_msp["menu"].add_command(label=port_name, command=tk._setit(self.port_var_msp, port_name))
    def toggle_serial_connect_M1_2(self):

        if self.ser_msp.is_open:
            try:
                self.ser_msp.close()
                self.baudrate_menu_msp.config(state="normal")
                self.delimiter_menu_msp.config(state="normal")
                self.toggle_button_msp.config(text="接続", background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp.get())
            try:
                #######################################
                self.ser_msp = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp.config(state="disabled")
                self.delimiter_menu_msp.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp = threading.Thread(target=self.Read_data_Serial_M1_2)
                self.com_thread_msp.daemon = True
                self.com_thread_msp.start()
                self.toggle_button_msp.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
    def READ32_M1_2(self):
        global p
        x = (inBuf[p] & 0xff) + ((inBuf[p + 1] & 0xff) << 8) + ((inBuf[p + 2] & 0xff) << 16) + ((inBuf[p + 3] & 0xff) << 24)
        p = p + 4
        return x
    def READ16_M1_2(self):
        global p
        y = (inBuf[p] & 0xff) + ((inBuf[p + 1]) << 8)
        p = p + 2
        return y
    def READ8_M1_2(self):
        global p
        z = inBuf[p] & 0xff
        p = p + 1
        return z
    def request_ask_msp_M1_2(self, msp: int) -> Optional[List[int]]:
        return self.ask_msp_payload_M1_2(msp, None)
    def request_msp_multiple(self, msps: List[int]) -> List[int]:
        s = []
        for m in msps:
            s.extend(self.ask_msp_payload_M1_2(m, None))
        return s
    def ask_msp_payload_M1_2(self, msp: int, payload: Optional[List[chr]]) -> Optional[List[int]]:
        if msp < 0:
            return None
        bf = [ord(c) for c in MSP_HEADER]

        checksum = 0
        pl_size = (len(payload) if payload is not None else 0) & 0xFF
        bf.append(pl_size)
        checksum ^= pl_size

        bf.append(msp & 0xFF)
        checksum ^= (msp & 0xFF)

        if payload is not None:
            for c in payload:
                bf.append(c & 0xFF)
                checksum ^= (c & 0xFF)
        bf.append(checksum)
        return bf
    def ask_msp_M1_2(self, msp: List[int]):
        arr = bytearray(msp)
        # print(arr)
        self.ser_msp.write(arr)  # send the complete byte sequence in one go
    def Operation_MSP_M1_2(self, cmd, data_size):
        icmd = cmd & 0xFF

        if icmd == MSP_MOTOR_MODE:
            error = self.READ8_M1_2()
            run_mode_val = self.READ8_M1_2()
            cur_motor_number = self.READ8_M1_2()

            motor1_resolution = self.READ8_M1_2()
            motor2_resolution = self.READ8_M1_2()
            motor1_pole = self.READ8_M1_2()
            motor2_pole = self.READ8_M1_2()

            data_motor1 = self.READ8_M1_2()
            data_motor2 = self.READ8_M1_2()
            control_motion_1 = self.READ8_M1_2()
            control_motion_2 = self.READ8_M1_2()
            # print(data_motor1)
            if run_mode_val == 1:
                self.mode_run_var.set(MODE_RUN[0])
                self.mode_menu.config(background="#F5B639")
            if run_mode_val == 2:
                self.mode_run_var.set(MODE_RUN[1])
                self.mode_menu.config(background="#F5B639")
            if run_mode_val == 3:
                self.mode_run_var.set(MODE_RUN[2])
                self.mode_menu.config(background="#4DEA42")
            #
            if cur_motor_number == 1:
                self.num_motor_var.set(NUM_MOTOR[0])
                self.num_motor.config(background="#F5B639")
            if cur_motor_number == 2:
                self.num_motor_var.set(NUM_MOTOR[1])
                self.num_motor.config(background="#4DEA42")
            # print(data_motor2)
            if (data_motor1 == 5 or data_motor1 == 14 or error == 3):
                self.toggle_servo_1.config(text="SERVO_1(Off)", background="#FA3559")
            if (data_motor1 == 4):
                self.toggle_servo_1.config(text="SERVO_1(On)", background="#17FA2F")
            ###
            if (data_motor2 == 5 or data_motor2 == 14 or error == 4):
                self.toggle_servo_2.config(text="SERVO_2(Off)", background="#FA3559")
            if (data_motor2 == 4):
                self.toggle_servo_2.config(text="SERVO_2(On)", background="#17FA2F")

        elif icmd == MSP_MOTOR_POS:
            current_pos_11 = self.READ32_M1_2()
            current_pos_12 = self.READ32_M1_2()
            current_pos_21 = self.READ32_M1_2()
            current_pos_22 = self.READ32_M1_2()
            # print(self.uint_to_float(current_pos_11))
            # print(self.uint_to_float(current_pos_12))
            # print(self.uint_to_float(current_pos_21))
            # print(self.uint_to_float(current_pos_22))
        elif icmd == MSP_MOTOR_VEL:
            current_vel11 = self.READ16_M1_2()
            current_vel12 = self.READ16_M1_2()
            current_vel21 = self.READ16_M1_2()
            current_vel22 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_ACCEL_DECCEL:
            current_accel1 = self.READ16_M1_2()
            current_accel2 = self.READ16_M1_2()
            current_deccel1 = self.READ16_M1_2()
            current_deccel2 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_VOL:
            if self.para_edit_1_is_on == False:
                current_vol_in1 = self.READ32_M1_2()
                current_vol_in2 = self.READ32_M1_2()
                self.spinbox_VOL_IN1.set(self.uint_to_float(current_vol_in1))
                self.spinbox_VOL_IN2.set(self.uint_to_float(current_vol_in2))

        elif icmd == MSP_MOTOR_VOLLIMIT:
            if self.para_edit_1_is_on == False:
                current_vol_limit1 = self.READ32_M1_2()
                current_vol_limit2 = self.READ32_M1_2()
                self.spinbox_VOL_LIMIT1.set(self.uint_to_float(current_vol_limit1))
                self.spinbox_VOL_LIMIT2.set(self.uint_to_float(current_vol_limit2))

        elif icmd == MSP_MOTOR_PID_P:
            if self.para_edit_1_is_on == False:
                pid_pos_p1 = self.READ32_M1_2()
                pid_pos_p2 = self.READ32_M1_2()
                self.spinbox_PID_PP_1.set(self.uint_to_float(pid_pos_p1))
                self.spinbox_PID_PP_2.set(self.uint_to_float(pid_pos_p2))

        elif icmd == MSP_MOTOR_PID_VEL_P:
            if self.para_edit_1_is_on == False:
                pid_vel_p1 = self.READ32_M1_2()
                pid_vel_p2 = self.READ32_M1_2()
                self.spinbox_PID_VP_1.set(self.uint_to_float(pid_vel_p1))
                self.spinbox_PID_VP_2.set(self.uint_to_float(pid_vel_p2))

        elif icmd == MSP_MOTOR_PID_VEL_I:
            if self.para_edit_1_is_on == False:
                pid_vel_i1 = self.READ32_M1_2()
                pid_vel_i2 = self.READ32_M1_2()
                self.spinbox_PID_VI_1.set(self.uint_to_float(pid_vel_i1))
                self.spinbox_PID_VI_2.set(self.uint_to_float(pid_vel_i2))

        elif icmd == MSP_MOTOR_PID_VEL_D:
            if self.para_edit_1_is_on == False:
                pid_vel_d1 = self.READ32_M1_2()
                pid_vel_d2 = self.READ32_M1_2()
                self.spinbox_PID_VD_1.set(self.uint_to_float(pid_vel_d1))
                self.spinbox_PID_VD_2.set(self.uint_to_float(pid_vel_d2))

        elif icmd == MSP_MOTOR_FB_ICHI:
            position_now_motor1 = self.READ32_M1_2()
            position_now_motor2 = self.READ32_M1_2()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M1:
            now_velocity_1 = self.READ16_M1_2()
            now_acceleration_1 = self.READ16_M1_2()
            now_decceleration_1 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M2:
            now_velocity_2 = self.READ16_M1_2()
            now_acceleration_2 = self.READ16_M1_2()
            now_decceleration_2 = self.READ16_M1_2()

        elif icmd == MSP_MOTOR_RUN_SET:
            pass

        else:
            print("Don't know how to handle reply " + str(icmd))
    def Operation_ASK_MSP_M1_2(self):
        global time2
        time0 = time.time() * 1000.0
        if ((time0 - time2) > 50):
            time2 = time0
            if (self.i_couter == 0):
                requests = {MSP_MOTOR_POS, MSP_MOTOR_VOL, MSP_MOTOR_PID_VEL_P}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 1):
                requests = {MSP_MOTOR_VEL, MSP_MOTOR_VOLLIMIT, MSP_MOTOR_PID_P}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 2):
                requests = {MSP_MOTOR_PID_VEL_D, MSP_MOTOR_MODE, MSP_MOTOR_ACCEL_DECCEL}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 3):
                requests = {MSP_MOTOR_PID_VEL_I}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 4):
                requests = {MSP_MOTOR_FB_ICHI}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            if (self.i_couter == 5):
                requests = {MSP_MOTOR_FB_VEL_ACC_DEC_M1, MSP_MOTOR_FB_VEL_ACC_DEC_M2}
                self.ask_msp_M1_2(self.request_msp_multiple(requests))

            self.i_couter += 1
            if (self.i_couter > 5):
                self.i_couter = 0
            if self.communicate_on == True:
                self.MSP_id = self.after(100, self.Operation_ASK_MSP_M1_2)
            else:
                self.MSP_id = None
    def send_data_M1_2(self):
        data = self.send_msp_text.get()
        requests = {MSP_MOTOR_POS}  # MSP_MOTOR_MEOBIET,MSP_MOTOR_POS
        # print(self.request_msp_multiple(requests))
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ask_msp_M1_2(self.request_msp_multiple(requests))
            # self.ser.write((data + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%M:%S')
            self.Text_msp_recieved.insert(tk.END, f"[S>{timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.Text_msp_recieved.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))
    def Read_data_Serial_M1_2(self):
        # 設定したデリミタを取得
        global c_state, err_rcvd, dataSize, p, offset, checksum, cmd, inBuf
        delimiter = DELIMITERS[self.delimiter_var_msp.get()]
        buf = ''
        while self.ser_msp.is_open:
            try:

                data = self.ser_msp.read(1).decode("latin1")  # .decode("CP932")//not UTF-8
                c = int.from_bytes(data.encode('latin1'), byteorder='big', signed=False)
                # print(c)
                ###########
                if c_state == IDLE:
                    if (c == 36):
                        c_state = HEADER_START
                    else:
                        c_state = IDLE
                elif c_state == HEADER_START:
                    c_state = HEADER_M if c == 77 else IDLE

                elif c_state == HEADER_M:
                    if c == 62:
                        c_state = HEADER_ARROW
                    elif c == '33':  # not use
                        c_state = HEADER_ERR
                    else:
                        c_state = IDLE
                elif c_state in (HEADER_ARROW, HEADER_ERR):
                    err_rcvd = (c_state == HEADER_ERR)  # now we are expecting the payload size
                    dataSize = c & 0xFF
                    # reset index variables
                    p = 0
                    offset = 0
                    checksum = 0
                    checksum ^= (c & 0xFF)

                    # the command is to follow
                    c_state = HEADER_SIZE
                elif c_state == HEADER_SIZE:
                    cmd = c & 0xFF
                    checksum ^= (c & 0xFF)
                    c_state = HEADER_CMD
                elif c_state == HEADER_CMD and offset < dataSize:
                    checksum ^= (c & 0xFF)
                    inBuf[offset] = c & 0xFF
                    offset += 1

                elif c_state == HEADER_CMD and offset >= dataSize:
                    # compare calculated and transferred checksum
                    if (checksum & 0xFF) == (c & 0xFF):
                        if err_rcvd:
                            print(f"Copter did not understand request type {c}")
                        else:
                            # we got a valid response packet, evaluate it
                            # print(cmd)
                            self.Operation_MSP_M1_2(cmd, int(dataSize))
                    else:
                        print(f"invalid checksum for command {cmd & 0xFF}: {checksum & 0xFF} expected, got {c & 0xFF}")
                        print(f"<{cmd & 0xFF} {dataSize & 0xFF}> {{", end=' ')
                        for i in range(dataSize):
                            if i != 0:
                                print(' ', end='')
                            print(inBuf[i] & 0xFF, end='')
                        print("} [", c, "]")
                        print(''.join(map(chr, inBuf[:dataSize])))
                    c_state = IDLE

                #######
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.Text_msp_recieved.see("end")
                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))
    def autoconnect_msp_comport_M1_2(self):
        self.port_var_msp.set('COM9')
        if self.ser_msp.is_open:
            try:
                self.ser_msp.close()
                self.baudrate_menu_msp.config(state="normal")
                self.delimiter_menu_msp.config(state="normal")
                self.toggle_button_msp.config(text="接続")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp.get()
            if not port:
                messagebox.showerror("接続エラー", "MSPシリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp.get())
            try:
                #######################################
                self.ser_msp = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp.config(state="disabled")
                self.delimiter_menu_msp.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp = threading.Thread(target=self.Read_data_Serial_M1_2)
                self.com_thread_msp.daemon = True
                self.com_thread_msp.start()
                self.toggle_button_msp.config(text="切断", background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))

    ################################################# MSP Protocol MOTOR 3 4 #####################################
    def refresh_serial_M3_4(self):

        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var_msp_1.set('')  # 初期値を空に設定
        self.port_menu_msp_1["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu_msp_1["menu"].add_command(label=port_name, command=tk._setit(self.port_var_msp_1, port_name))
    def toggle_serial_connect_M3_4(self):

        if self.ser_msp_1.is_open:
            try:
                self.ser_msp_1.close()
                self.baudrate_menu_msp_1.config(state="normal")
                self.delimiter_menu_msp_1.config(state="normal")
                self.toggle_button_msp_1.config(text="接続",background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_1.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_1.get())
            try:
                #######################################
                self.ser_msp_1 = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_1.config(state="disabled")
                self.delimiter_menu_msp_1.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_1 = threading.Thread(target=self.Read_data_Serial_M3_4)
                self.com_thread_msp_1.daemon = True
                self.com_thread_msp_1.start()
                self.toggle_button_msp_1.config(text="切断",background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
    def READ32_M3_4(self):
        global  p
        x = (inBuf[p] & 0xff) + ((inBuf[p+1] & 0xff) << 8) + ((inBuf[p+2] & 0xff) << 16) + ((inBuf[p+3] & 0xff) << 24)
        p = p+4
        return x
    def READ16_M3_4(self):
        global p
        y = (inBuf[p] & 0xff) + ((inBuf[p + 1]) << 8)
        p = p+2
        return y
    def READ8_M3_4(self):
        global p
        z = inBuf[p] & 0xff
        p = p+1
        return z
    def request_ask_msp_M3_4(self,msp: int) -> Optional[List[int]]:
        return self.ask_msp_payload_M3_4(msp, None)
    def request_multiple_M3_4(self,msps: List[int]) -> List[int]:
        s = []
        for m in msps:
            s.extend(self.ask_msp_payload_M3_4(m, None))
        return s
    def ask_msp_payload_M3_4(self,msp: int, payload: Optional[List[chr]]) -> Optional[List[int]]:
        if msp < 0:
            return None
        bf = [ord(c) for c in MSP_HEADER]

        checksum = 0
        pl_size = (len(payload) if payload is not None else 0) & 0xFF
        bf.append(pl_size)
        checksum ^= pl_size

        bf.append(msp & 0xFF)
        checksum ^= (msp & 0xFF)

        if payload is not None:
            for c in payload:
                bf.append(c & 0xFF)
                checksum ^= (c & 0xFF)
        bf.append(checksum)
        return bf
    def ask_msp_M3_4(self,msp: List[int]):
        arr = bytearray(msp)
        #print(arr)
        self.ser_msp_1.write(arr)  # send the complete byte sequence in one go
    def Operation_MSP_M3_4(self,cmd, data_size):
        icmd = cmd & 0xFF


        if icmd == MSP_MOTOR_MODE:
            error = self.READ8_M3_4()
            run_mode_val = self.READ8_M3_4()
            cur_motor_number = self.READ8_M3_4()

            motor1_resolution = self.READ8_M3_4()
            motor2_resolution = self.READ8_M3_4()
            motor1_pole = self.READ8_M3_4()
            motor2_pole = self.READ8_M3_4()

            data_motor1 = self.READ8_M3_4()
            data_motor2 = self.READ8_M3_4()
            control_motion_1 = self.READ8_M3_4()
            control_motion_2 = self.READ8_M3_4()
            #print(data_motor1)
            if run_mode_val==1:
                self.mode_run_var_1.set(MODE_RUN[0])
                self.mode_menu_1.config(background="#F5B639")
            if run_mode_val==2:
                self.mode_run_var_1.set(MODE_RUN[1])
                self.mode_menu_1.config(background="#F5B639")
            if run_mode_val == 3:
                self.mode_run_var_1.set(MODE_RUN[2])
                self.mode_menu_1.config(background="#4DEA42")
            #
            if cur_motor_number ==1:
                self.num_motor_var_1.set(NUM_MOTOR[0])
                self.num_motor_1.config(background="#F5B639")
            if cur_motor_number ==2:
                self.num_motor_var_1.set(NUM_MOTOR[1])
                self.num_motor_1.config(background="#4DEA42")
            #print(data_motor2)
            if (data_motor1 == 5 or data_motor1 == 14 or error == 3):
                self.toggle_servo_3.config(text="SERVO_3(Off)", background="#FA3559")
            if(data_motor1 == 4):
                self.toggle_servo_3.config(text="SERVO_3(On)", background="#17FA2F")
            ###
            if (data_motor2 == 5 or data_motor2 == 14 or error == 4):
                self.toggle_servo_4.config(text="SERVO_4(Off)", background="#FA3559")
            if (data_motor2 == 4):
                self.toggle_servo_4.config(text="SERVO_4(On)", background="#17FA2F")

        elif icmd == MSP_MOTOR_POS:
            current_pos_11 = self.READ32_M3_4()
            current_pos_12 = self.READ32_M3_4()
            current_pos_21 = self.READ32_M3_4()
            current_pos_22 = self.READ32_M3_4()
            # print(self.uint_to_float(current_pos_11))
            # print(self.uint_to_float(current_pos_12))
            # print(self.uint_to_float(current_pos_21))
            # print(self.uint_to_float(current_pos_22))
        elif icmd == MSP_MOTOR_VEL:
            current_vel11 = self.READ16_M3_4()
            current_vel12 = self.READ16_M3_4()
            current_vel21 = self.READ16_M3_4()
            current_vel22 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_ACCEL_DECCEL:
            current_accel1 = self.READ16_M3_4()
            current_accel2 = self.READ16_M3_4()
            current_deccel1 = self.READ16_M3_4()
            current_deccel2 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_VOL:
            if self.para_edit_2_is_on == False:
                current_vol_in1 = self.READ32_M3_4()
                current_vol_in2 = self.READ32_M3_4()
                self.spinbox_VOL_IN3.set(self.uint_to_float(current_vol_in1))
                self.spinbox_VOL_IN4.set(self.uint_to_float(current_vol_in2))

        elif icmd == MSP_MOTOR_VOLLIMIT:
            if self.para_edit_2_is_on == False:
                current_vol_limit1 = self.READ32_M3_4()
                current_vol_limit2 = self.READ32_M3_4()
                self.spinbox_VOL_LIMIT3.set(self.uint_to_float(current_vol_limit1))
                self.spinbox_VOL_LIMIT4.set(self.uint_to_float(current_vol_limit2))

        elif icmd == MSP_MOTOR_PID_P:
            if self.para_edit_2_is_on == False:
                pid_pos_p1 = self.READ32_M3_4()
                pid_pos_p2 = self.READ32_M3_4()
                self.spinbox_PID_PP_3.set(self.uint_to_float(pid_pos_p1))
                self.spinbox_PID_PP_4.set(self.uint_to_float(pid_pos_p2))

        elif icmd == MSP_MOTOR_PID_VEL_P:
            if self.para_edit_2_is_on == False:
                pid_vel_p1 = self.READ32_M3_4()
                pid_vel_p2 = self.READ32_M3_4()
                self.spinbox_PID_VP_3.set(self.uint_to_float(pid_vel_p1))
                self.spinbox_PID_VP_4.set(self.uint_to_float(pid_vel_p2))

        elif icmd == MSP_MOTOR_PID_VEL_I:
            if self.para_edit_2_is_on == False:
                pid_vel_i1 = self.READ32_M3_4()
                pid_vel_i2 = self.READ32_M3_4()
                self.spinbox_PID_VI_3.set(self.uint_to_float(pid_vel_i1))
                self.spinbox_PID_VI_4.set(self.uint_to_float(pid_vel_i2))

        elif icmd == MSP_MOTOR_PID_VEL_D:
            if self.para_edit_2_is_on == False:
                pid_vel_d1 = self.READ32_M3_4()
                pid_vel_d2 = self.READ32_M3_4()
                self.spinbox_PID_VD_3.set(self.uint_to_float(pid_vel_d1))
                self.spinbox_PID_VD_4.set(self.uint_to_float(pid_vel_d2))

        elif icmd == MSP_MOTOR_FB_ICHI:
            position_now_motor1 = self.READ32_M3_4()
            position_now_motor2 = self.READ32_M3_4()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M1:
            now_velocity_1 = self.READ16_M3_4()
            now_acceleration_1 = self.READ16_M3_4()
            now_decceleration_1 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M2:
            now_velocity_2 = self.READ16_M3_4()
            now_acceleration_2 = self.READ16_M3_4()
            now_decceleration_2 = self.READ16_M3_4()

        elif icmd == MSP_MOTOR_RUN_SET:
            pass

        else:
            print("Don't know how to handle reply " + str(icmd))
    def Operation_ASK_MSP_M3_4(self):
        global time3
        time1 = time.time() * 1000.0
        if ((time1 - time3) > 50):
            time3 = time1
            if (self.i_couter_1 == 0):
                requests = {MSP_MOTOR_POS, MSP_MOTOR_VOL, MSP_MOTOR_PID_VEL_P}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 1):
                requests = {MSP_MOTOR_VEL, MSP_MOTOR_VOLLIMIT, MSP_MOTOR_PID_P}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 2):
                requests = {MSP_MOTOR_PID_VEL_D, MSP_MOTOR_MODE, MSP_MOTOR_ACCEL_DECCEL}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 3):
                requests = {MSP_MOTOR_PID_VEL_I}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 4):

                requests = {MSP_MOTOR_FB_ICHI}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            if (self.i_couter_1 == 5):

                requests = {MSP_MOTOR_FB_VEL_ACC_DEC_M1, MSP_MOTOR_FB_VEL_ACC_DEC_M2}
                self.ask_msp_M3_4(self.request_multiple_M3_4(requests))

            self.i_couter_1 += 1
            if (self.i_couter_1 > 5):
                self.i_couter_1 = 0
            if self.communicate_on == True:
                self.MSP_id_1 = self.after(100, self.Operation_ASK_MSP_M3_4)
            else:
                self.MSP_id_1 =None
    def send_data_M3_4(self):
        data = self.send_msp_text.get()
        requests = {MSP_MOTOR_POS}#MSP_MOTOR_MEOBIET,MSP_MOTOR_POS
        #print(self.request_msp_multiple(requests))
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ask_msp_M3_4(self.request_multiple_M3_4(requests))
            #self.ser.write((data + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%M:%S')
            self.Text_msp_recieved.insert(tk.END, f"[S>{timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.Text_msp_recieved.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))
    def Read_data_Serial_M3_4(self):
        # 設定したデリミタを取得
        global c_state, err_rcvd, dataSize, p, offset, checksum, cmd, inBuf
        delimiter = DELIMITERS[self.delimiter_var_msp.get()]
        buf = ''
        while self.ser_msp_1.is_open:
            try:

                data = self.ser_msp_1.read(1).decode("latin1")#.decode("CP932")//not UTF-8
                c = int.from_bytes(data.encode('latin1'), byteorder='big', signed=False)
                #print(c)
                ###########
                if c_state == IDLE:
                    if (c ==36):
                        c_state = HEADER_START
                    else :
                        c_state = IDLE
                elif c_state == HEADER_START:
                    c_state = HEADER_M if c == 77 else IDLE

                elif c_state == HEADER_M:
                    if c == 62:
                        c_state = HEADER_ARROW
                    elif c == '33':#not use
                        c_state = HEADER_ERR
                    else:
                        c_state = IDLE
                elif c_state in (HEADER_ARROW, HEADER_ERR):
                    err_rcvd = (c_state == HEADER_ERR)  # now we are expecting the payload size
                    dataSize = c & 0xFF
                    # reset index variables
                    p = 0
                    offset = 0
                    checksum = 0
                    checksum ^= (c & 0xFF)

                    # the command is to follow
                    c_state = HEADER_SIZE
                elif c_state == HEADER_SIZE:
                    cmd = c & 0xFF
                    checksum ^= (c & 0xFF)
                    c_state = HEADER_CMD
                elif c_state == HEADER_CMD and offset < dataSize:
                    checksum ^= (c & 0xFF)
                    inBuf[offset] = c & 0xFF
                    offset += 1

                elif c_state == HEADER_CMD and offset >= dataSize:
                    # compare calculated and transferred checksum
                    if (checksum & 0xFF) == (c & 0xFF):
                        if err_rcvd:
                            print(f"Copter did not understand request type {c}")
                        else:
                            # we got a valid response packet, evaluate it
                            #print(cmd)
                            self.Operation_MSP_M3_4(cmd, int(dataSize))
                    else:
                        print(f"invalid checksum for command {cmd & 0xFF}: {checksum & 0xFF} expected, got {c & 0xFF}")
                        print(f"<{cmd & 0xFF} {dataSize & 0xFF}> {{", end=' ')
                        for i in range(dataSize):
                            if i != 0:
                                print(' ', end='')
                            print(inBuf[i] & 0xFF, end='')
                        print("} [", c, "]")
                        print(''.join(map(chr, inBuf[:dataSize])))
                    c_state = IDLE

                #######
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.Text_msp_recieved.see("end")
                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))
    def autoconnect_msp_comport_M3_4(self):
        self.port_var_msp_1.set('COM2')
        if self.ser_msp_1.is_open:
            try:
                self.ser_msp_1.close()
                self.baudrate_menu_msp_1.config(state="normal")
                self.delimiter_menu_msp_1.config(state="normal")
                self.toggle_button_msp_1.config(text="接続")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_1.get()
            if not port:
                messagebox.showerror("接続エラー", "MSPシリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_1.get())
            try:
                #######################################
                self.ser_msp_1 = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_1.config(state="disabled")
                self.delimiter_menu_msp_1.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_1 = threading.Thread(target=self.Read_data_Serial_M3_4)
                self.com_thread_msp_1.daemon = True
                self.com_thread_msp_1.start()
                self.toggle_button_msp_1.config(text="切断",background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
        #################################################MSP Protocol MOTOR 1 2 #####################################

        ################################################# MSP Protocol MOTOR 5 6 #####################################

    ################################################# MSP Protocol MOTOR 5 6 #####################################
    def refresh_serial_M5_6(self):

        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var_msp_2.set('')  # 初期値を空に設定
        self.port_menu_msp_2["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu_msp_2["menu"].add_command(label=port_name, command=tk._setit(self.port_var_msp_2, port_name))
    def toggle_serial_connect_M5_6(self):

        if self.ser_msp_2.is_open:
            try:
                self.ser_msp_2.close()
                self.baudrate_menu_msp_2.config(state="normal")
                self.delimiter_menu_msp_2.config(state="normal")
                self.toggle_button_msp_2.config(text="接続",background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_2.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_2.get())
            try:
                #######################################
                self.ser_msp_2 = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_2.config(state="disabled")
                self.delimiter_menu_msp_2.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_2 = threading.Thread(target=self.Read_data_Serial_M5_6)
                self.com_thread_msp_2.daemon = True
                self.com_thread_msp_2.start()
                self.toggle_button_msp_2.config(text="切断",background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
    def READ32_M5_6(self):
        global  p
        x = (inBuf[p] & 0xff) + ((inBuf[p+1] & 0xff) << 8) + ((inBuf[p+2] & 0xff) << 16) + ((inBuf[p+3] & 0xff) << 24)
        p = p+4
        return x
    def READ16_M5_6(self):
        global p
        y = (inBuf[p] & 0xff) + ((inBuf[p + 1]) << 8)
        p = p+2
        return y
    def READ8_M5_6(self):
        global p
        z = inBuf[p] & 0xff
        p = p+1
        return z
    def request_ask_msp_M5_6(self,msp: int) -> Optional[List[int]]:
        return self.ask_msp_payload_M5_6(msp, None)
    def request_multiple_M5_6(self,msps: List[int]) -> List[int]:
        s = []
        for m in msps:
            s.extend(self.ask_msp_payload_M5_6(m, None))
        return s
    def ask_msp_payload_M5_6(self,msp: int, payload: Optional[List[chr]]) -> Optional[List[int]]:
        if msp < 0:
            return None
        bf = [ord(c) for c in MSP_HEADER]

        checksum = 0
        pl_size = (len(payload) if payload is not None else 0) & 0xFF
        bf.append(pl_size)
        checksum ^= pl_size

        bf.append(msp & 0xFF)
        checksum ^= (msp & 0xFF)

        if payload is not None:
            for c in payload:
                bf.append(c & 0xFF)
                checksum ^= (c & 0xFF)
        bf.append(checksum)
        return bf
    def ask_msp_M5_6(self,msp: List[int]):
        arr = bytearray(msp)
        #print(arr)
        self.ser_msp_2.write(arr)  # send the complete byte sequence in one go
    def Operation_MSP_M5_6(self,cmd, data_size):
        icmd = cmd & 0xFF


        if icmd == MSP_MOTOR_MODE:
            error = self.READ8_M5_6()
            run_mode_val = self.READ8_M5_6()
            cur_motor_number = self.READ8_M5_6()

            motor1_resolution = self.READ8_M5_6()
            motor2_resolution = self.READ8_M5_6()
            motor1_pole = self.READ8_M5_6()
            motor2_pole = self.READ8_M5_6()

            data_motor1 = self.READ8_M5_6()
            data_motor2 = self.READ8_M5_6()
            control_motion_1 = self.READ8_M5_6()
            control_motion_2 = self.READ8_M5_6()
            #print(data_motor1)
            if run_mode_val==1:
                self.mode_run_var_2.set(MODE_RUN[0])
                self.mode_menu_2.config(background="#F5B639")
            if run_mode_val==2:
                self.mode_run_var_2.set(MODE_RUN[1])
                self.mode_menu_2.config(background="#F5B639")
            if run_mode_val == 3:
                self.mode_run_var_2.set(MODE_RUN[2])
                self.mode_menu_2.config(background="#4DEA42")
            #
            if cur_motor_number ==1:
                self.num_motor_var_2.set(NUM_MOTOR[0])
                self.num_motor_2.config(background="#F5B639")
            if cur_motor_number ==2:
                self.num_motor_var_2.set(NUM_MOTOR[1])
                self.num_motor_2.config(background="#4DEA42")
            #print(data_motor2)
            if (data_motor1 == 5 or data_motor1 == 14 or error == 3):
                self.toggle_servo_5.config(text="SERVO_5(Off)", background="#FA3559")
            if(data_motor1 == 4):
                self.toggle_servo_5.config(text="SERVO_5(On)", background="#17FA2F")
            ###
            if (data_motor2 == 5 or data_motor2 == 14 or error == 4):
                self.toggle_servo_6.config(text="SERVO_6(Off)", background="#FA3559")
            if (data_motor2 == 4):
                self.toggle_servo_6.config(text="SERVO_6(On)", background="#17FA2F")

        elif icmd == MSP_MOTOR_POS:
            current_pos_11 = self.READ32_M5_6()
            current_pos_12 = self.READ32_M5_6()
            current_pos_21 = self.READ32_M5_6()
            current_pos_22 = self.READ32_M5_6()
            # print(self.uint_to_float(current_pos_11))
            # print(self.uint_to_float(current_pos_12))
            # print(self.uint_to_float(current_pos_21))
            # print(self.uint_to_float(current_pos_22))
        elif icmd == MSP_MOTOR_VEL:
            current_vel11 = self.READ16_M5_6()
            current_vel12 = self.READ16_M5_6()
            current_vel21 = self.READ16_M5_6()
            current_vel22 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_ACCEL_DECCEL:
            current_accel1 = self.READ16_M5_6()
            current_accel2 = self.READ16_M5_6()
            current_deccel1 = self.READ16_M5_6()
            current_deccel2 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_VOL:
            if self.para_edit_3_is_on == False:
                current_vol_in1 = self.READ32_M5_6()
                current_vol_in2 = self.READ32_M5_6()
                self.spinbox_VOL_IN5.set(self.uint_to_float(current_vol_in1))
                self.spinbox_VOL_IN6.set(self.uint_to_float(current_vol_in2))

        elif icmd == MSP_MOTOR_VOLLIMIT:
            if self.para_edit_3_is_on == False:
                current_vol_limit1 = self.READ32_M5_6()
                current_vol_limit2 = self.READ32_M5_6()
                self.spinbox_VOL_LIMIT5.set(self.uint_to_float(current_vol_limit1))
                self.spinbox_VOL_LIMIT6.set(self.uint_to_float(current_vol_limit2))

        elif icmd == MSP_MOTOR_PID_P:
            if self.para_edit_3_is_on == False:
                pid_pos_p1 = self.READ32_M5_6()
                pid_pos_p2 = self.READ32_M5_6()
                self.spinbox_PID_PP_5.set(self.uint_to_float(pid_pos_p1))
                self.spinbox_PID_PP_6.set(self.uint_to_float(pid_pos_p2))

        elif icmd == MSP_MOTOR_PID_VEL_P:
            if self.para_edit_3_is_on == False:
                pid_vel_p1 = self.READ32_M5_6()
                pid_vel_p2 = self.READ32_M5_6()
                self.spinbox_PID_VP_5.set(self.uint_to_float(pid_vel_p1))
                self.spinbox_PID_VP_6.set(self.uint_to_float(pid_vel_p2))

        elif icmd == MSP_MOTOR_PID_VEL_I:
            if self.para_edit_3_is_on == False:
                pid_vel_i1 = self.READ32_M5_6()
                pid_vel_i2 = self.READ32_M5_6()
                self.spinbox_PID_VI_5.set(self.uint_to_float(pid_vel_i1))
                self.spinbox_PID_VI_6.set(self.uint_to_float(pid_vel_i2))

        elif icmd == MSP_MOTOR_PID_VEL_D:
            if self.para_edit_3_is_on == False:
                pid_vel_d1 = self.READ32_M5_6()
                pid_vel_d2 = self.READ32_M5_6()
                self.spinbox_PID_VD_5.set(self.uint_to_float(pid_vel_d1))
                self.spinbox_PID_VD_6.set(self.uint_to_float(pid_vel_d2))

        elif icmd == MSP_MOTOR_FB_ICHI:
            position_now_motor1 = self.READ32_M5_6()
            position_now_motor2 = self.READ32_M5_6()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M1:
            now_velocity_1 = self.READ16_M5_6()
            now_acceleration_1 = self.READ16_M5_6()
            now_decceleration_1 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_FB_VEL_ACC_DEC_M2:
            now_velocity_2 = self.READ16_M5_6()
            now_acceleration_2 = self.READ16_M5_6()
            now_decceleration_2 = self.READ16_M5_6()

        elif icmd == MSP_MOTOR_RUN_SET:
            pass

        else:
            print("Don't know how to handle reply " + str(icmd))
    def Operation_ASK_MSP_M5_6(self):
        global time4
        time5 = time.time() * 1000.0
        if ((time5 - time4) > 50):
            time4 = time5
            if (self.i_couter_2 == 0):
                requests = {MSP_MOTOR_POS, MSP_MOTOR_VOL, MSP_MOTOR_PID_VEL_P}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 1):
                requests = {MSP_MOTOR_VEL, MSP_MOTOR_VOLLIMIT, MSP_MOTOR_PID_P}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 2):
                requests = {MSP_MOTOR_PID_VEL_D, MSP_MOTOR_MODE, MSP_MOTOR_ACCEL_DECCEL}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 3):
                requests = {MSP_MOTOR_PID_VEL_I}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 4):

                requests = {MSP_MOTOR_FB_ICHI}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            if (self.i_couter_2 == 5):

                requests = {MSP_MOTOR_FB_VEL_ACC_DEC_M1, MSP_MOTOR_FB_VEL_ACC_DEC_M2}
                self.ask_msp_M5_6(self.request_multiple_M5_6(requests))

            self.i_couter_2 += 1
            if (self.i_couter_2 > 5):
                self.i_couter_2 = 0
            if self.communicate_on == True:
                self.MSP_id_2 = self.after(100, self.Operation_ASK_MSP_M5_6)
            else:
                self.MSP_id_2 =None
    def send_data_M5_6(self):
        data = self.send_msp_text.get()
        requests = {MSP_MOTOR_POS}#MSP_MOTOR_MEOBIET,MSP_MOTOR_POS
        #print(self.request_msp_multiple(requests))
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ask_msp_M5_6(self.request_multiple_M5_6(requests))
            #self.ser.write((data + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%M:%S')
            self.Text_msp_recieved.insert(tk.END, f"[S>{timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.Text_msp_recieved.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))
    def Read_data_Serial_M5_6(self):
        # 設定したデリミタを取得
        global c_state, err_rcvd, dataSize, p, offset, checksum, cmd, inBuf
        delimiter = DELIMITERS[self.delimiter_var_msp.get()]
        buf = ''
        while self.ser_msp_2.is_open:
            try:

                data = self.ser_msp_2.read(1).decode("latin1")#.decode("CP932")//not UTF-8
                c = int.from_bytes(data.encode('latin1'), byteorder='big', signed=False)
                #print(c)
                ###########
                if c_state == IDLE:
                    if (c ==36):
                        c_state = HEADER_START
                    else :
                        c_state = IDLE
                elif c_state == HEADER_START:
                    c_state = HEADER_M if c == 77 else IDLE

                elif c_state == HEADER_M:
                    if c == 62:
                        c_state = HEADER_ARROW
                    elif c == '33':#not use
                        c_state = HEADER_ERR
                    else:
                        c_state = IDLE
                elif c_state in (HEADER_ARROW, HEADER_ERR):
                    err_rcvd = (c_state == HEADER_ERR)  # now we are expecting the payload size
                    dataSize = c & 0xFF
                    # reset index variables
                    p = 0
                    offset = 0
                    checksum = 0
                    checksum ^= (c & 0xFF)

                    # the command is to follow
                    c_state = HEADER_SIZE
                elif c_state == HEADER_SIZE:
                    cmd = c & 0xFF
                    checksum ^= (c & 0xFF)
                    c_state = HEADER_CMD
                elif c_state == HEADER_CMD and offset < dataSize:
                    checksum ^= (c & 0xFF)
                    inBuf[offset] = c & 0xFF
                    offset += 1

                elif c_state == HEADER_CMD and offset >= dataSize:
                    # compare calculated and transferred checksum
                    if (checksum & 0xFF) == (c & 0xFF):
                        if err_rcvd:
                            print(f"Copter did not understand request type {c}")
                        else:
                            # we got a valid response packet, evaluate it
                            #print(cmd)
                            self.Operation_MSP_M5_6(cmd, int(dataSize))
                    else:
                        print(f"invalid checksum for command {cmd & 0xFF}: {checksum & 0xFF} expected, got {c & 0xFF}")
                        print(f"<{cmd & 0xFF} {dataSize & 0xFF}> {{", end=' ')
                        for i in range(dataSize):
                            if i != 0:
                                print(' ', end='')
                            print(inBuf[i] & 0xFF, end='')
                        print("} [", c, "]")
                        print(''.join(map(chr, inBuf[:dataSize])))
                    c_state = IDLE

                #######
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.Text_msp_recieved.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.Text_msp_recieved.see("end")
                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))
    def autoconnect_msp_comport_M5_6(self):
        self.port_var_msp_2.set('COM1')
        if self.ser_msp_2.is_open:
            try:
                self.ser_msp_2.close()
                self.baudrate_menu_msp_2.config(state="normal")
                self.delimiter_menu_msp_2.config(state="normal")
                self.toggle_button_msp_2.config(text="接続")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var_msp_2.get()
            if not port:
                messagebox.showerror("接続エラー", "MSPシリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var_msp_2.get())
            try:
                #######################################
                self.ser_msp_2 = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu_msp_2.config(state="disabled")
                self.delimiter_menu_msp_2.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread_msp_2 = threading.Thread(target=self.Read_data_Serial_M5_6)
                self.com_thread_msp_2.daemon = True
                self.com_thread_msp_2.start()
                self.toggle_button_msp_2.config(text="切断",background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
        #################################################MSP Protocol MOTOR 1 2 #####################################



    #################################################Main#####################################
    def Communication_serial_run(self):
        self.communicate_on = True
        self.MSP_id = self.after(100, self.Operation_ASK_MSP_M1_2)
    def Communication_serial_stop(self):
        self.communicate_on = False
        self.MSP_id = None
    ############################################################################################
    def riru_number_change(self):
        data = str(self.riru_henko.get())
        self.Socket_send("WR DM600 " + data)
    def reset_kikai(self):
        self.data_old = ''
        self.data_old1 = ''
        self.data_old2 = ''
        self.data_old_a = ''
        self.data_old_b = ''
        self.data_old_c = ''
    def add_fiter_to(self):
        self.so_fiter = self.so_fiter + 1
        if self.so_fiter == 2:
            self.filterchoosen1.place(x=5, y=40)
        if self.so_fiter == 3:
            self.filterchoosen1.place(x=5, y=40)
            self.filterchoosen2.place(x=5, y=80)
        if self.so_fiter == 4:
            self.filterchoosen1.place(x=5, y=40)
            self.filterchoosen2.place(x=5, y=80)
            self.filterchoosen3.place(x=5, y=120)
        if self.so_fiter == 5:
            self.filterchoosen1.place(x=5, y=40)
            self.filterchoosen2.place(x=5, y=80)
            self.filterchoosen3.place(x=5, y=120)
            self.filterchoosen4.place(x=5, y=160)
        if self.so_fiter >= 5:
            self.so_fiter = 5
    def minus_fiter_to(self):
        self.so_fiter = self.so_fiter - 1

        if self.so_fiter == 2:
            self.filterchoosen2.place(x=-100, y=-100)
        if self.so_fiter == 3:
            self.filterchoosen3.place(x=-100, y=-100)
        if self.so_fiter == 4:
            self.filterchoosen4.place(x=-100, y=-100)
        if self.so_fiter <= 1:
            self.filterchoosen1.place(x=-100, y=-100)
            self.so_fiter = 1
    def reset_button_on_click(self, event):
        self.Socket_send("ST R3003")
    def reset_button_off_click(self, event):
        self.Socket_send("RS R3003")
    def print_ryru_on_click(self, event):
        self.Socket_send("ST R9004")
    def print_ryru_off_click(self, event):
        self.Socket_send("RS R9004")
    def print_hako_on_click(self, event):
        self.Socket_send("ST R9005")
    def print_hako_off_click(self, event):
        self.Socket_send("RS R9005")
    def test_start_on_click(self, event):
        self.Socket_send("ST R4203")
        self.BAR_check.set('検査開始')
        self.barreceived_text.delete(1.0, tk.END)
    def test_start_off_click(self, event):
        self.Socket_send("RS R4203")
    def test_setok_on_click(self, event):
        self.Socket_send("ST R4204")
    def test_setok_off_click(self, event):
        self.Socket_send("RS R4204")
    def scan_start_on_click(self, event):
        self.Socket_send("ST R4205")
    def scan_start_off_click(self, event):
        self.Socket_send("RS R4205")
    def toridasi_ok_on_click(self, event):
        self.Socket_send("ST R4206")
        self.BAR_check.set('検査終わり')
    def toridasi_ok_off_click(self, event):
        self.Socket_send("RS R4206")
    def syudou_kirikae_switch(self):

        # Determine is on or off
        if self.syudou_kirikaeis_on:
            # self.syudou_kirikae.config(wraplength=90,text="手動モード",bg='yellow')
            self.Socket_send("ST MR3313")
            self.syudou_kirikaeis_on = False
        else:
            # self.syudou_kirikae.config(wraplength=90, text="手動モード", bg='yellow')
            # self.syudou_kirikae.config(wraplength=90, text="自動モード", height=2, width=9, bg='lightgreen',)
            self.Socket_send("RS MR3313")
            self.syudou_kirikaeis_on = True
#####################
    def slider_DM1000(self, a):
        data = str(a)
        self.Socket_send("WR DM1000 " + data)
    def slider_DM1002(self, a):
        data = str(a)
        self.Socket_send("WR DM1002 " + data)
    def slider_DM1004(self, a):
        data = str(a)
        self.Socket_send("WR DM1004 " + data)
    def slider_DM1006(self, a):
        data = str(a)
        self.Socket_send("WR DM1006 " + data)
    def slider_DM1008(self, a):
        data = str(a)
        self.Socket_send("WR DM1008 " + data)
    def slider_DM1010(self, a):
        data = str(a)
        self.Socket_send("WR DM1010 " + data)
    def slider_DM1012(self, a):
        data = str(a)
        self.Socket_send("WR DM1012 " + data)
    def slider_DM1014(self, a):
        data = str(a)
        self.Socket_send("WR DM1014 " + data)
    def slider_DM1016(self, a):
        data = str(a)
        self.Socket_send("WR DM1016 " + data)
    def slider_DM1018(self, a):
        data = str(a)
        self.Socket_send("WR DM1018 " + data)
    def slider_DM1020(self, a):
        data = str(a)
        self.Socket_send("WR DM1020 " + data)
    #########################

    def gettab_select(self, event):
        note = event.widget
        if note.tab(note.select(), "text") == "メイン":
            self.tab_select = 0
        if note.tab(note.select(), "text") == "手動":
            self.tab_select = 1
        if note.tab(note.select(), "text") == "保守１":
            self.tab_select = 2
        if note.tab(note.select(), "text") == "保守２":
            self.tab_select = 3
        if note.tab(note.select(), "text") == "設定１":
            self.tab_select = 4
        if note.tab(note.select(), "text") == "設定２":
            self.tab_select = 5
        if note.tab(note.select(), "text") == "通信":
            self.tab_select = 6
        if note.tab(note.select(), "text") == "カメラ":
            self.tab_select = 7
    def gettab1(self):
        self.tab_select = 1
        ''''''
    def gettab2(self):
        self.tab_select = 2
        ''''''
    def gettab3(self):
        self.tab_select = 3
        ''''''
    def gettab4(self):
        self.tab_select = 4
        ''''''
    def gettab5(self):
        self.tab_select = 5
        ''''''
    def gettab6(self):
        self.tab_select = 6
        ''''''
    def gettab7(self):
        self.tab_select = 7
        ''''''
    def seihin_ok(self):
        ''''''

    def clear_error_text(self):
        self.PLC_error_text.delete(1.0, tk.END)
        self.PLC_error_annai.delete(1.0, tk.END)

    ###########################################################
    def set_image_sita(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img_sita = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_sita(self.pil_img_sita.width, self.pil_img_sita.height)
        # 画像の表示
        self.draw_image_sita(self.pil_img_sita)
    def reset_transform_sita(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_sita = np.eye(3)  # 3x3の単位行列
    def translate_sita(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_sita = np.dot(mat, self.mat_affine_sita)
    def scale_sita(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_sita = np.dot(mat, self.mat_affine_sita)
    def scale_at_sita(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_sita(-cx, -cy)
        # 拡大縮小
        self.scale_sita(scale)
        # 元に戻す
        self.translate_sita(cx, cy)
    def zoom_fit_sita(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_sita.winfo_width()
        canvas_height = self.canvas_sita.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_sita()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_sita(scale)
        # あまり部分を中央に寄せる
        self.translate_sita(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------
    def draw_image_sita(self, pil_image):

        if pil_image == None:
            return

        self.canvas_sita.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_sita.winfo_width()
        canvas_height = self.canvas_sita.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_sita)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_sita = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_sita.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_sita  # 表示画像データ
        )

    ###########################################################
    def set_image_main(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_main(self.pil_img.width, self.pil_img.height)
        # 画像の表示
        self.draw_image_main(self.pil_img)
    def reset_transform_main(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_main = np.eye(3)  # 3x3の単位行列
    def translate_main(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_main = np.dot(mat, self.mat_affine_main)
    def scale_main(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_main = np.dot(mat, self.mat_affine_main)
    def scale_at_main(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_main(-cx, -cy)
        # 拡大縮小
        self.scale_main(scale)
        # 元に戻す
        self.translate_main(cx, cy)
    def zoom_fit_main(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_main()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_main(scale)
        # あまり部分を中央に寄せる
        self.translate_main(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------
    def draw_image_main(self, pil_image):

        if pil_image == None:
            return

        self.canvas.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_main)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_main = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_main  # 表示画像データ
        )

    ###########################################################
    def set_image_hoshu1(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img_hoshu1 = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_hoshu1(self.pil_img_hoshu1.width, self.pil_img_hoshu1.height)
        # 画像の表示
        self.draw_image_hoshu1(self.pil_img_hoshu1)
    def reset_transform_hoshu1(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_hoshu1 = np.eye(3)  # 3x3の単位行列
    def translate_hoshu1(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_hoshu1 = np.dot(mat, self.mat_affine_hoshu1)
    def scale_hoshu1(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_hoshu1 = np.dot(mat, self.mat_affine_hoshu1)
    def scale_at_hoshu1(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_hoshu1(-cx, -cy)
        # 拡大縮小
        self.scale_hoshu1(scale)
        # 元に戻す
        self.translate_hoshu1(cx, cy)
    def zoom_fit_hoshu1(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu1.winfo_width()
        canvas_height = self.canvas_hoshu1.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_hoshu1()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_hoshu1(scale)
        # あまり部分を中央に寄せる
        self.translate_hoshu1(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------
    def draw_image_hoshu1(self, pil_image):

        if pil_image == None:
            return

        self.canvas_hoshu1.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu1.winfo_width()
        canvas_height = self.canvas_hoshu1.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_hoshu1)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_hoshu1 = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_hoshu1.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_hoshu1  # 表示画像データ
        )

    ######################################################################
    def set_image_hoshu2(self, filename):
        ''' 画像ファイルを開く '''
        if not filename:
            return
        # PIL.Imageで開く
        self.pil_img_hoshu2 = Image.open(filename)
        # 画像全体に表示するようにアフィン変換行列を設定
        self.zoom_fit_hoshu2(self.pil_img_hoshu2.width, self.pil_img_hoshu2.height)
        # 画像の表示
        self.draw_image_hoshu2(self.pil_img_hoshu2)
    def reset_transform_hoshu2(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine_hoshu2 = np.eye(3)  # 3x3の単位行列
    def translate_hoshu2(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine_hoshu2 = np.dot(mat, self.mat_affine_hoshu2)
    def scale_hoshu2(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine_hoshu2 = np.dot(mat, self.mat_affine_hoshu2)
    def scale_at_hoshu2(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate_hoshu2(-cx, -cy)
        # 拡大縮小
        self.scale_hoshu2(scale)
        # 元に戻す
        self.translate_hoshu2(cx, cy)
    def zoom_fit_hoshu2(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu2.winfo_width()
        canvas_height = self.canvas_hoshu2.winfo_height()
        # print(canvas_width,canvas_height)
        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform_hoshu2()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale_hoshu2(scale)
        # あまり部分を中央に寄せる
        self.translate_hoshu2(offsetx, offsety)

        # -------------------------------------------------------------------------------
        # 描画
        # -------------------------------------------------------------------------------
    def draw_image_hoshu2(self, pil_image):

        if pil_image == None:
            return

        self.canvas_hoshu2.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_hoshu2.winfo_width()
        canvas_height = self.canvas_hoshu2.winfo_height()
        # print(canvas_width,canvas_height)
        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine_hoshu2)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor=self.back_color
        )

        # 表示用画像を保持
        self.image_hoshu2 = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_hoshu2.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image_hoshu2  # 表示画像データ
        )

    ######################### SSEERRIIAALL##########################
    def refresh_serial_ports(self):
        global port_var
        ports = serial.tools.list_ports.comports()
        port_names = [port.device for port in ports]
        self.port_var.set('')  # 初期値を空に設定
        self.port_menu["menu"].delete(0, tk.END)
        for port_name in port_names:
            self.port_menu["menu"].add_command(label=port_name, command=tk._setit(self.port_var, port_name))
    def autoconnect_comport(self):
        self.port_var.set('COM6')
        if self.ser.is_open:
            try:
                self.ser.close()
                self.baudrate_menu.config(state="normal")
                self.delimiter_menu.config(state="normal")
                self.toggle_button.config(text="接続",background="#E53333")

                # messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var.get())
            try:
                #######################################
                self.ser = Serial(port, baudrate)

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu.config(state="disabled")
                self.delimiter_menu.config(state="disabled")

                # messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread = threading.Thread(target=self.read_comport_data)
                self.com_thread.daemon = True
                self.com_thread.start()
                self.toggle_button.config(text="切断",background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
    def toggle_serial_port(self):

        if self.ser.is_open:
            try:
                self.ser.close()
                self.baudrate_menu.config(state="normal")
                self.delimiter_menu.config(state="normal")
                self.toggle_button.config(text="接続",background="#E53333")

                messagebox.showinfo("情報", "シリアルポートが切断されました")
            except Exception as e:
                messagebox.showerror("切断エラー", str(e))
        else:
            port = self.port_var.get()
            if not port:
                messagebox.showerror("接続エラー", "シリアルポートが選択されていません")
                return
            baudrate = int(self.baudrate_var.get())
            try:
                #######################################
                self.ser = Serial(port, baudrate, parity='N')  # N, parity= 'N'

                # ポート接続中はボーレート・デリミタを変更できないように
                self.baudrate_menu.config(state="disabled")
                self.delimiter_menu.config(state="disabled")

                messagebox.showinfo("情報", "シリアルポートに接続しました")
                # 受信スレッドを開始
                self.com_thread = threading.Thread(target=self.read_comport_data)
                self.com_thread.daemon = True
                self.com_thread.start()
                self.toggle_button.config(text="切断",background="#3AFA45")
            except Exception as e:
                messagebox.showerror("受信エラー", str(e))
    def send_data(self):
        # 送信ボックスの入力値を取得
        data = self.entry_text.get()
        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信
            self.ser.write((data + delimiter).encode())

            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%H:%M:%S')
            self.received_text.insert(tk.END, f"[Send -> {timestamp}] {data}{REV_DELIMITER[delimiter]}\n")
            self.received_text.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))
    def Serial_send_button(self, data_comsend):
        # 送信ボックスの入力値を取得

        # デリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        try:
            # 送信

            self.ser.write((data_comsend + delimiter).encode())
            # 履歴フィールドに時刻とともに表示
            timestamp = datetime.datetime.now().strftime('%H:%M:%S')
            self.received_text.insert(tk.END, f"[Send -> {timestamp}] {data_comsend}{REV_DELIMITER[delimiter]}\n")
            self.received_text.see("end")  # 自動スクロール

        except Exception as e:
            messagebox.showerror("送信エラー", str(e))
    def read_comport_data(self):
        # 設定したデリミタを取得
        delimiter = DELIMITERS[self.delimiter_var.get()]
        buf = ''
        while self.ser.is_open:
            try:
                data = self.ser.read(1).decode()
                if data:
                    buf += data
                    # デリミタを含む場合
                    if delimiter == '':
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.received_text.insert(tk.END, f">>{timestamp} {buf}\n")
                        buf = ''  # 受信バッファをクリア

                    elif delimiter in buf:

                        # 表示用にデリミタを変換（'\r'を'[CR]'に）
                        buf = buf.replace(delimiter, REV_DELIMITER[delimiter])
                        # 履歴フィールドにタイムスタンプと共に表示
                        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                        self.received_text.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.received_text.see("end")
                        self.barreceived_text.insert(tk.END, f">>{timestamp} {buf}\n")
                        self.barreceived_text.see("end")
                        self.barcoder_data = buf

                        buf = ''  # 受信バッファをクリア
            # finally:
            #    return
            except Exception as e:
                messagebox.showerror("Serial受信エラー", str(e))
    def data_caseread(self):
        ''''''
    def motor_para(self):
        ''''''

    # 履歴をクリア
    def clear_received_text(self):
        self.received_text.delete(1.0, tk.END)

    ################################################################ SOCKET ##########################
    def socket_connect(self):

        if self.PLC_CN_OK == 0:

            while True:
                try:
                    # 01. Preparing Socket : socket()
                    self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    self.sock.connect((self.IP_PLC.get(), self.PORT_PLC.get()))
                    mess = "ST MR0" + '\r'
                    self.sock.send(mess.encode('utf-8'))
                    rec_mess = self.sock.recv(MAX_MESSAGE).decode('utf-8')
                    # print(rec_mess)
                    self.sock.close()
                    time.sleep(0.01)

                    if rec_mess == 'OK' + '\r\n':
                        self.PLC_stt.delete(1.0, tk.END)
                        print('PLC connected')
                        self.PLC_stt.insert(tk.END, "OK")
                        self.PLC_CN_OK = 1
                        for i in range(1000, 1022, 2):
                            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                            self.sock.connect((self.IP_PLC.get(), self.PORT_PLC.get()))
                            # print(i)
                            mess = "RD DM" + str(i) + '\r'
                            self.sock.send(mess.encode('utf-8'))
                            # self.sock.settimeout(0.1)
                            rec_mess = self.sock.recv(MAX_MESSAGE).decode('utf-8')
                            int_mess = int(rec_mess)
                            # print(int_mess)
                            if i == 1000:
                                self.DM1000_var.set(int_mess)
                            if i == 1002:
                                self.DM1002_var.set(int_mess)
                            if i == 1004:
                                self.DM1004_var.set(int_mess)
                            if i == 1006:
                                self.DM1006_var.set(int_mess)
                            if i == 1008:
                                self.DM1008_var.set(int_mess)
                            if i == 1010:
                                self.DM1010_var.set(int_mess)
                            if i == 1012:
                                self.DM1012_var.set(int_mess)
                            if i == 1014:
                                self.DM1014_var.set(int_mess)
                            if i == 1016:
                                self.DM1016_var.set(int_mess)
                            if i == 1018:
                                self.DM1018_var.set(int_mess)
                            if i == 1020:
                                self.DM1020_var.set(int_mess)

                            self.sock.close()
                            time.sleep(0.01)

                    self.Main_update()
                    break
                except Exception as e:
                    self.PLC_stt.insert(tk.END, "NG")
                    messagebox.showerror("PLC 送信エラー", str(e))
                    self.sock.close()
                    break
    def Socket_send(self, mess):
        self.send_flag = 1
        while True:
            try:
                # 通信の確立
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.connect((self.IP_PLC.get(), self.PORT_PLC.get()))
                mess = mess + '\r'
                # sock.send(mess)

                # メッセージ送信
                sock.send(mess.encode('utf-8'))  # ASCII#'utf-8'
                timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                self.Text_sock_recieved.insert(tk.END, f"⏎{timestamp} {mess}\n")
                self.Text_sock_recieved.see("end")  # 自動スクロール
                ###
                rec_mess = sock.recv(MAX_MESSAGE).decode('utf-8')
                timestamp = datetime.datetime.now().strftime('%H:%M:%S')
                self.Text_sock_recieved.insert(tk.END, f"▶{timestamp} {rec_mess}\n")
                self.Text_sock_recieved.see("end")  # 自動スクロール

                # 通信の終了
                sock.close()
                self.send_flag = 0
                break

            except Exception as e:
                self.draw_id = None
                messagebox.showerror("PLC 送信エラー", str(e))
                sock.close()

                break
    def Socket_receive(self):

        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((HOST, PORT))
        # mess = ""
        mess = sock.recv(MAX_MESSAGE)
        mess = mess.decode('utf-8')
        timestamp = datetime.datetime.now().strftime('%H:%M:%S')
        self.Text_sock_recieved.insert(tk.END, f"▶{timestamp} {mess}\n")
        self.Text_sock_recieved.see("end")  # 自動スクロール

        self.sock.close()
        # self.recv_id = self.after(10, self.Socket_receive)


    ####################### CAMERA SET ##########################
    def camera_callback(self):
        # if self.read_defaut == False:
        #    self.camera_defaut()
        #    self.read_defaut = True

        ret, frame = self.capture.read()
        self.cv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        self.cv_image_base = self.cv_image
        # self.cv_background = self.cv_image.copy()
        if self.check_hani_ok.get() == 1:
            if self.check_patan6.get() ==1:
                #self.cv_image = self.area_select_special(self.cv_image)
                self.cv_image = self.area_select(self.cv_image)
                self.area_roi = self.color_map_select(self.area_roi)
                width, height = self.area_roi.shape[:2]
                self.area_roi = self.cam_filter(self.area_roi)
                if self.circle_select.get() == 1:
                    self.area_roi = self.cam_circle_elipse_detect(self.area_roi)
                if self.circle_select_adv.get() == 1:
                    self.area_roi = self.circle_detect(self.area_roi)
                if self.triang_select.get() == 1:
                    self.area_roi = self.triangle_detect(self.area_roi)
                if self.rect_select.get() == 1:
                    self.area_roi = self.rectange_detect(self.area_roi)
                if self.edge_select.get() == 1:
                    self.area_roi = self.edge_detect(self.area_roi)
                if self.draw_contour_select.get() == 1:
                    self.area_roi = self.draw_contour_object(self.area_roi)
                if self.connected_select.get() == 1:
                    self.area_roi = self.connected_component(self.area_roi)
                if self.rectangge_component_select.get() == 1:
                    self.area_roi = self.rectangge_component(self.area_roi)

                self.cv_image[self.pt1[1]:self.pt1[1] + width, self.pt1[0]:self.pt1[0] + height] = self.area_roi

            if self.check_patan2.get() == 1:
                self.cv_image = self.color_map_select(self.cv_image)
                w_3c = np.full_like(self.cv_image, fill_value=(255, 255, 255))#Create white image
                #
                roi_corners = np.array([[self.pt1_specal, self.pt2_specal, self.pt3_specal]], dtype=np.int32)
                d_3c = np.zeros_like(self.cv_image[:, :], dtype='uint8')#Create mask image
                cv2.fillPoly(d_3c, roi_corners, [255] * 3)
                d_1c = d_3c[:, :, 0]
                masked = cv2.bitwise_and(self.cv_image, w_3c, mask=d_1c)#black mask
                # Add white background
                d_3c_i = ~d_3c
                final = cv2.bitwise_or(self.cv_image, d_3c_i)

                masked_image = self.cam_filter(final)

                if self.circle_select.get() == 1:
                    masked_image = self.cam_circle_elipse_detect(masked_image)
                if self.circle_select_adv.get() == 1:
                    masked_image = self.circle_detect(masked_image)
                if self.triang_select.get() == 1:
                    masked_image = self.triangle_detect(masked_image)
                if self.rect_select.get() == 1:
                    masked_image = self.rectange_detect(masked_image)
                if self.edge_select.get() == 1:
                    masked_image = self.edge_detect(masked_image)
                if self.draw_contour_select.get() == 1:
                    masked_image = self.draw_contour_object(masked_image)
                if self.connected_select.get() == 1:
                    masked_image = self.connected_component(masked_image)
                if self.rectangge_component_select.get() == 1:
                    masked_image = self.rectangge_component(masked_image)



                cv_image_sel = self.area_select_3kaku(masked_image)
                final_select = cv2.bitwise_and(cv_image_sel, self.cv_image_base)

                self.cv_image = final_select

            if self.check_patan3.get() == 1:
                self.cv_image = self.color_map_select(self.cv_image)
                w_3c = np.full_like(self.cv_image, fill_value=(255, 255, 255))  # Create white image
                #
                roi_corners = np.array([[self.pt1_kaku, self.pt2_kaku, self.pt3_kaku,self.pt4_kaku]], dtype=np.int32)
                d_3c = np.zeros_like(self.cv_image[:, :], dtype='uint8')  # Create mask image
                cv2.fillPoly(d_3c, roi_corners, [255] * 3)
                d_1c = d_3c[:, :, 0]
                masked = cv2.bitwise_and(self.cv_image, w_3c, mask=d_1c)  # black mask
                # Add white background
                d_3c_i = ~d_3c
                final = cv2.bitwise_or(self.cv_image, d_3c_i)

                masked_image = self.cam_filter(final)
                if self.circle_select.get() == 1:
                    masked_image = self.cam_circle_elipse_detect(masked_image)
                if self.circle_select_adv.get() == 1:
                    masked_image = self.circle_detect(masked_image)
                if self.triang_select.get() == 1:
                    masked_image = self.triangle_detect(masked_image)
                if self.rect_select.get() == 1:
                    masked_image = self.rectange_detect(masked_image)
                if self.edge_select.get() == 1:
                    masked_image = self.edge_detect(masked_image)
                if self.draw_contour_select.get() == 1:
                    masked_image = self.draw_contour_object(masked_image)
                if self.connected_select.get() == 1:
                    masked_image = self.connected_component(masked_image)
                if self.rectangge_component_select.get() == 1:
                    masked_image = self.rectangge_component(masked_image)
                cv_image_sel = self.area_select_4kaku(masked_image)
                # self.cv_image  = cv2.bitwise_and(self.cv_image, masked_image)
                # self.cv_image = masked_image
                # self.cv_image[self.pt1[1]:self.pt1[1] + width, self.pt1[0]:self.pt1[0] + height] = masked_image
                final_select = cv2.bitwise_and(cv_image_sel, self.cv_image_base)

                self.cv_image = final_select

            if self.check_patan4.get() == 1:
                self.cv_image = self.color_map_select(self.cv_image)
                w_3c = np.full_like(self.cv_image, fill_value=(255, 255, 255))  # Create white image
                #
                roi_corners = np.array([[self.pt1_polygol5, self.pt2_polygol5, self.pt3_polygol5,self.pt4_polygol5,self.pt5_polygol5]], dtype=np.int32)
                d_3c = np.zeros_like(self.cv_image[:, :], dtype='uint8')  # Create mask image
                cv2.fillPoly(d_3c, roi_corners, [255] * 3)
                d_1c = d_3c[:, :, 0]
                masked = cv2.bitwise_and(self.cv_image, w_3c, mask=d_1c)  # black mask
                # Add white background
                d_3c_i = ~d_3c
                final = cv2.bitwise_or(self.cv_image, d_3c_i)

                masked_image = self.cam_filter(final)
                if self.circle_select.get() == 1:
                    masked_image = self.cam_circle_elipse_detect(masked_image)
                if self.circle_select_adv.get() == 1:
                    masked_image = self.circle_detect(masked_image)
                if self.triang_select.get() == 1:
                    masked_image = self.triangle_detect(masked_image)
                if self.rect_select.get() == 1:
                    masked_image = self.rectange_detect(masked_image)
                if self.edge_select.get() == 1:
                    masked_image = self.edge_detect(masked_image)
                if self.draw_contour_select.get() == 1:
                    masked_image = self.draw_contour_object(masked_image)
                if self.connected_select.get() == 1:
                    masked_image = self.connected_component(masked_image)
                if self.rectangge_component_select.get() == 1:
                    masked_image = self.rectangge_component(masked_image)
                cv_image_sel = self.area_select_5kaku(masked_image)
                # self.cv_image  = cv2.bitwise_and(self.cv_image, masked_image)
                # self.cv_image = masked_image
                # self.cv_image[self.pt1[1]:self.pt1[1] + width, self.pt1[0]:self.pt1[0] + height] = masked_image
                final_select = cv2.bitwise_and(cv_image_sel, self.cv_image_base)

                self.cv_image = final_select

        self.pil_image = Image.fromarray(self.cv_image)
        if self.zoom_load_defaut == False:
            self.zoom_fit(self.pil_image.width, self.pil_image.height)
            self.zoom_load_defaut = True

        self.redraw_image()
        # print(self.Camera_check)
        # self.disp_image(frame)
        if self.capture_switchis_on == True:
            self.disp_id = self.after(10, self.camera_callback)
        else:
            self.disp_id = None
    def camera_defaut(self):
        a = self.capture.get(cv2.CAP_PROP_FRAME_WIDTH)
        self.FRAME_WIDTH_var.set(a)
        b = self.capture.get(cv2.CAP_PROP_FRAME_HEIGHT)
        self.FRAME_HEIGHT_var.set(b)
        c = self.capture.get(cv2.CAP_PROP_FPS)
        self.FPS_var.set(30)
        d = self.capture.get(cv2.CAP_PROP_BRIGHTNESS)
        self.BRIGHTNESS_var.set(0)
        e = self.capture.get(cv2.CAP_PROP_CONTRAST)
        self.CONTRAST_var.set(0)
        f = self.capture.get(cv2.CAP_PROP_SATURATION)
        self.SATURATION_var.set(0)
        g = self.capture.get(cv2.CAP_PROP_HUE)
        self.HUE_var.set(0)
        h = self.capture.get(cv2.CAP_PROP_GAIN)
        self.GAIN_var.set(0)
        i = self.capture.get(cv2.CAP_PROP_EXPOSURE)
        self.GAIN_var.set(-2)
        j = self.capture.get(cv2.CAP_PROP_AUTO_EXPOSURE)
        self.GAIN_var.set(1)
        k = self.capture.get(cv2.CAP_PROP_AUTOFOCUS)
        self.GAIN_var.set(1)
        #
        # print(a,b,c,d,e,f,g,h)
    def camera_setting(self, a):
        self.capture.set(cv2.CAP_PROP_FRAME_WIDTH, self.FRAME_WIDTH_var.get())
    def camera_setting1(self, a):
        self.capture.set(cv2.CAP_PROP_FRAME_HEIGHT, self.FRAME_HEIGHT_var.get())
    def camera_setting2(self, a):
        self.capture.set(cv2.CAP_PROP_FPS, self.FPS_var.get())
    def camera_setting3(self, a):
        self.capture.set(cv2.CAP_PROP_BRIGHTNESS, self.BRIGHTNESS_var.get())
    def camera_setting4(self, a):
        self.capture.set(cv2.CAP_PROP_CONTRAST, self.CONTRAST_var.get())
    def camera_setting5(self, a):
        self.capture.set(cv2.CAP_PROP_SATURATION, self.SATURATION_var.get())
    def camera_setting6(self, a):
        self.capture.set(cv2.CAP_PROP_HUE, self.HUE_var.get())
    def camera_setting7(self, a):
        self.capture.set(cv2.CAP_PROP_GAIN, self.GAIN_var.get())
    def camera_setting8(self, a):
        self.capture.set(cv2.CAP_PROP_AUTO_EXPOSURE, self.auto_exposure_var.get())
    def camera_setting9(self, a):
        self.capture.set(cv2.CAP_PROP_AUTOFOCUS, self.auto_focus_var.get())
    def camera_setting10(self, a):
        self.capture.set(cv2.CAP_PROP_EXPOSURE, self.EXPOSE_var.get())
    def reset_data(self):
        self.capture.set(cv2.CAP_PROP_BRIGHTNESS, 0)
        self.capture.set(cv2.CAP_PROP_CONTRAST, 32)
        self.capture.set(cv2.CAP_PROP_SATURATION, 55)
        self.capture.set(cv2.CAP_PROP_HUE, 0)
        self.capture.set(cv2.CAP_PROP_GAIN, 0)
        #
        self.colorchoosen.current(0)
        self.boder_filter_map.current(0)
        self.boder_filter_map1.current(0)
        self.scale_var.set(50)
        self.adaptiveThreshold_var.set(0.01)
        self.canny1_var.set(100)
        self.canny2_var.set(500)
        self.gaussion_var.set(0)
        self.Blur_var.set(0)
        self.bold_var.set(0)
        self.debold_var.set(0)
        self.bold_var1.set(0)
        self.debold_var1.set(0)
        self.filterchoosen.current(0)
        self.filterchoosen1.current(0)
        self.filterchoosen2.current(0)
        self.filterchoosen3.current(0)
        self.filterchoosen4.current(0)
        self.so_fiter = 1
        self.fil_by_Area.set(1)
        self.filByCircularity.set(1)
        self.filByConvexity.set(1)
        self.filByInertia.set(1)
        self.minArea_var.set(50)
        self.minCircularity_var.set(0.1)
        self.minConvexity_var.set(0.1)
        self.minInertia_var.set(0.01)
        self.Kernel_var.set(3)
        self.param1_var.set(50)
        self.param2_var.set(30)
        self.minRadius_var.set(1)
        self.maxRadius_var.set(40)
        self.var_canny_a.set(250)
        self.var_canny_b.set(150)
        self.var_thr.set(50)
        self.len_size2.set(50)
        self.len_size3.set(15)
        self.len_size1.set(1)
        self.circle_select.set(0)
        self.circle_select_adv.set(0)
        self.triang_select.set(0)
        self.rect_select.set(0)
        self.edge_select.set(0)
        self.boder_map.current(0)
        self.contua_map.current(0)
        self.contua_map1.current(1)
        self.rect_para1.set(0.04)
        self.area_rect.set(150)
        self.chuvi_rect.set(50)
        self.rect_thres.set(150)
        #
        self.boder_map2.current(0)
        self.contua_map2.current(0)
        self.contua_map21.current(1)
        self.trian_para1.set(0.04)
        self.area_trian.set(150)
        self.chuvi_trian.set(50)
        self.trian_thres.set(150)
        self.add_fiter_to()
        self.minus_fiter_to()

        #
    def capture_switch(self):
        if self.capture_switchis_on:
            self.video_capture.config(bg='#FF4C00')
            self.video_capture.config(text='Continue_enable')
            self.camera_callback()
            self.capture_switchis_on = False
        else:

            self.video_capture.config(bg='gray')
            self.video_capture.config(text='Continue_disable')
            self.capture_switchis_on = True

    # $#################### CAMERA ZOOM ####################
    def mouse_move(self, event):
        ''' マウスの移動時
        # マウス座標
        self.mouse_position["text"] = f"mouse(x, y) = ({event.x: 4d}, {event.y: 4d})"

        if self.pil_image == None:
            return

        # 画像座標
        mouse_posi = np.array([event.x, event.y, 1])  # マウス座標(numpyのベクトル)
        mat_inv = np.linalg.inv(self.mat_affine)  # 逆行列（画像→Cancasの変換からCanvas→画像の変換へ）
        image_posi = np.dot(mat_inv, mouse_posi)  # 座標のアフィン変換
        x = int(np.floor(image_posi[0]))
        y = int(np.floor(image_posi[1]))
        if x >= 0 and x < self.pil_image.width and y >= 0 and y < self.pil_image.height:
            # 輝度値の取得
            value = self.pil_image.getpixel((x, y))
            self.image_position["text"] = f"image({x: 4d}, {y: 4d}) = {value}"
        else:
            self.image_position["text"] = "-------------------------" '''
    def mouse_move_mid(self, event):
        if self.pil_image == None:
            return
        self.translate(event.x - self.__old_event.x, event.y - self.__old_event.y)
        self.redraw_image()  # 再描画
        self.__old_event = event
    def mouse_down_mid(self, event):
        ''' マウスの左ボタンをドラッグ '''
        self.__old_event = event
    def mouse_move_right(self,event):
        ''''''
    def mouse_down_right(self,event):
        ''''''
    def mouse_up_right(self, event):
        ''''''
    def mouse_down_left(self, event):
        ''' マウスの左ボタンを押した '''
        if self.pil_image == None:
            return
        self.__old_event = event
        mouse_posi = np.array([event.x, event.y, 1])  # マウス座標(numpyのベクトル)
        mat_inv = np.linalg.inv(self.mat_affine)  # 逆行列（画像→Cancasの変換からCanvas→画像の変換へ）
        image_posi = np.dot(mat_inv, mouse_posi)  # 座標のアフィン変換
        x = int(np.floor(image_posi[0]))
        y = int(np.floor(image_posi[1]))
        if x >= 0 and x < self.pil_image.width and y >= 0 and y < self.pil_image.height and self.cam_edit_hani ==True:
            # 輝度値の取得
            if self.check_patan6.get() == 1:
                if abs(x - self.pt1[0]) < 16 and abs(y - self.pt1[1]) < 16:
                    self.edit_pt1 = True
                if abs(x - self.pt2[0]) < 16 and abs(y - self.pt2[1]) < 16:
                    self.edit_pt2 = True
            ####
            if self.check_patan2.get() == 1:
                if abs(x - self.pt1_specal[0]) < 16 and abs(y - self.pt1_specal[1]) < 16:
                    self.edit_pt1_specal = True
                if abs(x - self.pt2_specal[0]) < 16 and abs(y - self.pt2_specal[1]) < 16:
                    self.edit_pt2_specal = True
                if abs(x - self.pt3_specal[0]) < 16 and abs(y - self.pt3_specal[1]) < 16:
                    self.edit_pt3_specal = True
            ########
            if self.check_patan3.get() == 1:
                if abs(x - self.pt1_kaku[0]) < 16 and abs(y - self.pt1_kaku[1]) < 16:
                    self.edit_pt1_kaku = True
                if abs(x - self.pt2_kaku[0]) < 16 and abs(y - self.pt2_kaku[1]) < 16:
                    self.edit_pt2_kaku = True
                if abs(x - self.pt3_kaku[0]) < 16 and abs(y - self.pt3_kaku[1]) < 16:
                    self.edit_pt3_kaku = True
                if abs(x - self.pt4_kaku[0]) < 16 and abs(y - self.pt4_kaku[1]) < 16:
                    self.edit_pt4_kaku = True
            if self.check_patan4.get() ==1:
                if abs(x - self.pt1_polygol5[0]) < 16 and abs(y - self.pt1_polygol5[1]) < 16:
                    self.edit_pt1_polygol5 = True
                if abs(x - self.pt2_polygol5[0]) < 16 and abs(y - self.pt2_polygol5[1]) < 16:
                    self.edit_pt2_polygol5 = True
                if abs(x - self.pt3_polygol5[0]) < 16 and abs(y - self.pt3_polygol5[1]) < 16:
                    self.edit_pt3_polygol5 = True
                if abs(x - self.pt4_polygol5[0]) < 16 and abs(y - self.pt4_polygol5[1]) < 16:
                    self.edit_pt4_polygol5 = True
                if abs(x - self.pt5_polygol5[0]) < 16 and abs(y - self.pt5_polygol5[1]) < 16:
                    self.edit_pt5_polygol5 = True
    def mouse_up_left(self, event):
        self.edit_pt1 = False
        self.edit_pt2 = False
        self.edit_pt1_specal = False
        self.edit_pt2_specal = False
        self.edit_pt3_specal = False
        self.edit_pt1_kaku = False
        self.edit_pt2_kaku = False
        self.edit_pt3_kaku = False
        self.edit_pt4_kaku = False

        self.edit_pt1_polygol5 = False
        self.edit_pt2_polygol5 = False
        self.edit_pt3_polygol5 = False
        self.edit_pt4_polygol5 = False
        self.edit_pt5_polygol5 = False
    def mouse_move_left(self, event):
        if self.pil_image == None:
            return
        mouse_posi = np.array([event.x, event.y, 1])  # マウス座標(numpyのベクトル)
        mat_inv = np.linalg.inv(self.mat_affine)  # 逆行列（画像→Cancasの変換からCanvas→画像の変換へ）
        image_posi = np.dot(mat_inv, mouse_posi)  # 座標のアフィン変換
        x = int(np.floor(image_posi[0]))
        y = int(np.floor(image_posi[1]))
        if x >= 0 and x < self.pil_image.width and y >= 0 and y < self.pil_image.height:
            if self.edit_pt1 == True:
                # print("ohshit")
                self.pt1 = (x, y)
            if self.edit_pt2 == True:
                # print("ohshit")
                self.pt2 = (x, y)
            ###############
            if self.edit_pt1_specal == True:
                # print("ohshit")
                self.pt1_specal = (x, y)
            if self.edit_pt2_specal == True:
                # print("ohshit")
                self.pt2_specal = (x, y)
            if self.edit_pt3_specal == True:
                # print("ohshit")
                self.pt3_specal = (x, y)
            ###########
            if self.edit_pt1_kaku == True:
                self.pt1_kaku = (x, y)
            if self.edit_pt2_kaku == True:
                self.pt2_kaku = (x, y)
            if self.edit_pt3_kaku == True:
                self.pt3_kaku = (x, y)
            if self.edit_pt4_kaku == True:
                self.pt4_kaku = (x, y)
            ##########
            if self.edit_pt1_polygol5 == True:
                self.pt1_polygol5 = (x, y)
            if self.edit_pt2_polygol5 == True:
                self.pt2_polygol5 = (x, y)
            if self.edit_pt3_polygol5 == True:
                self.pt3_polygol5 = (x, y)
            if self.edit_pt4_polygol5 == True:
                self.pt4_polygol5 = (x, y)
            if self.edit_pt5_polygol5 == True:
                self.pt5_polygol5 = (x, y)
    def mouse_double_click_left(self, event):
        ''' マウスの左ボタンをダブルクリック '''
        if self.pil_image == None:
            return
        self.zoom_fit(self.pil_image.width, self.pil_image.height)
        self.redraw_image()  # 再描画
    def mouse_wheel(self, event):
        ''' マウスホイールを回した '''
        if self.pil_image == None:
            return

        if (event.delta < 0):
            # 上に回転の場合、縮小
            self.scale_at(0.8, event.x, event.y)
        else:
            # 下に回転の場合、拡大
            self.scale_at(1.25, event.x, event.y)

        self.redraw_image()  # 再描画
    def reset_transform(self):
        '''アフィン変換を初期化（スケール１、移動なし）に戻す'''
        self.mat_affine = np.eye(3)  # 3x3の単位行列
    def translate(self, offset_x, offset_y):
        ''' 平行移動 '''
        mat = np.eye(3)  # 3x3の単位行列
        mat[0, 2] = float(offset_x)
        mat[1, 2] = float(offset_y)

        self.mat_affine = np.dot(mat, self.mat_affine)
    def scale(self, scale: float):
        ''' 拡大縮小 '''
        mat = np.eye(3)  # 単位行列
        mat[0, 0] = scale
        mat[1, 1] = scale

        self.mat_affine = np.dot(mat, self.mat_affine)
    def scale_at(self, scale: float, cx: float, cy: float):
        ''' 座標(cx, cy)を中心に拡大縮小 '''

        # 原点へ移動
        self.translate(-cx, -cy)
        # 拡大縮小
        self.scale(scale)
        # 元に戻す
        self.translate(cx, cy)
    def zoom_fit(self, image_width, image_height):
        '''画像をウィジェット全体に表示させる'''

        # キャンバスのサイズ
        canvas_width = self.canvas_cam.winfo_width()
        canvas_height = self.canvas_cam.winfo_height()

        if (image_width * image_height <= 0) or (canvas_width * canvas_height <= 0):
            return

        # アフィン変換の初期化
        self.reset_transform()

        scale = 1.0
        offsetx = 0.0
        offsety = 0.0

        if (canvas_width * image_height) > (image_width * canvas_height):
            # ウィジェットが横長（画像を縦に合わせる）
            scale = canvas_height / image_height
            # あまり部分の半分を中央に寄せる
            offsetx = (canvas_width - image_width * scale) / 2
        else:
            # ウィジェットが縦長（画像を横に合わせる）
            scale = canvas_width / image_width
            # あまり部分の半分を中央に寄せる
            offsety = (canvas_height - image_height * scale) / 2

        # 拡大縮小
        self.scale(scale)
        # あまり部分を中央に寄せる
        self.translate(offsetx, offsety)
    def draw_image(self, pil_image):

        if pil_image == None:
            return

        self.canvas_cam.delete("all")

        # キャンバスのサイズ
        canvas_width = self.canvas_cam.winfo_width()
        canvas_height = self.canvas_cam.winfo_height()

        # キャンバスから画像データへのアフィン変換行列を求める
        # （表示用アフィン変換行列の逆行列を求める）
        mat_inv = np.linalg.inv(self.mat_affine)

        # PILの画像データをアフィン変換する
        dst = pil_image.transform(
            (canvas_width, canvas_height),  # 出力サイズ
            Image.AFFINE,  # アフィン変換
            tuple(mat_inv.flatten()),  # アフィン変換行列（出力→入力への変換行列）を一次元のタプルへ変換
            Image.NEAREST,  # 補間方法、ニアレストネイバー
            fillcolor="#CBD3C0"
        )

        # 表示用画像を保持
        self.image = ImageTk.PhotoImage(image=dst)

        # 画像の描画
        self.canvas_cam.create_image(
            0, 0,  # 画像表示位置(左上の座標)
            anchor='nw',  # アンカー、左上が原点
            image=self.image  # 表示画像データ
        )
    def redraw_image(self):
        ''' 画像の再描画 '''
        if self.pil_image == None:
            return
        self.draw_image(self.pil_image)

    ############## AREA SELECT ##############
    def area_select(self, cv_image):
        rect_x0 = self.pt1[0] - 5  # -3
        if rect_x0 < 0:
            rect_x0 = 0
        rect_y0 = self.pt1[1] - 5  # -3
        if rect_y0 < 0:
            rect_y0 = 0
        rect_ofset1 = [rect_x0, rect_y0]
        ###########3
        rect_x1 = self.pt2[0] + 5  # +3
        # if rect_x1 < 0:
        #    rect_x1 = 0
        rect_y1 = self.pt2[1] + 5  # +3
        # if rect_y1 > self.:
        #    rect_y1 = 0
        rect_ofset2 = [rect_x1, rect_y1]

        cv_image = cv2.rectangle(cv_image, rect_ofset1, rect_ofset2, (0, 255, 0), 2)
        # cv_image = cv2.rectangle(cv_image, self.pt1, self.pt2, (255, 255, 255), 1)
        cv_image = cv2.circle(cv_image, (self.pt1[0] - 5, self.pt1[1] - 5), 5, (255, 0, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2[0] + 5, self.pt2[1] + 5), 5, (255, 0, 0), -1)

        roi = (self.pt1[0] + 1, self.pt1[1] + 1, self.pt2[0] - 1, self.pt2[1] - 1)
        self.area_roi = cv_image[roi[1]: roi[3], roi[0]: roi[2]]

        return cv_image
    def area_select_3kaku(self, cv_image):

        cv_image = cv2.circle(cv_image, (self.pt1_specal[0] , self.pt1_specal[1] ), 5, (0, 255, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2_specal[0] , self.pt2_specal[1] ), 5, (0, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt3_specal[0] , self.pt3_specal[1] ), 5, (255, 0, 0), -1)
        cv_image = cv2.line(cv_image,self.pt1_specal,self.pt2_specal,(0,255,255),2)
        cv_image = cv2.line(cv_image, self.pt2_specal, self.pt3_specal, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt3_specal, self.pt1_specal, (0, 255, 255), 2)
        return cv_image
    def area_select_4kaku(self, cv_image):

        cv_image = cv2.circle(cv_image, (self.pt1_kaku[0] , self.pt1_kaku[1] ), 5, (0, 255, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2_kaku[0] , self.pt2_kaku[1] ), 5, (0, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt3_kaku[0] , self.pt3_kaku[1] ), 5, (255, 0, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt4_kaku[0], self.pt4_kaku[1]), 5, (255, 0, 255), -1)

        cv_image = cv2.line(cv_image,self.pt1_kaku,self.pt2_kaku,(0,255,255),2)
        cv_image = cv2.line(cv_image, self.pt2_kaku, self.pt3_kaku, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt3_kaku, self.pt4_kaku, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt4_kaku, self.pt1_kaku, (0, 255, 255), 2)
        return cv_image
    def area_select_5kaku(self, cv_image):

        cv_image = cv2.circle(cv_image, (self.pt1_polygol5[0] , self.pt1_polygol5[1] ), 5, (0, 255, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt2_polygol5[0] , self.pt2_polygol5[1] ), 5, (0, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt3_polygol5[0] , self.pt3_polygol5[1] ), 5, (255, 0, 0), -1)
        cv_image = cv2.circle(cv_image, (self.pt4_polygol5[0], self.pt4_polygol5[1]), 5, (255, 0, 255), -1)
        cv_image = cv2.circle(cv_image, (self.pt5_polygol5[0], self.pt5_polygol5[1]), 5, (255, 255, 0), -1)

        cv_image = cv2.line(cv_image,self.pt1_polygol5,self.pt2_polygol5,(0,255,255),2)
        cv_image = cv2.line(cv_image, self.pt2_polygol5, self.pt3_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt3_polygol5, self.pt4_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt4_polygol5, self.pt5_polygol5, (0, 255, 255), 2)
        cv_image = cv2.line(cv_image, self.pt5_polygol5, self.pt1_polygol5, (0, 255, 255), 2)
        return cv_image
    def color_map_select(self, cv_image):
        grayscale_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)

        if self.colormap.get() == "":
            return cv_image
        if self.colormap.get() == "NORMAL_NO_FILTER":
            return cv_image
        if self.colormap.get() == "GRAY_SCALE":
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.colormap.get() == "COLORMAP_AUTUMN":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_AUTUMN)
            # cv2.putText(self.cv_image, "COLORMAP_AUTUMN", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_JET":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_JET)
            # cv2.putText(self.cv_image, "COLORMAP_JET", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_WINTER":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_WINTER)
            # cv2.putText(self.cv_image, "COLORMAP_WINTER", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_RAINBOW":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_RAINBOW)
            # cv2.putText(self.cv_image, "COLORMAP_RAINBOW", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_OCEAN":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_OCEAN)
            # cv2.putText(self.cv_image, "COLORMAP_OCEAN", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_SUMMER":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_SUMMER)
            # cv2.putText(self.cv_image, "COLORMAP_SUMMER", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_SPRING":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_SPRING)
            # cv2.putText(self.cv_image, "COLORMAP_SPRING", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_COOL":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_COOL)
            # cv2.putText(self.cv_image, "COLORMAP_COOL", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_HSV":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_HSV)
            # cv2.putText(self.cv_image, "COLORMAP_HSV", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_PINK":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_PINK)
            # cv2.putText(self.cv_image, "COLORMAP_PINK", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        if self.colormap.get() == "COLORMAP_HOT":
            cv_image = cv2.applyColorMap(grayscale_image, cv2.COLORMAP_HOT)
            # cv2.putText(self.cv_image, "COLORMAP_HOT", (20, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 0), 2);
        return cv_image
    def cam_filter(self, cv_image):

        if self.filtermap.get() == 'None':
            cv_image = cv_image
        if self.filtermap.get() == 'Threshold':
            if self.boder_filter_map_sl.get() == 'THRESH_BINARY':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
            if self.boder_filter_map_sl.get() == 'THRESH_BINARY_INV':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY_INV)
            if self.boder_filter_map_sl.get() == 'THRESH_TRUNC':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_TRUNC)
            if self.boder_filter_map_sl.get() == 'THRESH_TOZERO':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_TOZERO)
            if self.boder_filter_map_sl.get() == 'THRESH_TOZERO_INV':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_TOZERO_INV)

        if self.filtermap.get() == 'Adapt_Thre':
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)

            if self.boder_filter_map_sl1.get() == 'THRESH_BINARY':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 51,
                                                 self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_BINARY_INV':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV,
                                                 51, self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_TRUNC':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_TRUNC, 51,
                                                 self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_TOZERO':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_TOZERO, 51,
                                                 self.adaptiveThreshold_var.get())
            if self.boder_filter_map_sl1.get() == 'THRESH_TOZERO_INV':
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_TOZERO_INV,
                                                 51, self.adaptiveThreshold_var.get())

            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.filtermap.get() == 'Canny_1':
            cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.filtermap.get() == 'Canny_2':
            cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
            cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        if self.filtermap.get() == 'Gaussion':
            ksize = int(self.gaussion_var.get())
            ksize = int(ksize / 2) * 2 + 1
            cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
        if self.filtermap.get() == 'Blur':
            ksize = int(self.Blur_var.get())
            cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
        if self.filtermap.get() == 'Bold':
            kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
            cv_image = cv2.erode(cv_image, kernel, iterations=1)
            # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
        if self.filtermap.get() == 'Debold':
            kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
            # cv_image = cv2.erode(cv_image, kernel, iterations=1)
            cv_image = cv2.dilate(cv_image, kernel, iterations=1)
        #########################################################
        if self.so_fiter >= 2:
            if self.filtermap1.get() == 'None':
                cv_image = cv_image
            if self.filtermap1.get() == 'Threshold':
                _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
            if self.filtermap1.get() == 'Adapt_Thre':
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 51,
                                                 self.adaptiveThreshold_var.get())
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
            if self.filtermap1.get() == 'Canny_1':
                cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
            if self.filtermap1.get() == 'Canny_2':
                cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
            if self.filtermap1.get() == 'Gaussion':
                ksize = int(self.gaussion_var.get())
                ksize = int(ksize / 2) * 2 + 1
                cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
            if self.filtermap1.get() == 'Blur':
                ksize = int(self.Blur_var.get())
                cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
            if self.filtermap1.get() == 'Bold':
                kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                cv_image = cv2.erode(cv_image, kernel, iterations=1)
                # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
            if self.filtermap1.get() == 'Debold':
                kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                ################################################
            if self.so_fiter >= 3:
                if self.filtermap2.get() == 'None':
                    cv_image = cv_image
                if self.filtermap2.get() == 'Threshold':
                    _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
                if self.filtermap2.get() == 'Adapt_Thre':
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                    cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
                                                     51,
                                                     self.adaptiveThreshold_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap2.get() == 'Canny_1':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap2.get() == 'Canny_2':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap2.get() == 'Gaussion':
                    ksize = int(self.gaussion_var.get())
                    ksize = int(ksize / 2) * 2 + 1
                    cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
                if self.filtermap2.get() == 'Blur':
                    ksize = int(self.Blur_var.get())
                    cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
                if self.filtermap2.get() == 'Bold':
                    kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                    cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                if self.filtermap2.get() == 'Debold':
                    kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                    # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    cv_image = cv2.dilate(cv_image, kernel, iterations=1)
            ############################################################
            if self.so_fiter >= 4:
                if self.filtermap3.get() == 'None':
                    cv_image = cv_image
                if self.filtermap3.get() == 'Threshold':
                    _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
                if self.filtermap3.get() == 'Adapt_Thre':
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                    cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
                                                     51,
                                                     self.adaptiveThreshold_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap3.get() == 'Canny_1':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap3.get() == 'Canny_2':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap3.get() == 'Gaussion':
                    ksize = int(self.gaussion_var.get())
                    ksize = int(ksize / 2) * 2 + 1
                    cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
                if self.filtermap3.get() == 'Blur':
                    ksize = int(self.Blur_var.get())
                    cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
                if self.filtermap3.get() == 'Bold':
                    kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                    cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                if self.filtermap3.get() == 'Debold':
                    kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                    # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    cv_image = cv2.dilate(cv_image, kernel, iterations=1)
            ####################################
            if self.so_fiter >= 5:
                if self.filtermap4.get() == 'None':
                    cv_image = cv_image
                if self.filtermap4.get() == 'Threshold':
                    _, cv_image = cv2.threshold(cv_image, int(self.scale_var.get()), 255, cv2.THRESH_BINARY)
                if self.filtermap4.get() == 'Adapt_Thre':
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
                    cv_image = cv2.adaptiveThreshold(cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY,
                                                     51,
                                                     self.adaptiveThreshold_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap4.get() == 'Canny_1':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny1_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap4.get() == 'Canny_2':
                    cv_image = cv2.Canny(cv_image, self.canny1_var.get(), self.canny2_var.get())
                    cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
                if self.filtermap4.get() == 'Gaussion':
                    ksize = int(self.gaussion_var.get())
                    ksize = int(ksize / 2) * 2 + 1
                    cv_image = cv2.GaussianBlur(cv_image, (ksize, ksize), 0)
                if self.filtermap4.get() == 'Blur':
                    ksize = int(self.Blur_var.get())
                    cv_image = cv2.blur(cv_image, (ksize, ksize), cv2.BORDER_ISOLATED)
                if self.filtermap4.get() == 'Bold':
                    kernel = np.ones((self.bold_var.get(), self.debold_var.get()), np.uint8)
                    cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    # cv_image = cv2.dilate(cv_image, kernel, iterations=1)
                if self.filtermap4.get() == 'Debold':
                    kernel = np.ones((self.bold_var1.get(), self.debold_var1.get()), np.uint8)
                    # cv_image = cv2.erode(cv_image, kernel, iterations=1)
                    cv_image = cv2.dilate(cv_image, kernel, iterations=1)

        return cv_image
    def cam_circle_elipse_detect(self, cv_image):
        params = cv2.SimpleBlobDetector_Params()

        # Set Area filtering parameters
        if self.fil_by_Area.get() == 1:
            params.filterByArea = True
            params.minArea = self.minArea_var.get()
        if self.filByCircularity.get() == 1:
            # Set Circularity filtering parameters
            params.filterByCircularity = True
            params.minCircularity = self.minCircularity_var.get()
        if self.filByConvexity.get() == 1:
            # Set Convexity filtering parameters
            params.filterByConvexity = True
            params.minConvexity = self.minConvexity_var.get()
        if self.filByInertia.get() == 1:
            # Set inertia filtering parameters
            params.filterByInertia = True
            params.minInertiaRatio = self.minInertia_var.get()

        # Create a detector with the parameters
        detector = cv2.SimpleBlobDetector_create(params)

        # Detect blobs
        keypoints = detector.detect(cv_image)

        # Draw blobs on our image as red circles
        blank = np.zeros((1, 1))
        blobs = cv2.drawKeypoints(cv_image, keypoints, blank, (0, 0, 255),
                                  cv2.DRAW_MATCHES_FLAGS_DRAW_RICH_KEYPOINTS)

        number_of_blobs = len(keypoints)
        text = "Number of Circular Blobs: " + str(len(keypoints))
        cv2.putText(blobs, text, (20, 550),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 100, 255), 2)
        # print(text)
        return blobs
    def circle_detect(self, cv_image):
        # Convert to grayscale.
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)

        # Blur using 3 * 3 kernel.
        gray_blurred = cv2.blur(gray, (self.Kernel_var.get(), self.Kernel_var.get()))

        # Apply Hough transform on the blurred image.
        detected_circles = cv2.HoughCircles(gray_blurred,
                                            cv2.HOUGH_GRADIENT, 1, 20, param1=self.param1_var.get(),
                                            param2=self.param2_var.get(), minRadius=self.minRadius_var.get(),
                                            maxRadius=self.maxRadius_var.get())

        # Draw circles that are detected.
        if detected_circles is not None:
            self.Camera_check = 1
            # Convert the circle parameters a, b and r to integers.
            detected_circles = np.uint16(np.around(detected_circles))

            for pt in detected_circles[0, :]:
                a, b, r = pt[0], pt[1], pt[2]

                # Draw the circumference of the circle.
                cv2.circle(cv_image, (a, b), r, (0, 255, 0), 2)

                # Draw a small circle (of radius 1) to show the center.
                cv2.circle(cv_image, (a, b), 1, (0, 0, 255), 3)
        else:
            self.Camera_check = 0
        return cv_image
    def rectange_detect(self, cv_image):
        # converting image into grayscale image
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        # setting threshold of gray image

        if self.border_sl.get() == 'THRESH_BINARY':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_BINARY)
        if self.border_sl.get() == 'THRESH_BINARY_INV':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_BINARY_INV)
        if self.border_sl.get() == 'THRESH_TRUNC':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_TRUNC)
        if self.border_sl.get() == 'THRESH_TOZERO':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_TOZERO)
        if self.border_sl.get() == 'THRESH_TOZERO_INV':
            ret, threshold = cv2.threshold(gray, self.rect_thres.get(), 255, cv2.THRESH_TOZERO_INV)

        # _, threshold = cv2.threshold(gray, self.rect_thres, 255, cv2.THRESH_BINARY)

        # using a findContours() function
        if self.contour_sl.get() == 'RETR_TREE':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl.get() == 'RETR_CCOMP':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl.get() == 'RETR_LIST':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl.get() == 'RETR_EXTERNAL':
            if self.contour_sl1.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl1.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_TC89_L1)

        i = 0

        # list for storing names of shapes
        for contour in contours:

            # here we are ignoring first counter because
            # findcontour function detects whole image as shape
            if i == 0:
                i = 1
                continue

            # cv2.approxPloyDP() function to approximate the shape
            approx = cv2.approxPolyDP(
                contour, self.rect_para1.get() * cv2.arcLength(contour, True), True)

            # using drawContours() function
            # cv2.drawContours(cv_image, [contour], 0, (0, 0, 255), 3)
            area = cv2.contourArea(contour)
            chuvi = cv2.arcLength(contour, True)
            # finding center point of shape
            M = cv2.moments(contour)
            if M['m00'] != 0.0:
                x = int(M['m10'] / M['m00'])
                y = int(M['m01'] / M['m00'])

                # putting shape name at center of each shape
            '''if len(approx) == 3:
                cv2.putText(cv_image, 'Triangle', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''

            if len(approx) == 4 and area >= self.area_rect.get() and chuvi >= self.chuvi_rect.get():
                x, y, w, h = cv2.boundingRect(approx)
                cv2.rectangle(cv_image, (x, y), (x + w, y + h), (0, 255, 0), 2)
                # print('Rectangle Detect')
                self.Camera_check = 1
                # cv2.putText(cv_image, 'Quadrilateral', (x, y),
                # cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
            else:
                self.Camera_check = 0
            '''elif len(approx) == 5:
                cv2.putText(cv_image, 'Pentagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            elif len(approx) == 6:
                cv2.putText(cv_image, 'Hexagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            else:
                print('co cai eo')
                #cv2.putText(cv_image, 'circle', (x, y),
                            #cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''
        return cv_image
    def triangle_detect(self, cv_image):
        self.Camera_check = 0
        # converting image into grayscale image
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        # setting threshold of gray image

        if self.border_sl2.get() == 'THRESH_BINARY':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_BINARY)
        if self.border_sl2.get() == 'THRESH_BINARY_INV':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_BINARY_INV)
        if self.border_sl2.get() == 'THRESH_TRUNC':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_TRUNC)
        if self.border_sl2.get() == 'THRESH_TOZERO':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_TOZERO)
        if self.border_sl2.get() == 'THRESH_TOZERO_INV':
            ret, threshold = cv2.threshold(gray, self.trian_thres.get(), 255, cv2.THRESH_TOZERO_INV)

        # _, threshold = cv2.threshold(gray, self.rect_thres, 255, cv2.THRESH_BINARY)

        # using a findContours() function
        if self.contour_sl2.get() == 'RETR_TREE':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl2.get() == 'RETR_CCOMP':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl2.get() == 'RETR_LIST':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_LIST, cv2.CHAIN_APPROX_TC89_L1)

        if self.contour_sl2.get() == 'RETR_EXTERNAL':
            if self.contour_sl21.get() == 'CHAIN_APPROX_NONE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_SIMPLE':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if self.contour_sl21.get() == 'CHAIN_APPROX_TC89_L1':
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_TC89_L1)

        i = 0

        # list for storing names of shapes
        for contour in contours:
            if contour is None:
                self.Camera_check = 0
                return cv_image
            # here we are ignoring first counter because
            # findcontour function detects whole image as shape
            if i == 0:
                i = 1
                continue

            # cv2.approxPloyDP() function to approximate the shape
            approx = cv2.approxPolyDP(
                contour, self.rect_para1.get() * cv2.arcLength(contour, True), True)

            # using drawContours() function
            # cv2.drawContours(cv_image, [contour], 0, (0, 0, 255), 3)
            area = cv2.contourArea(contour)
            chuvi = cv2.arcLength(contour, True)
            # finding center point of shape
            M = cv2.moments(contour)
            if M['m00'] != 0.0:
                x = int(M['m10'] / M['m00'])
                y = int(M['m01'] / M['m00'])

                # putting shape name at center of each shape
            '''if len(approx) == 3:
                cv2.putText(cv_image, 'Triangle', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''

            if len(approx) == 3 and area >= self.area_trian.get() and chuvi >= self.chuvi_trian.get():
                x, y, w, h = cv2.boundingRect(approx)
                cv2.rectangle(cv_image, (x, y), (x + w, y + h), (0, 255, 0), 2)
                # print('TriAngle Detect')
                self.Camera_check = 1
                # cv2.putText(cv_image, 'Quadrilateral', (x, y),
                # cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
                return cv_image
            '''elif len(approx) == 5:
                cv2.putText(cv_image, 'Pentagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            elif len(approx) == 6:
                cv2.putText(cv_image, 'Hexagon', (x, y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            else:
                print('co cai eo')
                #cv2.putText(cv_image, 'circle', (x, y),
                            #cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)'''
        return cv_image
    def draw_contour_object(self, cv_image):
        self.Camera_check = 0
        gray = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        # setting threshold of gray image
        self.now_area = 0
        self.now_chuvi = 0
        ret, threshold = cv2.threshold(gray, self.Threshole_contua_var.get(), 255, cv2.THRESH_BINARY)
        contours, _ = cv2.findContours(threshold, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

        i = 0

        # list for storing names of shapes
        for contour in contours:
            if contour is None:
                self.Camera_check = 0
                return cv_image
            # here we are ignoring first counter because
            # findcontour function detects whole image as shape
            if i == 0:
                i = 1
                continue
            area = cv2.contourArea(contour)
            chuvi = cv2.arcLength(contour, True)
            # cv2.approxPloyDP() function to approximate the shape
            approx = cv2.approxPolyDP(
                contour, self.contua_aprrox.get() * cv2.arcLength(contour, True), True)
            if chuvi >= self.chuvi_contua.get() and area > self.dientich_contua.get():
                # using drawContours() function
                self.Camera_check = 1
                self.now_area= cv2.contourArea(contour)
                self.now_chuvi = cv2.arcLength(contour, True)
                self.now_area_check.set(str(self.now_area))
                cv2.drawContours(cv_image, [contour], 0, (255, 0, 0), 2)
                return cv_image
        self.now_area_check.set(str(self.now_area))
        return cv_image
    def edge_detect(self, image):
        self.Camera_check = 0
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Use canny edge detection
        edges = cv2.Canny(gray, self.var_canny_a.get(), self.var_canny_b.get(), apertureSize=3)
        # Apply HoughLinesP method to
        # to directly obtain line end points
        lines_list = []
        lines = cv2.HoughLinesP(
            edges,  # Input edge image
            self.len_size1.get(),  # Distance resolution in pixels
            np.pi / 180,  # Angle resolution in radians
            self.var_thr.get(),  # Min number of votes for valid line
            minLineLength=self.len_size2.get(),  # Min allowed length of line
            maxLineGap=self.len_size3.get()  # Max allowed gap between line for joining them
        )

        # Iterate over points
        if lines is None:
            self.Camera_check = 0
            return image
        # print(lines)
        self.Camera_check = 1
        for points in lines:
            # Extracted points nested in the list
            x1, y1, x2, y2 = points[0]
            # Draw the lines joing the points
            # On the original image
            cv2.line(image, (x1, y1), (x2, y2), (0, 0, 255), 3, cv2.LINE_8)
            # Maintain a simples lookup list for points
            lines_list.append([(x1, y1), (x2, y2)])

        return image
    def connected_component(self, cv_image):
        # 画像表示用に入力画像をカラーデータに変換する
        image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2GRAY)
        cv_image = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
        if self.border_connect_sl.get() == 'THRESH_BINARY':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_BINARY)
        if self.border_connect_sl.get() == 'THRESH_BINARY_INV':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_BINARY_INV)
        if self.border_connect_sl.get() == 'THRESH_TRUNC':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_TRUNC)
        if self.border_connect_sl.get() == 'THRESH_TOZERO':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_TOZERO)
        if self.border_connect_sl.get() == 'THRESH_TOZERO_INV':
            retval, image = cv2.threshold(image, self.thres_connect.get(), 255, cv2.THRESH_TOZERO_INV)
        # 二値化(二値化して白黒反転)
        # retval, cv_image = cv2.threshold(cv_image, self.thres_connect.get(), 255, cv2.THRESH_BINARY_INV)

        # クロージング処理（細切れ状態を防ぐため）
        kernel = np.ones((self.kernel_connect.get(), self.kernel_connect.get()), np.uint8)
        image = cv2.morphologyEx(image, cv2.MORPH_CLOSE, kernel, iterations=self.iterations_var.get())

        # ラベリング
        retval, labels, stats, centroids = cv2.connectedComponentsWithStats(image)

        # 結果表示
        for i in range(1, retval):
            x, y, width, height, area = stats[i]  # x座標, y座標, 幅, 高さ, 面積

            if area > self.area_detect_connect.get():  # 面積が１０画素以上の部分
                cv2.rectangle(cv_image, (x, y), (x + width, y + height), (0, 0, 255), 2)
                cv2.putText(cv_image, f"[{i}]:{area}", (x, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (255, 0, 0), 1,
                            cv2.LINE_AA)
                # print(i, area)
        # cv_image = cv2.cvtColor(cv_image, cv2.COLOR_GRAY2RGB)
        return cv_image
    def rectangge_component(self, img):
        self.Camera_check = 0
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # dientich = self.var_Height_tab3.get() * self.var_Width_tab3.get()
        fill_change = self.fill_change_var.get() * 0.01

        # 白黒反転して二値化
        ret, img = cv2.threshold(img, self.thres_patan.get(), 255, cv2.THRESH_BINARY_INV)

        # 一番外側の輪郭のみを取得
        contours, hierarchy = cv2.findContours(img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # 画像表示用に入力画像をカラーデータに変換する
        img_disp = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

        # 全ての輪郭を描画
        cv2.drawContours(img_disp, contours, -1, (255, 0, 255), 1)

        # 輪郭の点の描画
        for i, contour in enumerate(contours):  ### for contour in contours:

            area = cv2.contourArea(contours[i])
            if area > self.area_patan.get():
                self.have_contour = 1
            else:
                self.have_contour = 0
                self.Camera_check = 0

                return img_disp
            if self.have_contour == 1:
                # print(area)
                # 傾いた外接する矩形領域の描画
                rect = cv2.minAreaRect(contour)
                box = cv2.boxPoints(rect)
                box = np.intp(box)
                cv2.drawContours(img_disp, [box], 0, (0, 255, 255), 1)
                # 矩形度の計算
                val = self.rectangularity(contour)
                # 輪郭の矩形領域
                x, y, w, h = cv2.boundingRect(contour)
                # 矩形度の描画
                if val is not None:

                    cv2.putText(img_disp, f"{val:.3f}", (x, y + 15), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0), 1,
                                cv2.LINE_AA)
                    # 円らしい領域（円形度が0.85以上）を囲う
                    if val > fill_change + fill_change * 0.1:
                        self.Camera_check = 1
                        cv2.rectangle(img_disp, (x - 0, y - 0), (x + w + 0, y + h + 0), (0, 255, 0), 2)  # 少し外側を囲う
                    if val >= fill_change - fill_change * 0.1 and val <= fill_change + fill_change * 0.1:
                        self.Camera_check = 1
                        cv2.rectangle(img_disp, (x - 0, y - 0), (x + w + 0, y + h + 0), (255, 255, 0), 2)  # 少し外側を囲う
                    if val < fill_change - fill_change * 0.1:
                        self.Camera_check = 0
                        cv2.rectangle(img_disp, (x - 0, y - 0), (x + w + 0, y + h + 0), (255, 0, 0), 2)  # 少し外側を囲う

        return img_disp
    def rectangularity(self, contour):
        '''
        矩形度を求める

        Parameters
        ----------
        contour : ndarray
            輪郭の(x,y)座標の配列

        Returns
        -------
            矩形度

        '''
        # 面積
        area = cv2.contourArea(contour)
        # 傾いた外接する矩形領域
        _, (width, height), _ = cv2.minAreaRect(contour)

        # 矩形度を返す
        if width == 0 or height == 0:
            return

        return area / width / height
    def hakotori_kakunin(self):
        '''モードレスダイアログボックスの作成'''
        self.hakotori_dlg = tk.Toplevel(self)
        self.hakotori_dlg.title("箱取り確認")  # ウィンドウタイトル
        self.hakotori_dlg.geometry("700x200")  # ウィンドウサイズ(幅x高さ)
        x = root.winfo_x()
        y = root.winfo_y()
        self.hakotori_dlg.geometry("+%d+%d" %(x+220,y+250))


        self.hakotori_ok = tk.Button(self.hakotori_dlg, font=("Arial", 15), wraplength=150, height=3, width=20, text="箱取り完了", command=self.hakotori_ok_clicked)
        self.hakotori_ok.place(x=50, y=70)
        self.hakomodosi_ok = tk.Button(self.hakotori_dlg, font=("Arial", 15), wraplength=150, height=3, width=20,text="箱入れ直し完了", command=self.close_hakotori)
        self.hakomodosi_ok.place(x=400, y=70)

        if self.hako_tori_kakunin == 0:
            self.lbl_hakotori = Label(self.hakotori_dlg, font=("Arial", 15),
                                      text="落とした空箱を一旦回収し、完了を押してください。")
            self.lbl_hakotori.place(x=0, y=0)
            self.hakotori_ok["state"] = "normal"#normal
            self.hakomodosi_ok["state"] = "disable"#disable
        #time.sleep(1)
        self.hakotori_dlg.grab_set()  # モーダルにする
        self.hakotori_dlg.focus_set()  # フォーカスを新しいウィンドウをへ移す
        self.hakotori_dlg.transient(self.master)  # タスクバーに表示しない
        app.wait_window(self.hakotori_dlg)
    def hakotori_ok_clicked(self):
        self.Socket_send("ST 4114")
        self.hako_tori_kakunin = 1
        self.lbl_hakotori = Label(self.hakotori_dlg, font=("Arial", 15),
                                  text="取り除いた箱の向きを合わせて入れ直し、入れ直し完了を押してください。")
        self.lbl_hakotori.place(x=0, y=0)
        self.hakotori_ok["state"] = "disable"
        self.hakomodosi_ok["state"] = "normal"
    ###########################DIALOG####################
    def close_hakotori(self):
        self.Socket_send("ST R4113")
        self.hako_tori_kakunin = 0
        self.hakotori_dlg.destroy()
    def create_anso_dialog_off(self):
        '''モードレスダイアログボックスの作成'''
        self.dlg_modeless = tk.Toplevel(self)
        self.dlg_modeless.title("リール表裏直し")  # ウィンドウタイトル
        self.dlg_modeless.geometry("600x200")  # ウィンドウサイズ(幅x高さ)
        #back_color
        x = root.winfo_x()
        y = root.winfo_y()
        self.dlg_modeless.geometry("+%d+%d" %(x+220,y+250))

        self.lbl_lock = Label(self.dlg_modeless, font=("Arial", 20), text="リールの表裏を修正し、OKを押してください。")
        self.lbl_lock.place(x=0, y=0)
        self.btnkakunin_ok = tk.Button(self.dlg_modeless, font=("Arial", 15), wraplength = 150,height=3, width=20,  text="リール表裏修正OK", command=self.close_messenger )
        self.btnkakunin_ok.place(x=190, y=70)
        self.skip_bt = tk.Button(self.dlg_modeless, font=("Arial", 10), wraplength=150, height=2, width=8,
                                       text="スキップ", command=self.skip_messenger)
        self.skip_bt.place(x=520, y=150)
        #time.sleep(1)
        self.dlg_modeless.grab_set()  # モーダルにする
        self.dlg_modeless.focus_set()  # フォーカスを新しいウィンドウをへ移す
        self.dlg_modeless.transient(self.master)  # タスクバーに表示しない

        # ダイアログが閉じられるまで待つ
        app.wait_window(self.dlg_modeless)
    def close_messenger(self):
        self.Socket_send("ST R4115")
        self.dlg_modeless.destroy()
    def skip_messenger(self):
        self.Socket_send("ST R4115")
        self.Socket_send("ST MR3502")
        self.dlg_modeless.destroy()
    def create_anso_dialog_on(self):
        '''モーダルダイアログボックスの作成'''
        self.dlg_modal = tk.Toplevel(self)
        self.dlg_modal.title("パスコードロック")  # ウィンドウタイトル
        self.dlg_modal.geometry("400x300")  # ウィンドウサイズ(幅x高さ)
        x = root.winfo_x()
        y = root.winfo_y()
        self.dlg_modal.geometry("+%d+%d" % (x + 450, y + 200))

        #####
        lbl_lock = Label(self.dlg_modal, font=("Arial", 15), text="パスコードを入力してください")
        lbl_lock.place(x=60, y=10)
        self.password = tk.StringVar()
        self.password.set("")
        self.txt_password = tk.Entry(self.dlg_modal, justify=tk.RIGHT, width=15, font=("Arial", 20), textvariable=self.password)
        self.txt_password.place(x=90, y=50)
        self.btn_modal = tk.Button(self.dlg_modal, font=("Arial", 15), text="ログイン", command=self.passcheck)
        self.btn_modal.place(x=160, y=100)
        self.lock_stt = tk.StringVar(value="")
        self.lbl_lock_stt = Label(self.dlg_modal, font=("Arial", 15), textvariable=self.lock_stt)
        self.lbl_lock_stt.place(x=95, y=170)
        # モーダルにする設定
        self.dlg_modal.grab_set()  # モーダルにする
        self.dlg_modal.focus_set()  # フォーカスを新しいウィンドウをへ移す
        self.dlg_modal.transient(self.master)  # タスクバーに表示しない

        # ダイアログが閉じられるまで待つ
        app.wait_window(self.dlg_modal)
        print("ダイアログが閉じられた")
    def passcheck(self):
        if self.password.get() != "1122":
            self.reset_data_cam["state"] = "disable"
            self.cam_edit_hani = False;
            #self.vitri_test["state"] = "disable"


            self.lock_stt.set("パスコードが違います。")
            for child in self.labelcamset.winfo_children():
                child.configure(state='disable')
            for child in self.labelfilter.winfo_children():
                child.configure(state='disable')
            for child in self.labelfilter1.winfo_children():
                child.configure(state='disable')
            for child in self.labelfilter2.winfo_children():
                child.configure(state='disable')
            #
            for child in self.labeldetect.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect1.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect2.winfo_children():
                child.configure(state='disable')
            for child in self.labeselectdetect.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect3.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect4.winfo_children():
                child.configure(state='disable')
            for child in self.labeldetect5.winfo_children():
                child.configure(state='disable')
            for child in self.objectdetect.winfo_children():
                child.configure(state='disable')
            for child in self.objectdetect1.winfo_children():
                child.configure(state='disable')
                self.setting_lock = 1
            return
        if self.password.get() == "1122":
            self.lock_stt.set("Ok")
            self.cam_edit_hani = True;
            self.reset_data_cam["state"] = "normal"

            for child in self.labelcamset.winfo_children():
                child.configure(state='normal')
            for child in self.labelfilter.winfo_children():
                child.configure(state='normal')
            for child in self.labelfilter1.winfo_children():
                child.configure(state='normal')
            for child in self.labelfilter2.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect1.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect2.winfo_children():
                child.configure(state='normal')
            for child in self.labeselectdetect.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect3.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect4.winfo_children():
                child.configure(state='normal')
            for child in self.labeldetect5.winfo_children():
                child.configure(state='normal')
            for child in self.objectdetect.winfo_children():
                child.configure(state='normal')
            for child in self.objectdetect1.winfo_children():
                child.configure(state='normal')
            self.setting_lock = 0
            self.dlg_modal.destroy()
    def enable(childList):
        for child in childList:
            child.configure(state='enable')
    def ending_program(self):
        self.Socket_send("ST R4015")
        sys.exit()
        root.destroy()
    def end_thread(self):
        sys.exit()
        root.destroy()
    ###########################SAVE DATA###############
    def save_data(self):
        # colum1
        self.write_data(1, 1, self.IP_PLC.get())
        self.write_data(2, 1, self.PORT_PLC.get())
        self.write_data(3, 1, self.port_var.get())
        self.write_data(4, 1, self.baudrate_var.get())
        self.write_data(5, 1, self.delimiter_var.get())
        self.write_data(6, 1, self.port_var_msp.get())
        self.write_data(7, 1, self.baudrate_var_msp.get())
        self.write_data(8, 1, self.delimiter_var_msp.get())
        # colum2
        self.write_data(1, 2, self.FRAME_WIDTH_var.get())
        self.write_data(2, 2, self.FRAME_HEIGHT_var.get())
        self.write_data(3, 2, self.FPS_var.get())
        self.write_data(4, 2, self.BRIGHTNESS_var.get())
        self.write_data(5, 2, self.CONTRAST_var.get())
        self.write_data(6, 2, self.SATURATION_var.get())
        self.write_data(7, 2, self.HUE_var.get())
        self.write_data(8, 2, self.GAIN_var.get())
        self.write_data(9, 2, self.check_hani_ok.get())
        self.write_data(10, 2, self.colorchoosen.current())

        self.write_data(11, 2, self.EXPOSE_var.get())
        self.write_data(12, 2, self.auto_exposure_var.get())
        self.write_data(13, 2, self.auto_focus_var.get())

        # colum3
        self.write_data(1, 3, self.boder_filter_map.current())
        self.write_data(2, 3, self.boder_filter_map1.current())
        self.write_data(3, 3, self.scale_var.get())
        self.write_data(4, 3, self.adaptiveThreshold_var.get())
        self.write_data(5, 3, self.canny1_var.get())
        self.write_data(6, 3, self.canny2_var.get())
        self.write_data(7, 3, self.gaussion_var.get())
        self.write_data(8, 3, self.Blur_var.get())
        self.write_data(9, 3, self.bold_var.get())
        self.write_data(10, 3, self.debold_var.get())
        self.write_data(11, 3, self.bold_var1.get())
        self.write_data(12, 3, self.debold_var1.get())
        # colum4
        self.write_data(1, 4, self.so_fiter)
        self.write_data(2, 4, self.filterchoosen.current())
        self.write_data(3, 4, self.filterchoosen1.current())
        self.write_data(4, 4, self.filterchoosen2.current())
        self.write_data(5, 4, self.filterchoosen3.current())
        self.write_data(6, 4, self.filterchoosen4.current())
        self.write_data(7, 4, self.pt1[0])
        self.write_data(8, 4, self.pt1[1])
        self.write_data(9, 4, self.pt2[0])
        self.write_data(10, 4, self.pt2[1])
        #
        self.write_data(1, 14, self.pt1_specal[0])
        self.write_data(2, 14, self.pt1_specal[1])
        self.write_data(3, 14, self.pt2_specal[0])
        self.write_data(4, 14, self.pt2_specal[1])
        self.write_data(5, 14, self.pt3_specal[0])
        self.write_data(6, 14, self.pt3_specal[1])
        #
        self.write_data(7, 14, self.pt1_kaku[0])
        self.write_data(8, 14, self.pt1_kaku[1])
        self.write_data(9, 14, self.pt2_kaku[0])
        self.write_data(10, 14, self.pt2_kaku[1])
        self.write_data(11, 14, self.pt3_kaku[0])
        self.write_data(12, 14, self.pt3_kaku[1])
        self.write_data(13, 14, self.pt4_kaku[0])
        self.write_data(14, 14, self.pt4_kaku[1])
        #
        self.write_data(15, 14, self.pt1_polygol5[0])
        self.write_data(16, 14, self.pt1_polygol5[1])
        self.write_data(17, 14, self.pt2_polygol5[0])
        self.write_data(18, 14, self.pt2_polygol5[1])
        self.write_data(19, 14, self.pt3_polygol5[0])
        self.write_data(20, 14, self.pt3_polygol5[1])
        self.write_data(21, 14, self.pt4_polygol5[0])
        self.write_data(22, 14, self.pt4_polygol5[1])
        self.write_data(23, 14, self.pt5_polygol5[0])
        self.write_data(24, 14, self.pt5_polygol5[1])

        # colum5

        self.write_data(1, 5, self.fil_by_Area.get())
        self.write_data(2, 5, self.filByCircularity.get())
        self.write_data(3, 5, self.filByConvexity.get())
        self.write_data(4, 5, self.filByInertia.get())
        self.write_data(5, 5, self.minArea_var.get())
        self.write_data(6, 5, self.minCircularity_var.get())
        self.write_data(7, 5, self.minConvexity_var.get())
        self.write_data(8, 5, self.minInertia_var.get())
        # colum6
        self.write_data(1, 6, self.Kernel_var.get())
        self.write_data(2, 6, self.param1_var.get())
        self.write_data(3, 6, self.param2_var.get())
        self.write_data(4, 6, self.minRadius_var.get())
        self.write_data(5, 6, self.maxRadius_var.get())
        self.write_data(6, 6, self.var_canny_a.get())
        self.write_data(7, 6, self.var_canny_b.get())
        self.write_data(8, 6, self.var_thr.get())
        self.write_data(9, 6, self.len_size2.get())
        self.write_data(10, 6, self.len_size3.get())
        self.write_data(11, 6, self.len_size1.get())
        # colum7

        self.write_data(1, 7, self.circle_select.get())
        self.write_data(2, 7, self.circle_select_adv.get())
        self.write_data(3, 7, self.triang_select.get())
        self.write_data(4, 7, self.rect_select.get())
        self.write_data(5, 7, self.edge_select.get())
        self.write_data(6, 7, self.draw_contour_select.get())
        self.write_data(7, 7, self.connected_select.get())
        self.write_data(8, 7, self.rectangge_component_select.get())

        # colum8
        self.write_data(1, 8, self.boder_map.current())
        self.write_data(2, 8, self.contua_map.current())
        self.write_data(3, 8, self.contua_map1.current())
        self.write_data(4, 8, self.rect_para1.get())
        self.write_data(5, 8, self.area_rect.get())
        self.write_data(6, 8, self.chuvi_rect.get())
        self.write_data(7, 8, self.rect_thres.get())
        # colum9
        self.write_data(1, 9, self.boder_map2.current())
        self.write_data(2, 9, self.contua_map2.current())
        self.write_data(3, 9, self.contua_map21.current())
        self.write_data(4, 9, self.trian_para1.get())
        self.write_data(5, 9, self.area_trian.get())
        self.write_data(6, 9, self.chuvi_trian.get())
        self.write_data(7, 9, self.trian_thres.get())
        ##colum 10

        self.write_data(1, 10, self.Threshole_contua_var.get())
        self.write_data(2, 10, self.contua_aprrox.get())
        self.write_data(3, 10, self.chuvi_contua.get())
        self.write_data(4, 10, self.dientich_contua.get())
        # colum 11
        self.write_data(1, 11, self.border_connect.current())
        self.write_data(2, 11, self.kernel_connect.get())
        self.write_data(3, 11, self.iterations_var.get())
        self.write_data(4, 11, self.area_detect_connect.get())
        self.write_data(5, 11, self.thres_connect.get())
        # colum 12
        self.write_data(1, 12, self.fill_change_var.get())
        self.write_data(2, 12, self.area_patan.get())
        self.write_data(3, 12, self.chuvi_patan.get())
        self.write_data(4, 12, self.thres_patan.get())

        self.write_data(1, 15, self.check_patan1.get())
        self.write_data(2, 15, self.check_patan2.get())
        self.write_data(3, 15, self.check_patan3.get())
        self.write_data(4, 15, self.check_patan4.get())
        self.write_data(5, 15, self.check_patan5.get())
        self.write_data(6, 15, self.check_patan6.get())
        # self.write_data(2, 2, self.AEC_Metering_Mode.current())
    def read_data(self, row, colum):
        cell = self.ws.cell(row=row, column=colum).value
        return cell
    def write_data(self, row, colum, value):
        self.ws.cell(row=row, column=colum, value=value)  #
        # 保存
        self.workbook.save('data.xlsx')
    def auto_load(self):
        self.excel_path = 'data.xlsx'
        self.workbook = load_workbook(filename=self.excel_path, read_only=False)
        self.ws = self.workbook.active
        # シートのロード
        self.sheet = self.workbook['Sheet1']
        # print(self.read_data(1,1))
        ###  LOAD DATA  TAB 1 ####################
        # self.write_data(1, 1, self.IP_PLC.get())
        self.IP_PLC.set(self.read_data(1, 1))
        # self.write_data(2, 1, self.PORT_PLC.get())
        self.PORT_PLC.set(self.read_data(2, 1))
        # self.write_data(3, 1, self.port_var.get())
        self.port_var.set(self.read_data(3, 1))
        # self.write_data(4, 1, self.baudrate_var.get())
        self.baudrate_var.set(self.read_data(4, 1))
        # self.write_data(5, 1, self.delimiter_var.get())
        self.delimiter_var.set(self.read_data(5, 1))
        self.port_var_msp.set(self.read_data(6, 1))
        # self.write_data(4, 1, self.baudrate_var.get())
        self.baudrate_var_msp.set(self.read_data(7, 1))
        # self.write_data(5, 1, self.delimiter_var.get())
        self.delimiter_var_msp.set(self.read_data(8, 1))

        # colum2
        # self.write_data(1, 2, self.FRAME_WIDTH_var.get())
        self.FRAME_WIDTH_var.set(self.read_data(1, 2))
        # self.write_data(2, 2, self.FRAME_HEIGHT_var.get())
        self.FRAME_HEIGHT_var.set(self.read_data(2, 2))
        # self.write_data(3, 2, self.FPS_var.get())
        self.FPS_var.set(self.read_data(3, 2))
        # self.write_data(4, 2, self.BRIGHTNESS_var.get())
        self.BRIGHTNESS_var.set(self.read_data(4, 2))
        # self.write_data(5, 2, self.CONTRAST_var.get())
        self.CONTRAST_var.set(self.read_data(5, 2))
        # self.write_data(6, 2, self.SATURATION_var.get())
        self.SATURATION_var.set(self.read_data(6, 2))
        # self.write_data(7, 2, self.HUE_var.get())
        self.HUE_var.set(self.read_data(7, 2))
        # self.write_data(8, 2, self.GAIN_var.get())
        self.GAIN_var.set(self.read_data(8, 2))
        # self.write_data(9, 2, self.check_hani_ok.get())
        self.check_hani_ok.set(self.read_data(9, 2))
        # self.write_data(10, 2, self.colorchoosen.current())
        self.colorchoosen.current(self.read_data(10, 2))
        self.EXPOSE_var.set(self.read_data(11, 2))
        self.auto_exposure_var.set(self.read_data(12, 2))
        self.auto_focus_var.set(self.read_data(13, 2))

        # colum3
        # self.write_data(1, 3, self.boder_filter_map.current())
        self.boder_filter_map.current(self.read_data(1, 3))
        # self.write_data(2, 3, self.boder_filter_map1.current())
        self.boder_filter_map1.current(self.read_data(2, 3))
        # self.write_data(3, 3, self.scale_var.get())
        self.scale_var.set(self.read_data(3, 3))
        # self.write_data(4, 3, self.adaptiveThreshold_var.get())
        self.adaptiveThreshold_var.set(self.read_data(4, 3))
        # self.write_data(5, 3, self.canny1_var.get())
        self.canny1_var.set(self.read_data(5, 3))
        # self.write_data(6, 3, self.canny2_var.get())
        self.canny2_var.set(self.read_data(6, 3))
        # self.write_data(7, 3, self.gaussion_var.get())
        self.gaussion_var.set(self.read_data(7, 3))
        # self.write_data(8, 3, self.Blur_var.get())
        self.Blur_var.set(self.read_data(8, 3))
        # self.write_data(9, 3, self.bold_var.get())
        self.bold_var.set(self.read_data(9, 3))
        # self.write_data(10, 3, self.debold_var.get())
        self.debold_var.set(self.read_data(10, 3))
        # self.write_data(11, 3, self.bold_var1.get())
        self.bold_var1.set(self.read_data(11, 3))
        # self.write_data(12, 3, self.debold_var1.get())
        self.debold_var1.set(self.read_data(12, 3))
        # colum4
        # self.write_data(1, 4, self.so_fiter)
        self.so_fiter = self.read_data(1, 4)
        # self.write_data(2, 4, self.filterchoosen.current())
        self.filterchoosen.current(self.read_data(2, 4))
        # self.write_data(3, 4, self.filterchoosen1.current())
        self.filterchoosen1.current(self.read_data(3, 4))
        # self.write_data(4, 4, self.filterchoosen2.current())
        self.filterchoosen2.current(self.read_data(4, 4))
        # self.write_data(5, 4, self.filterchoosen3.current())
        self.filterchoosen3.current(self.read_data(5, 4))
        # self.write_data(6, 4, self.filterchoosen4.current())
        self.filterchoosen4.current(self.read_data(6, 4))
        # self.pt1[0], self.pt1[1], self.pt2[0], self.pt2[1]
        self.pt1 = [self.read_data(7, 4), self.read_data(8, 4)]
        # self.pt1[1] = self.read_data(8, 4)
        self.pt2 = [self.read_data(9, 4), self.read_data(10, 4)]
        # self.pt2[1] = self.read_data(10, 4)

        # self..set(self.read_data(7, 3))
        # colum5

        # self.write_data(1, 5, self.fil_by_Area.get())
        self.fil_by_Area.set(self.read_data(1, 5))
        # self.write_data(2, 5, self.filByCircularity.get())
        self.filByCircularity.set(self.read_data(2, 5))
        # self.write_data(3, 5, self.filByConvexity.get())
        self.filByConvexity.set(self.read_data(3, 5))
        # self.write_data(4, 5, self.filByInertia.get())
        self.filByInertia.set(self.read_data(4, 5))
        # self.write_data(5, 5, self.minArea_var.get())
        self.minArea_var.set(self.read_data(5, 5))
        # self.write_data(6, 5, self.minCircularity_var.get())
        self.minCircularity_var.set(self.read_data(6, 5))
        # self.write_data(7, 5, self.minConvexity_var.get())
        self.minConvexity_var.set(self.read_data(7, 5))
        # self.write_data(8, 5, self.minInertia_var.get())
        self.minInertia_var.set(self.read_data(7, 5))

        # # colum6
        # self.write_data(1, 6, self.Kernel_var.get())
        self.Kernel_var.set(self.read_data(1, 6))
        # self.write_data(2, 6, self.param1_var.get())
        self.param1_var.set(self.read_data(2, 6))
        # self.write_data(3, 6, self.param2_var.get())
        self.param2_var.set(self.read_data(3, 6))
        # self.write_data(4, 6, self.minRadius_var.get())
        self.minRadius_var.set(self.read_data(4, 6))
        # self.write_data(5, 6, self.maxRadius_var.get())
        self.maxRadius_var.set(self.read_data(5, 6))
        # self.write_data(6, 6, self.var_canny_a.get())
        self.var_canny_a.set(self.read_data(6, 6))
        # self.write_data(7, 6, self.var_canny_b.get())
        self.var_canny_b.set(self.read_data(7, 6))
        # self.write_data(8, 6, self.var_thr.get())
        self.var_thr.set(self.read_data(8, 6))
        # self.write_data(9, 6, self.len_size2.get())
        self.len_size2.set(self.read_data(9, 6))
        # self.write_data(10, 6, self.len_size3.get())
        self.len_size3.set(self.read_data(10, 6))
        # self.write_data(11, 6, self.len_size1.get())
        self.len_size1.set(self.read_data(11, 6))
        # # colum7
        #
        # self.write_data(1, 7, self.circle_select.get())
        self.circle_select.set(self.read_data(1, 7))
        # self.write_data(2, 7, self.circle_select_adv.get())
        self.circle_select_adv.set(self.read_data(2, 7))
        # self.write_data(3, 7, self.triang_select.get())
        self.triang_select.set(self.read_data(3, 7))
        # self.write_data(4, 7, self.rect_select.get())
        self.rect_select.set(self.read_data(4, 7))
        # self.write_data(5, 7, self.edge_select.get())
        self.edge_select.set(self.read_data(5, 7))
        # self.draw_contour_select.set(0)
        self.draw_contour_select.set(self.read_data(6, 7))
        #
        self.connected_select.set(self.read_data(7, 7))
        #
        self.rectangge_component_select.set(self.read_data(8, 7))
        # # colum8
        # self.write_data(1, 8, self.boder_map.current())
        self.boder_map.current(self.read_data(1, 8))
        # self.write_data(2, 8, self.contua_map.current())
        self.contua_map.current(self.read_data(2, 8))
        # self.write_data(3, 8, self.contua_map1.current())
        self.contua_map1.current(self.read_data(3, 8))
        # self.write_data(4, 8, self.rect_para1.get())
        self.rect_para1.set(self.read_data(4, 8))
        # self.write_data(5, 8, self.area_rect.get())
        self.area_rect.set(self.read_data(5, 8))
        # self.write_data(6, 8, self.chuvi_rect.get())
        self.chuvi_rect.set(self.read_data(6, 8))
        # self.write_data(7, 8, self.rect_thres.get())
        self.rect_thres.set(self.read_data(7, 8))
        # colum9
        # self.write_data(1, 8, self.boder_map.current())
        self.boder_map2.current(self.read_data(1, 9))
        # self.write_data(2, 8, self.contua_map.current())
        self.contua_map2.current(self.read_data(2, 9))
        # self.write_data(3, 8, self.contua_map1.current())
        self.contua_map21.current(self.read_data(3, 9))
        # self.write_data(4, 8, self.rect_para1.get())
        self.trian_para1.set(self.read_data(4, 9))
        # self.write_data(5, 8, self.area_rect.get())
        self.area_trian.set(self.read_data(5, 9))
        # self.write_data(6, 8, self.chuvi_rect.get())
        self.chuvi_trian.set(self.read_data(6, 9))
        # self.write_data(7, 8, self.rect_thres.get())
        self.trian_thres.set(self.read_data(7, 9))
        ##########
        # self.write_data(1, 10, self.Threshole_contua_var.get())
        self.Threshole_contua_var.set(self.read_data(1, 10))
        # self.write_data(2, 10, self.contua_aprrox.get())
        self.contua_aprrox.set(self.read_data(2, 10))
        # self.write_data(3, 10, self.chuvi_contua.get())
        self.chuvi_contua.set(self.read_data(3, 10))
        # self.write_data(4, 10, self.dientich_contua.get())
        self.dientich_contua.set(self.read_data(4, 10))
        # colum 11
        # self.write_data(1, 11, self.border_connect.current())
        self.border_connect.current(self.read_data(1, 11))
        # self.write_data(2, 11, self.kernel_connect.get())
        self.kernel_connect.set(self.read_data(2, 11))
        # self.write_data(3, 11, self.iterations_var.get())
        self.iterations_var.set(self.read_data(3, 11))
        # self.write_data(4, 11, self.area_detect_connect.get())
        self.area_detect_connect.set(self.read_data(4, 11))
        # self.write_data(5, 11, self.thres_connect.get())
        self.thres_connect.set(self.read_data(5, 11))
        # colum 12
        # self.write_data(1, 12, self.fill_change_var.get())
        self.fill_change_var.set(self.read_data(1, 12))
        # self.write_data(2, 12, self.area_patan.get())
        self.area_patan.set(self.read_data(2, 12))
        # self.write_data(3, 12, self.chuvi_patan.get())
        self.chuvi_patan.set(self.read_data(3, 12))
        # self.write_data(4, 12, self.thres_patan.get())
        self.thres_patan.set(self.read_data(4, 12))

        self.stt = self.read_data(1, 13)
        self.last_day_on = self.read_data(2, 13)

        # self.write_data(11, 4, self.pt1_specal[0])
        self.pt1_specal = [self.read_data(1, 14), self.read_data(2, 14)]
        self.pt2_specal = [self.read_data(3, 14), self.read_data(4, 14)]
        self.pt3_specal = [self.read_data(5, 14), self.read_data(6, 14)]

        #
        self.pt1_kaku = [self.read_data(7, 14), self.read_data(8, 14)]
        self.pt2_kaku = [self.read_data(9, 14), self.read_data(10, 14)]
        self.pt3_kaku = [self.read_data(11, 14), self.read_data(12, 14)]
        self.pt4_kaku = [self.read_data(13, 14), self.read_data(14, 14)]
        #
        #
        self.pt1_polygol5 = [self.read_data(15, 14), self.read_data(16, 14)]
        self.pt2_polygol5 = [self.read_data(17, 14), self.read_data(18, 14)]
        self.pt3_polygol5 = [self.read_data(19, 14), self.read_data(20, 14)]
        self.pt4_polygol5 = [self.read_data(21, 14), self.read_data(22, 14)]
        self.pt5_polygol5 = [self.read_data(23, 14), self.read_data(24, 14)]

        self.check_patan1.set(self.read_data(1, 15))
        self.check_patan2.set(self.read_data(2, 15))
        self.check_patan3.set(self.read_data(3, 15))
        self.check_patan4.set(self.read_data(4, 15))
        self.check_patan5.set(self.read_data(5, 15))
        self.check_patan6.set(self.read_data(6, 15))

        self.add_fiter_to()
        self.minus_fiter_to()

        self.capture.set(cv2.CAP_PROP_BRIGHTNESS, self.BRIGHTNESS_var.get())
        self.capture.set(cv2.CAP_PROP_CONTRAST, self.CONTRAST_var.get())
        self.capture.set(cv2.CAP_PROP_SATURATION, self.SATURATION_var.get())
        self.capture.set(cv2.CAP_PROP_HUE, self.HUE_var.get())
        self.capture.set(cv2.CAP_PROP_GAIN, self.GAIN_var.get())
        self.capture.set(cv2.CAP_PROP_AUTO_EXPOSURE, self.auto_exposure_var.get())
        self.capture.set(cv2.CAP_PROP_AUTOFOCUS, self.auto_focus_var.get())
        self.capture.set(cv2.CAP_PROP_EXPOSURE, self.EXPOSE_var.get())

        self.day_log()

if __name__ == "__main__":
    root = tk.Tk()
    app = Application(master=root)
    #app.Disp_draw()
    # draw_thread = threading.Thread(target=app.Disp_draw)
    # draw_thread.start()
    app.auto_load()
    # app.camera_callback()
    app.autoconnect_comport()
    app.autoconnect_msp_comport_M1_2()
    app.autoconnect_msp_comport_M3_4()
    app.autoconnect_msp_comport_M5_6()
    #app.hakotori_kakunin()
    # app.camera_callback()
    #app.socket_connect()
    # time.sleep(10)
    # app.redraw_image()
    app.mainloop()
